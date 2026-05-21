"""
Import dữ liệu Khách Hàng từ file 'HoDanTienNga_2025.xlsx' tab 'Danh Sách Chat ID'
vào bảng customers. Cập nhật dữ liệu nếu đã tồn tại, thêm mới nếu chưa có.
"""
import openpyxl
from app.db.session import SessionLocal
from app.models.business import Customers, CollectionPoint

# === Config ===
EXCEL_FILE = "HoDanTienNga_2025_1.xlsx"
SHEET_NAME = "Danh Sách Chat ID"

def clean_chat_id(val) -> str | None:
    """Xử lý Chat ID: bỏ .0 nếu có."""
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str if val_str and val_str != "None" else None

def safe_int(val, default=0) -> int:
    """Convert value sang int, mặc định 0."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default

def main():
    print(f"📂 Đọc file: {EXCEL_FILE}, sheet: {SHEET_NAME}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    db = SessionLocal()
    try:
        # Pre-fetch collection points for fast mapping
        db_cps = db.query(CollectionPoint).all()
        cp_name_to_id = {cp.collection_name.strip().lower(): str(cp.id) for cp in db_cps if cp.collection_name}

        # Thu thập dữ liệu từ Excel
        rows_data = {}
        # Header ở dòng 2, data bắt đầu từ dòng 3
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            diem_thu_mua = row[2]  # C - Điểm Thu Mua
            ten_kh = row[3]        # D - Tên Khách Hàng
            cong_no = row[4]       # E - Công Nợ Hiện Tại
            ma_ho = row[5]         # F - Mã Hộ
            tro_gia = row[6]       # G - Trợ Giá
            chat_id = row[7]       # H - Chat ID

            if ma_ho is None or str(ma_ho).strip() == "" or str(ma_ho).strip() == "None":
                continue

            ma_ho_str = str(ma_ho).strip()
            
            # Xác định collection_point_id
            cp_id = None
            if diem_thu_mua:
                cp_name_lower = str(diem_thu_mua).strip().lower()
                cp_id = cp_name_to_id.get(cp_name_lower)
            
            if not cp_id:
                print(f"  ⚠️ Không tìm thấy collection_point cho '{diem_thu_mua}', Mã KH={ma_ho_str}")
                continue

            if ma_ho_str in rows_data:
                print(f"  ⚠️ Trùng lặp Mã KH={ma_ho_str}, ghi đè dữ liệu cũ.")

            rows_data[ma_ho_str] = {
                "id": ma_ho_str,
                "hoursehold_id": ma_ho_str,
                "fullname": str(ten_kh).strip() if ten_kh else "",
                "collection_point_id": cp_id,
                "total_debt": safe_int(cong_no),
                "is_subsidized": safe_int(tro_gia),
                "telegram_group": None,
                "amount_of_debt": 0,
                "cash_advance": 0,
                "status": "ACTIVE",
                "username": None,
                "ingredient": None,
                "number_phone": None,
                "address": None,
            }

        print(f"📊 Tổng dòng hợp lệ từ Excel: {len(rows_data)}")

        # 1. Update hoặc Insert (Upsert) dữ liệu
        print(f"\n📥 Xử lý {len(rows_data)} khách hàng từ Excel...")
        updated_count = 0
        inserted_count = 0
        for data in rows_data.values():
            existing_customer = db.query(Customers).filter(Customers.id == data["id"]).first()
            if existing_customer:
                # Update
                existing_customer.fullname = data["fullname"]
                existing_customer.collection_point_id = data["collection_point_id"]
                existing_customer.total_debt = data["total_debt"]
                existing_customer.is_subsidized = data["is_subsidized"]
                updated_count += 1
            else:
                # Insert
                new_customer = Customers(**data)
                db.add(new_customer)
                inserted_count += 1

        db.commit()
        print(f"✅ Import thành công! (Thêm mới: {inserted_count}, Cập nhật: {updated_count})")

        # 3. Verify
        new_count = db.query(Customers).count()
        print(f"\n📊 Verify: {new_count} khách hàng trong DB")

        # Thống kê theo collection point
        from sqlalchemy import func
        stats = (
            db.query(CollectionPoint.collection_name, func.count(Customers.id))
            .join(Customers, Customers.collection_point_id == CollectionPoint.id)
            .group_by(CollectionPoint.collection_name)
            .all()
        )
        print("\n📊 Phân bố theo Điểm Thu Mua:")
        for cp_name, count in stats:
            print(f"  • {cp_name}: {count} KH")

        # Spot check
        print("\n🔍 Spot check:")
        samples = db.query(Customers).limit(5).all()
        for s in samples:
            cp = db.query(CollectionPoint).filter(CollectionPoint.id == s.collection_point_id).first()
            print(f"  {s.id} | {s.fullname} | {cp.collection_name if cp else 'N/A'} | total_debt={s.total_debt} | telegram_group={s.telegram_group}")

    except Exception as e:
        db.rollback()
        print(f"❌ Lỗi: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()


if __name__ == "__main__":
    main()

"""
Script gửi lệnh /syncchat đến các nhóm Telegram từ file Test_Syncchat.xlsx
Sử dụng Telegram Bot API trực tiếp (aiohttp) để tránh lỗi Peer ID của Pyrogram
"""
import asyncio
import json
import sys
import io
import openpyxl
import aiohttp

# Fix Unicode output trên Windows console
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BOT_API_URL = "https://api.telegram.org/bot{token}/sendMessage"


def load_config() -> dict:
    """Đọc config từ appsettings.json"""
    with open("appsettings.json", "r", encoding="utf-8") as f:
        return json.load(f)


def load_chat_ids(file_path: str = "HoDanTienNga_2025.xlsx") -> list[dict]:
    """Đọc danh sách chat_id từ file Excel (tự tìm cột theo tên header)"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Danh Sách Chat ID"]

    # Đọc header row để tìm vị trí cột
    headers = [cell.value for cell in ws[1]]
    print(f"[DEBUG] Headers: {headers}")

    # Tìm index các cột cần thiết (case-insensitive)
    col_map = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_lower = str(h).strip().lower()
        if "chat" in h_lower and "id" in h_lower:
            col_map["chat_id"] = i
        elif "mã" in h_lower or "ma ho" in h_lower:
            col_map["ma_ho"] = i
        elif "khách" in h_lower or "ten" in h_lower:
            col_map["ten_kh"] = i
        elif "điểm" in h_lower or "thu mua" in h_lower:
            col_map["diem_thu_mua"] = i

    print(f"[DEBUG] Column mapping: {col_map}")

    if "chat_id" not in col_map:
        raise ValueError("Không tìm thấy cột 'Chat ID' trong file Excel!")

    groups = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        values = list(row)
        chat_id = values[col_map["chat_id"]] if col_map.get("chat_id") is not None else None
        ma_ho = values[col_map["ma_ho"]] if col_map.get("ma_ho") is not None else f"ROW{idx}"
        ten_kh = values[col_map["ten_kh"]] if col_map.get("ten_kh") is not None else "N/A"

        if chat_id is not None:
            groups.append({
                "stt": idx - 1,
                "ten_kh": ten_kh,
                "ma_ho": ma_ho,
                "chat_id": int(chat_id),
            })
    return groups


async def main():
    config = load_config()
    bot_token = config["Telegram"]["Bot_Token"]
    url = BOT_API_URL.format(token=bot_token)

    groups = load_chat_ids()

    print(f"=== Tìm thấy {len(groups)} nhóm từ file Excel ===")
    for g in groups:
        print(f"   - [{g['ma_ho']}] {g['ten_kh']} | Chat ID: {g['chat_id']}")
    print()

    async with aiohttp.ClientSession() as session:
        for group in groups:
            chat_id = group["chat_id"]
            ten_kh = group["ten_kh"]
            ma_ho = group["ma_ho"]
            try:
                async with session.post(url, json={
                    "chat_id": chat_id,
                    "text": "/syncchat"
                }) as resp:
                    result = await resp.json()
                    if result.get("ok"):
                        print(f"✅ Đã gửi /syncchat đến nhóm [{ma_ho}] {ten_kh} (chat_id: {chat_id})")
                    else:
                        print(f"❌ Lỗi từ API [{ma_ho}] {ten_kh} (chat_id: {chat_id}): {result.get('description')}")
            except Exception as e:
                print(f"❌ Lỗi khi gửi đến nhóm [{ma_ho}] {ten_kh} (chat_id: {chat_id}): {e}")

            # Delay giữa các tin nhắn để tránh rate limit
            await asyncio.sleep(2)

    print("\n🎉 Hoàn tất!")


if __name__ == "__main__":
    asyncio.run(main())

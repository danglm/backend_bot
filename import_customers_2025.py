"""
Import dữ liệu Khách Hàng từ file 'CS Tiến Nga 2025.xlsx' tab 'DATA KH'
vào bảng customers. Xóa hết dữ liệu cũ trước khi import.
Đồng thời update tên collection_points (ĐL → Đại lí).
"""
import openpyxl
import re
from app.db.session import SessionLocal
from app.models.business import Customers, CollectionPoint

# === Config ===
EXCEL_FILE = "CS Tiến Nga 2025.xlsx"
SHEET_NAME = "DATA KH"

# Mapping Mã TM → collection_point UUID
MA_TM_TO_CP = {
    "X":  "afaf045b-65f8-41f1-b6ad-1305cb9fba00",   # Xưởng Gia An
    "P":  "bbe93e72-a7a6-48a6-9ce1-c4aa3632232e",   # Xưởng Phê
    "LT": "3ace911d-c612-414a-8292-feef297f7362",    # Xưởng Lạc Tánh
}

# Mapping riêng cho nhóm DL (theo tên KH)
DL_NAME_TO_CP = {
    "DL001": "e3c3e719-4d94-4592-af32-5e80400f519a",   # Đại lí Vui (Chị Vui)
    "DL002": "0251c5de-358a-4b48-b74d-a4df056962d7",   # Đại lí Trang (Chị Trang)
    "DL003": "f048d641-826a-4392-b17f-54c39dc837ee",   # Đại lí Hải (Cô Hải)
}

# Update tên collection_points
CP_RENAME = {
    "f048d641-826a-4392-b17f-54c39dc837ee": "Đại lí Hải",
    "0251c5de-358a-4b48-b74d-a4df056962d7": "Đại lí Trang",
    "e3c3e719-4d94-4592-af32-5e80400f519a": "Đại lí Vui",
}


def strip_prefix(fullname: str, ma_tm: str) -> str:
    """Bỏ prefix Mã TM khỏi tên KH rồi uppercase.
    VD: 'X 2 Cấp' với ma_tm='X' → '2 CẤP'
        'LT 5 Chờm' với ma_tm='LT' → '5 CHỜM'
        'DL Chị Vui' với ma_tm='DL' → 'CHỊ VUI'
    """
    if fullname.startswith(ma_tm + " "):
        fullname = fullname[len(ma_tm) + 1:]
    return fullname.strip()


def clean_phone(phone_val) -> str | None:
    """Xử lý số điện thoại: strip dấu ' đầu, trả về None nếu rỗng."""
    if phone_val is None:
        return None
    phone = str(phone_val).strip().lstrip("'")
    return phone if phone else None


def safe_int(val, default=0) -> int:
    """Convert value sang int, mặc định 0."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def main():
    print(f"📂 Đọc file: {EXCEL_FILE}, sheet: {SHEET_NAME}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    # Thu thập dữ liệu từ Excel
    rows_data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        ma_kh = row[1]   # B - Mã KH
        if ma_kh is None:
            continue

        ten_kh = row[0]     # A - Tên KH
        cong_no = row[2]    # C - Công Nợ
        tro_gia = row[3]    # D - Trợ Giá
        ma_tm = row[4]      # E - Mã TM
        ung_tien = row[5]   # F - Ứng Tiền
        dia_chi = row[6]    # G - Địa Chỉ
        so_dt = row[7]      # H - Số ĐT

        # Xác định collection_point_id
        if ma_tm == "DL":
            cp_id = DL_NAME_TO_CP.get(str(ma_kh))
        else:
            cp_id = MA_TM_TO_CP.get(str(ma_tm))

        if cp_id is None:
            print(f"  ⚠️ Không tìm thấy collection_point cho Mã TM={ma_tm}, Mã KH={ma_kh}")
            continue

        # Xử lý tên
        fullname = strip_prefix(str(ten_kh), str(ma_tm)) if ten_kh else ""

        rows_data.append({
            "id": str(ma_kh),
            "hoursehold_id": str(ma_kh),
            "fullname": fullname,
            "collection_point_id": cp_id,
            "number_phone": clean_phone(so_dt),
            "address": str(dia_chi).strip() if dia_chi else None,
            "amount_of_debt": safe_int(cong_no),
            "is_subsidized": safe_int(tro_gia),
            "cash_advance": safe_int(ung_tien),
            "total_debt": 0,
            "ingredient": None,
            "username": None,
            "status": "ACTIVE",
        })

    print(f"📊 Tổng dòng hợp lệ từ Excel: {len(rows_data)}")

    # === Thao tác DB ===
    db = SessionLocal()
    try:
        # 1. Update tên collection_points
        print("\n🔄 Update tên Collection Points...")
        for cp_id, new_name in CP_RENAME.items():
            cp = db.query(CollectionPoint).filter(CollectionPoint.id == cp_id).first()
            if cp:
                old_name = cp.collection_name
                cp.collection_name = new_name
                print(f"  ✅ '{old_name}' → '{new_name}'")
            else:
                print(f"  ⚠️ Không tìm thấy CP id={cp_id}")

        # 2. Xóa hết dữ liệu cũ
        old_count = db.query(Customers).count()
        db.query(Customers).delete()
        print(f"\n🗑️ Đã xóa {old_count} khách hàng cũ")

        # 3. Insert dữ liệu mới
        print(f"\n📥 Insert {len(rows_data)} khách hàng mới...")
        for data in rows_data:
            customer = Customers(**data)
            db.add(customer)

        db.commit()
        print(f"✅ Import thành công!")

        # 4. Verify
        new_count = db.query(Customers).count()
        print(f"\n📊 Verify: {new_count} khách hàng trong DB")

        # Thống kê theo collection point
        from sqlalchemy import func
        stats = (
            db.query(CollectionPoint.collection_name, func.count(Customers.id))
            .join(Customers, Customers.collection_point_id == CollectionPoint.id)
            .group_by(CollectionPoint.collection_name)
            .all()
        )
        print("\n📊 Phân bố theo Xưởng:")
        for cp_name, count in stats:
            print(f"  • {cp_name}: {count} KH")

        # Spot check
        print("\n🔍 Spot check:")
        samples = db.query(Customers).filter(Customers.id.in_(["X001", "LT001", "P001", "DL001"])).all()
        for s in samples:
            cp = db.query(CollectionPoint).filter(CollectionPoint.id == s.collection_point_id).first()
            print(f"  {s.id} | {s.fullname} | {cp.collection_name if cp else 'N/A'} | is_subsidized={s.is_subsidized} | phone={s.number_phone}")

    except Exception as e:
        db.rollback()
        print(f"❌ Lỗi: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()


if __name__ == "__main__":
    main()

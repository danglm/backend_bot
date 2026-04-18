"""
Script import dữ liệu khách hàng từ List_Customer.xlsx vào bảng customers.
Sheet: "Danh Sách Khách Hàng"
- Dữ liệu bắt đầu từ Row 4 (Row 2 = header, Row 3 = trống)
- Nếu mã hộ đã tồn tại → ghi đè (upsert)
- Collection point: match tên trong Excel với bảng collection_points (bỏ prefix "Xưởng ")
- Username: để None
"""

import sys
import os

# Add parent directory to path to import app modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import unicodedata
import openpyxl
from app.db.session import SessionLocal
from app.models.business import Customers, CollectionPoint
from sqlalchemy.dialects.postgresql import insert as pg_insert


# ──────────────────────── Helpers ────────────────────────

def strip_diacritics(text: str) -> str:
    """Bỏ dấu tiếng Việt để so sánh.
    VD: 'Hải' → 'Hai', 'Lạc Tánh' → 'Lac Tanh'
    """
    nfkd = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def normalize_cp_name(name: str) -> str:
    """Chuẩn hoá tên điểm thu mua để so sánh.
    DB có thể chứa 'Xưởng Gia An', Excel chỉ có 'Gia An'.
    Cũng bỏ dấu tiếng Việt để match 'DL Hải' == 'DL Hai'.
    """
    if name is None:
        return ""
    name = name.strip()
    # Bỏ prefix "Xưởng " nếu có
    if name.startswith("Xưởng ") or name.startswith("Xuong "):
        name = name.split(" ", 1)[1]
    return strip_diacritics(name).upper()


def build_cp_lookup(db) -> dict:
    """Tạo dict mapping: tên chuẩn hoá → UUID id."""
    collection_points = db.query(CollectionPoint).all()
    lookup = {}
    for cp in collection_points:
        normalized = normalize_cp_name(cp.collection_name)
        lookup[normalized] = cp.id
        # Cũng thêm tên gốc (không bỏ prefix) để match trường hợp "DL Trang"
        original = cp.collection_name.strip().upper() if cp.collection_name else ""
        lookup[original] = cp.id
    return lookup


def format_phone(value) -> str | None:
    """Chuyển số điện thoại từ float → string có prefix '0'.
    VD: 396614465.0 → '0396614465'
    """
    if value is None:
        return None
    try:
        phone = str(int(float(value)))
        if not phone.startswith("0"):
            phone = "0" + phone
        return phone
    except (ValueError, TypeError):
        return str(value).strip() if value else None


def clean_money(value) -> int:
    """Chuyển giá trị tiền thành int. Xử lý None, '-', string."""
    if value is None:
        return 0
    if isinstance(value, str):
        value = value.strip()
        if value in ("-", "", "None"):
            return 0
        try:
            # Xoá dấu phẩy ngăn cách hàng nghìn nếu có
            value = value.replace(",", "").replace(".", "")
            return int(value)
        except ValueError:
            return 0
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0


# ──────────────────────── Main ────────────────────────

def main():
    print("=" * 60)
    print("  IMPORT CUSTOMERS FROM List_Customer.xlsx")
    print("=" * 60)

    # ── 1. Load Excel ──
    excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "List_Customer.xlsx")
    print(f"\n📂 Đọc file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Danh Sách Khách Hàng"]
    print(f"   Sheet: {ws.title} | Rows: {ws.max_row} | Cols: {ws.max_column}")

    # ── 2. Build Collection Point lookup ──
    db = SessionLocal()
    cp_lookup = build_cp_lookup(db)
    print(f"\n📍 Collection Points trong DB ({len(cp_lookup) // 2} điểm):")
    seen = set()
    for name, uid in cp_lookup.items():
        if uid not in seen:
            print(f"   - {name} → {uid}")
            seen.add(uid)

    # ── 3. Parse & Collect Data ──
    stats = {"imported": 0, "updated": 0, "skipped": 0, "errors": 0}
    DATA_START_ROW = 4
    # Dùng dict để loại trùng mã hộ trong cùng file Excel (giữ dòng cuối)
    records = {}

    print(f"\n🔄 Bắt đầu đọc dữ liệu từ Row {DATA_START_ROW}...\n")

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        stt = ws.cell(row_idx, 2).value       # Col B: STT
        ma_ho = ws.cell(row_idx, 7).value      # Col G: Mã Hộ

        # Bỏ qua dòng trống (không có STT hoặc Mã Hộ)
        if stt is None or ma_ho is None:
            continue

        ma_ho = str(ma_ho).strip()
        if not ma_ho:
            continue

        try:
            # Đọc dữ liệu từ Excel
            diem_thu_mua = ws.cell(row_idx, 3).value   # Col C: Điểm Thu Mua
            ten_kh = ws.cell(row_idx, 6).value          # Col F: Tên Khách Hàng
            sdt = ws.cell(row_idx, 8).value             # Col H: Số Điện Thoại
            dia_chi = ws.cell(row_idx, 12).value        # Col L: Địa Chỉ
            nguyen_lieu = ws.cell(row_idx, 16).value    # Col P: Nguyên Liệu
            so_tien_no = ws.cell(row_idx, 17).value     # Col Q: Số Tiền Nợ
            ung_tien = ws.cell(row_idx, 18).value       # Col R: Ứng Tiền Cuối Mùa
            tong_cong_no = ws.cell(row_idx, 19).value   # Col S: Tổng Công Nợ

            # Lookup collection point ID
            cp_id = None
            if diem_thu_mua:
                cp_key = normalize_cp_name(str(diem_thu_mua))
                cp_id = cp_lookup.get(cp_key)
                if cp_id is None:
                    print(f"   ⚠️  Row {row_idx}: Không tìm thấy điểm thu mua '{diem_thu_mua}' → bỏ qua collection_point_id")

            if ma_ho in records:
                print(f"   ⚠️  Row {row_idx}: Mã hộ '{ma_ho}' bị trùng trong Excel → ghi đè bằng dòng mới")
                stats["skipped"] += 1

            records[ma_ho] = {
                "id": ma_ho,
                "fullname": str(ten_kh).strip() if ten_kh else None,
                "hoursehold_id": ma_ho,
                "collection_point_id": cp_id,
                "number_phone": format_phone(sdt),
                "address": str(dia_chi).strip() if dia_chi else None,
                "ingredient": str(nguyen_lieu).strip() if nguyen_lieu else None,
                "amount_of_debt": clean_money(so_tien_no),
                "cash_advance": clean_money(ung_tien),
                "total_debt": clean_money(tong_cong_no),
                "status": "ACTIVE",
                "username": None,
            }

        except Exception as e:
            print(f"   ❌ Row {row_idx} (Mã hộ: {ma_ho}): {e}")
            stats["errors"] += 1
            continue

    print(f"\n📊 Đã đọc {len(records)} khách hàng duy nhất từ Excel")

    # ── 4. Upsert vào DB (INSERT ... ON CONFLICT UPDATE) ──
    print("\n💾 Đang ghi vào database...")
    try:
        for ma_ho, data in records.items():
            stmt = pg_insert(Customers).values(**data)
            stmt = stmt.on_conflict_do_update(
                index_elements=["id"],
                set_={
                    "fullname": stmt.excluded.fullname,
                    "hoursehold_id": stmt.excluded.hoursehold_id,
                    "collection_point_id": stmt.excluded.collection_point_id,
                    "number_phone": stmt.excluded.number_phone,
                    "address": stmt.excluded.address,
                    "ingredient": stmt.excluded.ingredient,
                    "amount_of_debt": stmt.excluded.amount_of_debt,
                    "cash_advance": stmt.excluded.cash_advance,
                    "total_debt": stmt.excluded.total_debt,
                    "status": stmt.excluded.status,
                    "username": stmt.excluded.username,
                },
            )
            result = db.execute(stmt)
            if result.rowcount == 1:
                stats["imported"] += 1
            else:
                stats["updated"] += 1

        db.commit()
        print("✅ Commit thành công!")
    except Exception as e:
        db.rollback()
        print(f"❌ Commit thất bại: {e}")
        return
    finally:
        db.close()

    # ── 5. Summary ──
    print("\n" + "=" * 60)
    print("  KẾT QUẢ IMPORT")
    print("=" * 60)
    print(f"  ✅ Imported (mới):  {stats['imported']}")
    print(f"  🔄 Updated (ghi đè): {stats['updated']}")
    print(f"  ⏭️  Skipped:         {stats['skipped']}")
    print(f"  ❌ Errors:           {stats['errors']}")
    print(f"  📊 Tổng xử lý:      {stats['imported'] + stats['updated'] + stats['skipped'] + stats['errors']}")
    print("=" * 60)


if __name__ == "__main__":
    main()

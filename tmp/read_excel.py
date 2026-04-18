import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

wb = openpyxl.load_workbook('List_Customer.xlsx', data_only=True)
ws = wb['Thu Mua Hàng Ngày']

output = []
output.append(f"Sheet: {ws.title}")
output.append(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# ALL headers (Row 2) - all columns
output.append("\n=== ALL HEADERS (Row 2) ===")
for c in range(1, ws.max_column + 1):
    output.append(f"  Col {c}: {ws.cell(2, c).value}")

# Sample data rows 3-8 - ALL columns
output.append("\n=== SAMPLE DATA (Rows 3-8) - ALL COLUMNS ===")
for r in range(3, 9):
    output.append(f"\n  Row {r}:")
    for c in range(1, ws.max_column + 1):
        val = ws.cell(r, c).value
        output.append(f"    Col {c}: {repr(val)}")

# Count actual data rows
count = 0
null_count = 0
na_count = 0
for r in range(3, ws.max_row + 1):
    stt = ws.cell(r, 2).value
    if stt is not None:
        count += 1
        ma_ho = ws.cell(r, 7).value
        if ma_ho is None:
            null_count += 1
        elif str(ma_ho).strip() in ('#N/A', 'N/A', ''):
            na_count += 1

output.append(f"\n\n=== DATA STATS ===")
output.append(f"Total rows with STT: {count}")
output.append(f"Rows with Ma Ho = None: {null_count}")
output.append(f"Rows with Ma Ho = #N/A: {na_count}")

# Check date range
min_date = None
max_date = None
for r in range(3, ws.max_row + 1):
    day = ws.cell(r, 4).value
    if day is not None:
        import datetime
        if isinstance(day, datetime.datetime):
            d = day.date()
        elif isinstance(day, datetime.date):
            d = day
        else:
            continue
        if min_date is None or d < min_date:
            min_date = d
        if max_date is None or d > max_date:
            max_date = d

output.append(f"\nDate range: {min_date} → {max_date}")

# Check collection points in this sheet
cp_set = set()
for r in range(3, ws.max_row + 1):
    cp = ws.cell(r, 5).value
    if cp is not None:
        cp_set.add(str(cp).strip())
output.append(f"\nCollection points: {sorted(cp_set)}")

# Last rows with data
output.append("\n=== LAST DATA ROWS ===")
last_data_row = 3
for r in range(ws.max_row, 2, -1):
    if ws.cell(r, 2).value is not None:
        last_data_row = r
        break
output.append(f"Last row with data: {last_data_row}")
for r in range(max(3, last_data_row - 3), last_data_row + 1):
    row_data = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    output.append(f"  Row {r}: {row_data}")

with open('tmp/daily_purchase_analysis.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

print("Done! Check tmp/daily_purchase_analysis.txt")

"""
Script import dữ liệu thu mua hàng ngày từ List_Customer.xlsx vào bảng daily_purchases.
Sheet: "Thu Mua Hàng Ngày"
- Dữ liệu bắt đầu từ Row 3 (Row 2 = header)
- Mã Hộ = #N/A → None
- Collection point: normalize tên + bỏ dấu để match. Tạo mới nếu chưa có.
- Chạy lại script → thêm tiếp (append)
- Batch insert 1000 records / commit
"""

import sys
import os
import unicodedata
import datetime

# Add parent directory to path to import app modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from app.db.session import SessionLocal
from app.models.business import DailyPurchases, CollectionPoint


# ──────────────────────── Helpers ────────────────────────

def strip_diacritics(text: str) -> str:
    """Bỏ dấu tiếng Việt để so sánh.
    VD: 'Hải' → 'Hai', 'Lạc Tánh' → 'Lac Tanh'
    """
    nfkd = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def normalize_cp_name(name: str) -> str:
    """Chuẩn hoá tên điểm thu mua để so sánh.
    Bỏ prefix 'Xưởng ', bỏ dấu, uppercase.
    """
    if name is None:
        return ""
    name = name.strip()
    prefixes = ["Xưởng ", "XƯỞNG ", "Xuong "]
    for prefix in prefixes:
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    return strip_diacritics(name).upper()


def build_cp_lookup(db) -> dict:
    """Tạo dict mapping: tên chuẩn hoá → UUID id."""
    collection_points = db.query(CollectionPoint).all()
    lookup = {}
    for cp in collection_points:
        normalized = normalize_cp_name(cp.collection_name)
        lookup[normalized] = cp.id
    return lookup


def safe_float(value, default=0.0) -> float:
    """Chuyển giá trị sang float an toàn."""
    if value is None:
        return default
    if isinstance(value, str):
        value = value.strip()
        if value in ('#N/A', 'N/A', '-', ''):
            return default
        try:
            return float(value.replace(',', '').replace('.', '').replace(' ', ''))
        except ValueError:
            return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def safe_int(value, default=0) -> int:
    """Chuyển giá trị sang int an toàn."""
    return int(safe_float(value, float(default)))


def clean_ma_ho(value) -> str | None:
    """Xử lý mã hộ: #N/A → None."""
    if value is None:
        return None
    s = str(value).strip()
    if s in ('#N/A', 'N/A', '', 'None'):
        return None
    return s


def to_date(value):
    """Chuyển datetime/date/string → date."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    # Try parse string
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ──────────────────────── Main ────────────────────────

def main():
    BATCH_SIZE = 1000

    print("=" * 60)
    print("  IMPORT DAILY PURCHASES FROM List_Customer.xlsx")
    print("=" * 60)

    # ── 1. Load Excel ──
    excel_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "List_Customer.xlsx",
    )
    print(f"\n📂 Đọc file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Thu Mua Hàng Ngày"]
    print(f"   Sheet: {ws.title} | Rows: {ws.max_row} | Cols: {ws.max_column}")

    # ── 2. Build Collection Point lookup ──
    db = SessionLocal()
    cp_lookup = build_cp_lookup(db)
    print(f"\n📍 Collection Points trong DB ({len(cp_lookup)} điểm):")
    for name, uid in cp_lookup.items():
        print(f"   - {name} → {uid}")

    # Track new CPs to create
    new_cps_created = []

    # ── 3. Read & Insert ──
    stats = {"imported": 0, "skipped": 0, "errors": 0, "na_houserhold": 0}
    DATA_START_ROW = 3
    batch = []

    print(f"\n🔄 Bắt đầu đọc dữ liệu từ Row {DATA_START_ROW}...")
    print(f"   Batch size: {BATCH_SIZE}\n")

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        stt = ws.cell(row_idx, 2).value  # Col B: STT
        if stt is None:
            continue

        try:
            # Đọc dữ liệu từ Excel
            tuan = ws.cell(row_idx, 3).value           # Col C: Tuần
            ngay = ws.cell(row_idx, 4).value            # Col D: Ngày
            diem_thu_mua = ws.cell(row_idx, 5).value    # Col E: Điểm Thu Mua
            ma_ho_raw = ws.cell(row_idx, 7).value       # Col G: Mã Hộ
            tro_gia = ws.cell(row_idx, 9).value         # Col I: Trợ Giá
            khoi_luong = ws.cell(row_idx, 10).value     # Col J: Khối Lượng
            tru_bi = ws.cell(row_idx, 11).value         # Col K: Trừ Bì
            kl_thuc_te = ws.cell(row_idx, 12).value     # Col L: KL Thực Tế
            so_do = ws.cell(row_idx, 13).value          # Col M: Số Độ
            mu_kho = ws.cell(row_idx, 14).value         # Col N: Mủ Khô
            don_gia = ws.cell(row_idx, 15).value        # Col O: Đơn Giá
            gia_ho_tro = ws.cell(row_idx, 16).value     # Col P: Giá Hỗ Trợ
            thanh_tien = ws.cell(row_idx, 17).value     # Col Q: Thành Tiền
            da_thanh_toan = ws.cell(row_idx, 18).value  # Col R: Đã Thanh Toán
            luu_so = ws.cell(row_idx, 19).value         # Col S: Lưu Sổ
            tam_ung = ws.cell(row_idx, 20).value        # Col T: Tạm Ứng
            da_kiem_tra = ws.cell(row_idx, 21).value    # Col U: Đã Kiểm Tra

            # Xử lý mã hộ
            ma_ho = clean_ma_ho(ma_ho_raw)
            if ma_ho is None:
                stats["na_houserhold"] += 1

            # Lookup collection point
            cp_id = None
            if diem_thu_mua:
                cp_key = normalize_cp_name(str(diem_thu_mua))
                cp_id = cp_lookup.get(cp_key)
                if cp_id is None:
                    # Tạo mới collection point
                    new_cp = CollectionPoint(
                        collection_name=str(diem_thu_mua).strip(),
                        address="",
                    )
                    db.add(new_cp)
                    db.flush()  # Get the new ID
                    cp_id = new_cp.id
                    cp_lookup[cp_key] = cp_id
                    new_cps_created.append(str(diem_thu_mua).strip())
                    print(f"   ✨ Tạo mới Collection Point: '{diem_thu_mua}' → {cp_id}")

            # Xử lý ngày
            day_value = to_date(ngay)
            if day_value is None:
                print(f"   ⚠️  Row {row_idx}: Ngày không hợp lệ ({ngay}) → bỏ qua")
                stats["skipped"] += 1
                continue

            # Build record
            record = DailyPurchases(
                hoursehold_id=ma_ho,
                collection_point_id=cp_id,
                week=safe_int(tuan),
                day=day_value,
                is_subsidized=safe_int(tro_gia),
                weight=safe_float(khoi_luong),
                tare_weight=safe_float(tru_bi),
                actual_weight=safe_float(kl_thuc_te),
                degree=safe_float(so_do),
                dry_rubber=safe_float(mu_kho),
                unit_price=safe_float(don_gia),
                subsidy_price=safe_float(gia_ho_tro),
                total_amount=safe_float(thanh_tien),
                paid_amount=safe_float(da_thanh_toan),
                saved_amount=safe_float(luu_so),
                advance_amount=safe_float(tam_ung),
                is_checked=bool(da_kiem_tra) if da_kiem_tra is not None else False,
            )
            batch.append(record)
            stats["imported"] += 1

            # Commit batch
            if len(batch) >= BATCH_SIZE:
                db.add_all(batch)
                db.commit()
                print(f"   ✅ Batch committed: {stats['imported']} records so far...")
                batch = []

        except Exception as e:
            print(f"   ❌ Row {row_idx}: {e}")
            stats["errors"] += 1
            continue

    # ── 4. Commit remaining ──
    if batch:
        try:
            db.add_all(batch)
            db.commit()
            print(f"   ✅ Final batch committed: {len(batch)} records")
        except Exception as e:
            db.rollback()
            print(f"   ❌ Final batch failed: {e}")
            stats["errors"] += len(batch)
            stats["imported"] -= len(batch)

    db.close()

    # ── 5. Summary ──
    print("\n" + "=" * 60)
    print("  KẾT QUẢ IMPORT")
    print("=" * 60)
    print(f"  ✅ Imported:           {stats['imported']:,}")
    print(f"  ⏭️  Skipped:            {stats['skipped']:,}")
    print(f"  ❌ Errors:              {stats['errors']:,}")
    print(f"  🔍 Mã Hộ = None (#N/A): {stats['na_houserhold']:,}")
    if new_cps_created:
        print(f"  ✨ Collection Points mới: {', '.join(set(new_cps_created))}")
    print(f"  📊 Tổng xử lý:         {stats['imported'] + stats['skipped'] + stats['errors']:,}")
    print("=" * 60)


if __name__ == "__main__":
    main()

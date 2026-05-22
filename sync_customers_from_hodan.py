"""
Script đồng bộ thành viên các nhóm khách hàng từ file HoDanTienNga_2025.xlsx
vào bảng telegram_project_members.
"""
import asyncio
import datetime
import openpyxl
from pyrogram.enums import ChatMemberStatus

from app.db.session import SessionLocal
from app.models.business import Projects
from app.models.telegram import TelegramProjectMember
from bot.utils.bot import bot
from bot.utils.logger import LogInfo, LogError, LogType

EXCEL_FILE = "HoDanTienNga_2025.xlsx"
SHEET_NAME = "Danh Sách Chat ID"

# Mapping Xưởng -> Parent ID
PARENT_MAP = {
    "gia an": "-1003843977286",
    "lạc tánh": "-1002380103246",
    "phê": "-1002365897939"
}

def clean_chat_id(val) -> str | None:
    """Xử lý Chat ID: bỏ .0 nếu có."""
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str if val_str and val_str != "None" else None

def get_parent_id(diem_thu_mua: str) -> str | None:
    """Xác định parent_id dựa vào Điểm Thu Mua (Xưởng)"""
    if not diem_thu_mua:
        return None
    diem_thu_mua_lower = str(diem_thu_mua).lower()
    for key, parent_id in PARENT_MAP.items():
        if key in diem_thu_mua_lower:
            return parent_id
    return None

async def main():
    print(f"📂 Đọc file: {EXCEL_FILE}, sheet: {SHEET_NAME}")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb[SHEET_NAME]
    except Exception as e:
        print(f"❌ Lỗi đọc file Excel: {e}")
        return

    try:
        db = SessionLocal()
        # 1. Tìm Project "Tiến Nga"
        project = db.query(Projects).filter(Projects.project_name.ilike("%Tiến Nga%")).first()
        if not project:
            print("❌ Không tìm thấy dự án 'Tiến Nga' trong DB.")
            db.close()
            return
        
        project_id = project.id
        db.close()
        print(f"✅ Đã tìm thấy dự án Tiến Nga, ID: {project_id}")

        # 2. Đọc danh sách Chat ID từ Excel
        headers = [cell.value for cell in ws[2]]
        
        col_map = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_lower = str(h).strip().lower()
            if "chat" in h_lower and "id" in h_lower:
                col_map["chat_id"] = i
            elif "điểm" in h_lower or "thu mua" in h_lower:
                col_map["diem_thu_mua"] = i
            elif "khách" in h_lower or "ten" in h_lower:
                col_map["ten_kh"] = i

        if "chat_id" not in col_map:
            print("❌ Không tìm thấy cột 'Chat ID' trong file Excel!")
            return

        groups = []
        for idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), start=3):
            values = list(row)
            raw_chat_id = values[col_map["chat_id"]] if col_map.get("chat_id") is not None else None
            diem_thu_mua = values[col_map["diem_thu_mua"]] if col_map.get("diem_thu_mua") is not None else None
            ten_kh = values[col_map["ten_kh"]] if col_map.get("ten_kh") is not None else f"Nhóm Dòng {idx}"

            chat_id = clean_chat_id(raw_chat_id)
            if chat_id:
                parent_id = get_parent_id(diem_thu_mua)
                groups.append({
                    "chat_id": chat_id,
                    "ten_kh": ten_kh,
                    "parent_id": parent_id
                })

        print(f"📊 Tìm thấy {len(groups)} nhóm cần đồng bộ.")

        # 3. Đồng bộ từ Telegram
        async with bot:
            for i, group in enumerate(groups, 1):
                chat_id = group["chat_id"]
                ten_kh = group["ten_kh"]
                parent_id = group["parent_id"]

                print(f"[{i}/{len(groups)}] Đang đồng bộ nhóm: {ten_kh} ({chat_id}) - Parent ID: {parent_id}")

                try:
                    # Lấy thông tin nhóm (để lấy title chính xác)
                    try:
                        chat_info = await bot.get_chat(int(chat_id))
                        chat_title = chat_info.title or ten_kh
                    except Exception as e:
                        print(f"  ⚠️ Không thể get_chat (có thể bot chưa vào nhóm hoặc chat_id sai): {e}")
                        chat_title = ten_kh
                        # Vẫn tiếp tục thử lấy members vì đôi khi get_chat lỗi nhưng get_chat_members vẫn được?
                        # Thực tế nếu get_chat lỗi thì get_chat_members cũng sẽ lỗi. Nhưng ta vẫn cho code chạy tiếp qua try/except.

                    db = SessionLocal()
                    try:
                        # Xóa dữ liệu cũ của nhóm này trong DB
                        old_records = db.query(TelegramProjectMember).filter(
                            TelegramProjectMember.chat_id == chat_id
                        ).all()
                        for old_record in old_records:
                            db.delete(old_record)
                        
                        if old_records:
                            print(f"  🗑 Đã xóa {len(old_records)} bản ghi cũ.")

                        # Lấy thành viên
                        synced_count = 0
                        admin_count = 0
                        member_count = 0
                    
                        owner_names = []
                        admin_names = []
                        member_names = []

                        async for member in bot.get_chat_members(int(chat_id)):
                            if not member.user or member.user.is_bot:
                                continue

                            user_id = str(member.user.id)
                            username = member.user.username
                            full_name = f"{member.user.first_name or ''} {member.user.last_name or ''}".strip()
                            status = member.status

                            role_type = "MEMBER"
                            slot_name = ""

                            if status == ChatMemberStatus.OWNER:
                                role_type = "OWNER"
                                slot_name = "owner"
                                owner_names.append(f"@{username}" if username else f"ID:{user_id}")
                            elif status == ChatMemberStatus.ADMINISTRATOR:
                                role_type = "ADMIN"
                                admin_count += 1
                                slot_name = f"admin_{admin_count:02d}"
                                admin_names.append(f"@{username}" if username else f"ID:{user_id}")
                            else:
                                role_type = "MEMBER"
                                member_count += 1
                                slot_name = f"member_{member_count:02d}"
                                member_names.append(f"@{username}" if username else f"ID:{user_id}")

                            new_member = TelegramProjectMember(
                                project_id=project_id,
                                chat_id=chat_id,
                                group_name=chat_title,
                                user_id=user_id,
                                user_name=username,
                                full_name=full_name,
                                role="member", # Vì đây là nhóm khách hàng, luôn là member role
                                slot_name=slot_name,
                                member_status=status.name if hasattr(status, 'name') else str(status),
                                is_bot=False,
                                custom_title="member_supplier",
                                parent_id=parent_id,
                                first_seen_at=datetime.datetime.now(),
                                last_seen_at=datetime.datetime.now()
                            )
                            db.add(new_member)
                            synced_count += 1

                        db.commit()
                        print(f"  ✅ Đã đồng bộ {synced_count} thành viên.")
                    except Exception as e:
                        db.rollback()
                        print(f"  ❌ Lỗi Database khi đồng bộ nhóm {chat_id}: {e}")
                    finally:
                        db.close()

                    # Gửi tin nhắn thông báo lên nhóm
                    from pyrogram.enums import ParseMode
                    msg_text = (
                        f"✅ <b>Đồng bộ hoàn tất!</b>\n\n"
                        f"Dự án: <b>Tiến Nga</b>\n"
                        f"Vai trò nhóm: <b>MEMBER</b>\n"
                        f"Custom title: <b>member_supplier</b>\n"
                        f"Nhóm Main: <code>{parent_id}</code>\n"
                        f"Tổng số: <b>{synced_count}</b> hội viên.\n"
                        f"- Chủ nhóm: {len(owner_names)} ({', '.join(owner_names)})\n"
                        f"- Quản trị viên: {admin_count} ({', '.join(admin_names)})\n"
                        f"- Thành viên: {member_count} ({', '.join(member_names)})\n"
                        f"🗑 <i>Đã xóa bản ghi cũ trước khi đồng bộ lại.</i>"
                    )
                    try:
                        await bot.send_message(int(chat_id), msg_text, parse_mode=ParseMode.HTML)
                    except Exception as e:
                        print(f"  ⚠️ Không thể gửi tin nhắn thông báo vào nhóm {chat_id}: {e}")

                except Exception as e:
                    print(f"  ❌ Lỗi khi xử lý nhóm {chat_id}: {e}")

                # Delay 2 giây giữa các nhóm theo yêu cầu để tránh bị nhầm là spam
                if i < len(groups):
                    print("  ⏳ Chờ 2 giây trước khi đồng bộ nhóm tiếp theo...")
                    await asyncio.sleep(2)

    except Exception as e:
        print(f"❌ Lỗi hệ thống: {e}")
    finally:
        print("🎉 Hoàn tất quá trình đồng bộ.")


if __name__ == "__main__":
    asyncio.run(main())

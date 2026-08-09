"""
Dựng file Excel danh sách nhân viên — dùng chung Tiến Nga và Ggomoosin.

Trước đây khối code này bị chép nguyên si ở tien_nga.py và ggomoosin.py.
"""

import datetime
import tempfile

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

EMPLOYEE_HEADERS = [
    "Mã nhân viên", "Họ và tên", "Giới tính", "Số điện thoại",
    "Username Telegram", "Nhóm nhân viên Telegram", "Địa chỉ",
    "CCCD", "Mức lương tháng", "Mức lương tuần", "Mức lương ngày",
    "Mức lương giờ", "Thời gian vào ca", "Thời gian kết thúc ca", "Công nợ",
]

EMPLOYEE_COL_WIDTHS = [15, 25, 12, 15, 20, 20, 30, 20, 18, 18, 18, 18, 20, 20, 18]

# Cột căn giữa / căn phải (1-based), còn lại căn trái
CENTER_COLS = {1, 3, 4, 8, 13, 14}
MONEY_COLS = {9, 10, 11, 12, 15}


def _fmt_money(val) -> str:
    if val is None or val == 0:
        return "0"
    return f"{int(val):,}".replace(",", ".")


def build_employee_excel(employees, project_name: str | None = None):
    """
    Ghi danh sách nhân viên ra file Excel tạm.

    Trả về (đường dẫn file tạm, tên file gửi đi, caption). Người gọi chịu trách
    nhiệm gửi file rồi xóa đường dẫn tạm.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DanhSachNhanVien"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    money_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    alt_fill = PatternFill("solid", fgColor="D9E2F3")

    for col_idx, header in enumerate(EMPLOYEE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, emp in enumerate(employees, 1):
        row = idx + 1
        row_fill = alt_fill if idx % 2 == 0 else None
        full_name = f"{emp.last_name or ''} {emp.first_name or ''}".strip()

        values = [
            emp.id,
            full_name,
            emp.gender or "—",
            emp.number_phone or "—",
            emp.username or "—",
            emp.telegram_group or "—",
            emp.address or "—",
            emp.identity_card or "—",
            _fmt_money(emp.monthly_salary),
            _fmt_money(emp.weekly_salary),
            _fmt_money(emp.daily_salary),
            _fmt_money(emp.hourly_salary),
            emp.start_time.strftime("%H:%M %d/%m/%Y") if emp.start_time else "—",
            emp.end_time.strftime("%H:%M %d/%m/%Y") if emp.end_time else "—",
            _fmt_money(emp.total_debt),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in CENTER_COLS:
                cell.alignment = center_align
            elif col_idx in MONEY_COLS:
                cell.alignment = money_align
            else:
                cell.alignment = left_align
            if row_fill:
                cell.fill = row_fill

    for col_idx, width in enumerate(EMPLOYEE_COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        temp_path = tmp.name
    wb.save(temp_path)

    now = datetime.datetime.now()
    title = f"DANH SÁCH NHÂN VIÊN{f' — {project_name}' if project_name else ''}"
    caption = (
        f"<b>{title}</b>\n\n"
        f"Tổng cộng: <b>{len(employees)}</b> nhân viên\n"
        f"<i>Xuất lúc: {now.strftime('%H:%M %d/%m/%Y')}</i>"
    )
    return temp_path, f"danh_sach_nhan_vien_{now.strftime('%Y%m%d_%H%M')}.xlsx", caption

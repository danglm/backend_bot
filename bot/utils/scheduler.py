import asyncio
import datetime
import io
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.db.session import SessionLocal
from app.models.telegram import TelegramProjectMember
from app.models.employee import Employee
from app.models.finance import Salary, Attendance
from bot.utils.logger import LogInfo, LogError, LogType
from pyrogram.enums import ParseMode
from bot.utils.enums import CustomTitle
from bot.utils.bot import bot
from bot.core.config import settings

# ──────────────────────────────────────────────────────────────────────────────
# Helper: build an Excel workbook for one month's attendance
# ──────────────────────────────────────────────────────────────────────────────
def _build_attendance_excel(
    employee_records: list[dict],
    year: int,
    month: int,
) -> io.BytesIO:
    """
    Build an Excel workbook with one sheet per employee.
    Each sheet contains the daily attendance rows for the given month.

    employee_records: list of dicts with keys:
        full_name, employee_id, rows (list of Attendance ORM objects)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Colour palette
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
    SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")  # medium blue
    ALT_FILL = PatternFill("solid", fgColor="D6E4F0")       # light blue
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

    thin = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    month_name = f"Tháng {month:02d}/{year}"
    num_days = calendar.monthrange(year, month)[1]
    day_names_vn = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "CN"]

    col_headers = [
        "Ngày", "Thứ", "Giờ vào", "Giờ ra", "Giờ làm việc (h)",
        "Đi muộn (phút)", "Tăng ca (h)", "Nghỉ phép năm", "Ghi chú",
    ]

    for record in employee_records:
        full_name = record["full_name"]
        attendance_map: dict[int, Attendance] = {a.day: a for a in record["rows"]}

        # Sheet name limited to 31 chars (Excel limit)
        ws = wb.create_sheet(title=full_name[:31])

        # ── Title row ─────────────────────────────────────────────────
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = f"BẢNG CHẤM CÔNG - {full_name.upper()} - {month_name.upper()}"
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = HEADER_FILL
        ws.row_dimensions[1].height = 28

        # ── Column-header row ─────────────────────────────────────────
        for col_idx, hdr in enumerate(col_headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = SUBHEADER_FILL
            cell.border = thin_border
        ws.row_dimensions[2].height = 22

        # ── Data rows ─────────────────────────────────────────────────
        total_working = 0.0
        total_late = 0.0
        total_overtime = 0.0
        total_leave_days = 0 

        for day in range(1, num_days + 1):
            row_num = day + 2  # data starts at row 3
            date_obj = datetime.date(year, month, day)
            weekday_vn = day_names_vn[date_obj.weekday()]  # Mon=0 … Sun=6

            att: Attendance | None = attendance_map.get(day)

            if att:
                check_in = att.check_in_time.strftime("%H:%M") if att.check_in_time else "—"
                check_out = att.check_out_time.strftime("%H:%M") if att.check_out_time else "—"
                working = round(att.working_time or 0, 2)
                late = round(att.late_time or 0, 2)
                overtime = round(att.overtime or 0, 2)
                
                is_annual_leave = ""
                if att.error and "nghỉ phép năm" in att.error.lower():
                    is_annual_leave = "X"
                    total_leave_days += 1
                
                note = att.error or ""
            else:
                # Weekend = CN / Thứ 7 - mark as "Nghỉ"
                if date_obj.weekday() >= 5:
                    check_in = check_out = "—"
                    working = late = overtime = 0.0
                    is_annual_leave = ""
                    note = "Nghỉ cuối tuần"
                else:
                    check_in = check_out = "—"
                    working = late = overtime = 0.0
                    is_annual_leave = ""
                    note = "Không có dữ liệu"

            total_working += working
            total_late += late
            total_overtime += overtime

            row_fill = ALT_FILL if day % 2 == 0 else WHITE_FILL
            values = [day, weekday_vn, check_in, check_out, working, late, overtime, is_annual_leave, note]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = row_fill
                cell.border = thin_border

        # ── Summary row ───────────────────────────────────────────────
        summary_row = num_days + 3
        ws.merge_cells(f"A{summary_row}:B{summary_row}")
        sum_label = ws.cell(row=summary_row, column=1, value="TỔNG CỘNG")
        sum_label.font = Font(bold=True, color="FFFFFF")
        sum_label.alignment = Alignment(horizontal="center", vertical="center")
        sum_label.fill = HEADER_FILL
        sum_label.border = thin_border

        for col_idx, val in [
            (5, round(total_working, 2)), 
            (6, round(total_late, 2)), 
            (7, round(total_overtime, 2)),
            (8, total_leave_days)
        ]:
            cell = ws.cell(row=summary_row, column=col_idx, value=val)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = thin_border

        # ── Column widths ─────────────────────────────────────────────
        col_widths = [8, 10, 12, 12, 20, 18, 16, 18, 30]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Save to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ──────────────────────────────────────────────────────────────────────────────
# Helper: build an Excel workbook for one employee's payroll
# ──────────────────────────────────────────────────────────────────────────────
def _build_payroll_excel(data: dict) -> io.BytesIO:
    """
    Build a Payroll Excel file for a single employee.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"

    # Style
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Mã nhân viên", "Tên nhân viên", "Tháng", "Năm", "Số ngày nghỉ phép", 
        "Nghỉ không phép", "Số phép năm còn lại", "Lương tháng", "Số ngày công", 
        "Thời gian tăng ca (h)", "Công tiêu chuẩn", "Thành tiền lương", 
        "Lương tăng ca", "Phạt", "Tổng lương"
    ]

    # Header Row
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = HEADER_FILL
        cell.border = thin_border
    
    # Data Row
    values = [
        data.get("employee_id"), data.get("full_name"), data.get("month"), data.get("year"),
        data.get("leave_days"), data.get("unpaid_leave"), data.get("leave_balance"),
        data.get("base_salary"), data.get("actual_working_days"), data.get("total_overtime"),
        data.get("standard_working_days"), data.get("salary_earned"), data.get("overtime_salary_earned"),
        data.get("penalty"), data.get("total_net_salary")
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Adjust widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def get_payroll_data(db: Session, employee_id: str, month: int, year: int) -> dict:
    """
    Gather and calculate payroll data for an employee.
    """
    from app.models.employee import Employee
    from app.models.finance import Salary, Attendance
    
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    salary = db.query(Salary).filter(Salary.employee_id == employee_id).first()
    attendances = db.query(Attendance).filter(
        Attendance.employee_id == employee_id,
        Attendance.month == month,
        Attendance.year == year
    ).all()
    
    if not employee: return {}
    
    full_name = f"{employee.last_name} {employee.first_name}".strip()
    
    # Calculate Standard Working Days (Mon-Fri)
    num_days = calendar.monthrange(year, month)[1]
    standard_days = 0
    for day in range(1, num_days + 1):
        if datetime.date(year, month, day).weekday() < 5: # 0-4 is Mon-Fri
            standard_days += 1
            
    # Aggregate Attendance
    actual_working_days = 0
    total_overtime = 0.0
    leave_days = 0
    unpaid_leave = 0
    
    attended_days = {a.day: a for a in attendances}
    for day in range(1, num_days + 1):
        is_weekday = datetime.date(year, month, day).weekday() < 5
        att = attended_days.get(day)
        
        if att:
            if (att.working_time or 0) > 0:
                actual_working_days += 1
            total_overtime += (att.overtime or 0.0)
            if att.error and "nghỉ phép năm" in att.error.lower():
                leave_days += 1
        elif is_weekday:
            # Weekday but no attendance record = Unpaid leave?
            unpaid_leave += 1
            
    base_salary = salary.monthly_salary if salary else None
    leave_balance = salary.leave_balance if salary else None
    
    # Simple Calculations (Null if missing base_salary)
    salary_earned = None
    overtime_salary_earned = None
    total_net_salary = None
    
    if base_salary and standard_days > 0:
        daily_rate = base_salary / standard_days
        salary_earned = round(daily_rate * actual_working_days, 2)
        
        # Overtime calculation: Assuming 1.5x hourly rate if overtime_salary not explicitly an hourly rate
        # Hourly rate = daily_rate / 8
        hourly_rate = daily_rate / 8
        overtime_rate = salary.overtime_salary if salary and salary.overtime_salary else (hourly_rate * 1.5)
        overtime_salary_earned = round(total_overtime * overtime_rate, 2)
        
        # Penalty (Placeholder for now)
        penalty = 0 # Or extract from Salary if exists
        
        total_net_salary = round(salary_earned + overtime_salary_earned - penalty, 2)

    return {
        "employee_id": employee_id,
        "full_name": full_name,
        "month": month,
        "year": year,
        "leave_days": leave_days,
        "unpaid_leave": unpaid_leave,
        "leave_balance": leave_balance,
        "base_salary": base_salary,
        "actual_working_days": actual_working_days,
        "total_overtime": round(total_overtime, 2),
        "standard_working_days": standard_days,
        "salary_earned": salary_earned,
        "overtime_salary_earned": overtime_salary_earned,
        "penalty": 0,
        "total_net_salary": total_net_salary
    }

async def generate_and_send_attendance_report(
    db, 
    employee: "Employee",
    report_month: int, 
    report_year: int,
) -> bool:
    """
    Generate an attendance Excel report for a single employee
    and send it to the MAIN (management) group of the employee's project.
    """
    from app.models.finance import Attendance
    from app.models.telegram import TelegramProjectMember
    
    try:
        if not employee.telegram_group:
            LogInfo(f"Employee {employee.id} has no telegram_group, skipping.", LogType.SYSTEM_STATUS)
            return False

        full_name = f"{employee.last_name} {employee.first_name}".strip() or employee.username or employee.id

        rows = db.query(Attendance).filter(
            Attendance.employee_id == employee.id,
            Attendance.year == report_year,
            Attendance.month == report_month,
        ).order_by(Attendance.day).all()

        if not rows:
            LogInfo(f"No attendance data for {employee.id} in {report_month:02d}/{report_year}.", LogType.SYSTEM_STATUS)
            return False

        # --- Resolve main group from employee's telegram_group ---
        tg_group = employee.telegram_group.strip()
        
        # Find the member group in telegram_project_members
        tpm = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.group_name == tg_group
        ).first()
        
        if not tpm:
            # Try as chat_id directly
            tpm = db.query(TelegramProjectMember).filter(
                TelegramProjectMember.chat_id == tg_group
            ).first()
        
        if not tpm or not tpm.project_id:
            LogInfo(f"Employee {employee.id}: cannot find project for telegram_group '{tg_group}', skipping.", LogType.SYSTEM_STATUS)
            return False

        # Find the MAIN group in the same project
        main_tpm = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == tpm.project_id,
            TelegramProjectMember.role == "main"
        ).first()

        if not main_tpm:
            LogInfo(f"Employee {employee.id}: no main group found for project {tpm.project_id}, skipping.", LogType.SYSTEM_STATUS)
            return False

        try:
            main_chat_id = int(main_tpm.chat_id)
        except (ValueError, TypeError):
            LogInfo(f"Employee {employee.id}: main group chat_id '{main_tpm.chat_id}' is invalid, skipping.", LogType.SYSTEM_STATUS)
            return False

        LogInfo(f"Generating attendance report for {employee.id} ({full_name}) -> main group {main_chat_id}.", LogType.SYSTEM_STATUS)

        # Build report
        employee_records = [{
            "full_name": full_name,
            "employee_id": employee.id,
            "rows": rows,
        }]

        buf = _build_attendance_excel(employee_records, report_year, report_month)
        
        safe_name = full_name.replace(" ", "_")
        filename = f"{safe_name}_BangChamCong_{report_month:02d}_{report_year}.xlsx"
        
        caption = (
            f"<b>BẢNG CHẤM CÔNG THÁNG {report_month:02d}/{report_year}</b>\n\n"
            f"<b>Nhân viên:</b> {full_name}\n"
            f"<b>Mã NV:</b> <code>{employee.id}</code>\n"
            f"<b>Nhóm:</b> {tg_group}\n"
            f"<b>Số ngày có dữ liệu:</b> {len(rows)} ngày\n\n"
            f"<i>Vui lòng kiểm tra file Excel đính kèm.</i>\n"
            f"<i>Sau khi xác nhận, nhập lệnh <code>/tien_nga_export_payroll {employee.id} {report_month:02d}/{report_year}</code> để xuất bảng lương.</i>"
        )

        from bot.utils.bot import bot
        client = bot

        # Send to MAIN group
        info_msg = await client.send_message(
            chat_id=main_chat_id,
            text=caption,
            parse_mode=ParseMode.HTML
        )
        
        await client.send_document(
            chat_id=main_chat_id,
            document=buf,
            file_name=filename,
            reply_to_message_id=info_msg.id
        )
        LogInfo(f"Sent attendance report for {employee.id} to main group {main_chat_id}.", LogType.SYSTEM_STATUS)
        return True
            
    except Exception as e:
        LogError(f"Error in generate_and_send_attendance_report for {employee.id}: {e}", LogType.SYSTEM_STATUS)
    return False


async def monthly_attendance_report_worker():
    """
    Background worker that generates monthly attendance reports.
    Queries Employee records directly using their telegram_group field.
    """
    LogInfo("Monthly attendance report worker started.", LogType.SYSTEM_STATUS)
    
    last_sent_month = None
    
    while True:
        try:
            now = datetime.datetime.now()

            # Trigger condition: 1st of month at 08:00
            cfg = settings.SCHEDULER_MONTHLY_REPORT
            if now.day == cfg.get('day', 1) and now.hour == cfg.get('hour', 8) and now.minute == cfg.get('minute', 0):
                report_month = now.month - 1 if now.month > 1 else 12
                report_year = (now.year if now.month > 1 else now.year - 1)

                if last_sent_month == report_month:
                    await asyncio.sleep(600)
                    continue

                LogInfo(f"Generating monthly reports for period {report_month:02d}/{report_year}...", LogType.SYSTEM_STATUS)

                db = SessionLocal()
                try:
                    from app.models.employee import Employee
                    
                    # Find all active employees with telegram_group set
                    employees = db.query(Employee).filter(
                        Employee.telegram_group != None,
                        Employee.telegram_group != "",
                        Employee.status != "inactive"
                    ).all()
                    
                    LogInfo(f"Found {len(employees)} employees with telegram_group.", LogType.SYSTEM_STATUS)
                    
                    for emp in employees:
                        await generate_and_send_attendance_report(
                            db, emp, report_month, report_year
                        )

                    last_sent_month = report_month
                except Exception as e:
                    LogError(f"Error in monthly_attendance_report_worker processing: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()

            # Wait before checking again
            await asyncio.sleep(60)
        except Exception as e:
            LogError(f"Error in monthly_attendance_report_worker loop: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)

async def checkin_reminder_worker():
    """
    Background worker that sends reminders to staff who haven't checked in.
    Reminders are sent every 15 minutes during the 30-minute window after their start time.
    """
    LogInfo("Check-in reminder worker started.", LogType.SYSTEM_STATUS)
    while True:
        try:
            now = datetime.datetime.now()
            
            # Wait for userbot client to be ready
            client = bot
            if not client.is_connected:
                await asyncio.sleep(30)
                continue
            
            db = SessionLocal()
            try:
                from app.models.employee import Employee

                # Fetch all employees who have a configured telegram_group
                employees = db.query(Employee).filter(
                    Employee.telegram_group != None,
                    Employee.telegram_group != ""
                ).all()

                if not employees:
                    db.close()
                    await asyncio.sleep(60)
                    continue

                # Group employees by their configured telegram_group
                groups = {}
                for emp in employees:
                    if not emp.username:
                        continue
                    gn = emp.telegram_group.strip()
                    if not gn:
                        continue
                    if gn not in groups:
                        groups[gn] = []
                    groups[gn].append(emp)

                for chat_id_str, group_emps in groups.items():
                    chat_id = chat_id_str
                    try:
                        chat_id = int(chat_id_str)
                    except ValueError:
                        pass
                        
                    late_checkin_mentions = []
                    late_checkout_mentions = []
                    
                    try:
                        for emp in group_emps:
                            username = emp.username
                            employee_id = emp.id
                            
                            # Get their scheduled times
                            salary_info = db.query(Salary).filter(Salary.employee_id == employee_id).first()
                            work_start = salary_info.start_time if salary_info and salary_info.start_time else datetime.time(8, 0)
                            work_end = salary_info.end_time if salary_info and salary_info.end_time else datetime.time(17, 30)
                            
                            # Calculate time differences in minutes
                            now_time = now.time()
                            now_total_min = now_time.hour * 60 + now_time.minute
                            
                            # --- Check-in Reminder Logic ---
                            start_total_min = work_start.hour * 60 + work_start.minute
                            diff_in = now_total_min - start_total_min
                            
                            if diff_in > 0 and diff_in <= 30 and diff_in % 15 == 0:
                                # Check if they already checked in today
                                att = db.query(Attendance).filter(
                                    Attendance.employee_id == employee_id,
                                    Attendance.year == now.year,
                                    Attendance.month == now.month,
                                    Attendance.day == now.day
                                ).first()
                                
                                if not att or not att.check_in_time:
                                    late_checkin_mentions.append(f"@{username}")
                            
                            # --- Check-out Reminder Logic ---
                            end_total_min = work_end.hour * 60 + work_end.minute
                            diff_out = now_total_min - end_total_min
                            
                            if diff_out > 0 and diff_out <= 30 and diff_out % 10 == 0:
                                # Check if they checked in but haven't checked out yet
                                att = db.query(Attendance).filter(
                                    Attendance.employee_id == employee_id,
                                    Attendance.year == now.year,
                                    Attendance.month == now.month,
                                    Attendance.day == now.day
                                ).first()
                                
                                if att and att.check_in_time and not att.check_out_time:
                                    late_checkout_mentions.append(f"@{username}")
                        
                        # Send Check-in Reminders
                        if late_checkin_mentions:
                            mentions_str = " ".join(late_checkin_mentions)
                            reminder_text = (
                                f"🔔 <b>NHẮC NHỞ CHẤM CÔNG (CHECK-IN)</b>\n\n"
                                f"Các bạn {mentions_str} ơi, vui lòng <b>Check-in</b> để ghi nhận ngày công nhé!\n"
                                f"⏰ Bây giờ là: <code>{now.strftime('%H:%M')}</code>"
                            )
                            client = bot
                            await client.send_message(chat_id=chat_id, text=reminder_text, parse_mode=ParseMode.HTML)
                            LogInfo(f"Sent check-in reminder to {chat_id} for {len(late_checkin_mentions)} users.", LogType.SYSTEM_STATUS)

                        # Send Check-out Reminders
                        if late_checkout_mentions:
                            mentions_str = " ".join(late_checkout_mentions)
                            reminder_text = (
                                f"🔔 <b>NHẮC NHỞ CHẤM CÔNG (CHECK-OUT)</b>\n\n"
                                f"Các bạn {mentions_str} ơi, đã đến giờ nghỉ rồi. Vui lòng <b>Check-out</b> trước khi về nhé!\n"
                                f"⏰ Bây giờ là: <code>{now.strftime('%H:%M')}</code>"
                            )
                            client = bot
                            await client.send_message(chat_id=chat_id, text=reminder_text, parse_mode=ParseMode.HTML)
                            LogInfo(f"Sent check-out reminder to {chat_id} for {len(late_checkout_mentions)} users.", LogType.SYSTEM_STATUS)

                    except Exception as e:
                        if "CHAT_ID_INVALID" in str(e) or "PEER_ID_INVALID" in str(e) or "USERNAME_INVALID" in str(e):
                            pass # Skip groups the bot is no longer in or invalid usernames
                        else:
                            LogError(f"Error checking group {chat_id}: {e}", LogType.SYSTEM_STATUS)

            except Exception as e:
                LogError(f"Error in reminder worker loop: {e}", LogType.SYSTEM_STATUS)
            finally:
                db.close()

            # Wait exactly until the next minute
            next_run = (datetime.datetime.now() + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
            sleep_time = (next_run - datetime.datetime.now()).total_seconds()
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
            else:
                await asyncio.sleep(1) # Safety

        except Exception as e:
            LogError(f"Critical error in checkin_reminder_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)

async def recurring_task_worker():
    """
    Background worker that checks for completed tasks with a recurring cycle
    and restarts them at the right time based on updated_at.
    """
    LogInfo("Recurring task worker started.", LogType.SYSTEM_STATUS)
    
    def add_months(sourcedate, months):
        month = sourcedate.month - 1 + months
        year = sourcedate.year + month // 12
        month = month % 12 + 1
        day = min(sourcedate.day, calendar.monthrange(year, month)[1])
        return datetime.date(year, month, day)

    while True:
        try:
            now = datetime.datetime.now()
            
            cfg = settings.SCHEDULER_RESTART_TASK
            if now.hour == cfg.get('hour', 8) and now.minute == cfg.get('minute', 0):
                db = SessionLocal()
                try:
                    from app.models.task import Task
                    from bot.utils.bot import bot
    
                    tasks = db.query(Task).filter(
                        Task.status != "CANCELLED",
                        Task.cycle != None,
                        Task.cycle != "Một lần",
                        ~Task.cycle.ilike("%(Đã lặp)%")
                    ).all()
                    
                    client = bot
                    
                    for task in tasks:
                        if not task.updated_at:
                            continue
                            
                        cycle_lower = task.cycle.lower()
                        ref_date = task.updated_at.date()
                        
                        should_trigger = False
                        new_start = None
                        new_end = None
                        
                        if "1 tháng" in cycle_lower or "hằng tháng" in cycle_lower:
                            target_date = add_months(ref_date, 1)
                            should_trigger = True
                            new_start = target_date.strftime("%d/%m/%Y")
                            try:
                                e_date = datetime.datetime.strptime(task.end_date, "%d/%m/%Y").date() if task.end_date else target_date
                                new_end = add_months(e_date, 1).strftime("%d/%m/%Y")
                            except Exception:
                                new_end = new_start
                                
                        elif "hằng tuần" in cycle_lower or "1 tuần" in cycle_lower:
                            target_date = ref_date + datetime.timedelta(days=7)
                            should_trigger = True
                            new_start = target_date.strftime("%d/%m/%Y")
                            try:
                                e_date = datetime.datetime.strptime(task.end_date, "%d/%m/%Y").date() if task.end_date else target_date
                                new_end = (e_date + datetime.timedelta(days=7)).strftime("%d/%m/%Y")
                            except Exception:
                                new_end = new_start
                                
                        elif "hằng ngày" in cycle_lower or "1 ngày" in cycle_lower:
                            target_date = ref_date + datetime.timedelta(days=1)
                            should_trigger = True
                            new_start = target_date.strftime("%d/%m/%Y")
                            try:
                                e_date = datetime.datetime.strptime(task.end_date, "%d/%m/%Y").date() if task.end_date else target_date
                                new_end = (e_date + datetime.timedelta(days=1)).strftime("%d/%m/%Y")
                            except Exception:
                                new_end = new_start
                    
                    if should_trigger:
                        # Create new task
                        new_task = Task(
                            employee_id=task.employee_id,
                            project_id=task.project_id,
                            group_chat_id=task.group_chat_id,
                            assigner=task.assigner,
                            assignee=task.assignee,
                            content=task.content,
                            start_date=new_start,
                            end_date=new_end,
                            cycle=task.cycle,
                            status="PENDING"
                        )
                        db.add(new_task)
                        
                        # Mark old task as restarted
                        task.cycle = f"{task.cycle} (Đã lặp)"
                        db.add(task)
                        
                        db.commit()
                        db.refresh(new_task)
                        
                        # Send notification
                        try:
                            from app.models.telegram import TelegramProjectMember
                            from bot.utils.enums import CustomTitle
                            
                            # Find MEMBER_HR group
                            hr_member = db.query(TelegramProjectMember).filter(
                                TelegramProjectMember.project_id == new_task.project_id,
                                TelegramProjectMember.role == "member",
                                TelegramProjectMember.custom_title == CustomTitle.MEMBER_HR.value
                            ).first()
                            
                            if hr_member and hr_member.chat_id:
                                target_chat_id = int(hr_member.chat_id)
                                new_task.group_chat_id = str(hr_member.chat_id)
                            else:
                                target_chat_id = int(new_task.group_chat_id)

                            msg_text = (
                                f"<b>CÔNG VIỆC ĐƯỢC TÁI KHỞI ĐỘNG KỲ MỚI</b>\n\n"
                                f"<b>Ngày bắt đầu:</b> {new_start}\n"
                                f"<b>Ngày kết thúc:</b> {new_end}\n"
                                f"<b>Chu kỳ:</b> {new_task.cycle}\n"
                                f"<b>Nhân viên:</b> {new_task.assignee}\n"
                                f"<b>Nội dung công việc:</b> {new_task.content}\n\n"
                                f"<i>Task đã được tái khởi động tự động theo chu kỳ.</i>"
                            )
                            sent_msg = await client.send_message(
                                chat_id=target_chat_id,
                                text=msg_text,
                                parse_mode=ParseMode.HTML
                            )
                            
                            new_task.message_id = sent_msg.id
                            db.commit()
                            LogInfo(f"Restarted task {task.id} -> new task {new_task.id} in {target_chat_id}", LogType.SYSTEM_STATUS)
                        except Exception as e:
                            LogError(f"Error sending notification for restarted task {task.id}: {e}", LogType.SYSTEM_STATUS)

                except Exception as e:
                    LogError(f"Error checking recurring tasks: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()
            
            # Wait exactly until the next minute
            next_run = (datetime.datetime.now() + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
            sleep_time = (next_run - datetime.datetime.now()).total_seconds()
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
            else:
                await asyncio.sleep(1) # Safety

        except Exception as e:
            LogError(f"Critical error in recurring_task_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)

async def bad_debt_notification_worker():
    """
    Background worker that checks for overdue contracts daily at 08:00 AM.
    If a contract is ACTIVE, has due_date < today - 7 days, and remaining_principal > 0:
    It finds the project main group and sends an alert.
    """
    LogInfo("Bad debt notification worker started.", LogType.SYSTEM_STATUS)
    
    last_sent_date = None
    
    while True:
        try:
            now = datetime.datetime.now()
            
            # Run at 08:00 system time
            cfg = settings.SCHEDULER_BAD_DEBT
            if now.hour == cfg.get('hour', 8) and now.minute == cfg.get('minute', 0):
                current_date = now.date()
                if last_sent_date == current_date:
                    await asyncio.sleep(60)
                    continue

                LogInfo("Checking for bad debts...", LogType.SYSTEM_STATUS)
                db = SessionLocal()
                try:
                    from app.models.credit import Credit, CreditCustomer, CreditStatus
                    from app.models.telegram import TelegramProjectMember
                    from bot.utils.bot import bot

                    
                    # Criteria: ACTIVE status, remaining > 0, due_date < today - 7
                    seven_days_ago = current_date - datetime.timedelta(days=7)
                    
                    overdue_contracts = db.query(Credit).filter(
                        Credit.credit_status == CreditStatus.ACTIVE.value,
                        Credit.remaining_principal > 0,
                        Credit.due_date <= seven_days_ago
                    ).all()
                    
                    if overdue_contracts:
                        client = bot
                        
                        for contract in overdue_contracts:
                            customer = contract.customer
                            if not customer:
                                continue
                                
                            from app.models.business import Projects
                            credit_project = db.query(Projects).filter(Projects.project_name == "Credit").first()
                            if not credit_project:
                                continue
                                
                            member_project_links = db.query(TelegramProjectMember).filter(
                                TelegramProjectMember.project_id == credit_project.id,
                                TelegramProjectMember.role == "member",
                                TelegramProjectMember.slot_name.like("member%")
                            ).all()

                            LogInfo(f"Found {len(member_project_links)} member project links.", LogType.SYSTEM_STATUS)
                            
                            target_project_id = None
                            parent_main_chat_id = None
                            for link in member_project_links:
                                if customer.group_name and link.group_name == customer.group_name:
                                    target_project_id = link.project_id
                                    parent_main_chat_id = link.parent_id
                                    break
                                    
                            if not target_project_id:
                                continue
                                
                            main_group = None
                            if parent_main_chat_id:
                                main_group = db.query(TelegramProjectMember).filter(
                                    TelegramProjectMember.project_id == target_project_id,
                                    TelegramProjectMember.role == "main",
                                    TelegramProjectMember.chat_id == parent_main_chat_id
                                ).first()
                                
                            if not main_group:
                                main_group = db.query(TelegramProjectMember).filter(
                                    TelegramProjectMember.project_id == target_project_id,
                                    TelegramProjectMember.role == "main"
                                ).first()
                                
                            if main_group:
                                days_overdue = (current_date - contract.due_date).days
                                
                                def fmt_num(val):
                                    if val is None: return 0
                                    return int(val) if val == int(val) else val
                                
                                msg_text = (
                                    f"🚨 <b>CẢNH BÁO NỢ XẤU (QUÁ HẠN {days_overdue} NGÀY)</b> 🚨\n\n"
                                    f"<b>Mã Hợp Đồng:</b> <code>{contract.contract_id}</code>\n"
                                    f"<b>Khách hàng:</b> {customer.customer_name}\n"
                                    f"<b>Mã Khách Hàng:</b> {customer.customer_id or 'N/A'}\n"
                                    f"<b>Liên hệ:</b> {customer.contact_info}\n"
                                    f"<b>Còn nợ gốc:</b> {fmt_num(contract.remaining_principal):,} VND\n"
                                    f"<b>Ngày đáo hạn gốc:</b> {contract.due_date.strftime('%d/%m/%Y')}\n\n"
                                    f"<i>=> Admin/Owner hãy <b>Reply</b> tin nhắn này với lệnh <code>/bad_debt</code> để đưa vào BLACKLIST!</i>"
                                )
                                
                                try:
                                    await client.send_message(
                                        chat_id=int(main_group.chat_id),
                                        text=msg_text,
                                        parse_mode=ParseMode.HTML
                                    )
                                    LogInfo(f"Sent bad debt alert for {contract.contract_id} to chat {main_group.chat_id}", LogType.SYSTEM_STATUS)
                                except Exception as e:
                                    LogError(f"Failed to send bad debt alert to {main_group.chat_id}: {e}", LogType.SYSTEM_STATUS)

                    last_sent_date = current_date

                except Exception as e:
                    LogError(f"Error checking bad debts: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()
                    
                await asyncio.sleep(61)
            
            next_run = (datetime.datetime.now() + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
            sleep_time = (next_run - datetime.datetime.now()).total_seconds()
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
            else:
                await asyncio.sleep(1)

        except Exception as e:
            LogError(f"Critical error in bad_debt_notification_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)

async def interest_payment_notification_worker():
    """
    Background worker that checks for due interest payments daily at 08:00 AM.
    If a contract is ACTIVE, remaining_principal > 0, and today's day matches the interest_start_date's day:
    It finds the member group and sends an alert for them to pay interest.
    """
    LogInfo("Interest payment notification worker started.", LogType.SYSTEM_STATUS)
    
    last_sent_date = None
    
    while True:
        try:
            now = datetime.datetime.now()
            
            # Run at 08:00 system time
            cfg = settings.SCHEDULER_INTEREST
            if now.hour == cfg.get('hour', 8) and now.minute == cfg.get('minute', 0):
                current_date = now.date()
                if last_sent_date == current_date:
                    await asyncio.sleep(60)
                    continue

                LogInfo("Checking for interest payments due today...", LogType.SYSTEM_STATUS)
                db = SessionLocal()
                try:
                    from app.models.credit import Credit, CreditCustomer, CreditStatus
                    from app.models.telegram import TelegramProjectMember
                    from bot.utils.bot import bot

                    
                    active_contracts = db.query(Credit).filter(
                        Credit.credit_status == CreditStatus.ACTIVE.value,
                        Credit.remaining_principal > 0,
                        Credit.interest_start_date != None
                    ).all()
                    
                    if active_contracts:
                        client = bot
                        from sqlalchemy import extract
                        from app.models.credit import CreditInterest
                        
                        for contract in active_contracts:
                            interest_day = contract.interest_start_date.day
                            
                            # Calculate the due_date for the current cycle
                            if current_date.day >= interest_day:
                                due_year, due_month = current_date.year, current_date.month
                            else:
                                due_year, due_month = (current_date.year, current_date.month - 1) if current_date.month > 1 else (current_date.year - 1, 12)
                            
                            try:
                                due_date = datetime.date(due_year, due_month, interest_day)
                            except ValueError:
                                due_date = datetime.date(due_year, due_month, calendar.monthrange(due_year, due_month)[1])
                                
                            days_late = (current_date - due_date).days
                            
                            if days_late < 0 or days_late > 7:
                                continue
                                
                            # Check if we should charge interest for this cycle
                            int_rate = contract.monthly_interest_rate or 0
                            int_amt = contract.monthly_interest_amount or 0
                            if int_amt == 0 and int_rate > 0:
                                int_amt = (contract.remaining_principal * int_rate) / 100
                                
                            if contract.last_interest_charged_date is None or contract.last_interest_charged_date < due_date:
                                contract.interest_debt = (contract.interest_debt or 0) + int_amt
                                contract.last_interest_charged_date = due_date
                                db.commit()
                                
                            if contract.interest_debt is not None and contract.interest_debt <= 0:
                                continue
                                
                            customer = contract.customer
                            if not customer:
                                continue
                                
                            from app.models.business import Projects
                            credit_project = db.query(Projects).filter(Projects.project_name == "Credit").first()
                            if not credit_project:
                                continue

                            # Find the target project and member group using target customer contact info
                            customer_links = db.query(TelegramProjectMember).filter(
                                TelegramProjectMember.project_id == credit_project.id,
                                TelegramProjectMember.role == "member",
                                TelegramProjectMember.slot_name.like("member%")
                            ).all()
                            
                            target_project_id = None
                            member_chat_id = None
                            parent_main_chat_id = None
                            for link in customer_links:
                                if customer.group_name and link.group_name == customer.group_name:
                                    target_project_id = link.project_id
                                    member_chat_id = link.chat_id
                                    parent_main_chat_id = link.parent_id
                                    break
                                    
                            if days_late == 7:
                                # Overdue! Mark as BAD DEBT
                                contract.credit_status = CreditStatus.BAD_DEBT.value
                                if contract.notes:
                                    if "[BLACKLIST]" not in contract.notes:
                                        contract.notes = f"[BLACKLIST] {contract.notes}"
                                else:
                                    contract.notes = "[BLACKLIST]"
                                    
                                db.commit()
                                
                                # Notify main group
                                if target_project_id:
                                    main_group = None
                                    if parent_main_chat_id:
                                        main_group = db.query(TelegramProjectMember).filter(
                                            TelegramProjectMember.project_id == target_project_id,
                                            TelegramProjectMember.role == "main",
                                            TelegramProjectMember.chat_id == parent_main_chat_id
                                        ).first()
                                        
                                    if not main_group:
                                        main_group = db.query(TelegramProjectMember).filter(
                                            TelegramProjectMember.project_id == target_project_id,
                                            TelegramProjectMember.role == "main"
                                        ).first()
                                    if main_group:
                                        msg_text = (
                                            f"🚨 <b>CẢNH BÁO NỢ XẤU TỰ ĐỘNG (QUÁ HẠN LÃI 7 NGÀY)</b> 🚨\n\n"
                                            f"<b>Mã Hợp Đồng:</b> <code>{contract.contract_id}</code>\n"
                                            f"<b>Khách hàng:</b> {customer.customer_name}\n"
                                            f"<b>Mã Khách Hàng:</b> {customer.customer_id or 'N/A'}\n"
                                            f"<b>Liên hệ:</b> {customer.contact_info}\n"
                                            f"<b>Trạng thái:</b> Hệ thống đã tự động chuyển khách hàng vào BLACKLIST NỢ XẤU!"
                                        )
                                        try:
                                            await client.send_message(
                                                chat_id=int(main_group.chat_id),
                                                text=msg_text,
                                                parse_mode=ParseMode.HTML
                                            )
                                            LogInfo(f"Sent auto bad debt alert for {contract.contract_id} to {main_group.chat_id}", LogType.SYSTEM_STATUS)
                                        except Exception as e:
                                            LogError(f"Failed to send auto bad debt alert to {main_group.chat_id}: {e}", LogType.SYSTEM_STATUS)
                                continue
                                    
                            if member_chat_id:
                                def fmt_num(val):
                                    if val is None: return 0
                                    return int(val) if val == int(val) else val
                                
                                principal = fmt_num(contract.remaining_principal)
                                current_interest_debt = fmt_num(contract.interest_debt or 0)
                                days_text = f"(Trễ hạn {days_late} ngày)" if days_late > 0 else "(Đến hạn hôm nay)"
                                
                                msg_text = (
                                    f"🔔 <b>THÔNG BÁO ĐÓNG TIỀN LÃI {days_text}</b> 🔔\n\n"
                                    f"<b>Khách hàng:</b> {customer.customer_name}\n"
                                    f"<b>Mã Khách Hàng:</b> {customer.customer_id or 'N/A'}\n"
                                    f"<b>Liên hệ:</b> {customer.contact_info}\n"
                                    f"<b>Mã Hợp Đồng:</b> <code>{contract.contract_id}</code>\n"
                                    f"<b>Số tiền lãi cần đóng:</b> <b>{current_interest_debt:,} VND</b>\n"
                                    f"---------------------------\n"
                                    f"<i>Quý khách vui lòng thanh toán đúng hạn. Cảm ơn Quý Khách Hàng.</i>\n"
                                    f"<i>Sau khi Quý Khách Hàng đã thanh toán, vui lòng gửi lại biên lai thanh toán để Admin được xác nhận.</i>\n"
                                    f"<i>Admin vui lòng nhập <pre>/credit_payment_confirmed [Số tiền thanh toán]</pre> để xác nhận thanh toán.</i>"
                                )
                                
                                try:
                                    await client.send_message(
                                        chat_id=int(member_chat_id),
                                        text=msg_text,
                                        parse_mode=ParseMode.HTML
                                    )
                                    LogInfo(f"Sent interest payment alert for contract {contract.contract_id} to {member_chat_id}", LogType.SYSTEM_STATUS)
                                except Exception as e:
                                    LogError(f"Failed to send interest alert: {e}", LogType.SYSTEM_STATUS)

                    last_sent_date = current_date

                except Exception as e:
                    LogError(f"Error checking interest payments: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()
                    
                await asyncio.sleep(61)
            
            next_run = (datetime.datetime.now() + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
            sleep_time = (next_run - datetime.datetime.now()).total_seconds()
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
            else:
                await asyncio.sleep(1)

        except Exception as e:
            LogError(f"Critical error in interest_payment_notification_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)

async def rental_payment_notification_worker():
    """
    Background worker that checks for rental payment due dates daily at 08:00 AM.
    Logic:
    - Day 0 (due day): Accumulate monthly_rental into rental_debt + send notification
    - Day 1-6: If rental_debt > 0, send daily reminder
    - Day 7: Send final warning, no more reminders after this
    - If rental_debt <= 0 (already paid), skip notification
    """
    LogInfo("Rental payment notification worker started.", LogType.SYSTEM_STATUS)
    
    last_sent_date = None
    
    while True:
        try:
            now = datetime.datetime.now()
            
            # Run at 08:00 system time
            cfg = settings.SCHEDULER_RENTAL
            if now.hour == cfg.get('hour', 8) and now.minute == cfg.get('minute', 0):
                current_date = now.date()
                if last_sent_date == current_date:
                    await asyncio.sleep(60)
                    continue

                LogInfo("Checking for rental payments due...", LogType.SYSTEM_STATUS)
                db = SessionLocal()
                try:
                    from app.models.rental import Rental, RentalCustomer, RentalStatus
                    from app.models.telegram import TelegramProjectMember
                    
                    active_contracts = db.query(Rental).filter(
                        Rental.status == RentalStatus.ACTIVE.value,
                        Rental.start_rental != None
                    ).all()
                    
                    if active_contracts:
                        client = bot
                        
                        for contract in active_contracts:
                            rental_day = contract.start_rental.day
                            
                            # Calculate due date for the current month cycle
                            # The due day is the same day as start_rental each month
                            if current_date.day >= rental_day:
                                due_year, due_month = current_date.year, current_date.month
                            else:
                                due_year, due_month = (current_date.year, current_date.month - 1) if current_date.month > 1 else (current_date.year - 1, 12)
                            
                            try:
                                due_date = datetime.date(due_year, due_month, rental_day)
                            except ValueError:
                                due_date = datetime.date(due_year, due_month, calendar.monthrange(due_year, due_month)[1])
                            
                            days_late = (current_date - due_date).days
                            
                            # Only process if within 0-7 day window
                            if days_late < 0 or days_late > 7:
                                continue
                            
                            customer = db.query(RentalCustomer).filter(
                                RentalCustomer.id == contract.customer_id
                            ).first()
                            if not customer:
                                continue
                            
                            # Day 0: Accumulate monthly rental into debt
                            if days_late == 0:
                                contract.rental_debt = (contract.rental_debt or 0.0) + (contract.monthly_rental or 0.0)
                                db.commit()
                            
                            # If already paid (debt <= 0), skip notification
                            if (contract.rental_debt or 0) <= 0:
                                continue
                            
                            from app.models.business import Projects
                            rental_project = db.query(Projects).filter(Projects.project_name == "Rental").first()
                            if not rental_project:
                                continue

                            # Find the member group for this customer
                            customer_links = db.query(TelegramProjectMember).filter(
                                TelegramProjectMember.project_id == rental_project.id,
                                TelegramProjectMember.role == "member",
                                TelegramProjectMember.slot_name.like("member%")
                            ).all()
                            
                            member_chat_id = None
                            for link in customer_links:
                                if customer.group_name and link.group_name == customer.group_name:
                                    member_chat_id = link.chat_id
                                    break
                                    
                            if not member_chat_id:
                                continue

                            def fmt_num(val):
                                if val is None: return 0
                                return int(val) if val == int(val) else val
                            
                            def fmt_dt(dt):
                                if not dt: return "N/A"
                                return dt.strftime('%d/%m/%Y')
                            
                            monthly = fmt_num(contract.monthly_rental)
                            debt = fmt_num(contract.rental_debt)
                            
                            # Build notification message based on days_late
                            if days_late == 0:
                                days_text = "Đến hạn hôm nay"
                            elif days_late == 7:
                                days_text = f"CẢNH BÁO CUỐI - Trễ hạn {days_late} ngày"
                            else:
                                days_text = f"Nhắc nhở lần {days_late} - Trễ hạn {days_late} ngày"
                            
                            msg_text = (
                                f"🔔 <b>THÔNG BÁO ĐÓNG TIỀN THUÊ ({days_text})</b> 🔔\n\n"
                                f"<b>Khách hàng:</b> {customer.customer_name}\n"
                                f"<b>Mã Khách Hàng:</b> {customer.customer_id or 'N/A'}\n"
                                f"<b>Liên hệ:</b> {customer.contact_info}\n"
                                f"<b>Mã Hợp Đồng:</b> <code>{contract.contract_id}</code>\n"
                                f"<b>Loại Hợp Đồng:</b> {contract.type_contract or 'N/A'}\n"
                                f"<b>Mã BĐS:</b> {contract.real_estate_id or 'N/A'}\n"
                                f"<b>Tiền thuê cần đóng:</b> <b>{monthly:,} VNĐ</b>\n"
                                f"<b>Công nợ hiện tại:</b> <b>{debt:,} VNĐ</b>\n"
                                f"<b>Thời gian thuê:</b> {fmt_dt(contract.start_rental)} - {fmt_dt(contract.end_rental)}\n"
                                f"---------------------------\n"
                                f"<i>Quý khách vui lòng thanh toán đúng hạn. Cảm ơn Quý Khách Hàng.</i>\n"
                                f"<i>Sau khi Quý Khách Hàng đã thanh toán, vui lòng gửi lại biên lai thanh toán để Admin được xác nhận.</i>\n"
                                f"<i>Admin vui lòng nhập <pre>/rental_payment_confirmed [Số tiền thanh toán]</pre> để xác nhận thanh toán.</i>"
                            )
                            
                            try:
                                await client.send_message(
                                    chat_id=int(member_chat_id),
                                    text=msg_text,
                                    parse_mode=ParseMode.HTML
                                )
                                LogInfo(f"Sent rental payment alert (day {days_late}) for contract {contract.contract_id} to {member_chat_id}", LogType.SYSTEM_STATUS)
                            except Exception as e:
                                LogError(f"Failed to send rental payment alert: {e}", LogType.SYSTEM_STATUS)
                            
                            # Day 7: Escalate to main group (admin/owner)
                            if days_late == 7:
                                # Find project_id and parent_main_chat_id from member link
                                target_project_id = None
                                parent_main_chat_id = None
                                for link in customer_links:
                                    if customer.group_name and link.group_name == customer.group_name:
                                        target_project_id = link.project_id
                                        parent_main_chat_id = link.parent_id
                                        break
                                
                                if target_project_id:
                                    main_group = None
                                    if parent_main_chat_id:
                                        main_group = db.query(TelegramProjectMember).filter(
                                            TelegramProjectMember.project_id == target_project_id,
                                            TelegramProjectMember.role == "main",
                                            TelegramProjectMember.chat_id == parent_main_chat_id
                                        ).first()
                                        
                                    if not main_group:
                                        main_group = db.query(TelegramProjectMember).filter(
                                            TelegramProjectMember.project_id == target_project_id,
                                            TelegramProjectMember.role == "main"
                                        ).first()
                                    
                                    if main_group:
                                        alert_text = (
                                            f"🚨 <b>CẢNH BÁO: KHÁCH HÀNG CHƯA ĐÓNG TIỀN THUÊ (QUÁ HẠN 7 NGÀY)</b> 🚨\n\n"
                                            f"<b>Khách hàng:</b> {customer.customer_name}\n"
                                            f"<b>Mã Khách Hàng:</b> {customer.customer_id or 'N/A'}\n"
                                            f"<b>Liên hệ:</b> {customer.contact_info}\n"
                                            f"<b>Mã Hợp Đồng:</b> <code>{contract.contract_id}</code>\n"
                                            f"<b>Loại Hợp Đồng:</b> {contract.type_contract or 'N/A'}\n"
                                            f"<b>Mã BĐS:</b> {contract.real_estate_id or 'N/A'}\n"
                                            f"<b>Tiền thuê / tháng:</b> <b>{monthly:,} VNĐ</b>\n"
                                            f"<b>Công nợ hiện tại:</b> <b>{debt:,} VNĐ</b>\n"
                                            f"---------------------------\n"
                                            f"<i>Hệ thống đã nhắc nhở 7 ngày liên tục nhưng khách hàng vẫn chưa thanh toán.\n"
                                            f"Vui lòng liên hệ trực tiếp để xử lý.</i>"
                                        )
                                        try:
                                            await client.send_message(
                                                chat_id=int(main_group.chat_id),
                                                text=alert_text,
                                                parse_mode=ParseMode.HTML
                                            )
                                            LogInfo(f"Sent rental overdue escalation for {contract.contract_id} to main group {main_group.chat_id}", LogType.SYSTEM_STATUS)
                                        except Exception as e:
                                            LogError(f"Failed to send rental overdue escalation: {e}", LogType.SYSTEM_STATUS)


                    last_sent_date = current_date

                except Exception as e:
                    LogError(f"Error checking rental payments: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()
                    
                await asyncio.sleep(61)
            
            next_run = (datetime.datetime.now() + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
            sleep_time = (next_run - datetime.datetime.now()).total_seconds()
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
            else:
                await asyncio.sleep(1)

        except Exception as e:
            LogError(f"Critical error in rental_payment_notification_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)


async def monthly_attendance_summary_worker():
    """
    Background worker: ngày 1 hàng tháng lúc 08:00.
    Tổng hợp chấm công tháng trước cho mỗi thành viên thuộc nhóm member_ns.
    Tạo ảnh HTML+CSS -> lưu file -> gửi lên nhóm main để admin xác nhận.
    """
    LogInfo("Monthly attendance summary worker started.", LogType.SYSTEM_STATUS)
    last_sent_month = None

    while True:
        try:
            now = datetime.datetime.now()

            cfg = settings.SCHEDULER_MONTHLY_SUMMARY
            if now.day == cfg.get('day', 1) and now.hour == cfg.get('hour', 8) and now.minute == cfg.get('minute', 0):
                if now.month == 1:
                    report_month = 12
                    report_year = now.year - 1
                else:
                    report_month = now.month - 1
                    report_year = now.year

                if last_sent_month == report_month:
                    await asyncio.sleep(600)
                    continue

                LogInfo(f"[AttSummary] Generating for {report_month:02d}/{report_year}...", LogType.SYSTEM_STATUS)

                db = SessionLocal()
                try:
                    from app.models.telegram import TelegramProjectMember
                    from app.models.employee import Employee
                    from app.models.finance import Attendance
                    from app.models.business import Projects
                    from bot.utils.attendance_generator import generate_attendance_image
                    from bot.utils.bot import bot

                    import os

                    project = db.query(Projects).filter(Projects.project_name == "Tiến Nga").first()
                    if not project:
                        last_sent_month = report_month
                        continue

                    project_id = project.id

                    main_members = db.query(TelegramProjectMember).filter(
                        TelegramProjectMember.project_id == project_id,
                        TelegramProjectMember.custom_title.in_([CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR]),
                    ).all()
                    main_chat_ids = list(set(m.chat_id for m in main_members))
                    if not main_chat_ids:
                        last_sent_month = report_month
                        continue

                    main_chat_id = main_chat_ids[0]

                    members = db.query(TelegramProjectMember).filter(
                        TelegramProjectMember.project_id == project_id,
                        TelegramProjectMember.custom_title == CustomTitle.MEMBER_HR,
                        TelegramProjectMember.is_bot == False,
                    ).all()

                    processed = set()
                    for member in members:
                        username = member.user_name
                        if not username or username in processed:
                            continue
                        processed.add(username)

                        employee = db.query(Employee).filter(Employee.username == username).first()
                        if not employee:
                            continue

                        employee_id = employee.id
                        full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()

                        records = db.query(Attendance).filter(
                            Attendance.employee_id == employee_id,
                            Attendance.year == report_year,
                            Attendance.month == report_month,
                        ).order_by(Attendance.day).all()

                        data_rows = []
                        for att in records:
                            ci = att.check_in_time.strftime("%H:%M") if att.check_in_time else "—"
                            co = att.check_out_time.strftime("%H:%M") if att.check_out_time else "—"
                            ot = "—"
                            wh = "—"
                            if att.start_overtime and att.end_overtime:
                                ot = f"{att.start_overtime.strftime('%H:%M')}-{att.end_overtime.strftime('%H:%M')}"
                            elif att.overtime:
                                ot = f"{att.overtime}h"
                            if att.working_time is not None:
                                wh = f"{att.working_time:.1f}h"
                            data_rows.append({
                                "day": att.day, "month": att.month, "date_str": att.date_str or "",
                                "check_in": ci, "check_out": co, "overtime": ot,
                                "work_hours": wh, "error": att.error or "",
                            })

                        img_buf = await generate_attendance_image(
                            employee_name=full_name, employee_id=employee_id,
                            month=report_month, year=report_year, records=data_rows,
                        )

                        folder = os.path.join("images", str(report_year), f"{report_month:02d}")
                        os.makedirs(folder, exist_ok=True)
                        filepath = os.path.join(folder, f"{employee_id}_{report_month:02d}_{report_year}_tong_hop_cong.png")
                        with open(filepath, "wb") as f:
                            f.write(img_buf.getvalue())

                        img_buf.seek(0)
                        await bot.send_photo(
                            chat_id=int(main_chat_id), photo=img_buf,
                            caption=(
                                f"<b>Tổng hợp chấm công tháng {report_month:02d}/{report_year}</b>\n"
                                f"Nhân viên: <b>{full_name}</b> (Mã NV: <code>{employee_id}</code>)\n"
                                f"Tổng: <b>{len(data_rows)}</b> ngày\n\n"
                                f"<i>Reply /tien_nga_export_payroll {employee_id} {report_month:02d}/{report_year} để xuất bảng lương.</i>"
                            ),
                            parse_mode=ParseMode.HTML,
                        )

                    last_sent_month = report_month
                    LogInfo(f"[AttSummary] Done. {len(processed)} members.", LogType.SYSTEM_STATUS)

                except Exception as e:
                    LogError(f"Error in monthly_attendance_summary_worker: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()

            await asyncio.sleep(60)

        except Exception as e:
            LogError(f"Critical error in monthly_attendance_summary_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)


async def send_factory_purchase_report(db, project_id, current_date, client, specific_cp_id=None):
    from app.models.business import DailyPurchases, CollectionPoint
    from app.models.telegram import TelegramProjectMember
    from bot.utils.receipt_generator import fmt_money_vn
    from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    from pyrogram.enums import ParseMode
    from bot.utils.logger import LogInfo, LogError, LogType

    main_groups = db.query(TelegramProjectMember.chat_id, TelegramProjectMember.group_name, TelegramProjectMember.custom_title).filter(
        TelegramProjectMember.project_id == project_id,
        TelegramProjectMember.role == "main",
        TelegramProjectMember.custom_title.in_(["super_main", "main_supplier"])
    ).distinct().all()
    
    if not main_groups:
        return

    query = db.query(DailyPurchases).filter(DailyPurchases.day == current_date)
    if specific_cp_id:
        query = query.filter(DailyPurchases.collection_point_id == specific_cp_id)
        
    purchases = query.all()
    if not purchases:
        return
        
    from collections import defaultdict
    by_cp = defaultdict(list)
    for p in purchases:
        by_cp[p.collection_point_id].append(p)
        
    for cp_id, cp_purchases in by_cp.items():
        cp = db.query(CollectionPoint).filter(CollectionPoint.id == cp_id).first()
        cp_name = cp.collection_name if cp else str(cp_id)
        
        from bot.core.config import settings
        
        target_chat_ids = []
        
        # Lấy tên group được map cho xưởng này (nếu có)
        mapped_group_name = settings.FACTORY_GROUP_MAPPING.get(cp_name)
        
        if mapped_group_name:
            for chat_id, group_name in main_groups:
                if group_name == mapped_group_name:
                    target_chat_ids.append(chat_id)
                    
        if not target_chat_ids:
            # Fallback nếu không có mapping hoặc không tìm thấy group: gửi cho tất cả các nhóm main, trừ super_main
            target_chat_ids = [g[0] for g in main_groups if g[2] != "super_main"]
        else:
            # Loại bỏ trùng lặp nếu có
            target_chat_ids = list(set(target_chat_ids))
            
        by_product = defaultdict(list)
        for p in cp_purchases:
            code = p.product_code or "N/A"
            by_product[code].append(p)
            
        report_text = f"<b>TỔNG HỢP MUA MỦ - {cp_name.upper()}</b>\n"
        report_text += f"<b>Ngày:</b> {current_date.strftime('%d/%m/%Y')}\n"
        report_text += f"-----------------------------------\n"
        
        total_all_amount = 0
        
        for p_code, p_list in by_product.items():
            t_wet = sum(p.actual_weight or 0 for p in p_list)
            t_dry = sum(p.dry_rubber or 0 for p in p_list)
            t_amount = sum(p.total_amount or 0 for p in p_list)
            total_all_amount += t_amount
            
            degrees = [p.degree for p in p_list if p.degree is not None]
            avg_deg = sum(degrees) / len(degrees) if degrees else 0
            
            prices = [p.unit_price for p in p_list if p.unit_price is not None]
            avg_price = sum(prices) / len(prices) if prices else 0
            
            subsidy_prices = [p.subsidy_price for p in p_list if p.subsidy_price is not None]
            avg_total_price = sum(subsidy_prices) / len(subsidy_prices) if subsidy_prices else 0
            
            report_text += f"<b>Mã lô hàng:</b> <code>{p_code}</code>\n"
            report_text += f" Tổng khối lượng mủ nước: <b>{t_wet:,.2f} kg</b>\n"
            report_text += f" Tổng số lượng Mủ khô: <b>{t_dry:,.2f} kg</b>\n"
            report_text += f" Số độ trung bình: <b>{avg_deg:.2f}%</b>\n"
            report_text += f" Đơn giá: <b>{fmt_money_vn(avg_price)} VNĐ</b>\n"
            report_text += f" Giá thu mua trung bình: <b>{fmt_money_vn(avg_total_price)} VNĐ</b>\n"
            report_text += f" Tổng thành tiền: <b>{fmt_money_vn(t_amount)} VNĐ</b>\n\n"
            
        report_text += f"-----------------------------------\n"
        report_text += f"<b>TỔNG THÀNH TIỀN: {fmt_money_vn(total_all_amount)} VNĐ</b>\n\n"
        report_text += f"<b>Nhắc nhở:</b> Admin vui lòng thực hiện lệnh <code>/tien_nga_kiem_soat_hao_hut</code> nếu đã thấy ok. Nếu chưa thấy ok thì hãy làm các lệnh liên quan để hoàn thiện rồi nhấn vào nút <b>Tạo lại báo cáo</b>."
        
        buttons = [
            [
                InlineKeyboardButton("Xác nhận", callback_data=f"confirm_fac_rep_{cp_id}_{current_date}"),
                InlineKeyboardButton("Tạo lại báo cáo", callback_data=f"regen_fac_rep_{cp_id}_{current_date}")
            ]
        ]
        
        reply_markup = InlineKeyboardMarkup(buttons)
        
        for chat_id in target_chat_ids:
            try:
                await client.send_message(
                    chat_id=int(chat_id),
                    text=report_text,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML
                )
            except Exception as e:
                LogError(f"Error sending factory report to {chat_id}: {e}", LogType.SYSTEM_STATUS)

async def daily_factory_purchase_summary_worker():
    """
    Background worker: chạy lúc 17:30 hàng ngày.
    Tổng hợp kết quả thu mua mủ của Xưởng và gửi vào nhóm main.
    """
    from bot.utils.logger import LogInfo, LogError, LogType
    from bot.core.config import settings
    from bot.utils.bot import bot
    import asyncio
    import datetime

    LogInfo("Daily factory purchase summary worker started.", LogType.SYSTEM_STATUS)
    last_sent_date = None

    while True:
        try:
            now = datetime.datetime.now()

            cfg = settings.SCHEDULER_DAILY_FACTORY_PURCHASE
            if now.hour == cfg.get('hour', 17) and now.minute == cfg.get('minute', 30):
                current_date = now.date()
                if last_sent_date == current_date:
                    await asyncio.sleep(60)
                    continue

                LogInfo(f"[DailyFactoryPurchase] Generating summaries for {current_date}...", LogType.SYSTEM_STATUS)
                from app.db.session import SessionLocal
                db = SessionLocal()
                try:
                    from app.models.business import Projects

                    project = db.query(Projects).filter(Projects.project_name == "Tiến Nga").first()
                    if not project:
                        LogInfo("[DailyFactoryPurchase] Không tìm thấy project Tiến Nga, bỏ qua.", LogType.SYSTEM_STATUS)
                        last_sent_date = current_date
                        continue

                    await send_factory_purchase_report(db, project.id, current_date, bot)
                    last_sent_date = current_date

                except Exception as e:
                    LogError(f"Error in daily_factory_purchase_summary_worker inner: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()
            
            await asyncio.sleep(60)

        except Exception as e:
            LogError(f"Error in daily_factory_purchase_summary_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)


async def daily_purchase_summary_worker():
    """
    Background worker: chạy lúc 20:00 hàng ngày.
    Tổng hợp daily_purchases trong ngày, cho mỗi hộ dân tạo ảnh biên lai
    và gửi vào nhóm member_ns tương ứng (dựa trên customer.username).
    """
    LogInfo("Daily purchase summary worker started.", LogType.SYSTEM_STATUS)
    last_sent_date = None

    while True:
        try:
            now = datetime.datetime.now()

            cfg = settings.SCHEDULER_DAILY_PURCHASE
            if now.hour == cfg.get('hour', 20) and now.minute == cfg.get('minute', 0):
                current_date = now.date()
                if last_sent_date == current_date:
                    await asyncio.sleep(60)
                    continue

                LogInfo(f"[DailyPurchase] Generating summaries for {current_date}...", LogType.SYSTEM_STATUS)
                db = SessionLocal()
                try:
                    from app.models.business import DailyPurchases, Customers, CollectionPoint, Projects
                    from app.models.telegram import TelegramProjectMember
                    from bot.utils.receipt_generator import generate_chotso_ketoan_image, fmt_money_vn
                    from bot.utils.bot import bot

                    from sqlalchemy import func

                    # Tìm project Tiến Nga
                    project = db.query(Projects).filter(Projects.project_name == "Tiến Nga").first()
                    if not project:
                        LogInfo("[DailyPurchase] Không tìm thấy project Tiến Nga, bỏ qua.", LogType.SYSTEM_STATUS)
                        last_sent_date = current_date
                        continue

                    # Lấy tất cả purchases trong ngày hôm nay
                    today_purchases = db.query(DailyPurchases).filter(
                        DailyPurchases.day == current_date
                    ).all()

                    if not today_purchases:
                        LogInfo("[DailyPurchase] Không có dữ liệu mua mủ ngày hôm nay.", LogType.SYSTEM_STATUS)
                        last_sent_date = current_date
                        continue

                    # Nhóm theo hoursehold_id
                    from collections import defaultdict
                    grouped = defaultdict(list)
                    for p in today_purchases:
                        grouped[p.hoursehold_id].append(p)

                    sent_count = 0
                    for hh_id, purchases in grouped.items():
                        try:
                            customer = db.query(Customers).filter(Customers.hoursehold_id == hh_id).first()
                            if not customer:
                                continue

                            # Tìm nhóm member_ns của khách hàng dựa trên username
                            cust_username = customer.username
                            if not cust_username:
                                continue

                            # Tìm member trong project Tiến Nga có user_name khớp
                            member = db.query(TelegramProjectMember).filter(
                                TelegramProjectMember.project_id == project.id,
                                TelegramProjectMember.custom_title == CustomTitle.MEMBER_SUPPLIER,
                                TelegramProjectMember.user_name == cust_username or TelegramProjectMember.user_name == "@" + cust_username or "@" + TelegramProjectMember.user_name == cust_username,
                            ).first()

                            LogInfo(f"[DailyPurchase] Project id: {project.id}, customer username: {cust_username}, member: {member}", LogType.SYSTEM_STATUS)

                            if not member:
                                continue

                            member_chat_id = member.chat_id

                            # Tính toán tổng hợp
                            cp_name = ""
                            if customer.collection_point_id:
                                cp = db.query(CollectionPoint).filter(CollectionPoint.id == customer.collection_point_id).first()
                                if cp:
                                    cp_name = cp.collection_name

                            total_amount_today = sum(p.total_amount or 0 for p in purchases)

                            # Tổng hợp tháng (cùng tháng của current_date)
                            month_start = current_date.replace(day=1)
                            month_purchases = db.query(DailyPurchases).filter(
                                DailyPurchases.hoursehold_id == hh_id,
                                DailyPurchases.day >= month_start,
                                DailyPurchases.day <= current_date
                            ).all()

                            tong_mua_mu_thang = sum(p.total_amount or 0 for p in month_purchases)
                            tong_tam_ung_thang = sum(p.advance_amount or 0 for p in month_purchases)
                            tong_da_thanh_toan_thang = sum(p.paid_amount or 0 for p in month_purchases)

                            cong_no_hien_tai = tong_mua_mu_thang - tong_da_thanh_toan_thang - tong_tam_ung_thang

                            ngay_str = current_date.strftime("%d/%m/%Y")

                            # Extract detail fields
                            so_luong = sum(p.weight or 0 for p in purchases)
                            mu_kho = sum(p.dry_rubber or 0 for p in purchases)
                            so_do = [str(p.degree) for p in purchases if p.degree is not None]
                            tro_gia = [str(p.is_subsidized) for p in purchases if p.is_subsidized is not None]
                            don_gia = [p.unit_price for p in purchases if p.unit_price is not None]
                            gia_ho_tro = [p.subsidy_price for p in purchases if p.subsidy_price is not None]

                            receipt_data = {
                                "ngay": ngay_str,
                                "diem_thu_mua": cp_name or "N/A",
                                "ten_kh": customer.fullname or hh_id,
                                "ma_ho": hh_id,
                                "tien_mua_mu_ngay": total_amount_today,
                                "cong_no_cuoi_ky": total_amount_today,
                                "tong_tam_ung_thang": -abs(tong_tam_ung_thang) if tong_tam_ung_thang else 0,
                                "tong_mua_mu_thang": tong_mua_mu_thang,
                                "tong_da_thanh_toan_thang": -abs(tong_da_thanh_toan_thang) if tong_da_thanh_toan_thang else 0,
                                "cong_no_hien_tai": cong_no_hien_tai,
                                "so_luong": round(so_luong, 2),
                                "mu_kho": round(mu_kho, 2),
                                "so_do": ", ".join(so_do) if so_do else "0",
                                "tro_gia": ", ".join(tro_gia) if tro_gia else "0",
                                "don_gia": don_gia,
                                "gia_ho_tro": gia_ho_tro,
                            }

                            buf = await generate_chotso_ketoan_image(receipt_data)

                            caption = (
                                f"<b>TỔNG HỢP MUA MỦ NGÀY {ngay_str}</b>\n"
                                f"Khách hàng: <b>{customer.fullname}</b> ({hh_id})\n"
                                f"Xưởng: <b>{cp_name or 'N/A'}</b>\n"
                                f"Thành tiền hôm nay: <b>{fmt_money_vn(total_amount_today)} VNĐ</b>\n"
                                f"Công Nợ Hiện Tại: <b>{fmt_money_vn(cong_no_hien_tai)} VNĐ</b>"
                            )

                            await bot.send_photo(
                                chat_id=int(member_chat_id),
                                photo=buf,
                                caption=caption,
                                parse_mode=ParseMode.HTML
                            )

                            sent_count += 1
                            LogInfo(f"[DailyPurchase] Sent summary for {hh_id} to chat {member_chat_id}", LogType.SYSTEM_STATUS)

                        except Exception as e:
                            LogError(f"[DailyPurchase] Error processing {hh_id}: {e}", LogType.SYSTEM_STATUS)

                    last_sent_date = current_date
                    LogInfo(f"[DailyPurchase] Done. Sent {sent_count} summaries.", LogType.SYSTEM_STATUS)

                except Exception as e:
                    LogError(f"Error in daily_purchase_summary_worker: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()

                await asyncio.sleep(61)

            next_run = (datetime.datetime.now() + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
            sleep_time = (next_run - datetime.datetime.now()).total_seconds()
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
            else:
                await asyncio.sleep(1)

        except Exception as e:
            LogError(f"Critical error in daily_purchase_summary_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)

async def daily_fund_summary_worker():
    """
    Background worker that runs daily to aggregate fund income and expenses
    and sends a report to the configured Telegram groups.
    """
    LogInfo("Daily fund summary worker started.", LogType.SYSTEM_STATUS)
    last_sent_date = None

    while True:
        try:
            now = datetime.datetime.now()
            cfg = settings.SCHEDULER_DAILY_FUND
            
            if now.hour == cfg.get('hour', 17) and now.minute == cfg.get('minute', 30):
                current_date = now.date()
                if last_sent_date == current_date:
                    await asyncio.sleep(60)
                    continue

                LogInfo("Generating daily fund summary...", LogType.SYSTEM_STATUS)
                db = SessionLocal()
                try:
                    from app.models.business import Investment, DailyPayment
                    from app.models.telegram import TelegramProjectMember
                    from bot.utils.bot import bot
                    from sqlalchemy import func
                    from bot.utils.utils import fmt_money

                    for fund_name, group_name in settings.FUND_GROUP_MAPPING.items():
                        LogInfo(f"Checking fund mapping: {fund_name} -> {group_name}", LogType.SYSTEM_STATUS)
                        investment = db.query(Investment).filter(Investment.name == fund_name).first()
                        if not investment:
                            LogInfo(f"Fund '{fund_name}' not found in DB.", LogType.SYSTEM_STATUS)
                            continue
                            
                        # Find telegram group
                        tpm = db.query(TelegramProjectMember).filter(TelegramProjectMember.group_name == group_name).first()
                        if not tpm or not tpm.chat_id:
                            LogInfo(f"Telegram group '{group_name}' not found or chat_id is missing.", LogType.SYSTEM_STATUS)
                            continue
                            
                        # Calculate today's income/expense
                        today_payments = db.query(DailyPayment).filter(
                            DailyPayment.investment_id == investment.id,
                            DailyPayment.day == current_date,
                            DailyPayment.status == "APPROVED"
                        ).all()
                        
                        tong_thu = sum(p.amount for p in today_payments if p.payment_type.lower() == 'thu')
                        tong_chi = sum(p.amount for p in today_payments if p.payment_type.lower() == 'chi')
                        
                        so_du = (investment.total_income or 0) - (investment.total_expense or 0) + (investment.initial_capital or 0)
                        
                        from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
                        from pyrogram.enums import ParseMode
                        markup = InlineKeyboardMarkup([
                            [InlineKeyboardButton("Xác nhận", callback_data=f"fund_confirm_{investment.id}")],
                            [InlineKeyboardButton("Tạo lại báo cáo", callback_data=f"fund_regen_{investment.id}")]
                        ])
                        
                        text = (
                            f"<b>BÁO CÁO TỔNG HỢP THU CHI QUỸ: {fund_name}</b>\n\n"
                            f"<b>Ngày:</b> {current_date.strftime('%d/%m/%Y')}\n"
                            f"<b>Tổng Thu (Hôm nay):</b> <code>{fmt_money(tong_thu)}</code> VNĐ\n"
                            f"<b>Tổng Chi (Hôm nay):</b> <code>{fmt_money(tong_chi)}</code> VNĐ\n"
                            f"<b>Số tiền trong Quỹ hiện tại:</b> <code>{fmt_money(so_du)}</code> VNĐ\n\n"
                            f"<i>Vui lòng kiểm tra và xác nhận.</i>"
                        )
                        
                        try:
                            await bot.send_message(
                                chat_id=int(tpm.chat_id),
                                text=text,
                                parse_mode=ParseMode.HTML,
                                reply_markup=markup
                            )
                            LogInfo(f"Successfully sent fund summary to {group_name}", LogType.SYSTEM_STATUS)
                        except Exception as e:
                            LogError(f"Failed to send fund summary to {group_name}: {e}", LogType.SYSTEM_STATUS)

                    last_sent_date = current_date

                except Exception as e:
                    LogError(f"Error generating daily fund summary: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()
                    
                await asyncio.sleep(61)
                
            next_run = (datetime.datetime.now() + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
            sleep_time = (next_run - datetime.datetime.now()).total_seconds()
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
            else:
                await asyncio.sleep(1)

        except Exception as e:
            LogError(f"Critical error in daily_fund_summary_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)

async def generate_and_send_inventory_report(client, target_date: datetime.date = None, specific_inv_id: str = None):
    if target_date is None:
        target_date = datetime.datetime.now().date()
        
    db = SessionLocal()
    try:
        from app.models.inventory import Inventory, ProductTransaction
        from app.models.telegram import TelegramProjectMember
        from bot.core.config import settings
        from pyrogram.types import InlineKeyboardButton, InlineKeyboardMarkup
        from pyrogram.enums import ParseMode
        
        query = db.query(Inventory)
        if specific_inv_id:
            query = query.filter(Inventory.id == specific_inv_id)
            
        inventories = query.all()
        if not inventories:
            return
            
        main_groups = db.query(TelegramProjectMember.chat_id, TelegramProjectMember.group_name).filter(
            TelegramProjectMember.role == "main",
            TelegramProjectMember.is_bot == False
        ).all()
        
        for inv in inventories:
            inv_name = inv.storage_name or "N/A"
            inv_id = inv.id
            
            transactions = db.query(ProductTransaction).filter(
                ProductTransaction.storage_id == inv_id,
                ProductTransaction.transaction_date == target_date
            ).all()
            
            total_export = 0.0
            total_import = 0.0
            
            for txn in transactions:
                if txn.transaction_type == "Xuất":
                    total_export += txn.quantity or 0.0
                elif txn.transaction_type == "Nhập":
                    total_import += txn.quantity or 0.0
            
            current_stock = inv.quantity or 0.0
            capacity = inv.capacity or 0.0
            
            mapped_group_name = settings.INVENTORY_GROUP_MAPPING.get(inv_name)
            target_chat_ids = []
            
            if mapped_group_name:
                for chat_id, group_name in main_groups:
                    if group_name == mapped_group_name:
                        target_chat_ids.append(chat_id)
                        
            if not target_chat_ids:
                inventory_groups = db.query(TelegramProjectMember.chat_id).filter(
                    TelegramProjectMember.role == "main",
                    TelegramProjectMember.is_bot == False,
                    TelegramProjectMember.custom_title.in_(["main_inventory", "main_product"])
                ).distinct().all()
                target_chat_ids = [g[0] for g in inventory_groups]
            
            target_chat_ids = list(set(target_chat_ids))
                
            report_text = f"<b>TỔNG HỢP KẾT QUẢ XUẤT/NHẬP KHO - {inv_name.upper()}</b>\n"
            report_text += f"<b>Ngày:</b> {target_date.strftime('%d/%m/%Y')}\n"
            report_text += f"-----------------------------------\n"
            report_text += f" Tổng nhập kho: <b>{total_import:,.2f} Kg</b>\n"
            report_text += f" Tổng xuất kho: <b>{total_export:,.2f} Kg</b>\n"
            report_text += f" Số lượng Tồn kho: <b>{current_stock:,.2f} Kg</b>\n"
            report_text += f" Sức chứa của kho: <b>{capacity:,.2f} Kg</b>\n\n"
            report_text += f"<b>Nhắc nhở:</b> Admin vui lòng kiểm tra, nếu đã thấy ok thì nhấn nút <b>Xác nhận</b>. Nếu chưa thấy ok thì hãy làm các lệnh liên quan để hoàn thiện rồi nhấn vào nút <b>Tạo lại báo cáo</b>."
            
            buttons = [
                [
                    InlineKeyboardButton("Xác nhận", callback_data=f"confirm_inv_rep_{inv_id}_{target_date}"),
                    InlineKeyboardButton("Tạo lại báo cáo", callback_data=f"regen_inv_rep_{inv_id}_{target_date}")
                ]
            ]
            
            reply_markup = InlineKeyboardMarkup(buttons)
            
            from pyrogram.errors import FloodWait
            import asyncio
            
            for chat_id in target_chat_ids:
                try:
                    await client.send_message(
                        chat_id=int(chat_id),
                        text=report_text,
                        reply_markup=reply_markup,
                        parse_mode=ParseMode.HTML
                    )
                    await asyncio.sleep(0.5)
                except FloodWait as e:
                    LogError(f"FloodWait hit, sleeping for {e.value} seconds...", LogType.SYSTEM_STATUS)
                    await asyncio.sleep(e.value + 1)
                    try:
                        await client.send_message(
                            chat_id=int(chat_id),
                            text=report_text,
                            reply_markup=reply_markup,
                            parse_mode=ParseMode.HTML
                        )
                    except Exception as ex:
                        LogError(f"Error sending inventory report to {chat_id} after retry: {ex}", LogType.SYSTEM_STATUS)
                except Exception as e:
                    LogError(f"Error sending inventory report to {chat_id}: {e}", LogType.SYSTEM_STATUS)
                    
    except Exception as e:
        LogError(f"Error generating inventory report: {e}", LogType.SYSTEM_STATUS)
    finally:
        db.close()

async def daily_inventory_summary_worker():
    from bot.utils.logger import LogInfo, LogError, LogType
    from bot.core.config import settings
    from bot.utils.bot import bot
    import asyncio
    import datetime

    LogInfo("Daily inventory summary worker started.", LogType.SYSTEM_STATUS)
    last_sent_date = None

    while True:
        try:
            now = datetime.datetime.now()

            cfg = settings.SCHEDULER_DAILY_INVENTORY
            if now.hour == cfg.get('hour', 17) and now.minute == cfg.get('minute', 30):
                current_date = now.date()
                if last_sent_date == current_date:
                    await asyncio.sleep(60)
                    continue

                LogInfo("Generating daily inventory reports...", LogType.SYSTEM_STATUS)
                
                await generate_and_send_inventory_report(bot, current_date)

                last_sent_date = current_date

            next_run = (datetime.datetime.now() + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
            sleep_time = (next_run - datetime.datetime.now()).total_seconds()
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
            else:
                await asyncio.sleep(1)
        except Exception as e:
            LogError(f"Error in daily_inventory_summary_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)

async def send_harvest_summary_report(db, project_id, current_date, client, specific_affiliation=None):
    from app.models.business import DailyHarvest, DailyPurchases, AgriculturalLand, Households
    from app.models.telegram import TelegramProjectMember
    from bot.core.config import settings
    from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    from pyrogram.enums import ParseMode
    from bot.utils.receipt_generator import fmt_money_vn
    from sqlalchemy import func

    # Gather data by affiliation
    affiliations = settings.HARVEST_GROUP_MAPPING.keys()
    if specific_affiliation:
        affiliations = [specific_affiliation]

    main_groups = db.query(TelegramProjectMember.chat_id, TelegramProjectMember.group_name).filter(
        TelegramProjectMember.project_id == project_id,
        TelegramProjectMember.role == "main",
        TelegramProjectMember.is_bot == False
    ).all()

    for aff in affiliations:
        mapped_group_name = settings.HARVEST_GROUP_MAPPING.get(aff)
        target_chat_ids = []
        if mapped_group_name:
            for chat_id, group_name in main_groups:
                if group_name == mapped_group_name:
                    target_chat_ids.append(chat_id)
                    
        if not target_chat_ids:
            continue
            
        target_chat_ids = list(set(target_chat_ids))

        # Query lands for this affiliation
        lands = db.query(AgriculturalLand).filter(AgriculturalLand.affiliation == aff).all()
        land_codes = [l.land_code for l in lands]
        
        # Query households associated with these lands
        households = db.query(Households).filter(Households.land_code.in_(land_codes)).all()
        hh_codes = [h.household_code for h in households]
        purchase_codes = [h.purchase_code for h in households]

        # 1. Total trees harvested
        harvests = db.query(DailyHarvest).filter(
            DailyHarvest.day == current_date,
            DailyHarvest.land_code.in_(land_codes)
        ).all()
        
        total_trees = sum(h.tree_count or 0 for h in harvests)
        total_harvest_amount = sum(h.total_amount or 0 for h in harvests)

        # 2. Total purchased
        purchases = db.query(DailyPurchases).filter(
            DailyPurchases.day == current_date,
            DailyPurchases.hoursehold_id.in_(purchase_codes)
        ).all()

        total_actual_weight = sum(p.actual_weight or 0 for p in purchases)
        total_dry_rubber = sum(p.dry_rubber or 0 for p in purchases)
        total_purchase_amount = sum(p.total_amount or 0 for p in purchases)
        
        # Averages
        if total_actual_weight > 0:
            avg_degree = sum((p.degree or 0) * (p.actual_weight or 0) for p in purchases) / total_actual_weight
        else:
            avg_degree = 0.0

        prices = [p.unit_price for p in purchases if p.unit_price is not None]
        avg_unit_price = sum(prices) / len(prices) if prices else 0.0

        report_text = f"<b>TỔNG HỢP KẾT QUẢ THU HOẠCH CAO SU - {aff.upper()}</b>\n"
        report_text += f"<b>Ngày:</b> {current_date.strftime('%d/%m/%Y')}\n"
        report_text += f"-----------------------------------\n"
        report_text += f" Tổng số lượng cây thu hoạch: <b>{total_trees:,.0f} cây</b>\n"
        report_text += f" Tổng tiền trả cạo mủ: <b>{fmt_money_vn(total_harvest_amount)} VNĐ</b>\n"
        report_text += f" Tổng SL Mủ nước thu mua: <b>{total_actual_weight:,.2f} kg</b>\n"
        report_text += f" Tổng SL Mủ khô thu mua: <b>{total_dry_rubber:,.2f} kg</b>\n"
        report_text += f" Số độ trung bình: <b>{avg_degree:.2f}%</b>\n"
        report_text += f" Đơn giá trung bình: <b>{fmt_money_vn(avg_unit_price)} VNĐ</b>\n"
        report_text += f" Tổng thành tiền: <b>{fmt_money_vn(total_purchase_amount)} VNĐ</b>\n\n"
        report_text += f"<b>Nhắc nhở:</b> Admin vui lòng kiểm tra, nếu đã thấy ok thì nhấn nút <b>Xác nhận</b>. Nếu chưa thấy ok thì hãy làm các lệnh liên quan để hoàn thiện rồi nhấn vào nút <b>Tạo lại báo cáo</b>."

        buttons = [
            [
                InlineKeyboardButton("Xác nhận", callback_data=f"confirm_harv_rep_{aff}_{current_date}"),
                InlineKeyboardButton("Tạo lại báo cáo", callback_data=f"regen_harv_rep_{aff}_{current_date}")
            ]
        ]
        
        reply_markup = InlineKeyboardMarkup(buttons)
        
        from pyrogram.errors import FloodWait
        import asyncio
        
        for chat_id in target_chat_ids:
            try:
                await client.send_message(
                    chat_id=int(chat_id),
                    text=report_text,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML
                )
                await asyncio.sleep(0.5)
            except FloodWait as e:
                LogError(f"FloodWait hit, sleeping for {e.value} seconds...", LogType.SYSTEM_STATUS)
                await asyncio.sleep(e.value + 1)
                await client.send_message(
                    chat_id=int(chat_id),
                    text=report_text,
                    reply_markup=reply_markup,
                    parse_mode=ParseMode.HTML
                )
            except Exception as e:
                LogError(f"Error sending harvest report to {chat_id}: {e}", LogType.SYSTEM_STATUS)


async def daily_harvest_summary_worker():
    from bot.utils.logger import LogInfo, LogError, LogType
    from bot.core.config import settings
    from bot.utils.bot import bot
    import asyncio
    import datetime

    LogInfo("Daily harvest summary worker started.", LogType.SYSTEM_STATUS)
    last_sent_date = None

    while True:
        try:
            now = datetime.datetime.now()

            cfg = settings.SCHEDULER_DAILY_HARVEST
            if now.hour == cfg.get('hour', 17) and now.minute == cfg.get('minute', 30):
                current_date = now.date()
                if last_sent_date == current_date:
                    await asyncio.sleep(60)
                    continue

                LogInfo(f"Generating daily harvest reports for {current_date}...", LogType.SYSTEM_STATUS)
                
                from app.db.session import SessionLocal
                from app.models.business import Projects
                db = SessionLocal()
                try:
                    project = db.query(Projects).filter(Projects.project_name == "Tiến Nga").first()
                    if project:
                        await send_harvest_summary_report(db, project.id, current_date, bot)
                except Exception as e:
                    LogError(f"Error in daily harvest summary generation: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()

                last_sent_date = current_date

            next_run = (datetime.datetime.now() + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
            sleep_time = (next_run - datetime.datetime.now()).total_seconds()
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
            else:
                await asyncio.sleep(1)
        except Exception as e:
            LogError(f"Error in daily_harvest_summary_worker: {e}", LogType.SYSTEM_STATUS)
            await asyncio.sleep(60)


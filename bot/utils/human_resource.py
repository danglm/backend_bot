"""
Human Resource Utilities
Các hàm xử lý dùng chung cho quản lý nhân sự, dùng được cho nhiều dự án (Tiến Nga, GGoMoonSin, ...).
"""
import uuid
import datetime
from io import BytesIO
from pyrogram.types import Message, InlineKeyboardMarkup
from bot.utils.enums import CustomTitle
from pyrogram.enums import ParseMode
from app.db.session import SessionLocal
from app.models.employee import Employee
from app.models.telegram import TelegramProjectMember
from bot.utils.logger import LogInfo, LogError, LogType


EMPLOYEE_FORM_TEMPLATE = """<b>FORM THÊM NHÂN VIÊN</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>{cmd}
Mã NV: 
Họ: 
Tên: 
Username: 
Ủy quyền: 
Nhóm Telegram: 
Giới tính: 
Ngày sinh: 
SĐT: 
Email: 
Địa chỉ: 
CCCD/CMND: 
Nơi cấp: 
Quốc tịch: Việt Nam
Tình trạng hôn nhân: 
Trình độ học vấn: 
Chuyên ngành: 
Chứng chỉ: 
Kinh nghiệm: 
Phòng ban: 
Chức vụ: 
Loại hợp đồng: 
Ảnh nhân viên: 
Giờ vào ca (hh:mm): 
Giờ tan ca (hh:mm): 
Giờ vào ca T7 (hh:mm): 
Giờ tan ca T7 (hh:mm): 
Số giờ làm việc (giờ/ngày): 
Lương cơ bản (VNĐ): 
Lương tháng (VNĐ): 
Lương tuần (VNĐ): 
Lương ngày (VNĐ): 
Lương giờ (VNĐ): 
Lương làm thêm giờ (VNĐ): 
Tiền thưởng (VNĐ): 
Phúc lợi: 
Số ngày phép năm: 
Bảo hiểm: 
Tỷ lệ BHXH (%): 
Mục tiêu nghề nghiệp: 
Đánh giá hiệu suất: 
Ngân hàng: 
Số tài khoản: 
Mã thanh toán: 
SĐT khẩn cấp: 
Người liên hệ khẩn cấp: 
Auto chấm công (có/không): có
Loại công (1-4): 3</pre>

<i>Các trường không bắt buộc có thể để trống.
Bắt buộc: <b>Mã NV</b>, <b>Họ</b>, <b>Tên</b>, <b>Username</b> (Telegram)
Loại công: 1=T2-T6, 2=T2-T7(sáng), 3=T2-T7, 4=T2-CN</i>"""

# Danh sách các field label chứa dấu ":" → cần match đặc biệt khi parse form
_KNOWN_FIELD_LABELS = [
    "Giờ vào ca (hh:mm)",
    "Giờ tan ca (hh:mm)",
    "Số giờ làm việc (giờ/ngày)",
    "Lương cơ bản (VNĐ)",
    "Lương tháng (VNĐ)",
    "Lương tuần (VNĐ)",
    "Lương ngày (VNĐ)",
    "Lương giờ (VNĐ)",
    "Lương làm thêm giờ (VNĐ)",
    "Tiền thưởng (VNĐ)",
    "Tỷ lệ BHXH (%)",
    "Auto chấm công (có/không)",
    "Loại công (1-4)",
    "Giờ vào ca T7 (hh:mm)",
    "Giờ tan ca T7 (hh:mm)",
]


def _parse_form_lines(lines: list[str]) -> dict[str, str]:
    """
    Parse danh sách dòng form thành dict {field_label: value}.
    Xử lý đúng các field label chứa dấu ":" (vd: "Giờ vào ca (hh:mm): 13:30").
    """
    data = {}
    for line in lines:
        line_stripped = line.strip()
        if not line_stripped:
            continue
        matched = False
        # Thử match với các known field labels (chứa dấu ":")
        for label in _KNOWN_FIELD_LABELS:
            prefix = label + ":"
            if line_stripped.startswith(prefix):
                val = line_stripped[len(prefix):].strip()
                data[label] = val
                matched = True
                break
        if not matched and ":" in line_stripped:
            key, val = line_stripped.split(":", 1)
            data[key.strip()] = val.strip()
    return data


async def _parse_float_or_reply(message: Message, field_name: str, value: str):
    if not value: return None
    try:
        return float(value.replace(".", "").replace(",", "."))
    except ValueError:
        await message.reply_text(f"⚠️ <b>{field_name}</b> nhập sai kiểu dữ liệu. Vui lòng nhập đúng kiểu số (VD: 10000000).", parse_mode=ParseMode.HTML)
        raise ValueError()

async def _parse_int_or_reply(message: Message, field_name: str, value: str):
    if not value: return None
    try:
        return int(value)
    except ValueError:
        await message.reply_text(f"⚠️ <b>{field_name}</b> nhập sai kiểu dữ liệu. Vui lòng nhập nguyên (VD: 12).", parse_mode=ParseMode.HTML)
        raise ValueError()

async def _parse_time_or_reply(message: Message, field_name: str, value: str):
    if not value: return None
    try:
        parts = value.split(":")
        return datetime.datetime(2000, 1, 1, int(parts[0]), int(parts[1]) if len(parts) > 1 else 0)
    except (ValueError, IndexError):
        await message.reply_text(f"⚠️ <b>{field_name}</b> nhập sai kiểu dữ liệu. Vui lòng nhập đúng định dạng giờ (VD: 08:00).", parse_mode=ParseMode.HTML)
        raise ValueError()



async def handle_create_employee(client, message: Message, command_name: str) -> None:
    """
    Xử lý logic thêm nhân viên - dùng chung cho mọi dự án.
    
    Args:
        client: Pyrogram client
        message: Tin nhắn từ user
        command_name: Tên lệnh hiển thị trên form (vd: "/tien_nga_add_employee")
    """
    lines = message.text.strip().split("\n")
    
    # Nếu chỉ gõ lệnh không tham số -> hiển thị form mẫu
    if len(lines) < 3:
        form = EMPLOYEE_FORM_TEMPLATE.format(cmd=command_name)
        await message.reply_text(form, parse_mode=ParseMode.HTML)
        return

    # Parse form data
    data = _parse_form_lines(lines[1:])

    # Validate required fields
    emp_id = data.get("Mã NV", "").strip()
    last_name = data.get("Họ", "").strip()
    first_name = data.get("Tên", "").strip()
    username = data.get("Username", "").strip().lstrip("@")

    if not emp_id or not last_name or not first_name or not username:
        await message.reply_text(
            "⚠️ <b>Mã NV</b>, <b>Họ</b>, <b>Tên</b> và <b>Username</b> là bắt buộc.",
            parse_mode=ParseMode.HTML
        )
        return

    # Parse ngày sinh
    birthday = None
    birthday_str = data.get("Ngày sinh", "").strip()
    if birthday_str:
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d"):
            try:
                birthday = datetime.datetime.strptime(birthday_str, fmt).date()
                break
            except ValueError:
                continue

    try:
        base_salary = await _parse_float_or_reply(message, "Lương cơ bản", data.get("Lương cơ bản (VNĐ)", data.get("Lương cơ bản", "")).strip())
        bonus = await _parse_float_or_reply(message, "Tiền thưởng", data.get("Tiền thưởng (VNĐ)", data.get("Tiền thưởng", "")).strip())
        monthly_salary = await _parse_float_or_reply(message, "Lương tháng", data.get("Lương tháng (VNĐ)", data.get("Lương tháng", "")).strip())
        weekly_salary = await _parse_float_or_reply(message, "Lương tuần", data.get("Lương tuần (VNĐ)", data.get("Lương tuần", "")).strip())
        daily_salary = await _parse_float_or_reply(message, "Lương ngày", data.get("Lương ngày (VNĐ)", data.get("Lương ngày", "")).strip())
        hourly_salary = await _parse_float_or_reply(message, "Lương giờ", data.get("Lương giờ (VNĐ)", data.get("Lương giờ", "")).strip())
        overtime_salary = await _parse_float_or_reply(message, "Lương làm thêm giờ", data.get("Lương làm thêm giờ (VNĐ)", data.get("Lương làm thêm giờ", "")).strip())
        rate_bhxh = await _parse_float_or_reply(message, "Tỷ lệ BHXH", data.get("Tỷ lệ BHXH (%)", data.get("Tỷ lệ BHXH", "")).strip())
        working_hours = await _parse_float_or_reply(message, "Số giờ làm việc", data.get("Số giờ làm việc (giờ/ngày)", data.get("Số giờ làm việc", "")).strip())
        leave_balance = await _parse_int_or_reply(message, "Số ngày phép năm", data.get("Số ngày phép năm", "").strip())
        work_type = await _parse_int_or_reply(message, "Loại công", data.get("Loại công (1-4)", data.get("Loại công", "")).strip())
        
        start_time = await _parse_time_or_reply(message, "Giờ vào ca", data.get("Giờ vào ca (hh:mm)", data.get("Giờ vào ca", "")).strip())
        end_time = await _parse_time_or_reply(message, "Giờ tan ca", data.get("Giờ tan ca (hh:mm)", data.get("Giờ tan ca", "")).strip())
        sat_start_time = await _parse_time_or_reply(message, "Giờ vào ca T7", data.get("Giờ vào ca T7 (hh:mm)", data.get("Giờ vào ca T7", "")).strip())
        sat_end_time = await _parse_time_or_reply(message, "Giờ tan ca T7", data.get("Giờ tan ca T7 (hh:mm)", data.get("Giờ tan ca T7", "")).strip())
    except ValueError:
        return

    db = SessionLocal()
    try:
        chat_id = str(message.chat.id)
        
        # Lấy project_id từ nhóm hiện tại
        current_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()

        if not current_member:
            await message.reply_text(
                "⚠️ Nhóm này chưa được đồng bộ. Vui lòng sử dụng lệnh /syncchat trước.",
                parse_mode=ParseMode.HTML
            )
            return

        company_id = current_member.project_id

        # Check duplicate mã NV
        existing_id = db.query(Employee).filter(Employee.id == emp_id).first()
        if existing_id:
            await message.reply_text(
                f"⚠️ Mã nhân viên <b>{emp_id}</b> đã tồn tại trong hệ thống.",
                parse_mode=ParseMode.HTML
            )
            return

        # Check duplicate username
        existing = db.query(Employee).filter(Employee.username == username).first()
        if existing:
            await message.reply_text(
                f"⚠️ Nhân viên có username <b>@{username}</b> đã tồn tại trong hệ thống.",
                parse_mode=ParseMode.HTML
            )
            return

        new_employee = Employee(
            id=emp_id,
            username=username,
            authority=data.get("Ủy quyền", "").strip().lstrip("@") or None,
            telegram_group=data.get("Nhóm Telegram", "").strip() or None,
            last_name=last_name,
            first_name=first_name,
            gender=data.get("Giới tính", "").strip() or None,
            birthday=birthday,
            number_phone=data.get("SĐT", "").strip() or None,
            email=data.get("Email", "").strip() or None,
            address=data.get("Địa chỉ", "").strip() or None,
            identity_card=data.get("CCCD/CMND", "").strip() or None,
            place_of_issue=data.get("Nơi cấp", "").strip() or None,
            nationality=data.get("Quốc tịch", "").strip() or None,
            marital_status=data.get("Tình trạng hôn nhân", "").strip() or None,
            status="active",
            experience=data.get("Kinh nghiệm", "").strip() or None,
            company_id=company_id,
            education_level=data.get("Trình độ học vấn", "").strip() or None,
            major=data.get("Chuyên ngành", "").strip() or None,
            certificates=data.get("Chứng chỉ", "").strip() or None,
            position=data.get("Chức vụ", "").strip() or None,
            department=data.get("Phòng ban", "").strip() or None,
            contract_type=data.get("Loại hợp đồng", "").strip() or None,
            base_salary=base_salary,
            leave_balance=leave_balance,
            insurance=data.get("Bảo hiểm", "").strip() or None,
            bank_name=data.get("Ngân hàng", "").strip() or None,
            bank_account_number=data.get("Số tài khoản", "").strip() or None,
            code_payment=data.get("Mã thanh toán", "").strip() or None,
            emergency_phone=data.get("SĐT khẩn cấp", "").strip() or None,
            emergency_contact=data.get("Người liên hệ khẩn cấp", "").strip() or None,
            employee_photo=data.get("Ảnh nhân viên", "").strip() or None,
            working_hours=working_hours,
            performance_evaluation=data.get("Đánh giá hiệu suất", "").strip() or None,
            career_goal=data.get("Mục tiêu nghề nghiệp", "").strip() or None,
            benefits=data.get("Phúc lợi", "").strip() or None,
            bonus=bonus,
            monthly_salary=monthly_salary,
            weekly_salary=weekly_salary,
            daily_salary=daily_salary,
            hourly_salary=hourly_salary,
            overtime_salary=overtime_salary,
            rate_bhxh=rate_bhxh,
            auto_attendance=data.get("Auto chấm công (có/không)", "có").strip().lower() in ("có", "co", "yes", "true", "1"),
            work_type=work_type if work_type and 1 <= work_type <= 4 else 3,
            total_debt=0,
            start_time=start_time,
            end_time=end_time,
            sat_start_time=sat_start_time,
            sat_end_time=sat_end_time,
            created_at=datetime.datetime.now(),
            updated_at=datetime.datetime.now(),
        )
        db.add(new_employee)
        db.commit()

        # Build result message
        full_name = f"{last_name} {first_name}"
        result = (
            f"<b>Thêm nhân viên thành công!</b>\n\n"
            f"<b>Họ tên:</b> {full_name}\n"
            f"<b>Username:</b> @{username}\n"
        )
        if data.get("Phòng ban", "").strip():
            result += f"<b>Phòng ban:</b> {data['Phòng ban'].strip()}\n"
        if data.get("Chức vụ", "").strip():
            result += f"<b>Chức vụ:</b> {data['Chức vụ'].strip()}\n"
        if data.get("SĐT", "").strip():
            result += f"<b>SĐT:</b> {data['SĐT'].strip()}\n"
        if data.get("Email", "").strip():
            result += f"<b>Email:</b> {data['Email'].strip()}\n"
        
        result += f"\n<i>Mã NV: <code>{emp_id}</code></i>"
        
        await message.reply_text(result, parse_mode=ParseMode.HTML)
        LogInfo(
            f"[AddEmployee] Created {full_name} (@{username}) by user {message.from_user.id}",
            LogType.SYSTEM_STATUS
        )

    except Exception as e:
        db.rollback()
        LogError(f"Error in handle_create_employee: {e}", LogType.SYSTEM_STATUS)
        from sqlalchemy.exc import IntegrityError
        if isinstance(e, IntegrityError) and "ix_employee_email" in str(e):
            await message.reply_text("❌ Email này đã được sử dụng cho nhân viên khác. Vui lòng kiểm tra lại.", parse_mode=ParseMode.HTML)
        elif isinstance(e, IntegrityError) and "ix_employee_number_phone" in str(e):
            await message.reply_text("❌ Số điện thoại này đã được sử dụng cho nhân viên khác. Vui lòng kiểm tra lại.", parse_mode=ParseMode.HTML)
        elif isinstance(e, IntegrityError):
            await message.reply_text("❌ Lỗi dữ liệu trùng lặp (có thể email, SĐT, CCCD đã tồn tại). Vui lòng kiểm tra lại.", parse_mode=ParseMode.HTML)
        else:
            await message.reply_text(
                "❌ Có lỗi xảy ra khi thêm nhân viên. Vui lòng thử lại.",
                parse_mode=ParseMode.HTML
            )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════
# UPDATE EMPLOYEE
# ══════════════════════════════════════════════════════════════

# Mapping key form -> field Employee model
FIELD_MAP = {
    "Họ": "last_name",
    "Tên": "first_name",
    "Ủy quyền": "authority",
    "Nhóm Telegram": "telegram_group",
    "Giới tính": "gender",
    "SĐT": "number_phone",
    "Email": "email",
    "Địa chỉ": "address",
    "CCCD/CMND": "identity_card",
    "Nơi cấp": "place_of_issue",
    "Quốc tịch": "nationality",
    "Tình trạng hôn nhân": "marital_status",
    "Trình độ học vấn": "education_level",
    "Chuyên ngành": "major",
    "Chứng chỉ": "certificates",
    "Kinh nghiệm": "experience",
    "Phòng ban": "department",
    "Chức vụ": "position",
    "Loại hợp đồng": "contract_type",
    "Ảnh nhân viên": "employee_photo",
    "Phúc lợi": "benefits",
    "Bảo hiểm": "insurance",
    "Mục tiêu nghề nghiệp": "career_goal",
    "Đánh giá hiệu suất": "performance_evaluation",
    "Ngân hàng": "bank_name",
    "Số tài khoản": "bank_account_number",
    "Mã thanh toán": "code_payment",
    "SĐT khẩn cấp": "emergency_phone",
    "Người liên hệ khẩn cấp": "emergency_contact",
    "Trạng thái": "status",
}


def _build_prefilled_update_form(emp: "Employee", command_name: str) -> str:
    """Tạo form cập nhật đã điền sẵn dữ liệu hiện tại từ DB."""
    birthday_str = emp.birthday.strftime("%d-%m-%Y") if emp.birthday else ""

    def _fmt_salary(val):
        if val is None: return ""
        int_val = int(val) if val == int(val) else val
        return f"{int_val:,}".replace(",", ".")

    salary_str = _fmt_salary(emp.base_salary)
    monthly_salary_str = _fmt_salary(emp.monthly_salary)
    weekly_salary_str = _fmt_salary(emp.weekly_salary)
    daily_salary_str = _fmt_salary(emp.daily_salary)
    hourly_salary_str = _fmt_salary(emp.hourly_salary)
    overtime_salary_str = _fmt_salary(emp.overtime_salary)
    bonus_str = _fmt_salary(emp.bonus)
    rate_bhxh_str = ""
    if emp.rate_bhxh is not None:
        rate_bhxh_str = str(int(emp.rate_bhxh) if emp.rate_bhxh == int(emp.rate_bhxh) else emp.rate_bhxh)
    working_hours_str = ""
    if emp.working_hours is not None:
        working_hours_str = str(int(emp.working_hours) if emp.working_hours == int(emp.working_hours) else emp.working_hours)
    
    start_time_str = emp.start_time.strftime("%H:%M") if emp.start_time else ""
    end_time_str = emp.end_time.strftime("%H:%M") if emp.end_time else ""
    sat_start_time_str = emp.sat_start_time.strftime("%H:%M") if emp.sat_start_time else ""
    sat_end_time_str = emp.sat_end_time.strftime("%H:%M") if emp.sat_end_time else ""

    form = f"""<b>FORM CẬP NHẬT NHÂN VIÊN</b>
Dữ liệu hiện tại đã được điền sẵn. Chỉ sửa các trường cần thay đổi rồi gửi lại:

<pre>{command_name}
Username: @{emp.username or ''}
Ủy quyền: @{emp.authority or ''}
Họ: {emp.last_name or ''}
Tên: {emp.first_name or ''}
Nhóm Telegram: {emp.telegram_group or ''}
Giới tính: {emp.gender or ''}
Ngày sinh: {birthday_str}
SĐT: {emp.number_phone or ''}
Email: {emp.email or ''}
Địa chỉ: {emp.address or ''}
CCCD/CMND: {emp.identity_card or ''}
Nơi cấp: {emp.place_of_issue or ''}
Quốc tịch: {emp.nationality or ''}
Tình trạng hôn nhân: {emp.marital_status or ''}
Trình độ học vấn: {emp.education_level or ''}
Chuyên ngành: {emp.major or ''}
Chứng chỉ: {emp.certificates or ''}
Kinh nghiệm: {emp.experience or ''}
Phòng ban: {emp.department or ''}
Chức vụ: {emp.position or ''}
Loại hợp đồng: {emp.contract_type or ''}
Ảnh nhân viên: {emp.employee_photo or ''}
Giờ vào ca (hh:mm): {start_time_str}
Giờ tan ca (hh:mm): {end_time_str}
Giờ vào ca T7 (hh:mm): {sat_start_time_str}
Giờ tan ca T7 (hh:mm): {sat_end_time_str}
Số giờ làm việc (giờ/ngày): {working_hours_str}
Lương cơ bản (VNĐ): {salary_str}
Lương tháng (VNĐ): {monthly_salary_str}
Lương tuần (VNĐ): {weekly_salary_str}
Lương ngày (VNĐ): {daily_salary_str}
Lương giờ (VNĐ): {hourly_salary_str}
Lương làm thêm giờ (VNĐ): {overtime_salary_str}
Tiền thưởng (VNĐ): {bonus_str}
Phúc lợi: {emp.benefits or ''}
Số ngày phép năm: {emp.leave_balance if emp.leave_balance is not None else ''}
Bảo hiểm: {emp.insurance or ''}
Tỷ lệ BHXH (%): {rate_bhxh_str}
Mục tiêu nghề nghiệp: {emp.career_goal or ''}
Đánh giá hiệu suất: {emp.performance_evaluation or ''}
Ngân hàng: {emp.bank_name or ''}
Số tài khoản: {emp.bank_account_number or ''}
Mã thanh toán: {emp.code_payment or ''}
SĐT khẩn cấp: {emp.emergency_phone or ''}
Người liên hệ khẩn cấp: {emp.emergency_contact or ''}
Auto chấm công (có/không): {'có' if emp.auto_attendance else 'không'}
Loại công (1-4): {emp.work_type if emp.work_type is not None else 3}
Trạng thái: {emp.status or ''}</pre>

<i>Mã NV: <b>{emp.id}</b> | Chỉ sửa trường cần thay đổi.</i>"""
    return form


async def handle_update_employee(client, message: Message, command_name: str) -> None:
    """
    Xử lý cập nhật thông tin nhân viên - dùng chung.
    
    Flow:
    1. /command              -> Hiện hướng dẫn
    2. /command @username    -> Tra cứu DB, hiện form đã điền sẵn dữ liệu hiện tại
    3. /command + form data  -> Thực hiện cập nhật
    """
    lines = message.text.strip().split("\n")
    parts = lines[0].split()

    # Case 1: Chỉ gõ lệnh, không tham số -> hướng dẫn
    if len(lines) < 3 and len(parts) < 2:
        await message.reply_text(
            f"<b>FORM CẬP NHẬT NHÂN VIÊN</b>\n\n"
            f"Cú pháp: <code>{command_name} @username</code> hoặc <code>{command_name} [Mã nhân viên]</code>\n\n"
            f"<i>Bot sẽ tra cứu dữ liệu nhân viên và hiện form đã điền sẵn để bạn chỉnh sửa.</i>",
            parse_mode=ParseMode.HTML
        )
        return

    # Case 2: /command @username -> tra cứu và hiện form pre-filled
    if len(lines) < 3 and len(parts) >= 2:
        lookup = parts[1].strip().lstrip("@")
        db = SessionLocal()
        try:
            emp = db.query(Employee).filter(
                (Employee.username == lookup) | (Employee.id == lookup)
            ).first()
            if not emp:
                await message.reply_text(
                    f"⚠️ Không tìm thấy nhân viên <b>{lookup}</b>.",
                    parse_mode=ParseMode.HTML
                )
                return
            form = _build_prefilled_update_form(emp, command_name)
            await message.reply_text(form, parse_mode=ParseMode.HTML)
            return
        except Exception as e:
            LogError(f"Error looking up employee: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Có lỗi khi tra cứu nhân viên.", parse_mode=ParseMode.HTML)
            return
        finally:
            db.close()

    # Parse form
    data = _parse_form_lines(lines[1:])

    username = data.get("Username", "").strip().lstrip("@")
    if not username:
        await message.reply_text(
            "⚠️ <b>Username</b> là bắt buộc để tìm nhân viên cần cập nhật.",
            parse_mode=ParseMode.HTML
        )
        return

    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.username == username).first()
        if not employee:
            await message.reply_text(
                f"⚠️ Không tìm thấy nhân viên có username <b>@{username}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        # Cập nhật các trường có giá trị
        updated_fields = []

        for form_key, model_field in FIELD_MAP.items():
            val = data.get(form_key, "").strip()
            if model_field == "authority":
                val_clean = val.lstrip("@").strip()
                if val_clean:
                    employee.authority = val_clean
                    updated_fields.append(form_key)
                elif val == "" or val == "@":
                    employee.authority = None
                    updated_fields.append(form_key)
            else:
                if val:
                    setattr(employee, model_field, val)
                    updated_fields.append(form_key)

        # Xử lý riêng: Ngày sinh
        birthday_str = data.get("Ngày sinh", "").strip()
        if birthday_str:
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d"):
                try:
                    employee.birthday = datetime.datetime.strptime(birthday_str, fmt).date()
                    updated_fields.append("Ngày sinh")
                    break
                except ValueError:
                    continue

        try:
            # Xử lý riêng: Lương cơ bản
            salary_str = data.get("Lương cơ bản (VNĐ)", data.get("Lương cơ bản", "")).strip()
            if salary_str:
                employee.base_salary = await _parse_float_or_reply(message, "Lương cơ bản", salary_str)
                updated_fields.append("Lương cơ bản")

            # Xử lý riêng: Lương tháng
            monthly_str = data.get("Lương tháng (VNĐ)", data.get("Lương tháng", "")).strip()
            if monthly_str:
                employee.monthly_salary = await _parse_float_or_reply(message, "Lương tháng", monthly_str)
                updated_fields.append("Lương tháng")

            # Xử lý riêng: Lương tuần
            weekly_str = data.get("Lương tuần (VNĐ)", data.get("Lương tuần", "")).strip()
            if weekly_str:
                employee.weekly_salary = await _parse_float_or_reply(message, "Lương tuần", weekly_str)
                updated_fields.append("Lương tuần")

            # Xử lý riêng: Lương ngày
            daily_str = data.get("Lương ngày (VNĐ)", data.get("Lương ngày", "")).strip()
            if daily_str:
                employee.daily_salary = await _parse_float_or_reply(message, "Lương ngày", daily_str)
                updated_fields.append("Lương ngày")

            # Xử lý riêng: Lương giờ
            hourly_str = data.get("Lương giờ (VNĐ)", data.get("Lương giờ", "")).strip()
            if hourly_str:
                employee.hourly_salary = await _parse_float_or_reply(message, "Lương giờ", hourly_str)
                updated_fields.append("Lương giờ")

            # Xử lý riêng: Lương làm thêm giờ
            overtime_str = data.get("Lương làm thêm giờ (VNĐ)", data.get("Lương làm thêm giờ", "")).strip()
            if overtime_str:
                employee.overtime_salary = await _parse_float_or_reply(message, "Lương làm thêm giờ", overtime_str)
                updated_fields.append("Lương làm thêm giờ")

            # Xử lý riêng: Tiền thưởng
            bonus_str = data.get("Tiền thưởng (VNĐ)", data.get("Tiền thưởng", "")).strip()
            if bonus_str:
                employee.bonus = await _parse_float_or_reply(message, "Tiền thưởng", bonus_str)
                updated_fields.append("Tiền thưởng")

            # Xử lý riêng: Tỷ lệ BHXH
            bhxh_str = data.get("Tỷ lệ BHXH (%)", data.get("Tỷ lệ BHXH", "")).strip()
            if bhxh_str:
                employee.rate_bhxh = await _parse_float_or_reply(message, "Tỷ lệ BHXH", bhxh_str)
                updated_fields.append("Tỷ lệ BHXH")

            # Xử lý riêng: Số giờ làm việc
            working_hours_str = data.get("Số giờ làm việc (giờ/ngày)", data.get("Số giờ làm việc", "")).strip()
            if working_hours_str:
                employee.working_hours = await _parse_float_or_reply(message, "Số giờ làm việc", working_hours_str)
                updated_fields.append("Số giờ làm việc")

            # Xử lý riêng: Giờ vào ca
            start_str = data.get("Giờ vào ca (hh:mm)", data.get("Giờ vào ca", "")).strip()
            if start_str:
                employee.start_time = await _parse_time_or_reply(message, "Giờ vào ca", start_str)
                updated_fields.append("Giờ vào ca")

            # Xử lý riêng: Giờ tan ca
            end_str = data.get("Giờ tan ca (hh:mm)", data.get("Giờ tan ca", "")).strip()
            if end_str:
                employee.end_time = await _parse_time_or_reply(message, "Giờ tan ca", end_str)
                updated_fields.append("Giờ tan ca")

            # Xử lý riêng: Giờ vào ca T7
            sat_start_str = data.get("Giờ vào ca T7 (hh:mm)", data.get("Giờ vào ca T7", "")).strip()
            if sat_start_str:
                employee.sat_start_time = await _parse_time_or_reply(message, "Giờ vào ca T7", sat_start_str)
                updated_fields.append("Giờ vào ca T7")

            # Xử lý riêng: Giờ tan ca T7
            sat_end_str = data.get("Giờ tan ca T7 (hh:mm)", data.get("Giờ tan ca T7", "")).strip()
            if sat_end_str:
                employee.sat_end_time = await _parse_time_or_reply(message, "Giờ tan ca T7", sat_end_str)
                updated_fields.append("Giờ tan ca T7")

            # Xử lý riêng: Số ngày phép năm
            leave_str = data.get("Số ngày phép năm", "").strip()
            if leave_str:
                employee.leave_balance = await _parse_int_or_reply(message, "Số ngày phép năm", leave_str)
                updated_fields.append("Số ngày phép năm")

            # Xử lý riêng: Auto chấm công
            auto_att_str = data.get("Auto chấm công (có/không)", "").strip()
            if auto_att_str:
                employee.auto_attendance = auto_att_str.lower() in ("có", "co", "yes", "true", "1")
                updated_fields.append("Auto chấm công")

            # Xử lý riêng: Loại công
            work_type_str = data.get("Loại công (1-4)", data.get("Loại công", "")).strip()
            if work_type_str:
                wt = await _parse_int_or_reply(message, "Loại công", work_type_str)
                if wt and 1 <= wt <= 4:
                    employee.work_type = wt
                    updated_fields.append("Loại công")
                else:
                    await message.reply_text("⚠️ <b>Loại công</b> phải từ 1 đến 4 (1=T2-T6, 2=T2-T7 sáng, 3=T2-T7, 4=T2-CN).", parse_mode=ParseMode.HTML)
                    return
        except ValueError:
            return

        if not updated_fields:
            await message.reply_text(
                "⚠️ Không có trường nào được điền để cập nhật. Vui lòng điền ít nhất 1 trường.",
                parse_mode=ParseMode.HTML
            )
            return

        employee.updated_at = datetime.datetime.now()
        db.commit()

        fields_str = ", ".join([f"<b>{f}</b>" for f in updated_fields])
        full_name = f"{employee.last_name} {employee.first_name}"
        result = (
            f"<b>Cập nhật nhân viên thành công!</b>\n\n"
            f"<b>Nhân viên:</b> {full_name} (@{username})\n"
            f"<b>Đã cập nhật {len(updated_fields)} trường:</b> {fields_str}"
        )
        await message.reply_text(result, parse_mode=ParseMode.HTML)
        LogInfo(
            f"[UpdateEmployee] Updated @{username} fields: {', '.join(updated_fields)} by user {message.from_user.id}",
            LogType.SYSTEM_STATUS
        )

    except Exception as e:
        db.rollback()
        LogError(f"Error in handle_update_employee: {e}", LogType.SYSTEM_STATUS)
        from sqlalchemy.exc import IntegrityError
        if isinstance(e, IntegrityError) and "ix_employee_email" in str(e):
            await message.reply_text("❌ Email này đã được sử dụng cho nhân viên khác. Vui lòng kiểm tra lại.", parse_mode=ParseMode.HTML)
        elif isinstance(e, IntegrityError) and "ix_employee_number_phone" in str(e):
            await message.reply_text("❌ Số điện thoại này đã được sử dụng cho nhân viên khác. Vui lòng kiểm tra lại.", parse_mode=ParseMode.HTML)
        elif isinstance(e, IntegrityError):
            await message.reply_text("❌ Lỗi dữ liệu trùng lặp (có thể email, SĐT, CCCD đã tồn tại). Vui lòng kiểm tra lại.", parse_mode=ParseMode.HTML)
        else:
            await message.reply_text(
                "❌ Có lỗi xảy ra khi cập nhật nhân viên.",
                parse_mode=ParseMode.HTML
            )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════
# DELETE EMPLOYEE (soft delete -> set status = 'inactive')
# ══════════════════════════════════════════════════════════════

# Pending delete-employee confirmations: {msg_id: emp_id}
_PENDING_DEL_EMP: dict = {}

async def handle_delete_employee(client, message: Message, command_name: str) -> None:
    """
    Xóa nhân viên (soft delete) - chỉ chuyển trạng thái sang 'inactive'.
    Có bước xác nhận bằng Inline Button.
    Cú pháp: /command MÃ_NV
    """
    from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton

    lines = message.text.strip().split("\n")
    parts = lines[0].split()

    if len(parts) < 2:
        await message.reply_text(
            f"<b>XÓA NHÂN VIÊN</b>\n\n"
            f"Cú pháp: <code>{command_name} MÃ_NV</code>\n"
            f"Ví dụ: <code>{command_name} TN001</code>\n\n"
            f"<i>Lưu ý: Lệnh này chỉ chuyển trạng thái nhân viên sang <b>inactive</b>, "
            f"không xóa hoàn toàn khỏi hệ thống.</i>",
            parse_mode=ParseMode.HTML
        )
        return

    emp_id = parts[1].strip()

    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.id == emp_id).first()
        if not employee:
            await message.reply_text(
                f"⚠️ Không tìm thấy nhân viên có mã <b>{emp_id}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        if employee.status == "inactive":
            await message.reply_text(
                f"⚠️ Nhân viên <b>{emp_id}</b> đã ở trạng thái <b>inactive</b> rồi.",
                parse_mode=ParseMode.HTML
            )
            return

        full_name = f"{employee.last_name} {employee.first_name}"
        username = employee.username or "N/A"

        confirm_text = (
            f"<b>XÁC NHẬN XÓA NHÂN VIÊN</b>\n\n"
            f"<b>Mã NV:</b> {emp_id}\n"
            f"<b>Họ tên:</b> {full_name} (@{username})\n"
            f"<b>Trạng thái hiện tại:</b> {employee.status or 'active'}\n\n"
            f"<i>Nhân viên sẽ được chuyển sang trạng thái <b>inactive</b> (nghỉ việc).\n"
            f"Thao tác này không xóa hoàn toàn khỏi hệ thống.</i>\n\n"
            f"Bạn có chắc chắn muốn thực hiện?"
        )
        markup = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("Xác nhận", callback_data=f"del_emp_confirm|{emp_id}"),
                InlineKeyboardButton("Huỷ", callback_data=f"del_emp_cancel|{emp_id}"),
            ]
        ])
        sent = await message.reply_text(confirm_text, reply_markup=markup, parse_mode=ParseMode.HTML)
        _PENDING_DEL_EMP[sent.id] = emp_id

    except Exception as e:
        LogError(f"Error in handle_delete_employee: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra khi tra cứu nhân viên.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def handle_delete_employee_callback(client, callback_query) -> None:
    """Xử lý callback xác nhận / huỷ xóa nhân viên."""
    data = callback_query.data  # del_emp_confirm|TN001 hoặc del_emp_cancel|TN001
    action, emp_id = data.split("|", 1)
    msg_id = callback_query.message.id

    if action == "del_emp_cancel":
        _PENDING_DEL_EMP.pop(msg_id, None)
        await callback_query.message.edit_text(
            f"❌ <b>Đã huỷ thao tác xóa nhân viên {emp_id}.</b>",
            parse_mode=ParseMode.HTML
        )
        await callback_query.answer()
        return

    # action == del_emp_confirm
    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.id == emp_id).first()
        if not employee:
            await callback_query.message.edit_text(
                f"⚠️ Không tìm thấy nhân viên <b>{emp_id}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        full_name = f"{employee.last_name} {employee.first_name}"
        username = employee.username or "N/A"
        old_status = employee.status or "active"

        employee.status = "inactive"
        employee.updated_at = datetime.datetime.now()
        db.commit()

        result = (
            f"✅ <b>Đã xóa nhân viên!</b>\n\n"
            f"👤 <b>Mã NV:</b> {emp_id}\n"
            f"👤 <b>Họ tên:</b> {full_name} (@{username})\n"
            f"📋 <b>Trạng thái:</b> {old_status} → <b>inactive</b>\n\n"
            f"<i>Nhân viên không bị xóa khỏi hệ thống, chỉ chuyển sang trạng thái nghỉ việc.</i>"
        )
        await callback_query.message.edit_text(result, parse_mode=ParseMode.HTML)
        _PENDING_DEL_EMP.pop(msg_id, None)
        LogInfo(
            f"[DeleteEmployee] Deactivated {emp_id} - {full_name} (@{username}) by user {callback_query.from_user.id}",
            LogType.SYSTEM_STATUS
        )

    except Exception as e:
        db.rollback()
        LogError(f"Error in handle_delete_employee_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.message.edit_text("❌ Có lỗi xảy ra khi xóa nhân viên.", parse_mode=ParseMode.HTML)
    finally:
        db.close()
    await callback_query.answer()




# ══════════════════════════════════════════════════════════════
# AUTHORITY DELEGATION (Ủy quyền)
# ══════════════════════════════════════════════════════════════

# Lưu trữ tạm dữ liệu khi user chọn nhân viên ủy quyền từ Inline Keyboard
# key: message_id của tin nhắn chứa Inline Keyboard
# value: {"action": str, "form_data": dict|None, "command_name": str,
#          "chat_id": int, "from_user_id": int, "original_message": Message}
_PENDING_AUTHORITY_ACTIONS: dict[int, dict] = {}


async def resolve_employee_for_command(
    client, message: Message, db, callback_prefix: str, command_name: str,
    form_data: dict | None = None,
) -> Employee | None:
    """
    Xác định Employee cho lệnh chấm công/nghỉ phép, hỗ trợ ủy quyền (authority).

    Logic:
    1. Lấy telegram_username từ message.from_user.username
    2. Tìm Employee.username == telegram_username (chính chủ)
    3. Tìm tất cả Employee.authority == telegram_username (được ủy quyền)
    4. Nếu chỉ có 1 kết quả → return trực tiếp
    5. Nếu có nhiều kết quả → hiện Inline Keyboard chọn nhân viên → return None
    6. Nếu không có kết quả → báo lỗi → return None

    Args:
        client: Pyrogram client
        message: Tin nhắn gốc từ user
        db: Database session
        callback_prefix: Prefix cho callback_data (ci/co/lv/ov)
        command_name: Tên lệnh gốc (vd: "/tien_nga_check_in")
        form_data: Dữ liệu form đã parse (cho request_leave/request_overtime), None cho check_in/check_out

    Returns:
        Employee nếu chỉ có 1 candidate, None nếu cần chờ callback hoặc lỗi
    """
    from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton

    username = message.from_user.username
    if not username:
        await message.reply_text(
            "⚠️ Bạn chưa cài đặt username Telegram. Vui lòng cài đặt username trước.",
            parse_mode=ParseMode.HTML
        )
        return None

    candidates = []

    # 1. Tìm chính chủ (Employee.username == telegram_username)
    employee_self = db.query(Employee).filter(
        Employee.username == username,
        Employee.status != "inactive"
    ).first()
    if employee_self:
        candidates.append(("self", employee_self))

    # 2. Tìm nhân viên ủy quyền (Employee.authority == telegram_username)
    delegated = db.query(Employee).filter(
        Employee.authority == username,
        Employee.status != "inactive"
    ).all()
    for emp in delegated:
        candidates.append(("delegated", emp))

    # 3. Xử lý kết quả
    if not candidates:
        await message.reply_text(
            f"⚠️ Không tìm thấy nhân viên nào liên kết với <b>@{username}</b> trong hệ thống.\n"
            "Vui lòng liên hệ quản lý để được thêm vào hệ thống.",
            parse_mode=ParseMode.HTML
        )
        return None

    if len(candidates) == 1:
        return candidates[0][1]

    # 4. Nhiều candidates → hiện Inline Keyboard
    ACTION_LABELS = {
        "ci": "check-in",
        "co": "check-out",
        "lv": "xin nghỉ phép",
        "ov": "đăng ký tăng ca",
        "au": "yêu cầu cập nhật công",
    }
    action_label = ACTION_LABELS.get(callback_prefix, callback_prefix)

    buttons = []
    for tag, emp in candidates:
        full_name = f"{emp.last_name} {emp.first_name}"
        label_suffix = " (Bạn)" if tag == "self" else " (UQ)"
        btn_text = f"{full_name} - {emp.id}{label_suffix}"
        buttons.append([
            InlineKeyboardButton(
                btn_text,
                callback_data=f"auth_{callback_prefix}|{emp.id}"
            )
        ])
    buttons.append([InlineKeyboardButton("❌ Hủy", callback_data=f"auth_{callback_prefix}|cancel")])

    markup = InlineKeyboardMarkup(buttons)
    sent = await message.reply_text(
        f"<b>Chọn nhân viên để {action_label}:</b>",
        reply_markup=markup,
        parse_mode=ParseMode.HTML
    )

    # Lưu context để callback xử lý tiếp
    _PENDING_AUTHORITY_ACTIONS[sent.id] = {
        "action": callback_prefix,
        "form_data": form_data,
        "command_name": command_name,
        "chat_id": message.chat.id,
        "from_user_id": message.from_user.id,
        "original_message": message,
    }

    return None


async def _execute_check_in_for_employee(client, message: Message, employee: Employee, db, acting_username: str) -> None:
    """Thực hiện logic check-in cho 1 Employee cụ thể (tách riêng từ handle_check_in)."""
    from app.models.finance import Attendance

    now_vn = datetime.datetime.now(VN_TZ)
    today = now_vn.date()
    current_time = now_vn.time()

    # Kiểm tra start_time (giờ vào ca)
    if not employee.start_time:
        await message.reply_text(
            f"⚠️ Nhân viên <b>{employee.last_name} {employee.first_name}</b> chưa được cài đặt giờ vào ca. Vui lòng liên hệ quản lý.",
            parse_mode=ParseMode.HTML
        )
        return

    # Lấy giờ vào ca từ start_time (DateTime -> Time)
    shift_time = employee.start_time.time() if isinstance(employee.start_time, datetime.datetime) else employee.start_time
    shift_start_str = shift_time.strftime("%H:%M")

    # Tính cửa sổ check-in: start_time - 30 phút → start_time + 30 phút
    shift_dt = datetime.datetime.combine(today, shift_time)
    window_start = (shift_dt - datetime.timedelta(minutes=30)).time()
    window_end = (shift_dt + datetime.timedelta(minutes=30)).time()

    if not (window_start <= current_time <= window_end):
        await message.reply_text(
            f"⚠️ <b>Ngoài thời gian check-in!</b>\n\n"
            f"Ca của nhân viên: <b>{shift_start_str}</b>\n"
            f"Thời gian cho phép: <b>{window_start.strftime('%H:%M')} - {window_end.strftime('%H:%M')}</b>\n"
            f"Hiện tại: <b>{current_time.strftime('%H:%M')}</b>",
            parse_mode=ParseMode.HTML
        )
        return

    # Kiểm tra đã check-in hôm nay chưa
    existing = db.query(Attendance).filter(
        Attendance.employee_id == employee.id,
        Attendance.year == today.year,
        Attendance.month == today.month,
        Attendance.day == today.day,
        Attendance.check_in_time.isnot(None)
    ).first()

    if existing:
        checkin_str = existing.check_in_time.strftime("%H:%M:%S") if existing.check_in_time else "N/A"
        full_name = f"{employee.last_name} {employee.first_name}"
        await message.reply_text(
            f"⚠️ Nhân viên <b>{full_name}</b> đã check-in hôm nay rồi!\n\n"
            f"Thời gian check-in: <b>{checkin_str}</b>",
            parse_mode=ParseMode.HTML
        )
        return

    # Tính trạng thái
    late_minutes = 0
    status = "on_time"

    if current_time > shift_time:
        now_dt = datetime.datetime.combine(today, current_time.replace(second=0, microsecond=0))
        shift_dt_full = datetime.datetime.combine(today, shift_time.replace(second=0, microsecond=0))
        diff = (now_dt - shift_dt_full).total_seconds() / 60
        late_minutes = int(diff)
        if late_minutes > 0:
            status = "late"
    elif current_time < shift_time:
        status = "early"

    # Tạo Attendance record
    day_name = DAY_NAMES_VN.get(today.weekday(), "")

    new_attendance = Attendance(
        employee_id=employee.id,
        year=today.year,
        month=today.month,
        day=today.day,
        date_str=day_name,
        check_in_time=now_vn.replace(tzinfo=None),
        late_time=float(late_minutes),
    )
    db.add(new_attendance)
    db.commit()

    # Reply
    full_name = f"{employee.last_name} {employee.first_name}"
    emp_username = employee.username or employee.id
    delegated_note = ""
    if acting_username != employee.username:
        delegated_note = f"\n<b>Người thực hiện:</b> @{acting_username} (ủy quyền)"

    result = (
        f"<b>CHECK-IN THÀNH CÔNG</b>\n\n"
        f"<b>Nhân viên:</b> {full_name} (@{emp_username})\n"
        f"<b>Mã NV:</b> {employee.id}\n"
        f"<b>Ngày:</b> {day_name}, {today.strftime('%d-%m-%Y')}\n"
        f"<b>Giờ check-in:</b> {current_time.strftime('%H:%M:%S')}\n"
        f"<b>Ca vào:</b> {shift_start_str}\n"
        f"{delegated_note}"
    )

    await message.reply_text(result, parse_mode=ParseMode.HTML)
    LogInfo(
        f"[CheckIn] {full_name} (@{emp_username}) checked in at {current_time.strftime('%H:%M')} - {status} (by @{acting_username})",
        LogType.SYSTEM_STATUS
    )


async def _execute_check_out_for_employee(client, message: Message, employee: Employee, db, acting_username: str) -> None:
    """Thực hiện logic check-out cho 1 Employee cụ thể (tách riêng từ handle_check_out)."""
    from app.models.finance import Attendance

    now_vn = datetime.datetime.now(VN_TZ)
    today = now_vn.date()
    current_time = now_vn.time()

    # Kiểm tra end_time (giờ tan ca)
    if not employee.end_time:
        await message.reply_text(
            f"⚠️ Nhân viên <b>{employee.last_name} {employee.first_name}</b> chưa được cài đặt giờ tan ca. Vui lòng liên hệ quản lý.",
            parse_mode=ParseMode.HTML
        )
        return

    end_time_obj = employee.end_time.time() if isinstance(employee.end_time, datetime.datetime) else employee.end_time
    end_time_str = end_time_obj.strftime("%H:%M")

    # Check-out chỉ được phép từ end_time đến 23:59 cùng ngày
    if current_time < end_time_obj:
        await message.reply_text(
            f"⚠️ <b>Chưa đến giờ tan ca!</b>\n\n"
            f"Giờ tan ca: <b>{end_time_str}</b>\n"
            f"Hiện tại: <b>{current_time.strftime('%H:%M')}</b>\n\n"
            f"<i>Bạn chỉ có thể check-out từ <b>{end_time_str}</b> đến <b>23:59</b>.</i>",
            parse_mode=ParseMode.HTML
        )
        return

    # Tìm record check-in hôm nay
    attendance = db.query(Attendance).filter(
        Attendance.employee_id == employee.id,
        Attendance.year == today.year,
        Attendance.month == today.month,
        Attendance.day == today.day,
        Attendance.check_in_time.isnot(None)
    ).first()

    if not attendance:
        full_name = f"{employee.last_name} {employee.first_name}"
        await message.reply_text(
            f"⚠️ Nhân viên <b>{full_name}</b> chưa check-in hôm nay. Vui lòng check-in trước.",
            parse_mode=ParseMode.HTML
        )
        return

    if attendance.check_out_time:
        checkout_str = attendance.check_out_time.strftime("%H:%M:%S")
        full_name = f"{employee.last_name} {employee.first_name}"
        await message.reply_text(
            f"⚠️ Nhân viên <b>{full_name}</b> đã check-out hôm nay rồi!\n\n"
            f"Thời gian check-out: <b>{checkout_str}</b>",
            parse_mode=ParseMode.HTML
        )
        return

    # Cập nhật check-out
    checkout_dt = now_vn.replace(tzinfo=None)
    attendance.check_out_time = checkout_dt

    # Tính overtime nếu check-out sau end_time
    overtime_minutes = 0
    if current_time > end_time_obj:
        now_dt = datetime.datetime.combine(today, current_time)
        end_dt_full = datetime.datetime.combine(today, end_time_obj)
        overtime_minutes = round((now_dt - end_dt_full).total_seconds() / 60)
        attendance.overtime = round(overtime_minutes / 60, 2)

    # Tính working_time (giờ làm việc)
    if attendance.check_in_time:
        diff = (checkout_dt - attendance.check_in_time).total_seconds() / 3600
        if overtime_minutes > 0:
            diff -= (overtime_minutes / 60)
        attendance.working_time = round(diff, 2)

    db.commit()

    # Reply
    full_name = f"{employee.last_name} {employee.first_name}"
    emp_username = employee.username or employee.id
    checkin_str = attendance.check_in_time.strftime("%H:%M:%S") if attendance.check_in_time else "N/A"
    day_name = DAY_NAMES_VN.get(today.weekday(), "")
    working_hours = attendance.working_time or 0

    delegated_note = ""
    if acting_username != employee.username:
        delegated_note = f"\n<b>Người thực hiện:</b> @{acting_username} (ủy quyền)"

    result = (
        f"<b>CHECK-OUT THÀNH CÔNG</b>\n\n"
        f"<b>Nhân viên:</b> {full_name} (@{emp_username})\n"
        f"<b>Mã NV:</b> {employee.id}\n"
        f"<b>Ngày:</b> {day_name}, {today.strftime('%d-%m-%Y')}\n"
        f"<b>Giờ check-in:</b> {checkin_str}\n"
        f"<b>Giờ check-out:</b> {current_time.strftime('%H:%M:%S')}\n"
        f"<b>Ca tan:</b> {end_time_str}\n"
        f"<b>Tổng giờ làm:</b> {working_hours:.1f}h\n"
        f"{delegated_note}"
    )
    # if overtime_minutes > 0:
    #     result += f"<b>Tăng ca:</b> {overtime_minutes} phút\n"

    await message.reply_text(result, parse_mode=ParseMode.HTML)
    LogInfo(
        f"[CheckOut] {full_name} (@{emp_username}) checked out at {current_time.strftime('%H:%M')} - {working_hours:.1f}h (by @{acting_username})",
        LogType.SYSTEM_STATUS
    )


async def handle_authority_callback(client, callback_query) -> None:
    """
    Xử lý callback khi user chọn nhân viên từ Inline Keyboard ủy quyền.
    Callback data format: auth_{prefix}|{employee_id}  hoặc  auth_{prefix}|cancel
    """
    data = callback_query.data
    parts = data.split("|", 1)
    if len(parts) != 2:
        await callback_query.answer("⚠️ Dữ liệu không hợp lệ.", show_alert=True)
        return

    prefix_part = parts[0]  # auth_ci, auth_co, auth_lv, auth_ov
    emp_id = parts[1]

    action = prefix_part.replace("auth_", "")  # ci, co, lv, ov

    msg_id = callback_query.message.id
    pending = _PENDING_AUTHORITY_ACTIONS.pop(msg_id, None)

    if emp_id == "cancel":
        await callback_query.message.edit_text(
            "❌ <b>Đã hủy thao tác.</b>",
            parse_mode=ParseMode.HTML
        )
        await callback_query.answer()
        return

    # Kiểm tra người nhấn nút đúng là người gõ lệnh
    if pending and callback_query.from_user.id != pending["from_user_id"]:
        await callback_query.answer("⚠️ Chỉ người gõ lệnh mới được chọn.", show_alert=True)
        return

    acting_username = callback_query.from_user.username

    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.id == emp_id).first()
        if not employee:
            await callback_query.message.edit_text(
                f"⚠️ Không tìm thấy nhân viên <b>{emp_id}</b>.",
                parse_mode=ParseMode.HTML
            )
            await callback_query.answer()
            return

        if employee.status == "inactive":
            await callback_query.message.edit_text(
                "⚠️ Tài khoản nhân viên đã bị vô hiệu hóa.",
                parse_mode=ParseMode.HTML
            )
            await callback_query.answer()
            return

        full_name = f"{employee.last_name} {employee.first_name}"
        await callback_query.message.edit_text(
            f"✅ Đã chọn: <b>{full_name} - {employee.id}</b>",
            parse_mode=ParseMode.HTML
        )

        # Dispatch theo action
        original_message = pending["original_message"] if pending else callback_query.message

        if action == "ci":
            await _execute_check_in_for_employee(client, original_message, employee, db, acting_username)
        elif action == "co":
            await _execute_check_out_for_employee(client, original_message, employee, db, acting_username)
        elif action == "lv":
            form_data = pending.get("form_data") if pending else None
            command_name = pending.get("command_name", "") if pending else ""
            if form_data:
                await _execute_request_leave_for_employee(client, original_message, employee, db, acting_username, form_data, command_name)
            else:
                await callback_query.message.reply_text(
                    "⚠️ Dữ liệu form nghỉ phép đã hết hạn. Vui lòng gửi lại lệnh.",
                    parse_mode=ParseMode.HTML
                )
        elif action == "ov":
            form_data = pending.get("form_data") if pending else None
            command_name = pending.get("command_name", "") if pending else ""
            if form_data:
                await _execute_request_overtime_for_employee(client, original_message, employee, db, acting_username, form_data, command_name)
            else:
                await callback_query.message.reply_text(
                    "⚠️ Dữ liệu form tăng ca đã hết hạn. Vui lòng gửi lại lệnh.",
                    parse_mode=ParseMode.HTML
                )
        elif action == "au":
            form_data = pending.get("form_data") if pending else None
            command_name = pending.get("command_name", "") if pending else ""
            if form_data:
                await _execute_request_attendance_update_for_employee(client, original_message, employee, db, acting_username, form_data, command_name)
            else:
                await callback_query.message.reply_text(
                    "⚠️ Dữ liệu form cập nhật công đã hết hạn. Vui lòng gửi lại lệnh.",
                    parse_mode=ParseMode.HTML
                )

        await callback_query.answer()

    except Exception as e:
        db.rollback()
        LogError(f"Error in handle_authority_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.message.reply_text(
            "❌ Có lỗi xảy ra. Vui lòng thử lại.",
            parse_mode=ParseMode.HTML
        )
        await callback_query.answer()
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════
# CHECK-IN (Chấm công)
# ══════════════════════════════════════════════════════════════

# Timezone Việt Nam (UTC+7)
VN_TZ = datetime.timezone(datetime.timedelta(hours=7))

DAY_NAMES_VN = {
    0: "Thứ 2", 1: "Thứ 3", 2: "Thứ 4",
    3: "Thứ 5", 4: "Thứ 6", 5: "Thứ 7", 6: "Chủ Nhật"
}


async def handle_check_in(client, message: Message, command_name: str) -> None:
    """
    Xử lý chấm công (check-in) cho nhân viên.
    Hỗ trợ ủy quyền: nếu user được nhiều nhân viên ủy quyền, hiện Inline Keyboard chọn.
    """
    db = SessionLocal()
    try:
        employee = await resolve_employee_for_command(
            client, message, db, callback_prefix="ci", command_name=command_name
        )
        if not employee:
            return  # Đã hiển thị lỗi hoặc Inline Keyboard chờ chọn

        acting_username = message.from_user.username
        await _execute_check_in_for_employee(client, message, employee, db, acting_username)

    except Exception as e:
        db.rollback()
        LogError(f"Error in handle_check_in: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(
            "❌ Có lỗi xảy ra khi chấm công. Vui lòng thử lại.",
            parse_mode=ParseMode.HTML
        )
    finally:
        db.close()



# ══════════════════════════════════════════════════════════════
# CHECK-OUT (Tan ca)
# ══════════════════════════════════════════════════════════════

async def handle_check_out(client, message: Message, command_name: str) -> None:
    """
    Xử lý check-out (tan ca) cho nhân viên.
    Hỗ trợ ủy quyền: nếu user được nhiều nhân viên ủy quyền, hiện Inline Keyboard chọn.
    """
    db = SessionLocal()
    try:
        employee = await resolve_employee_for_command(
            client, message, db, callback_prefix="co", command_name=command_name
        )
        if not employee:
            return  # Đã hiển thị lỗi hoặc Inline Keyboard chờ chọn

        acting_username = message.from_user.username
        await _execute_check_out_for_employee(client, message, employee, db, acting_username)

    except Exception as e:
        db.rollback()
        LogError(f"Error in handle_check_out: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(
            "❌ Có lỗi xảy ra khi check-out. Vui lòng thử lại.",
            parse_mode=ParseMode.HTML
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════
# REQUEST LEAVE (Xin nghỉ phép)
# ══════════════════════════════════════════════════════════════

LEAVE_TYPE_OPTIONS = [
    "Nghỉ phép năm",
    "Nghỉ không phép",
    "Nghỉ kết hôn",
    "Nghỉ thai sản",
    "Khác",
]

LEAVE_FORM_TEMPLATE = """<b>MẪU ĐƠN XIN NGHỈ PHÉP</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>{cmd}
Thời gian: dd/mm/yyyy - dd/mm/yyyy
Loại nghỉ: 
Người duyệt: @username
Người hỗ trợ: @username
Lý do: </pre>

<i>Loại nghỉ gồm: {leave_types}
Số ngày nghỉ phép năm còn lại: <b>{leave_balance}</b></i>"""


async def _execute_request_leave_for_employee(
    client, message: Message, employee: Employee, db, acting_username: str,
    form_data: dict, command_name: str,
) -> None:
    """Thực hiện logic request_leave cho 1 Employee cụ thể."""
    from app.models.finance import Attendance

    dates = form_data["dates"]
    leave_type = form_data["leave_type"]
    approver = form_data["approver"]
    supporter = form_data["supporter"]
    reason = form_data["reason"]
    start_dt = form_data["start_dt"]
    end_dt = form_data["end_dt"]
    num_days = form_data["num_days"]
    display_dates = form_data["display_dates"]

    # Kiểm tra nghỉ phép năm
    from bot.utils.utils import get_best_match
    is_annual_leave = (
        leave_type == "Nghỉ phép năm"
        or get_best_match(leave_type, ["Nghỉ phép năm"], threshold=0.7) is not None
    )

    full_name = f"{employee.last_name} {employee.first_name}"
    emp_username = employee.username or employee.id

    if is_annual_leave:
        balance = employee.leave_balance or 0
        if balance == 0:
            await message.reply_text(
                f"⚠️ Nhân viên <b>{full_name}</b> đã <b>hết ngày nghỉ phép năm</b> để sử dụng.",
                parse_mode=ParseMode.HTML,
            )
            return
        if balance < num_days:
            await message.reply_text(
                f"⚠️ Nhân viên <b>{full_name}</b> chỉ còn <b>{balance}</b> ngày phép năm, "
                f"không đủ để đăng ký <b>{num_days}</b> ngày.",
                parse_mode=ParseMode.HTML,
            )
            return

        # Kiểm tra trùng ngày đã đăng ký hoặc đã chấm công
        for i in range(num_days):
            curr_date = start_dt + datetime.timedelta(days=i)
            existing_att = db.query(Attendance).filter(
                Attendance.employee_id == employee.id,
                Attendance.year == curr_date.year,
                Attendance.month == curr_date.month,
                Attendance.day == curr_date.day,
            ).first()

            if existing_att:
                if existing_att.error and get_best_match(
                    existing_att.error, ["Nghỉ phép năm"], threshold=0.8
                ):
                    await message.reply_text(
                        f"⚠️ Nhân viên <b>{full_name}</b> đã đăng ký phép năm cho ngày "
                        f"<b>{curr_date.strftime('%d/%m/%Y')}</b> rồi.",
                        parse_mode=ParseMode.HTML,
                    )
                    return

                if existing_att.check_in_time or existing_att.check_out_time:
                    await message.reply_text(
                        f"⚠️ Nhân viên <b>{full_name}</b>, Ngày <b>{curr_date.strftime('%d/%m/%Y')}</b> "
                        "đã có dữ liệu chấm công. Không thể xin nghỉ phép năm.",
                        parse_mode=ParseMode.HTML,
                    )
                    return

    # Build response
    delegated_note = ""
    if acting_username != employee.username:
        delegated_note = f"\nNgười thực hiện: <b>@{acting_username}</b> (ủy quyền)"

    response = (
        f"<b>ĐƠN XIN NGHỈ PHÉP</b>\n\n"
        f"Người nghỉ: <b>{full_name}</b> (<b>@{emp_username}</b>)\n"
        f"Thời gian: <code>{display_dates}</code>\n"
        f"Loại nghỉ: <b>{leave_type}</b>\n"
        f"Người hỗ trợ: <b>{supporter}</b>\n"
        f"Lý do: <i>{reason}</i>\n"
        f"Người duyệt: <b>{approver}</b>"
        f"{delegated_note}"
    )

    # Tìm HR group
    hr_chat_id = None
    try:
        from app.models.telegram import TelegramProjectMember
        from bot.utils.enums import CustomTitle
        current_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == str(message.chat.id)
        ).first()

        main_hr_group = None
        if current_member:
            main_hr_group = db.query(TelegramProjectMember).filter(
                TelegramProjectMember.project_id == current_member.project_id,
                TelegramProjectMember.role == "main",
                TelegramProjectMember.custom_title.in_([CustomTitle.MAIN_HR.value, CustomTitle.SUPER_MAIN.value])
            ).first()

        hr_chat_id = main_hr_group.chat_id if main_hr_group else None
    except Exception as e:
        LogError(f"Error finding HR group: {e}", LogType.SYSTEM_STATUS)

    LogInfo(
        f"[RequestLeave] {full_name} (@{emp_username}) submitted leave: "
        f"{display_dates} | {leave_type} | {reason} (by @{acting_username})",
        LogType.SYSTEM_STATUS,
    )

    if hr_chat_id:
        from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        markup = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("Chấp thuận", callback_data=f"lv_req|ok|{message.chat.id}|{message.id}"),
                InlineKeyboardButton("Từ chối", callback_data=f"lv_req|no|{message.chat.id}|{message.id}")
            ]
        ])

        try:
            await client.send_message(
                chat_id=int(hr_chat_id),
                text=response,
                parse_mode=ParseMode.HTML,
                reply_markup=markup
            )
            await message.reply_text("✅ <b>Đơn xin nghỉ phép đã được gửi đến bộ phận Nhân sự để xét duyệt.</b>", parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error forwarding leave request to HR group: {e}", LogType.SYSTEM_STATUS)
            response += (
                f"\n\n<b>Yêu cầu người duyệt ({approver}) hoặc người hỗ trợ ({supporter}) "
                f"phản hồi đơn này bằng cách reply:</b>\n"
                f"- <code>/confirmed</code> để chấp thuận.\n"
                f"- <code>/denied</code> để từ chối."
            )
            await message.reply_text(response, parse_mode=ParseMode.HTML)
    else:
        # Fallback to local reply
        response += (
            f"\n\n<b>Yêu cầu người duyệt ({approver}) hoặc người hỗ trợ ({supporter}) "
            f"phản hồi đơn này bằng cách reply:</b>\n"
            f"- <code>/confirmed</code> để chấp thuận.\n"
            f"- <code>/denied</code> để từ chối."
        )
        await message.reply_text(response, parse_mode=ParseMode.HTML)


async def handle_request_leave(client, message: Message, command_name: str) -> None:
    """
    Xử lý yêu cầu xin nghỉ phép - dùng chung cho mọi dự án.
    Hỗ trợ ủy quyền: nếu user được nhiều nhân viên ủy quyền, hiện Inline Keyboard chọn.
    """
    from app.models.finance import Attendance

    lines = message.text.strip().split("\n")

    # Nếu chỉ gõ lệnh không tham số -> hiển thị form mẫu
    if len(lines) < 3:
        leave_balance = "N/A"
        username = message.from_user.username
        if username:
            db = SessionLocal()
            try:
                employee = db.query(Employee).filter(
                    Employee.username == username
                ).first()
                if employee and employee.leave_balance is not None:
                    leave_balance = str(employee.leave_balance)
            except Exception:
                pass
            finally:
                db.close()

        leave_types_str = ", ".join(LEAVE_TYPE_OPTIONS)
        form = LEAVE_FORM_TEMPLATE.format(
            cmd=command_name,
            leave_types=leave_types_str,
            leave_balance=leave_balance,
        )
        await message.reply_text(form, parse_mode=ParseMode.HTML)
        return

    # Parse form data
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    # Lấy các trường
    dates = data.get("Thời gian", "").strip()
    leave_type = data.get("Loại nghỉ", "").strip()
    approver = data.get("Người duyệt", "").strip()
    supporter = data.get("Người hỗ trợ", "").strip()
    reason = data.get("Lý do", "").strip()

    # Validate required fields
    if not all([dates, leave_type, approver, supporter, reason]):
        await message.reply_text(
            "⚠️ <b>Thiếu thông tin hoặc sai định dạng!</b>\n\n"
            "Vui lòng nhập đủ các trường theo mẫu sau:\n"
            "<code>Thời gian: dd/mm/yyyy - dd/mm/yyyy</code>\n"
            "<code>Loại nghỉ: ...</code>\n"
            "<code>Người duyệt: @username</code>\n"
            "<code>Người hỗ trợ: @username</code>\n"
            "<code>Lý do: ...</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    # Parse dates
    start_dt = None
    end_dt = None
    try:
        if " - " in dates:
            start_str, end_str = dates.split(" - ", 1)
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
                try:
                    start_dt = datetime.datetime.strptime(start_str.strip(), fmt)
                    end_dt = datetime.datetime.strptime(end_str.strip(), fmt)
                    break
                except ValueError:
                    continue
        else:
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
                try:
                    start_dt = datetime.datetime.strptime(dates.strip(), fmt)
                    end_dt = start_dt
                    break
                except ValueError:
                    continue
    except Exception:
        pass

    if not start_dt or not end_dt:
        await message.reply_text(
            "⚠️ <b>Sai định dạng ngày!</b>\n"
            "Vui lòng nhập theo format: <code>dd/mm/yyyy - dd/mm/yyyy</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    if end_dt < start_dt:
        await message.reply_text(
            "⚠️ <b>Ngày kết thúc phải sau ngày bắt đầu!</b>",
            parse_mode=ParseMode.HTML,
        )
        return

    num_days = (end_dt - start_dt).days + 1
    display_dates = f"{start_dt.strftime('%d/%m/%Y')} - {end_dt.strftime('%d/%m/%Y')}"

    # Đóng gói form data
    form_data = {
        "dates": dates,
        "leave_type": leave_type,
        "approver": approver,
        "supporter": supporter,
        "reason": reason,
        "start_dt": start_dt,
        "end_dt": end_dt,
        "num_days": num_days,
        "display_dates": display_dates,
    }

    # Resolve employee (hỗ trợ ủy quyền)
    db = SessionLocal()
    try:
        employee = await resolve_employee_for_command(
            client, message, db, callback_prefix="lv", command_name=command_name,
            form_data=form_data,
        )
        if not employee:
            return  # Đã hiển thị lỗi hoặc Inline Keyboard chờ chọn

        acting_username = message.from_user.username
        await _execute_request_leave_for_employee(
            client, message, employee, db, acting_username, form_data, command_name
        )

    except Exception as e:
        LogError(f"Error in handle_request_leave: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(
            "❌ Có lỗi xảy ra khi kiểm tra thông tin. Vui lòng thử lại.",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


async def handle_leave_request_callback(client, callback_query) -> None:
    data = callback_query.data  # format: lv_req|ok|<orig_chat_id>|<orig_msg_id>
    parts = data.split("|")
    if len(parts) != 4:
        await callback_query.answer("⚠️ Dữ liệu callback không hợp lệ.", show_alert=True)
        return

    action = parts[1]  # "ok" or "no"
    orig_chat_id = parts[2]
    orig_msg_id = parts[3]

    message = callback_query.message
    content = message.text
    sender_username = callback_query.from_user.username

    import re
    approver_match = re.search(r"Người duyệt:\s*(?:<b>)?@([^\s<]+)(?:</b>)?", content)
    designated_approver = approver_match.group(1).strip() if approver_match else ""

    from bot.utils.enums import UserType, has_flag

    is_authorized = False
    if sender_username == designated_approver:
        is_authorized = True
    else:
        # Check if the clicker is SUPER_MAIN or ADMIN/OWNER
        from app.models.employee import Employee
        db = SessionLocal()
        try:
            emp = db.query(Employee).filter(Employee.username == sender_username).first()
            if emp and has_flag(emp.user_mask, UserType.OWNER | UserType.ADMIN):
                is_authorized = True
            else:
                from app.models.telegram import TelegramProjectMember
                from bot.utils.enums import CustomTitle
                member = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.chat_id == str(message.chat.id),
                    TelegramProjectMember.user_name == sender_username
                ).first()
                if member and member.custom_title == CustomTitle.SUPER_MAIN:
                    is_authorized = True
        except Exception as e:
            LogError(f"Error checking authorization: {e}", LogType.SYSTEM_STATUS)
        finally:
            db.close()

    if not is_authorized:
        await callback_query.answer(f"⚠️ Chỉ người duyệt (@{designated_approver}) hoặc Quản lý mới có quyền thực hiện.", show_alert=True)
        return

    requester_match = re.search(r"Người nghỉ:\s*(?:<b>)?(.*?)(?:</b>)?(?:\n|$)", content)
    requester_name = requester_match.group(1).strip() if requester_match else "Nhân viên"

    if action == "ok":
        status_text = "<b>ĐÃ CHẤP THUẬN</b>"
        color_msg = "đã được chấp thuận"
    else:
        status_text = "<b>BỊ TỪ CHỐI</b>"
        color_msg = "đã bị từ chối"

    response = (
        f"{status_text}\n\n"
        f"Yêu cầu của <b>{requester_name}</b> {color_msg} bởi <b>@{sender_username}</b>."
    )

    # Automated Leave Balance Deduction
    if action == "ok" and "ĐƠN XIN NGHỈ PHÉP" in content:
        leave_type_match = re.search(r"Loại nghỉ:\s*(?:<b>)?(.*?)(?:</b>)?(?:\n|$)", content)
        dates_match = re.search(r"Thời gian:\s*(?:<code>)?(.*?)(?:</code>)?(?:\n|$)", content)
        requester_user_match = re.search(r"Người nghỉ:.*?\(\s*(?:<b>)?@([^\s<]+)(?:</b>)?\s*\)", content)
        
        if leave_type_match and dates_match and requester_user_match:
            l_type = leave_type_match.group(1).strip()
            l_dates = dates_match.group(1).strip()
            l_username = requester_user_match.group(1).strip()
            
            from bot.utils.utils import get_best_match
            if l_type == "Nghỉ phép năm" or get_best_match(l_type, ["Nghỉ phép năm"], threshold=0.7):
                db = SessionLocal()
                try:
                    num_days = 0
                    if " - " in l_dates:
                        start_str, end_str = l_dates.split(" - ")
                        start_dt = datetime.datetime.strptime(start_str.strip(), "%d/%m/%Y")
                        end_dt = datetime.datetime.strptime(end_str.strip(), "%d/%m/%Y")
                        num_days = (end_dt - start_dt).days + 1
                    else:
                        num_days = 1
                        
                    if num_days > 0:
                        from app.models.employee import Employee
                        from app.models.finance import Attendance
                        employee = db.query(Employee).filter(Employee.username == l_username).first()
                        if employee:
                            old_balance = employee.leave_balance or 0
                            new_balance = old_balance - num_days
                            if new_balance < 0:
                                response += (
                                    f"\n\n⚠️ <b>Không thể trừ phép năm!</b>"
                                    f"\nSố ngày phép còn lại: <code>{old_balance}</code>"
                                    f"\nSố ngày xin nghỉ: <code>{num_days}</code>"
                                    f"\n<i>Nhân viên không đủ phép năm. Vui lòng chuyển sang loại nghỉ khác (nghỉ không lương, ...).</i>"
                                )
                            else:
                                employee.leave_balance = new_balance
                                db.commit()
                                response += f"\n\n💡 <b>Số ngày nghỉ phép năm còn lại:</b> <code>{new_balance}</code>"
                            
                                try:
                                    start_dt_parsed = datetime.datetime.strptime(l_dates.split(" - ")[0].strip() if " - " in l_dates else l_dates.strip(), "%d/%m/%Y")
                                    days_vn = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
                                    
                                    for i in range(num_days):
                                        current_date = start_dt_parsed + datetime.timedelta(days=i)
                                        d, m, y = current_date.day, current_date.month, current_date.year
                                        d_str = days_vn[current_date.weekday()]
                                        
                                        existing_att = db.query(Attendance).filter(
                                            Attendance.employee_id == employee.id,
                                            Attendance.year == y,
                                            Attendance.month == m,
                                            Attendance.day == d,
                                        ).first()
                                        if existing_att:
                                            existing_att.error = l_type
                                            db.add(existing_att)
                                        else:
                                            new_att = Attendance(
                                                employee_id=employee.id,
                                                year=y,
                                                month=m,
                                                day=d,
                                                date_str=d_str,
                                                error=l_type
                                            )
                                            db.add(new_att)
                                    db.commit()
                                except Exception as att_err:
                                    db.rollback()
                                    LogError(f"Error recording attendance for leave: {att_err}", LogType.SYSTEM_STATUS)
                except Exception as e:
                    LogError(f"Error updating leave balance: {e}", LogType.SYSTEM_STATUS)
                finally:
                    db.close()

    try:
        await callback_query.message.edit_text(
            f"{content}\n\n====================\n{response}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error editing original message: {e}", LogType.SYSTEM_STATUS)

    try:
        await client.send_message(
            chat_id=int(orig_chat_id),
            reply_to_message_id=int(orig_msg_id),
            text=response,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error notifying employee: {e}", LogType.SYSTEM_STATUS)

    await callback_query.answer()

async def handle_overtime_request_callback(client, callback_query) -> None:
    data = callback_query.data  # format: ov_req|ok|<orig_chat_id>|<orig_msg_id>
    parts = data.split("|")
    if len(parts) != 4:
        await callback_query.answer("⚠️ Dữ liệu callback không hợp lệ.", show_alert=True)
        return

    action = parts[1]  # "ok" or "no"
    orig_chat_id = parts[2]
    orig_msg_id = parts[3]

    message = callback_query.message
    content = message.text
    sender_username = callback_query.from_user.username

    import re
    approver_match = re.search(r"Người duyệt:\s*(?:<b>)?@([^\s<]+)(?:</b>)?", content)
    designated_approver = approver_match.group(1).strip() if approver_match else ""

    from bot.utils.enums import UserType, has_flag

    is_authorized = False
    if sender_username == designated_approver:
        is_authorized = True
    else:
        # Check if the clicker is SUPER_MAIN or ADMIN/OWNER
        from app.models.employee import Employee
        db = SessionLocal()
        try:
            emp = db.query(Employee).filter(Employee.username == sender_username).first()
            if emp and has_flag(emp.user_mask, UserType.OWNER | UserType.ADMIN):
                is_authorized = True
            else:
                from app.models.telegram import TelegramProjectMember
                from bot.utils.enums import CustomTitle
                member = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.chat_id == str(message.chat.id),
                    TelegramProjectMember.user_name == sender_username
                ).first()
                if member and member.custom_title == CustomTitle.SUPER_MAIN:
                    is_authorized = True
        except Exception as e:
            LogError(f"Error checking authorization: {e}", LogType.SYSTEM_STATUS)
        finally:
            db.close()

    if not is_authorized:
        await callback_query.answer(f"⚠️ Chỉ người duyệt (@{designated_approver}) hoặc Quản lý mới có quyền thực hiện.", show_alert=True)
        return

    requester_match = re.search(r"Người yêu cầu:\s*(?:<b>)?(.*?)(?:</b>)?(?:\n|$)", content)
    requester_name = requester_match.group(1).strip() if requester_match else "Nhân viên"

    if action == "ok":
        status_text = "<b>ĐÃ CHẤP THUẬN</b>"
        color_msg = "đã được chấp thuận"
    else:
        status_text = "<b>BỊ TỪ CHỐI</b>"
        color_msg = "đã bị từ chối"

    response = (
        f"{status_text}\n\n"
        f"Yêu cầu của <b>{requester_name}</b> {color_msg} bởi <b>@{sender_username}</b>."
    )

    # Automated Overtime Recording
    if action == "ok" and "YÊU CẦU ĐĂNG KÝ TĂNG CA" in content:
        d_match = re.search(r"Ngày:\s*(?:<code>)?(.*?)(?:</code>)?(?:\n|$)", content)
        t_match = re.search(r"Thời gian:\s*(?:<b>)?(.*?)(?:</b>)?(?:\n|$)", content)
        u_match = re.search(r"Người yêu cầu:.*?\(\s*(?:<b>)?@([^\s<]+)(?:</b>)?\s*\)", content)

        if d_match and t_match and u_match:
            ot_date_str = d_match.group(1).strip()
            ot_time_range = t_match.group(1).strip()
            ot_username = u_match.group(1).strip()

            db = SessionLocal()
            try:
                from app.models.employee import Employee
                from app.models.finance import Attendance
                from app.crud.attendance import get_attendance
                # Parse date
                ot_date = None
                if "/" in ot_date_str:
                    d_parts = ot_date_str.split("/")
                    try:
                        if len(d_parts[0]) == 4: # yyyy/mm/dd
                            ot_date = datetime.datetime.strptime(ot_date_str, "%Y/%m/%d")
                        else: # dd/mm/yyyy
                            ot_date = datetime.datetime.strptime(ot_date_str, "%d/%m/%Y")
                    except Exception as e:
                        LogError(f"Date parse error in OT recording: {e}", LogType.SYSTEM_STATUS)

                # Parse time range (TimeA - TimeB)
                if ot_date and "-" in ot_time_range:
                    time_parts = ot_time_range.split("-")
                    if len(time_parts) == 2:
                        s_t_str = time_parts[0].strip()
                        e_t_str = time_parts[1].strip()
                        
                        try:
                            start_t = datetime.datetime.strptime(s_t_str, "%H:%M").time()
                            end_t = datetime.datetime.strptime(e_t_str, "%H:%M").time()
                            
                            start_dt = datetime.datetime.combine(ot_date.date(), start_t)
                            end_dt = datetime.datetime.combine(ot_date.date(), end_t)
                            
                            if end_dt < start_dt:
                                end_dt += datetime.timedelta(days=1)

                            employee = db.query(Employee).filter(Employee.username == ot_username).first()
                            if employee:
                                s_id = employee.id
                                y, m, d = ot_date.year, ot_date.month, ot_date.day
                                days_vn = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
                                ds_vn = days_vn[ot_date.weekday()]

                                existing_att = get_attendance(db, s_id, y, m, d)
                                if existing_att:
                                    existing_att.start_overtime = start_dt
                                    existing_att.end_overtime = end_dt
                                    db.add(existing_att)
                                else:
                                    new_att = Attendance(
                                        employee_id=s_id,
                                        year=y,
                                        month=m,
                                        day=d,
                                        date_str=ds_vn,
                                        start_overtime=start_dt,
                                        end_overtime=end_dt
                                    )
                                    db.add(new_att)
                                
                                db.commit()
                                response += f"\n\n✅ <b>Đã cập nhật giờ tăng ca vào hệ thống chấm công.</b>"
                        except Exception as e:
                            LogError(f"Time parse error in OT recording: {e}", LogType.SYSTEM_STATUS)
            except Exception as e:
                LogError(f"Error recording overtime: {e}", LogType.SYSTEM_STATUS)
            finally:
                db.close()

    try:
        await callback_query.message.edit_text(
            f"{content}\n\n====================\n{response}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error editing original message: {e}", LogType.SYSTEM_STATUS)

    try:
        await client.send_message(
            chat_id=int(orig_chat_id),
            reply_to_message_id=int(orig_msg_id),
            text=response,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error notifying employee: {e}", LogType.SYSTEM_STATUS)

    await callback_query.answer()


# ══════════════════════════════════════════════════════════════
# REQUEST OVERTIME (Đăng ký tăng ca)
# ══════════════════════════════════════════════════════════════

OVERTIME_FORM_TEMPLATE = """<b>MẪU ĐĂNG KÝ TĂNG CA</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>{cmd}
Ngày: dd/mm/yyyy
Thời gian: hh:mm - hh:mm
Người duyệt: @username
Người hỗ trợ: @username
Lý do: </pre>

<i>Ví dụ thời gian: 18:00 - 21:00</i>"""


async def _execute_request_overtime_for_employee(
    client, message: Message, employee: Employee, db, acting_username: str,
    form_data: dict, command_name: str,
) -> None:
    """Thực hiện logic request_overtime cho 1 Employee cụ thể."""
    date_str = form_data["date_str"]
    time_str = form_data["time_str"]
    approver = form_data["approver"]
    supporter = form_data["supporter"]
    reason = form_data["reason"]
    display_date = form_data["display_date"]

    full_name = f"{employee.last_name} {employee.first_name}"
    emp_username = employee.username or employee.id

    delegated_note = ""
    if acting_username != employee.username:
        delegated_note = f"\nNgười thực hiện: <b>@{acting_username}</b> (ủy quyền)"

    response = (
        f"<b>YÊU CẦU ĐĂNG KÝ TĂNG CA</b>\n\n"
        f"Người yêu cầu: <b>{full_name}</b> (<b>@{emp_username}</b>)\n"
        f"Ngày: <code>{display_date}</code>\n"
        f"Thời gian: <b>{time_str}</b>\n"
        f"Người duyệt: <b>{approver}</b>\n"
        f"Người hỗ trợ: <b>{supporter}</b>\n"
        f"Lý do: <i>{reason}</i>"
        f"{delegated_note}"
    )

    # Tìm HR group
    hr_chat_id = None
    try:
        from app.models.telegram import TelegramProjectMember
        from bot.utils.enums import CustomTitle
        current_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == str(message.chat.id)
        ).first()

        main_hr_group = None
        if current_member:
            main_hr_group = db.query(TelegramProjectMember).filter(
                TelegramProjectMember.project_id == current_member.project_id,
                TelegramProjectMember.role == "main",
                TelegramProjectMember.custom_title.in_([CustomTitle.MAIN_HR.value, CustomTitle.SUPER_MAIN.value])
            ).first()

        hr_chat_id = main_hr_group.chat_id if main_hr_group else None
    except Exception as e:
        LogError(f"Error finding HR group for overtime: {e}", LogType.SYSTEM_STATUS)

    LogInfo(
        f"[RequestOvertime] {full_name} (@{emp_username}) submitted overtime: "
        f"{display_date} | {time_str} | {reason} (by @{acting_username})",
        LogType.SYSTEM_STATUS,
    )

    if hr_chat_id:
        from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        markup = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("Chấp thuận", callback_data=f"ov_req|ok|{message.chat.id}|{message.id}"),
                InlineKeyboardButton("Từ chối", callback_data=f"ov_req|no|{message.chat.id}|{message.id}")
            ]
        ])

        try:
            await client.send_message(
                chat_id=int(hr_chat_id),
                text=response,
                parse_mode=ParseMode.HTML,
                reply_markup=markup
            )
            await message.reply_text("✅ <b>Đơn xin tăng ca đã được gửi đến bộ phận Nhân sự để xét duyệt.</b>", parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error forwarding overtime request to HR group: {e}", LogType.SYSTEM_STATUS)
            response += (
                f"\n\n<b>Yêu cầu người duyệt ({approver}) hoặc người hỗ trợ ({supporter}) "
                f"phản hồi đơn này bằng cách reply:</b>\n"
                f"- <code>/confirmed</code> để chấp thuận.\n"
                f"- <code>/denied</code> để từ chối."
            )
            await message.reply_text(response, parse_mode=ParseMode.HTML)
    else:
        # Fallback to local reply
        response += (
            f"\n\n<b>Yêu cầu người duyệt ({approver}) hoặc người hỗ trợ ({supporter}) "
            f"phản hồi đơn này bằng cách reply:</b>\n"
            f"- <code>/confirmed</code> để chấp thuận.\n"
            f"- <code>/denied</code> để từ chối."
        )
        await message.reply_text(response, parse_mode=ParseMode.HTML)


async def handle_request_overtime(client, message: Message, command_name: str) -> None:
    """
    Xử lý yêu cầu đăng ký tăng ca - dùng chung cho mọi dự án.
    Hỗ trợ ủy quyền: nếu user được nhiều nhân viên ủy quyền, hiện Inline Keyboard chọn.
    """
    lines = message.text.strip().split("\n")

    # Nếu chỉ gõ lệnh không tham số -> hiển thị form mẫu
    if len(lines) < 3:
        form = OVERTIME_FORM_TEMPLATE.format(cmd=command_name)
        await message.reply_text(form, parse_mode=ParseMode.HTML)
        return

    # Parse form data
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    # Lấy các trường
    date_str = data.get("Ngày", "").strip()
    time_str = data.get("Thời gian", "").strip()
    approver = data.get("Người duyệt", "").strip()
    supporter = data.get("Người hỗ trợ", "").strip()
    reason = data.get("Lý do", "").strip()

    # Validate required fields
    if not all([date_str, time_str, approver, supporter, reason]):
        await message.reply_text(
            "⚠️ <b>Thiếu thông tin hoặc sai định dạng!</b>\n\n"
            "Vui lòng nhập đủ các trường theo mẫu sau:\n"
            "<code>Ngày: dd/mm/yyyy</code>\n"
            "<code>Thời gian: hh:mm - hh:mm</code>\n"
            "<code>Người duyệt: @username</code>\n"
            "<code>Người hỗ trợ: @username</code>\n"
            "<code>Lý do: ...</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    # Parse date
    parsed_date = None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            parsed_date = datetime.datetime.strptime(date_str, fmt)
            break
        except ValueError:
            continue

    if not parsed_date:
        await message.reply_text(
            "⚠️ <b>Sai định dạng ngày!</b>\n"
            "Vui lòng nhập theo format: <code>dd/mm/yyyy</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    display_date = parsed_date.strftime("%d/%m/%Y")

    # Validate time format (hh:mm - hh:mm)
    if " - " not in time_str:
        await message.reply_text(
            "⚠️ <b>Sai định dạng thời gian!</b>\n"
            "Vui lòng nhập theo format: <code>hh:mm - hh:mm</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    time_parts = time_str.split(" - ")
    if len(time_parts) != 2:
        await message.reply_text(
            "⚠️ <b>Sai định dạng thời gian!</b>\n"
            "Vui lòng nhập theo format: <code>hh:mm - hh:mm</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    try:
        datetime.datetime.strptime(time_parts[0].strip(), "%H:%M")
        datetime.datetime.strptime(time_parts[1].strip(), "%H:%M")
    except ValueError:
        await message.reply_text(
            "⚠️ <b>Sai định dạng thời gian!</b>\n"
            "Vui lòng nhập theo format: <code>hh:mm - hh:mm</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    # Đóng gói form data
    form_data = {
        "date_str": date_str,
        "time_str": time_str,
        "approver": approver,
        "supporter": supporter,
        "reason": reason,
        "display_date": display_date,
    }

    # Resolve employee (hỗ trợ ủy quyền)
    db = SessionLocal()
    try:
        employee = await resolve_employee_for_command(
            client, message, db, callback_prefix="ov", command_name=command_name,
            form_data=form_data,
        )
        if not employee:
            return  # Đã hiển thị lỗi hoặc Inline Keyboard chờ chọn

        acting_username = message.from_user.username
        await _execute_request_overtime_for_employee(
            client, message, employee, db, acting_username, form_data, command_name
        )

    except Exception as e:
        LogError(f"Error in handle_request_overtime: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(
            "❌ Có lỗi xảy ra khi kiểm tra thông tin. Vui lòng thử lại.",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════
# REQUEST ATTENDANCE UPDATE (Yêu cầu cập nhật chấm công)
# ══════════════════════════════════════════════════════════════

ATTENDANCE_UPDATE_FORM_TEMPLATE = """<b>MẪU YÊU CẦU CẬP NHẬT CHẤM CÔNG</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>{cmd}
Thời gian: dd/mm/yyyy
Ca làm: hh:mm - hh:mm
Người duyệt: @username
Người hỗ trợ: @username
Lý do: </pre>

<i>Ví dụ ca làm: 08:00 - 17:00
Mã NV: <b>{emp_id}</b></i>"""


async def _execute_request_attendance_update_for_employee(
    client, message: Message, employee: Employee, db, acting_username: str,
    form_data: dict, command_name: str,
) -> None:
    """Thực hiện logic request_attendance_update cho 1 Employee cụ thể."""
    date_str = form_data["date_str"]
    shift_str = form_data["shift_str"]
    approver = form_data["approver"]
    supporter = form_data["supporter"]
    reason = form_data["reason"]
    display_date = form_data["display_date"]

    full_name = f"{employee.last_name} {employee.first_name}"
    emp_username = employee.username or employee.id

    delegated_note = ""
    if acting_username != employee.username:
        delegated_note = f"\nNgười thực hiện: <b>@{acting_username}</b> (ủy quyền)"

    response = (
        f"<b>YÊU CẦU CẬP NHẬT CHẤM CÔNG</b>\n\n"
        f"Người yêu cầu: <b>{full_name}</b> (<b>@{emp_username}</b>)\n"
        f"Mã NV: <code>{employee.id}</code>\n"
        f"Thời gian: <code>{display_date}</code>\n"
        f"Ca làm: <b>{shift_str}</b>\n"
        f"Người duyệt: <b>{approver}</b>\n"
        f"Người hỗ trợ: <b>{supporter}</b>\n"
        f"Lý do: <i>{reason}</i>"
        f"{delegated_note}"
    )

    # Tìm HR group thông qua parent_id (liên kết trực tiếp Member → Main)
    hr_chat_id = None
    try:
        from app.models.telegram import TelegramProjectMember
        current_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == str(message.chat.id),
            TelegramProjectMember.role == "member"
        ).first()

        if current_member and current_member.parent_id:
            hr_chat_id = current_member.parent_id
        elif current_member:
            # Fallback: tìm theo project_id nếu không có parent_id
            from bot.utils.enums import CustomTitle
            main_hr_group = db.query(TelegramProjectMember).filter(
                TelegramProjectMember.project_id == current_member.project_id,
                TelegramProjectMember.role == "main",
                TelegramProjectMember.custom_title.in_([CustomTitle.MAIN_HR.value, CustomTitle.SUPER_MAIN.value])
            ).first()
            hr_chat_id = main_hr_group.chat_id if main_hr_group else None
    except Exception as e:
        LogError(f"Error finding HR group for attendance update: {e}", LogType.SYSTEM_STATUS)

    LogInfo(
        f"[RequestAttendanceUpdate] {full_name} (@{emp_username}) submitted attendance update: "
        f"{display_date} | {shift_str} | {reason} (by @{acting_username})",
        LogType.SYSTEM_STATUS,
    )

    if hr_chat_id:
        from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        markup = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("Xác nhận", callback_data=f"att_req|ok|{message.chat.id}|{message.id}"),
                InlineKeyboardButton("Từ chối", callback_data=f"att_req|no|{message.chat.id}|{message.id}")
            ]
        ])

        try:
            await client.send_message(
                chat_id=int(hr_chat_id),
                text=response,
                parse_mode=ParseMode.HTML,
                reply_markup=markup
            )
            await message.reply_text("✅ <b>Yêu cầu cập nhật chấm công đã được gửi đến bộ phận Nhân sự để xét duyệt.</b>", parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error forwarding attendance update request to HR group: {e}", LogType.SYSTEM_STATUS)
            # Fallback: gửi ngay tại nhóm member với inline buttons
            from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
            fallback_markup = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("Xác nhận", callback_data=f"att_req|ok|{message.chat.id}|{message.id}"),
                    InlineKeyboardButton("Từ chối", callback_data=f"att_req|no|{message.chat.id}|{message.id}")
                ]
            ])
            await message.reply_text(response, parse_mode=ParseMode.HTML, reply_markup=fallback_markup)
    else:
        # Fallback: không tìm được nhóm HR → gửi tại nhóm hiện tại với inline buttons
        from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        fallback_markup = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("Xác nhận", callback_data=f"att_req|ok|{message.chat.id}|{message.id}"),
                InlineKeyboardButton("Từ chối", callback_data=f"att_req|no|{message.chat.id}|{message.id}")
            ]
        ])
        await message.reply_text(response, parse_mode=ParseMode.HTML, reply_markup=fallback_markup)


async def handle_request_attendance_update(client, message: Message, command_name: str) -> None:
    """
    Xử lý yêu cầu cập nhật chấm công - dùng chung cho mọi dự án.
    Hỗ trợ ủy quyền: nếu user được nhiều nhân viên ủy quyền, hiện Inline Keyboard chọn.
    """
    lines = message.text.strip().split("\n")

    # Nếu chỉ gõ lệnh không tham số -> hiển thị form mẫu
    if len(lines) < 3:
        emp_id = "N/A"
        username = message.from_user.username
        if username:
            db = SessionLocal()
            try:
                employee = db.query(Employee).filter(
                    Employee.username == username
                ).first()
                if employee:
                    emp_id = employee.id
            except Exception:
                pass
            finally:
                db.close()

        form = ATTENDANCE_UPDATE_FORM_TEMPLATE.format(
            cmd=command_name,
            emp_id=emp_id,
        )
        await message.reply_text(form, parse_mode=ParseMode.HTML)
        return

    # Parse form data
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    # Lấy các trường
    date_str = data.get("Thời gian", "").strip()
    shift_str = data.get("Ca làm", "").strip()
    approver = data.get("Người duyệt", "").strip()
    supporter = data.get("Người hỗ trợ", "").strip()
    reason = data.get("Lý do", "").strip()

    # Validate required fields
    if not all([date_str, shift_str, approver, supporter, reason]):
        await message.reply_text(
            "⚠️ <b>Thiếu thông tin hoặc sai định dạng!</b>\n\n"
            "Vui lòng nhập đủ các trường theo mẫu sau:\n"
            "<code>Thời gian: dd/mm/yyyy</code>\n"
            "<code>Ca làm: hh:mm - hh:mm</code>\n"
            "<code>Người duyệt: @username</code>\n"
            "<code>Người hỗ trợ: @username</code>\n"
            "<code>Lý do: ...</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    # Parse date
    parsed_date = None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            parsed_date = datetime.datetime.strptime(date_str, fmt)
            break
        except ValueError:
            continue

    if not parsed_date:
        await message.reply_text(
            "⚠️ <b>Sai định dạng ngày!</b>\n"
            "Vui lòng nhập theo format: <code>dd/mm/yyyy</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    display_date = parsed_date.strftime("%d/%m/%Y")

    # Validate shift format (hh:mm - hh:mm)
    if " - " not in shift_str:
        await message.reply_text(
            "⚠️ <b>Sai định dạng ca làm!</b>\n"
            "Vui lòng nhập theo format: <code>hh:mm - hh:mm</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    shift_parts = shift_str.split(" - ")
    if len(shift_parts) != 2:
        await message.reply_text(
            "⚠️ <b>Sai định dạng ca làm!</b>\n"
            "Vui lòng nhập theo format: <code>hh:mm - hh:mm</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    try:
        datetime.datetime.strptime(shift_parts[0].strip(), "%H:%M")
        datetime.datetime.strptime(shift_parts[1].strip(), "%H:%M")
    except ValueError:
        await message.reply_text(
            "⚠️ <b>Sai định dạng ca làm!</b>\n"
            "Vui lòng nhập theo format: <code>hh:mm - hh:mm</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    # Đóng gói form data
    form_data = {
        "date_str": date_str,
        "shift_str": shift_str,
        "approver": approver,
        "supporter": supporter,
        "reason": reason,
        "display_date": display_date,
    }

    # Resolve employee (hỗ trợ ủy quyền)
    db = SessionLocal()
    try:
        employee = await resolve_employee_for_command(
            client, message, db, callback_prefix="au", command_name=command_name,
            form_data=form_data,
        )
        if not employee:
            return  # Đã hiển thị lỗi hoặc Inline Keyboard chờ chọn

        acting_username = message.from_user.username
        await _execute_request_attendance_update_for_employee(
            client, message, employee, db, acting_username, form_data, command_name
        )

    except Exception as e:
        LogError(f"Error in handle_request_attendance_update: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(
            "❌ Có lỗi xảy ra khi kiểm tra thông tin. Vui lòng thử lại.",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


async def handle_attendance_update_request_callback(client, callback_query) -> None:
    """Xử lý callback xác nhận / từ chối yêu cầu cập nhật chấm công."""
    data = callback_query.data  # format: att_req|ok|<orig_chat_id>|<orig_msg_id>
    parts = data.split("|")
    if len(parts) != 4:
        await callback_query.answer("⚠️ Dữ liệu callback không hợp lệ.", show_alert=True)
        return

    action = parts[1]  # "ok" or "no"
    orig_chat_id = parts[2]
    orig_msg_id = parts[3]

    message = callback_query.message
    content = message.text
    sender_username = callback_query.from_user.username

    import re
    approver_match = re.search(r"Người duyệt:\s*(?:<b>)?@([^\s<]+)(?:</b>)?", content)
    designated_approver = approver_match.group(1).strip() if approver_match else ""

    from bot.utils.enums import UserType, has_flag

    is_authorized = False
    if sender_username == designated_approver:
        is_authorized = True
    else:
        from app.models.employee import Employee
        db = SessionLocal()
        try:
            emp = db.query(Employee).filter(Employee.username == sender_username).first()
            if emp and has_flag(emp.user_mask, UserType.OWNER | UserType.ADMIN):
                is_authorized = True
            else:
                from app.models.telegram import TelegramProjectMember
                from bot.utils.enums import CustomTitle
                member = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.chat_id == str(message.chat.id),
                    TelegramProjectMember.user_name == sender_username
                ).first()
                if member and member.custom_title == CustomTitle.SUPER_MAIN:
                    is_authorized = True
        except Exception as e:
            LogError(f"Error checking authorization: {e}", LogType.SYSTEM_STATUS)
        finally:
            db.close()

    if not is_authorized:
        await callback_query.answer(f"⚠️ Chỉ người duyệt (@{designated_approver}) hoặc Quản lý mới có quyền thực hiện.", show_alert=True)
        return

    requester_match = re.search(r"Người yêu cầu:\s*(?:<b>)?(.*?)(?:</b>)?(?:\n|$)", content)
    requester_name = requester_match.group(1).strip() if requester_match else "Nhân viên"

    if action == "ok":
        status_text = "<b>ĐÃ XÁC NHẬN</b>"
        color_msg = "đã được xác nhận"
    else:
        status_text = "<b>BỊ TỪ CHỐI</b>"
        color_msg = "đã bị từ chối"

    response = (
        f"{status_text}\n\n"
        f"Yêu cầu cập nhật chấm công của <b>{requester_name}</b> {color_msg} bởi <b>@{sender_username}</b>."
    )

    # Nếu xác nhận → cập nhật DB Attendance
    if action == "ok" and "YÊU CẦU CẬP NHẬT CHẤM CÔNG" in content:
        d_match = re.search(r"Thời gian:\s*(?:<code>)?(.*?)(?:</code>)?(?:\n|$)", content)
        s_match = re.search(r"Ca làm:\s*(?:<b>)?(.*?)(?:</b>)?(?:\n|$)", content)
        u_match = re.search(r"Người yêu cầu:.*?\(\s*(?:<b>)?@([^\s<]+)(?:</b>)?\s*\)", content)

        if d_match and s_match and u_match:
            att_date_str = d_match.group(1).strip()
            att_shift = s_match.group(1).strip()
            att_username = u_match.group(1).strip()

            db = SessionLocal()
            try:
                # Parse date
                att_date = None
                if "/" in att_date_str:
                    try:
                        att_date = datetime.datetime.strptime(att_date_str, "%d/%m/%Y")
                    except Exception as e:
                        LogError(f"Date parse error in Attendance update recording: {e}", LogType.SYSTEM_STATUS)

                # Parse time range (hh:mm - hh:mm)
                if att_date and "-" in att_shift:
                    time_parts = att_shift.split("-")
                    if len(time_parts) == 2:
                        s_t_str = time_parts[0].strip()
                        e_t_str = time_parts[1].strip()

                        try:
                            start_t = datetime.datetime.strptime(s_t_str, "%H:%M").time()
                            end_t = datetime.datetime.strptime(e_t_str, "%H:%M").time()

                            start_dt = datetime.datetime.combine(att_date.date(), start_t)
                            end_dt = datetime.datetime.combine(att_date.date(), end_t)

                            if end_dt < start_dt:
                                end_dt += datetime.timedelta(days=1)

                            from app.models.employee import Employee
                            from app.models.finance import Attendance
                            from app.crud.attendance import get_attendance
                            emp = db.query(Employee).filter(Employee.username == att_username).first()
                            if emp:
                                s_id = emp.id
                                y, m, d = att_date.year, att_date.month, att_date.day
                                days_vn = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
                                ds_vn = days_vn[att_date.weekday()]

                                existing_att = get_attendance(db, s_id, y, m, d)
                                if existing_att:
                                    existing_att.check_in_time = start_dt
                                    existing_att.check_out_time = end_dt
                                    db.add(existing_att)
                                else:
                                    new_att = Attendance(
                                        employee_id=s_id,
                                        year=y,
                                        month=m,
                                        day=d,
                                        date_str=ds_vn,
                                        check_in_time=start_dt,
                                        check_out_time=end_dt
                                    )
                                    db.add(new_att)

                                db.commit()
                                response += f"\n\n✅ <b>Đã cập nhật dữ liệu chấm công.</b>"
                                LogInfo(
                                    f"[AttendanceUpdate] Updated attendance for @{att_username} on {att_date_str}: {att_shift} by @{sender_username}",
                                    LogType.SYSTEM_STATUS,
                                )
                        except Exception as e:
                            LogError(f"Time parse error in Attendance update recording: {e}", LogType.SYSTEM_STATUS)
            except Exception as e:
                LogError(f"Error recording attendance update: {e}", LogType.SYSTEM_STATUS)
            finally:
                db.close()

    try:
        await callback_query.message.edit_text(
            f"{content}\n\n====================\n{response}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error editing original message: {e}", LogType.SYSTEM_STATUS)

    try:
        await client.send_message(
            chat_id=int(orig_chat_id),
            reply_to_message_id=int(orig_msg_id),
            text=response,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error notifying employee: {e}", LogType.SYSTEM_STATUS)

    await callback_query.answer()


# ══════════════════════════════════════════════════════════════
# LIST CHECK-IN (Xem chấm công)
# ══════════════════════════════════════════════════════════════

def _build_attendance_image(
    employee_name: str,
    employee_id: str,
    month: int,
    year: int,
    records: list,
) -> BytesIO:
    """
    Tạo ảnh bảng chấm công từ danh sách Attendance records.
    Returns BytesIO chứa ảnh PNG.
    """
    from PIL import Image, ImageDraw, ImageFont
    from io import BytesIO

    # --- Config ---
    HEADER_BG = (30, 41, 59)        # slate-800
    HEADER_TEXT = (255, 255, 255)
    TITLE_BG = (15, 23, 42)         # slate-900
    TITLE_TEXT = (56, 189, 248)      # sky-400
    ROW_EVEN = (241, 245, 249)      # slate-100
    ROW_ODD = (255, 255, 255)
    BORDER = (203, 213, 225)        # slate-300
    TEXT_COLOR = (30, 41, 59)
    ERROR_COLOR = (220, 38, 38)     # red-600
    SUCCESS_COLOR = (22, 163, 74)   # green-600

    # Column widths
    col_widths = [50, 90, 90, 100, 100, 80, 120]
    headers = ["Ngày", "Thứ", "Check-in", "Check-out", "Tăng ca", "Giờ làm", "Ghi chú"]

    row_height = 36
    title_height = 70
    header_height = 40
    padding = 20

    total_width = sum(col_widths) + padding * 2
    num_rows = len(records)
    total_height = title_height + header_height + (num_rows * row_height) + padding + 30  # 30 for footer

    # --- Try to load a nice font, fallback to default ---
    try:
        font = ImageFont.truetype("arial.ttf", 14)
        font_bold = ImageFont.truetype("arialbd.ttf", 14)
        font_title = ImageFont.truetype("arialbd.ttf", 18)
        font_small = ImageFont.truetype("arial.ttf", 11)
    except OSError:
        font = ImageFont.load_default()
        font_bold = font
        font_title = font
        font_small = font

    img = Image.new("RGB", (total_width, total_height), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    # --- Title bar ---
    draw.rectangle([0, 0, total_width, title_height], fill=TITLE_BG)
    title_text = f"BẢNG CHẤM CÔNG THÁNG {month:02d}/{year}"
    draw.text((padding, 12), title_text, fill=TITLE_TEXT, font=font_title)
    sub_text = f"{employee_name}  |  Mã NV: {employee_id}"
    draw.text((padding, 40), sub_text, fill=(148, 163, 184), font=font_small)

    # --- Header row ---
    y = title_height
    draw.rectangle([0, y, total_width, y + header_height], fill=HEADER_BG)
    x = padding
    for i, header in enumerate(headers):
        draw.text((x + 4, y + 10), header, fill=HEADER_TEXT, font=font_bold)
        x += col_widths[i]

    # --- Data rows ---
    y = title_height + header_height
    for idx, rec in enumerate(records):
        bg = ROW_EVEN if idx % 2 == 0 else ROW_ODD
        draw.rectangle([0, y, total_width, y + row_height], fill=bg)
        # Border line
        draw.line([(0, y + row_height), (total_width, y + row_height)], fill=BORDER, width=1)

        x = padding
        day_str = f"{rec['day']:02d}/{rec['month']:02d}"
        check_in = rec.get("check_in", "—")
        check_out = rec.get("check_out", "—")
        overtime = rec.get("overtime", "—")
        work_h = rec.get("work_hours", "—")
        error = rec.get("error", "")
        date_label = rec.get("date_str", "")

        values = [day_str, date_label, check_in, check_out, overtime, work_h, error]
        for i, val in enumerate(values):
            color = TEXT_COLOR
            if i == 6 and val:  # error/ghi chú column
                color = ERROR_COLOR
            if i == 2 and val not in ("—", ""):  # check-in
                color = SUCCESS_COLOR
            draw.text((x + 4, y + 9), str(val), fill=color, font=font)
            x += col_widths[i]

        y += row_height

    # --- Footer ---
    draw.text((padding, y + 5), f"Tổng: {num_rows} ngày  |  Tạo lúc: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", fill=(100, 116, 139), font=font_small)

    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    buf.name = f"chamcong_{employee_id}_{month:02d}_{year}.png"
    return buf


async def handle_list_check_in(client, message: Message, command_name: str) -> None:
    """
    Xem bảng chấm công tháng hiện tại, xuất ảnh.

    Flow:
    1. Lấy username người gọi → tìm trong TelegramProjectMember → lấy mã NV
    2. Query Attendance theo employee_id, tháng hiện tại
    3. Render ảnh bảng chấm công → gửi vào nhóm
    """
    from app.models.finance import Attendance
    from app.models.telegram import TelegramProjectMember

    username = message.from_user.username
    if not username:
        await message.reply_text(
            "⚠️ Bạn chưa cài đặt username Telegram.",
            parse_mode=ParseMode.HTML,
        )
        return

    now = datetime.datetime.now()
    target_month = now.month
    target_year = now.year

    # Cho phép truyền tham số tháng/năm: /cmd 03/2026
    lines = message.text.strip().split()
    if len(lines) >= 2:
        param = lines[-1]
        try:
            if "/" in param:
                parts = param.split("/")
                target_month = int(parts[0])
                if len(parts) > 1:
                    target_year = int(parts[1])
        except (ValueError, IndexError):
            pass

    db = SessionLocal()
    try:
        # Tìm employee qua username
        employee = db.query(Employee).filter(Employee.username == username).first()
        if not employee:
            await message.reply_text(
                f"⚠️ Không tìm thấy nhân viên <b>@{username}</b> trong hệ thống.",
                parse_mode=ParseMode.HTML,
            )
            return

        employee_id = employee.id
        full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()

        # Query attendance records
        records = (
            db.query(Attendance)
            .filter(
                Attendance.employee_id == employee_id,
                Attendance.year == target_year,
                Attendance.month == target_month,
            )
            .order_by(Attendance.day)
            .all()
        )

        if not records:
            await message.reply_text(
                f"ℹ️ Không có dữ liệu chấm công tháng <b>{target_month:02d}/{target_year}</b> "
                f"cho nhân viên <b>{full_name}</b>.",
                parse_mode=ParseMode.HTML,
            )
            return

        # Build data for image
        data_rows = []
        for att in records:
            check_in = ""
            check_out = ""
            overtime_str = ""
            work_h_str = ""

            if att.check_in_time:
                check_in = att.check_in_time.strftime("%H:%M")
            if att.check_out_time:
                check_out = att.check_out_time.strftime("%H:%M")
            if att.start_overtime and att.end_overtime:
                overtime_str = f"{att.start_overtime.strftime('%H:%M')}-{att.end_overtime.strftime('%H:%M')}"
            elif att.overtime:
                overtime_str = f"{att.overtime}h"
            if att.working_time is not None:
                work_h_str = f"{att.working_time:.1f}h"
            elif att.check_in_time and att.check_out_time:
                diff = (att.check_out_time - att.check_in_time).total_seconds() / 3600
                work_h_str = f"{diff:.1f}h"

            data_rows.append({
                "day": att.day,
                "month": att.month,
                "date_str": att.date_str or "",
                "check_in": check_in or "—",
                "check_out": check_out or "—",
                "overtime": overtime_str or "—",
                "work_hours": work_h_str or "—",
                "error": att.error or "",
            })

        # Loading message
        loading_msg = await message.reply_text(
            "⏳ Đang tạo bảng chấm công, vui lòng chờ...",
            parse_mode=ParseMode.HTML,
        )

        # Generate image (HTML+CSS → Playwright screenshot)
        from bot.utils.attendance_generator import generate_attendance_image
        img_buf = await generate_attendance_image(
            employee_name=full_name,
            employee_id=employee_id,
            month=target_month,
            year=target_year,
            records=data_rows,
        )

        await message.reply_photo(
            photo=img_buf,
            caption=(
                f"📋 <b>Bảng chấm công tháng {target_month:02d}/{target_year}</b>\n"
                f"Nhân viên: <b>{full_name}</b> (Mã NV: <code>{employee_id}</code>)"
            ),
            parse_mode=ParseMode.HTML,
        )
        await loading_msg.delete()

    except Exception as e:
        LogError(f"Error in handle_list_check_in: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(
            "❌ Có lỗi xảy ra khi lấy dữ liệu chấm công. Vui lòng thử lại.",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════
# LIST REQUEST LEAVE (Xem danh sách nghỉ phép)
# ══════════════════════════════════════════════════════════════

async def handle_list_request_leave(client, message: Message, command_name: str) -> None:
    """
    Xem danh sách nghỉ phép trong tháng, xuất ảnh (HTML+CSS → Playwright screenshot).

    Flow:
    1. Lấy username → tìm Employee → lấy mã NV
    2. Query Attendance theo employee_id WHERE error IS NOT NULL (chứa loại nghỉ)
    3. Render ảnh danh sách nghỉ phép → gửi vào nhóm
    """
    from app.models.finance import Attendance

    username = message.from_user.username
    if not username:
        await message.reply_text(
            "⚠️ Bạn chưa cài đặt username Telegram.",
            parse_mode=ParseMode.HTML,
        )
        return

    now = datetime.datetime.now()
    target_month = now.month
    target_year = now.year

    # Cho phép truyền tham số tháng/năm: /cmd 03/2026
    lines = message.text.strip().split()
    if len(lines) >= 2:
        param = lines[-1]
        try:
            if "/" in param:
                parts = param.split("/")
                target_month = int(parts[0])
                if len(parts) > 1:
                    target_year = int(parts[1])
        except (ValueError, IndexError):
            pass

    db = SessionLocal()
    try:
        # Tìm employee qua username
        employee = db.query(Employee).filter(Employee.username == username).first()
        if not employee:
            await message.reply_text(
                f"⚠️ Không tìm thấy nhân viên <b>@{username}</b> trong hệ thống.",
                parse_mode=ParseMode.HTML,
            )
            return

        employee_id = employee.id
        full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()
        leave_balance = employee.leave_balance

        # Query attendance records có error (nghỉ phép)
        records = (
            db.query(Attendance)
            .filter(
                Attendance.employee_id == employee_id,
                Attendance.year == target_year,
                Attendance.month == target_month,
                Attendance.error.isnot(None),
                Attendance.error != "",
            )
            .order_by(Attendance.day)
            .all()
        )

        if not records:
            await message.reply_text(
                f"ℹ️ Không có dữ liệu nghỉ phép tháng <b>{target_month:02d}/{target_year}</b> "
                f"cho nhân viên <b>{full_name}</b>.",
                parse_mode=ParseMode.HTML,
            )
            return

        # Build data for image
        data_rows = []
        for att in records:
            data_rows.append({
                "day": att.day,
                "month": att.month,
                "year": att.year,
                "date_str": att.date_str or "",
                "leave_type": att.error or "",
                "status": "Đã duyệt",
            })

        # Loading message
        loading_msg = await message.reply_text(
            "⏳ Đang tạo danh sách nghỉ phép, vui lòng chờ...",
            parse_mode=ParseMode.HTML,
        )

        # Generate image (HTML+CSS → Playwright screenshot)
        from bot.utils.leave_generator import generate_leave_image
        img_buf = await generate_leave_image(
            employee_name=full_name,
            employee_id=employee_id,
            month=target_month,
            year=target_year,
            records=data_rows,
            leave_balance=leave_balance,
        )

        await message.reply_photo(
            photo=img_buf,
            caption=(
                f"📋 <b>Danh sách nghỉ phép tháng {target_month:02d}/{target_year}</b>\n"
                f"Nhân viên: <b>{full_name}</b> (Mã NV: <code>{employee_id}</code>)\n"
                f"Tổng: <b>{len(data_rows)}</b> ngày nghỉ | Phép năm còn: <b>{leave_balance or 0}</b> ngày"
            ),
            parse_mode=ParseMode.HTML,
        )
        await loading_msg.delete()

    except Exception as e:
        LogError(f"Error in handle_list_request_leave: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(
            "❌ Có lỗi xảy ra khi lấy dữ liệu nghỉ phép. Vui lòng thử lại.",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════
# CREATE TASK (Giao việc)
# ══════════════════════════════════════════════════════════════

TASK_FORM_TEMPLATE = """<b>MẪU GIAO VIỆC</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>{cmd}
Người nhận: [mã nhân viên]
Nội dung: 
Thời gian: hh:mm - hh:mm
Ngày bắt đầu: dd/mm/yyyy
Ngày kết thúc: dd/mm/yyyy
Chu kỳ: </pre>

<i>Chu kỳ: Hàng ngày, Hàng tuần, Hàng tháng, Một lần, ...</i>"""


async def handle_create_task(client, message: Message, command_name: str) -> None:
    """
    Xử lý tạo task giao việc cho nhân viên.

    Flow:
    1. Gõ lệnh không tham số → hiển thị form mẫu
    2. Gõ lệnh với form data → validate, lưu vào bảng tasks
    3. Thông báo kết quả
    """
    from app.models.task import Task

    lines = message.text.strip().split("\n")

    # Nếu chỉ gõ lệnh không tham số -> hiển thị form mẫu
    if len(lines) < 3:
        form = TASK_FORM_TEMPLATE.format(cmd=command_name)
        await message.reply_text(form, parse_mode=ParseMode.HTML)
        return

    # Parse form data
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    # Lấy các trường
    assignee_raw = data.get("Người nhận", "").strip()
    content = data.get("Nội dung", "").strip()
    time_range = data.get("Thời gian", "").strip()
    start_date = data.get("Ngày bắt đầu", "").strip()
    end_date = data.get("Ngày kết thúc", "").strip()
    cycle = data.get("Chu kỳ", "").strip()

    # Validate required fields
    if not all([assignee_raw, content, time_range, start_date, end_date]):
        await message.reply_text(
            "⚠️ <b>Thiếu thông tin hoặc sai định dạng!</b>\n\n"
            "Vui lòng nhập đủ các trường theo mẫu sau:\n"
            "<code>Người nhận: [mã nhân viên]</code>\n"
            "<code>Nội dung: ...</code>\n"
            "<code>Thời gian: hh:mm - hh:mm</code>\n"
            "<code>Ngày bắt đầu: dd/mm/yyyy</code>\n"
            "<code>Ngày kết thúc: dd/mm/yyyy</code>\n"
            "<code>Chu kỳ: ...</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    # Parse assignee (mã nhân viên)
    emp_code = assignee_raw.strip()
    if not emp_code:
        await message.reply_text(
            "⚠️ Vui lòng nhập mã nhân viên (VD: TN001).",
            parse_mode=ParseMode.HTML,
        )
        return

    # Validate time format
    if " - " not in time_range:
        await message.reply_text(
            "⚠️ <b>Sai định dạng thời gian!</b>\n"
            "Vui lòng nhập theo format: <code>hh:mm - hh:mm</code>",
            parse_mode=ParseMode.HTML,
        )
        return
    try:
        t1, t2 = time_range.split(" - ")
        datetime.datetime.strptime(t1.strip(), "%H:%M")
        datetime.datetime.strptime(t2.strip(), "%H:%M")
    except ValueError:
        await message.reply_text(
            "⚠️ <b>Sai định dạng thời gian!</b>\n"
            "Vui lòng nhập theo format: <code>hh:mm - hh:mm</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    # Append to content
    full_content = f"{content}\nThời gian: {time_range}"

    # Validate dates
    for label, date_str in [("Ngày bắt đầu", start_date), ("Ngày kết thúc", end_date)]:
        parsed = None
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                parsed = datetime.datetime.strptime(date_str, fmt)
                break
            except ValueError:
                continue
        if not parsed:
            await message.reply_text(
                f"⚠️ <b>Sai định dạng {label}!</b>\n"
                f"Vui lòng nhập theo format: <code>dd/mm/yyyy</code>",
                parse_mode=ParseMode.HTML,
            )
            return

    # Validate assignee exists in Employee
    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.id == emp_code).first()
        if not employee:
            await message.reply_text(
                f"⚠️ Không tìm thấy nhân viên mã <b>{emp_code}</b> trong hệ thống.",
                parse_mode=ParseMode.HTML,
            )
            return

        assignee_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()
        employee_id = employee.id
        assignee_username = employee.username or ""

        # Assigner info
        user = message.from_user
        assigner_name = f"{user.first_name} {user.last_name or ''}".strip()

        # Get project_id
        project_id = None
        from app.models.business import Projects
        project = db.query(Projects).filter(
            Projects.project_name == "Tiến Nga"
        ).first()
        if project:
            project_id = project.id

        # Create task
        new_task = Task(
            employee_id=employee_id,
            project_id=project_id,
            group_chat_id=str(message.chat.id),
            assigner=f"{assigner_name} (@{user.username})" if user.username else assigner_name,
            assignee=f"{assignee_name} ({employee_id})",
            content=full_content,
            start_date=start_date,
            end_date=end_date,
            cycle=cycle or "Một lần",
            status="PENDING",
        )
        db.add(new_task)
        db.commit()

        response = (
            f"<b>GIAO VIỆC THÀNH CÔNG</b>\n\n"
            f"Người giao: <b>{assigner_name}</b>\n"
            f"Người nhận: <b>{assignee_name}</b> (Mã NV: <code>{employee_id}</code>)\n"
            f"Nội dung: <i>{content}</i>\n"
            f"Thời gian: <b>{time_range}</b>\n"
            f"Bắt đầu: <code>{start_date}</code>\n"
            f"Kết thúc: <code>{end_date}</code>\n"
            f"Chu kỳ: <b>{cycle or 'Một lần'}</b>\n"
            f"Trạng thái: <b>PENDING</b>\n\n"
            f"<i>Mã task: <code>{new_task.id}</code></i>"
        )

        summary_msg = await message.reply_text(response, parse_mode=ParseMode.HTML)

        # Update message_id
        new_task.message_id = summary_msg.id
        db.commit()

        LogInfo(
            f"[CreateTask] {assigner_name} assigned task to {assignee_name}: {full_content}",
            LogType.SYSTEM_STATUS,
        )

        # Notify HR member group
        from app.models.telegram import TelegramProjectMember
        from bot.utils.enums import CustomTitle
        hr_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project_id,
            TelegramProjectMember.role == "member",
            TelegramProjectMember.custom_title == CustomTitle.MEMBER_HR.value
        ).first()
        
        if hr_member and hr_member.chat_id:
            try:
                await client.send_message(
                    chat_id=int(hr_member.chat_id),
                    text=f"🔔 <b>CÓ CÔNG VIỆC MỚI ĐƯỢC GIAO</b>\n\n{response}",
                    parse_mode=ParseMode.HTML
                )
            except Exception as e:
                LogError(f"Error forwarding task to HR member group: {e}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in handle_create_task: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(
            "❌ Có lỗi xảy ra khi tạo task. Vui lòng thử lại.",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════
# LIST TASKS (Xem danh sách task)
# ══════════════════════════════════════════════════════════════

TASK_STATUS_MAP = {
    "PENDING": "pending - Chờ xử lý",
    "IN_PROGRESS": "in_progress - Đang thực hiện",
    "COMPLETED": "completed - Hoàn thành",
    "CANCELLED": "cancelled - Đã hủy",
}


def _build_task_list_buttons(tasks: list) -> InlineKeyboardMarkup:
    """Tạo inline keyboard cho danh sách task."""
    from pyrogram.types import InlineKeyboardButton, InlineKeyboardMarkup

    buttons = []
    for t in tasks:
        status_icon = TASK_STATUS_MAP.get(t.status, t.status).split(" ")[0]
        label = f"{status_icon} - {t.content[:35]}{'...' if len(t.content) > 35 else ''}"
        buttons.append([InlineKeyboardButton(label, callback_data=f"tsk_sel_{t.id}")])

    buttons.append([InlineKeyboardButton("Đóng", callback_data="tsk_cancel")])
    return InlineKeyboardMarkup(buttons)


def _build_task_action_buttons(task_id: str) -> InlineKeyboardMarkup:
    """Tạo inline keyboard cho actions của 1 task."""
    from pyrogram.types import InlineKeyboardButton, InlineKeyboardMarkup

    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Xem chi tiết", callback_data=f"tsk_detail_{task_id}")],
        [InlineKeyboardButton("Cập nhật trạng thái", callback_data=f"tsk_status_{task_id}")],
        [
            InlineKeyboardButton("Quay lại", callback_data="tsk_back_list"),
            InlineKeyboardButton("Đóng", callback_data="tsk_cancel"),
        ],
    ])


def _build_status_buttons(task_id: str) -> "InlineKeyboardMarkup":
    """Tạo inline keyboard cho update trạng thái."""
    from pyrogram.types import InlineKeyboardButton, InlineKeyboardMarkup

    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Chờ xử lý", callback_data=f"tsk_set_PENDING_{task_id}")],
        [InlineKeyboardButton("Đang thực hiện", callback_data=f"tsk_set_IN_PROGRESS_{task_id}")],
        [InlineKeyboardButton("Hoàn thành", callback_data=f"tsk_set_COMPLETED_{task_id}")],
        [InlineKeyboardButton("Đã hủy", callback_data=f"tsk_set_CANCELLED_{task_id}")],
        [
            InlineKeyboardButton("Quay lại", callback_data=f"tsk_sel_{task_id}"),
            InlineKeyboardButton("Đóng", callback_data="tsk_cancel"),
        ],
    ])


def _format_task_detail(task) -> str:
    """Format chi tiết task thành text."""
    status_label = TASK_STATUS_MAP.get(task.status, task.status)
    return (
        f"<b>CHI TIẾT CÔNG VIỆC</b>\n\n"
        f"Người giao: <b>{task.assigner}</b>\n"
        f"Người nhận: <b>{task.assignee}</b>\n"
        f"Nội dung: <i>{task.content}</i>\n"
        f"Bắt đầu: <code>{task.start_date}</code>\n"
        f"Kết thúc: <code>{task.end_date}</code>\n"
        f"Chu kỳ: <b>{task.cycle or 'Một lần'}</b>\n"
        f"Trạng thái: <b>{status_label}</b>\n\n"
        f"<i>Mã task: <code>{task.id}</code></i>"
    )


async def handle_list_tasks(client, message: Message, command_name: str) -> None:
    """
    Hiển thị danh sách task của nhân viên dưới dạng inline buttons.
    """
    from app.models.task import Task

    username = message.from_user.username
    if not username:
        await message.reply_text(
            "⚠️ Bạn chưa cài đặt username Telegram.",
            parse_mode=ParseMode.HTML,
        )
        return

    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.username == username).first()
        if not employee:
            await message.reply_text(
                f"⚠️ Không tìm thấy nhân viên <b>@{username}</b> trong hệ thống.",
                parse_mode=ParseMode.HTML,
            )
            return

        tasks = (
            db.query(Task)
            .filter(Task.employee_id == employee.id)
            .order_by(Task.created_at.desc())
            .all()
        )

        if not tasks:
            await message.reply_text(
                "ℹ️ Bạn chưa có task nào được giao.",
                parse_mode=ParseMode.HTML,
            )
            return

        full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()
        markup = _build_task_list_buttons(tasks)

        await message.reply_text(
            f"<b>DANH SÁCH CÔNG VIỆC</b>\n"
            f"Nhân viên: <b>{full_name}</b>\n"
            f"Tổng: <b>{len(tasks)}</b> task\n\n"
            f"Chọn task để xem chi tiết:",
            parse_mode=ParseMode.HTML,
            reply_markup=markup,
        )

    except Exception as e:
        LogError(f"Error in handle_list_tasks: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(
            "❌ Có lỗi xảy ra khi lấy danh sách task.",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════
# CHECK TASKS (Admin xem task của nhân viên cụ thể)
# ══════════════════════════════════════════════════════════════

STATUS_ICON = {
    "PENDING": "⏳",
    "IN_PROGRESS": "🔄",
    "COMPLETED": "✅",
    "CANCELLED": "❌",
}

STATUS_LABEL = {
    "PENDING": "pending",
    "IN_PROGRESS": "in_progress",
    "COMPLETED": "completed",
    "CANCELLED": "cancelled",
}


async def handle_check_tasks(client, message: Message, command_name: str) -> None:
    """
    Addmin see all task of a employee.
    Syntax: /command [Mã NV] [Số tháng (optional, default=3)]
    Example:   /tien_nga_check_tasks TN001
             /tien_nga_check_tasks TN001 6
    """
    from app.models.task import Task

    parts = message.text.strip().split()

    if len(parts) < 2:
        await message.reply_text(
            f"<b>XEM CÔNG VIỆC NHÂN VIÊN</b>\n\n"
            f"Cú pháp: <code>{command_name} [Mã NV] [Số tháng]</code>\n"
            f"Ví dụ: <code>{command_name} TN001</code> (mặc định 3 tháng)\n"
            f"       <code>{command_name} TN001 6</code> (6 tháng gần nhất)",
            parse_mode=ParseMode.HTML
        )
        return

    emp_id = parts[1].strip()
    months = 3
    if len(parts) >= 3:
        try:
            months = int(parts[2])
            if months < 1 or months > 24:
                months = 3
        except ValueError:
            months = 3

    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.id == emp_id).first()
        if not employee:
            await message.reply_text(
                f"⚠️ Không tìm thấy nhân viên có mã <b>{emp_id}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        cutoff = datetime.datetime.now() - datetime.timedelta(days=months * 30)

        tasks = (
            db.query(Task)
            .filter(
                Task.employee_id == employee.id,
                Task.created_at >= cutoff
            )
            .order_by(Task.created_at.desc())
            .all()
        )

        full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()

        if not tasks:
            await message.reply_text(
                f"ℹ️ Nhân viên <b>{full_name}</b> ({emp_id}) không có task nào trong <b>{months} tháng</b> gần nhất.",
                parse_mode=ParseMode.HTML
            )
            return

        # Group by status
        groups: dict[str, list] = {
            "IN_PROGRESS": [],
            "PENDING": [],
            "COMPLETED": [],
            "CANCELLED": [],
        }
        for t in tasks:
            s = (t.status or "PENDING").upper()
            if s in groups:
                groups[s].append(t)
            else:
                groups.setdefault(s, []).append(t)

        lines = [
            f"<b>CÔNG VIỆC CỦA NHÂN VIÊN</b>",
            f"<b>{full_name}</b> | Mã: <code>{emp_id}</code>",
            f"{months} tháng gần nhất | Tổng: <b>{len(tasks)}</b> task",
            "─" * 30,
        ]

        for status_key in ("IN_PROGRESS", "PENDING", "COMPLETED", "CANCELLED"):
            group = groups.get(status_key, [])
            if not group:
                continue
            icon = STATUS_ICON.get(status_key, "❓")
            label = STATUS_LABEL.get(status_key, status_key)
            lines.append(f"\n{icon} <b>{label} ({len(group)})</b>")
            for t in group:
                content_preview = (t.content or "")[:60].replace("\n", " ")
                if len(t.content or "") > 60:
                    content_preview += "..."
                date_str = ""
                if t.end_date:
                    date_str = f" | HH: <code>{t.end_date}</code>"
                lines.append(f"  • {content_preview}{date_str}")

        text = "\n".join(lines)

        # Telegram message limit 4096 chars
        if len(text) > 4000:
            text = text[:3990] + "\n\n<i>...(còn thêm task, vui lòng lọc theo khoảng thời gian ngắn hơn)</i>"

        await message.reply_text(text, parse_mode=ParseMode.HTML)
        LogInfo(
            f"[CheckTasks] Admin {message.from_user.id} viewed tasks of {emp_id} ({months} months)",
            LogType.SYSTEM_STATUS
        )

    except Exception as e:
        LogError(f"Error in handle_check_tasks: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra khi lấy danh sách công việc.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# --- Callback: Chọn 1 task → hiển thị actions ---

async def task_select_callback(client, callback_query):
    from app.models.task import Task

    task_id = callback_query.data.replace("tsk_sel_", "")
    db = SessionLocal()
    try:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            await callback_query.answer("Task không tồn tại!", show_alert=True)
            return

        status_label = TASK_STATUS_MAP.get(task.status, task.status)
        text = (
            f"<b>CÔNG VIỆC</b>\n\n"
            f"{task.content[:60]}{'...' if len(task.content) > 60 else ''}\n"
            f"Trạng thái: <b>{status_label}</b>\n\n"
            f"Chọn thao tác:"
        )
        markup = _build_task_action_buttons(task_id)
        await callback_query.message.edit_text(text, parse_mode=ParseMode.HTML, reply_markup=markup)
        await callback_query.answer()
    finally:
        db.close()


# --- Callback: Xem chi tiết task ---
async def task_detail_callback(client, callback_query):
    from app.models.task import Task

    task_id = callback_query.data.replace("tsk_detail_", "")
    db = SessionLocal()
    try:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            await callback_query.answer("Task không tồn tại!", show_alert=True)
            return

        from pyrogram.types import InlineKeyboardButton, InlineKeyboardMarkup
        text = _format_task_detail(task)
        markup = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("Quay lại", callback_data=f"tsk_sel_{task_id}"),
                InlineKeyboardButton("Đóng", callback_data="tsk_cancel"),
            ],
        ])
        await callback_query.message.edit_text(text, parse_mode=ParseMode.HTML, reply_markup=markup)
        await callback_query.answer()
    finally:
        db.close()


# --- Callback: Hiển thị nút chọn trạng thái ---
async def task_status_callback(client, callback_query):
    from app.models.task import Task

    task_id = callback_query.data.replace("tsk_status_", "")
    db = SessionLocal()
    try:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            await callback_query.answer("Task không tồn tại!", show_alert=True)
            return

        current = TASK_STATUS_MAP.get(task.status, task.status)
        text = (
            f"<b>CẬP NHẬT TRẠNG THÁI</b>\n\n"
            f"{task.content[:60]}\n"
            f"Trạng thái hiện tại: <b>{current}</b>\n\n"
            f"Chọn trạng thái mới:"
        )
        markup = _build_status_buttons(task_id)
        await callback_query.message.edit_text(text, parse_mode=ParseMode.HTML, reply_markup=markup)
        await callback_query.answer()
    finally:
        db.close()


# --- Callback: Set trạng thái ---
async def task_set_status_callback(client, callback_query):
    import re
    from app.models.task import Task

    match = re.match(r"tsk_set_(PENDING|IN_PROGRESS|COMPLETED|CANCELLED)_(.+)", callback_query.data)
    if not match:
        await callback_query.answer("Lỗi dữ liệu!", show_alert=True)
        return

    new_status = match.group(1)
    task_id = match.group(2)

    db = SessionLocal()
    try:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            await callback_query.answer("Task không tồn tại!", show_alert=True)
            return

        old_status = task.status
        task.status = new_status
        task.updated_at = datetime.datetime.now()
        db.commit()

        old_label = TASK_STATUS_MAP.get(old_status, old_status)
        new_label = TASK_STATUS_MAP.get(new_status, new_status)

        text = (
            f"<b>CẬP NHẬT THÀNH CÔNG</b>\n\n"
            f"{task.content[:60]}\n"
            f"Trạng thái: {old_label} → <b>{new_label}</b>"
        )
        from pyrogram.types import InlineKeyboardButton, InlineKeyboardMarkup
        markup = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("Quay lại", callback_data=f"tsk_sel_{task_id}"),
                InlineKeyboardButton("Đóng", callback_data="tsk_cancel"),
            ],
        ])
        await callback_query.message.edit_text(text, parse_mode=ParseMode.HTML, reply_markup=markup)
        await callback_query.answer(f"Đã cập nhật: {new_label}")

        LogInfo(
            f"[TaskStatus] {callback_query.from_user.first_name} updated task {task_id}: {old_status} → {new_status}",
            LogType.SYSTEM_STATUS,
        )
    except Exception as e:
        db.rollback()
        LogError(f"Error in task_set_status_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("Lỗi cập nhật!", show_alert=True)
    finally:
        db.close()


# --- Callback: Quay lại danh sách ---
async def task_back_list_callback(client, callback_query):
    from app.models.task import Task

    username = callback_query.from_user.username
    if not username:
        await callback_query.answer("Không tìm thấy username!", show_alert=True)
        return

    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.username == username).first()
        if not employee:
            await callback_query.answer("Không tìm thấy nhân viên!", show_alert=True)
            return

        tasks = (
            db.query(Task)
            .filter(Task.employee_id == employee.id)
            .order_by(Task.created_at.desc())
            .all()
        )

        full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()
        markup = _build_task_list_buttons(tasks)

        await callback_query.message.edit_text(
            f"<b>DANH SÁCH CÔNG VIỆC</b>\n"
            f"Nhân viên: <b>{full_name}</b>\n"
            f"Tổng: <b>{len(tasks)}</b> task\n\n"
            f"Chọn task để xem chi tiết:",
            parse_mode=ParseMode.HTML,
            reply_markup=markup,
        )
        await callback_query.answer()
    finally:
        db.close()


# --- Callback: Đóng / Hủy ---
async def task_cancel_callback(client, callback_query):
    await callback_query.message.delete()
    await callback_query.answer()


# ══════════════════════════════════════════════════════════════
# CANCEL TASK BY REPLY (Hủy task bằng reply /cancel)
# ══════════════════════════════════════════════════════════════

async def handle_cancel_task_reply(client, message: Message) -> None:
    """
    Admin reply /cancel vào tin nhắn task → hủy task đó.
    Tìm task qua Mã task (UUID) trong tin nhắn được reply.
    """
    import re
    from app.models.task import Task

    replied_msg = message.reply_to_message
    if not replied_msg:
        return

    content = replied_msg.text or replied_msg.caption or ""

    # Tìm Mã task UUID trong tin nhắn
    match = re.search(r"Mã task:\s*([0-9a-fA-F\-]{36})", content)
    if not match:
        return  # Không phải tin nhắn task, bỏ qua

    task_id = match.group(1)

    db = SessionLocal()
    try:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            await message.reply_text(
                f"⚠️ Không tìm thấy task <code>{task_id}</code>.",
                parse_mode=ParseMode.HTML,
            )
            return

        if task.status == "CANCELLED":
            await message.reply_text(
                "ℹ️ Task này đã được hủy trước đó.",
                parse_mode=ParseMode.HTML,
            )
            return

        old_status = task.status
        task.status = "CANCELLED"
        task.updated_at = datetime.datetime.now()
        db.commit()

        old_label = TASK_STATUS_MAP.get(old_status, old_status)
        new_label = TASK_STATUS_MAP.get("CANCELLED", "CANCELLED")

        user = message.from_user
        admin_name = f"{user.first_name} {user.last_name or ''}".strip()

        await message.reply_text(
            f"<b>ĐÃ HỦY TASK</b>\n\n"
            f"Nội dung: <i>{task.content}</i>\n"
            f"Người nhận: <b>{task.assignee}</b>\n"
            f"Trạng thái: {old_label} → <b>{new_label}</b>\n"
            f"Hủy bởi: <b>{admin_name}</b>\n\n"
            f"<i>Mã task: <code>{task_id}</code></i>",
            parse_mode=ParseMode.HTML,
        )

        LogInfo(
            f"[CancelTask] {admin_name} cancelled task {task_id}: {task.content}",
            LogType.SYSTEM_STATUS,
        )

    except Exception as e:
        db.rollback()
        LogError(f"Error in handle_cancel_task_reply: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(
            "❌ Có lỗi xảy ra khi hủy task.",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════
# EXPORT PAYROLL
# ══════════════════════════════════════════════════════════════

async def handle_export_payroll(client, message, command_name: str) -> None:
    """
    Tạo ảnh payroll dựa trên tháng và năm. Lưu vào disk và gửi cho nhóm member.
    Quyền: Owner / Admin (Đã check ở handler).
    Cú pháp: /export_payroll [Mã NV] [MM/YYYY]
    """
    import os
    import calendar
    from app.models.employee import Employee
    from app.models.finance import Attendance, Payroll
    from app.models.telegram import TelegramProjectMember
    from bot.utils.payroll_generator import generate_payroll_image

    args = message.text.split()
    if len(args) != 3:
        await message.reply_text(
            f"⚠️ Lệnh không hợp lệ.\nCú pháp: <code>{command_name} [Mã nhân viên] [MM/YYYY]</code>\n"
            f"VD: <code>{command_name} TN001 04/2026</code>",
            parse_mode=ParseMode.HTML
        )
        return

    emp_code = args[1]
    month_year_str = args[2]

    try:
        month_str, year_str = month_year_str.split("/")
        month = int(month_str)
        year = int(year_str)
    except:
        await message.reply_text("⚠️ Định dạng tháng/năm phải là MM/YYYY.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        # Tìm Employee
        employee = db.query(Employee).filter(Employee.id == emp_code).first()
        if not employee:
            await message.reply_text(f"⚠️ Không tìm thấy nhân viên mã <code>{emp_code}</code>.", parse_mode=ParseMode.HTML)
            return

        full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()

        # Query attendance
        attendances = db.query(Attendance).filter(
            Attendance.employee_id == emp_code,
            Attendance.month == month,
            Attendance.year == year
        ).all()

        # Calculate standard days based on employee's work_type
        from bot.utils.scheduler import _is_working_day
        emp_work_type = employee.work_type if employee.work_type else 3
        num_days_in_month = calendar.monthrange(year, month)[1]
        standard_days = sum(1 for d in range(1, num_days_in_month + 1) if _is_working_day(datetime.date(year, month, d).weekday(), emp_work_type))

        actual_working_days = 0
        total_overtime = 0.0
        leave_days = 0

        attended_days = {a.day: a for a in attendances}
        for d in range(1, num_days_in_month + 1):
            att = attended_days.get(d)
            if att:
                # Actual working hours or if both checkin/out exist
                has_worked = False
                if (att.working_time or 0) > 0:
                    has_worked = True
                elif att.check_in_time and att.check_out_time:
                    has_worked = True
                    
                if has_worked:
                    actual_working_days += 1
                
                # Overtime
                if att.overtime:
                    total_overtime += att.overtime
                elif att.start_overtime and att.end_overtime:
                    diff = (att.end_overtime - att.start_overtime).total_seconds() / 3600
                    total_overtime += diff
                
                if att.error and "phép" in att.error.lower() and "không" not in att.error.lower():
                    leave_days += 1

        unpaid_leave = standard_days - actual_working_days - leave_days
        if unpaid_leave < 0:
            unpaid_leave = 0

        # Calculations
        base_salary = employee.monthly_salary or employee.base_salary or 0.0
        daily_salary = employee.daily_salary
        if not daily_salary and standard_days > 0:
            daily_salary = base_salary / standard_days
            
        salary_earned = round(daily_salary * actual_working_days, 2)
        
        overtime_rate = employee.overtime_salary or ((daily_salary / 8) * 1.5 if daily_salary else 0.0)
        overtime_salary_earned = round(total_overtime * overtime_rate, 2)
        
        bonus = employee.bonus or 0.0
        bhxh = employee.rate_bhxh or 0.0
        # penalty could be calculated here or passed directly
        penalty = 0.0
        
        total_net_salary = round(salary_earned + overtime_salary_earned + bonus - bhxh - penalty, 2)
        
        payroll_record = db.query(Payroll).filter(
            Payroll.employee_id == emp_code,
            Payroll.month == month,
            Payroll.year == year
        ).first()
        
        if not payroll_record:
            payroll_record = Payroll(employee_id=emp_code, month=month, year=year)
            db.add(payroll_record)
            old_salary = 0.0
        else:
            old_salary = payroll_record.total_salary or 0.0
            
        payroll_record.leave = leave_days
        payroll_record.unapproved_leave = unpaid_leave
        payroll_record.base_salary_amount = salary_earned
        payroll_record.overtime_salary_amount = overtime_salary_earned
        payroll_record.late_penalty = penalty
        payroll_record.total_salary = total_net_salary
        payroll_record.note = f"Thưởng: {bonus} | BHXH: {bhxh} | Chốt: {datetime.datetime.now().strftime('%d/%m/%Y')}"
        
        # Cập nhật cộng dồn công nợ lương cho nhân viên
        diff = total_net_salary - old_salary
        if diff != 0:
            if employee.total_debt is None:
                employee.total_debt = 0
            employee.total_debt += diff
            
        db.commit()

        data = {
            "standard_working_days": standard_days,
            "actual_working_days": actual_working_days,
            "leave_days": leave_days,
            "unpaid_leave": unpaid_leave,
            "total_overtime": round(total_overtime, 2),
            "leave_balance": employee.leave_balance or 0,
            
            "base_salary": base_salary,
            "salary_earned": salary_earned,
            "overtime_salary_earned": overtime_salary_earned,
            "bonus": bonus,
            "penalty": penalty,
            "bhxh": bhxh,
            "total_net_salary": total_net_salary,
            "position": employee.position or "Nhân sự",
            "department": employee.department or ""
        }

        # Validate generating image
        processing_msg = await message.reply_text("⏳ Đang tính toán và xuất bảng lương...")
        
        img_buf = await generate_payroll_image(
            employee_name=full_name,
            employee_id=emp_code,
            month=month,
            year=year,
            data=data
        )
        
        folder = os.path.join("images", str(year), f"{month:02d}")
        os.makedirs(folder, exist_ok=True)
        filename = f"{emp_code}_{month:02d}_{year}_bang_luong.png"
        filepath = os.path.join(folder, filename)

        with open(filepath, "wb") as f:
            f.write(img_buf.getvalue())
        
        # Tìm nhóm member của nhân viên qua Employee.telegram_group
        target_chat_id = None
        if employee.telegram_group:
            tg_group = employee.telegram_group.strip()
            # Try as numeric chat_id first
            try:
                target_chat_id = str(int(tg_group))
            except ValueError:
                # It's a group name, resolve from telegram_project_members
                tpm = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.group_name == tg_group
                ).first()
                if tpm:
                    target_chat_id = tpm.chat_id

        img_buf.seek(0)
        
        caption=(
            f"<b>BẢNG LƯONG THÁNG {month:02d}/{year}</b>\n"
            f"Nhân viên: <a href='tg://user?id={employee.username}'>Thuộc {full_name}</a>\n"
            f"Mã NV: <code>{emp_code}</code>\n\n"
            f"<i>Lưu ý: Nếu có thắc mắc vui lòng liên hệ admin.</i>"
        )
        
        if target_chat_id:
            await client.send_photo(
                chat_id=int(target_chat_id),
                photo=img_buf,
                caption=caption,
                parse_mode=ParseMode.HTML
            )
            # Re-upload for the admin/requesting group to confirm it's sent
            img_buf.seek(0)
            await processing_msg.delete()
            await message.reply_photo(
                photo=img_buf,
                caption=f"✅ Đã gửi bảng lương cho nhân viên <b>{full_name}</b> xuống nhóm Member.",
                parse_mode=ParseMode.HTML
            )
        else:
            await processing_msg.delete()
            await message.reply_photo(
                photo=img_buf,
                caption=f"✅ Đã tạo bảng lương cho nhân viên <b>{full_name}</b>.\n⚠️ Tuy nhiên, không tìm thấy nhóm Member_ns có chứa nhân viên này để gửi.",
                parse_mode=ParseMode.HTML
            )
            
        LogInfo(f"[Payroll] Exported payroll for {emp_code} - {month}/{year}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error in handle_export_payroll: {e}", LogType.SYSTEM_STATUS)
        try:
             await processing_msg.delete()
        except:
             pass
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình xuất bảng lương.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def handle_list_payroll_excel(client, message, command_name: str) -> None:
    from app.models.employee import Employee
    from app.models.finance import Payroll
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import os
    import tempfile

    args = message.text.split()
    if len(args) != 2:
        clean_cmd = command_name.split('@')[0]
        await message.reply_text(
            f"⚠️ Lệnh thiếu hoặc sai cú pháp.\nCú pháp: <code>{clean_cmd} [MM/YYYY]</code>\n"
            f"VD: <code>{clean_cmd} 04/2026</code>",
            parse_mode=ParseMode.HTML
        )
        return

    month_year_str = args[1]
    try:
        month_str, year_str = month_year_str.split("/")
        month = int(month_str)
        year = int(year_str)
    except:
        await message.reply_text("⚠️ Định dạng tháng/năm phải là MM/YYYY.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        processing_msg = await message.reply_text("⏳ Đang truy xuất thông tin DB và tạo file Excel...")
        
        # We fetch all payrolls for this month/year, joining with Employee to get names and department
        # Here we only get employees belonging to the project "Tiến Nga". 
        # For simplicity since this endpoint relies on project decorators, we trust the DB structure or check for non-None.
        results = db.query(Payroll, Employee).join(
            Employee, Payroll.employee_id == Employee.id
        ).filter(
            Payroll.month == month,
            Payroll.year == year
        ).all()

        if not results:
            await processing_msg.edit_text(f"⚠️ Không tìm thấy bất kỳ báo cáo lương nào trong tháng {month:02d}/{year}.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Payroll_{month:02d}_{year}"

        headers = [
            "STT", "Mã NV", "Họ tên", "Phòng ban", 
            "Ngày phép", "Nghỉ không phép", 
            "Lương cơ bản", "Lương tăng ca", "Tiền phạt", "Lương thực nhận", "Ghi chú"
        ]

        # Formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        def format_currency(val):
            if val is None: return 0
            return float(val)

        row_idx = 2
        total_salary_all = 0.0
        for i, (payroll, employee) in enumerate(results, 1):
            full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()
            
            ws.cell(row=row_idx, column=1, value=i).border = border
            ws.cell(row=row_idx, column=2, value=employee.id).border = border
            ws.cell(row=row_idx, column=3, value=full_name).border = border
            ws.cell(row=row_idx, column=4, value=employee.department or "").border = border
            
            ws.cell(row=row_idx, column=5, value=payroll.leave or 0).border = border
            ws.cell(row=row_idx, column=6, value=payroll.unapproved_leave or 0).border = border
            
            base_cell = ws.cell(row=row_idx, column=7, value=format_currency(payroll.base_salary_amount))
            base_cell.border = border
            base_cell.number_format = '#,##0'
            
            ot_cell = ws.cell(row=row_idx, column=8, value=format_currency(payroll.overtime_salary_amount))
            ot_cell.border = border
            ot_cell.number_format = '#,##0'
            
            penalty_cell = ws.cell(row=row_idx, column=9, value=format_currency(payroll.late_penalty))
            penalty_cell.border = border
            penalty_cell.number_format = '#,##0'
            
            net = format_currency(payroll.total_salary)
            total_salary_all += net
            
            total_cell = ws.cell(row=row_idx, column=10, value=net)
            total_cell.border = border
            total_cell.number_format = '#,##0'
            
            ws.cell(row=row_idx, column=11, value=payroll.note or "").border = border
            
            row_idx += 1

        # Dòng tổng cộng
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
        total_label = ws.cell(row=row_idx, column=1, value="TỔNG CỘNG:")
        total_label.font = Font(bold=True)
        total_label.alignment = Alignment(horizontal="right", vertical="center")
        for col_idx in range(1, 10):
            ws.cell(row=row_idx, column=col_idx).border = border
        
        total_sum_cell = ws.cell(row=row_idx, column=10, value=total_salary_all)
        total_sum_cell.font = Font(bold=True)
        total_sum_cell.border = border
        total_sum_cell.number_format = '#,##0'
        ws.cell(row=row_idx, column=11, value="").border = border

        # Column widths
        widths = [5, 12, 25, 20, 15, 15, 18, 18, 18, 20, 35]
        for col_num, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Dump
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix=f"payrolls_{month:02d}_{year}_")
        os.close(fd)
        wb.save(temp_path)

        await client.send_document(
            chat_id=message.chat.id,
            document=temp_path,
            file_name=f"Danh_sach_luong_{month:02d}_{year}.xlsx",
            caption=f"✅ <b>Báo cáo Danh Sách Bảng Lương tháng {month:02d}/{year}</b>\n\nTổng cộng: <b>{len(results)}</b> nhân viên\nTổng chi: <b>{total_salary_all:,.0f} đ</b>".replace(",", "."),
            parse_mode=ParseMode.HTML,
            reply_to_message_id=message.id
        )
        
        await processing_msg.delete()
        os.remove(temp_path)
            
    except Exception as e:
        LogError(f"Error in handle_list_payroll_excel: {e}", LogType.SYSTEM_STATUS)
        try:
             await processing_msg.edit_text("❌ Có lỗi xảy ra trong quá trình xuất bảng lương Excel.")
        except:
             await message.reply_text("❌ Có lỗi xảy ra trong quá trình xuất bảng lương Excel.")
    finally:
        db.close()

# ══════════════════════════════════════════════════════════════
# RECREATE ATTENDANCE REPORT (IMAGE)
# ══════════════════════════════════════════════════════════════

async def handle_recreate_attendance_report(client, message, command_name: str) -> None:
    """
    Tạo lại ảnh Bảng chấm công cho một nhân viên cụ thể theo tháng và gửi về nhóm yêu cầu (main).
    Quyền: Owner / Admin (Đã check ở handler).
    Cú pháp: /recreate_attendance_report [Mã NV] [MM/YYYY]
    """
    from app.models.employee import Employee
    from app.models.finance import Attendance
    from bot.utils.attendance_generator import generate_attendance_image
    import os

    args = message.text.split()
    if len(args) != 3:
        clean_cmd = command_name.split('@')[0]
        await message.reply_text(
            f"⚠️ Lệnh không hợp lệ.\nCú pháp: <code>{clean_cmd} [Mã nhân viên] [MM/YYYY]</code>\n"
            f"VD: <code>{clean_cmd} TN001 04/2026</code>",
            parse_mode=ParseMode.HTML
        )
        return

    emp_code = args[1]
    month_year_str = args[2]

    try:
        month_str, year_str = month_year_str.split("/")
        month = int(month_str)
        year = int(year_str)
    except:
        await message.reply_text("⚠️ Định dạng tháng/năm phải là MM/YYYY.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.id == emp_code).first()
        if not employee:
            await message.reply_text(f"⚠️ Không tìm thấy nhân viên mã <code>{emp_code}</code>.", parse_mode=ParseMode.HTML)
            return

        full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()

        records = db.query(Attendance).filter(
            Attendance.employee_id == emp_code,
            Attendance.month == month,
            Attendance.year == year
        ).order_by(Attendance.day).all()

        if not records:
            await message.reply_text(f"ℹ️ Không có dữ liệu chấm công cho nhân viên {full_name} trong tháng {month:02d}/{year}.", parse_mode=ParseMode.HTML)
            return

        processing_msg = await message.reply_text("⏳ Đang tạo lại ảnh bảng chấm công...")

        data_rows = []
        for att in records:
            ci = att.check_in_time.strftime("%H:%M") if att.check_in_time else "—"
            co = att.check_out_time.strftime("%H:%M") if att.check_out_time else "—"
            ot = "—"
            wh = "—"
            if att.start_overtime and att.end_overtime:
                ot = f"{att.start_overtime.strftime('%H:%M')}-{att.end_overtime.strftime('%H:%M')}"
            elif att.overtime:
                ot = f"{att.overtime}h"
            if att.working_time is not None:
                wh = f"{att.working_time:.1f}h"
            data_rows.append({
                "day": att.day, "month": att.month, "date_str": att.date_str or "",
                "check_in": ci, "check_out": co, "overtime": ot,
                "work_hours": wh, "error": att.error or "",
            })

        img_buf = await generate_attendance_image(
            employee_name=full_name, employee_id=emp_code,
            month=month, year=year, records=data_rows,
        )

        folder = os.path.join("images", str(year), f"{month:02d}")
        os.makedirs(folder, exist_ok=True)
        filepath = os.path.join(folder, f"{emp_code}_{month:02d}_{year}_tong_hop_cong.png")
        with open(filepath, "wb") as f:
            f.write(img_buf.getvalue())

        admin_username = message.from_user.username or "Admin"
        
        caption = (
            f"<b>BẢNG CHẤM CÔNG CẬP NHẬT/TẠO LẠI</b>\n"
            f"Nhân viên: <a href='tg://user?id={employee.username}'>Thuộc {full_name}</a>\n"
            f"Mã NV: <code>{emp_code}</code>\n"
            f"Tháng: <b>{month:02d}/{year}</b>\n"
            f"Yêu cầu bởi: @{admin_username}\n\n"
            f"<i>Reply /tien_nga_export_payroll {emp_code} {month:02d}/{year} để xuất bảng lương.</i>"
        )

        img_buf.seek(0)
        
        await processing_msg.delete()
        await message.reply_photo(
            photo=img_buf,
            caption=caption,
            parse_mode=ParseMode.HTML
        )
        
        LogInfo(f"[RecreateAttendanceImage] @{admin_username} recreated image for {emp_code} - {month}/{year}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error in handle_recreate_attendance_report: {e}", LogType.SYSTEM_STATUS)
        try:
             await processing_msg.delete()
        except:
             pass
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình tạo lại ảnh bảng chấm công.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

# ══════════════════════════════════════════════════════════════
# LIST ATTENDANCE (EXCEL)
# ══════════════════════════════════════════════════════════════

async def handle_list_attendance_excel(client, message, command_name: str) -> None:
    from app.models.employee import Employee
    from app.models.finance import Attendance
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import os
    import tempfile
    from pyrogram.enums import ParseMode

    args = message.text.split()
    if len(args) != 3:
        clean_cmd = command_name.split('@')[0]
        await message.reply_text(
            f"⚠️ Lệnh thiếu hoặc sai cú pháp.\nCú pháp: <code>{clean_cmd} [Mã nhân viên] [MM/YYYY]</code>\n"
            f"VD: <code>{clean_cmd} TN001 04/2026</code>",
            parse_mode=ParseMode.HTML
        )
        return

    emp_code = args[1]
    month_year_str = args[2]
    try:
        month_str, year_str = month_year_str.split("/")
        month = int(month_str)
        year = int(year_str)
    except:
        await message.reply_text("⚠️ Định dạng tháng/năm phải là MM/YYYY.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        processing_msg = await message.reply_text("⏳ Đang truy xuất thông tin DB và tạo file Excel chấm công...")
        
        employee = db.query(Employee).filter(Employee.id == emp_code).first()
        if not employee:
            await processing_msg.edit_text(f"⚠️ Không tìm thấy nhân viên mã <code>{emp_code}</code>.", parse_mode=ParseMode.HTML)
            return

        full_name = f"{employee.last_name or ''} {employee.first_name or ''}".strip()

        records = db.query(Attendance).filter(
            Attendance.employee_id == emp_code,
            Attendance.month == month,
            Attendance.year == year
        ).order_by(Attendance.day).all()

        if not records:
            await processing_msg.edit_text(f"⚠️ Không tìm thấy bất kỳ dữ liệu chấm công nào của {full_name} trong tháng {month:02d}/{year}.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"ChamCong_{month:02d}_{year}"

        headers = [
            "STT", "Ngày", "Thời gian", "Check In", "Check Out", "Tăng Ca", "Số Giờ Làm", "Ghi chú/Lỗi"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        row_idx = 2
        for i, att in enumerate(records, 1):
            ci = att.check_in_time.strftime("%H:%M") if att.check_in_time else "—"
            co = att.check_out_time.strftime("%H:%M") if att.check_out_time else "—"
            ot = "—"
            wh = "—"
            if att.start_overtime and att.end_overtime:
                ot = f"{att.start_overtime.strftime('%H:%M')}-{att.end_overtime.strftime('%H:%M')}"
            elif att.overtime:
                ot = f"{att.overtime}h"
            if att.working_time is not None:
                wh = f"{att.working_time:.1f}h"
                
            ws.cell(row=row_idx, column=1, value=i).border = border
            ws.cell(row=row_idx, column=2, value=att.date_str or f"{att.day:02d}/{att.month:02d}/{att.year}").border = border
            ws.cell(row=row_idx, column=3, value=f"{att.day:02d}/{att.month:02d}/{att.year}").border = border
            ws.cell(row=row_idx, column=4, value=ci).border = border
            ws.cell(row=row_idx, column=5, value=co).border = border
            ws.cell(row=row_idx, column=6, value=ot).border = border
            ws.cell(row=row_idx, column=7, value=wh).border = border
            ws.cell(row=row_idx, column=8, value=att.error or "").border = border
            
            row_idx += 1

        widths = [5, 15, 15, 15, 15, 15, 15, 40]
        for col_num, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        file_name = f"bang_cham_cong_{emp_code}_{month:02d}_{year}.xlsx"
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix=f"att_{emp_code}_{month:02d}_{year}_")
        os.close(fd)
        wb.save(temp_path)

        await client.send_document(
            chat_id=message.chat.id,
            document=temp_path,
            file_name=file_name,
            caption=f"✅ <b>Bảng Chấm Công tháng {month:02d}/{year}</b>\nNhân viên: <b>{full_name}</b> ({emp_code})",
            parse_mode=ParseMode.HTML,
            reply_to_message_id=message.id
        )
        
        await processing_msg.delete()
        os.remove(temp_path)
            
    except Exception as e:
        LogError(f"Error in handle_list_attendance_excel: {e}", LogType.SYSTEM_STATUS)
        try:
             await processing_msg.edit_text("❌ Có lỗi xảy ra trong quá trình xuất bảng chấm công.")
        except:
             pass
    finally:
        db.close()

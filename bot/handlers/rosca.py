from pyrogram import filters
from pyrogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from pyrogram.enums import ParseMode
from bot.utils.bot import bot
from bot.utils.utils import check_command_target, require_user_type, require_group_role, require_project_name
from bot.utils.enums import UserType
from bot.utils.logger import LogInfo, LogError, LogType
from app.db.session import SessionLocal
import uuid

# --- Rosca: Create User ---
@bot.on_message(filters.command(["hui_tao_nguoi_choi", "rosca_create_user"]) | filters.regex(r"^@\w+\s+/(hui_tao_nguoi_choi|rosca_create_user)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def rosca_create_user_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_tao_nguoi_choi", "rosca_create_user"])
    if args is None: return

    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        form_template = """<b>FORM TẠO NGƯỜI CHƠI / CHỦ HỤI</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/hui_tao_nguoi_choi
Mã ID: 
Họ và Tên: 
Username Telegram (không bắt buộc): 
Số Điện Thoại: 
Số CCCD: 
Vai Trò (Owner/Player): Player
</pre>

<i>Lưu ý: Mã ID là bắt buộc (VD: NC01, CH01). Vai trò mặc định là Player (Người chơi). Username điền định dạng @username hoặc bỏ trống.</i>"""
        await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        return

    # Parse Form
    data = {}
    for line in lines[1:]:
        if "):" in line:
            key, val = line.split("):", 1)
            key += ")"
        elif ":" in line:
            key, val = line.split(":", 1)
        else:
            continue
        data[key.strip()] = val.strip()

    user_id = data.get("Mã ID", "")
    full_name = data.get("Họ và Tên", "")
    username = data.get("Username Telegram (không bắt buộc)", "").replace("@", "")
    phone_number = data.get("Số Điện Thoại", "")
    cccd = data.get("Số CCCD", "")
    role = data.get("Vai Trò (Owner/Player)", "Player").capitalize()

    if not user_id:
        await message.reply_text("⚠️ <b>Mã ID</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    if not full_name:
        await message.reply_text("⚠️ <b>Họ và Tên</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    if role not in ["Owner", "Player"]:
        await message.reply_text("⚠️ <b>Vai Trò</b> chỉ được phép là <b>Owner</b> hoặc <b>Player</b>.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        from app.models.rosca import UserRosca

        # Check duplicate by ID
        existing_id = db.query(UserRosca).filter(UserRosca.id == user_id).first()
        if existing_id:
            await message.reply_text(f"⚠️ Mã ID <b>{user_id}</b> đã tồn tại trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        # Check duplicate by phone number or CCCD if provided
        if phone_number:
            existing = db.query(UserRosca).filter(UserRosca.phone_number == phone_number).first()
            if existing:
                await message.reply_text(f"⚠️ Người chơi với số điện thoại <b>{phone_number}</b> đã tồn tại trong hệ thống.", parse_mode=ParseMode.HTML)
                return
                
        if cccd:
            existing = db.query(UserRosca).filter(UserRosca.cccd == cccd).first()
            if existing:
                await message.reply_text(f"⚠️ Người chơi với CCCD <b>{cccd}</b> đã tồn tại trong hệ thống.", parse_mode=ParseMode.HTML)
                return

        new_user = UserRosca(
            id=user_id,
            full_name=full_name,
            username=username if username else None,
            phone_number=phone_number if phone_number else None,
            cccd=cccd if cccd else None,
            role=role,
            status="Active"
        )
        db.add(new_user)
        db.commit()

        await message.reply_text(f"✅ Đã tạo hồ sơ <b>{full_name}</b> (Vai trò: {role}) thành công!", parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaCreateUser] Created {role} {full_name} by {message.from_user.id}", LogType.SYSTEM_STATUS)
    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_create_user_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình lưu thông tin.")
    finally:
        db.close()

# --- Rosca: Update User ---
@bot.on_message(filters.command(["hui_cap_nhat_nguoi_choi", "rosca_update_user"]) | filters.regex(r"^@\w+\s+/(hui_cap_nhat_nguoi_choi|rosca_update_user)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def rosca_update_user_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_cap_nhat_nguoi_choi", "rosca_update_user"])
    if args is None: return

    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        if len(args) < 2:
            await message.reply_text("⚠️ Vui lòng cung cấp Mã ID của người chơi cần cập nhật.\nVí dụ: <code>/hui_cap_nhat_nguoi_choi NC01</code>", parse_mode=ParseMode.HTML)
            return
            
        user_id = args[1]
        db = SessionLocal()
        try:
            from app.models.rosca import UserRosca
            user = db.query(UserRosca).filter(UserRosca.id == user_id).first()
            if not user:
                await message.reply_text(f"⚠️ Không tìm thấy người chơi có Mã ID: <b>{user_id}</b>", parse_mode=ParseMode.HTML)
                return
                
            form_template = f"""<b>FORM CẬP NHẬT NGƯỜI CHƠI / CHỦ HỤI</b>
Vui lòng sao chép form dưới đây, chỉnh sửa thông tin và gửi lại:

<pre>/hui_cap_nhat_nguoi_choi {user_id}
Họ và Tên: {user.full_name or ''}
Username Telegram (không bắt buộc): {user.username or ''}
Số Điện Thoại: {user.phone_number or ''}
Số CCCD: {user.cccd or ''}
Vai Trò (Owner/Player): {user.role or 'Player'}
</pre>
"""
            await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error checking user in rosca_update_user: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Có lỗi xảy ra trong quá trình truy xuất thông tin.")
        finally:
            db.close()
        return

    # Parse Form
    if len(args) < 2:
        await message.reply_text("⚠️ Lệnh cập nhật không hợp lệ, thiếu Mã ID.", parse_mode=ParseMode.HTML)
        return
        
    user_id = args[1]
    data = {}
    for line in lines[1:]:
        if "):" in line:
            key, val = line.split("):", 1)
            key += ")"
        elif ":" in line:
            key, val = line.split(":", 1)
        else:
            continue
        data[key.strip()] = val.strip()

    full_name = data.get("Họ và Tên", "")
    username = data.get("Username Telegram (không bắt buộc)", "").replace("@", "")
    phone_number = data.get("Số Điện Thoại", "")
    cccd = data.get("Số CCCD", "")
    role = data.get("Vai Trò (Owner/Player)", "Player").capitalize()

    if not full_name:
        await message.reply_text("⚠️ <b>Họ và Tên</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    if role not in ["Owner", "Player"]:
        await message.reply_text("⚠️ <b>Vai Trò</b> chỉ được phép là <b>Owner</b> hoặc <b>Player</b>.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        from app.models.rosca import UserRosca
        user = db.query(UserRosca).filter(UserRosca.id == user_id).first()
        if not user:
            await message.reply_text(f"⚠️ Không tìm thấy người chơi có Mã ID: <b>{user_id}</b>", parse_mode=ParseMode.HTML)
            return

        # Check duplicate by phone number or CCCD if provided and changed
        if phone_number and phone_number != user.phone_number:
            existing = db.query(UserRosca).filter(UserRosca.phone_number == phone_number).first()
            if existing:
                await message.reply_text(f"⚠️ Số điện thoại <b>{phone_number}</b> đã được sử dụng bởi người khác.", parse_mode=ParseMode.HTML)
                return
                
        if cccd and cccd != user.cccd:
            existing = db.query(UserRosca).filter(UserRosca.cccd == cccd).first()
            if existing:
                await message.reply_text(f"⚠️ CCCD <b>{cccd}</b> đã được sử dụng bởi người khác.", parse_mode=ParseMode.HTML)
                return

        user.full_name = full_name
        user.username = username if username else None
        user.phone_number = phone_number if phone_number else None
        user.cccd = cccd if cccd else None
        user.role = role
        
        db.commit()

        await message.reply_text(f"✅ Đã cập nhật hồ sơ <b>{full_name}</b> (Vai trò: {role}) thành công!", parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaUpdateUser] Updated user {user_id} ({full_name}) by {message.from_user.id}", LogType.SYSTEM_STATUS)
    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_update_user_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình lưu thông tin.")
    finally:
        db.close()


# --- Rosca: Delete User ---
@bot.on_message(filters.command(["hui_xoa_nguoi_choi", "rosca_delete_user"]) | filters.regex(r"^@\w+\s+/(hui_xoa_nguoi_choi|rosca_delete_user)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def rosca_delete_user_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_xoa_nguoi_choi", "rosca_delete_user"])
    if args is None: return

    if len(args) < 2:
        await message.reply_text("⚠️ Vui lòng cung cấp Mã ID của người chơi cần xóa.\nVí dụ: <code>/hui_xoa_nguoi_choi NC01</code>", parse_mode=ParseMode.HTML)
        return
        
    user_id = args[1]
    db = SessionLocal()
    try:
        from app.models.rosca import UserRosca
        user = db.query(UserRosca).filter(UserRosca.id == user_id).first()
        if not user:
            await message.reply_text(f"⚠️ Không tìm thấy người chơi có Mã ID: <b>{user_id}</b>", parse_mode=ParseMode.HTML)
            return

        text = (
            f"<b>XÁC NHẬN XÓA NGƯỜI CHƠI / CHỦ HỤI</b>\n\n"
            f"- Mã ID: <b>{user.id}</b>\n"
            f"- Họ và Tên: <b>{user.full_name}</b>\n"
            f"- Vai trò: <b>{user.role}</b>\n"
            f"- SĐT: <b>{user.phone_number or 'N/A'}</b>\n\n"
            f"Bạn có chắc chắn muốn xóa hồ sơ này không?"
        )

        buttons = [
            [
                InlineKeyboardButton("Xác nhận", callback_data=f"rosca_deluser_confirm_{user.id}"),
                InlineKeyboardButton("Hủy", callback_data="rosca_deluser_cancel")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(buttons)
        await message.reply_text(text, reply_markup=reply_markup, parse_mode=ParseMode.HTML)
    except Exception as e:
        LogError(f"Error in rosca_delete_user_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình truy xuất thông tin.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rosca_deluser_confirm_(.+)$"))
async def rosca_delete_user_confirm_callback(client, callback_query: CallbackQuery):
    user_id = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rosca import UserRosca
        
        user = db.query(UserRosca).filter(UserRosca.id == user_id).first()
        if not user:
            await callback_query.message.edit_text("⚠️ Hồ sơ không tồn tại hoặc đã bị xóa.")
            return

        full_name = user.full_name
        db.delete(user)
        db.commit()

        await callback_query.message.edit_text(f"✅ Đã xóa thành công hồ sơ <b>{full_name}</b> (Mã ID: {user_id}).", parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaDeleteUser] User {user_id} ({full_name}) was deleted by {callback_query.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_deluser_confirm callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi xác nhận xóa.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rosca_deluser_cancel$"))
async def rosca_delete_user_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("❌ Đã hủy thao tác xóa hồ sơ.")


# --- Rosca: Create Rosca (Dây Hụi) ---
@bot.on_message(filters.command(["hui_tao_day_hui", "roscas_create_roscas"]) | filters.regex(r"^@\w+\s+/(hui_tao_day_hui|roscas_create_roscas)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def rosca_create_roscas_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_tao_day_hui", "roscas_create_roscas"])
    if args is None: return

    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        form_template = """<b>FORM TẠO DÂY HỤI</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/hui_tao_day_hui
Mã Dây Hụi: 
ID Chủ Hụi (Mã ID): 
Số Tiền Gốc 1 Chân (VNĐ): 
Mức Bỏ Hụi Tối Thiểu (VNĐ): 
Mức Bỏ Hụi Tối Đa (VNĐ): 
Tổng Số Chân Hụi: 
Tiền Thảo (VNĐ): 
Ngày Bắt Đầu (DD/MM/YYYY): 
Ngày Kết Thúc (DD/MM/YYYY): 
Ngày Đóng Hụi Hàng Kỳ (1-31): 
Giờ Khui Hụi (HH:MM): 
Loại Hụi (Hụi ngày/Hụi tuần/Hụi 2 tuần/Hụi Tháng): Hụi Tháng
Trạng Thái (Draft/Active/Closed): Active
Ghi Chú: 
</pre>

<i>Lưu ý: Bạn có thể bỏ trống một số trường nếu chưa xác định, nhưng Mã Dây Hụi, ID Chủ Hụi và Số Tiền Gốc là bắt buộc. Số tiền viết liền không dấu phẩy hoặc có chấm phẩy đều được (VD: 10000000 hoặc 10.000.000).</i>"""
        await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        return

    # Parse Form
    data = {}
    for line in lines[1:]:
        if "):" in line:
            key, val = line.split("):", 1)
            key += ")"
        elif ":" in line:
            key, val = line.split(":", 1)
        else:
            continue
        data[key.strip()] = val.strip()

    code = data.get("Mã Dây Hụi", "")
    user_id = data.get("ID Chủ Hụi (Mã ID)", "")
    base_amount_str = data.get("Số Tiền Gốc 1 Chân (VNĐ)", "")
    min_bid_amount_str = data.get("Mức Bỏ Hụi Tối Thiểu (VNĐ)", "")
    max_bid_amount_str = data.get("Mức Bỏ Hụi Tối Đa (VNĐ)", "")
    total_parts_str = data.get("Tổng Số Chân Hụi", "")
    commission_fee_str = data.get("Tiền Thảo (VNĐ)", "")
    start_date_str = data.get("Ngày Bắt Đầu (DD/MM/YYYY)", "")
    end_date_str = data.get("Ngày Kết Thúc (DD/MM/YYYY)", "")
    payment_day_str = data.get("Ngày Đóng Hụi Hàng Kỳ (1-31)", "")
    bidding_time_str = data.get("Giờ Khui Hụi (HH:MM)", "")
    period_type_str = data.get("Loại Hụi (Hụi ngày/Hụi tuần/Hụi 2 tuần/Hụi Tháng)", "Hụi Tháng")
    status = data.get("Trạng Thái (Draft/Active/Closed)", "Active").capitalize()
    note = data.get("Ghi Chú", "")

    if not code:
        await message.reply_text("⚠️ <b>Mã Dây Hụi</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not user_id:
        await message.reply_text("⚠️ <b>ID Chủ Hụi</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not base_amount_str:
        await message.reply_text("⚠️ <b>Số Tiền Gốc</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    def parse_float(val_str):
        if not val_str: return None
        try:
            return float(val_str.replace(",", "").replace(".", "").replace(" ", ""))
        except:
            return None

    def parse_int(val_str):
        if not val_str: return None
        try:
            return int(val_str.replace(",", "").replace(".", "").replace(" ", ""))
        except:
            return None

    from datetime import datetime
    def parse_date(val_str):
        if not val_str: return None
        try:
            return datetime.strptime(val_str, "%d/%m/%Y").date()
        except:
            return None

    def parse_time(val_str):
        if not val_str: return None
        try:
            return datetime.strptime(val_str, "%H:%M").time()
        except:
            return None

    base_amount = parse_float(base_amount_str)
    min_bid_amount = parse_float(min_bid_amount_str)
    max_bid_amount = parse_float(max_bid_amount_str)
    total_parts = parse_int(total_parts_str)
    commission_fee = parse_float(commission_fee_str)
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    payment_day = parse_int(payment_day_str)
    bidding_time = parse_time(bidding_time_str)
    
    if base_amount is None:
        await message.reply_text("⚠️ <b>Số Tiền Gốc</b> không hợp lệ.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        from app.models.rosca import Rosca, UserRosca
        
        # Check if owner exists
        owner = db.query(UserRosca).filter(UserRosca.id == user_id, UserRosca.role == "Owner").first()
        if not owner:
            await message.reply_text(f"⚠️ Không tìm thấy Chủ hụi có ID: <b>{user_id}</b> (hoặc người này không phải Owner).", parse_mode=ParseMode.HTML)
            return

        # Check duplicate code
        existing_rosca = db.query(Rosca).filter(Rosca.code == code).first()
        if existing_rosca:
            await message.reply_text(f"⚠️ Dây hụi có mã <b>{code}</b> đã tồn tại.", parse_mode=ParseMode.HTML)
            return

        # Determine period type Enum
        from app.models.rosca import RoscaPeriodType
        period_type_enum = None
        for pt in RoscaPeriodType:
            if pt.value.lower() == period_type_str.lower():
                period_type_enum = pt
                break
        
        if not period_type_enum:
            period_type_enum = RoscaPeriodType.MONTHLY

        new_rosca = Rosca(
            id=str(uuid.uuid4()),
            code=code,
            user_id=user_id,
            base_amount=base_amount,
            min_bid_amount=min_bid_amount,
            max_bid_amount=max_bid_amount,
            total_parts=total_parts,
            commission_fee=commission_fee,
            start_date=start_date,
            end_date=end_date,
            payment_day=payment_day,
            bidding_time=bidding_time,
            period_type=period_type_enum,
            status=status,
            note=note
        )
        db.add(new_rosca)
        db.commit()

        await message.reply_text(f"✅ Đã tạo thành công Dây hụi <b>{code}</b> (Chủ hụi: {owner.full_name}).\nTrạng thái: <b>{status}</b>", parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaCreate] Created Rosca {code} by {message.from_user.id}", LogType.SYSTEM_STATUS)
    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_create_roscas_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình lưu thông tin Dây hụi.")
    finally:
        db.close()


# --- Rosca: Update Rosca (Dây Hụi) ---
@bot.on_message(filters.command(["hui_cap_nhat_day_hui", "roscas_update_roscas"]) | filters.regex(r"^@\w+\s+/(hui_cap_nhat_day_hui|roscas_update_roscas)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def rosca_update_roscas_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_cap_nhat_day_hui", "roscas_update_roscas"])
    if args is None: return

    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        if len(args) < 2:
            await message.reply_text("⚠️ Vui lòng cung cấp Mã Dây Hụi cần cập nhật.\nVí dụ: <code>/hui_cap_nhat_day_hui DH01</code>", parse_mode=ParseMode.HTML)
            return
            
        code = args[1]
        db = SessionLocal()
        try:
            from app.models.rosca import Rosca
            rosca = db.query(Rosca).filter(Rosca.code == code).first()
            if not rosca:
                await message.reply_text(f"⚠️ Không tìm thấy Dây hụi có mã: <b>{code}</b>", parse_mode=ParseMode.HTML)
                return
                
            from bot.utils.utils import fmt_num
            form_template = f"""<b>FORM CẬP NHẬT DÂY HỤI</b>
Vui lòng sao chép form dưới đây, chỉnh sửa thông tin và gửi lại:

<pre>/hui_cap_nhat_day_hui {code}
ID Chủ Hụi (Mã ID): {rosca.user_id or ''}
Số Tiền Gốc 1 Chân (VNĐ): {fmt_num(rosca.base_amount) if rosca.base_amount else ''}
Mức Bỏ Hụi Tối Thiểu (VNĐ): {fmt_num(rosca.min_bid_amount) if rosca.min_bid_amount else ''}
Mức Bỏ Hụi Tối Đa (VNĐ): {fmt_num(rosca.max_bid_amount) if rosca.max_bid_amount else ''}
Tổng Số Chân Hụi: {rosca.total_parts or ''}
Tiền Thảo (VNĐ): {fmt_num(rosca.commission_fee) if rosca.commission_fee else ''}
Ngày Bắt Đầu (DD/MM/YYYY): {rosca.start_date.strftime('%d/%m/%Y') if rosca.start_date else ''}
Ngày Kết Thúc (DD/MM/YYYY): {rosca.end_date.strftime('%d/%m/%Y') if rosca.end_date else ''}
Ngày Đóng Hụi Hàng Kỳ (1-31): {rosca.payment_day or ''}
Giờ Khui Hụi (HH:MM): {rosca.bidding_time.strftime('%H:%M') if rosca.bidding_time else ''}
Loại Hụi (Hụi ngày/Hụi tuần/Hụi 2 tuần/Hụi Tháng): {rosca.period_type.value if rosca.period_type else 'Hụi Tháng'}
Trạng Thái (Draft/Active/Closed): {rosca.status or 'Active'}
Ghi Chú: {rosca.note or ''}
</pre>"""
            await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error checking rosca in rosca_update_roscas: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Có lỗi xảy ra trong quá trình truy xuất thông tin.")
        finally:
            db.close()
        return

    # Parse Form
    if len(args) < 2:
        await message.reply_text("⚠️ Lệnh cập nhật không hợp lệ, thiếu Mã Dây Hụi.", parse_mode=ParseMode.HTML)
        return
        
    code = args[1]
    data = {}
    for line in lines[1:]:
        if "):" in line:
            key, val = line.split("):", 1)
            key += ")"
        elif ":" in line:
            key, val = line.split(":", 1)
        else:
            continue
        data[key.strip()] = val.strip()

    user_id = data.get("ID Chủ Hụi (Mã ID)", "")
    base_amount_str = data.get("Số Tiền Gốc 1 Chân (VNĐ)", "")
    min_bid_amount_str = data.get("Mức Bỏ Hụi Tối Thiểu (VNĐ)", "")
    max_bid_amount_str = data.get("Mức Bỏ Hụi Tối Đa (VNĐ)", "")
    total_parts_str = data.get("Tổng Số Chân Hụi", "")
    commission_fee_str = data.get("Tiền Thảo (VNĐ)", "")
    start_date_str = data.get("Ngày Bắt Đầu (DD/MM/YYYY)", "")
    end_date_str = data.get("Ngày Kết Thúc (DD/MM/YYYY)", "")
    payment_day_str = data.get("Ngày Đóng Hụi Hàng Kỳ (1-31)", "")
    bidding_time_str = data.get("Giờ Khui Hụi (HH:MM)", "")
    period_type_str = data.get("Loại Hụi (Hụi ngày/Hụi tuần/Hụi 2 tuần/Hụi Tháng)", "Hụi Tháng")
    status = data.get("Trạng Thái (Draft/Active/Closed)", "Active").capitalize()
    note = data.get("Ghi Chú", "")

    if not user_id:
        await message.reply_text("⚠️ <b>ID Chủ Hụi</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not base_amount_str:
        await message.reply_text("⚠️ <b>Số Tiền Gốc</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    def parse_float(val_str):
        if not val_str: return None
        try:
            return float(val_str.replace(",", "").replace(".", "").replace(" ", ""))
        except:
            return None

    def parse_int(val_str):
        if not val_str: return None
        try:
            return int(val_str.replace(",", "").replace(".", "").replace(" ", ""))
        except:
            return None

    from datetime import datetime
    def parse_date(val_str):
        if not val_str: return None
        try:
            return datetime.strptime(val_str, "%d/%m/%Y").date()
        except:
            return None

    def parse_time(val_str):
        if not val_str: return None
        try:
            return datetime.strptime(val_str, "%H:%M").time()
        except:
            return None

    base_amount = parse_float(base_amount_str)
    min_bid_amount = parse_float(min_bid_amount_str)
    max_bid_amount = parse_float(max_bid_amount_str)
    total_parts = parse_int(total_parts_str)
    commission_fee = parse_float(commission_fee_str)
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    payment_day = parse_int(payment_day_str)
    bidding_time = parse_time(bidding_time_str)
    
    if base_amount is None:
        await message.reply_text("⚠️ <b>Số Tiền Gốc</b> không hợp lệ.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        from app.models.rosca import Rosca, UserRosca
        
        rosca = db.query(Rosca).filter(Rosca.code == code).first()
        if not rosca:
            await message.reply_text(f"⚠️ Không tìm thấy Dây hụi có mã: <b>{code}</b>", parse_mode=ParseMode.HTML)
            return

        # Check if owner exists
        owner = db.query(UserRosca).filter(UserRosca.id == user_id, UserRosca.role == "Owner").first()
        if not owner:
            await message.reply_text(f"⚠️ Không tìm thấy Chủ hụi có ID: <b>{user_id}</b> (hoặc người này không phải Owner).", parse_mode=ParseMode.HTML)
            return

        # Determine period type Enum
        from app.models.rosca import RoscaPeriodType
        period_type_enum = None
        for pt in RoscaPeriodType:
            if pt.value.lower() == period_type_str.lower():
                period_type_enum = pt
                break
        
        if not period_type_enum:
            period_type_enum = RoscaPeriodType.MONTHLY

        rosca.user_id = user_id
        rosca.base_amount = base_amount
        rosca.min_bid_amount = min_bid_amount
        rosca.max_bid_amount = max_bid_amount
        rosca.total_parts = total_parts
        rosca.commission_fee = commission_fee
        rosca.start_date = start_date
        rosca.end_date = end_date
        rosca.payment_day = payment_day
        rosca.bidding_time = bidding_time
        rosca.period_type = period_type_enum
        rosca.status = status
        rosca.note = note
        
        db.commit()

        await message.reply_text(f"✅ Đã cập nhật thành công Dây hụi <b>{code}</b>.", parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaUpdate] Updated Rosca {code} by {message.from_user.id}", LogType.SYSTEM_STATUS)
    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_update_roscas_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình cập nhật thông tin Dây hụi.")
    finally:
        db.close()


# --- Rosca: Delete Rosca (Dây Hụi) ---
@bot.on_message(filters.command(["hui_xoa_day_hui", "roscas_delete_roscas"]) | filters.regex(r"^@\w+\s+/(hui_xoa_day_hui|roscas_delete_roscas)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def rosca_delete_roscas_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_xoa_day_hui", "roscas_delete_roscas"])
    if args is None: return

    if len(args) < 2:
        await message.reply_text("⚠️ Vui lòng cung cấp Mã Dây Hụi cần xóa.\nVí dụ: <code>/hui_xoa_day_hui DH01</code>", parse_mode=ParseMode.HTML)
        return
        
    code = args[1]
    db = SessionLocal()
    try:
        from app.models.rosca import Rosca, UserRosca
        from bot.utils.utils import fmt_num
        
        rosca = db.query(Rosca).filter(Rosca.code == code).first()
        if not rosca:
            await message.reply_text(f"⚠️ Không tìm thấy Dây hụi có mã: <b>{code}</b>", parse_mode=ParseMode.HTML)
            return
            
        owner = db.query(UserRosca).filter(UserRosca.id == rosca.user_id).first()
        owner_name = owner.full_name if owner else "Không xác định"

        text = (
            f"<b>XÁC NHẬN XÓA DÂY HỤI</b>\n\n"
            f"- Mã Dây Hụi: <b>{rosca.code}</b>\n"
            f"- Số Tiền Gốc: <b>{fmt_num(rosca.base_amount)} VNĐ</b>\n"
            f"- Tổng Số Chân: <b>{rosca.total_parts or 'N/A'}</b>\n"
            f"- Chủ Hụi: <b>{owner_name} ({rosca.user_id})</b>\n\n"
            f"Bạn có chắc chắn muốn xóa (chuyển sang trạng thái Deleted) dây hụi này không?"
        )

        buttons = [
            [
                InlineKeyboardButton("Xác nhận", callback_data=f"rosca_delrosca_confirm_{rosca.id}"),
                InlineKeyboardButton("Hủy", callback_data="rosca_delrosca_cancel")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(buttons)
        await message.reply_text(text, reply_markup=reply_markup, parse_mode=ParseMode.HTML)
    except Exception as e:
        LogError(f"Error in rosca_delete_roscas_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình truy xuất thông tin.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rosca_delrosca_confirm_(.+)$"))
async def rosca_delete_rosca_confirm_callback(client, callback_query: CallbackQuery):
    rosca_id = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rosca import Rosca
        
        rosca = db.query(Rosca).filter(Rosca.id == rosca_id).first()
        if not rosca:
            await callback_query.message.edit_text("⚠️ Dây hụi không tồn tại hoặc đã bị xóa.")
            return

        code = rosca.code
        rosca.status = "Deleted"
        db.commit()

        await callback_query.message.edit_text(f"✅ Đã xóa thành công Dây hụi <b>{code}</b> (chuyển trạng thái sang Deleted).", parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaDelete] Rosca {code} was deleted by {callback_query.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_delrosca_confirm callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi xác nhận xóa.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rosca_delrosca_cancel$"))
async def rosca_delete_rosca_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("❌ Đã hủy thao tác xóa Dây hụi.")


# --- Rosca: Create Rosca Member (Chân Hụi) ---
@bot.on_message(filters.command(["hui_tao_chan_hui", "roscas_create_member"]) | filters.regex(r"^@\w+\s+/(hui_tao_chan_hui|roscas_create_member)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def rosca_create_member_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_tao_chan_hui", "roscas_create_member"])
    if args is None: return

    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        db = SessionLocal()
        try:
            from app.models.rosca import Rosca, UserRosca
            from bot.utils.utils import fmt_num
            active_roscas = db.query(Rosca).filter(Rosca.status.in_(["Draft", "Active"])).all()
            if not active_roscas:
                await message.reply_text("⚠️ Hiện tại chưa có Dây hụi nào đang hoạt động.", parse_mode=ParseMode.HTML)
                return
            
            buttons = []
            for r in active_roscas:
                owner = db.query(UserRosca).filter(UserRosca.id == r.user_id).first()
                owner_name = owner.full_name if owner else "Chưa rõ"
                btn_text = f"Mã: {r.code} | {owner_name} ({r.total_parts or '?'} chân)"
                cb_data = f"cb_crmem_r_{r.code}"
                buttons.append([InlineKeyboardButton(btn_text, callback_data=cb_data)])
            
            buttons.append([InlineKeyboardButton("Hủy", callback_data="rosca_delrosca_cancel")])
            
            await message.reply_text("<b>CHỌN DÂY HỤI</b>\nVui lòng chọn Dây hụi mà bạn muốn thêm chân:", reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error fetching roscas for create member: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Có lỗi xảy ra trong quá trình lấy danh sách Dây hụi.")
        finally:
            db.close()
        return

    # Parse Form
    data = {}
    for line in lines[1:]:
        if "):" in line:
            key, val = line.split("):", 1)
            key += ")"
        elif ":" in line:
            key, val = line.split(":", 1)
        else:
            continue
        data[key.strip()] = val.strip()

    member_id = data.get("Mã Chân Hụi", "")
    rosca_code = data.get("Mã Dây Hụi", "")
    user_id = data.get("ID Người Chơi (Mã ID)", "")
    parts_count_str = data.get("Số Lượng Chân", "1")
    total_contributed_str = data.get("Tổng Tiền Đã Đóng (VNĐ)", "0")
    total_received_str = data.get("Tổng Tiền Đã Nhận (VNĐ)", "0")
    total_profit_str = data.get("Tổng Tiền Lãi (VNĐ)", "0")
    profit_rate_str = data.get("Tỷ Suất Lợi Nhuận (%)", "0")
    status = data.get("Trạng Thái (Playing/Defaulted)", "Playing")
    telegram_group = data.get("Nhóm Chat Telegram (tùy chọn)", "")
    note = data.get("Ghi Chú", "")

    if not member_id:
        await message.reply_text("⚠️ <b>Mã Chân Hụi</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not rosca_code:
        await message.reply_text("⚠️ <b>Mã Dây Hụi</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not user_id:
        await message.reply_text("⚠️ <b>ID Người Chơi</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    def parse_float(val_str):
        if not val_str: return 0.0
        try:
            return float(val_str.replace(",", "").replace(".", "").replace(" ", ""))
        except:
            return 0.0

    try:
        parts_count = int(parts_count_str.replace(",", "").replace(".", "").replace(" ", ""))
        if parts_count <= 0:
            raise ValueError("Số lượng chân phải lớn hơn 0.")
    except:
        await message.reply_text("⚠️ <b>Số Lượng Chân</b> không hợp lệ.", parse_mode=ParseMode.HTML)
        return

    total_contributed = parse_float(total_contributed_str)
    total_received = parse_float(total_received_str)
    total_profit = parse_float(total_profit_str)
    profit_rate = parse_float(profit_rate_str)

    db = SessionLocal()
    try:
        from app.models.rosca import Rosca, UserRosca, RoscaMember
        
        # Check if member ID exists
        existing_id = db.query(RoscaMember).filter(RoscaMember.id == member_id).first()
        if existing_id:
            await message.reply_text(f"⚠️ Mã Chân Hụi <b>{member_id}</b> đã tồn tại trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        # Look up Rosca by code
        rosca = db.query(Rosca).filter(Rosca.code == rosca_code).first()
        if not rosca:
            await message.reply_text(f"⚠️ Không tìm thấy Dây hụi có mã: <b>{rosca_code}</b>", parse_mode=ParseMode.HTML)
            return

        # Check if user exists
        player = db.query(UserRosca).filter(UserRosca.id == user_id).first()
        if not player:
            await message.reply_text(f"⚠️ Không tìm thấy Người chơi có ID: <b>{user_id}</b>.", parse_mode=ParseMode.HTML)
            return

        # Check if user is already a member in this Rosca
        existing_member = db.query(RoscaMember).filter(
            RoscaMember.rosca_id == rosca.id,
            RoscaMember.user_id == user_id
        ).first()
        
        if existing_member:
            await message.reply_text(f"⚠️ Người chơi <b>{player.full_name}</b> (ID: {user_id}) đã tham gia Dây hụi này với <b>{existing_member.parts_count}</b> chân. Vui lòng cập nhật thay vì tạo mới.", parse_mode=ParseMode.HTML)
            return

        # Sum of current parts to check if it exceeds total_parts of Rosca
        current_members = db.query(RoscaMember).filter(RoscaMember.rosca_id == rosca.id).all()
        current_total_parts = sum([m.parts_count for m in current_members if m.parts_count])
        
        if rosca.total_parts and (current_total_parts + parts_count > rosca.total_parts):
            await message.reply_text(f"⚠️ Không thể thêm. Dây hụi này chỉ còn <b>{rosca.total_parts - current_total_parts}</b> chân trống (Tổng số: {rosca.total_parts}).", parse_mode=ParseMode.HTML)
            return

        new_member = RoscaMember(
            id=member_id,
            rosca_id=rosca.id,
            user_id=user_id,
            parts_count=parts_count,
            total_contributed=total_contributed,
            total_received=total_received,
            total_profit=total_profit,
            profit_rate=profit_rate,
            status=status,
            note=note,
            telegram_group=telegram_group if telegram_group else None
        )
        db.add(new_member)
        db.commit()

        await message.reply_text(f"✅ Đã thêm <b>{player.full_name}</b> (ID: {user_id}) vào Dây hụi <b>{rosca_code}</b> với <b>{parts_count}</b> chân.\nMã Chân Hụi: <b>{member_id}</b>", parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaAddMember] Added user {user_id} (ID: {member_id}) to Rosca {rosca_code} with {parts_count} parts by {message.from_user.id}", LogType.SYSTEM_STATUS)
    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_create_member_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình thêm Chân hụi.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^cb_crmem_r_(.+)$"))
async def rosca_create_member_select_callback(client, callback_query: CallbackQuery):
    rosca_code = callback_query.matches[0].group(1)
    
    form_template = f"""<b>FORM TẠO CHÂN HỤI</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/hui_tao_chan_hui
Mã Chân Hụi: 
Mã Dây Hụi: {rosca_code}
ID Người Chơi (Mã ID): 
Số Lượng Chân: 1
Tổng Tiền Đã Đóng (VNĐ): 0
Tổng Tiền Đã Nhận (VNĐ): 0
Tổng Tiền Lãi (VNĐ): 0
Tỷ Suất Lợi Nhuận (%): 0
Trạng Thái (Playing/Defaulted): Playing
Nhóm Chat Telegram (tùy chọn): 
Ghi Chú: 
</pre>

<i>Lưu ý: Mã Chân Hụi, Mã Dây Hụi và ID Người Chơi là bắt buộc. ID Người Chơi phải tồn tại trong hệ thống. Nếu bạn tạo mới từ đầu, hãy giữ nguyên các số tiền là 0.</i>"""
    await callback_query.message.edit_text(form_template, parse_mode=ParseMode.HTML)


# --- Rosca: Update Rosca Member (Chân Hụi) ---
@bot.on_message(filters.command(["hui_cap_nhat_chan_hui", "roscas_update_member"]) | filters.regex(r"^@\w+\s+/(hui_cap_nhat_chan_hui|roscas_update_member)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def rosca_update_member_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_cap_nhat_chan_hui", "roscas_update_member"])
    if args is None: return

    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        if len(args) < 2:
            await message.reply_text("⚠️ Vui lòng cung cấp Mã Chân Hụi cần cập nhật.\nVí dụ: <code>/hui_cap_nhat_chan_hui AT001</code>", parse_mode=ParseMode.HTML)
            return
            
        member_id = args[1]
        db = SessionLocal()
        try:
            from app.models.rosca import RoscaMember, Rosca
            from bot.utils.utils import fmt_num
            member = db.query(RoscaMember).filter(RoscaMember.id == member_id).first()
            if not member:
                await message.reply_text(f"⚠️ Không tìm thấy Chân hụi có mã: <b>{member_id}</b>", parse_mode=ParseMode.HTML)
                return
            
            rosca = db.query(Rosca).filter(Rosca.id == member.rosca_id).first()
            rosca_code = rosca.code if rosca else ""
            
            form_template = f"""<b>FORM CẬP NHẬT CHÂN HỤI</b>
Vui lòng sao chép form dưới đây, chỉnh sửa thông tin và gửi lại:

<pre>/hui_cap_nhat_chan_hui {member_id}
Mã Dây Hụi: {rosca_code}
ID Người Chơi (Mã ID): {member.user_id or ''}
Số Lượng Chân: {member.parts_count or 1}
Tổng Tiền Đã Đóng (VNĐ): {fmt_num(member.total_contributed)}
Tổng Tiền Đã Nhận (VNĐ): {fmt_num(member.total_received)}
Tổng Tiền Lãi (VNĐ): {fmt_num(member.total_profit)}
Tỷ Suất Lợi Nhuận (%): {fmt_num(member.profit_rate)}
Trạng Thái (Playing/Defaulted): {member.status or 'Playing'}
Nhóm Chat Telegram (tùy chọn): {member.telegram_group or ''}
Ghi Chú: {member.note or ''}
</pre>"""
            await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error checking member in rosca_update_member: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Có lỗi xảy ra trong quá trình truy xuất thông tin.")
        finally:
            db.close()
        return

    # Parse Form
    if len(args) < 2:
        await message.reply_text("⚠️ Lệnh cập nhật không hợp lệ, thiếu Mã Chân Hụi.", parse_mode=ParseMode.HTML)
        return
        
    member_id = args[1]
    data = {}
    for line in lines[1:]:
        if "):" in line:
            key, val = line.split("):", 1)
            key += ")"
        elif ":" in line:
            key, val = line.split(":", 1)
        else:
            continue
        data[key.strip()] = val.strip()

    rosca_code = data.get("Mã Dây Hụi", "")
    user_id = data.get("ID Người Chơi (Mã ID)", "")
    parts_count_str = data.get("Số Lượng Chân", "1")
    total_contributed_str = data.get("Tổng Tiền Đã Đóng (VNĐ)", "0")
    total_received_str = data.get("Tổng Tiền Đã Nhận (VNĐ)", "0")
    total_profit_str = data.get("Tổng Tiền Lãi (VNĐ)", "0")
    profit_rate_str = data.get("Tỷ Suất Lợi Nhuận (%)", "0")
    status = data.get("Trạng Thái (Playing/Defaulted)", "Playing")
    telegram_group = data.get("Nhóm Chat Telegram (tùy chọn)", "")
    note = data.get("Ghi Chú", "")

    if not rosca_code:
        await message.reply_text("⚠️ <b>Mã Dây Hụi</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not user_id:
        await message.reply_text("⚠️ <b>ID Người Chơi</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    def parse_float(val_str):
        if not val_str: return 0.0
        try:
            return float(str(val_str).replace(",", "").replace(".", "").replace(" ", ""))
        except:
            return 0.0

    try:
        parts_count = int(str(parts_count_str).replace(",", "").replace(".", "").replace(" ", ""))
        if parts_count <= 0:
            raise ValueError("Số lượng chân phải lớn hơn 0.")
    except:
        await message.reply_text("⚠️ <b>Số Lượng Chân</b> không hợp lệ.", parse_mode=ParseMode.HTML)
        return

    total_contributed = parse_float(total_contributed_str)
    total_received = parse_float(total_received_str)
    total_profit = parse_float(total_profit_str)
    profit_rate = parse_float(profit_rate_str)

    db = SessionLocal()
    try:
        from app.models.rosca import Rosca, UserRosca, RoscaMember
        
        member = db.query(RoscaMember).filter(RoscaMember.id == member_id).first()
        if not member:
            await message.reply_text(f"⚠️ Không tìm thấy Chân hụi có mã: <b>{member_id}</b>", parse_mode=ParseMode.HTML)
            return

        # Look up Rosca by code
        rosca = db.query(Rosca).filter(Rosca.code == rosca_code).first()
        if not rosca:
            await message.reply_text(f"⚠️ Không tìm thấy Dây hụi có mã: <b>{rosca_code}</b>", parse_mode=ParseMode.HTML)
            return

        # Check if user exists
        player = db.query(UserRosca).filter(UserRosca.id == user_id).first()
        if not player:
            await message.reply_text(f"⚠️ Không tìm thấy Người chơi có ID: <b>{user_id}</b>.", parse_mode=ParseMode.HTML)
            return

        # Check if they are changing to another user or rosca which already has an entry
        if member.rosca_id != rosca.id or member.user_id != user_id:
            existing_member = db.query(RoscaMember).filter(
                RoscaMember.rosca_id == rosca.id,
                RoscaMember.user_id == user_id,
                RoscaMember.id != member_id
            ).first()
            if existing_member:
                await message.reply_text(f"⚠️ Người chơi <b>{player.full_name}</b> đã tham gia Dây hụi này ở một Chân hụi khác (Mã: {existing_member.id}).", parse_mode=ParseMode.HTML)
                return

        # Sum of current parts to check if it exceeds total_parts of Rosca
        if member.rosca_id == rosca.id:
            current_members = db.query(RoscaMember).filter(RoscaMember.rosca_id == rosca.id).all()
            current_total_parts = sum([m.parts_count for m in current_members if m.parts_count and m.id != member_id])
        else:
            current_members = db.query(RoscaMember).filter(RoscaMember.rosca_id == rosca.id).all()
            current_total_parts = sum([m.parts_count for m in current_members if m.parts_count])

        if rosca.total_parts and (current_total_parts + parts_count > rosca.total_parts):
            await message.reply_text(f"⚠️ Không thể cập nhật. Dây hụi này chỉ còn <b>{rosca.total_parts - current_total_parts}</b> chân trống (Tổng số: {rosca.total_parts}).", parse_mode=ParseMode.HTML)
            return

        member.rosca_id = rosca.id
        member.user_id = user_id
        member.parts_count = parts_count
        member.total_contributed = total_contributed
        member.total_received = total_received
        member.total_profit = total_profit
        member.profit_rate = profit_rate
        member.status = status
        member.note = note
        member.telegram_group = telegram_group if telegram_group else None
        
        db.commit()

        await message.reply_text(f"✅ Đã cập nhật thành công thông tin Chân hụi <b>{member_id}</b>.", parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaUpdateMember] Updated member {member_id} in Rosca {rosca_code} by {message.from_user.id}", LogType.SYSTEM_STATUS)
    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_update_member_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình cập nhật thông tin Chân hụi.")
    finally:
        db.close()


# --- Rosca: Delete Rosca Member (Chân Hụi) ---
@bot.on_message(filters.command(["hui_xoa_chan_hui", "roscas_delete_member"]) | filters.regex(r"^@\w+\s+/(hui_xoa_chan_hui|roscas_delete_member)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def rosca_delete_member_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_xoa_chan_hui", "roscas_delete_member"])
    if args is None: return

    if len(args) < 2:
        await message.reply_text("⚠️ Vui lòng cung cấp Mã Chân Hụi cần xóa.\nVí dụ: <code>/hui_xoa_chan_hui AT001</code>", parse_mode=ParseMode.HTML)
        return
        
    member_id = args[1]
    db = SessionLocal()
    try:
        from app.models.rosca import RoscaMember, Rosca, UserRosca
        from bot.utils.utils import fmt_num
        
        member = db.query(RoscaMember).filter(RoscaMember.id == member_id).first()
        if not member:
            await message.reply_text(f"⚠️ Không tìm thấy Chân hụi có mã: <b>{member_id}</b>", parse_mode=ParseMode.HTML)
            return
            
        if member.status == "Deleted":
            await message.reply_text(f"⚠️ Chân hụi <b>{member_id}</b> đã bị xóa trước đó.", parse_mode=ParseMode.HTML)
            return

        rosca = db.query(Rosca).filter(Rosca.id == member.rosca_id).first()
        player = db.query(UserRosca).filter(UserRosca.id == member.user_id).first()
        
        rosca_code = rosca.code if rosca else "N/A"
        player_name = player.full_name if player else "N/A"

        info_text = f"""<b>⚠️ XÁC NHẬN XÓA CHÂN HỤI</b>

Bạn có chắc chắn muốn xóa Chân hụi này không?
- <b>Mã Chân Hụi:</b> {member.id}
- <b>Mã Dây Hụi:</b> {rosca_code}
- <b>Người Chơi:</b> {player_name}
- <b>Số Lượng Chân:</b> {member.parts_count or 1}
- <b>Tổng Tiền Đã Đóng:</b> {fmt_num(member.total_contributed)} đ

<i>Lưu ý: Thao tác này sẽ chuyển trạng thái của chân hụi sang Deleted và không xóa vĩnh viễn khỏi hệ thống.</i>"""

        buttons = [
            [
                InlineKeyboardButton("Xác nhận", callback_data=f"rosca_delmem_confirm_{member.id}"),
                InlineKeyboardButton("Hủy", callback_data="rosca_delmem_cancel")
            ]
        ]
        
        await message.reply_text(info_text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
    except Exception as e:
        LogError(f"Error checking member for deletion: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình truy xuất thông tin.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rosca_delmem_confirm_(.+)$"))
async def rosca_delmember_confirm_callback(client, callback_query: CallbackQuery):
    member_id = callback_query.matches[0].group(1)
    
    db = SessionLocal()
    try:
        from app.models.rosca import RoscaMember
        member = db.query(RoscaMember).filter(RoscaMember.id == member_id).first()
        
        if not member:
            await callback_query.answer("⚠️ Chân hụi không tồn tại hoặc đã bị xóa.", show_alert=True)
            return

        if member.status == "Deleted":
            await callback_query.answer("⚠️ Chân hụi này đã bị xóa trước đó.", show_alert=True)
            return

        member.status = "Deleted"
        db.commit()

        await callback_query.message.edit_text(f"✅ Đã xóa thành công Chân hụi <b>{member_id}</b> (chuyển trạng thái sang Deleted).", parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaDeleteMember] Member {member_id} was deleted by {callback_query.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_delmember_confirm callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi xác nhận xóa.", show_alert=True)
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^rosca_delmem_cancel$"))
async def rosca_delete_member_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("❌ Đã hủy thao tác xóa Chân hụi.")

# --- Rosca: Check Member (Kiểm Tra Chân Hụi) ---
@bot.on_message(filters.command(["hui_kiem_tra_chan_hui", "roscas_check_member"]) | filters.regex(r"^@\w+\s+/(hui_kiem_tra_chan_hui|roscas_check_member)\b"))
@require_group_role("main", "member")
async def rosca_check_member_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_kiem_tra_chan_hui", "roscas_check_member"])
    if args is None: return

    if len(args) < 2:
        await message.reply_text("⚠️ Vui lòng cung cấp Mã Chân Hụi cần kiểm tra.\nVí dụ: <code>/hui_kiem_tra_chan_hui AT2506_01</code>", parse_mode=ParseMode.HTML)
        return
        
    member_id = args[1].strip()
    db = SessionLocal()
    try:
        from app.models.rosca import RoscaMember, Rosca, UserRosca, RoscaContribution
        from bot.utils.utils import fmt_num
        
        member = db.query(RoscaMember).filter(RoscaMember.id == member_id).first()
        if not member:
            await message.reply_text(f"⚠️ Không tìm thấy Chân hụi có mã: <b>{member_id}</b>", parse_mode=ParseMode.HTML)
            return
                
        rosca = db.query(Rosca).filter(Rosca.id == member.rosca_id).first()
        player = db.query(UserRosca).filter(UserRosca.id == member.user_id).first()
        
        # Đếm số kỳ đã đóng và tổng tiền đã đóng thực tế (dựa vào bảng RoscaContribution)
        from sqlalchemy.sql import func
        contrib_stats = db.query(
            func.count(RoscaContribution.id)
        ).filter(
            RoscaContribution.member_id == member.id,
            RoscaContribution.status == "Paid"
        ).first()
        
        paid_rounds = contrib_stats[0] or 0
        
        # Chỉ tính tổng các khoản đóng thực tế (mang dấu âm)
        total_paid_query = db.query(
            func.sum(RoscaContribution.amount)
        ).filter(
            RoscaContribution.member_id == member.id,
            RoscaContribution.status == "Paid",
            RoscaContribution.amount < 0
        ).first()
        actual_total_contributed = total_paid_query[0] or 0
        
        rosca_code = rosca.code if rosca else "N/A"
        player_name = player.full_name if player else "N/A"
        
        suggested_profit_text = ""
        if rosca and rosca.total_parts and member.status not in ["Dead", "Withdrawn", "Completed"]:
            max_payment = rosca.base_amount if rosca.base_amount and rosca.base_amount > (rosca.max_bid_amount or 0) else (rosca.max_bid_amount or 0)
            remaining_rounds = rosca.total_parts - paid_rounds - 1
            if remaining_rounds > 0 and max_payment > 0:
                min_receive = abs(actual_total_contributed) + (remaining_rounds * max_payment)
                suggested_profit_text = f"\n\n💡 <i>Gợi ý: Mức hốt tối thiểu để có lời: {fmt_num(min_receive)} đ (Không bao gồm Tiền thảo)</i>"
        
        text = f"""<b>THÔNG TIN CHÂN HỤI</b>
- <b>Mã Chân Hụi:</b> <code>{member.id}</code>
- <b>Người Chơi:</b> {player_name} (ID: <code>{member.user_id}</code>)
- <b>Thuộc Dây Hụi:</b> <b>{rosca_code}</b>
- <b>Số Lượng Chân:</b> {member.parts_count or 1}
- <b>Số Kỳ Đã Đóng:</b> {paid_rounds}
- <b>Tổng Tiền Đã Đóng:</b> {fmt_num(actual_total_contributed)} đ
- <b>Tổng Tiền Đã Nhận:</b> {fmt_num(member.total_received)} đ
- <b>Tổng Lợi Nhuận:</b> {fmt_num(member.total_profit)} đ
- <b>Tỷ Suất LN:</b> {member.profit_rate or 0.0}%
- <b>Trạng Thái:</b> {member.status or 'Playing'}
- <b>Nhóm Nhận TB:</b> {member.telegram_group or 'Không có'}
- <b>Ghi Chú:</b> {member.note or 'Không'}""" + suggested_profit_text

        await message.reply_text(text, parse_mode=ParseMode.HTML)
    except Exception as e:
        LogError(f"Error checking rosca member: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình kiểm tra.")
    finally:
        db.close()



# --- Rosca: Check Contributions History (Kiểm Tra Đóng Hụi) ---
@bot.on_message(filters.command(["hui_kiem_tra_dong_hui", "roscas_check_contributions"]) | filters.regex(r"^@\w+\s+/(hui_kiem_tra_dong_hui|roscas_check_contributions)\b"))
@require_group_role("main", "member")
async def rosca_check_contributions_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_kiem_tra_dong_hui", "roscas_check_contributions"])
    if args is None: return

    if len(args) < 2:
        await message.reply_text("⚠️ Vui lòng cung cấp Mã Chân Hụi cần kiểm tra.\nVí dụ: <code>/hui_kiem_tra_dong_hui AT001</code>", parse_mode=ParseMode.HTML)
        return

    member_id = args[1].strip()
    db = SessionLocal()
    try:
        from app.models.rosca import RoscaMember, Rosca, UserRosca, RoscaContribution
        from bot.utils.utils import fmt_num, fmt_vn

        member = db.query(RoscaMember).filter(RoscaMember.id == member_id).first()
        if not member:
            await message.reply_text(f"⚠️ Không tìm thấy Chân hụi có mã: <b>{member_id}</b>", parse_mode=ParseMode.HTML)
            return

        rosca = db.query(Rosca).filter(Rosca.id == member.rosca_id).first()
        player = db.query(UserRosca).filter(UserRosca.id == member.user_id).first()

        rosca_code = rosca.code if rosca else "N/A"
        player_name = player.full_name if player else "N/A"

        # Lấy tất cả lịch sử đóng hụi, sắp xếp theo kỳ rồi theo ngày
        contributions = db.query(RoscaContribution).filter(
            RoscaContribution.member_id == member.id
        ).order_by(
            RoscaContribution.round_number.asc(),
            RoscaContribution.actual_payment_date.asc()
        ).all()

        if not contributions:
            await message.reply_text(
                f"<b>LỊCH SỬ ĐÓNG HỤI</b>\n"
                f"- Chân hụi: <b>{member.id}</b>\n"
                f"- Người chơi: <b>{player_name}</b>\n"
                f"- Dây hụi: <b>{rosca_code}</b>\n\n"
                f"<i>Chưa có lịch sử đóng hụi nào.</i>",
                parse_mode=ParseMode.HTML
            )
            return

        total_paid = 0
        total_unpaid = 0
        count_paid = 0
        count_unpaid = 0
        last_paid_contrib = None

        # Loop over contributions to compute metrics first
        for idx, c in enumerate(contributions, start=1):
            if c.status == "Paid":
                count_paid += 1
                if c.amount and c.amount < 0:
                    total_paid += abs(c.amount)
                    last_paid_contrib = c
            elif c.status == "Late":
                count_paid += 1
                if c.amount and c.amount < 0:
                    total_paid += abs(c.amount)
                    last_paid_contrib = c
            else:
                count_unpaid += 1
                if c.amount and c.amount < 0:
                    total_unpaid += abs(c.amount)

        # Simulation calculations
        sim_text = ""
        base_amount = rosca.base_amount or 0.0
        total_parts = rosca.total_parts or 0
        commission_fee = rosca.commission_fee or 0.0
        total_parts_excluding_me = max(0, total_parts - 1)

        if last_paid_contrib and member.status not in ["Dead", "Withdrawn", "Completed"]:
            last_paid_amount = last_paid_contrib.amount
            upcoming_bid = base_amount + last_paid_amount
            upcoming_bid = max(0.0, upcoming_bid)
            
            # Tính Tổng hốt hụi giả
            dead_rounds_count = count_paid
            living_rounds_count = max(0, total_parts_excluding_me - dead_rounds_count)
            total_withdrawn_amount = (dead_rounds_count * base_amount) + ((base_amount - upcoming_bid) * living_rounds_count) - commission_fee
            
            # Tính Mức hốt tối thiểu để có lời
            max_payment = base_amount
            if rosca.max_bid_amount is not None and rosca.max_bid_amount > base_amount:
                max_payment = rosca.max_bid_amount
            remaining_rounds = total_parts - dead_rounds_count - 1
            min_receive = total_paid + (max_payment * max(0, remaining_rounds))
            
            simulated_profit = total_withdrawn_amount - min_receive
            
            sim_text = (
                f"{'━' * 20}\n"
                f"- <b>Lợi nhuận giả lập:</b> <b>{fmt_vn(simulated_profit)}</b>\n"
                f"  <i>(Tiền bỏ hụi giả định: {fmt_vn(upcoming_bid)})</i>\n"
                f"  <i>(Tổng hốt hụi giả định: {fmt_vn(total_withdrawn_amount)})</i>\n"
                f"  <i>(Mức hốt tối thiểu để có lời: {fmt_vn(min_receive)})</i>\n"
            )
        else:
            sim_text = (
                f"{'━' * 20}\n"
                f"- <b>Lợi nhuận giả lập:</b> N/A <i>(Chưa có kỳ đóng hụi nào được ghi nhận hoặc chân hụi đã hốt)</i>\n"
            )

        start_date_str = rosca.start_date.strftime("%d/%m/%Y") if rosca.start_date else "N/A"
        end_date_str = rosca.end_date.strftime("%d/%m/%Y") if rosca.end_date else "N/A"

        # Build header
        text = (
            f"<b>LỊCH SỬ ĐÓNG HỤI</b>\n"
            f"- Chân hụi: <code>{member.id}</code>\n"
            f"- Người chơi: <b>{player_name}</b>\n"
            f"- Dây hụi: <b>{rosca_code}</b>\n"
            f"- Tiền hụi tháng: <b>{fmt_vn(base_amount)}</b>\n"
            f"- Số chân hụi: <b>{total_parts} chân</b>\n"
            f"- Tiền thảo: <b>{fmt_vn(commission_fee)}</b>\n"
            f"- Ngày bắt đầu: <b>{start_date_str}</b>\n"
            f"- Ngày kết thúc: <b>{end_date_str}</b>\n"
            f"{sim_text}"
            f"{'━' * 20}\n\n"
        )

        for idx, c in enumerate(contributions, start=1):
            # Determine status emoji
            status_icon = ""
            round_label = f"Kỳ {c.round_number}" if c.round_number else f"#{idx}"
            date_str = c.actual_payment_date.strftime("%d/%m/%Y %H:%M") if c.actual_payment_date else "Chưa đóng"
            amount_str = fmt_vn(c.amount) if c.amount else "0 VNĐ"
            note_str = f" — {c.note}" if c.note else ""

            text += (
                f"{status_icon} <b>{round_label}</b>\n"
                f"{date_str}  |  {amount_str}  |  {c.status or 'N/A'}{note_str}\n"
            )

        # Summary
        text += (
            f"\n{'━' * 20}\n"
            f"<b>TỔNG KẾT:</b>\n"
            f"- Số kỳ đã đóng: <b>{count_paid}</b>\n"
            f"- Số kỳ chưa đóng: <b>{count_unpaid}</b>\n"
            f"- Tổng tiền đã đóng: <b>{fmt_vn(total_paid)}</b>\n"
        )
        if total_unpaid > 0:
            text += f"- Tổng tiền chưa đóng: <b>{fmt_vn(total_unpaid)}</b>\n"

        # Split message if too long (Telegram limit 4096 chars)
        if len(text) > 4096:
            parts = []
            current = ""
            for line in text.split("\n"):
                if len(current) + len(line) + 1 > 4000:
                    parts.append(current)
                    current = line + "\n"
                else:
                    current += line + "\n"
            if current:
                parts.append(current)

            for i, part in enumerate(parts):
                if i == 0:
                    await message.reply_text(part, parse_mode=ParseMode.HTML)
                else:
                    await message.reply_text(part, parse_mode=ParseMode.HTML)
        else:
            await message.reply_text(text, parse_mode=ParseMode.HTML)

    except Exception as e:
        LogError(f"Error checking rosca contributions: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình kiểm tra lịch sử đóng hụi.")
    finally:
        db.close()


# --- Rosca: Pay Contribution (Đóng Tiền Chân Hụi) ---
@bot.on_message(filters.command(["hui_dong_tien_chan_hui", "roscas_pay_contribution"]) | filters.regex(r"^@\w+\s+/(hui_dong_tien_chan_hui|roscas_pay_contribution)\b"))
@require_group_role("member", "main")
async def rosca_pay_contribution_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_dong_tien_chan_hui", "roscas_pay_contribution"])
    if args is None: return

    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        db = SessionLocal()
        try:
            from app.models.rosca import Rosca, UserRosca
            active_roscas = db.query(Rosca).filter(Rosca.status == "Active").all()
            if not active_roscas:
                await message.reply_text("⚠️ Hiện tại chưa có Dây hụi nào đang hoạt động.", parse_mode=ParseMode.HTML)
                return
            
            buttons = []
            for r in active_roscas:
                owner = db.query(UserRosca).filter(UserRosca.id == r.user_id).first()
                owner_name = owner.full_name if owner else "Chưa rõ"
                btn_text = f"Mã: {r.code} | {owner_name} ({r.total_parts or '?'} chân)"
                cb_data = f"cb_rosca_pay_{r.code}"
                buttons.append([InlineKeyboardButton(btn_text, callback_data=cb_data)])
            
            buttons.append([InlineKeyboardButton("Hủy", callback_data="cb_rosca_pay_cancel")])
            
            await message.reply_text("<b>CHỌN DÂY HỤI ĐỂ ĐÓNG TIỀN</b>\nVui lòng chọn Dây hụi:", reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error fetching roscas for pay contribution: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Có lỗi xảy ra trong quá trình lấy danh sách Dây hụi.")
        finally:
            db.close()
        return

    # Parse Form
    data = {}
    for line in lines[1:]:
        if "):" in line:
            key, val = line.split("):", 1)
            key += ")"
        elif ":" in line:
            key, val = line.split(":", 1)
        else:
            continue
        data[key.strip()] = val.strip()

    rosca_code = data.get("Mã Dây Hụi", "")
    round_id = data.get("Kỳ Khui Hụi (round_id) (Lưu ý có thể để trống)", "")
    round_number_str = data.get("Kỳ Thu Hụi (Chỉ ghi số, vd: 1, 2...)", "")
    member_id = data.get("Mã Chân Hụi", "")
    amount_str = data.get("Số Tiền Đóng (VNĐ)", "")
    payment_date_str = data.get("Ngày Giờ Đóng Tiền (DD/MM/YYYY HH:MM)", "")
    status = data.get("Trạng Thái (Unpaid/Paid/Late)", "Paid").capitalize()
    note = data.get("Ghi Chú", "")

    if not rosca_code:
        await message.reply_text("⚠️ <b>Mã Dây Hụi</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not member_id:
        await message.reply_text("⚠️ <b>Mã Chân Hụi</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not amount_str:
        await message.reply_text("⚠️ <b>Số Tiền Đóng</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    def parse_float(val_str):
        if not val_str: return 0.0
        try:
            cleaned = val_str.replace("–", "-").replace("—", "-").replace(",", "").replace(".", "").replace(" ", "")
            return float(cleaned)
        except:
            return 0.0

    from datetime import datetime
    def parse_datetime(val_str):
        if not val_str: return None
        try:
            return datetime.strptime(val_str, "%d/%m/%Y %H:%M")
        except:
            return None

    def parse_int(val_str):
        if not val_str: return None
        try:
            return int(val_str.replace(",", "").replace(".", "").replace(" ", ""))
        except:
            return None

    round_number = parse_int(round_number_str)
    amount = parse_float(amount_str)
    if amount == 0:
        await message.reply_text("⚠️ <b>Số Tiền Đóng</b> không hợp lệ (phải khác 0).", parse_mode=ParseMode.HTML)
        return

    payment_date = parse_datetime(payment_date_str)
    if not payment_date:
        payment_date = datetime.now()

    db = SessionLocal()
    try:
        from app.models.rosca import Rosca, RoscaMember, UserRosca
        
        rosca = db.query(Rosca).filter(Rosca.code == rosca_code).first()
        if not rosca:
            await message.reply_text(f"⚠️ Không tìm thấy Dây hụi có mã: <b>{rosca_code}</b>", parse_mode=ParseMode.HTML)
            return

        member = db.query(RoscaMember).filter(RoscaMember.id == member_id).first()
        if not member:
            # Fallback: Check if they accidentally put user_id instead of member_id
            member = db.query(RoscaMember).filter(
                RoscaMember.user_id == member_id, 
                RoscaMember.rosca_id == rosca.id
            ).first()

        if not member:
            await message.reply_text(f"⚠️ Không tìm thấy Chân hụi hoặc Người chơi có mã: <b>{member_id}</b> trong Dây hụi này.", parse_mode=ParseMode.HTML)
            return

        is_dead = member.status == "Dead"

        if not is_dead:
            if amount < 0:
                base_amt = rosca.base_amount or 0.0
                check_val = amount + base_amt
                
                min_bid = rosca.min_bid_amount
                max_bid = rosca.max_bid_amount if rosca.max_bid_amount is not None else rosca.base_amount
                
                if min_bid is not None and check_val < min_bid:
                    await message.reply_text(
                        f"⚠️ <b>Số Tiền Đóng + Số Tiền Gốc</b> ({check_val:,.0f} đ) đang nhỏ hơn Mức Bỏ Hụi Tối Thiểu ({min_bid:,.0f} đ).", 
                        parse_mode=ParseMode.HTML
                    )
                    return
                    
                if max_bid is not None and check_val > max_bid:
                    await message.reply_text(
                        f"⚠️ <b>Số Tiền Đóng + Số Tiền Gốc</b> ({check_val:,.0f} đ) đang lớn hơn Mức Bỏ Hụi Tối Đa ({max_bid:,.0f} đ).", 
                        parse_mode=ParseMode.HTML
                    )
                    return
            else:
                if rosca.min_bid_amount is not None and amount < rosca.min_bid_amount:
                    await message.reply_text(f"⚠️ <b>Số Tiền Đóng</b> ({amount:,.0f} đ) đang nhỏ hơn mức tối thiểu quy định ({rosca.min_bid_amount:,.0f} đ).", parse_mode=ParseMode.HTML)
                    return

                max_allowed = rosca.base_amount if rosca.base_amount and rosca.base_amount > (rosca.max_bid_amount or 0) else rosca.max_bid_amount
                if max_allowed is not None and amount > max_allowed:
                    await message.reply_text(f"⚠️ <b>Số Tiền Đóng</b> ({amount:,.0f} đ) không được vượt quá mức tối đa ({max_allowed:,.0f} đ).", parse_mode=ParseMode.HTML)
                    return

        player = db.query(UserRosca).filter(UserRosca.id == member.user_id).first()
        player_name = player.full_name if player else "N/A"

        cache_key = str(uuid.uuid4())[:8]
        
        if not hasattr(bot, 'temp_rosca_contributions'):
            bot.temp_rosca_contributions = {}
            
        bot.temp_rosca_contributions[cache_key] = {
            "rosca_id": rosca.id,
            "rosca_code": rosca_code,
            "round_id": round_id if round_id else None,
            "round_number": round_number,
            "member_id": member.id,
            "amount": amount,
            "payment_date": payment_date,
            "status": status,
            "note": note
        }

        info_text = f"""<b>⚠️ XÁC NHẬN ĐÓNG TIỀN CHÂN HỤI</b>

Bạn có chắc chắn muốn ghi nhận giao dịch đóng hụi này không?
- <b>Mã Dây Hụi:</b> {rosca_code}
- <b>Kỳ Khui Hụi:</b> {round_id or 'N/A'}
- <b>Kỳ Thu Hụi:</b> Kỳ {round_number}
- <b>Mã Chân Hụi:</b> {member.id}
- <b>Người Chơi:</b> {player_name}
- <b>Số Tiền Đóng:</b> {amount:,.0f} đ
- <b>Ngày Đóng:</b> {payment_date.strftime('%d/%m/%Y %H:%M')}
- <b>Trạng Thái:</b> {status}
- <b>Ghi Chú:</b> {note or 'N/A'}"""

        buttons = [
            [
                InlineKeyboardButton("Xác nhận", callback_data=f"cb_rosca_pay_confirm_{cache_key}"),
                InlineKeyboardButton("Hủy", callback_data="cb_rosca_pay_cancel")
            ]
        ]
        
        await message.reply_text(info_text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)

    except Exception as e:
        LogError(f"Error preparing rosca contribution confirmation: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình kiểm tra thông tin.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^cb_rosca_pay_confirm_(.+)$"))
async def rosca_pay_contribution_confirm_callback(client, callback_query: CallbackQuery):
    cache_key = callback_query.matches[0].group(1)
    
    if not hasattr(bot, 'temp_rosca_contributions') or cache_key not in bot.temp_rosca_contributions:
        await callback_query.message.edit_text("⚠️ Thông tin giao dịch đã hết hạn hoặc không tồn tại. Vui lòng thực hiện lại.")
        return
        
    data = bot.temp_rosca_contributions[cache_key]
    
    db = SessionLocal()
    try:
        from app.models.rosca import RoscaMember, RoscaContribution
        
        payment_date = data["payment_date"]
        start_of_day = payment_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = payment_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        existing_payment = db.query(RoscaContribution).filter(
            RoscaContribution.member_id == data["member_id"],
            RoscaContribution.rosca_id == data["rosca_id"],
            RoscaContribution.actual_payment_date >= start_of_day,
            RoscaContribution.actual_payment_date <= end_of_day
        ).first()

        if existing_payment:
            await callback_query.message.edit_text(f"⚠️ Chân hụi <b>{data['member_id']}</b> đã được ghi nhận đóng tiền trong ngày {payment_date.strftime('%d/%m/%Y')} rồi!\nThao tác bị hủy để tránh trùng lặp.", parse_mode=ParseMode.HTML)
            del bot.temp_rosca_contributions[cache_key]
            return

        member = db.query(RoscaMember).filter(RoscaMember.id == data["member_id"]).first()
        if not member:
            await callback_query.answer("⚠️ Không tìm thấy Chân hụi.", show_alert=True)
            return

        # Update member's total_contributed
        if data["status"] == "Paid":
            member.total_contributed = (member.total_contributed or 0) + data["amount"]

        new_contribution = RoscaContribution(
            id=str(uuid.uuid4()),
            rosca_id=data["rosca_id"],
            round_id=data["round_id"],
            round_number=data["round_number"],
            member_id=data["member_id"],
            amount=data["amount"],
            actual_payment_date=data["payment_date"],
            status=data["status"],
            note=data["note"]
        )
        db.add(new_contribution)
        db.commit()

        round_info = f"\n- Kỳ thu hụi: <b>Kỳ {data['round_number']}</b>" if data.get('round_number') else ""
        msg_text = f"✅ Đã ghi nhận đóng hụi thành công!\n- Mã GD: <code>{new_contribution.id}</code>\n- Dây hụi: <b>{data['rosca_code']}</b>{round_info}\n- Chân hụi: <b>{data['member_id']}</b>\n- Số tiền: <b>{data['amount']:,.0f} VNĐ</b>\n\n<i>Lưu ý: Bạn có thể Reply tin nhắn này kèm lệnh /hui_huy_dong_tien để hủy giao dịch.</i>"
        await callback_query.message.edit_text(msg_text, parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaContribution] Member {data['member_id']} paid {data['amount']} in rosca {data['rosca_code']} (Round {data.get('round_number')}) by {callback_query.from_user.id}", LogType.SYSTEM_STATUS)
        
        del bot.temp_rosca_contributions[cache_key]
    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_pay_contribution_confirm_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra trong quá trình lưu giao dịch.", show_alert=True)
    finally:
        db.close()



@bot.on_callback_query(filters.regex(r"^cb_rosca_pay_(.+)$"))
async def rosca_pay_contribution_callback(client, callback_query: CallbackQuery):
    action = callback_query.matches[0].group(1)
    
    if action == "cancel":
        await callback_query.message.edit_text("❌ Đã hủy thao tác đóng tiền hụi.")
        return

    rosca_code = action
    from datetime import datetime
    current_time = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    form_template = f"""<b>FORM ĐÓNG TIỀN CHÂN HỤI</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/hui_dong_tien_chan_hui
Mã Dây Hụi: {rosca_code}
Kỳ Khui Hụi (round_id) (Lưu ý có thể để trống): 
Kỳ Thu Hụi (Chỉ ghi số, vd: 1, 2...): 
Mã Chân Hụi: 
Số Tiền Đóng (VNĐ): 
Ngày Giờ Đóng Tiền (DD/MM/YYYY HH:MM): {current_time}
Trạng Thái (Unpaid/Paid/Late): Paid
Ghi Chú: 
</pre>
"""
    await callback_query.message.edit_text(form_template, parse_mode=ParseMode.HTML)


@bot.on_callback_query(filters.regex(r"^rosca_delmem_cancel$"))
async def rosca_delmember_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("❌ Đã hủy thao tác xóa Chân hụi.")

@bot.on_message(filters.command(["hui_huy_dong_tien"]) & filters.reply)
@require_project_name("Hụi")
@require_group_role("member", "main")
async def rosca_cancel_contribution_handler(client, message: Message) -> None:
    replied_msg = message.reply_to_message
    if not replied_msg or not replied_msg.text:
        return
        
    text = replied_msg.text
    if "Đã ghi nhận đóng hụi thành công!" not in text:
        return
        
    import re
    match = re.search(r"- Mã GD:\s*([a-zA-Z0-9\-]+)", text)
    if not match:
        await message.reply_text("⚠️ Không tìm thấy Mã Giao Dịch trong tin nhắn này.")
        return
        
    tx_id = match.group(1)
    
    db = SessionLocal()
    try:
        from app.models.rosca import RoscaContribution, RoscaMember
        
        contribution = db.query(RoscaContribution).filter(RoscaContribution.id == tx_id).first()
        if not contribution:
            await message.reply_text(f"⚠️ Giao dịch <b>{tx_id}</b> không tồn tại hoặc đã bị hủy trước đó.", parse_mode=ParseMode.HTML)
            db.close()
            from pyrogram import StopPropagation
            raise StopPropagation
            
        member = db.query(RoscaMember).filter(RoscaMember.id == contribution.member_id).first()
        
        if contribution.status == "Paid" and member:
            member.total_contributed = (member.total_contributed or 0) - contribution.amount
            if member.total_contributed < 0:
                member.total_contributed = 0
                
        db.delete(contribution)
        db.commit()
        
        await message.reply_text(f"✅ Đã hủy giao dịch đóng hụi <b>{tx_id}</b> thành công và trừ lại số tiền đã cộng.", parse_mode=ParseMode.HTML)
        
        new_text = text.replace("✅ Đã ghi nhận đóng hụi thành công!", "❌ [ĐÃ HỦY] Giao dịch đóng hụi")
        await replied_msg.edit_text(new_text, parse_mode=ParseMode.HTML)
        
    except Exception as e:
        db.rollback()
        LogError(f"Error cancelling rosca contribution: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình hủy giao dịch.")
    finally:
        db.close()


# --- Rosca: Rút Dây Hụi (Hốt Hụi) ---
@bot.on_message(filters.command(["hui_rut_day_hui", "roscas_withdraw"]) | filters.regex(r"^@\w+\s+/(hui_rut_day_hui|roscas_withdraw)\b"))
@require_group_role("main", "member")
async def rosca_withdraw_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_rut_day_hui", "roscas_withdraw"])
    if args is None: return

    if len(args) < 3:
        await message.reply_text("⚠️ Vui lòng cung cấp Mã Chân Hụi và Số Tiền.\nVí dụ: <code>/hui_rut_day_hui AT001 50000000</code>", parse_mode=ParseMode.HTML)
        return

    member_id = args[1].strip()
    try:
        amount = float(args[2].replace(",", "").replace(".", "").replace(" ", ""))
    except ValueError:
        await message.reply_text("⚠️ Số tiền không hợp lệ. Vui lòng nhập số.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        from app.models.rosca import RoscaMember, Rosca, UserRosca, RoscaContribution
        from bot.utils.utils import fmt_vn
        from sqlalchemy.sql import func
        
        member = db.query(RoscaMember).filter(RoscaMember.id == member_id).first()
        if not member:
            await message.reply_text(f"⚠️ Không tìm thấy Chân hụi có mã: <b>{member_id}</b>", parse_mode=ParseMode.HTML)
            return

        rosca = db.query(Rosca).filter(Rosca.id == member.rosca_id).first()
        player = db.query(UserRosca).filter(UserRosca.id == member.user_id).first()
        
        if not rosca or not player:
            await message.reply_text("⚠️ Lỗi dữ liệu: Không tìm thấy dây hụi hoặc người chơi tương ứng.", parse_mode=ParseMode.HTML)
            return
            
        if member.status in ["Dead", "Withdrawn", "Completed"]:
            await message.reply_text(f"⚠️ Chân hụi này đã ở trạng thái <b>{member.status}</b> (đã hốt hoặc đã rút)!", parse_mode=ParseMode.HTML)
            return

        # Tính min_receive
        contrib_stats = db.query(
            func.count(RoscaContribution.id)
        ).filter(
            RoscaContribution.member_id == member.id,
            RoscaContribution.status == "Paid"
        ).first()
        
        paid_rounds = contrib_stats[0] or 0
        
        # Chỉ tính tổng các khoản đóng thực tế (mang dấu âm)
        total_paid_query = db.query(
            func.sum(RoscaContribution.amount)
        ).filter(
            RoscaContribution.member_id == member.id,
            RoscaContribution.status == "Paid",
            RoscaContribution.amount < 0
        ).first()
        actual_total_contributed = total_paid_query[0] or 0
        
        max_payment = rosca.base_amount if rosca.base_amount and rosca.base_amount > (rosca.max_bid_amount or 0) else (rosca.max_bid_amount or 0)
        remaining_rounds = (rosca.total_parts or 0) - paid_rounds - 1
        
        commission_fee = rosca.commission_fee or 0
        actual_receive = amount
        min_receive = abs(actual_total_contributed) + (remaining_rounds * max_payment)
        profit = actual_receive - min_receive
        
        cache_key = str(uuid.uuid4())[:8]
        if not hasattr(bot, 'temp_rosca_withdraws'):
            bot.temp_rosca_withdraws = {}
            
        bot.temp_rosca_withdraws[cache_key] = {
            "member_id": member.id,
            "amount": actual_receive,
            "actual_receive": actual_receive,
            "profit": profit,
            "rosca_id": rosca.id,
            "next_round_number": paid_rounds + 1
        }

        profit_status = "LỜI" if profit >= 0 else "LỖ"

        info_text = f"""<b>⚠️ XÁC NHẬN RÚT/HỐT HỤI</b>

<b>Thông tin chi tiết:</b>
- <b>Mã Chân Hụi:</b> <code>{member.id}</code>
- <b>Dây Hụi:</b> {rosca.code}
- <b>Người Chơi:</b> {player.full_name}
- <b>Số kỳ đã đóng:</b> {paid_rounds} kỳ
- <b>Số tiền đã đóng:</b> {fmt_vn(actual_total_contributed)}
- <b>Số tiền hốt (Đã trừ thảo):</b> {fmt_vn(actual_receive)}
- <b>Tiền thảo:</b> {fmt_vn(commission_fee)}

💡 <b>Mức hốt tối thiểu để có lời:</b> {fmt_vn(min_receive)} (Không bao gồm Tiền thảo)
👉 <i>Kết quả nếu hốt với giá này:</i> Bạn sẽ <b>{profit_status}</b> {fmt_vn(abs(profit))}.

Bạn có chắc chắn muốn xác nhận Rút/Hốt hụi và chuyển trạng thái Chân hụi sang "Dead"?
"""

        buttons = [
            [
                InlineKeyboardButton("Xác nhận", callback_data=f"cb_rw_confirm_{cache_key}"),
                InlineKeyboardButton("Hủy", callback_data="cb_rw_cancel")
            ]
        ]
        
        await message.reply_text(info_text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)

    except Exception as e:
        LogError(f"Error checking rosca withdraw: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình kiểm tra thông tin.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^cb_rw_confirm_(.+)$"))
async def rosca_withdraw_confirm_callback(client, callback_query: CallbackQuery):
    cache_key = callback_query.matches[0].group(1)
    
    if not hasattr(bot, 'temp_rosca_withdraws') or cache_key not in bot.temp_rosca_withdraws:
        await callback_query.message.edit_text("⚠️ Thông tin giao dịch đã hết hạn hoặc không tồn tại. Vui lòng thực hiện lại.")
        return
        
    data = bot.temp_rosca_withdraws[cache_key]
    member_id = data["member_id"]
    actual_receive = data["actual_receive"]
    profit = data["profit"]
    rosca_id = data["rosca_id"]
    next_round_number = data["next_round_number"]
    amount = data["amount"]
    
    db = SessionLocal()
    try:
        from app.models.rosca import RoscaMember, RoscaContribution
        from bot.utils.utils import fmt_vn
        from datetime import datetime
        
        member = db.query(RoscaMember).filter(RoscaMember.id == member_id).first()
        if not member:
            await callback_query.answer("⚠️ Không tìm thấy Chân hụi.", show_alert=True)
            return

        member.total_received = (member.total_received or 0) + actual_receive
        member.total_profit = (member.total_profit or 0) + profit
        member.status = "Dead"
        
        # Lưu giao dịch rút/hốt hụi với giá trị dương vào rosca_contributions (đã trừ tiền thảo)
        new_contribution = RoscaContribution(
            id=str(uuid.uuid4()),
            rosca_id=rosca_id,
            round_number=next_round_number,
            member_id=member_id,
            amount=amount,
            actual_payment_date=datetime.now(),
            status="Paid",
            note=f"Hốt hụi kỳ {next_round_number}"
        )
        db.add(new_contribution)
        db.commit()

        msg_text = f"✅ Đã xác nhận Rút/Hốt hụi thành công!\n- Chân hụi: <b>{member_id}</b>\n- Tiền nhận được (sau trừ thảo): <b>{fmt_vn(actual_receive)}</b>\n- Lợi nhuận: <b>{fmt_vn(profit)}</b>\n- Trạng thái: <b>Dead</b>\n- Ghi nhận giao dịch hốt hụi: <b>+{fmt_vn(amount)}</b> (Kỳ {next_round_number})"
        await callback_query.message.edit_text(msg_text, parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaWithdraw] Member {member_id} withdrew. Actual receive: {actual_receive}. Profit: {profit}. By {callback_query.from_user.id}", LogType.SYSTEM_STATUS)
        
        del bot.temp_rosca_withdraws[cache_key]
    except Exception as e:
        db.rollback()
        LogError(f"Error in rosca_withdraw_confirm_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra trong quá trình lưu giao dịch.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^cb_rw_cancel$"))
async def rosca_withdraw_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("❌ Đã hủy thao tác rút/hốt hụi.")



# --- Rosca: Export Statistics to Excel (Thống Kê Hụi) ---
@bot.on_message(filters.command(["hui_thong_ke_hui", "roscas_export_stats"]) | filters.regex(r"^@\w+\s+/(hui_thong_ke_hui|roscas_export_stats)\b"))
@require_group_role("main", "member")
async def roscas_export_stats_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_thong_ke_hui", "roscas_export_stats"])
    if args is None: return

    player_id = args[1].strip() if len(args) >= 2 else None
    db = SessionLocal()
    try:
        from app.models.rosca import UserRosca, RoscaMember, Rosca, RoscaContribution
        import io
        import os
        import openpyxl
        from openpyxl.utils import get_column_letter
        from copy import copy
        import datetime
        from bot.utils.logger import LogInfo, LogError, LogType

        # 1. Fetch user(s)
        if player_id:
            user = db.query(UserRosca).filter(
                (UserRosca.id == player_id) | (UserRosca.username == player_id)
            ).first()
            if not user:
                await message.reply_text(f"⚠️ Không tìm thấy người chơi có mã hoặc username: <b>{player_id}</b>", parse_mode=ParseMode.HTML)
                return
            users = [user]
        else:
            # Get all users who have at least one RoscaMember position
            users = db.query(UserRosca).filter(
                UserRosca.id.in_(db.query(RoscaMember.user_id))
            ).all()
            if not users:
                await message.reply_text("⚠️ Hiện tại chưa có người chơi nào tham gia bát/dây hụi.", parse_mode=ParseMode.HTML)
                return

        # 2. Load template
        template_path = "Template_Hui.xlsx"
        if not os.path.exists(template_path):
            await message.reply_text("❌ Không tìm thấy tệp mẫu Template_Hui.xlsx trên hệ thống.", parse_mode=ParseMode.HTML)
            return

        wb = openpyxl.load_workbook(template_path, data_only=False)
        ws_template = wb.active # The template sheet

        # Keep a list of sheets to keep/generated
        generated_sheets = []

        # Styles copier helper
        def copy_cell_style(src_cell, dst_cell):
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

        # Scheduled date generator
        def get_scheduled_date(start_date, period_type, round_idx):
            if not start_date:
                return None
            delta = round_idx - 1
            if period_type == "Hụi ngày" or period_type == "DAILY":
                return start_date + datetime.timedelta(days=delta)
            elif period_type == "Hụi tuần" or period_type == "WEEKLY":
                return start_date + datetime.timedelta(weeks=delta)
            elif period_type == "Hụi 2 tuần" or period_type == "BIWEEKLY":
                return start_date + datetime.timedelta(weeks=delta * 2)
            else: # MONTHLY or default
                year = start_date.year + (start_date.month + delta - 1) // 12
                month = (start_date.month + delta - 1) % 12 + 1
                day = min(start_date.day, 28)
                return datetime.date(year, month, day)

        # Process each user
        users_processed = 0
        for user in users:
            # Query members (positions)
            members = db.query(RoscaMember).filter(RoscaMember.user_id == user.id).all()
            if not members:
                continue

            # Group members by rosca_id
            rosca_member_map = {}
            for m in members:
                if m.rosca_id not in rosca_member_map:
                    rosca_member_map[m.rosca_id] = []
                rosca_member_map[m.rosca_id].append(m)

            # Sort positions for each rosca
            for rid in rosca_member_map:
                rosca_member_map[rid].sort(key=lambda x: x.id)

            # Get Roscas
            roscas_in_db = {r.id: r for r in db.query(Rosca).filter(Rosca.id.in_(rosca_member_map.keys())).all()}
            sorted_rosca_ids = sorted(rosca_member_map.keys(), key=lambda rid: roscas_in_db[rid].code if rid in roscas_in_db else "")

            if not sorted_rosca_ids:
                continue

            # Sanitize and limit sheet title (limit to 30 chars, remove special excel chars)
            title = user.full_name or user.id or "N/A"
            for char in ['\\', '/', '?', '*', '[', ']']:
                title = title.replace(char, '')
            title = title[:30].strip()

            # Create sheet
            ws = wb.create_sheet(title=title)
            generated_sheets.append(ws)
            users_processed += 1

            # 3. Copy Column A and B styles and values from template
            ws.column_dimensions["A"].width = ws_template.column_dimensions["A"].width or 19.5
            ws.column_dimensions["B"].width = ws_template.column_dimensions["B"].width or 13.0

            for r in range(1, 41):
                # Column A
                cell_src_a = ws_template.cell(row=r, column=1)
                cell_dst_a = ws.cell(row=r, column=1)
                cell_dst_a.value = cell_src_a.value
                copy_cell_style(cell_src_a, cell_dst_a)

                # Column B
                cell_src_b = ws_template.cell(row=r, column=2)
                cell_dst_b = ws.cell(row=r, column=2)
                cell_dst_b.value = cell_src_b.value
                copy_cell_style(cell_src_b, cell_dst_b)

            # Find maximum rounds among user's roscas to dynamically size the history grid
            max_rounds = 12
            for rid in sorted_rosca_ids:
                rosca = roscas_in_db.get(rid)
                if rosca and rosca.total_parts and rosca.total_parts > max_rounds:
                    max_rounds = rosca.total_parts

            # Ensure column A & B are styled down to the dynamic SUM row if max_rounds > 12
            for r in range(28, 28 + max_rounds):
                cell_src_a = ws_template.cell(row=28, column=1)
                cell_dst_a = ws.cell(row=r, column=1)
                copy_cell_style(cell_src_a, cell_dst_a)

                cell_src_b = ws_template.cell(row=28, column=2)
                cell_dst_b = ws.cell(row=r, column=2)
                copy_cell_style(cell_src_b, cell_dst_b)

            sum_row_idx = 28 + max_rounds
            # Write and style dynamic SUM label in Column A
            cell_src_sum_a = ws_template.cell(row=40, column=1)
            cell_dst_sum_a = ws.cell(row=sum_row_idx, column=1)
            cell_dst_sum_a.value = cell_src_sum_a.value
            copy_cell_style(cell_src_sum_a, cell_dst_sum_a)

            cell_src_sum_b = ws_template.cell(row=40, column=2)
            cell_dst_sum_b = ws.cell(row=sum_row_idx, column=2)
            copy_cell_style(cell_src_sum_b, cell_dst_sum_b)

            # Build columns dynamically starting at Column C (index 3)
            curr_col = 3

            for rid_idx, rid in enumerate(sorted_rosca_ids):
                rosca = roscas_in_db.get(rid)
                if not rosca:
                    continue

                rosca_members = rosca_member_map[rid]
                num_positions = len(rosca_members)

                # Add a 1-column spacer if not the first Rosca
                if rid_idx > 0:
                    col_letter = get_column_letter(curr_col)
                    ws.column_dimensions[col_letter].width = 2.0
                    for r in range(1, sum_row_idx + 1):
                        src_cell = ws_template.cell(row=min(r, 40), column=5) # Col E style in template
                        dst_cell = ws.cell(row=r, column=curr_col)
                        copy_cell_style(src_cell, dst_cell)
                    curr_col += 1

                start_col = curr_col
                end_col = start_col + 2 * num_positions - 1

                # Copy styles for all columns of this Rosca
                for p in range(num_positions):
                    c_left = start_col + 2 * p
                    c_right = c_left + 1

                    col_letter_left = get_column_letter(c_left)
                    col_letter_right = get_column_letter(c_right)

                    ws.column_dimensions[col_letter_left].width = ws_template.column_dimensions["C"].width or 12.0
                    ws.column_dimensions[col_letter_right].width = ws_template.column_dimensions["D"].width or 15.0

                    for r in range(1, sum_row_idx + 1):
                        if r < 28:
                            src_l = ws_template.cell(row=r, column=3)
                            src_r = ws_template.cell(row=r, column=4)
                        elif r < sum_row_idx:
                            src_l = ws_template.cell(row=28, column=3)
                            src_r = ws_template.cell(row=28, column=4)
                        else:
                            src_l = ws_template.cell(row=40, column=3)
                            src_r = ws_template.cell(row=40, column=4)

                        dst_l = ws.cell(row=r, column=c_left)
                        dst_r = ws.cell(row=r, column=c_right)

                        copy_cell_style(src_l, dst_l)
                        copy_cell_style(src_r, dst_r)

                # Merge General Rosca Info headers (Rows 1 to 15)
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                for r in range(2, 16):
                    ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=end_col)

                # Fetch Rosca owner full name
                owner = db.query(UserRosca).filter(UserRosca.id == rosca.user_id).first()
                owner_name = owner.full_name if owner else "N/A"

                # Populate Rosca properties
                ws.cell(row=1, column=start_col).value = f"Dây hụi {rosca.code or ''}"
                ws.cell(row=2, column=start_col).value = owner_name
                ws.cell(row=3, column=start_col).value = rosca.code or "N/A"
                ws.cell(row=4, column=start_col).value = rosca.base_amount
                ws.cell(row=5, column=start_col).value = rosca.min_bid_amount
                ws.cell(row=6, column=start_col).value = rosca.max_bid_amount
                ws.cell(row=7, column=start_col).value = rosca.total_parts
                ws.cell(row=8, column=start_col).value = rosca.commission_fee
                ws.cell(row=9, column=start_col).value = rosca.start_date
                ws.cell(row=10, column=start_col).value = rosca.end_date
                ws.cell(row=11, column=start_col).value = rosca.payment_day
                ws.cell(row=12, column=start_col).value = rosca.bidding_time
                p_type = rosca.period_type
                if hasattr(p_type, 'value'):
                    p_type = p_type.value
                ws.cell(row=13, column=start_col).value = str(p_type or "N/A")
                ws.cell(row=14, column=start_col).value = rosca.status or "N/A"
                ws.cell(row=15, column=start_col).value = rosca.note or ""

                # Populate individual positions
                for p, member in enumerate(rosca_members):
                    c_left = start_col + 2 * p
                    c_right = c_left + 1

                    # Merge Position details headers (Rows 17 to 26)
                    for r in range(17, 27):
                        ws.merge_cells(start_row=r, start_column=c_left, end_row=r, end_column=c_right)

                    # Get contributions
                    contribs = db.query(RoscaContribution).filter(
                        RoscaContribution.member_id == member.id
                    ).all()

                    # Sort contributions sequentially/chronologically
                    round_nums = [c.round_number for c in contribs if c.round_number]
                    if len(set(round_nums)) == len(contribs) and all(r > 0 for r in round_nums):
                        contribs.sort(key=lambda x: x.round_number)
                        contrib_map = {c.round_number: c for c in contribs}
                    else:
                        # Fallback for seeded duplicate round numbers
                        contribs.sort(key=lambda x: (x.actual_payment_date or datetime.datetime.max, x.id))
                        contrib_map = {idx + 1: c for idx, c in enumerate(contribs)}

                    # Fill Position stats
                    ws.cell(row=17, column=c_left).value = member.id
                    ws.cell(row=18, column=c_left).value = member.parts_count or 1
                    paid_count = sum(1 for c in contribs if c.status in ["Paid", "Late"])
                    ws.cell(row=19, column=c_left).value = paid_count
                    total_contributed = sum(c.amount for c in contribs if c.status in ["Paid", "Late"])
                    ws.cell(row=20, column=c_left).value = total_contributed or member.total_contributed
                    ws.cell(row=21, column=c_left).value = member.total_received or 0.0
                    ws.cell(row=22, column=c_left).value = member.total_profit or 0.0
                    ws.cell(row=23, column=c_left).value = member.profit_rate or 0.0
                    ws.cell(row=24, column=c_left).value = member.status or "Playing"
                    ws.cell(row=25, column=c_left).value = member.telegram_group or "N/A"
                    ws.cell(row=26, column=c_left).value = member.note or ""

                    # Fill Round History
                    total_parts = rosca.total_parts or 12
                    for round_idx in range(1, total_parts + 1):
                        r = 27 + round_idx
                        c_data = contrib_map.get(round_idx)

                        if c_data:
                            p_date = c_data.actual_payment_date
                            if p_date:
                                ws.cell(row=r, column=c_left).value = p_date.date()
                            else:
                                ws.cell(row=r, column=c_left).value = "Chưa đóng"
                            ws.cell(row=r, column=c_right).value = c_data.amount
                        else:
                            # Project next payment dates
                            sched_date = get_scheduled_date(rosca.start_date, rosca.period_type, round_idx)
                            if sched_date:
                                ws.cell(row=r, column=c_left).value = sched_date
                            else:
                                ws.cell(row=r, column=c_left).value = ""
                            ws.cell(row=r, column=c_right).value = None

                    # Set Sum Formula
                    col_letter_right = get_column_letter(c_right)
                    formula_str = f"=SUM({col_letter_right}28:{col_letter_right}{sum_row_idx - 1})"
                    ws.cell(row=sum_row_idx, column=c_right).value = formula_str

                curr_col = end_col + 1

        if users_processed == 0:
            await message.reply_text("⚠️ Không tìm thấy người chơi nào có chân hụi hợp lệ để thống kê.", parse_mode=ParseMode.HTML)
            return

        # Delete the default template sheet
        if "Trang tính1" in wb.sheetnames:
            wb.remove(wb["Trang tính1"])

        # Save to BytesIO stream
        out_stream = io.BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)

        # File name
        if player_id:
            filename = f"Thong_Ke_Hui_{player_id}.xlsx"
            caption_text = f"📊 Thống kê hụi chi tiết cho người chơi: <b>{users[0].full_name}</b>"
        else:
            filename = "Thong_Ke_Hui_Tat_Ca.xlsx"
            caption_text = "📊 Thống kê hụi tổng hợp cho tất cả người chơi."

        # Send to Telegram
        await message.reply_document(
            document=out_stream,
            file_name=filename,
            caption=caption_text,
            parse_mode=ParseMode.HTML
        )
        LogInfo(f"[RoscaExportStats] Exported excel stats for {player_id or 'all'}. Users processed: {users_processed}. By {message.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error in roscas_export_stats_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình kết xuất thống kê.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# --- Rosca: Simulate Interest (Tính Lãi Giả Lập) ---
@bot.on_message(filters.command(["hui_tinh_lai_gia_lap"]) | filters.regex(r"^@\w+\s+/hui_tinh_lai_gia_lap\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main", "member")
async def rosca_tinh_lai_gia_lap_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["hui_tinh_lai_gia_lap"])
    if args is None: return

    if len(args) < 3:
        await message.reply_text(
            "⚠️ <b>Cú pháp lệnh không hợp lệ.</b>\n"
            "Vui lòng nhập: <code>/hui_tinh_lai_gia_lap [mã dây hụi] [Số tiền bỏ sắp tới]</code>\n"
            "Ví dụ: <code>/hui_tinh_lai_gia_lap DH01 100000</code>",
            parse_mode=ParseMode.HTML
        )
        return

    rosca_code = args[1].strip()
    upcoming_bid_str = args[2].strip()

    def parse_float(val_str):
        if not val_str: return None
        try:
            return float(val_str.replace(",", "").replace(".", "").replace(" ", ""))
        except:
            return None

    upcoming_bid = parse_float(upcoming_bid_str)
    if upcoming_bid is None or upcoming_bid < 0:
        await message.reply_text("⚠️ <b>Số tiền bỏ sắp tới</b> không hợp lệ. Vui lòng nhập một số dương.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        from app.models.rosca import Rosca, RoscaMember, RoscaContribution, UserRosca
        from bot.utils.utils import fmt_vn

        # 1. Tìm dây hụi
        rosca = db.query(Rosca).filter(Rosca.code == rosca_code).first()
        if not rosca:
            await message.reply_text(f"⚠️ Không tìm thấy Dây hụi có mã: <b>{rosca_code}</b>", parse_mode=ParseMode.HTML)
            return

        # 2. Các thông số gốc
        total_parts = rosca.total_parts or 0
        base_amount = rosca.base_amount or 0.0
        commission_fee = rosca.commission_fee or 0.0

        # 3. Tính toán các chỉ số
        total_parts_excluding_me = max(0, total_parts - 1)

        # Lấy chân hụi (RoscaMember) của người dùng hiện tại trong dây hụi này
        user_id_str = str(message.from_user.id) if message.from_user else None
        username_str = message.from_user.username if message.from_user else None

        member = db.query(RoscaMember).filter(
            RoscaMember.rosca_id == rosca.id
        ).first()

        # Số hụi chết = số lượng lần đóng hụi của chân hụi đó. Nếu không có thì = 0
        dead_rounds_count = 0
        if member:
            dead_rounds_count = db.query(RoscaContribution).filter(
                RoscaContribution.member_id == member.id,
                RoscaContribution.status == "Paid"
            ).count()

        # Số hụi sống = Tổng chân ngoài mình - Số hụi chết
        living_rounds_count = max(0, total_parts_excluding_me - dead_rounds_count)

        # Tổng tiền hốt = (Số hụi chết * Số tiền gốc 1 chân) + ((Số tiền gốc 1 chân - Số tiền bỏ sắp tới) * Số hụi sống) - tiền thảo
        total_withdrawn_amount = (dead_rounds_count * base_amount) + ((base_amount - upcoming_bid) * living_rounds_count) - commission_fee

        # Tổng hụi chuẩn = Số tiền gốc 1 chân * Tổng chân ngoài mình
        total_standard_amount = base_amount * total_parts_excluding_me

        # Chi phí (%) = 100 - (Tổng tiền hốt / Tổng hụi chuẩn)*100
        if total_standard_amount > 0:
            cost_percentage = 100.0 - (total_withdrawn_amount / total_standard_amount) * 100.0
        else:
            cost_percentage = 0.0

        # 4. Hiển thị kết quả
        result_text = f"""<b>TÍNH LÃI GIẢ LẬP DÂY HỤI: {rosca_code}</b>
━━━━━━━━━━━━━━━━━━━━━━
- <b>Tổng số chân:</b> {total_parts} chân
- <b>Tổng chân ngoài mình:</b> {total_parts_excluding_me} chân
- <b>Số hụi chết (Kỳ đã qua):</b> {dead_rounds_count} kỳ
- <b>Số hụi sống:</b> {living_rounds_count} kỳ
 
- <b>Số tiền gốc 1 chân:</b> {fmt_vn(base_amount)}
- <b>Số tiền bỏ sắp tới:</b> {fmt_vn(upcoming_bid)}
- <b>Tiền thảo (Commission):</b> {fmt_vn(commission_fee)}
━━━━━━━━━━━━━━━━━━━━━━
👉 <b>KẾT QUẢ GIẢ LẬP:</b>
- <b>Tổng tiền hốt:</b> <b>{fmt_vn(total_withdrawn_amount)}</b>
- <b>Tổng hụi chuẩn:</b> <b>{fmt_vn(total_standard_amount)}</b>
- <b>Chi phí:</b> <b>{cost_percentage:.2f}%</b>
 
💡 <i>Lưu ý: Kết quả trên mang tính chất giả lập dựa theo số tiền bỏ sắp tới được cung cấp.</i>"""

        await message.reply_text(result_text, parse_mode=ParseMode.HTML)
        LogInfo(f"[RoscaSimulateInterest] Simulated rosca {rosca_code} with bid {upcoming_bid} by {message.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error in rosca_tinh_lai_gia_lap_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình tính lãi giả lập.")
    finally:
        db.close()


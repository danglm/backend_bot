from pyrogram import filters
from pyrogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from pyrogram.enums import ParseMode
from bot.utils.bot import bot
from bot.utils.utils import check_command_target, require_user_type, require_project_name, require_custom_title, require_group_role, fmt_vn, fmt_money, fmt_num, fmt_weight
from bot.utils.enums import UserType, CustomTitle
from bot.utils.logger import LogInfo, LogError, LogType
from app.db.session import SessionLocal
import re
from datetime import datetime, timedelta
import traceback

# TienNga project handlers


# --- Thêm nhân viên ---
@bot.on_message(filters.command(["tien_nga_create_employee", "tien_nga_tao_nhan_vien"]) | filters.regex(r"^@\w+\s+/(tien_nga_create_employee|tien_nga_tao_nhan_vien)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def tien_nga_create_employee_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_create_employee", "tien_nga_tao_nhan_vien"])
    if args is None: return

    from bot.utils.human_resource import handle_create_employee
    await handle_create_employee(client, message, "/tien_nga_create_employee")


# --- Cập nhật nhân viên ---
@bot.on_message(filters.command(["tien_nga_update_employee", "tien_nga_cap_nhat_nhan_vien"]) | filters.regex(r"^@\w+\s+/(tien_nga_update_employee|tien_nga_cap_nhat_nhan_vien)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def tien_nga_update_employee_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_update_employee", "tien_nga_cap_nhat_nhan_vien"])
    if args is None: return

    from bot.utils.human_resource import handle_update_employee
    await handle_update_employee(client, message, "/tien_nga_update_employee")


# --- Xóa nhân viên (soft delete) ---
@bot.on_message(filters.command(["tien_nga_delete_employee", "tien_nga_xoa_nhan_vien"]) | filters.regex(r"^@\w+\s+/(tien_nga_delete_employee|tien_nga_xoa_nhan_vien)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def tien_nga_delete_employee_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_delete_employee", "tien_nga_xoa_nhan_vien"])
    if args is None: return

    from bot.utils.human_resource import handle_delete_employee
    await handle_delete_employee(client, message, "/tien_nga_delete_employee")


@bot.on_callback_query(filters.regex(r"^del_emp_(confirm|cancel)\|(.+)$"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def tien_nga_delete_employee_cb_handler(client, callback_query):
    from bot.utils.human_resource import handle_delete_employee_callback
    await handle_delete_employee_callback(client, callback_query)

@bot.on_message(filters.command(["tien_nga_check_in", "tien_nga_cham_cong"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_in|tien_nga_cham_cong)\b"))
@require_project_name("Tiến Nga")
@require_group_role("member")
@require_custom_title(CustomTitle.MEMBER_HR)
async def tien_nga_check_in_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_check_in", "tien_nga_cham_cong"])
    if args is None: return

    from bot.utils.human_resource import handle_check_in
    await handle_check_in(client, message, "/tien_nga_check_in")

@bot.on_message(filters.command(["tien_nga_check_out", "tien_nga_tan_ca"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_out|tien_nga_tan_ca)\b"))
@require_project_name("Tiến Nga")
@require_group_role("member")
@require_custom_title(CustomTitle.MEMBER_HR)
async def tien_nga_check_out_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_check_out", "tien_nga_tan_ca"])
    if args is None: return

    from bot.utils.human_resource import handle_check_out
    await handle_check_out(client, message, "/tien_nga_check_out")

@bot.on_message(filters.command(["tien_nga_request_leave", "tien_nga_xin_nghi_phep"]) | filters.regex(r"^@\w+\s+/(tien_nga_request_leave|tien_nga_xin_nghi_phep)\b"))
@require_project_name("Tiến Nga")
@require_group_role("member")
@require_custom_title(CustomTitle.MEMBER_HR)
async def tien_nga_request_leave_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_request_leave", "tien_nga_xin_nghi_phep"])
    if args is None: return

    from bot.utils.human_resource import handle_request_leave
    await handle_request_leave(client, message, "/tien_nga_request_leave")

@bot.on_message(filters.command(["tien_nga_request_overtime", "tien_nga_dang_ky_tang_ca"]) | filters.regex(r"^@\w+\s+/(tien_nga_request_overtime|tien_nga_dang_ky_tang_ca)\b"))
@require_project_name("Tiến Nga")
@require_group_role("member")
@require_custom_title(CustomTitle.MEMBER_HR)
async def tien_nga_request_overtime_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_request_overtime", "tien_nga_dang_ky_tang_ca"])
    if args is None: return

    from bot.utils.human_resource import handle_request_overtime
    await handle_request_overtime(client, message, "/tien_nga_request_overtime")

@bot.on_message(filters.command(["tien_nga_list_check_in", "tien_nga_xem_cham_cong"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_check_in|tien_nga_xem_cham_cong)\b"))
@require_project_name("Tiến Nga")
@require_group_role("member")
@require_custom_title(CustomTitle.MEMBER_HR)
async def tien_nga_list_check_in_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_list_check_in", "tien_nga_xem_cham_cong"])
    if args is None: return

    from bot.utils.human_resource import handle_list_check_in
    await handle_list_check_in(client, message, "/tien_nga_list_check_in")

@bot.on_message(filters.command(["tien_nga_list_request_leave", "tien_nga_xem_nghi_phep"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_request_leave|tien_nga_xem_nghi_phep)\b"))
@require_project_name("Tiến Nga")
@require_group_role("member")
@require_custom_title(CustomTitle.MEMBER_HR)
async def tien_nga_list_request_leave_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_list_request_leave", "tien_nga_xem_nghi_phep"])
    if args is None: return

    from bot.utils.human_resource import handle_list_request_leave
    await handle_list_request_leave(client, message, "/tien_nga_list_request_leave")

@bot.on_message(filters.command(["tien_nga_create_task", "tien_nga_giao_viec"]) | filters.regex(r"^@\w+\s+/(tien_nga_create_task|tien_nga_giao_viec)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def tien_nga_create_task_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_create_task", "tien_nga_giao_viec"])
    if args is None: return

    from bot.utils.human_resource import handle_create_task
    await handle_create_task(client, message, "/tien_nga_create_task")

@bot.on_message(filters.command(["tien_nga_list_tasks", "tien_nga_danh_sach_cong_viec"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_tasks|tien_nga_danh_sach_cong_viec)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def tien_nga_list_tasks_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_list_tasks", "tien_nga_danh_sach_cong_viec"])
    if args is None: return

    from bot.utils.human_resource import handle_check_tasks
    cmd = args[0] if args else "tien_nga_danh_sach_cong_viec"
    clean_cmd = cmd.split('@')[0]
    message.text = cmd + " " + " ".join(args[1:]) if len(args) > 1 else cmd
    await handle_check_tasks(client, message, clean_cmd)

# --- Nhân viên xem công việc của mình ---
@bot.on_message(filters.command(["tien_nga_xem_cong_viec"]) | filters.regex(r"^@\w+\s+/tien_nga_xem_cong_viec\b"))
@require_project_name("Tiến Nga")
@require_group_role("member")
@require_custom_title(CustomTitle.MEMBER_HR)
async def tien_nga_xem_cong_viec_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_xem_cong_viec"])
    if args is None: return

    from bot.utils.human_resource import handle_list_tasks
    await handle_list_tasks(client, message, "/tien_nga_xem_cong_viec")

# --- Task callback handlers ---
@bot.on_callback_query(filters.regex(r"^tsk_sel_(.+)$"))
async def _tsk_sel(client, callback_query):
    from bot.utils.human_resource import task_select_callback
    await task_select_callback(client, callback_query)

@bot.on_callback_query(filters.regex(r"^tsk_detail_(.+)$"))
async def _tsk_detail(client, callback_query):
    from bot.utils.human_resource import task_detail_callback
    await task_detail_callback(client, callback_query)

@bot.on_callback_query(filters.regex(r"^tsk_status_(.+)$"))
async def _tsk_status(client, callback_query):
    from bot.utils.human_resource import task_status_callback
    await task_status_callback(client, callback_query)

@bot.on_callback_query(filters.regex(r"^tsk_set_(PENDING|IN_PROGRESS|COMPLETED|CANCELLED)_(.+)$"))
async def _tsk_set(client, callback_query):
    from bot.utils.human_resource import task_set_status_callback
    await task_set_status_callback(client, callback_query)

@bot.on_callback_query(filters.regex(r"^tsk_back_list$"))
async def _tsk_back(client, callback_query):
    from bot.utils.human_resource import task_back_list_callback
    await task_back_list_callback(client, callback_query)

@bot.on_callback_query(filters.regex(r"^tsk_cancel$"))
async def _tsk_cancel(client, callback_query):
    from bot.utils.human_resource import task_cancel_callback
    await task_cancel_callback(client, callback_query)

@bot.on_message(filters.command(["cancel", "huy_task"]) & filters.reply)
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
async def tien_nga_cancel_handler(client, message: Message) -> None:
    replied = message.reply_to_message
    if not replied or not replied.text:
        await message.reply_text("⚠️ Vui lòng reply vào tin nhắn cần hủy.", parse_mode=ParseMode.HTML)
        return
    
    replied_text = replied.text or ""
    
    # --- Hủy nhập mua mủ ---
    if "NHẬP MUA MỦ THÀNH CÔNG" in replied_text:
        # Parse purchase ID from the replied message
        id_match = re.search(r"🆔\s*([a-f0-9\-]{36})", replied_text)
        if not id_match:
            await message.reply_text(
                "⚠️ Không tìm thấy mã giao dịch trong tin nhắn. Không thể hủy.",
                parse_mode=ParseMode.HTML
            )
            return
        
        purchase_id = id_match.group(1)
        
        from app.models.business import DailyPurchases
        db = SessionLocal()
        try:
            import uuid as uuid_lib
            purchase = db.query(DailyPurchases).filter(
                DailyPurchases.id == uuid_lib.UUID(purchase_id)
            ).first()
            
            if not purchase:
                await message.reply_text(
                    f"⚠️ Không tìm thấy giao dịch <b>{purchase_id}</b> trong hệ thống (có thể đã bị xóa trước đó).",
                    parse_mode=ParseMode.HTML
                )
                return
            
            # Lưu thông tin trước khi xóa để hiển thị
            ho_id = purchase.hoursehold_id
            ngay = purchase.day.strftime("%d/%m/%Y") if purchase.day else "—"
            total = purchase.total_amount or 0
            saved = purchase.saved_amount or 0
            
            def fmt_money(val):
                if val is None: return "0 đ"
                try:
                    return f"{int(val):,} đ".replace(",", ".")
                except:
                    return str(val)
            
            # Update total debt for customer if there was a saved amount
            if saved > 0:
                from app.models.business import Customers
                customer = db.query(Customers).filter(Customers.hoursehold_id == ho_id).first()
                if customer:
                    if customer.total_debt is None:
                        customer.total_debt = 0
                    customer.total_debt -= int(saved)
                    if customer.total_debt < 0:
                        customer.total_debt = 0
                        
            db.delete(purchase)
            db.commit()
            
            LogInfo(
                f"[TienNga] Cancelled daily purchase '{purchase_id}' for '{ho_id}' on {ngay} by user {message.from_user.id}",
                LogType.SYSTEM_STATUS
            )
            
            await message.reply_text(
                f"<b>ĐÃ HỦY NHẬP MUA MỦ</b>\n\n"
                f"<b>Mã Hộ:</b> {ho_id}\n"
                f"<b>Ngày:</b> {ngay}\n"
                f"<b>Thành Tiền:</b> <code>{fmt_money(total)}</code>\n\n"
                f"<i>Giao dịch đã được xóa khỏi hệ thống.</i>",
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            db.rollback()
            LogError(f"Error cancelling daily purchase: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text(
                f"❌ Có lỗi xảy ra khi hủy giao dịch: {e}",
                parse_mode=ParseMode.HTML
            )
        finally:
            db.close()
        return
    
    from bot.utils.human_resource import handle_cancel_task_reply
    await handle_cancel_task_reply(client, message)

@bot.on_message(filters.command(["tien_nga_check_tasks", "tien_nga_xem_cong_viec"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_tasks|tien_nga_xem_cong_viec)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def tien_nga_check_tasks_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_check_tasks", "tien_nga_xem_cong_viec"])
    if args is None: return

    from bot.utils.human_resource import handle_check_tasks
    await handle_check_tasks(client, message, "/tien_nga_check_tasks")

@bot.on_message(filters.command(["tien_nga_list_payroll", "tien_nga_xuat_danh_sach_luong"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_payroll|tien_nga_xuat_danh_sach_luong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def tien_nga_list_payroll_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_list_payroll", "tien_nga_xuat_danh_sach_luong"])
    if args is None: return

    from bot.utils.human_resource import handle_list_payroll_excel
    import re
    cmd = args[0] if args else "tien_nga_list_payroll"
    message.text = cmd + " " + " ".join(args[1:]) if len(args) > 1 else cmd
    await handle_list_payroll_excel(client, message, cmd)

@bot.on_message(filters.command(["export_payroll", "tien_nga_xuat_luong", "tien_nga_export_payroll"]) | filters.regex(r"^@\w+\s+/(export_payroll|tien_nga_xuat_luong|tien_nga_export_payroll)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def tien_nga_export_payroll_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["export_payroll", "tien_nga_xuat_luong", "tien_nga_export_payroll"])
    if args is None: return

    from bot.utils.human_resource import handle_export_payroll
    # Re-construct message.text for the actual handler to parse args
    import re
    cmd = args[0] if args else "export_payroll"
    message.text = cmd + " " + " ".join(args[1:]) if len(args) > 1 else cmd
    await handle_export_payroll(client, message, cmd)

@bot.on_message(filters.command(["tien_nga_recreate_attendance_report", "tien_nga_tao_lai_bang_cham_cong"]) | filters.regex(r"^@\w+\s+/(tien_nga_recreate_attendance_report|tien_nga_tao_lai_bang_cham_cong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def tien_nga_recreate_attendance_report_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_recreate_attendance_report", "tien_nga_tao_lai_bang_cham_cong"])
    if args is None: return

    from bot.utils.human_resource import handle_recreate_attendance_report
    import re
    cmd = args[0] if args else "tien_nga_recreate_attendance_report"
    message.text = cmd + " " + " ".join(args[1:]) if len(args) > 1 else cmd
    await handle_recreate_attendance_report(client, message, cmd)


#############  Supplier #############
@bot.on_message(filters.command(["tien_nga_create_collection_point", "tien_nga_tao_diem_thu_mua"]) | filters.regex(r"^@\w+\s+/(tien_nga_create_collection_point|tien_nga_tao_diem_thu_mua)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_create_collection_point_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    
    if len(lines) < 2:
        form_template = """<b>FORM TẠO ĐIỂM THU MUA MỚI</b>
Vui lòng sao chép form dưới đây, điền đầy đủ thông tin và gửi lại:

<pre>/tien_nga_create_collection_point
Tên Điểm Thu Mua: 
Địa Chỉ: </pre>"""
        await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        return

    # Parse form data
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    collection_name = data.get("Tên Điểm Thu Mua", "").strip()
    address = data.get("Địa Chỉ", "").strip()

    if not collection_name or not address:
        await message.reply_text(
            "⚠️ <b>Tên Điểm Thu Mua</b> và <b>Địa Chỉ</b> là bắt buộc.",
            parse_mode=ParseMode.HTML
        )
        return

    from app.db.session import SessionLocal
    from app.models.business import CollectionPoint
    
    db = SessionLocal()
    try:
        # Check if already exists
        existing = db.query(CollectionPoint).filter(CollectionPoint.collection_name == collection_name).first()
        if existing:
            await message.reply_text(
                f"⚠️ Điểm thu mua <b>{collection_name}</b> đã tồn tại trong hệ thống.",
                parse_mode=ParseMode.HTML
            )
            return
            
        new_point = CollectionPoint(
            collection_name=collection_name,
            address=address
        )
        db.add(new_point)
        db.commit()
        
        LogInfo(f"[TienNga] Created collection point '{collection_name}' by user {message.from_user.id}", LogType.SYSTEM_STATUS)
        
        await message.reply_text(
            f"<b>TẠO ĐIỂM THU MUA THÀNH CÔNG</b>\n\n"
            f"<b>Tên:</b> {collection_name}\n"
            f"<b>Địa Chỉ:</b> {address}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error creating collection point: {e}", LogType.SYSTEM_STATUS)
        db.rollback()
        await message.reply_text("❌ Có lỗi xảy ra khi lưu vào database.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_list_collection_point", "tien_nga_danh_sach_diem_thu_mua"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_collection_point|tien_nga_danh_sach_diem_thu_mua)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_list_collection_point_handler(client, message: Message) -> None:
    from app.db.session import SessionLocal
    from app.models.business import CollectionPoint
    
    db = SessionLocal()
    try:
        points = db.query(CollectionPoint).all()
        if not points:
            await message.reply_text("⚠️ <b>Không có Điểm Thu Mua nào trong hệ thống.</b>", parse_mode=ParseMode.HTML)
            return
            
        text = "<b>DANH SÁCH ĐIỂM THU MUA</b>\n\n"
        for idx, p in enumerate(points, 1):
            text += f"<b>{idx}. {p.collection_name}</b>\n"
            text += f"   Mã: <code>{p.id}</code>\n"
            text += f"   Địa chỉ: {p.address}\n\n"
            
        await message.reply_text(text, parse_mode=ParseMode.HTML)
    except Exception as e:
        LogError(f"Error listing collection points: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra khi tải danh sách Điểm Thu Mua.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_create_customer", "tien_nga_tao_khach_hang"]) | filters.regex(r"^@\w+\s+/(tien_nga_create_customer|tien_nga_tao_khach_hang)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_create_customer_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    
    if len(lines) < 2:
        from app.db.session import SessionLocal
        from app.models.business import CollectionPoint
        
        db = SessionLocal()
        try:
            points = db.query(CollectionPoint).all()
            if not points:
                await message.reply_text("⚠️ Chưa có Điểm Thu Mua nào. Vui lòng tạo trước bằng lệnh /tien_nga_create_collection_point.")
                return
                
            buttons = []
            for p in points:
                buttons.append([InlineKeyboardButton(p.collection_name, callback_data=f"tn_selcp_{p.id}")])
                
            markup = InlineKeyboardMarkup(buttons)
            await message.reply_text("<b>Vui lòng chọn Điểm Thu Mua (Xưởng) cho Khách Hàng:</b>", reply_markup=markup, parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error listing collection points: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Có lỗi xảy ra.")
        finally:
            db.close()
        return

    # Parse form data
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    cp_id = data.get("Mã Điểm Thu", "").strip()
    fullname = data.get("Tên Khách Hàng", "").strip()
    hoursehold_id = data.get("Mã Hộ", "").strip()
    phone = data.get("SĐT Khách", "").strip()
    address = data.get("Địa Chỉ", "").strip()
    ingredient = data.get("Nguyên Liệu", "").strip()
    username = data.get("Username TG", "").strip()
    telegram_group = data.get("Nhóm Telegram", "").strip()
    bank_name = data.get("Ngân Hàng", "").strip()
    number_bank = data.get("STK Ngân Hàng", "").strip()
    
    amount_of_debt = parse_float_vn(data.get("Số Tiền Nợ", "0"))
    cash_advance = parse_float_vn(data.get("Ứng Tiền Cuối Mùa", "0"))
    total_debt = parse_float_vn(data.get("Tổng Công Nợ", "0"))

    if not cp_id or not fullname or not hoursehold_id:
        await message.reply_text("⚠️ <b>Mã Điểm Thu</b>, <b>Tên Khách Hàng</b> và <b>Mã Hộ</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    from app.db.session import SessionLocal
    from app.models.business import Customers
    import uuid
    
    db = SessionLocal()
    try:
        existing = db.query(Customers).filter(Customers.hoursehold_id == hoursehold_id).first()
        if existing:
            await message.reply_text(f"⚠️ Khách hàng mã hộ <b>{hoursehold_id}</b> đã tồn tại.", parse_mode=ParseMode.HTML)
            return
            
        new_customer = Customers(
            id=str(uuid.uuid4()),
            fullname=fullname,
            hoursehold_id=hoursehold_id,
            collection_point_id=cp_id,
            number_phone=phone,
            address=address,
            ingredient=ingredient,
            telegram_group=telegram_group if telegram_group else None,
            bank_name=bank_name if bank_name else None,
            number_bank=number_bank if number_bank else None,
            amount_of_debt=int(amount_of_debt),
            cash_advance=int(cash_advance),
            total_debt=int(total_debt),
            username=username if username else None
        )
        db.add(new_customer)
        db.commit()
        
        LogInfo(f"[TienNga] Created customer '{hoursehold_id}' by user {message.from_user.id}", LogType.SYSTEM_STATUS)
        
        await message.reply_text(f"✅ <b>TẠO KHÁCH HÀNG THÀNH CÔNG</b>\n\n<b>Mã Hộ:</b> {hoursehold_id}\n<b>Tên KH:</b> {fullname}", parse_mode=ParseMode.HTML)
    except Exception as e:
        LogError(f"Error creating customer: {e}", LogType.SYSTEM_STATUS)
        db.rollback()
        await message.reply_text("❌ Có lỗi xảy ra khi lưu vào database.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^tn_selcp_(.+)$"))
async def tien_nga_select_collection_point_callback(client, callback_query):
    cp_id = callback_query.matches[0].group(1)
    from app.db.session import SessionLocal
    from app.models.business import CollectionPoint
    
    db = SessionLocal()
    try:
        cp = db.query(CollectionPoint).filter(CollectionPoint.id == cp_id).first()
        if not cp:
            await callback_query.answer("⚠️ Không tìm thấy Điểm Thu Mua", show_alert=True)
            return
            
        cp_name = cp.collection_name
        form_template = f"""<b>FORM TẠO KHÁCH HÀNG MỚI</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/tien_nga_create_customer
Mã Điểm Thu: {cp_id}
Tên Khách Hàng: 
Mã Hộ: 
SĐT Khách: 
Địa Chỉ: 
Nguyên Liệu: 
Số Tiền Nợ: 
Ứng Tiền Cuối Mùa: 
Tổng Công Nợ: 
Username TG: 
Nhóm Telegram: 
Ngân Hàng: 
STK Ngân Hàng: </pre>

<i>(*Ghi chú: Đang thêm khách hàng cho Xưởng: <b>{cp_name}</b>)</i>"""
        await callback_query.message.reply_text(form_template, parse_mode=ParseMode.HTML)
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in select collection point: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống", show_alert=True)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_update_customer", "tien_nga_cap_nhat_khach_hang"]) | filters.regex(r"^@\w+\s+/(tien_nga_update_customer|tien_nga_cap_nhat_khach_hang)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_update_customer_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    
    if len(lines) < 2:
        args = lines[0].split()
        if len(args) != 2:
            await message.reply_text("⚠️ Cú pháp: <code>/tien_nga_update_customer [Mã Hộ]</code>\n\n<i>Ví dụ: <code>/tien_nga_update_customer KH001</code></i>", parse_mode=ParseMode.HTML)
            return
            
        hoursehold_id = args[1]
        
        from app.db.session import SessionLocal
        from app.models.business import Customers
        
        db = SessionLocal()
        try:
            customer = db.query(Customers).filter(Customers.hoursehold_id == hoursehold_id).first()
            if not customer:
                await message.reply_text(f"⚠️ Không tìm thấy Khách hàng với mã hộ <b>{hoursehold_id}</b>.", parse_mode=ParseMode.HTML)
                return
                
            def fmt_num(val):
                if val is None: return 0
                return int(val) if val == int(val) else val
            
            # Format number output to Vietnam string but keeping the numeric in form to be copy pasted easily
            form_template = f"""<b>FORM CẬP NHẬT KHÁCH HÀNG</b>
Vui lòng sao chép form dưới đây, sửa thông tin và gửi lại:

<pre>/tien_nga_update_customer
Mã Hộ: {customer.hoursehold_id}
Tên Khách Hàng: {customer.fullname or ""}
SĐT Khách: {customer.number_phone or ""}
Địa Chỉ: {customer.address or ""}
Nguyên Liệu: {customer.ingredient or ""}
Số Tiền Nợ: {fmt_num(customer.amount_of_debt)}
Ứng Tiền Cuối Mùa: {fmt_num(customer.cash_advance)}
Tổng Công Nợ: {fmt_num(customer.total_debt)}
Username: {customer.username or ''}
Nhóm Telegram: {customer.telegram_group or ''}
Ngân Hàng: {customer.bank_name or ''}
STK Ngân Hàng: {customer.number_bank or ''}</pre>"""
            await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error fetching customer for update: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    # Parse form data (nhiều dòng)
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    hoursehold_id = data.get("Mã Hộ", "").strip()
    if not hoursehold_id:
        await message.reply_text("⚠️ <b>Mã Hộ</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
        
    fullname = data.get("Tên Khách Hàng", "").strip()
    phone = data.get("SĐT Khách", "").strip()
    address = data.get("Địa Chỉ", "").strip()
    ingredient = data.get("Nguyên Liệu", "").strip()
    username = data.get("Username", "").strip()
    telegram_group = data.get("Nhóm Telegram", "").strip()
    bank_name = data.get("Ngân Hàng", "").strip()
    number_bank = data.get("STK Ngân Hàng", "").strip()
    
    amount_of_debt = parse_float_vn(data.get("Số Tiền Nợ", "0"))
    cash_advance = parse_float_vn(data.get("Ứng Tiền Cuối Mùa", "0"))
    total_debt = parse_float_vn(data.get("Tổng Công Nợ", "0"))

    from app.db.session import SessionLocal
    from app.models.business import Customers
    
    db = SessionLocal()
    try:
        customer = db.query(Customers).filter(Customers.hoursehold_id == hoursehold_id).first()
        if not customer:
            await message.reply_text(f"⚠️ Khách hàng mã hộ <b>{hoursehold_id}</b> không tồn tại.", parse_mode=ParseMode.HTML)
            return
            
        if fullname: customer.fullname = fullname
        customer.number_phone = phone
        customer.address = address
        customer.ingredient = ingredient
        customer.telegram_group = telegram_group if telegram_group else None
        customer.bank_name = bank_name if bank_name else None
        customer.number_bank = number_bank if number_bank else None
        customer.amount_of_debt = int(amount_of_debt)
        customer.cash_advance = int(cash_advance)
        customer.total_debt = int(total_debt)
        customer.username = username
        
        db.commit()
        
        LogInfo(f"[TienNga] Updated customer '{hoursehold_id}' by user {message.from_user.id}", LogType.SYSTEM_STATUS)
            
        await message.reply_text(
            f"✅ <b>CẬP NHẬT KHÁCH HÀNG THÀNH CÔNG</b>\n\n"
            f"<b>Mã Hộ:</b> {hoursehold_id}\n"
            f"<b>Tên KH:</b> {customer.fullname}\n"
            f"<b>SĐT:</b> {customer.number_phone or '—'}\n"
            f"<b>Địa Chỉ:</b> {customer.address or '—'}\n"
            f"<b>Nguyên Liệu:</b> {customer.ingredient or '—'}\n"
            f"<b>Nhóm Telegram:</b> {customer.telegram_group or '—'}\n"
            f"<b>Ngân hàng:</b> {customer.bank_name or '—'} <i>({customer.number_bank or '—'})</i>\n"
            f"<b>Số Tiền Nợ:</b> <code>{fmt_vn(customer.amount_of_debt)}</code>\n"
            f"<b>Ứng Cuối Mùa:</b> <code>{fmt_vn(customer.cash_advance)}</code>\n"
            f"<b>Tổng Nợ:</b> <code>{fmt_vn(customer.total_debt)}</code>\n"
            f"<b>Username:</b> {customer.username or '—'}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error updating customer: {e}", LogType.SYSTEM_STATUS)
        db.rollback()
        await message.reply_text("❌ Có lỗi xảy ra khi lưu vào database.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_check_customer", "tien_nga_kiem_tra_khach_hang"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_customer|tien_nga_kiem_tra_khach_hang)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_check_customer_handler(client, message: Message) -> None:
    args = message.text.split()
    if len(args) != 2:
        await message.reply_text(
            "⚠️ Cú pháp: <code>/tien_nga_check_customer [Mã Hộ]</code>\n\n"
            "<i>Ví dụ: <code>/tien_nga_check_customer KH001</code></i>",
            parse_mode=ParseMode.HTML
        )
        return
        
    hoursehold_id = args[1]
    
    from app.db.session import SessionLocal
    from app.models.business import Customers, CollectionPoint
    
    db = SessionLocal()
    try:
        customer = db.query(Customers).filter(Customers.hoursehold_id == hoursehold_id).first()
        if not customer:
            await message.reply_text(f"⚠️ Không tìm thấy Khách hàng mã hộ <b>{hoursehold_id}</b>.", parse_mode=ParseMode.HTML)
            return

        cp_name = "Chưa rõ"
        if customer.collection_point_id:
            cp = db.query(CollectionPoint).filter(CollectionPoint.id == customer.collection_point_id).first()
            if cp:
                cp_name = cp.collection_name
                
        status_str = "Hoạt động" if customer.status == "ACTIVE" else ("Đã ẩn/xóa" if customer.status == "DELETED" else customer.status)
        
        await message.reply_text(
            f"<b>THÔNG TIN KHÁCH HÀNG: {customer.fullname or '—'}</b>\n\n"
            f"<b>Mã Hộ:</b> {customer.hoursehold_id}\n"
            f"<b>Tên Khách Hàng:</b> {customer.fullname or '—'}\n"
            f"<b>Điểm Thu Mua:</b> {cp_name}\n"
            f"<b>SĐT:</b> {customer.number_phone or '—'}\n"
            f"<b>Địa Chỉ:</b> {customer.address or '—'}\n"
            f"<b>Nguyên Liệu:</b> {customer.ingredient or '—'}\n"
            f"<b>Số Tiền Nợ:</b> <code>{fmt_vn(customer.amount_of_debt)}</code>\n"
            f"<b>Ứng Cuối Mùa:</b> <code>{fmt_vn(customer.cash_advance)}</code>\n"
            f"<b>Tổng Công Nợ:</b> <code>{fmt_vn(customer.total_debt)}</code>\n"
            f"<b>Trạng Thái:</b> {status_str}\n"
            f"<b>Tài khoản TG:</b> {customer.username or '—'}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error checking customer: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Lỗi hệ thống khi tìm kiếm khách hàng.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_delete_customer", "tien_nga_xoa_khach_hang"]) | filters.regex(r"^@\w+\s+/(tien_nga_delete_customer|tien_nga_xoa_khach_hang)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_delete_customer_handler(client, message: Message) -> None:
    args = message.text.split()
    if len(args) != 2:
        await message.reply_text(
            "⚠️ Cú pháp: <code>/tien_nga_delete_customer [Mã Hộ]</code>\n\n"
            "<i>Ví dụ: <code>/tien_nga_delete_customer KH001</code></i>",
            parse_mode=ParseMode.HTML
        )
        return
        
    hoursehold_id = args[1]
    
    from app.db.session import SessionLocal
    from app.models.business import Customers
    
    db = SessionLocal()
    try:
        customer = db.query(Customers).filter(Customers.hoursehold_id == hoursehold_id).first()
        if not customer:
            await message.reply_text(f"⚠️ Không tìm thấy Khách hàng mã hộ <b>{hoursehold_id}</b>.", parse_mode=ParseMode.HTML)
            return
            
        if customer.status == "INACTIVE" or customer.status == "DELETED":
            await message.reply_text(f"ℹ️ Khách hàng mã hộ <b>{hoursehold_id}</b> đã ở trạng thái khóa / xóa từ trước.", parse_mode=ParseMode.HTML)
            return

        customer.status = "DELETED"
        db.commit()
        
        LogInfo(f"[TienNga] Soft deleted customer '{hoursehold_id}' by user {message.from_user.id}", LogType.SYSTEM_STATUS)
        
        await message.reply_text(
            f"✅ <b>XÓA KHÁCH HÀNG THÀNH CÔNG</b>\n\n"
            f"Mã Hộ: <b>{hoursehold_id}</b>\n"
            f"Tên KH: {customer.fullname}\n"
            f"<i>Khách hàng này đã được ẩn (chuyển sang trạng thái DELETED).</i>",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error soft deleting customer: {e}", LogType.SYSTEM_STATUS)
        db.rollback()
        await message.reply_text("❌ Có lỗi xảy ra khi cập nhật database.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_list_employee", "tien_nga_danh_sach_nhan_vien"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_employee|tien_nga_danh_sach_nhan_vien)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def tien_nga_list_employee_handler(client, message: Message) -> None:
    db = SessionLocal()
    try:
        from app.models.employee import Employee
        import tempfile
        import os
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        employees = db.query(Employee).order_by(Employee.last_name, Employee.first_name).all()

        if not employees:
            await message.reply_text("⚠️ Không có nhân viên nào trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DanhSachNhanVien"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        money_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        alt_fill = PatternFill("solid", fgColor="D9E2F3")

        headers = [
            "Mã nhân viên", "Họ và tên", "Giới tính", "Số điện thoại", 
            "Username Telegram", "Nhóm nhân viên Telegram", "Địa chỉ", 
            "CCCD", "Mức lương tháng", "Mức lương tuần", "Mức lương ngày", 
            "Mức lương giờ", "Thời gian vào ca", "Thời gian kết thúc ca", "Công nợ"
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        def fmt_money(val):
            if val is None or val == 0:
                return "0"
            return f"{int(val):,}".replace(",", ".")

        for idx, emp in enumerate(employees, 1):
            row = idx + 1
            row_fill = alt_fill if idx % 2 == 0 else None
            
            full_name = f"{emp.last_name or ''} {emp.first_name or ''}".strip()
            
            # Format times gracefully
            start_date_str = emp.start_time.strftime('%H:%M %d/%m/%Y') if emp.start_time else '—'
            end_date_str = emp.end_time.strftime('%H:%M %d/%m/%Y') if emp.end_time else '—'

            values = [
                emp.id,
                full_name,
                emp.gender or "—",
                emp.number_phone or "—",
                emp.username or "—",
                emp.telegram_group or "—",
                emp.address or "—",
                emp.identity_card or "—",
                fmt_money(emp.monthly_salary),
                fmt_money(emp.weekly_salary),
                fmt_money(emp.daily_salary),
                fmt_money(emp.hourly_salary),
                start_date_str,
                end_date_str,
                fmt_money(emp.total_debt)
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                
                # Alignments
                if col_idx in [1, 3, 4, 8, 13, 14]:
                    cell.alignment = center_align
                elif col_idx in [9, 10, 11, 12, 15]:
                    cell.alignment = money_align
                else:
                    cell.alignment = left_align
                    
                if row_fill:
                    cell.fill = row_fill

        col_widths = [15, 25, 12, 15, 20, 20, 30, 20, 18, 18, 18, 18, 20, 20, 18]
        for col_idx, width in enumerate(col_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        now_str = datetime.now().strftime("%Y%m%d_%H%M")
        await message.reply_document(
            document=tmp_path,
            file_name=f"danh_sach_nhan_vien_{now_str}.xlsx",
            caption=(
                f"<b>DANH SÁCH NHÂN VIÊN</b>\n\n"
                f"Tổng cộng: <b>{len(employees)}</b> nhân viên\n"
                f"<i>Xuất lúc: {datetime.now().strftime('%H:%M %d/%m/%Y')}</i>"
            ),
            parse_mode=ParseMode.HTML,
        )

        os.remove(tmp_path)
        LogInfo(f"[TienNga] Exported {len(employees)} employees by @{message.from_user.username}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error in tien_nga_list_employee_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(f"❌ Có lỗi xảy ra khi xuất danh sách: {e}", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_list_customers", "tien_nga_ds_khach_hang"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_customers|tien_nga_ds_khach_hang)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_list_customers_handler(client, message: Message) -> None:
    args = message.text.strip().split()

    db = SessionLocal()
    try:
        from app.models.business import Customers, CollectionPoint
        import tempfile
        import os
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Query all active customers with their collection point info
        customers = (
            db.query(Customers, CollectionPoint.collection_name)
            .outerjoin(CollectionPoint, Customers.collection_point_id == CollectionPoint.id)
            .filter(Customers.status == "ACTIVE")
            .order_by(CollectionPoint.collection_name, Customers.id)
            .all()
        )

        if not customers:
            await message.reply_text(
                "⚠️ Không có khách hàng nào trong hệ thống.",
                parse_mode=ParseMode.HTML,
            )
            return

        # Group customers by collection point
        grouped = {}
        for cust, cp_name in customers:
            group_key = cp_name or "Chưa phân loại"
            if group_key not in grouped:
                grouped[group_key] = []
            grouped[group_key].append(cust)

        # Build Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        money_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        alt_fill = PatternFill("solid", fgColor="D9E2F3")

        headers = [
            "STT",
            "Mã Hộ",
            "Tên Khách Hàng",
            "Trợ Giá",
            "Số Điện Thoại",
            "Địa Chỉ",
            "Nguyên Liệu",
            "Số Tiền Nợ",
            "Ứng Tiền Cuối Mùa",
            "Tổng Công Nợ",
            "Username TG",
            "Trạng Thái",
        ]

        col_widths = [6, 10, 22, 10, 16, 25, 15, 18, 20, 18, 18, 12]

        total_customers = 0

        for group_name, custs in grouped.items():
            # Excel sheet name max 31 chars
            sheet_name = group_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Write data rows
            def fmt_money(val):
                if val is None or val == 0:
                    return "0"
                return f"{int(val):,}".replace(",", ".")

            for idx, cust in enumerate(custs, 1):
                row = idx + 1
                row_fill = alt_fill if idx % 2 == 0 else None

                values = [
                    idx,
                    cust.hoursehold_id or cust.id,
                    cust.fullname,
                    cust.is_subsidized or 0,
                    cust.number_phone,
                    cust.address,
                    cust.ingredient,
                    fmt_money(cust.amount_of_debt),
                    fmt_money(cust.cash_advance),
                    fmt_money(cust.total_debt),
                    cust.username,
                    cust.status,
                ]

                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx in (8, 9, 10):  # Money columns
                        cell.alignment = money_align
                    elif col_idx == 1:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                    if row_fill:
                        cell.fill = row_fill

            total_customers += len(custs)

            # Set column widths
            for col_idx, width in enumerate(col_widths, 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width

            # Freeze header row
            ws.freeze_panes = "A2"

        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        now_str = datetime.now().strftime("%Y%m%d_%H%M")
        summary_lines = "\n".join(
            f"  • <b>{name}</b>: {len(custs)} KH" for name, custs in grouped.items()
        )

        await message.reply_document(
            document=tmp_path,
            file_name=f"danh_sach_khach_hang_{now_str}.xlsx",
            caption=(
                f"<b>DANH SÁCH KHÁCH HÀNG</b>\n\n"
                f"Tổng: <b>{total_customers}</b> khách hàng\n"
                f"Số Xưởng: <b>{len(grouped)}</b>\n\n"
                f"{summary_lines}\n\n"
                f"<i>Xuất lúc: {datetime.now().strftime('%H:%M %d/%m/%Y')}</i>"
            ),
            parse_mode=ParseMode.HTML,
        )

        os.remove(tmp_path)
        LogInfo(
            f"[TienNga] Exported {total_customers} customers by @{message.from_user.username or message.from_user.id}",
            LogType.SYSTEM_STATUS,
        )

    except Exception as e:
        LogError(f"Error in tien_nga_list_customers_handler: {e}", LogType.SYSTEM_STATUS)
        import traceback
        traceback.print_exc()
        await message.reply_text(
            f"❌ Có lỗi xảy ra khi xuất danh sách: {e}",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


def parse_float_vn(val_str: str) -> float:
    """Parse số tiền kiểu Việt Nam (1.000.000,5 hoặc -3.000.000)"""
    if not val_str:
        return 0.0
    v = val_str.strip()
    
    negative = v.startswith("-")
    v = v.lstrip("-").strip()
    
    # Remove currency symbols
    v = re.sub(r'[^0-9.,-]', '', v)
    if not v:
        return 0.0
    
    # Vietnamese format: 1.000.000,5 -> dots are thousand separators, comma is decimal
    if "," in v and "." in v:
        # 5.898.728,5 -> remove dots, replace comma with dot
        v = v.replace(".", "").replace(",", ".")
    elif "," in v:
        # Could be 1,5 (decimal) or 1,000 (thousands)
        parts = v.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            v = v.replace(",", ".")
        else:
            v = v.replace(",", "")
    elif "." in v:
        # Could be 1.000.000 (VN thousands) or 1.5 (decimal)
        parts = v.split(".")
        if all(len(p) == 3 for p in parts[1:]):
            v = v.replace(".", "")
        # else keep as is (decimal dot)
    
    try:
        result = float(v)
        return -result if negative else result
    except ValueError:
        return 0.0

@bot.on_message(filters.command(["tien_nga_daily_purchase", "tien_nga_thu_mua_hang_ngay"]) | filters.regex(r"^@\w+\s+/(tien_nga_daily_purchase|tien_nga_thu_mua_hang_ngay)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_daily_purchase_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    
    # Nếu chỉ có 1 dòng (lệnh + có thể mã hộ)
    if len(lines) < 2:
        args = lines[0].split()
        if len(args) < 2:
            await message.reply_text(
                "⚠️ Cú pháp: <code>/tien_nga_daily_purchase [Mã Hộ]</code>\n\n"
                "<i>Ví dụ: <code>/tien_nga_daily_purchase X051</code></i>",
                parse_mode=ParseMode.HTML
            )
            return
            
        hoursehold_id = args[1].upper()
        
        from app.db.session import SessionLocal
        from app.models.business import Customers, CollectionPoint
        
        db = SessionLocal()
        try:
            customer = db.query(Customers).filter(Customers.hoursehold_id == hoursehold_id).first()
            if not customer:
                await message.reply_text(
                    f"⚠️ Không tìm thấy Khách hàng với mã hộ <b>{hoursehold_id}</b>.",
                    parse_mode=ParseMode.HTML
                )
                return
            
            cp_name = ""
            cp_id_str = ""
            if customer.collection_point_id:
                cp = db.query(CollectionPoint).filter(CollectionPoint.id == customer.collection_point_id).first()
                if cp:
                    cp_name = cp.collection_name
                    cp_id_str = str(cp.id)
            
            today_str = datetime.now().strftime("%d/%m/%Y")
            tro_gia = customer.is_subsidized or 0
            
            form_template = f"""<b>📋 FORM NHẬP MUA MỦ NGÀY</b>
Khách hàng: <b>{customer.fullname}</b> (<code>{hoursehold_id}</code>)
Xưởng: <b>{cp_name or 'Chưa có'}</b>

Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/tien_nga_daily_purchase
Mã Hộ: {hoursehold_id}
Điểm Thu Mua: {cp_id_str}
Ngày: {today_str}
Trợ Giá (VNĐ): {tro_gia}
Khối Lượng (Kg): 
Trừ Bì (Kg): 0
Số Độ (%): 
Đơn Giá (VNĐ): 
Có Lưu Sổ: yes</pre>

<i>Ghi chú: Số tiền ví dụ 424.080 hoặc 5.898.728,5
Có Lưu Sổ: yes (lưu sổ) hoặc no (thanh toán)</i>"""
            await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error fetching customer for daily purchase: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    # Parse form data (nhiều dòng)
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    hoursehold_id = data.get("Mã Hộ", "").strip().upper()
    if not hoursehold_id:
        await message.reply_text("⚠️ <b>Mã Hộ</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    cp_id = data.get("Điểm Thu Mua", "").strip()
    ngay_str = data.get("Ngày", "").strip()
    
    if not ngay_str:
        await message.reply_text("⚠️ <b>Ngày</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    
    try:
        ngay = datetime.strptime(ngay_str, "%d/%m/%Y").date()
    except:
        await message.reply_text(
            "⚠️ Định dạng <b>Ngày</b> không hợp lệ. Vui lòng nhập DD/MM/YYYY.",
            parse_mode=ParseMode.HTML
        )
        return
    
    # Tuần tự động từ ngày
    week_val = ngay.isocalendar()[1]
    
    is_subsidized = int(parse_float_vn(data.get("Trợ Giá (VNĐ)", data.get("Trợ Giá", "0"))))
    weight = parse_float_vn(data.get("Khối Lượng (Kg)", data.get("Khối Lượng", "0")))
    tare_weight = parse_float_vn(data.get("Trừ Bì (Kg)", data.get("Trừ Bì", "0")))
    degree = parse_float_vn(data.get("Số Độ (%)", data.get("Số Độ", "0")))
    unit_price = parse_float_vn(data.get("Đơn Giá (VNĐ)", data.get("Đơn Giá", "0")))
    co_luu_so = data.get("Có Lưu Sổ", "yes").strip().lower()

    # Validate required fields
    if weight <= 0:
        await message.reply_text("⚠️ <b>Khối Lượng</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return
    if degree <= 0:
        await message.reply_text("⚠️ <b>Số Độ</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return
    if unit_price <= 0:
        await message.reply_text("⚠️ <b>Đơn Giá</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return

    # Auto-calculate
    actual_weight = weight - tare_weight
    subsidy_price = unit_price + is_subsidized
    dry_rubber = round(actual_weight * degree / 100, 2)
    total_amount = round(dry_rubber * subsidy_price, 0)

    # Phân bổ theo lưu sổ
    if co_luu_so in ("yes", "y", "có", "co"):
        saved_amount = total_amount
        paid_amount = 0
    else:
        saved_amount = 0
        paid_amount = total_amount

    from app.db.session import SessionLocal
    from app.models.business import DailyPurchases, Customers, CollectionPoint
    import uuid as uuid_lib
    
    db = SessionLocal()
    try:
        # Verify customer exists
        customer = db.query(Customers).filter(Customers.hoursehold_id == hoursehold_id).first()
        if not customer:
            await message.reply_text(
                f"⚠️ Mã hộ <b>{hoursehold_id}</b> không tồn tại trong hệ thống.",
                parse_mode=ParseMode.HTML
            )
            return

        # Lấy tên xưởng
        final_cp_id = cp_id if cp_id else customer.collection_point_id
        cp_name = ""
        if final_cp_id:
            cp = db.query(CollectionPoint).filter(CollectionPoint.id == str(final_cp_id)).first()
            if cp:
                cp_name = cp.collection_name
        
        new_purchase = DailyPurchases(
            id=uuid_lib.uuid4(),
            hoursehold_id=hoursehold_id,
            collection_point_id=final_cp_id,
            week=week_val,
            day=ngay,
            is_subsidized=is_subsidized,
            weight=weight,
            tare_weight=tare_weight,
            actual_weight=actual_weight,
            degree=degree,
            dry_rubber=dry_rubber,
            unit_price=unit_price,
            subsidy_price=subsidy_price,
            total_amount=total_amount,
            paid_amount=paid_amount,
            saved_amount=saved_amount,
            advance_amount=0,
            is_checked=False
        )
        db.add(new_purchase)
        
        # Cập nhật Tổng Công Nợ cho khách hàng nếu có Lưu Sổ
        if saved_amount > 0:
            if customer.total_debt is None:
                customer.total_debt = 0
            customer.total_debt += int(saved_amount)
            
        db.commit()
        
        luu_so_label = "Có" if saved_amount > 0 else "Không"
        
        LogInfo(f"[TienNga] Created daily purchase for '{hoursehold_id}' on {ngay_str} by user {message.from_user.id}", LogType.SYSTEM_STATUS)
        
        await message.reply_text(
            f"✅ <b>NHẬP MUA MỦ THÀNH CÔNG</b>\n\n"
            f"<b>Mã Hộ:</b> {hoursehold_id}\n"
            f"<b>Tên KH:</b> {customer.fullname}\n"
            f"<b>Điểm Thu Mua:</b> {cp_name or '—'}\n"
            f"<b>Ngày:</b> {ngay_str}\n"
            f"<b>Tuần:</b> {week_val}\n"
            f"<b>Trợ Giá:</b> {is_subsidized}\n"
            f"<b>Khối Lượng:</b> {fmt_num(weight)} kg\n"
            f"<b>Trừ Bì:</b> {fmt_num(tare_weight)} kg\n"
            f"<b>KL Thực Tế:</b> {fmt_num(actual_weight)} kg\n"
            f"<b>Số Độ:</b> {fmt_num(degree)}%\n"
            f"<b>Mủ Khô:</b> {fmt_num(dry_rubber)} kg\n"
            f"<b>Đơn Giá:</b> {fmt_money(unit_price)}\n"
            f"<b>Giá Hỗ Trợ:</b> {fmt_money(subsidy_price)}\n"
            f"<b>Thành Tiền:</b> <code>{fmt_money(total_amount)}</code>\n"
            f"<b>Lưu Sổ:</b> {luu_so_label} — <code>{fmt_money(saved_amount)}</code>\n"
            f"<b>Thanh Toán:</b> <code>{fmt_money(paid_amount)}</code>\n\n"
            f"🆔 {str(new_purchase.id)}\n"
            f"<i>Reply /cancel để hủy giao dịch này.</i>",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error creating daily purchase: {e}", LogType.SYSTEM_STATUS)
        db.rollback()
        await message.reply_text("❌ Có lỗi xảy ra khi lưu dữ liệu.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_export_daily_purchase", "tien_nga_xuat_bao_cao_thu_mua"]) | filters.regex(r"^@\w+\s+/(tien_nga_export_daily_purchase|tien_nga_xuat_bao_cao_thu_mua)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_export_daily_purchase_handler(client, message: Message) -> None:
    args = message.text.strip().split()

    # Validate cú pháp
    if len(args) < 3:
        await message.reply_text(
            "⚠️ <b>Cú pháp:</b>\n"
            "<code>/tien_nga_export_daily_purchase [Mã Hộ] [dd/mm/yyyy]</code>\n"
            "<code>/tien_nga_export_daily_purchase [Mã Hộ] [dd/mm/yyyy - dd/mm/yyyy]</code>\n\n"
            "<i>Ví dụ:\n"
            "<code>/tien_nga_export_daily_purchase X001 14/04/2026</code>\n"
            "<code>/tien_nga_export_daily_purchase X001 01/04/2026 - 14/04/2026</code></i>",
            parse_mode=ParseMode.HTML
        )
        return

    hoursehold_id = args[1].upper()

    # Parse date(s)
    date_part = " ".join(args[2:])
    try:
        if "-" in date_part and len(args) >= 5:
            # Range: dd/mm/yyyy - dd/mm/yyyy
            parts = date_part.split("-")
            start_date = datetime.strptime(parts[0].strip(), "%d/%m/%Y").date()
            end_date = datetime.strptime(parts[1].strip(), "%d/%m/%Y").date()
            timeframe = f"{start_date.strftime('%d/%m/%Y')} — {end_date.strftime('%d/%m/%Y')}"
        else:
            # Single date
            start_date = datetime.strptime(args[2].strip(), "%d/%m/%Y").date()
            end_date = start_date
            timeframe = start_date.strftime("%d/%m/%Y")
    except Exception:
        await message.reply_text(
            "⚠️ Định dạng ngày không hợp lệ. Vui lòng nhập <b>DD/MM/YYYY</b>.",
            parse_mode=ParseMode.HTML
        )
        return

    loading_msg = await message.reply_text("⏳ Đang tạo báo cáo, vui lòng chờ...", parse_mode=ParseMode.HTML)

    from app.models.business import DailyPurchases, Customers, CollectionPoint

    db = SessionLocal()
    try:
        # Verify customer
        customer = db.query(Customers).filter(Customers.hoursehold_id == hoursehold_id).first()
        if not customer:
            await loading_msg.delete()
            await message.reply_text(
                f"⚠️ Không tìm thấy Khách hàng với mã hộ <b>{hoursehold_id}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        # Lấy tên xưởng
        cp_name = "—"
        if customer.collection_point_id:
            cp = db.query(CollectionPoint).filter(CollectionPoint.id == customer.collection_point_id).first()
            if cp:
                cp_name = cp.collection_name

        # Query purchases
        purchases = (
            db.query(DailyPurchases)
            .filter(
                DailyPurchases.hoursehold_id == hoursehold_id,
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date
            )
            .order_by(DailyPurchases.day)
            .all()
        )

        if not purchases:
            await loading_msg.delete()
            await message.reply_text(
                f"ℹ️ Không có dữ liệu mua mủ cho <b>{customer.fullname}</b> ({hoursehold_id}) trong khoảng <b>{timeframe}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        # Build records
        records = []
        tong_kl = 0
        tong_kl_tt = 0
        tong_mu_kho = 0
        tong_thanh_tien = 0
        tong_luu_so = 0
        tong_thanh_toan = 0

        for p in purchases:
            kl = p.weight or 0
            bi = p.tare_weight or 0
            kl_tt = p.actual_weight or 0
            mu_kho = p.dry_rubber or 0
            thanh_tien = p.total_amount or 0
            luu_so = p.saved_amount or 0
            thanh_toan = p.paid_amount or 0

            tong_kl += kl
            tong_kl_tt += kl_tt
            tong_mu_kho += mu_kho
            tong_thanh_tien += thanh_tien
            tong_luu_so += luu_so
            tong_thanh_toan += thanh_toan

            records.append({
                "ngay": p.day.strftime("%d/%m") if p.day else "—",
                "tuan": p.week or "—",
                "tro_gia": p.is_subsidized or 0,
                "kl": kl,
                "bi": bi,
                "kl_tt": kl_tt,
                "so_do": p.degree or 0,
                "mu_kho": mu_kho,
                "don_gia": p.unit_price or 0,
                "gia_ht": p.subsidy_price or 0,
                "thanh_tien": thanh_tien,
                "luu_so": luu_so,
                "thanh_toan": thanh_toan,
            })

        report_data = {
            "ten_kh": customer.fullname,
            "ma_ho": hoursehold_id,
            "diem_thu_mua": cp_name,
            "timeframe": timeframe,
            "records": records,
            "tong_kl": tong_kl,
            "tong_kl_tt": tong_kl_tt,
            "tong_mu_kho": tong_mu_kho,
            "tong_thanh_tien": tong_thanh_tien,
            "tong_luu_so": tong_luu_so,
            "tong_thanh_toan": tong_thanh_toan,
        }

        from bot.utils.daily_purchase_report_generator import generate_daily_purchase_report_image
        img_buf = await generate_daily_purchase_report_image(report_data)

        caption_text = (
                f"<b>BÁO CÁO MUA MỦ</b>\n"
                f"<b>Khách hàng:</b> {customer.fullname} ({hoursehold_id})\n"
                f"<b>Xưởng:</b> {cp_name}\n"
                f"<b>Thời gian:</b> {timeframe}\n"
                f"<b>Số lần mua:</b> {len(records)}"
            )

        try:
            await message.reply_photo(
                photo=img_buf,
                caption=caption_text,
                parse_mode=ParseMode.HTML,
            )
        except Exception:
            # Fallback: send as document if photo dimensions are rejected
            img_buf.seek(0)
            await message.reply_document(
                document=img_buf,
                file_name=f"bao_cao_mua_mu_{hoursehold_id}.png",
                caption=caption_text,
                parse_mode=ParseMode.HTML,
            )

        await loading_msg.delete()
        LogInfo(f"[TienNga] Exported daily purchase report for '{hoursehold_id}' ({timeframe}) by user {message.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error exporting daily purchase: {e}", LogType.SYSTEM_STATUS)
        import traceback
        traceback.print_exc()
        await loading_msg.delete()
        await message.reply_text("❌ Có lỗi xảy ra khi tạo báo cáo.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_export_info", "tien_nga_truy_xuat_thong_tin"]) | filters.regex(r"^@\w+\s+/(tien_nga_export_info|tien_nga_truy_xuat_thong_tin)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER, CustomTitle.MAIN_HR, CustomTitle.MAIN_PRODUCT, CustomTitle.MAIN_PARTNER)
async def tien_nga_export_info_handler(client, message: Message) -> None:
    args = message.text.strip().split()
    time_preset = ""
    
    if len(args) == 3:
        try:
            start = datetime.strptime(args[1], "%d/%m/%Y").date()
            end = datetime.strptime(args[2], "%d/%m/%Y").date()
            time_preset = f"_time_c_{start.strftime('%d%m%y')}_{end.strftime('%d%m%y')}"
        except:
            await message.reply_text("⚠️ Định dạng ngày không hợp lệ. Vui lòng nhập DD/MM/YYYY.\nVí dụ: <code>/tien_nga_export_info 01/05/2025 15/05/2025</code>", parse_mode=ParseMode.HTML)
            return

    # If time_preset is set, we bypass the time-selection menu and embed the time into the module callbacks
    cb_ncc = f"tn_exp_ncc{time_preset}"
    
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Nhà cung cấp", callback_data=cb_ncc),
         InlineKeyboardButton("Kinh doanh", callback_data="tn_exp_dummy")],
        [InlineKeyboardButton("Nhân sự", callback_data="tn_exp_dummy"),
         InlineKeyboardButton("Thành phẩm", callback_data="tn_exp_dummy")],
        [InlineKeyboardButton("Đối tác", callback_data="tn_exp_dummy")]
    ])
    
    msg_text = "<b>TRUNG TÂM XUẤT BÁO CÁO Tiến Nga</b>\n\n"
    if time_preset:
        msg_text += f"🗓 Chế độ Tùy chọn ngày: <b>{args[1]} - {args[2]}</b>\n"
    
    msg_text += "Vui lòng chọn module báo cáo tương ứng:"
    
    await message.reply_text(
        msg_text,
        reply_markup=keyboard,
        parse_mode=ParseMode.HTML
    )

# --- Xử lý Callback Xuất Báo Cáo ---
@bot.on_callback_query(filters.regex(r"^tn_exp_"))
async def tien_nga_export_callback(client, callback_query: CallbackQuery):
    data = callback_query.data
    
    if data == "tn_exp_main":
        # Quay lại menu chính
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Nhà cung cấp", callback_data="tn_exp_ncc"),
             InlineKeyboardButton("Kinh doanh", callback_data="tn_exp_dummy")],
            [InlineKeyboardButton("Nhân sự", callback_data="tn_exp_dummy"),
             InlineKeyboardButton("Thành phẩm", callback_data="tn_exp_dummy")],
            [InlineKeyboardButton("Đối tác", callback_data="tn_exp_dummy")]
        ])
        await callback_query.message.edit_text(
            "<b>TRUNG TÂM XUẤT BÁO CÁO Tiến Nga</b>\n\n"
            "Vui lòng chọn module báo cáo tương ứng:",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
        
    elif data == "tn_exp_cancel":
        await callback_query.message.delete()
        
    elif data == "tn_exp_ncc":
        # Menu chọn Date Range cho Nhà Cung Cấp
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("7 ngày", callback_data="tn_exp_ncc_time_7d"),
             InlineKeyboardButton("14 ngày", callback_data="tn_exp_ncc_time_14d"),
             InlineKeyboardButton("21 ngày", callback_data="tn_exp_ncc_time_21d")],
            [InlineKeyboardButton("1 tháng", callback_data="tn_exp_ncc_time_1m"),
             InlineKeyboardButton("1 quý", callback_data="tn_exp_ncc_time_1q")],
            [InlineKeyboardButton("2 quý", callback_data="tn_exp_ncc_time_2q"),
             InlineKeyboardButton("1 năm", callback_data="tn_exp_ncc_time_1y")],
            [InlineKeyboardButton("Năm trước", callback_data="tn_exp_ncc_time_py")],
            [InlineKeyboardButton("Quay lại", callback_data="tn_exp_main"),
             InlineKeyboardButton("Hủy", callback_data="tn_exp_cancel")]
        ])
        await callback_query.message.edit_text(
            "<b>BÁO CÁO [NHÀ CUNG CẤP]</b>\n\n"
            "Vui lòng chọn chu kỳ dữ liệu. Hoặc nếu muốn khoảng thời gian tùy thích, hãy sử dụng lệnh:\n"
            "<code>/tien_nga_export_info DD/MM/YYYY DD/MM/YYYY</code>\n"
            "<i>(Ví dụ: /tien_nga_export_info 01/05/2025 10/05/2025)</i>",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
        
    elif data.startswith("tn_exp_ncc_time_"):
        # Lấy được time được chọn
        time_code = data[len("tn_exp_ncc_time_"):]
        
        # Menu chọn Loại Nhà cung cấp cho chu kỳ đó
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Nhà cung cấp Mủ", callback_data=f"tn_exp_ncc_mu_{time_code}")],
            [InlineKeyboardButton("Nhà cung cấp Củi", callback_data=f"tn_exp_ncc_cui_{time_code}")],
            [InlineKeyboardButton("Nhà cung cấp Acid", callback_data=f"tn_exp_dummy")],
            [InlineKeyboardButton("Quay lại", callback_data="tn_exp_ncc"),
             InlineKeyboardButton("Hủy", callback_data="tn_exp_cancel")]
        ])
        await callback_query.message.edit_text(
            f"<b>CHUYÊN MỤC NHÀ CUNG CẤP</b>\n\n"
            f"Bạn muốn xuất báo cáo cho nguồn cấp nào?",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
        
    elif data.startswith("tn_exp_ncc_mu_"):
        time_code = data[len("tn_exp_ncc_mu_"):]
        from app.db.session import SessionLocal
        from app.models.business import CollectionPoint
        db = SessionLocal()
        try:
            points = db.query(CollectionPoint).all()
            buttons = []
            buttons.append([InlineKeyboardButton("Báo cáo tổng", callback_data=f"tn_exp_rall_{time_code}")])
            for p in points:
                short_id = str(p.id)[:8]
                buttons.append([InlineKeyboardButton(p.collection_name, callback_data=f"tn_exp_rcp_{short_id}_{time_code}")])
            buttons.append([InlineKeyboardButton("Chưa phân xưởng", callback_data=f"tn_exp_rnon_{time_code}")])
            
            buttons.append([
                InlineKeyboardButton("Quay lại", callback_data=f"tn_exp_ncc_time_{time_code}"),
                InlineKeyboardButton("Hủy", callback_data="tn_exp_cancel")
            ])
            markup = InlineKeyboardMarkup(buttons)
            await callback_query.message.edit_text(
                "<b>CHỌN ĐIỂM THU MUA (XƯỞNG)</b>\n\nBạn muốn xem báo cáo của Xưởng nào?",
                reply_markup=markup,
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            LogError(f"Error fetching points: {e}", LogType.SYSTEM_STATUS)
            await callback_query.answer("Lỗi hệ thống.", show_alert=True)
        finally:
            db.close()

    elif data.startswith("tn_exp_rall_") or data.startswith("tn_exp_rnon_") or data.startswith("tn_exp_rcp_"):
        await callback_query.message.edit_text("⏳ <i>Đang tính toán dữ liệu, vui lòng chờ...</i>", parse_mode=ParseMode.HTML)
        
        # parse the time_code and mode
        if data.startswith("tn_exp_rall_"):
            mode = "all"
            target_id = None
            time_code = data[len("tn_exp_rall_"):]
        elif data.startswith("tn_exp_rnon_"):
            mode = "none"
            target_id = None
            time_code = data[len("tn_exp_rnon_"):]
        else:
            mode = "cp"
            # tn_exp_rcp_12345678_...
            parts = data.split("_", 4)
            target_id = parts[3]
            time_code = parts[4]
            
        today = datetime.now().date()
        start_date = today
        end_date = today
        
        if time_code == "7d":
            start_date = today - timedelta(days=7)
        elif time_code == "14d":
            start_date = today - timedelta(days=14)
        elif time_code == "21d":
            start_date = today - timedelta(days=21)
        elif time_code == "1m":
            start_date = today - timedelta(days=30)
        elif time_code == "1q":
            start_date = today - timedelta(days=90)
        elif time_code == "2q":
            start_date = today - timedelta(days=180)
        elif time_code == "1y":
            start_date = today - timedelta(days=365)
        elif time_code == "py":
            last_year = today.year - 1
            start_date = datetime(last_year, 1, 1).date()
            end_date = datetime(last_year, 12, 31).date()
        elif time_code.startswith("c_"):
            parts = time_code.split("_")
            if len(parts) == 3:
                try:
                    start_date = datetime.strptime(parts[1], "%d%m%y").date()
                    end_date = datetime.strptime(parts[2], "%d%m%y").date()
                except:
                    pass
            
        from app.db.session import SessionLocal
        from app.models.business import DailyPurchases, CollectionPoint, Customers
        from sqlalchemy import func, String
        db = SessionLocal()
        try:
            def format_vn_currency(value):
                try:
                    return f"{int(value):,} đ".replace(",", ".")
                except:
                    return "0 đ"
                    
            def format_vn_float(value):
                try:
                    return f"{value:,.1f} kg".replace(",", "X").replace(".", ",").replace("X", ".")
                except:
                    return "0.0 kg"

            # Query base filter
            base_query = db.query(DailyPurchases).filter(
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date
            )

            # Apply mode filter
            cp_entity = None
            if mode == "none":
                base_query = base_query.filter(DailyPurchases.collection_point_id == None)
            elif mode == "cp":
                # Find the actual cp to filter
                cp_entity = db.query(CollectionPoint).filter(CollectionPoint.id.cast(String).startswith(target_id)).first()
                if cp_entity:
                    base_query = base_query.filter(DailyPurchases.collection_point_id == cp_entity.id)
                else:
                    base_query = base_query.filter(DailyPurchases.collection_point_id == None)

            if mode == "all":
                # Aggregate Total Info
                total_stats = db.query(
                    func.sum(DailyPurchases.actual_weight).label('total_weight'),
                    func.sum(DailyPurchases.total_amount).label('total_amount')
                ).filter(
                    DailyPurchases.day >= start_date,
                    DailyPurchases.day <= end_date
                ).first()
                
                t_weight = total_stats.total_weight or 0.0
                t_amount = total_stats.total_amount or 0.0
                
                # Aggregate By Collection Point (Xưởng)
                cp_stats = db.query(
                    CollectionPoint.collection_name,
                    func.sum(DailyPurchases.actual_weight).label('weight'),
                    func.sum(DailyPurchases.total_amount).label('amount')
                ).outerjoin(
                    CollectionPoint, CollectionPoint.id == DailyPurchases.collection_point_id
                ).filter(
                    DailyPurchases.day >= start_date,
                    DailyPurchases.day <= end_date
                ).group_by(
                    CollectionPoint.collection_name
                ).all()

                report = f"<b>BÁO CÁO THU MUA MỦ TỔNG HỢP</b>\n"
                if time_code == "py":
                    report += f"<b>Năm trước ({start_date.year})</b>\n\n"
                else:
                    report += f"<b>Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}</b>\n\n"
                    
                report += f"<b>TỔNG CỘNG HỆ THỐNG:</b>\n"
                report += f"- Khối lượng thực: <b>{format_vn_float(t_weight)}</b>\n"
                report += f"- Thành tiền: <b>{format_vn_currency(t_amount)}</b>\n\n"
                
                if cp_stats:
                    report += f"<b>CHI TIẾT THEO TỪNG XƯỞNG:</b>\n"
                    for cp in cp_stats:
                        cp_w = cp.weight or 0.0
                        cp_a = cp.amount or 0.0
                        name = cp.collection_name or "Chưa phân xưởng"
                        report += f"<b>{name}</b>:\n"
                        report += f"  • Khối lượng: {format_vn_float(cp_w)}\n"
                        report += f"  • Thành tiền: {format_vn_currency(cp_a)}\n\n"
                else:
                    report += "<i>Không phát sinh dữ liệu mua mủ trong chu kỳ này.</i>\n"
                
                kb = InlineKeyboardMarkup([[InlineKeyboardButton("Quay lại", callback_data=f"tn_exp_ncc_mu_{time_code}")]])
                await callback_query.message.edit_text(report, parse_mode=ParseMode.HTML, reply_markup=kb)

            else:
                # MODE CP OR NONE (Report By Xưởng detailed by Hoursehold)
                target_name = cp_entity.collection_name if cp_entity else "Chưa phân xưởng"

                # Aggregate Total Info for this specific point
                total_stats = base_query.with_entities(
                    func.sum(DailyPurchases.actual_weight).label('total_weight'),
                    func.sum(DailyPurchases.total_amount).label('total_amount')
                ).first()
                
                t_weight = total_stats.total_weight or 0.0
                t_amount = total_stats.total_amount or 0.0

                # Aggregate by hoursehold_id
                hh_stats = base_query.with_entities(
                    DailyPurchases.hoursehold_id,
                    func.sum(DailyPurchases.actual_weight).label('weight'),
                    func.sum(DailyPurchases.total_amount).label('amount')
                ).group_by(
                    DailyPurchases.hoursehold_id
                ).all()

                if time_code == "py":
                    timeframe_str = f"Năm trước ({start_date.year})"
                else:
                    timeframe_str = f"Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}"
                    
                records = []
                if hh_stats:
                    for hh in hh_stats:
                        hh_id = hh.hoursehold_id or "Không rõ mã hộ"
                        hh_w = hh.weight or 0.0
                        hh_a = hh.amount or 0.0
                        
                        cust = db.query(Customers).filter(Customers.hoursehold_id == hh_id).first()
                        cust_name = cust.fullname if cust else ""
                        
                        records.append({
                            "id": hh_id,
                            "name": cust_name,
                            "weight": format_vn_float(hh_w),
                            "amount": format_vn_currency(hh_a)
                        })

                from bot.utils.report_generator import generate_report_image
                kb = InlineKeyboardMarkup([[InlineKeyboardButton("Quay lại", callback_data=f"tn_exp_ncc_mu_{time_code}")]])
                
                if records:
                    buf = await generate_report_image(
                        target_name=target_name.upper(),
                        timeframe=timeframe_str,
                        total_weight=format_vn_float(t_weight),
                        total_amount=format_vn_currency(t_amount),
                        records=records
                    )
                    cap = f"<b>BÁO CÁO THU MUA: {target_name.upper()}</b>\n{timeframe_str}\n\nTổng KL: <b>{format_vn_float(t_weight)}</b>\nTổng Tiền: <b>{format_vn_currency(t_amount)}</b>"
                    await callback_query.message.delete()
                    await client.send_document(chat_id=callback_query.message.chat.id, document=buf, caption=cap, parse_mode=ParseMode.HTML, reply_markup=kb)
                else:
                    report = f"<b>BÁO CÁO THU MUA: {target_name.upper()}</b>\n{timeframe_str}\n\n<i>Không phát sinh dữ liệu mua mủ tại xưởng này trong chu kỳ.</i>\n"
                    await callback_query.message.edit_text(report, parse_mode=ParseMode.HTML, reply_markup=kb)

        except Exception as e:
            import traceback
            LogError(f"[TienNga] Error export mu: {traceback.format_exc()}", LogType.SYSTEM_STATUS)
            await callback_query.message.edit_text("❌ Có lỗi xảy ra khi truy vấn dữ liệu.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
            
    elif data.startswith("tn_exp_ncc_cui_"):
        time_code = data[len("tn_exp_ncc_cui_"):]
        await callback_query.message.edit_text("⏳ <i>Đang tính toán dữ liệu củi, vui lòng chờ...</i>", parse_mode=ParseMode.HTML)

        today = datetime.now().date()
        start_date = today
        end_date = today

        if time_code == "7d":
            start_date = today - timedelta(days=7)
        elif time_code == "14d":
            start_date = today - timedelta(days=14)
        elif time_code == "21d":
            start_date = today - timedelta(days=21)
        elif time_code == "1m":
            start_date = today - timedelta(days=30)
        elif time_code == "1q":
            start_date = today - timedelta(days=90)
        elif time_code == "2q":
            start_date = today - timedelta(days=180)
        elif time_code == "1y":
            start_date = today - timedelta(days=365)
        elif time_code == "py":
            last_year = today.year - 1
            start_date = datetime(last_year, 1, 1).date()
            end_date = datetime(last_year, 12, 31).date()
        elif time_code.startswith("c_"):
            parts = time_code.split("_")
            if len(parts) == 3:
                try:
                    start_date = datetime.strptime(parts[1], "%d%m%y").date()
                    end_date = datetime.strptime(parts[2], "%d%m%y").date()
                except:
                    pass

        from app.models.business import FirewoodPurchases, Customers
        from sqlalchemy import func
        from app.db.session import SessionLocal
        db = SessionLocal()
        try:
            def format_vn_currency(value):
                try:
                    return f"{int(value):,} đ".replace(",", ".")
                except:
                    return "0 đ"

            # Tổng hợp theo KH
            stats = db.query(
                FirewoodPurchases.hoursehold_id,
                func.sum(FirewoodPurchases.trip_count).label('total_trips'),
                func.sum(FirewoodPurchases.firewood_weight).label('total_weight'),
                func.sum(FirewoodPurchases.total_amount).label('total_amount'),
                func.sum(FirewoodPurchases.advance_amount).label('total_advance'),
            ).filter(
                FirewoodPurchases.day >= start_date,
                FirewoodPurchases.day <= end_date
            ).group_by(
                FirewoodPurchases.hoursehold_id
            ).all()

            # Tổng cộng
            grand_total = db.query(
                func.sum(FirewoodPurchases.trip_count).label('trips'),
                func.sum(FirewoodPurchases.firewood_weight).label('weight'),
                func.sum(FirewoodPurchases.total_amount).label('amount'),
                func.sum(FirewoodPurchases.advance_amount).label('advance'),
            ).filter(
                FirewoodPurchases.day >= start_date,
                FirewoodPurchases.day <= end_date
            ).first()

            if time_code == "py":
                timeframe_str = f"Năm trước ({start_date.year})"
            else:
                timeframe_str = f"Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}"

            g_trips = grand_total.trips or 0
            g_weight = grand_total.weight or 0
            g_amount = grand_total.amount or 0
            g_advance = grand_total.advance or 0

            report = f"<b>BÁO CÁO NHÀ CUNG CẤP CỦI</b>\n"
            report += f"<b>{timeframe_str}</b>\n\n"

            report += f"<b>TỔNG CỘNG:</b>\n"
            report += f"  • Số chuyến: <b>{g_trips:,}</b>\n"
            report += f"  • Khối lượng: <b>{g_weight:,.1f}</b>\n"
            report += f"  • Thành tiền: <b>{format_vn_currency(g_amount)}</b>\n"
            report += f"  • Tạm ứng: <b>{format_vn_currency(g_advance)}</b>\n\n"

            if stats:
                report += f"<b>CHI TIẾT THEO KHÁCH HÀNG:</b>\n"
                for s in stats:
                    hh_id = s.hoursehold_id or "—"
                    cust = db.query(Customers).filter(Customers.hoursehold_id == hh_id).first()
                    cust_name = cust.fullname if cust else hh_id

                    s_trips = s.total_trips or 0
                    s_weight = s.total_weight or 0
                    s_amount = s.total_amount or 0
                    s_advance = s.total_advance or 0

                    report += f"\n<b>{cust_name}</b> (<code>{hh_id}</code>):\n"
                    report += f"  • Số chuyến: {s_trips:,}\n"
                    report += f"  • Khối lượng: {s_weight:,.1f}\n"
                    report += f"  • Thành tiền: {format_vn_currency(s_amount)}\n"
                    report += f"  • Tạm ứng: {format_vn_currency(s_advance)}\n"
            else:
                report += "<i>Không phát sinh dữ liệu mua củi trong chu kỳ này.</i>\n"

            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("Quay lại", callback_data=f"tn_exp_ncc_time_{time_code}"),
                 InlineKeyboardButton("Hủy", callback_data="tn_exp_cancel")]
            ])
            await callback_query.message.edit_text(report, parse_mode=ParseMode.HTML, reply_markup=kb)

        except Exception as e:
            import traceback
            LogError(f"[TienNga] Error export cui: {traceback.format_exc()}", LogType.SYSTEM_STATUS)
            await callback_query.message.edit_text("❌ Có lỗi xảy ra khi truy vấn dữ liệu củi.", parse_mode=ParseMode.HTML)
        finally:
            db.close()

    elif data == "tn_exp_dummy":
        await callback_query.answer("🚀 Nhánh module này sẽ được phát triển tiếp sau nhé!", show_alert=True)

### Report paid amount
@bot.on_message(filters.command(["tien_nga_paid_amount_report", "tien_nga_bao_cao_da_thanh_toan"]) | filters.regex(r"^@\w+\s+/(tien_nga_paid_amount_report|tien_nga_bao_cao_da_thanh_toan)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_paid_amount_report_handler(client, message: Message) -> None:
    """Bước 1: Hiển thị button chọn khoảng thời gian hoặc nhận ngày tùy chọn."""
    args = message.text.strip().split()
    time_preset = ""

    # /tien_nga_paid_amount MÃ_HỘ DD/MM/YYYY DD/MM/YYYY
    if len(args) == 4:
        hh_id = args[1].upper()
        try:
            start = datetime.strptime(args[2], "%d/%m/%Y").date()
            end = datetime.strptime(args[3], "%d/%m/%Y").date()
        except Exception:
            await message.reply_text(
                "⚠️ Định dạng không hợp lệ.\nVí dụ: <code>/tien_nga_paid_amount_report_report X001 01/05/2025 15/05/2025</code>",
                parse_mode=ParseMode.HTML
            )
            return

        from app.db.session import SessionLocal
        from app.models.business import DailyPurchases, Customers, CollectionPoint
        from sqlalchemy import func
        db = SessionLocal()
        try:
            customer = db.query(Customers).filter(Customers.hoursehold_id == hh_id).first()
            cust_name = customer.fullname if customer else "Không rõ"
            cp_name = ""
            if customer and customer.collection_point_id:
                cp = db.query(CollectionPoint).filter(CollectionPoint.id == customer.collection_point_id).first()
                cp_name = cp.collection_name if cp else ""

            stats = db.query(
                func.sum(DailyPurchases.paid_amount).label("paid"),
                func.sum(DailyPurchases.total_amount).label("total"),
                func.sum(DailyPurchases.saved_amount).label("saved"),
                func.sum(DailyPurchases.advance_amount).label("advance"),
                func.sum(DailyPurchases.actual_weight).label("weight"),
                func.count(DailyPurchases.id).label("cnt")
            ).filter(
                DailyPurchases.hoursehold_id == hh_id,
                DailyPurchases.day >= start,
                DailyPurchases.day <= end
            ).first()

            paid = stats.paid or 0
            total = stats.total or 0
            saved = stats.saved or 0
            advance = stats.advance or 0
            weight = stats.weight or 0
            cnt = stats.cnt or 0

            report = (
                f"<b>THỐNG KÊ ĐÃ THANH TOÁN - HỘ DÂN</b>\n"
                f"🗓 <b>{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}</b>\n\n"
                f"<b>Mã Hộ:</b> <code>{hh_id}</code>\n"
                f"<b>Tên KH:</b> {cust_name}\n"
                f"<b>Xưởng:</b> {cp_name or 'N/A'}\n"
                f"{'━' * 20}\n"
                f"<b>Số lượt mua:</b> <code>{cnt}</code>\n"
                f"<b>Tổng KL thực tế:</b> <code>{weight:,.1f} kg</code>\n"
                f"<b>Tổng thành tiền:</b> <code>{fmt_vn(total)}</code>\n"
                f"<b>Đã thanh toán:</b> <code>{fmt_vn(paid)}</code>\n"
                f"<b>Lưu sổ:</b> <code>{fmt_vn(saved)}</code>\n"
                f"<b>Tạm ứng:</b> <code>{fmt_vn(advance)}</code>\n"
                f"{'━' * 20}\n"
                f"<b>Còn lại:</b> <code>{fmt_vn(total - paid - saved - advance)}</code>"
            )

            if cnt == 0:
                report = (
                    f"<b>THỐNG KÊ ĐÃ THANH TOÁN - HỘ DÂN</b>\n"
                    f"🗓 <b>{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}</b>\n\n"
                    f"<b>Mã Hộ:</b> <code>{hh_id}</code> | {cust_name}\n\n"
                    f"<i>Không có dữ liệu mua mủ trong khoảng thời gian này.</i>"
                )

            await message.reply_text(report, parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"[PaidAmount] Error hh search: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Có lỗi xảy ra.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    if len(args) == 3:
        try:
            start = datetime.strptime(args[1], "%d/%m/%Y").date()
            end = datetime.strptime(args[2], "%d/%m/%Y").date()
            time_preset = f"c_{start.strftime('%d%m%y')}_{end.strftime('%d%m%y')}"
        except Exception:
            await message.reply_text(
                "⚠️ Định dạng ngày không hợp lệ.\nVí dụ: <code>/tien_nga_paid_amount_report 01/05/2025 15/05/2025</code>",
                parse_mode=ParseMode.HTML
            )
            return

    if time_preset:
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Tất cả (Tổng hợp)", callback_data=f"tn_pa_all_{time_preset}")],
            [InlineKeyboardButton("Từng Xưởng (Chi tiết)", callback_data=f"tn_pa_sel_{time_preset}")],
            [InlineKeyboardButton("Hủy", callback_data="tn_pa_cancel")]
        ])
        await message.reply_text(
            f"<b>THỐNG KÊ ĐÃ THANH TOÁN</b>\n"
            f"🗓 <b>{args[1]} - {args[2]}</b>\n\n"
            f"Chọn phạm vi hiển thị:",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
    else:
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Hôm nay", callback_data="tn_pa_time_0d"),
             InlineKeyboardButton("7 Ngày", callback_data="tn_pa_time_7d"),
             InlineKeyboardButton("14 Ngày", callback_data="tn_pa_time_14d")],
            [InlineKeyboardButton("21 Ngày", callback_data="tn_pa_time_21d"),
             InlineKeyboardButton("1 Tháng", callback_data="tn_pa_time_1m"),
             InlineKeyboardButton("1 Quý", callback_data="tn_pa_time_1q")],
            [InlineKeyboardButton("2 Quý", callback_data="tn_pa_time_2q"),
             InlineKeyboardButton("1 Năm", callback_data="tn_pa_time_1y"),
             InlineKeyboardButton("Năm trước", callback_data="tn_pa_time_py")],
            [InlineKeyboardButton("Hủy", callback_data="tn_pa_cancel")]
        ])
        await message.reply_text(
            "<b>THỐNG KÊ ĐÃ THANH TOÁN</b>\n\n"
            "Vui lòng chọn khoảng thời gian: \n"
            "Hoặc nhập: <code>/tien_nga_paid_amount_report DD/MM/YYYY DD/MM/YYYY</code> \n"
            "Hoặc nhập: <code>/tien_nga_paid_amount_report [MÃ_HỘ] DD/MM/YYYY DD/MM/YYYY</code>",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )


@bot.on_callback_query(filters.regex(r"^tn_pa_"))
async def tien_nga_paid_amount_callback(client, callback_query: CallbackQuery):
    data = callback_query.data

    if data == "tn_pa_cancel":
        await callback_query.message.delete()
        return

    # ── Bước 2: Đã chọn time range → hiển thị "Tất cả" / "Từng xưởng" ──
    if data.startswith("tn_pa_time_"):
        time_code = data[len("tn_pa_time_"):]
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Tất cả (Tổng hợp)", callback_data=f"tn_pa_all_{time_code}")],
            [InlineKeyboardButton("Từng Xưởng (Chi tiết)", callback_data=f"tn_pa_sel_{time_code}")],
            [InlineKeyboardButton("Quay lại", callback_data="tn_pa_back_time"),
             InlineKeyboardButton("Hủy", callback_data="tn_pa_cancel")]
        ])
        await callback_query.message.edit_text(
            "<b>THỐNG KÊ ĐÃ THANH TOÁN</b>\n\n"
            "Chọn phạm vi hiển thị:",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
        return

    # ── Quay lại chọn thời gian ──
    if data == "tn_pa_back_time":
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Hôm nay", callback_data="tn_pa_time_0d"),
             InlineKeyboardButton("7 Ngày", callback_data="tn_pa_time_7d"),
             InlineKeyboardButton("14 Ngày", callback_data="tn_pa_time_14d")],
            [InlineKeyboardButton("21 Ngày", callback_data="tn_pa_time_21d"),
             InlineKeyboardButton("1 Tháng", callback_data="tn_pa_time_1m"),
             InlineKeyboardButton("1 Quý", callback_data="tn_pa_time_1q")],
            [InlineKeyboardButton("2 Quý", callback_data="tn_pa_time_2q"),
             InlineKeyboardButton("1 Năm", callback_data="tn_pa_time_1y"),
             InlineKeyboardButton("Năm trước", callback_data="tn_pa_time_py")],
            [InlineKeyboardButton("Hủy", callback_data="tn_pa_cancel")]
        ])
        await callback_query.message.edit_text(
            "<b>THỐNG KÊ ĐÃ THANH TOÁN</b>\n\n"
            "Vui lòng chọn khoảng thời gian:",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
        return

    # ── Chọn từng xưởng: hiển thị danh sách ──
    if data.startswith("tn_pa_sel_"):
        time_code = data[len("tn_pa_sel_"):]
        from app.db.session import SessionLocal
        from app.models.business import CollectionPoint
        db = SessionLocal()
        try:
            points = db.query(CollectionPoint).all()
            buttons = []
            for p in points:
                short_id = str(p.id)[:8]
                buttons.append([InlineKeyboardButton(p.collection_name, callback_data=f"tn_pa_cp_{short_id}_{time_code}")])
            buttons.append([
                InlineKeyboardButton("Quay lại", callback_data=f"tn_pa_time_{time_code}"),
                InlineKeyboardButton("Hủy", callback_data="tn_pa_cancel")
            ])
            markup = InlineKeyboardMarkup(buttons)
            await callback_query.message.edit_text(
                "<b>CHỌN XƯỞNG</b>\n\nBạn muốn xem thống kê thanh toán của Xưởng nào?",
                reply_markup=markup,
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            LogError(f"Error listing CPs for paid_amount: {e}", LogType.SYSTEM_STATUS)
            await callback_query.answer("Lỗi hệ thống.", show_alert=True)
        finally:
            db.close()
        return

    # ── Helper: tính date range từ time_code ──
    def _parse_time_range(time_code):
        today = datetime.now().date()
        start_date = today
        end_date = today
        label = "Hôm nay"

        if time_code == "0d":
            label = "Hôm nay"
        elif time_code == "7d":
            start_date = today - timedelta(days=7)
            label = "7 ngày gần nhất"
        elif time_code == "14d":
            start_date = today - timedelta(days=14)
            label = "14 ngày gần nhất"
        elif time_code == "21d":
            start_date = today - timedelta(days=21)
            label = "21 ngày gần nhất"
        elif time_code == "1m":
            start_date = today - timedelta(days=30)
            label = "1 tháng gần nhất"
        elif time_code == "1q":
            start_date = today - timedelta(days=90)
            label = "1 quý gần nhất"
        elif time_code == "2q":
            start_date = today - timedelta(days=180)
            label = "2 quý gần nhất"
        elif time_code == "1y":
            start_date = today - timedelta(days=365)
            label = "1 năm gần nhất"
        elif time_code == "py":
            last_year = today.year - 1
            start_date = datetime(last_year, 1, 1).date()
            end_date = datetime(last_year, 12, 31).date()
            label = f"Năm trước ({last_year})"
        elif time_code.startswith("c_"):
            parts = time_code.split("_")
            if len(parts) == 3:
                try:
                    start_date = datetime.strptime(parts[1], "%d%m%y").date()
                    end_date = datetime.strptime(parts[2], "%d%m%y").date()
                    label = f"Tùy chọn ({start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')})"
                except Exception:
                    pass
        return start_date, end_date, label

    # ── Bước 3a: TẤT CẢ → Tổng hợp text ──
    if data.startswith("tn_pa_all_"):
        time_code = data[len("tn_pa_all_"):]
        start_date, end_date, time_label = _parse_time_range(time_code)

        await callback_query.message.edit_text("⏳ <i>Đang tính toán dữ liệu...</i>", parse_mode=ParseMode.HTML)

        from app.db.session import SessionLocal
        from app.models.business import DailyPurchases, CollectionPoint
        from sqlalchemy import func
        db = SessionLocal()
        try:
            # Tổng toàn hệ thống
            total = db.query(
                func.sum(DailyPurchases.paid_amount).label("paid")
            ).filter(
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date
            ).first()
            total_paid = total.paid or 0

            # Theo từng xưởng
            cp_stats = db.query(
                CollectionPoint.collection_name,
                func.sum(DailyPurchases.paid_amount).label("paid"),
                func.count(DailyPurchases.id).label("cnt")
            ).outerjoin(
                CollectionPoint, CollectionPoint.id == DailyPurchases.collection_point_id
            ).filter(
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date
            ).group_by(
                CollectionPoint.collection_name
            ).order_by(
                func.sum(DailyPurchases.paid_amount).desc()
            ).all()

            report = (
                f"<b>THỐNG KÊ ĐÃ THANH TOÁN</b>\n"
                f"<b>{time_label}</b>\n"
                f"<i>Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}</i>\n\n"
                f"<b>TỔNG ĐÃ THANH TOÁN:</b> <code>{fmt_vn(total_paid)}</code>\n"
                f"{'━' * 30}\n\n"
                f"<b>CHI TIẾT THEO TỪNG XƯỞNG:</b>\n\n"
            )

            if cp_stats:
                for idx, cp in enumerate(cp_stats, 1):
                    name = cp.collection_name or "Chưa phân xưởng"
                    paid = cp.paid or 0
                    cnt = cp.cnt or 0
                    pct = (paid / total_paid * 100) if total_paid > 0 else 0
                    report += (
                        f"<b>{idx}. {name}</b>\n"
                        f"   Đã thanh toán: <code>{fmt_vn(paid)}</code>\n"
                        f"   Tỉ lệ: <code>{pct:.1f}%</code> | Số lượt: <code>{cnt}</code>\n\n"
                    )
            else:
                report += "<i>Không có dữ liệu trong khoảng thời gian này.</i>\n"

            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("Quay lại", callback_data=f"tn_pa_time_{time_code}"),
                 InlineKeyboardButton("Hủy", callback_data="tn_pa_cancel")]
            ])
            await callback_query.message.edit_text(report, parse_mode=ParseMode.HTML, reply_markup=kb)

        except Exception as e:
            import traceback
            LogError(f"[PaidAmount] Error all: {traceback.format_exc()}", LogType.SYSTEM_STATUS)
            await callback_query.message.edit_text("❌ Có lỗi xảy ra.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    # ── Bước 3b: TỪNG XƯỞNG → Tạo ảnh báo cáo ──
    if data.startswith("tn_pa_cp_"):
        # tn_pa_cp_12345678_7d
        parts = data.split("_", 4)
        target_id = parts[3]
        time_code = parts[4]
        start_date, end_date, time_label = _parse_time_range(time_code)

        await callback_query.message.edit_text("⏳ <i>Đang tạo báo cáo thanh toán...</i>", parse_mode=ParseMode.HTML)

        from app.db.session import SessionLocal
        from app.models.business import DailyPurchases, CollectionPoint, Customers
        from sqlalchemy import func, String
        db = SessionLocal()
        try:
            def fmt_weight(val):
                try:
                    return f"{val:,.1f} kg".replace(",", "X").replace(".", ",").replace("X", ".")
                except:
                    return "0,0 kg"

            # Find collection point
            cp_entity = db.query(CollectionPoint).filter(
                CollectionPoint.id.cast(String).startswith(target_id)
            ).first()

            if not cp_entity:
                await callback_query.message.edit_text("⚠️ Không tìm thấy Xưởng.", parse_mode=ParseMode.HTML)
                return

            target_name = cp_entity.collection_name

            # Total paid for this CP
            total_stats = db.query(
                func.sum(DailyPurchases.paid_amount).label("paid"),
                func.sum(DailyPurchases.actual_weight).label("weight")
            ).filter(
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date,
                DailyPurchases.collection_point_id == cp_entity.id
            ).first()

            t_paid = total_stats.paid or 0
            t_weight = total_stats.weight or 0

            # Aggregate by household
            hh_stats = db.query(
                DailyPurchases.hoursehold_id,
                func.sum(DailyPurchases.paid_amount).label("paid"),
                func.sum(DailyPurchases.actual_weight).label("weight")
            ).filter(
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date,
                DailyPurchases.collection_point_id == cp_entity.id
            ).group_by(
                DailyPurchases.hoursehold_id
            ).order_by(
                func.sum(DailyPurchases.paid_amount).desc()
            ).all()

            timeframe_str = f"{time_label}\n{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"

            records = []
            for hh in hh_stats:
                hh_id = hh.hoursehold_id or "N/A"
                hh_paid = hh.paid or 0
                hh_w = hh.weight or 0

                cust = db.query(Customers).filter(Customers.hoursehold_id == hh_id).first()
                cust_name = cust.fullname if cust else ""

                records.append({
                    "id": hh_id,
                    "name": cust_name,
                    "weight": fmt_weight(hh_w),
                    "amount": fmt_vn(hh_paid)
                })

            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("Quay lại", callback_data=f"tn_pa_sel_{time_code}"),
                 InlineKeyboardButton("Hủy", callback_data="tn_pa_cancel")]
            ])

            if records:
                from bot.utils.paid_report_generator import generate_paid_report_image
                buf = await generate_paid_report_image(
                    target_name=target_name.upper(),
                    timeframe=timeframe_str,
                    total_weight=fmt_weight(t_weight),
                    total_paid=fmt_vn(t_paid),
                    records=records
                )

                cap = (
                    f"<b>ĐÃ THANH TOÁN: {target_name.upper()}</b>\n"
                    f"{time_label}\n\n"
                    f"Tổng KL: <b>{fmt_weight(t_weight)}</b>\n"
                    f"Tổng Đã TT: <b>{fmt_vn(t_paid)}</b>\n"
                    f"Số hộ: <b>{len(records)}</b>"
                )
                chat_id = callback_query.message.chat.id
                await callback_query.message.delete()
                await client.send_document(
                    chat_id=chat_id,
                    document=buf,
                    caption=cap,
                    parse_mode=ParseMode.HTML,
                    reply_markup=kb
                )
            else:
                await callback_query.message.edit_text(
                    f"<b>ĐÃ THANH TOÁN: {target_name.upper()}</b>\n"
                    f"{time_label}\n\n"
                    f"<i>Không phát sinh thanh toán tại xưởng này trong chu kỳ.</i>",
                    parse_mode=ParseMode.HTML,
                    reply_markup=kb
                )

        except Exception as e:
            import traceback
            LogError(f"[PaidAmount] Error cp: {traceback.format_exc()}", LogType.SYSTEM_STATUS)
            try:
                await callback_query.message.edit_text("❌ Có lỗi xảy ra.", parse_mode=ParseMode.HTML)
            except Exception:
                await client.send_message(chat_id=callback_query.message.chat.id, text="❌ Có lỗi xảy ra.")
        finally:
            db.close()
        return


# ===================== THỐNG KÊ LƯU SỔ (Save Amount) =====================

@bot.on_message(filters.command(["tien_nga_save_amount_report", "tien_nga_bao_cao_luu_so"]) | filters.regex(r"^@\w+\s+/(tien_nga_save_amount_report|tien_nga_bao_cao_luu_so)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_save_amount_report_handler(client, message: Message) -> None:
    """Bước 1: Hiển thị button chọn khoảng thời gian hoặc nhận ngày tùy chọn."""
    args = message.text.strip().split()
    time_preset = ""

    # /tien_nga_save_amount MÃ_HỘ DD/MM/YYYY DD/MM/YYYY
    if len(args) == 4:
        hh_id = args[1].upper()
        try:
            start = datetime.strptime(args[2], "%d/%m/%Y").date()
            end = datetime.strptime(args[3], "%d/%m/%Y").date()
        except Exception:
            await message.reply_text(
                "⚠️ Định dạng không hợp lệ.\nVí dụ: <code>/tien_nga_save_amount_report_report X001 01/05/2025 15/05/2025</code>",
                parse_mode=ParseMode.HTML
            )
            return

        from app.db.session import SessionLocal
        from app.models.business import DailyPurchases, Customers, CollectionPoint
        from sqlalchemy import func
        db = SessionLocal()
        try:
            customer = db.query(Customers).filter(Customers.hoursehold_id == hh_id).first()
            cust_name = customer.fullname if customer else "Không rõ"
            cp_name = ""
            if customer and customer.collection_point_id:
                cp = db.query(CollectionPoint).filter(CollectionPoint.id == customer.collection_point_id).first()
                cp_name = cp.collection_name if cp else ""

            stats = db.query(
                func.sum(DailyPurchases.saved_amount).label("saved"),
                func.sum(DailyPurchases.total_amount).label("total"),
                func.sum(DailyPurchases.paid_amount).label("paid"),
                func.sum(DailyPurchases.advance_amount).label("advance"),
                func.sum(DailyPurchases.actual_weight).label("weight"),
                func.count(DailyPurchases.id).label("cnt")
            ).filter(
                DailyPurchases.hoursehold_id == hh_id,
                DailyPurchases.day >= start,
                DailyPurchases.day <= end
            ).first()

            saved = stats.saved or 0
            total = stats.total or 0
            paid = stats.paid or 0
            advance = stats.advance or 0
            weight = stats.weight or 0
            cnt = stats.cnt or 0

            report = (
                f"<b>THỐNG KÊ LƯU SỔ - HỘ DÂN</b>\n"
                f"<b>{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}</b>\n\n"
                f"<b>Mã Hộ:</b> <code>{hh_id}</code>\n"
                f"<b>Tên KH:</b> {cust_name}\n"
                f"<b>Xưởng:</b> {cp_name or 'N/A'}\n"
                f"{'━' * 30}\n"
                f"<b>Số lượt mua:</b> <code>{cnt}</code>\n"
                f"<b>Tổng KL thực tế:</b> <code>{weight:,.1f} kg</code>\n"
                f"<b>Tổng thành tiền:</b> <code>{fmt_vn(total)}</code>\n"
                f"<b>Lưu sổ:</b> <code>{fmt_vn(saved)}</code>\n"
                f"<b>Đã thanh toán:</b> <code>{fmt_vn(paid)}</code>\n"
                f"<b>Tạm ứng:</b> <code>{fmt_vn(advance)}</code>\n"
                f"{'━' * 30}\n"
                f"<b>Còn lại:</b> <code>{fmt_vn(total - paid - saved - advance)}</code>"
            )

            if cnt == 0:
                report = (
                    f"<b>THỐNG KÊ LƯU SỔ - HỘ DÂN</b>\n"
                    f"<b>{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}</b>\n\n"
                    f"<b>Mã Hộ:</b> <code>{hh_id}</code> | {cust_name}\n\n"
                    f"<i>Không có dữ liệu trong khoảng thời gian này.</i>"
                )

            await message.reply_text(report, parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"[SaveAmount] Error hh search: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Có lỗi xảy ra.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    if len(args) == 3:
        try:
            start = datetime.strptime(args[1], "%d/%m/%Y").date()
            end = datetime.strptime(args[2], "%d/%m/%Y").date()
            time_preset = f"c_{start.strftime('%d%m%y')}_{end.strftime('%d%m%y')}"
        except Exception:
            await message.reply_text(
                "⚠️ Định dạng ngày không hợp lệ.\nVí dụ: <code>/tien_nga_save_amount_report 01/05/2025 15/05/2025</code>",
                parse_mode=ParseMode.HTML
            )
            return

    if time_preset:
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Tất cả (Tổng hợp)", callback_data=f"tn_sa_all_{time_preset}")],
            [InlineKeyboardButton("Từng Xưởng (Chi tiết)", callback_data=f"tn_sa_sel_{time_preset}")],
            [InlineKeyboardButton("Hủy", callback_data="tn_sa_cancel")]
        ])
        await message.reply_text(
            f"<b>THỐNG KÊ LƯU SỔ</b>\n"
            f"<b>{args[1]} - {args[2]}</b>\n\n"
            f"Chọn phạm vi hiển thị:",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
    else:
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Hôm nay", callback_data="tn_sa_time_0d"),
             InlineKeyboardButton("7 Ngày", callback_data="tn_sa_time_7d"),
             InlineKeyboardButton("14 Ngày", callback_data="tn_sa_time_14d")],
            [InlineKeyboardButton("21 Ngày", callback_data="tn_sa_time_21d"),
             InlineKeyboardButton("1 Tháng", callback_data="tn_sa_time_1m"),
             InlineKeyboardButton("1 Quý", callback_data="tn_sa_time_1q")],
            [InlineKeyboardButton("2 Quý", callback_data="tn_sa_time_2q"),
             InlineKeyboardButton("1 Năm", callback_data="tn_sa_time_1y"),
             InlineKeyboardButton("Năm trước", callback_data="tn_sa_time_py")],
            [InlineKeyboardButton("Hủy", callback_data="tn_sa_cancel")]
        ])
        await message.reply_text(
            "<b>THỐNG KÊ LƯU SỔ</b>\n\n"
            "Vui lòng chọn khoảng thời gian: \n"
            "Hoặc nhập: <code>/tien_nga_save_amount_report DD/MM/YYYY DD/MM/YYYY</code> \n"
            "Hoặc nhập: <code>/tien_nga_save_amount_report [MÃ_HỘ] DD/MM/YYYY DD/MM/YYYY</code>",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )


@bot.on_callback_query(filters.regex(r"^tn_sa_"))
async def tien_nga_save_amount_callback(client, callback_query: CallbackQuery):
    data = callback_query.data

    if data == "tn_sa_cancel":
        await callback_query.message.delete()
        return

    # ── Bước 2: Đã chọn time range → hiển thị "Tất cả" / "Từng xưởng" ──
    if data.startswith("tn_sa_time_"):
        time_code = data[len("tn_sa_time_"):]
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Tất cả (Tổng hợp)", callback_data=f"tn_sa_all_{time_code}")],
            [InlineKeyboardButton("Từng Xưởng (Chi tiết)", callback_data=f"tn_sa_sel_{time_code}")],
            [InlineKeyboardButton("Quay lại", callback_data="tn_sa_back_time"),
             InlineKeyboardButton("Hủy", callback_data="tn_sa_cancel")]
        ])
        await callback_query.message.edit_text(
            "<b>THỐNG KÊ LƯU SỔ</b>\n\n"
            "Chọn phạm vi hiển thị:",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
        return

    # ── Quay lại chọn thời gian ──
    if data == "tn_sa_back_time":
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Hôm nay", callback_data="tn_sa_time_0d"),
             InlineKeyboardButton("7 Ngày", callback_data="tn_sa_time_7d"),
             InlineKeyboardButton("14 Ngày", callback_data="tn_sa_time_14d")],
            [InlineKeyboardButton("21 Ngày", callback_data="tn_sa_time_21d"),
             InlineKeyboardButton("1 Tháng", callback_data="tn_sa_time_1m"),
             InlineKeyboardButton("1 Quý", callback_data="tn_sa_time_1q")],
            [InlineKeyboardButton("2 Quý", callback_data="tn_sa_time_2q"),
             InlineKeyboardButton("1 Năm", callback_data="tn_sa_time_1y"),
             InlineKeyboardButton("Năm trước", callback_data="tn_sa_time_py")],
            [InlineKeyboardButton("Hủy", callback_data="tn_sa_cancel")]
        ])
        await callback_query.message.edit_text(
            "<b>THỐNG KÊ LƯU SỔ</b>\n\n"
            "Vui lòng chọn khoảng thời gian:",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
        return

    # ── Chọn từng xưởng: hiển thị danh sách ──
    if data.startswith("tn_sa_sel_"):
        time_code = data[len("tn_sa_sel_"):]
        from app.db.session import SessionLocal
        from app.models.business import CollectionPoint
        db = SessionLocal()
        try:
            points = db.query(CollectionPoint).all()
            buttons = []
            for p in points:
                short_id = str(p.id)[:8]
                buttons.append([InlineKeyboardButton(p.collection_name, callback_data=f"tn_sa_cp_{short_id}_{time_code}")])
            buttons.append([
                InlineKeyboardButton("Quay lại", callback_data=f"tn_sa_time_{time_code}"),
                InlineKeyboardButton("Hủy", callback_data="tn_sa_cancel")
            ])
            markup = InlineKeyboardMarkup(buttons)
            await callback_query.message.edit_text(
                "<b>CHỌN XƯỞNG</b>\n\nBạn muốn xem thống kê lưu sổ của Xưởng nào?",
                reply_markup=markup,
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            LogError(f"Error listing CPs for save_amount: {e}", LogType.SYSTEM_STATUS)
            await callback_query.answer("Lỗi hệ thống.", show_alert=True)
        finally:
            db.close()
        return

    # ── Helper: tính date range từ time_code ──
    def _parse_time_range_sa(time_code):
        today = datetime.now().date()
        start_date = today
        end_date = today
        label = "Hôm nay"

        if time_code == "0d":
            label = "Hôm nay"
        elif time_code == "7d":
            start_date = today - timedelta(days=7)
            label = "7 ngày gần nhất"
        elif time_code == "14d":
            start_date = today - timedelta(days=14)
            label = "14 ngày gần nhất"
        elif time_code == "21d":
            start_date = today - timedelta(days=21)
            label = "21 ngày gần nhất"
        elif time_code == "1m":
            start_date = today - timedelta(days=30)
            label = "1 tháng gần nhất"
        elif time_code == "1q":
            start_date = today - timedelta(days=90)
            label = "1 quý gần nhất"
        elif time_code == "2q":
            start_date = today - timedelta(days=180)
            label = "2 quý gần nhất"
        elif time_code == "1y":
            start_date = today - timedelta(days=365)
            label = "1 năm gần nhất"
        elif time_code == "py":
            last_year = today.year - 1
            start_date = datetime(last_year, 1, 1).date()
            end_date = datetime(last_year, 12, 31).date()
            label = f"Năm trước ({last_year})"
        elif time_code.startswith("c_"):
            parts = time_code.split("_")
            if len(parts) == 3:
                try:
                    start_date = datetime.strptime(parts[1], "%d%m%y").date()
                    end_date = datetime.strptime(parts[2], "%d%m%y").date()
                    label = f"Tùy chọn ({start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')})"
                except Exception:
                    pass
        return start_date, end_date, label

    # ── Bước 3a: TẤT CẢ → Tổng hợp text ──
    if data.startswith("tn_sa_all_"):
        time_code = data[len("tn_sa_all_"):]
        start_date, end_date, time_label = _parse_time_range_sa(time_code)

        await callback_query.message.edit_text("⏳ <i>Đang tính toán dữ liệu...</i>", parse_mode=ParseMode.HTML)

        from app.db.session import SessionLocal
        from app.models.business import DailyPurchases, CollectionPoint
        from sqlalchemy import func
        db = SessionLocal()
        try:
            # Tổng toàn hệ thống
            total = db.query(
                func.sum(DailyPurchases.saved_amount).label("saved")
            ).filter(
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date
            ).first()
            total_saved = total.saved or 0

            # Theo từng xưởng
            cp_stats = db.query(
                CollectionPoint.collection_name,
                func.sum(DailyPurchases.saved_amount).label("saved"),
                func.count(DailyPurchases.id).label("cnt")
            ).outerjoin(
                CollectionPoint, CollectionPoint.id == DailyPurchases.collection_point_id
            ).filter(
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date
            ).group_by(
                CollectionPoint.collection_name
            ).order_by(
                func.sum(DailyPurchases.saved_amount).desc()
            ).all()

            report = (
                f"<b>THỐNG KÊ LƯU SỔ</b>\n"
                f"<b>{time_label}</b>\n"
                f"<i>Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}</i>\n\n"
                f"<b>TỔNG LƯU SỔ:</b> <code>{fmt_vn(total_saved)}</code>\n"
                f"{'━' * 30}\n\n"
                f"<b>CHI TIẾT THEO TỪNG XƯỞNG:</b>\n\n"
            )

            if cp_stats:
                for idx, cp in enumerate(cp_stats, 1):
                    name = cp.collection_name or "Chưa phân xưởng"
                    saved = cp.saved or 0
                    cnt = cp.cnt or 0
                    pct = (saved / total_saved * 100) if total_saved > 0 else 0
                    report += (
                        f"<b>{idx}. {name}</b>\n"
                        f"   Lưu sổ: <code>{fmt_vn(saved)}</code>\n"
                        f"   Tỉ lệ: <code>{pct:.1f}%</code> | Số lượt: <code>{cnt}</code>\n\n"
                    )
            else:
                report += "<i>Không có dữ liệu trong khoảng thời gian này.</i>\n"

            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("Quay lại", callback_data=f"tn_sa_time_{time_code}"),
                 InlineKeyboardButton("Hủy", callback_data="tn_sa_cancel")]
            ])
            await callback_query.message.edit_text(report, parse_mode=ParseMode.HTML, reply_markup=kb)

        except Exception as e:
            import traceback
            LogError(f"[SaveAmount] Error all: {traceback.format_exc()}", LogType.SYSTEM_STATUS)
            await callback_query.message.edit_text("❌ Có lỗi xảy ra.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    # ── Bước 3b: TỪNG XƯỞNG → Tạo ảnh báo cáo ──
    if data.startswith("tn_sa_cp_"):
        # tn_sa_cp_12345678_7d
        parts = data.split("_", 4)
        target_id = parts[3]
        time_code = parts[4]
        start_date, end_date, time_label = _parse_time_range_sa(time_code)

        await callback_query.message.edit_text("⏳ <i>Đang tạo báo cáo lưu sổ...</i>", parse_mode=ParseMode.HTML)

        from app.db.session import SessionLocal
        from app.models.business import DailyPurchases, CollectionPoint, Customers
        from sqlalchemy import func, String
        db = SessionLocal()
        try:
            # Find collection point
            cp_entity = db.query(CollectionPoint).filter(
                CollectionPoint.id.cast(String).startswith(target_id)
            ).first()

            if not cp_entity:
                await callback_query.message.edit_text("⚠️ Không tìm thấy Xưởng.", parse_mode=ParseMode.HTML)
                return

            target_name = cp_entity.collection_name

            # Total saved for this CP
            total_stats = db.query(
                func.sum(DailyPurchases.saved_amount).label("saved"),
                func.sum(DailyPurchases.actual_weight).label("weight")
            ).filter(
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date,
                DailyPurchases.collection_point_id == cp_entity.id
            ).first()

            t_saved = total_stats.saved or 0
            t_weight = total_stats.weight or 0

            # Aggregate by household
            hh_stats = db.query(
                DailyPurchases.hoursehold_id,
                func.sum(DailyPurchases.saved_amount).label("saved"),
                func.sum(DailyPurchases.actual_weight).label("weight")
            ).filter(
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date,
                DailyPurchases.collection_point_id == cp_entity.id
            ).group_by(
                DailyPurchases.hoursehold_id
            ).order_by(
                func.sum(DailyPurchases.saved_amount).desc()
            ).all()

            timeframe_str = f"{time_label}\n{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"

            records = []
            for hh in hh_stats:
                hh_id = hh.hoursehold_id or "N/A"
                hh_saved = hh.saved or 0
                hh_w = hh.weight or 0

                cust = db.query(Customers).filter(Customers.hoursehold_id == hh_id).first()
                cust_name = cust.fullname if cust else ""

                records.append({
                    "id": hh_id,
                    "name": cust_name,
                    "weight": fmt_weight(hh_w),
                    "amount": fmt_vn(hh_saved)
                })

            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("Quay lại", callback_data=f"tn_sa_sel_{time_code}"),
                 InlineKeyboardButton("Hủy", callback_data="tn_sa_cancel")]
            ])

            if records:
                from bot.utils.save_report_generator import generate_save_report_image
                buf = await generate_save_report_image(
                    target_name=target_name.upper(),
                    timeframe=timeframe_str,
                    total_weight=fmt_weight(t_weight),
                    total_saved=fmt_vn(t_saved),
                    records=records
                )

                cap = (
                    f"<b>LƯU SỔ: {target_name.upper()}</b>\n"
                    f"{time_label}\n\n"
                    f"Tổng KL: <b>{fmt_weight(t_weight)}</b>\n"
                    f"Tổng Lưu Sổ: <b>{fmt_vn(t_saved)}</b>\n"
                    f"Số hộ: <b>{len(records)}</b>"
                )
                chat_id = callback_query.message.chat.id
                await callback_query.message.delete()
                await client.send_document(
                    chat_id=chat_id,
                    document=buf,
                    caption=cap,
                    parse_mode=ParseMode.HTML,
                    reply_markup=kb
                )
            else:
                await callback_query.message.edit_text(
                    f"<b>LƯU SỔ: {target_name.upper()}</b>\n"
                    f"{time_label}\n\n"
                    f"<i>Không phát sinh lưu sổ tại xưởng này trong chu kỳ.</i>",
                    parse_mode=ParseMode.HTML,
                    reply_markup=kb
                )

        except Exception as e:
            import traceback
            LogError(f"[SaveAmount] Error cp: {traceback.format_exc()}", LogType.SYSTEM_STATUS)
            try:
                await callback_query.message.edit_text("❌ Có lỗi xảy ra.", parse_mode=ParseMode.HTML)
            except Exception:
                await client.send_message(chat_id=callback_query.message.chat.id, text="❌ Có lỗi xảy ra.")
        finally:
            db.close()
        return


# ===================== GỬI TIN NHẮN (Send Message) =====================
# In-memory state for multi-select: {user_id: {"message_id": ..., "chat_id": ..., "selected": set(), "orig_msg": Message}}
_send_msg_state = {}

@bot.on_message(filters.command(["send_message"]) | filters.regex(r"^@\w+\s+/send_message\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
# @require_custom_title(CustomTitle.SUPER_MAIN, "main")
async def send_message_handler(client, message: Message) -> None:
    """
    /send_message [Nội dung]
    Gửi tin nhắn/ảnh/video đến các nhóm thành viên theo dự án.
    Nếu reply một tin nhắn thì nội dung reply sẽ được chuyển tiếp.
    """
    # Kiểm tra nội dung
    has_content = False
    text_content = ""
    
    # Lấy args sau lệnh
    first_line = message.text.strip().split("\n")[0] if message.text else ""
    all_text = message.text or message.caption or ""
    
    # Bỏ phần lệnh, lấy nội dung
    for cmd in ["send_message"]:
        pattern = f"/{cmd}"
        if pattern in all_text:
            idx = all_text.index(pattern) + len(pattern)
            text_content = all_text[idx:].strip()
            break
    
    # Check reply message
    reply_msg = message.reply_to_message
    
    if text_content:
        has_content = True
    elif reply_msg:
        has_content = True
    elif message.photo or message.video or message.document:
        has_content = True
    
    if not has_content:
        await message.reply_text(
            "⚠️ Vui lòng nhập nội dung hoặc reply một tin nhắn.\n\n"
            "Cú pháp:\n"
            "• <code>/send_message [Nội dung văn bản]</code>\n"
            "• Gửi ảnh/video kèm caption <code>/send_message</code>\n"
            "• Reply một tin nhắn rồi gõ <code>/send_message</code>",
            parse_mode=ParseMode.HTML
        )
        return
    
    # Lưu state
    user_id = message.from_user.id
    _send_msg_state[user_id] = {
        "text": text_content,
        "orig_msg": message,
        "reply_msg": reply_msg,
        "selected": set(),
    }
    
    # Hiển thị button chọn Dự án
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("Tiến Nga", callback_data="sm_proj_tiennga")],
        [InlineKeyboardButton("Credit", callback_data="sm_proj_dummy"),
         InlineKeyboardButton("Rental", callback_data="sm_proj_dummy")],
        [InlineKeyboardButton("Ggomoonsin", callback_data="sm_proj_dummy"),
         InlineKeyboardButton("Other", callback_data="sm_proj_dummy")],
        [InlineKeyboardButton("Hủy", callback_data="sm_cancel")],
    ])
    
    preview = text_content[:100] + "..." if len(text_content) > 100 else text_content
    if reply_msg:
        preview = "[Reply tin nhắn]"
    elif message.photo:
        preview = f"[Ảnh] {preview}" if preview else "[Ảnh]"
    elif message.video:
        preview = f"[Video] {preview}" if preview else "[Video]"
    
    await message.reply_text(
        f"<b>GỬI TIN NHẮN ĐẾN NHÓM</b>\n\n"
        f"Nội dung: <i>{preview or '(không có text)'}</i>\n\n"
        f"Chọn Dự án:",
        parse_mode=ParseMode.HTML,
        reply_markup=kb
    )


@bot.on_callback_query(filters.regex(r"^sm_"))
async def send_message_callback(client, callback_query: CallbackQuery):
    data = callback_query.data
    user_id = callback_query.from_user.id
    
    if data == "sm_cancel":
        _send_msg_state.pop(user_id, None)
        await callback_query.message.edit_text("❌ Đã hủy gửi tin nhắn.", parse_mode=ParseMode.HTML)
        return
    
    state = _send_msg_state.get(user_id)
    if not state:
        await callback_query.answer("⚠️ Phiên đã hết hạn. Vui lòng gõ lại /send_message", show_alert=True)
        return
    
    # === Chọn Dự án: Tiến Nga ===
    if data == "sm_proj_tiennga":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("Gửi tất cả nhóm", callback_data="sm_tn_all")],
            [InlineKeyboardButton("Nhà cung cấp", callback_data="sm_tn_ncc")],
            [InlineKeyboardButton("Kinh doanh", callback_data="sm_tn_dummy")],
            [InlineKeyboardButton("Nhân sự", callback_data="sm_tn_dummy")],
            [InlineKeyboardButton("QL Thành phẩm", callback_data="sm_tn_dummy")],
            [InlineKeyboardButton("QL Đối tác", callback_data="sm_tn_dummy")],
            [InlineKeyboardButton("Hủy", callback_data="sm_cancel"),
             InlineKeyboardButton("Quay lại", callback_data="sm_back_proj")],
        ])
        await callback_query.message.edit_text(
            "<b>DỰ ÁN: TIẾN NGA</b>\n\nChọn nhóm nhận tin:",
            parse_mode=ParseMode.HTML, reply_markup=kb
        )
    
    # === Quay lại chọn Dự án ===
    elif data == "sm_back_proj":
        state["selected"] = set()
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("Tiến Nga", callback_data="sm_proj_tiennga")],
            [InlineKeyboardButton("Credit", callback_data="sm_proj_dummy"),
             InlineKeyboardButton("Rental", callback_data="sm_proj_dummy")],
            [InlineKeyboardButton("Ggomoonsin", callback_data="sm_proj_dummy"),
             InlineKeyboardButton("Other", callback_data="sm_proj_dummy")],
            [InlineKeyboardButton("Hủy", callback_data="sm_cancel")],
        ])
        await callback_query.message.edit_text(
            "<b>GỬI TIN NHẮN ĐẾN NHÓM</b>\n\nChọn Dự án:",
            parse_mode=ParseMode.HTML, reply_markup=kb
        )
    
    # === Nhà cung cấp -> Multi-select sub-categories ===
    elif data == "sm_tn_ncc":
        state["selected"] = set()
        await _render_ncc_select(callback_query, state)
    
    # === Toggle NCC sub-category ===
    elif data.startswith("sm_tn_ncc_tog_"):
        ncc_key = data.replace("sm_tn_ncc_tog_", "")
        selected = state.get("selected", set())
        if ncc_key in selected:
            selected.discard(ncc_key)
        else:
            selected.add(ncc_key)
        state["selected"] = selected
        await _render_ncc_select(callback_query, state)
    
    # === Gửi đến NCC đã chọn ===
    elif data == "sm_tn_ncc_send":
        selected = state.get("selected", set())
        if not selected:
            await callback_query.answer("⚠️ Vui lòng chọn ít nhất 1 nhà cung cấp!", show_alert=True)
            return
        
        await callback_query.message.edit_text("⏳ <i>Đang gửi tin nhắn...</i>", parse_mode=ParseMode.HTML)
        
        # Map NCC keys to custom_titles
        ncc_map = {
            "mu": CustomTitle.MEMBER_SUPPLIER,
            "cui": "member_ncc_cui",
            "acid": "member_ncc_acid",
        }
        
        target_titles = [ncc_map[k] for k in selected if k in ncc_map]
        sent, failed = await _do_send_to_groups(client, state, "Tiến Nga", target_titles)
        
        _send_msg_state.pop(user_id, None)
        
        ncc_labels = {"mu": "Mủ", "cui": "Củi", "acid": "Acid"}
        selected_names = ", ".join(ncc_labels.get(k, k) for k in selected)
        
        await callback_query.message.edit_text(
            f"✅ <b>ĐÃ GỬI TIN NHẮN</b>\n\n"
            f"Nhóm: <b>NCC ({selected_names})</b>\n"
            f"Gửi thành công: <b>{sent}</b> nhóm\n"
            f"Thất bại: <b>{failed}</b> nhóm",
            parse_mode=ParseMode.HTML
        )
    
    # === Gửi tất cả nhóm Tiến Nga ===
    elif data == "sm_tn_all":
        await callback_query.message.edit_text("⏳ <i>Đang gửi tin nhắn đến tất cả nhóm...</i>", parse_mode=ParseMode.HTML)
        
        sent, failed = await _do_send_to_groups(client, state, "Tiến Nga", None)
        
        _send_msg_state.pop(user_id, None)
        
        await callback_query.message.edit_text(
            f"✅ <b>ĐÃ GỬI TIN NHẮN</b>\n\n"
            f"Nhóm: <b>Tất cả nhóm Tiến Nga</b>\n"
            f"Gửi thành công: <b>{sent}</b> nhóm\n"
            f"Thất bại: <b>{failed}</b> nhóm",
            parse_mode=ParseMode.HTML
        )
    
    # === Dummy buttons ===
    elif data == "sm_tn_dummy" or data == "sm_proj_dummy":
        await callback_query.answer("🚀 Nhánh này sẽ được phát triển tiếp sau!", show_alert=True)


async def _render_ncc_select(callback_query, state):
    """Render NCC multi-select buttons."""
    selected = state.get("selected", set())
    
    ncc_items = [
        ("mu", "Nhà cung cấp Mủ"),
        ("cui", "Nhà cung cấp Củi"),
        ("acid", "Nhà cung cấp Acid"),
    ]
    
    buttons = []
    for key, label in ncc_items:
        check = "✅" if key in selected else "⬜"
        buttons.append([InlineKeyboardButton(f"{check} {label}", callback_data=f"sm_tn_ncc_tog_{key}")])
    
    buttons.append([InlineKeyboardButton("Gửi", callback_data="sm_tn_ncc_send")])
    buttons.append([
        InlineKeyboardButton("Hủy", callback_data="sm_cancel"),
        InlineKeyboardButton("Quay lại", callback_data="sm_proj_tiennga"),
    ])
    
    selected_text = ", ".join(
        {"mu": "Mủ", "cui": "Củi", "acid": "Acid"}.get(k, k) for k in selected
    ) or "Chưa chọn"
    
    kb = InlineKeyboardMarkup(buttons)
    await callback_query.message.edit_text(
        f"<b>NHÀ CUNG CẤP</b>\n\n"
        f"Đã chọn: <b>{selected_text}</b>\n\n"
        f"Nhấn để chọn/bỏ chọn, sau đó nhấn <b>Gửi</b>:",
        parse_mode=ParseMode.HTML, reply_markup=kb
    )


async def _do_send_to_groups(client, state, project_name: str, target_titles: list = None):
    """
    Gửi tin nhắn/ảnh/video đến các nhóm member.
    target_titles=None -> gửi tất cả nhóm member trong dự án.
    Returns (sent_count, failed_count).
    """
    from app.db.session import SessionLocal
    from app.models.business import Projects
    from app.models.telegram import TelegramProjectMember
    
    db = SessionLocal()
    sent = 0
    failed = 0
    
    try:
        project = db.query(Projects).filter(Projects.project_name == project_name).first()
        if not project:
            return 0, 0
        
        query = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project.id,
        )
        
        if target_titles:
            query = query.filter(TelegramProjectMember.custom_title.in_(target_titles))
        else:
            # Tất cả nhóm member (custom_title bắt đầu bằng member_)
            query = query.filter(TelegramProjectMember.custom_title.like("member_%"))
        
        members = query.all()
        chat_ids = list(set(m.chat_id for m in members))
        
        if not chat_ids:
            return 0, 0
        
        orig_msg = state.get("orig_msg")
        reply_msg = state.get("reply_msg")
        text_content = state.get("text", "")
        
        # Clean caption: remove /send_message command from original caption
        clean_caption = text_content
        if not clean_caption and orig_msg:
            raw_caption = orig_msg.caption or ""
            # Loại bỏ /send_message khỏi caption
            import re as _re
            clean_caption = _re.sub(r'/send_message\s*', '', raw_caption).strip()
        
        for chat_id in chat_ids:
            try:
                if reply_msg:
                    # Forward reply message
                    await reply_msg.forward(int(chat_id))
                elif orig_msg and orig_msg.photo:
                    await client.send_photo(
                        chat_id=int(chat_id),
                        photo=orig_msg.photo.file_id,
                        caption=clean_caption,
                        parse_mode=ParseMode.HTML
                    )
                elif orig_msg and orig_msg.video:
                    await client.send_video(
                        chat_id=int(chat_id),
                        video=orig_msg.video.file_id,
                        caption=clean_caption,
                        parse_mode=ParseMode.HTML
                    )
                elif orig_msg and orig_msg.document:
                    await client.send_document(
                        chat_id=int(chat_id),
                        document=orig_msg.document.file_id,
                        caption=clean_caption,
                        parse_mode=ParseMode.HTML
                    )
                elif text_content:
                    await client.send_message(
                        chat_id=int(chat_id),
                        text=text_content,
                        parse_mode=ParseMode.HTML
                    )
                sent += 1
            except Exception as e:
                LogError(f"[SendMsg] Failed to send to {chat_id}: {e}", LogType.SYSTEM_STATUS)
                failed += 1
    except Exception as e:
        LogError(f"[SendMsg] Error: {e}", LogType.SYSTEM_STATUS)
    finally:
        db.close()
    
    return sent, failed


#############  Củi (Firewood) #############
@bot.on_message(filters.command(["tien_nga_firewood_purcharse", "tien_nga_nhap_mua_cui"]) | filters.regex(r"^@\w+\s+/tien_nga_firewood_purcharse|/tien_nga_nhap_mua_cui\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_firewood_purcharse_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")

    # Nếu chỉ có 1 dòng (lệnh + có thể mã hộ)
    if len(lines) < 2:
        args = lines[0].split()
        if len(args) < 2:
            await message.reply_text(
                "⚠️ Cú pháp: <code>/tien_nga_firewood_purcharse [Mã Khách Hàng]</code>\n\n"
                "<i>Ví dụ: <code>/tien_nga_firewood_purcharse KH001</code></i>",
                parse_mode=ParseMode.HTML
            )
            return

        hoursehold_id = args[1].upper()

        from app.models.business import Customers

        db = SessionLocal()
        try:
            customer = db.query(Customers).filter(
                Customers.hoursehold_id == hoursehold_id,
                Customers.status == "ACTIVE"
            ).first()
            if not customer:
                await message.reply_text(
                    f"⚠️ Không tìm thấy Khách hàng với mã <b>{hoursehold_id}</b>.",
                    parse_mode=ParseMode.HTML
                )
                return

            today_str = datetime.now().strftime("%d/%m/%Y")

            form_template = f"""<b>FORM NHẬP MUA CỦI</b>
Khách hàng: <b>{customer.fullname}</b> (<code>{hoursehold_id}</code>)

Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/tien_nga_firewood_purcharse
Mã Khách Hàng: {hoursehold_id}
Ngày: {today_str}
Kích Thước Xe: 
Số Lượng Chuyến: 
Khối Lượng Củi: 
Đơn Giá: 
Tạm Ứng: 0</pre>

<i>Ghi chú:
- Thành Tiền tự tính = Khối Lượng Củi x Đơn Giá
- Số tiền ví dụ: 1.500.000 hoặc 500.000</i>"""
            await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        except Exception as e:
            LogError(f"Error fetching customer for firewood purchase: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    # Parse form data (nhiều dòng)
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    hoursehold_id = data.get("Mã Khách Hàng", "").strip().upper()
    ngay_str = data.get("Ngày", "").strip()
    vehicle_size = data.get("Kích Thước Xe", "").strip()

    if not hoursehold_id:
        await message.reply_text("⚠️ <b>Mã Khách Hàng</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    if not ngay_str:
        await message.reply_text("⚠️ <b>Ngày</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    try:
        ngay = datetime.strptime(ngay_str, "%d/%m/%Y").date()
    except:
        await message.reply_text(
            "⚠️ Định dạng <b>Ngày</b> không hợp lệ. Vui lòng nhập DD/MM/YYYY.",
            parse_mode=ParseMode.HTML
        )
        return

    trip_count = int(parse_float_vn(data.get("Số Lượng Chuyến", "0")))
    firewood_weight = parse_float_vn(data.get("Khối Lượng Củi", "0"))
    unit_price = parse_float_vn(data.get("Đơn Giá", "0"))
    advance_amount = parse_float_vn(data.get("Tạm Ứng", "0"))

    # Validate
    if not vehicle_size:
        await message.reply_text("⚠️ <b>Kích Thước Xe</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if trip_count <= 0:
        await message.reply_text("⚠️ <b>Số Lượng Chuyến</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return
    if firewood_weight <= 0:
        await message.reply_text("⚠️ <b>Khối Lượng Củi</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return
    if unit_price <= 0:
        await message.reply_text("⚠️ <b>Đơn Giá</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return

    # Auto-calc thành tiền
    total_amount = round(firewood_weight * unit_price, 0)

    from app.models.business import Customers, FirewoodPurchases
    import uuid as uuid_lib

    db = SessionLocal()
    try:
        # Verify customer exists
        customer = db.query(Customers).filter(
            Customers.hoursehold_id == hoursehold_id,
            Customers.status == "ACTIVE"
        ).first()
        if not customer:
            await message.reply_text(
                f"⚠️ Mã khách hàng <b>{hoursehold_id}</b> không tồn tại hoặc đã bị xóa.",
                parse_mode=ParseMode.HTML
            )
            return

        new_purchase = FirewoodPurchases(
            id=uuid_lib.uuid4(),
            day=ngay,
            hoursehold_id=hoursehold_id,
            vehicle_size=vehicle_size,
            trip_count=trip_count,
            firewood_weight=firewood_weight,
            unit_price=unit_price,
            total_amount=total_amount,
            advance_amount=advance_amount
        )
        db.add(new_purchase)
        db.commit()
        db.refresh(new_purchase)

        def fmt_money(val):
            if val is None: return "0 đ"
            try:
                return f"{int(val):,} đ".replace(",", ".")
            except:
                return str(val)

        LogInfo(
            f"[TienNga] Created firewood purchase for '{customer.fullname}' ({hoursehold_id}) on {ngay_str} by user {message.from_user.id}",
            LogType.SYSTEM_STATUS
        )

        await message.reply_text(
            f"<b>NHẬP MUA CỦI THÀNH CÔNG</b>\n\n"
            f"<b>THÔNG TIN KHÁCH HÀNG</b>\n"
            f"<b>Mã KH:</b> <code>{hoursehold_id}</code>\n"
            f"<b>Tên KH:</b> {customer.fullname}\n"
            f"<b>SĐT:</b> {customer.number_phone or '—'}\n"
            f"<b>Địa Chỉ:</b> {customer.address or '—'}\n\n"
            f"<b>CHI TIẾT MUA CỦI</b>\n"
            f"<code>{new_purchase.id}</code>\n"
            f"<b>Ngày:</b> {ngay_str}\n"
            f"<b>Kích Thước Xe:</b> {vehicle_size}\n"
            f"<b>Số Chuyến:</b> {trip_count}\n"
            f"<b>Khối Lượng Củi:</b> <code>{firewood_weight:,.1f}</code>\n"
            f"<b>Đơn Giá:</b> <code>{fmt_money(unit_price)}</code>\n"
            f"<b>Thành Tiền:</b> <code>{fmt_money(total_amount)}</code>\n"
            f"<b>Tạm Ứng:</b> <code>{fmt_money(advance_amount)}</code>",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error creating firewood purchase: {e}", LogType.SYSTEM_STATUS)
        db.rollback()
        traceback.print_exc()
        await message.reply_text("❌ Có lỗi xảy ra khi lưu vào database.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# ===================== XUẤT BÁO CÁO TỔNG HỢP (Export Summary) =====================

@bot.on_message(filters.command(["tien_nga_export_summary", "tien_nga_xuat_bao_cao_tong_hop"]) | filters.regex(r"^@\w+\s+/tien_nga_export_summary|/tien_nga_xuat_bao_cao_tong_hop\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_export_summary_handler(client, message: Message) -> None:
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("1 ngày", callback_data="tn_sum_1d"),
         InlineKeyboardButton("7 ngày", callback_data="tn_sum_7d"),
         InlineKeyboardButton("14 ngày", callback_data="tn_sum_14d")],
        [InlineKeyboardButton("21 ngày", callback_data="tn_sum_21d"),
         InlineKeyboardButton("1 tháng", callback_data="tn_sum_1m"),
         InlineKeyboardButton("3 tháng", callback_data="tn_sum_3m")],
        [InlineKeyboardButton("6 tháng", callback_data="tn_sum_6m"),
         InlineKeyboardButton("1 năm", callback_data="tn_sum_1y"),
         InlineKeyboardButton("Năm trước", callback_data="tn_sum_py")],
        [InlineKeyboardButton("Hủy", callback_data="tn_sum_cancel")]
    ])
    await message.reply_text(
        "<b>XUẤT BÁO CÁO TỔNG HỢP MUA MỦ</b>\n\n"
        "Vui lòng chọn khoảng thời gian:",
        reply_markup=keyboard,
        parse_mode=ParseMode.HTML
    )


@bot.on_callback_query(filters.regex(r"^tn_sum_"))
async def tien_nga_export_summary_callback(client, callback_query: CallbackQuery):
    data = callback_query.data

    if data == "tn_sum_cancel":
        await callback_query.message.delete()
        return

    time_code = data[len("tn_sum_"):]

    await callback_query.message.edit_text("⏳ <i>Đang tạo báo cáo tổng hợp Excel, vui lòng chờ...</i>", parse_mode=ParseMode.HTML)

    today = datetime.now().date()
    start_date = today
    end_date = today

    if time_code == "1d":
        start_date = today
    elif time_code == "7d":
        start_date = today - timedelta(days=6)
    elif time_code == "14d":
        start_date = today - timedelta(days=13)
    elif time_code == "21d":
        start_date = today - timedelta(days=20)
    elif time_code == "1m":
        start_date = today - timedelta(days=29)
    elif time_code == "3m":
        start_date = today - timedelta(days=89)
    elif time_code == "6m":
        start_date = today - timedelta(days=179)
    elif time_code == "1y":
        start_date = today - timedelta(days=364)
    elif time_code == "py":
        last_year = today.year - 1
        start_date = datetime(last_year, 1, 1).date()
        end_date = datetime(last_year, 12, 31).date()

    from app.db.session import SessionLocal
    from app.models.business import DailyPurchases, CollectionPoint
    from sqlalchemy import func
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    import tempfile
    import os

    db = SessionLocal()
    try:
        # Lấy tất cả Điểm Thu Mua
        points = db.query(CollectionPoint).order_by(CollectionPoint.collection_name).all()

        # Tất cả các điểm
        cp_list = [("ALL", "TỔNG HỢP TẤT CẢ")]
        # Thêm từng điểm
        cp_list.extend([(p.id, p.collection_name) for p in points])
        # Thêm 1 entry cho "Chưa phân xưởng"
        cp_list.append((None, "Chưa phân xưởng"))

        # ==== Styles ====
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill("solid", fgColor="FFC000")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        alt_fill = PatternFill("solid", fgColor="D9E2F3")
        number_format_vn = '#,##0'
        number_format_kg = '#,##0.0'

        headers = [
            "Ngày",
            "Tổng Số Kg",
            "Tổng Kg Trừ Bì",
            "Tổng Mủ Khô",
            "Tổng Thành Tiền",
            "Tổng Đã Thanh Toán",
            "Tổng Lưu Sổ",
        ]
        col_widths = [14, 16, 16, 16, 20, 20, 20]

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        sheets_with_data = 0

        for cp_id, cp_name in cp_list:
            # Query aggregated by day
            q = db.query(
                DailyPurchases.day,
                func.sum(DailyPurchases.weight).label('total_weight'),
                func.sum(DailyPurchases.actual_weight).label('total_tare'),
                func.sum(DailyPurchases.dry_rubber).label('total_dry'),
                func.sum(DailyPurchases.total_amount).label('total_amount'),
                func.sum(DailyPurchases.paid_amount).label('total_paid'),
                func.sum(DailyPurchases.saved_amount).label('total_saved'),
            ).filter(
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date
            )

            if cp_id == "ALL":
                pass # Không filter điểm thu mua
            elif cp_id is not None:
                q = q.filter(DailyPurchases.collection_point_id == cp_id)
            else:
                q = q.filter(DailyPurchases.collection_point_id == None)

            q = q.group_by(DailyPurchases.day).order_by(DailyPurchases.day)
            rows_data = q.all()

            if not rows_data:
                continue

            sheets_with_data += 1
            sheet_name = cp_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Header row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Data rows
            sum_weight = 0
            sum_tare = 0
            sum_dry = 0
            sum_amount = 0
            sum_paid = 0
            sum_saved = 0

            for idx, row_data in enumerate(rows_data):
                row_num = idx + 2
                row_fill = alt_fill if idx % 2 == 0 else None

                w = row_data.total_weight or 0
                t = row_data.total_tare or 0
                d = row_data.total_dry or 0
                a = row_data.total_amount or 0
                p = row_data.total_paid or 0
                s = row_data.total_saved or 0

                sum_weight += w
                sum_tare += t
                sum_dry += d
                sum_amount += a
                sum_paid += p
                sum_saved += s

                values = [
                    row_data.day.strftime("%d/%m/%Y"),
                    w, t, d, a, p, s
                ]

                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill
                    if col_idx == 1:
                        cell.alignment = center_align
                    else:
                        cell.alignment = right_align
                        if col_idx in (2, 3, 4):
                            cell.number_format = number_format_kg
                        else:
                            cell.number_format = number_format_vn

            # Total row
            total_row = len(rows_data) + 2
            total_values = ["TỔNG CỘNG", sum_weight, sum_tare, sum_dry, sum_amount, sum_paid, sum_saved]

            for col_idx, val in enumerate(total_values, 1):
                cell = ws.cell(row=total_row, column=col_idx, value=val)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = center_align
                else:
                    cell.alignment = right_align
                    if col_idx in (2, 3, 4):
                        cell.number_format = number_format_kg
                    else:
                        cell.number_format = number_format_vn

            # Column widths + freeze
            for col_idx, width in enumerate(col_widths, 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width
            ws.freeze_panes = "A2"

        if sheets_with_data == 0:
            await callback_query.message.edit_text(
                "⚠️ Không có dữ liệu mua mủ trong khoảng thời gian đã chọn.",
                parse_mode=ParseMode.HTML
            )
            return

        # Save & send
        file_name = f"summary_{today.strftime('%Y_%m_%d')}.xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        timeframe_str = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"

        await callback_query.message.delete()
        await client.send_document(
            chat_id=callback_query.message.chat.id,
            document=tmp_path,
            file_name=file_name,
            caption=(
                f"<b>BÁO CÁO TỔNG HỢP MUA MỦ</b>\n\n"
                f"<b>Thời gian:</b> {timeframe_str}\n"
                f"<b>Số xưởng:</b> {sheets_with_data} tab\n"
                f"<b>File:</b> <code>{file_name}</code>"
            ),
            parse_mode=ParseMode.HTML
        )

        try:
            os.unlink(tmp_path)
        except:
            pass

    except Exception as e:
        import traceback as tb
        LogError(f"[TienNga] Error export summary: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        await callback_query.message.edit_text("❌ Có lỗi xảy ra khi tạo báo cáo.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

# ===================== ĐỐI TÁC =====================

@bot.on_message(filters.command(["tien_nga_create_partner", "tien_nga_tao_doi_tac"]) | filters.regex(r"^@\w+\s+/tien_nga_create_partner|/tien_nga_tao_doi_tac\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_PARTNER)
async def tien_nga_create_partner_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")

    if len(lines) < 2:
        form_template = """<b>FORM TẠO ĐỐI TÁC</b>

Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/tien_nga_create_partner
Mã Đối Tác: 
Tên Đối Tác: 
Công Nợ: 0
Username TG: 
Nhóm Telegram: 
Ngân Hàng: 
Số TK: </pre>

<i>Ghi chú: Mã đối tác nên viết liền không dấu (VD: DT001).</i>"""
        await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        return

    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    partner_id = data.get("Mã Đối Tác", "").strip().upper()
    partner_name = data.get("Tên Đối Tác", "").strip()
    total_debt = parse_float_vn(data.get("Công Nợ", "0"))
    username = data.get("Username TG", "").strip()
    telegram_group = data.get("Nhóm Telegram", "").strip()
    bank_name = data.get("Ngân Hàng", "").strip()
    bank_account = data.get("Số TK", "").strip()

    if not partner_id:
        await message.reply_text("⚠️ <b>Mã Đối Tác</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not partner_name:
        await message.reply_text("⚠️ <b>Tên Đối Tác</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    from app.db.session import SessionLocal
    from app.models.business import Partners
    import uuid as uuid_lib

    db = SessionLocal()
    try:
        # Kiểm tra xem mã đã tồn tại chưa
        existing = db.query(Partners).filter(
            Partners.partner_id == partner_id,
            Partners.status == "ACTIVE"
        ).first()
        if existing:
            await message.reply_text(f"⚠️ Mã đối tác <b>{partner_id}</b> đã tồn tại trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        new_partner = Partners(
            id=uuid_lib.uuid4(),
            partner_id=partner_id,
            partner_name=partner_name,
            total_debt=total_debt,
            username=username if username else None,
            telegram_group=telegram_group if telegram_group else None,
            bank_name=bank_name if bank_name else None,
            bank_account=bank_account if bank_account else None,
            status="ACTIVE"
        )
        db.add(new_partner)
        db.commit()

        def fmt_money(val):
            if val is None or val == 0:
                return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        await message.reply_text(
            f"✅ <b>Tạo Đối Tác Thành Công!</b>\n\n"
            f"<b>Mã Đối Tác:</b> <code>{partner_id}</code>\n"
            f"<b>Tên Đối Tác:</b> {partner_name}\n"
            f"<b>Công Nợ:</b> <code>{fmt_money(total_debt)}</code>\n"
            f"<b>Username TG:</b> {username or '—'}\n"
            f"<b>Nhóm Telegram:</b> {telegram_group or '—'}\n"
            f"<b>Ngân Hàng:</b> {bank_name or '—'}\n"
            f"<b>Số TK:</b> <code>{bank_account or '—'}</code>\n"
            f"<b>Trạng Thái:</b> ACTIVE",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        import traceback as tb
        LogError(f"Error creating partner: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        db.rollback()
        await message.reply_text("❌ Có lỗi hệ thống khi lưu thông tin đối tác.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_update_partner", "tien_nga_cap_nhat_doi_tac"]) | filters.regex(r"^@\w+\s+/(tien_nga_update_partner|tien_nga_cap_nhat_doi_tac)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_PARTNER)
async def tien_nga_update_partner_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")

    if len(lines) < 2:
        args = lines[0].split()
        if len(args) < 2:
            await message.reply_text(
                "⚠️ Cú pháp: <code>/tien_nga_update_partner [Mã Đối Tác]</code>\n\n"
                "<i>Ví dụ: <code>/tien_nga_update_partner DT001</code></i>",
                parse_mode=ParseMode.HTML
            )
            return

        partner_id = args[1].upper()

        from app.db.session import SessionLocal
        from app.models.business import Partners

        db = SessionLocal()
        try:
            partner = db.query(Partners).filter(Partners.partner_id == partner_id).first()
            if not partner:
                await message.reply_text(
                    f"⚠️ Không tìm thấy Đối tác với mã <b>{partner_id}</b>.",
                    parse_mode=ParseMode.HTML
                )
                return

            def fmt_debt(val):
                if val is None or val == 0:
                    return "0"
                return f"{int(val):,}".replace(",", ".")

            form_template = (
                f"<b>FORM CẬP NHẬT ĐỐI TÁC</b>\n"
                f"Vui lòng sao chép form dưới đây, sửa thông tin và gửi lại:\n\n"
                f"<pre>/tien_nga_update_partner\n"
                f"Mã Đối Tác: {partner.partner_id}\n"
                f"Tên Đối Tác: {partner.partner_name or ''}\n"
                f"Công Nợ: {fmt_debt(partner.total_debt)}\n"
                f"Username TG: {partner.username or ''}\n"
                f"Nhóm Telegram: {partner.telegram_group or ''}\n"
                f"Ngân Hàng: {partner.bank_name or ''}\n"
                f"Số TK: {partner.bank_account or ''}\n"
                f"Trạng Thái: {partner.status or 'ACTIVE'}</pre>"
            )
            await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        except Exception as e:
            import traceback as tb
            LogError(f"Error fetching partner for update: {tb.format_exc()}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    # Parse form data (nhiều dòng)
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    partner_id = data.get("Mã Đối Tác", "").strip().upper()
    if not partner_id:
        await message.reply_text("⚠️ <b>Mã Đối Tác</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    partner_name = data.get("Tên Đối Tác", "").strip()
    total_debt_str = data.get("Công Nợ", "").strip()
    username = data.get("Username TG", "").strip()
    telegram_group = data.get("Nhóm Telegram", "").strip()
    bank_name = data.get("Ngân Hàng", "").strip()
    bank_account = data.get("Số TK", "").strip()
    status = data.get("Trạng Thái", "").strip().upper()

    from app.db.session import SessionLocal
    from app.models.business import Partners

    db = SessionLocal()
    try:
        partner = db.query(Partners).filter(Partners.partner_id == partner_id).first()
        if not partner:
            await message.reply_text(
                f"⚠️ Mã đối tác <b>{partner_id}</b> không tồn tại.",
                parse_mode=ParseMode.HTML
            )
            return

        if partner_name:
            partner.partner_name = partner_name
        if total_debt_str:
            partner.total_debt = parse_float_vn(total_debt_str)
        partner.username = username if username else None
        partner.telegram_group = telegram_group if telegram_group else None
        partner.bank_name = bank_name if bank_name else None
        partner.bank_account = bank_account if bank_account else None
        if status in ("ACTIVE", "DELETED", "INACTIVE"):
            partner.status = status

        db.commit()

        def fmt_money(val):
            if val is None or val == 0:
                return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        LogInfo(f"[TienNga] Updated partner '{partner_id}' by user {message.from_user.id}", LogType.SYSTEM_STATUS)

        await message.reply_text(
            f"✅ <b>CẬP NHẬT ĐỐI TÁC THÀNH CÔNG</b>\n\n"
            f"<b>Mã Đối Tác:</b> <code>{partner.partner_id}</code>\n"
            f"<b>Tên Đối Tác:</b> {partner.partner_name}\n"
            f"<b>Công Nợ:</b> <code>{fmt_money(partner.total_debt)}</code>\n"
            f"<b>Username TG:</b> {partner.username or '—'}\n"
            f"<b>Nhóm Telegram:</b> {partner.telegram_group or '—'}\n"
            f"<b>Ngân Hàng:</b> {partner.bank_name or '—'}\n"
            f"<b>Số TK:</b> <code>{partner.bank_account or '—'}</code>\n"
            f"<b>Trạng Thái:</b> {partner.status}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        import traceback as tb
        LogError(f"Error updating partner: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        db.rollback()
        await message.reply_text("❌ Có lỗi hệ thống khi lưu thông tin đối tác.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_delete_partner", "tien_nga_xoa_doi_tac"]) | filters.regex(r"^@\w+\s+/(tien_nga_delete_partner|tien_nga_xoa_doi_tac)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_PARTNER)
async def tien_nga_delete_partner_handler(client, message: Message) -> None:
    args = message.text.split()
    if len(args) < 2:
        await message.reply_text(
            "⚠️ Cú pháp: <code>/tien_nga_delete_partner [Mã Đối Tác]</code>\n\n"
            "<i>Ví dụ: <code>/tien_nga_delete_partner DT001</code></i>",
            parse_mode=ParseMode.HTML
        )
        return

    partner_id = args[1].upper()

    from app.db.session import SessionLocal
    from app.models.business import Partners

    db = SessionLocal()
    try:
        partner = db.query(Partners).filter(Partners.partner_id == partner_id).first()
        if not partner:
            await message.reply_text(
                f"⚠️ Không tìm thấy Đối tác với mã <b>{partner_id}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        if partner.status in ("DELETED", "INACTIVE"):
            await message.reply_text(
                f"ℹ️ Đối tác <b>{partner_id}</b> đã ở trạng thái khóa / xóa từ trước.",
                parse_mode=ParseMode.HTML
            )
            return

        partner.status = "DELETED"
        db.commit()

        LogInfo(f"[TienNga] Soft deleted partner '{partner_id}' by user {message.from_user.id}", LogType.SYSTEM_STATUS)

        await message.reply_text(
            f"✅ <b>XÓA ĐỐI TÁC THÀNH CÔNG</b>\n\n"
            f"<b>Mã Đối Tác:</b> <code>{partner_id}</code>\n"
            f"<b>Tên Đối Tác:</b> {partner.partner_name}\n"
            f"<i>Đối tác này đã được ẩn (chuyển sang trạng thái DELETED).</i>",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        import traceback as tb
        LogError(f"Error soft deleting partner: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        db.rollback()
        await message.reply_text("❌ Có lỗi xảy ra khi xóa đối tác.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_list_partner", "tien_nga_ds_doi_tac"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_partner|tien_nga_ds_doi_tac)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_PARTNER)
async def tien_nga_list_partner_handler(client, message: Message) -> None:
    from app.db.session import SessionLocal
    from app.models.business import Partners

    db = SessionLocal()
    try:
        partners = db.query(Partners).filter(Partners.status == "ACTIVE").order_by(Partners.partner_id).all()

        if not partners:
            await message.reply_text(
                "⚠️ <b>Không có đối tác nào trong hệ thống.</b>",
                parse_mode=ParseMode.HTML
            )
            return

        def fmt_money(val):
            if val is None or val == 0:
                return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        # ── ≤ 10: hiển thị text ──
        if len(partners) <= 10:
            text = f"<b>DANH SÁCH ĐỐI TÁC ({len(partners)})</b>\n\n"
            for idx, p in enumerate(partners, 1):
                debt_val = p.total_debt or 0
                text += (
                    f"<b>{idx}. {p.partner_name}</b>\n"
                    f"   Mã ĐT: <code>{p.partner_id}</code>\n"
                    f"   Công Nợ: <code>{fmt_money(debt_val)}</code>\n"
                    f"   Username TG: {p.username or '—'}\n"
                    f"   Nhóm Telegram: {p.telegram_group or '—'}\n"
                    f"   Ngân Hàng: {p.bank_name or '—'}\n"
                    f"   Số TK: <code>{p.bank_account or '—'}</code>\n"
                    f"   Trạng Thái: {p.status}\n\n"
                )
            await message.reply_text(text, parse_mode=ParseMode.HTML)
            return

        # ── > 10: xuất Excel ──
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import tempfile, os

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DanhSachDoiTac"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        money_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        alt_fill = PatternFill("solid", fgColor="D9E2F3")

        headers = ["STT", "Mã Đối Tác", "Tên Đối Tác", "Công Nợ (VNĐ)", "Username TG", "Nhóm Telegram", "Ngân Hàng", "Số TK", "Trạng Thái"]

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for idx, p in enumerate(partners, 1):
            row = idx + 1
            row_fill = alt_fill if idx % 2 == 0 else None
            debt_val = p.total_debt or 0

            values = [
                idx,
                p.partner_id,
                p.partner_name or "—",
                int(debt_val),
                p.username or "—",
                p.telegram_group or "—",
                p.bank_name or "—",
                p.bank_account or "—",
                p.status or "ACTIVE",
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx in (1, 2, 8, 9):
                    cell.alignment = center_align
                elif col_idx == 4:
                    cell.alignment = money_align
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = left_align
                if row_fill:
                    cell.fill = row_fill

        col_widths = [6, 15, 25, 20, 18, 20, 15, 15, 14]
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        now_str = datetime.now().strftime("%Y%m%d_%H%M")
        await message.reply_document(
            document=tmp_path,
            file_name=f"danh_sach_doi_tac_{now_str}.xlsx",
            caption=(
                f"<b>DANH SÁCH ĐỐI TÁC</b>\n\n"
                f"Tổng cộng: <b>{len(partners)}</b> đối tác\n"
                f"<i>Xuất lúc: {datetime.now().strftime('%H:%M %d/%m/%Y')}</i>"
            ),
            parse_mode=ParseMode.HTML,
        )

        os.remove(tmp_path)
        LogInfo(
            f"[TienNga] Listed {len(partners)} partners by @{message.from_user.username or message.from_user.id}",
            LogType.SYSTEM_STATUS
        )

    except Exception as e:
        import traceback as tb
        LogError(f"Error listing partners: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        await message.reply_text(f"❌ Có lỗi xảy ra: {e}", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_check_transaction", "tien_nga_kiem_tra_giao_dich"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_transaction|tien_nga_kiem_tra_giao_dich)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN, UserType.MEMBER)
@require_project_name("Tiến Nga")
async def tien_nga_check_transaction_handler(client, message: Message) -> None:
    import re
    text = message.text
    dates = re.findall(r'(\d{1,2}/\d{1,2}/\d{4})', text)
    
    start_date = "ALL"
    end_date = "ALL"
    if len(dates) >= 2:
        start_date = dates[0].replace('/', '')
        end_date = dates[1].replace('/', '')
    elif len(dates) == 1:
        start_date = dates[0].replace('/', '')
        end_date = dates[0].replace('/', '')
        
    text_no_dates = re.sub(r'\d{1,2}/\d{1,2}/\d{4}', '', text).replace('-', '')
    args = text_no_dates.split()
    partner_id = None
    
    from app.db.session import SessionLocal
    from app.models.business import Partners
    from app.models.telegram import TelegramProjectMember
    
    db = SessionLocal()
    try:
        if len(args) > 1:
            partner_id = args[1].upper()
        else:
            chat_id = str(message.chat.id)
            group_member = db.query(TelegramProjectMember).filter(
                TelegramProjectMember.chat_id == chat_id
            ).first()
            if group_member and group_member.role == "member" and group_member.group_name:
                partner = db.query(Partners).filter(Partners.telegram_group == group_member.group_name).first()
                if partner:
                    partner_id = partner.partner_id
        
        if not partner_id:
            await message.reply_text("⚠️ Không xác định được mã đối tác. Vui lòng truyền mã: <code>/tien_nga_kiem_tra_giao_dich [Mã]</code>", parse_mode=ParseMode.HTML)
            return
            
        partner = db.query(Partners).filter(Partners.partner_id == partner_id).first()
        if not partner:
            await message.reply_text(f"⚠️ Không tìm thấy đối tác với mã <b>{partner_id}</b>.", parse_mode=ParseMode.HTML)
            return
            
        from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Nhập", callback_data=f"check_partner_txn:import:{partner_id}:{start_date}:{end_date}")],
            [InlineKeyboardButton("Xuất", callback_data=f"check_partner_txn:export:{partner_id}:{start_date}:{end_date}")],
            [InlineKeyboardButton("Hủy", callback_data="check_partner_txn:cancel")]
        ])
        
        date_info = f" (Từ {dates[0]} đến {dates[1]})" if len(dates) >= 2 else (f" (Ngày {dates[0]})" if len(dates) == 1 else "")
        
        await message.reply_text(
            f"<b>KIỂM TRA GIAO DỊCH ĐỐI TÁC</b>\n\n"
            f"Đối tác: <b>{partner.partner_name}</b> (<code>{partner_id}</code>)\n"
            f"Thời gian: <b>{date_info.strip(' ()') or 'Tất cả'}</b>\n"
            f"<code>/tien_nga_kiem_tra_giao_dich [dd/mm/yyyy] [dd/mm/yyyy]</code>\n\n"
            f"Vui lòng chọn loại giao dịch bạn muốn kiểm tra:",
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        import traceback as tb
        LogError(f"Error in tien_nga_check_transaction_handler: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^check_partner_txn:"))
async def check_partner_txn_callback(client, callback_query) -> None:
    data = callback_query.data.split(":")
    action = data[1]
    
    if action == "cancel":
        await callback_query.message.delete()
        return
        
    partner_id = data[2]
    start_date_str = data[3] if len(data) > 3 else "ALL"
    end_date_str = data[4] if len(data) > 4 else "ALL"
    
    txn_type = "Nhập" if action == "import" else "Xuất"
    
    from app.db.session import SessionLocal
    from app.models.business import PartnerBusinesses, Partners
    import tempfile
    import os
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    db = SessionLocal()
    try:
        partner = db.query(Partners).filter(Partners.partner_id == partner_id).first()
        if not partner:
            await callback_query.answer("⚠️ Không tìm thấy đối tác.", show_alert=True)
            return
            
        from datetime import datetime
        start_dt = None
        end_dt = None
        if start_date_str != "ALL" and end_date_str != "ALL":
            try:
                start_dt = datetime.strptime(start_date_str, "%d%m%Y").date()
                end_dt = datetime.strptime(end_date_str, "%d%m%Y").date()
            except:
                pass

        query = db.query(PartnerBusinesses).filter(PartnerBusinesses.partner_id == partner_id)
        if txn_type == "Nhập":
            query = query.filter(PartnerBusinesses.import_amount > 0)
        else:
            query = query.filter(PartnerBusinesses.export_amount > 0)
            
        if start_dt and end_dt:
            query = query.filter(PartnerBusinesses.day >= start_dt, PartnerBusinesses.day <= end_dt)
            
        txns = query.order_by(PartnerBusinesses.day.desc()).all()
        
        if not txns:
            await callback_query.answer(f"⚠️ Không có giao dịch {txn_type} nào.", show_alert=True)
            return
            
        await callback_query.answer()
        
        def fmt_money(val):
            if val is None or val == 0: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")
            
        if len(txns) <= 10:
            text = f"<b>DANH SÁCH GIAO DỊCH {txn_type.upper()}</b>\n"
            text += f"Đối tác: <b>{partner.partner_name}</b> (<code>{partner_id}</code>)\n"
            if start_dt and end_dt:
                text += f"Thời gian: {start_dt.strftime('%d/%m/%Y')} - {end_dt.strftime('%d/%m/%Y')}\n\n"
            else:
                text += "\n"
            for idx, txn in enumerate(txns, 1):
                date_str = txn.day.strftime("%d/%m/%Y") if txn.day else "N/A"
                quantity = txn.import_amount if txn_type == "Nhập" else txn.export_amount
                text += (
                    f"<b>{idx}. Ngày: {date_str}</b>\n"
                    f"   Số Lượng: {quantity} Kg\n"
                    f"   Đơn Giá: {fmt_money(txn.unit_price)}\n"
                    f"   Thành Tiền: {fmt_money(txn.total_amount)}\n"
                    f"   Mã Đơn: {txn.order_code or '—'}\n\n"
                )
            await callback_query.message.reply_text(text, parse_mode=ParseMode.HTML)
            await callback_query.message.delete()
            return
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"GiaoDich{txn_type}"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        money_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        alt_fill = PatternFill("solid", fgColor="D9E2F3")
        
        headers = ["STT", "Ngày GD", "Mã Đơn Hàng", "Số Lượng (Kg)", "Đơn Giá (VNĐ)", "Thành Tiền (VNĐ)", "Ghi Chú"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        total_qty = 0
        total_amt = 0
        
        for idx, txn in enumerate(txns, 1):
            row = idx + 1
            row_fill = alt_fill if idx % 2 == 0 else None
            date_str = txn.day.strftime("%d/%m/%Y") if txn.day else "N/A"
            quantity = txn.import_amount if txn_type == "Nhập" else txn.export_amount
            
            total_qty += (quantity or 0)
            total_amt += (txn.total_amount or 0)
            
            values = [
                idx, date_str, txn.order_code or "—",
                quantity or 0, txn.unit_price or 0, txn.total_amount or 0,
                txn.notes or "—"
            ]
            
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx in (1, 2, 3):
                    cell.alignment = center_align
                elif col_idx in (4, 5, 6):
                    cell.alignment = money_align
                    cell.number_format = '#,##0'
                if row_fill:
                    cell.fill = row_fill
                    
        summary_row = len(txns) + 2
        ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
        sum_cell = ws.cell(row=summary_row, column=1, value="TỔNG CỘNG")
        sum_cell.font = Font(bold=True)
        sum_cell.alignment = center_align
        sum_cell.border = thin_border
        
        qty_cell = ws.cell(row=summary_row, column=4, value=total_qty)
        qty_cell.font = Font(bold=True)
        qty_cell.alignment = money_align
        qty_cell.number_format = '#,##0'
        qty_cell.border = thin_border
        
        ws.cell(row=summary_row, column=5, value="").border = thin_border
        
        amt_cell = ws.cell(row=summary_row, column=6, value=total_amt)
        amt_cell.font = Font(bold=True)
        amt_cell.alignment = money_align
        amt_cell.number_format = '#,##0'
        amt_cell.border = thin_border
        
        ws.cell(row=summary_row, column=7, value="").border = thin_border
        
        col_widths = [6, 15, 15, 15, 15, 20, 25]
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
            
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)
        
        from datetime import datetime
        now_str = datetime.now().strftime("%Y%m%d_%H%M")
        
        caption_text = f"<b>GIAO DỊCH {txn_type.upper()}</b>\nĐối tác: <b>{partner.partner_name}</b>"
        if start_dt and end_dt:
            caption_text += f"\nThời gian: {start_dt.strftime('%d/%m/%Y')} - {end_dt.strftime('%d/%m/%Y')}"
            
        await callback_query.message.reply_document(
            document=tmp_path,
            file_name=f"gd_{'nhap' if action == 'import' else 'xuat'}_{partner_id}_{now_str}.xlsx",
            caption=caption_text,
            parse_mode=ParseMode.HTML
        )
        await callback_query.message.delete()
        os.remove(tmp_path)
        
    except Exception as e:
        import traceback as tb
        LogError(f"Error in check_partner_txn_callback: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        await callback_query.message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_check_debt", "tien_nga_kiem_tra_cong_no"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_debt|tien_nga_kiem_tra_cong_no)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN, UserType.MEMBER)
@require_project_name("Tiến Nga")
async def tien_nga_check_debt_handler(client, message: Message) -> None:
    args = message.text.split()
    target_id = None
    
    from app.db.session import SessionLocal
    from app.models.telegram import TelegramProjectMember
    from app.models.business import Partners, Customers
    from app.models.employee import Employee
    
    db = SessionLocal()
    try:
        if len(args) > 1:
            target_id = args[1].upper()
        else:
            chat_id = str(message.chat.id)
            group_member = db.query(TelegramProjectMember).filter(
                TelegramProjectMember.chat_id == chat_id
            ).first()
            
            if group_member and group_member.role == "member" and group_member.group_name:
                group_name = group_member.group_name
                # Check Partner
                partner = db.query(Partners).filter(Partners.telegram_group == group_name).first()
                if partner:
                    target_id = partner.partner_id
                else:
                    # Check Customer
                    customer = db.query(Customers).filter(Customers.telegram_group == group_name).first()
                    if customer:
                        target_id = customer.hoursehold_id
                    else:
                        # Check Employee
                        employee = db.query(Employee).filter(Employee.telegram_group == group_name).first()
                        if employee:
                            target_id = employee.id
                            
        if not target_id:
            await message.reply_text("⚠️ Không xác định được mã đối tượng. Vui lòng truyền tham số: <code>/tien_nga_kiem_tra_cong_no [Mã DT/KH/NV]</code>", parse_mode=ParseMode.HTML)
            return
            
        def fmt_money(val):
            if val is None or val == 0: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")
            
        if target_id.startswith("DT"):
            partner = db.query(Partners).filter(Partners.partner_id == target_id).first()
            if partner:
                await message.reply_text(
                    f"<b>THÔNG TIN CÔNG NỢ ĐỐI TÁC</b>\n\n"
                    f"<b>Mã ĐT:</b> <code>{partner.partner_id}</code>\n"
                    f"<b>Tên ĐT:</b> {partner.partner_name}\n"
                    f"<b>Công Nợ:</b> <code>{fmt_money(partner.total_debt)}</code>\n"
                    f"<b>Trạng Thái:</b> {partner.status}",
                    parse_mode=ParseMode.HTML
                )
                return
        elif target_id.startswith("KH"):
            customer = db.query(Customers).filter(Customers.hoursehold_id == target_id).first()
            if customer:
                await message.reply_text(
                    f"<b>THÔNG TIN CÔNG NỢ KHÁCH HÀNG</b>\n\n"
                    f"<b>Mã KH:</b> <code>{customer.hoursehold_id}</code>\n"
                    f"<b>Tên KH:</b> {customer.fullname}\n"
                    f"<b>Công Nợ:</b> <code>{fmt_money(customer.total_debt)}</code>\n"
                    f"<b>Trạng Thái:</b> {customer.status}",
                    parse_mode=ParseMode.HTML
                )
                return
        elif target_id.startswith("NV"):
            employee = db.query(Employee).filter(Employee.id == target_id).first()
            if employee:
                await message.reply_text(
                    f"<b>THÔNG TIN CÔNG NỢ NHÂN VIÊN</b>\n\n"
                    f"<b>Mã NV:</b> <code>{employee.id}</code>\n"
                    f"<b>Tên NV:</b> {employee.last_name} {employee.first_name}\n"
                    f"<b>Công Nợ:</b> <code>{fmt_money(employee.total_debt)}</code>\n"
                    f"<b>Trạng Thái:</b> {employee.status}",
                    parse_mode=ParseMode.HTML
                )
                return
                
        await message.reply_text(f"⚠️ Không tìm thấy thông tin công nợ cho mã <b>{target_id}</b>.", parse_mode=ParseMode.HTML)
        
    except Exception as e:
        import traceback as tb
        LogError(f"Error in tien_nga_check_debt_handler: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

# =========================================================================================
# YÊU CẦU THANH TOÁN CÔNG NỢ (MEMBER ĐỐI TÁC)
# =========================================================================================

@bot.on_message(filters.command(["tien_nga_partner_payment", "tien_nga_doi_tac_thanh_toan"]) | filters.regex(r"^@\w+\s+/(tien_nga_partner_payment|tien_nga_doi_tac_thanh_toan)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN, UserType.MEMBER)
@require_project_name("Tiến Nga")
async def tien_nga_partner_payment_handler(client, message: Message) -> None:
    args = message.text.split()
    if len(args) < 2:
        await message.reply_text(
            "⚠️ Cú pháp: <code>/tien_nga_doi_tac_thanh_toan [Số Tiền]</code>\n\n"
            "<i>(Ví dụ: <code>/tien_nga_doi_tac_thanh_toan 500000</code>)</i>",
            parse_mode=ParseMode.HTML
        )
        return

    try:
        amount_str = args[1].replace(".", "").replace(",", "")
        amount = int(amount_str)
    except ValueError:
        await message.reply_text("⚠️ <b>Số Tiền</b> phải là một số hợp lệ.", parse_mode=ParseMode.HTML)
        return

    if amount <= 0:
        await message.reply_text("⚠️ <b>Số Tiền</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return

    from app.models.business import Partners, Projects
    from app.models.telegram import TelegramProjectMember

    db = SessionLocal()
    try:
        chat_id = str(message.chat.id)
        group_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()

        if not group_member or group_member.role != "member" or not group_member.group_name:
            await message.reply_text("⚠️ Lệnh này chỉ sử dụng được trong nhóm đối tác.", parse_mode=ParseMode.HTML)
            return

        partner = db.query(Partners).filter(Partners.telegram_group == group_member.group_name).first()
        if not partner:
            await message.reply_text("⚠️ Không tìm thấy đối tác liên kết với nhóm này.", parse_mode=ParseMode.HTML)
            return

        old_debt = partner.total_debt or 0
        name = partner.partner_name or partner.partner_id

        def fmt_money(val):
            if val is None or val == 0: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        if old_debt >= 0:
            await message.reply_text(
                f"⚠️ <b>Không thể thanh toán!</b>\n\n"
                f"<b>Đối tác:</b> {name}\n"
                f"<b>Công nợ hiện tại:</b> <code>{fmt_money(old_debt)}</code>\n\n"
                f"<i>Đối tác không có công nợ cần thanh toán.</i>",
                parse_mode=ParseMode.HTML
            )
            return

        abs_debt = abs(old_debt)
        if amount > abs_debt:
            await message.reply_text(
                f"⚠️ <b>Số tiền thanh toán vượt quá công nợ!</b>\n\n"
                f"<b>Đối tác:</b> {name}\n"
                f"<b>Công nợ hiện tại:</b> <code>{fmt_money(old_debt)}</code>\n"
                f"<b>Số tiền yêu cầu:</b> <code>{fmt_money(amount)}</code>",
                parse_mode=ParseMode.HTML
            )
            return

        user_display = message.from_user.first_name or message.from_user.username or str(message.from_user.id)

        # Gửi yêu cầu xác nhận vào nhóm main
        project = db.query(Projects).filter(Projects.project_name == "Tiến Nga").first()
        if not project:
            await message.reply_text("⚠️ Không tìm thấy dự án Tiến Nga.", parse_mode=ParseMode.HTML)
            return

        main_chat = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project.id,
            TelegramProjectMember.role == "main"
        ).first()

        if not main_chat or not main_chat.chat_id:
            await message.reply_text("⚠️ Không tìm thấy nhóm chính để gửi yêu cầu.", parse_mode=ParseMode.HTML)
            return

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Xác Nhận", callback_data=f"partner_pay_confirm:{partner.partner_id}:{amount}:{chat_id}")],
            [InlineKeyboardButton("Từ Chối", callback_data=f"partner_pay_deny:{partner.partner_id}:{amount}:{chat_id}")]
        ])

        await client.send_message(
            chat_id=int(main_chat.chat_id),
            text=(
                f"🔔 <b>YÊU CẦU THANH TOÁN CÔNG NỢ ĐỐI TÁC</b>\n\n"
                f"<b>Người yêu cầu:</b> {user_display}\n"
                f"<b>Đối tác:</b> {name} (<code>{partner.partner_id}</code>)\n"
                f"<b>Số Tiền:</b> <code>{fmt_money(amount)}</code>\n"
                f"<b>Công Nợ Hiện Tại:</b> <code>{fmt_money(old_debt)}</code>\n\n"
                f"<i>Vui lòng xác nhận hoặc từ chối yêu cầu này.</i>"
            ),
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )

        await message.reply_text(
            f"✅ <b>Đã gửi yêu cầu thanh toán!</b>\n\n"
            f"<b>Số Tiền:</b> <code>{fmt_money(amount)}</code>\n"
            f"<b>Trạng Thái:</b> <i>Đang chờ xác nhận từ quản lý.</i>",
            parse_mode=ParseMode.HTML
        )

        LogInfo(f"[TienNga] Partner payment request {partner.partner_id} amount={amount} by {message.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        import traceback as tb
        LogError(f"Error in tien_nga_partner_payment: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^partner_pay_(confirm|deny):"))
async def partner_pay_callback(client, callback_query) -> None:
    data = callback_query.data.split(":")
    action = data[0].replace("partner_pay_", "")  # "confirm" or "deny"
    partner_id = data[1]
    amount = int(data[2])
    member_chat_id = data[3]

    from app.models.business import Partners, Projects
    from app.models.telegram import TelegramProjectMember

    def fmt_money(val):
        if val is None or val == 0: return "0 VNĐ"
        return f"{int(val):,} VNĐ".replace(",", ".")

    db = SessionLocal()
    try:
        partner = db.query(Partners).filter(Partners.partner_id == partner_id).first()
        if not partner:
            await callback_query.answer("⚠️ Không tìm thấy đối tác.", show_alert=True)
            return

        name = partner.partner_name or partner_id
        approver = callback_query.from_user.first_name or callback_query.from_user.username or str(callback_query.from_user.id)

        if action == "deny":
            await callback_query.answer()
            await callback_query.message.edit_text(
                f"❌ <b>TỪ CHỐI THANH TOÁN CÔNG NỢ</b>\n\n"
                f"<b>Đối tác:</b> {name} (<code>{partner_id}</code>)\n"
                f"<b>Số Tiền:</b> <code>{fmt_money(amount)}</code>\n"
                f"<b>Người xử lý:</b> {approver}",
                parse_mode=ParseMode.HTML
            )

            # Thông báo từ chối vào nhóm member
            try:
                await client.send_message(
                    chat_id=int(member_chat_id),
                    text=(
                        f"❌ <b>YÊU CẦU THANH TOÁN BỊ TỪ CHỐI</b>\n\n"
                        f"<b>Số Tiền:</b> <code>{fmt_money(amount)}</code>\n"
                        f"<b>Người xử lý:</b> {approver}\n\n"
                        f"<i>Vui lòng liên hệ quản lý để biết thêm chi tiết.</i>"
                    ),
                    parse_mode=ParseMode.HTML
                )
            except Exception as e:
                LogError(f"Error notifying member about denied payment: {e}", LogType.SYSTEM_STATUS)
            return

        # action == "confirm"
        old_debt = partner.total_debt or 0

        if amount > abs(old_debt):
            await callback_query.answer("⚠️ Số tiền vượt quá công nợ hiện tại!", show_alert=True)
            return

        new_debt = old_debt + amount  # old_debt < 0, thêm amount để giảm nợ
        partner.total_debt = new_debt
        db.commit()

        await callback_query.answer()
        
        status_note = ""
        if new_debt < 0:
            status_note = f"\n<i>Còn nợ <code>{fmt_money(new_debt)}</code>.</i>"
        elif new_debt == 0:
            status_note = f"\n<i>Đã thanh toán hết công nợ!</i>"

        await callback_query.message.edit_text(
            f"✅ <b>XÁC NHẬN THANH TOÁN CÔNG NỢ THÀNH CÔNG</b>\n\n"
            f"<b>Đối tác:</b> {name} (<code>{partner_id}</code>)\n"
            f"<b>Số Tiền:</b> <code>{fmt_money(amount)}</code>\n"
            f"<b>Công Nợ Cũ:</b> <code>{fmt_money(old_debt)}</code>\n"
            f"<b>Công Nợ Mới:</b> <code>{fmt_money(new_debt)}</code>\n"
            f"<b>Người xử lý:</b> {approver}"
            f"{status_note}",
            parse_mode=ParseMode.HTML
        )

        # Thông báo vào nhóm member
        try:
            await client.send_message(
                chat_id=int(member_chat_id),
                text=(
                    f"🔔 <b>THANH TOÁN CÔNG NỢ ĐƯỢC XÁC NHẬN</b>\n\n"
                    f"<b>Số Tiền Thanh Toán:</b> <code>{fmt_money(amount)}</code>\n"
                    f"<b>Công Nợ Hiện Tại:</b> <code>{fmt_money(new_debt)}</code>\n\n"
                    f"<i>Cảm ơn quý đối tác!</i>"
                ),
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            LogError(f"Error notifying member about confirmed payment: {e}", LogType.SYSTEM_STATUS)

        LogInfo(f"[TienNga] Partner payment confirmed {partner_id} amount={amount} by {callback_query.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        import traceback as tb
        LogError(f"Error in partner_pay_callback: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        await callback_query.message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_partner_transaction", "tien_nga_giao_dich_doi_tac"]) | filters.regex(r"^@\w+\s+/tien_nga_partner_transaction|/tien_nga_giao_dich_doi_tac\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_PARTNER)
async def tien_nga_partner_transaction_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")

    # ── Step 1: Nếu chỉ có 1 dòng → hiện buttons Nhập / Xuất / Hủy ──
    if len(lines) < 2:
        args = lines[0].split()
        if len(args) < 2:
            await message.reply_text(
                "⚠️ Cú pháp: <code>/tien_nga_partner_transaction [Mã Đối Tác]</code>\n\n"
                "<i>Ví dụ: <code>/tien_nga_partner_transaction DT001</code></i>",
                parse_mode=ParseMode.HTML
            )
            return

        partner_id = args[1].upper()

        from app.models.business import Partners
        from app.db.session import SessionLocal

        db = SessionLocal()
        try:
            partner = db.query(Partners).filter(
                Partners.partner_id == partner_id,
                Partners.status == "ACTIVE"
            ).first()
            if not partner:
                await message.reply_text(
                    f"⚠️ Không tìm thấy Đối tác với mã <b>{partner_id}</b>.",
                    parse_mode=ParseMode.HTML
                )
                return

            def fmt_money(val):
                return f"{int(val):,} VNĐ".replace(",", ".")

            debt_val = partner.total_debt or 0
            debt_str = fmt_money(abs(debt_val))
            if debt_val > 0:
                debt_display = f"{debt_str} (nợ)"
            elif debt_val < 0:
                debt_display = f"{debt_str} (dư)"
            else:
                debt_display = "0 đ"

            buttons = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("Nhập", callback_data=f"tn_ptxn_{partner_id}_Nhập"),
                    InlineKeyboardButton("Xuất", callback_data=f"tn_ptxn_{partner_id}_Xuất"),
                ],
                [InlineKeyboardButton("Hủy", callback_data="tn_ptxn_cancel")]
            ])

            await message.reply_text(
                f"<b>GIAO DỊCH ĐỐI TÁC</b>\n\n"
                f"<b>Mã Đối Tác:</b> <code>{partner_id}</code>\n"
                f"<b>Tên Đối Tác:</b> {partner.partner_name}\n"
                f"<b>Công Nợ:</b> {debt_display}\n\n"
                f"Vui lòng chọn loại giao dịch:",
                reply_markup=buttons,
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            import traceback as tb
            LogError(f"Error fetching partner for transaction: {tb.format_exc()}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    # ── Step 3: Parse form data (nhiều dòng) → lưu giao dịch ──
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    partner_id = data.get("Mã Đối Tác", "").strip().upper()
    txn_type = data.get("Loại Giao Dịch", "").strip()
    ngay_str = data.get("Ngày", "").strip()
    order_code = data.get("Mã Đơn Hàng", "").strip()
    notes = data.get("Ghi Chú", "").strip()

    if not partner_id:
        await message.reply_text("⚠️ <b>Mã Đối Tác</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    if txn_type not in ("Nhập", "Xuất"):
        await message.reply_text("⚠️ <b>Loại Giao Dịch</b> phải là Nhập hoặc Xuất.", parse_mode=ParseMode.HTML)
        return

    if not ngay_str:
        await message.reply_text("⚠️ <b>Ngày</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    try:
        ngay = datetime.strptime(ngay_str, "%d/%m/%Y").date()
    except:
        await message.reply_text(
            "⚠️ Định dạng <b>Ngày</b> không hợp lệ. Vui lòng nhập DD/MM/YYYY.",
            parse_mode=ParseMode.HTML
        )
        return

    qty = parse_float_vn(data.get("Số Lượng (Kg)", "") or data.get("Số Lượng", "0"))
    unit_price = parse_float_vn(data.get("Đơn Giá (VNĐ)", "") or data.get("Đơn Giá", "0"))

    if qty <= 0:
        await message.reply_text("⚠️ <b>Số Lượng</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return

    # Auto-calc
    total_amount = round(qty * unit_price, 0)

    # Tính import/export amount và công nợ
    if txn_type == "Nhập":
        import_amt = qty
        export_amt = 0.0
        debt_change = total_amount       # Nhập → công nợ tăng (ta nợ đối tác)
    else:
        import_amt = 0.0
        export_amt = qty
        debt_change = -total_amount      # Xuất → công nợ giảm (đối tác trả hàng/ta thu tiền)

    from app.models.business import Partners, PartnerBusinesses
    from app.db.session import SessionLocal
    import uuid as uuid_lib

    db = SessionLocal()
    try:
        partner = db.query(Partners).filter(
            Partners.partner_id == partner_id,
            Partners.status == "ACTIVE"
        ).first()
        if not partner:
            await message.reply_text(
                f"⚠️ Mã đối tác <b>{partner_id}</b> không tồn tại hoặc đã bị xóa.",
                parse_mode=ParseMode.HTML
            )
            return

        new_business = PartnerBusinesses(
            id=uuid_lib.uuid4(),
            day=ngay,
            partner_id=partner_id,
            import_amount=import_amt,
            export_amount=export_amt,
            order_code=order_code,
            unit_price=unit_price,
            total_amount=total_amount,
            notes=notes
        )
        db.add(new_business)

        # Cập nhật công nợ đối tác
        if partner.total_debt is None:
            partner.total_debt = 0
        partner.total_debt += debt_change

        db.commit()

        def fmt_money(val):
            return f"{int(val):,} VNĐ".replace(",", ".")

        debt_val = partner.total_debt or 0
        debt_str = fmt_money(debt_val)

        await message.reply_text(
            f"✅ <b>GIAO DỊCH {txn_type.upper()} THÀNH CÔNG</b>\n\n"
            f"<b>Đối tác:</b> {partner.partner_name} (<code>{partner_id}</code>)\n"
            f"<b>Loại:</b> {txn_type}\n"
            f"<b>Ngày:</b> {ngay_str}\n"
            f"<b>Số Lượng:</b> <code>{qty:,.2f} Kg</code>\n"
            f"<b>Đơn Giá:</b> <code>{fmt_money(unit_price)}</code>\n"
            f"<b>Thành Tiền:</b> <code>{fmt_money(total_amount)}</code>\n"
            f"<b>Mã Đơn Hàng:</b> {order_code or '—'}\n"
            f"<b>Ghi Chú:</b> {notes or '—'}\n\n"
            f"<b>Công Nợ Hiện Tại:</b> {debt_str}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        import traceback as tb
        LogError(f"Error creating partner business: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        db.rollback()
        await message.reply_text("❌ Có lỗi hệ thống khi lưu giao dịch.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# ── Callback: Chọn loại giao dịch Đối Tác (Nhập/Xuất) ──
@bot.on_callback_query(filters.regex(r"^tn_ptxn_(.+)$"))
async def tien_nga_partner_txn_type_callback(client, callback_query):
    action = callback_query.matches[0].group(1)

    if action == "cancel":
        await callback_query.message.edit_text("Giao dịch đã bị hủy.")
        await callback_query.answer()
        return

    parts = action.rsplit("_", 1)
    if len(parts) < 2:
        await callback_query.answer("⚠️ Dữ liệu không hợp lệ.", show_alert=True)
        return

    partner_id, txn_type = parts[0], parts[1]

    from app.db.session import SessionLocal
    from app.models.business import Partners

    db = SessionLocal()
    try:
        partner = db.query(Partners).filter(
            Partners.partner_id == partner_id,
            Partners.status == "ACTIVE"
        ).first()
        if not partner:
            await callback_query.answer("⚠️ Không tìm thấy Đối tác.", show_alert=True)
            return

        today_str = datetime.now().strftime("%d/%m/%Y")

        form = (
            f"<b>FORM GIAO DỊCH {txn_type.upper()} ĐỐI TÁC</b>\n"
            f"Đối tác: <b>{partner.partner_name}</b> (<code>{partner_id}</code>)\n\n"
            f"Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:\n\n"
            f"<pre>/tien_nga_partner_transaction\n"
            f"Mã Đối Tác: {partner_id}\n"
            f"Loại Giao Dịch: {txn_type}\n"
            f"Ngày: {today_str}\n"
            f"Số Lượng (Kg): 0\n"
            f"Đơn Giá (VNĐ): 0\n"
            f"Mã Đơn Hàng: \n"
            f"Ghi Chú: </pre>\n\n"
            f"<i>Ghi chú:\n"
            f"- Thành Tiền = Số Lượng x Đơn Giá (tự tính).\n"
            f"- Công nợ sẽ tự cập nhật sau giao dịch.\n"
            f"- Mã Đơn Hàng có thể bỏ trống.</i>"
        )
        await callback_query.message.reply_text(form, parse_mode=ParseMode.HTML)
        await callback_query.answer()
    except Exception as e:
        import traceback as tb
        LogError(f"Error in partner txn callback: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống.", show_alert=True)
    finally:
        db.close()



# ── Báo cáo tổng hợp Đối Tác (Excel) ──
@bot.on_message(filters.command(["tien_nga_partner_report", "tien_nga_bao_cao_doi_tac"]) | filters.regex(r"^@\w+\s+/(tien_nga_partner_report|tien_nga_bao_cao_doi_tac)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_PARTNER)
async def tien_nga_partner_report_handler(client, message: Message) -> None:
    args = message.text.split(maxsplit=1)
    if len(args) < 2:
        await message.reply_text(
            "⚠️ Cú pháp: <code>/tien_nga_partner_report MM/YYYY - MM/YYYY</code>\n\n"
            "<i>Ví dụ: <code>/tien_nga_partner_report 01/2026 - 04/2026</code></i>",
            parse_mode=ParseMode.HTML
        )
        return

    # Parse date range mm/yyyy - mm/yyyy
    range_str = args[1].strip()
    m = re.match(r"(\d{1,2})/(\d{4})\s*-\s*(\d{1,2})/(\d{4})", range_str)
    if not m:
        await message.reply_text(
            "⚠️ Định dạng không hợp lệ. Vui lòng nhập <code>MM/YYYY - MM/YYYY</code>.",
            parse_mode=ParseMode.HTML
        )
        return

    start_month, start_year = int(m.group(1)), int(m.group(2))
    end_month, end_year = int(m.group(3)), int(m.group(4))

    if not (1 <= start_month <= 12 and 1 <= end_month <= 12):
        await message.reply_text("⚠️ Tháng phải từ 1-12.", parse_mode=ParseMode.HTML)
        return

    from datetime import date
    start_date = date(start_year, start_month, 1)
    # End date = last day of end_month
    if end_month == 12:
        end_date = date(end_year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(end_year, end_month + 1, 1) - timedelta(days=1)

    if start_date > end_date:
        await message.reply_text("⚠️ Ngày bắt đầu không thể lớn hơn ngày kết thúc.", parse_mode=ParseMode.HTML)
        return

    from app.models.business import Partners, PartnerBusinesses
    from app.db.session import SessionLocal
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile, os

    db = SessionLocal()
    try:
        # Lấy tất cả giao dịch trong khoảng thời gian
        txns = db.query(PartnerBusinesses).filter(
            PartnerBusinesses.day >= start_date,
            PartnerBusinesses.day <= end_date
        ).order_by(PartnerBusinesses.day).all()

        # Lấy tất cả partner (để tra tên)
        partners = db.query(Partners).filter(Partners.status == "ACTIVE").all()
        partner_map = {p.partner_id: p for p in partners}

        if not txns:
            await message.reply_text(
                f"⚠️ Không có giao dịch đối tác nào từ "
                f"<b>{start_month:02d}/{start_year}</b> đến <b>{end_month:02d}/{end_year}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        # ── Styles ──
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        title_font = Font(bold=True, size=14, color="2F5496")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        money_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        alt_fill = PatternFill("solid", fgColor="D9E2F3")
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill("solid", fgColor="FFF2CC")

        def fmt_money_xl(val):
            if val is None or val == 0:
                return "0"
            return f"{int(val):,}".replace(",", ".")

        def write_sheet(ws, title, data_rows, headers, col_widths):
            """Helper: ghi tiêu đề, header, data, tổng vào sheet."""
            # Title row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            period_text = f"Từ {start_month:02d}/{start_year} đến {end_month:02d}/{end_year}"
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            period_cell = ws.cell(row=2, column=1, value=period_text)
            period_cell.font = Font(italic=True, size=10)
            period_cell.alignment = Alignment(horizontal="center", vertical="center")

            header_row = 4
            # Headers
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Data
            total_import_qty = 0
            total_export_qty = 0
            total_import_amt = 0
            total_export_amt = 0
            total_amount_sum = 0

            for idx, row_data in enumerate(data_rows):
                row_num = header_row + 1 + idx
                row_fill = alt_fill if idx % 2 == 1 else None
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx in (1, 3, 4):  # STT, Ngày, Loại GD
                        cell.alignment = center_align
                    elif col_idx >= 5:  # Số liệu / tiền
                        cell.alignment = money_align
                    else:
                        cell.alignment = left_align
                    if row_fill:
                        cell.fill = row_fill

                # Accumulate totals from raw values
                total_import_qty += row_data[4] if isinstance(row_data[4], (int, float)) else 0
                total_export_qty += row_data[5] if isinstance(row_data[5], (int, float)) else 0
                total_import_amt += row_data[7] if len(row_data) > 7 and isinstance(row_data[7], (int, float)) else 0
                total_export_amt += row_data[8] if len(row_data) > 8 and isinstance(row_data[8], (int, float)) else 0
                total_amount_sum += row_data[9] if len(row_data) > 9 and isinstance(row_data[9], (int, float)) else 0

            # Total row
            total_row_num = header_row + 1 + len(data_rows)
            total_vals = ["", "", "", "TỔNG CỘNG"]

            if len(headers) == 11:  # Tab Tổng
                total_vals += [total_import_qty, total_export_qty, "", total_import_amt, total_export_amt, total_amount_sum, ""]
            elif len(headers) == 8:  # Tab Nhập / Xuất
                # Re-calc for sub-tabs
                sub_qty = sum(r[4] for r in data_rows if isinstance(r[4], (int, float)))
                sub_amt = sum(r[6] for r in data_rows if isinstance(r[6], (int, float)))
                total_vals += [sub_qty, "", sub_amt, ""]

            for col_idx, val in enumerate(total_vals, 1):
                if col_idx > len(headers):
                    break
                cell = ws.cell(row=total_row_num, column=col_idx, value=val)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
                if col_idx >= 5:
                    cell.alignment = money_align
                elif col_idx == 4:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.font = total_font

            # Column widths
            for col_idx, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

            ws.freeze_panes = f"A{header_row + 1}"

        # ── Build data ──
        all_rows = []      # For "Tổng" sheet
        import_rows = []   # For "Nhập" sheet
        export_rows = []   # For "Xuất" sheet

        for idx, t in enumerate(txns, 1):
            p = partner_map.get(t.partner_id)
            partner_name = p.partner_name if p else t.partner_id
            day_str = t.day.strftime("%d/%m/%Y") if t.day else "—"
            txn_type = "Nhập" if (t.import_amount or 0) > 0 else "Xuất"
            import_qty = t.import_amount or 0
            export_qty = t.export_amount or 0
            unit_price = t.unit_price or 0
            import_amt = import_qty * unit_price
            export_amt = export_qty * unit_price
            total_amt = t.total_amount or 0

            all_rows.append([
                idx,
                partner_name,
                t.partner_id,
                day_str,
                import_qty,
                export_qty,
                unit_price,
                import_amt,
                export_amt,
                total_amt,
                t.notes or ""
            ])

            if txn_type == "Nhập":
                import_rows.append([
                    len(import_rows) + 1,
                    partner_name,
                    t.partner_id,
                    day_str,
                    import_qty,
                    unit_price,
                    import_amt,
                    t.notes or ""
                ])
            else:
                export_rows.append([
                    len(export_rows) + 1,
                    partner_name,
                    t.partner_id,
                    day_str,
                    export_qty,
                    unit_price,
                    export_amt,
                    t.notes or ""
                ])

        # ── Create workbook ──
        wb = openpyxl.Workbook()

        # Tab 1: Tổng
        ws_total = wb.active
        ws_total.title = "Tổng"
        total_headers = [
            "STT", "Tên Đối Tác", "Mã ĐT", "Ngày",
            "SL Nhập (Kg)", "SL Xuất (Kg)", "Đơn Giá (VNĐ)",
            "Tiền Nhập (VNĐ)", "Tiền Xuất (VNĐ)", "Thành Tiền (VNĐ)", "Ghi Chú"
        ]
        total_widths = [6, 22, 12, 14, 14, 14, 16, 18, 18, 18, 20]
        write_sheet(ws_total, "BÁO CÁO TỔNG HỢP ĐỐI TÁC", all_rows, total_headers, total_widths)

        # Tab 2: Nhập
        ws_import = wb.create_sheet("Nhập")
        import_headers = [
            "STT", "Tên Đối Tác", "Mã ĐT", "Ngày",
            "Số Lượng (Kg)", "Đơn Giá (VNĐ)", "Thành Tiền (VNĐ)", "Ghi Chú"
        ]
        import_widths = [6, 22, 12, 14, 16, 16, 18, 20]
        write_sheet(ws_import, "BÁO CÁO NHẬP ĐỐI TÁC", import_rows, import_headers, import_widths)

        # Tab 3: Xuất
        ws_export = wb.create_sheet("Xuất")
        export_headers = [
            "STT", "Tên Đối Tác", "Mã ĐT", "Ngày",
            "Số Lượng (Kg)", "Đơn Giá (VNĐ)", "Thành Tiền (VNĐ)", "Ghi Chú"
        ]
        export_widths = [6, 22, 12, 14, 16, 16, 18, 20]
        write_sheet(ws_export, "BÁO CÁO XUẤT ĐỐI TÁC", export_rows, export_headers, export_widths)

        # ── Format number columns as VN-style in all sheets ──
        for ws in [ws_total, ws_import, ws_export]:
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
                for cell in row:
                    if isinstance(cell.value, float) and cell.value == int(cell.value):
                        cell.value = int(cell.value)

        # ── Save & send ──
        file_name = f"bao_cao_tong_hop_doi_tac_{start_month:02d}_{start_year}_{end_month:02d}_{end_year}.xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        total_import = sum(r[7] for r in all_rows if isinstance(r[7], (int, float)))
        total_export = sum(r[8] for r in all_rows if isinstance(r[8], (int, float)))

        await message.reply_document(
            document=tmp_path,
            file_name=file_name,
            caption=(
                f"<b>BÁO CÁO TỔNG HỢP ĐỐI TÁC</b>\n\n"
                f"<b>Kỳ:</b> {start_month:02d}/{start_year} - {end_month:02d}/{end_year}\n"
                f"<b>Tổng GD:</b> {len(txns)} giao dịch\n"
                f"<b>Nhập:</b> {len(import_rows)} GD | <b>Xuất:</b> {len(export_rows)} GD\n"
                f"<b>Tổng Tiền Nhập:</b> <code>{fmt_money_xl(total_import)} VNĐ</code>\n"
                f"<b>Tổng Tiền Xuất:</b> <code>{fmt_money_xl(total_export)} VNĐ</code>\n\n"
                f"<i>Xuất lúc: {datetime.now().strftime('%H:%M %d/%m/%Y')}</i>"
            ),
            parse_mode=ParseMode.HTML,
        )

        os.remove(tmp_path)
        LogInfo(
            f"[TienNga] Partner report exported {start_month:02d}/{start_year}-{end_month:02d}/{end_year} "
            f"by @{message.from_user.username or message.from_user.id}",
            LogType.SYSTEM_STATUS
        )

    except Exception as e:
        import traceback as tb
        LogError(f"Error in partner report: {tb.format_exc()}", LogType.SYSTEM_STATUS)
        await message.reply_text(f"❌ Có lỗi xảy ra khi xuất báo cáo: {e}", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# =========================================================================================
# Biểu đồ Mua Mủ (Chart Purchases)
# =========================================================================================

@bot.on_message(filters.command(["tien_nga_chart_purcharse", "tien_nga_bieu_do_thu_mua"]) | filters.regex(r"^@\w+\s+/tien_nga_chart_purcharse|/tien_nga_bieu_do_thu_mua\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_chart_purcharse_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_chart_purcharse", "tien_nga_bieu_do_thu_mua"])
    if args is None: return

    # Parse custom date range and optional household_id
    command_text = message.text.split(maxsplit=1)
    custom_date_range = ""
    hhd_id = None
    if len(command_text) > 1:
        import re
        m = re.match(r"(?:([A-Za-z0-9_-]+)\s+)?(\d{1,2})/(\d{4})\s*-\s*(\d{1,2})/(\d{4})", command_text[1].strip())
        if m:
            hhd_id = m.group(1)
            m1, y1, m2, y2 = m.group(2), m.group(3), m.group(4), m.group(5)
            custom_date_range = f"{int(m1):02d}{y1[2:]}_{int(m2):02d}{y2[2:]}"
        else:
            await message.reply_text("⚠️ Định dạng ngày không hợp lệ. Vui lòng thử:\n<code>/tien_nga_chart_purcharse 01/2026 - 05/2026</code>\nHoặc:\n<code>/tien_nga_chart_purcharse KH001 01/2026 - 05/2026</code>", parse_mode=ParseMode.HTML)
            return

    db = SessionLocal()
    try:
        if hhd_id:
            status_msg = await message.reply_text("⏳ Đang tổng hợp dữ liệu cá nhân...", parse_mode=ParseMode.HTML)
            from app.models.business import Customers, DailyPurchases
            from sqlalchemy import func
            import calendar
            from datetime import datetime
            
            customer = db.query(Customers).filter(Customers.hoursehold_id == hhd_id).first()
            customer_name = customer.fullname if customer else "Khách vãng lai"
            
            start_m, start_y = int(m1), int(y1)
            end_m, end_y = int(m2), int(y2)
            start_date = datetime(start_y, start_m, 1, 0, 0, 0)
            last_day = calendar.monthrange(end_y, end_m)[1]
            end_date = datetime(end_y, end_m, last_day, 23, 59, 59, 999999)
            
            query = db.query(
                DailyPurchases.day,
                func.sum(DailyPurchases.weight).label('t_weight'),
                func.sum(DailyPurchases.actual_weight).label('t_actual_weight'),
                func.sum(DailyPurchases.dry_rubber).label('t_dry_rubber')
            ).filter(
                DailyPurchases.is_checked == True,
                DailyPurchases.hoursehold_id == hhd_id,
                DailyPurchases.day >= start_date.date(),
                DailyPurchases.day <= end_date.date()
            )
            group_results = query.group_by(DailyPurchases.day).order_by(DailyPurchases.day.asc()).all()
            
            if not group_results:
                await status_msg.edit_text(f"❌ Không có dữ liệu thu mua nào của hộ <b>{hhd_id}</b> trong khoảng thời gian này.", parse_mode=ParseMode.HTML)
                return
                
            months_dict = {}
            for r in group_results:
                if r.day:
                    m_key = r.day.strftime("Tháng %m/%Y")
                    if m_key not in months_dict:
                        months_dict[m_key] = {"labels": [], "weight": [], "actual_weight": [], "dry_rubber": []}
                    months_dict[m_key]["labels"].append(r.day.strftime("%d/%m"))
                    months_dict[m_key]["weight"].append(float(r.t_weight or 0))
                    months_dict[m_key]["actual_weight"].append(float(r.t_actual_weight or 0))
                    months_dict[m_key]["dry_rubber"].append(float(r.t_dry_rubber or 0))
                    
            charts_list = []
            for m_key, m_data in months_dict.items():
                charts_list.append({
                    "month_label": m_key,
                    "labels": m_data["labels"],
                    "weight": m_data["weight"],
                    "actual_weight": m_data["actual_weight"],
                    "dry_rubber": m_data["dry_rubber"]
                })
                    
            from bot.utils.chart_generator import generate_chart_image
            chart_data = {
                "title": f"Biểu đồ thu mua: Hộ {customer_name}",
                "subtitle": f"Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}",
                "charts": charts_list
            }
            img_buf = await generate_chart_image(chart_data)
            await message.reply_document(
                document=img_buf,
                caption=f"📈 <b>BIỂU ĐỒ BÁN MỦ - {customer_name.upper()}</b>\nMã hộ: <b>{hhd_id}</b>\nThời gian: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
                parse_mode=ParseMode.HTML
            )
            await status_msg.delete()
            return
            
        from app.models.business import CollectionPoint
        cps = db.query(CollectionPoint).all()
        
        keyboard = []
        if custom_date_range:
            keyboard.append([InlineKeyboardButton("Tổng (Toàn hệ thống)", callback_data=f"cb_tnchart_c_all_{custom_date_range}")])
            row = []
            for cp in cps:
                if cp.collection_name:
                    row.append(InlineKeyboardButton(cp.collection_name, callback_data=f"cb_tnchart_c_{cp.id}_{custom_date_range}"))
                    if len(row) == 2:
                        keyboard.append(row)
                        row = []
            if len(row) > 0:
                keyboard.append(row)
            keyboard.append([InlineKeyboardButton("Hủy", callback_data="cb_tnchart_scope_cancel")])
        else:
            keyboard.append([InlineKeyboardButton("Tổng (Toàn hệ thống)", callback_data="cb_tnchart_scope_all")])
            row = []
            for cp in cps:
                if cp.collection_name:
                    row.append(InlineKeyboardButton(cp.collection_name, callback_data=f"cb_tnchart_scope_{cp.id}"))
                    if len(row) == 2:
                        keyboard.append(row)
                        row = []
            if len(row) > 0:
                keyboard.append(row)
            keyboard.append([InlineKeyboardButton("Hủy", callback_data="cb_tnchart_scope_cancel")])
            
        reply_markup = InlineKeyboardMarkup(keyboard)
        txt = "<b>BIỂU ĐỒ THU MUA MỦ</b>\n"
        txt += "<i>Gợi ý tra cứu nhanh:</i>\n<code>/tien_nga_chart_purcharse [mã hộ] mm/yyyy - mm/yyyy</code>\n"
        txt += "<i>Gợi ý tra cứu nhanh:</i>\n<code>/tien_nga_chart_purcharse mm/yyyy - mm/yyyy</code>\n"
        
        if custom_date_range:
            txt += f"<i>Đã chọn thời gian: {m1}/{y1} - {m2}/{y2}</i>\n"
        txt += "Vui lòng chọn phạm vi lên biểu đồ:"
        
        await message.reply_text(
            text=txt,
            reply_markup=reply_markup,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error in chart_purcharse_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^cb_tnchart_scope_(.*)$"))
async def tien_nga_chart_scope_callback(client, callback_query: CallbackQuery):
    scope_id = callback_query.matches[0].group(1)
    
    if scope_id == "cancel":
        await callback_query.message.delete()
        return

    # Menu thời gian
    keyboard = [
        [
            InlineKeyboardButton("1 Tháng", callback_data=f"cb_tnchart_time_{scope_id}_1m"),
            InlineKeyboardButton("3 Tháng", callback_data=f"cb_tnchart_time_{scope_id}_3m")
        ],
        [
            InlineKeyboardButton("6 Tháng", callback_data=f"cb_tnchart_time_{scope_id}_6m"),
            InlineKeyboardButton("1 Năm", callback_data=f"cb_tnchart_time_{scope_id}_1y")
        ],
        [
            InlineKeyboardButton("Năm hiện tại", callback_data=f"cb_tnchart_time_{scope_id}_ty"),
            InlineKeyboardButton("Năm trước", callback_data=f"cb_tnchart_time_{scope_id}_ly")
        ],
        [
            InlineKeyboardButton("Quay Lại", callback_data="cb_tnchart_back"),
            InlineKeyboardButton("Hủy", callback_data="cb_tnchart_scope_cancel")
        ]
    ]
    
    await callback_query.message.edit_text(
        "<b>Khung thời gian</b>\n\nBạn muốn vẽ biểu đồ trong khoảng thời gian nào?",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode=ParseMode.HTML
    )

@bot.on_callback_query(filters.regex(r"^cb_tnchart_back$"))
async def tien_nga_chart_back_callback(client, callback_query: CallbackQuery):
    db = SessionLocal()
    try:
        from app.models.business import CollectionPoint
        cps = db.query(CollectionPoint).all()
        keyboard = []
        keyboard.append([InlineKeyboardButton("Tổng (Toàn hệ thống)", callback_data="cb_tnchart_scope_all")])
        row = []
        for cp in cps:
            if cp.collection_name:
                row.append(InlineKeyboardButton(cp.collection_name, callback_data=f"cb_tnchart_scope_{cp.id}"))
                if len(row) == 2:
                    keyboard.append(row)
                    row = []
        if len(row) > 0:
            keyboard.append(row)
            
        keyboard.append([InlineKeyboardButton("Hủy", callback_data="cb_tnchart_scope_cancel")])
        
        await callback_query.message.edit_text(
            "<b>BIỂU ĐỒ THU MUA MỦ</b>\n\n"
            "Vui lòng chọn phạm vi lên biểu đồ:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        await callback_query.message.edit_text("❌ Lỗi hệ thống.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^cb_tnchart_time_(.*)_(.*)$"))
async def tien_nga_chart_time_callback(client, callback_query: CallbackQuery):
    scope_id = callback_query.matches[0].group(1)
    time_code = callback_query.matches[0].group(2)
    
    await callback_query.message.edit_text("⏳ Đang tổng hợp dữ liệu...\n<i>Quá trình này có thể mất vài giây.</i>", parse_mode=ParseMode.HTML)
    
    from datetime import datetime, timedelta
    from sqlalchemy import func
    from app.models.business import DailyPurchases, CollectionPoint
    
    today = datetime.now()
    if time_code == "1m":
        start_date = today - timedelta(days=30)
    elif time_code == "3m":
        start_date = today - timedelta(days=90)
    elif time_code == "6m":
        start_date = today - timedelta(days=180)
    elif time_code == "1y":
        start_date = today - timedelta(days=365)
    elif time_code == "ty":
        start_date = datetime(today.year, 1, 1)
    elif time_code == "ly":
        start_date = datetime(today.year - 1, 1, 1)
        today = datetime(today.year - 1, 12, 31)
    else:
        start_date = today - timedelta(days=30)
        
    start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    db = SessionLocal()
    try:
        query = db.query(
            DailyPurchases.day,
            func.sum(DailyPurchases.weight).label('t_weight'),
            func.sum(DailyPurchases.actual_weight).label('t_actual_weight'),
            func.sum(DailyPurchases.dry_rubber).label('t_dry_rubber')
        ).filter(
            DailyPurchases.is_checked == True,
            DailyPurchases.day >= start_date.date(),
            DailyPurchases.day <= end_date.date()
        )
        
        scope_label = "Tổng Toàn Hệ Thống"
        if scope_id != "all":
            cp = db.query(CollectionPoint).filter(CollectionPoint.id == scope_id).first()
            if cp:
                scope_label = cp.collection_name
            query = query.filter(DailyPurchases.collection_point_id == scope_id)
            
        group_results = query.group_by(DailyPurchases.day).order_by(DailyPurchases.day.asc()).all()
        
        if not group_results:
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Quay Lại", callback_data=f"cb_tnchart_scope_{scope_id}")]])
            await callback_query.message.edit_text(
                "❌ Không có dữ liệu thu mua nào trong khoảng thời gian đã chọn.", 
                reply_markup=kb
            )
            return
            
        months_dict = {}
        for r in group_results:
            if r.day:
                m_key = r.day.strftime("Tháng %m/%Y")
                if m_key not in months_dict:
                    months_dict[m_key] = {"labels": [], "weight": [], "actual_weight": [], "dry_rubber": []}
                
                months_dict[m_key]["labels"].append(r.day.strftime("%d/%m"))
                months_dict[m_key]["weight"].append(float(r.t_weight or 0))
                months_dict[m_key]["actual_weight"].append(float(r.t_actual_weight or 0))
                months_dict[m_key]["dry_rubber"].append(float(r.t_dry_rubber or 0))
                
        charts_list = []
        for m_key, m_data in months_dict.items():
            charts_list.append({
                "month_label": m_key,
                "labels": m_data["labels"],
                "weight": m_data["weight"],
                "actual_weight": m_data["actual_weight"],
                "dry_rubber": m_data["dry_rubber"]
            })
                
        # Gọi module vẽ đồ thị
        from bot.utils.chart_generator import generate_chart_image
        chart_data = {
            "title": f"Biểu đồ thu mua: {scope_label}",
            "subtitle": f"Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}",
            "charts": charts_list
        }
        
        img_buf = await generate_chart_image(chart_data)
        
        await callback_query.message.reply_document(
            document=img_buf,
            caption=f"<b>BIỂU ĐỒ THU MUA MỦ</b>\nXưởng: <b>{scope_label}</b>\nThời gian: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            parse_mode=ParseMode.HTML
        )
        await callback_query.message.delete()
        
    except Exception as e:
        LogError(f"Error drawing chart: {e}", LogType.SYSTEM_STATUS)
        try:
            await callback_query.message.edit_text("❌ Có lỗi xảy ra trong quá trình tổng hợp dữ liệu.")
        except Exception:
            pass
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^cb_tnchart_c_(.*)_(\d{4})_(\d{4})$"))
async def tien_nga_chart_custom_callback(client, callback_query: CallbackQuery):
    scope_id = callback_query.matches[0].group(1)
    start_mmyy = callback_query.matches[0].group(2)
    end_mmyy = callback_query.matches[0].group(3)
    
    await callback_query.message.edit_text("⏳ Đang tổng hợp dữ liệu (Tùy chọn)...\n<i>Quá trình này có thể mất vài giây.</i>", parse_mode=ParseMode.HTML)
    
    import calendar
    from datetime import datetime
    from sqlalchemy import func
    from app.models.business import DailyPurchases, CollectionPoint
    
    # Parse mmyy
    start_m = int(start_mmyy[:2])
    start_y = int("20" + start_mmyy[2:])
    end_m = int(end_mmyy[:2])
    end_y = int("20" + end_mmyy[2:])
    
    start_date = datetime(start_y, start_m, 1, 0, 0, 0)
    last_day = calendar.monthrange(end_y, end_m)[1]
    end_date = datetime(end_y, end_m, last_day, 23, 59, 59, 999999)
    
    db = SessionLocal()
    try:
        query = db.query(
            DailyPurchases.day,
            func.sum(DailyPurchases.weight).label('t_weight'),
            func.sum(DailyPurchases.actual_weight).label('t_actual_weight'),
            func.sum(DailyPurchases.dry_rubber).label('t_dry_rubber')
        ).filter(
            DailyPurchases.is_checked == True,
            DailyPurchases.day >= start_date.date(),
            DailyPurchases.day <= end_date.date()
        )
        
        scope_label = "Tổng Toàn Hệ Thống"
        if scope_id != "all":
            cp = db.query(CollectionPoint).filter(CollectionPoint.id == scope_id).first()
            if cp:
                scope_label = cp.collection_name
            query = query.filter(DailyPurchases.collection_point_id == scope_id)
            
        group_results = query.group_by(DailyPurchases.day).order_by(DailyPurchases.day.asc()).all()
        
        if not group_results:
            await callback_query.message.edit_text(
                f"❌ Không có dữ liệu thu mua nào trong khoảng thời gian {start_m:02d}/{start_y} đến {end_m:02d}/{end_y}."
            )
            return
            
        months_dict = {}
        for r in group_results:
            if r.day:
                m_key = r.day.strftime("Tháng %m/%Y")
                if m_key not in months_dict:
                    months_dict[m_key] = {"labels": [], "weight": [], "actual_weight": [], "dry_rubber": []}
                
                months_dict[m_key]["labels"].append(r.day.strftime("%d/%m"))
                months_dict[m_key]["weight"].append(float(r.t_weight or 0))
                months_dict[m_key]["actual_weight"].append(float(r.t_actual_weight or 0))
                months_dict[m_key]["dry_rubber"].append(float(r.t_dry_rubber or 0))
                
        charts_list = []
        for m_key, m_data in months_dict.items():
            charts_list.append({
                "month_label": m_key,
                "labels": m_data["labels"],
                "weight": m_data["weight"],
                "actual_weight": m_data["actual_weight"],
                "dry_rubber": m_data["dry_rubber"]
            })
                
        # Gọi module vẽ đồ thị
        from bot.utils.chart_generator import generate_chart_image
        chart_data = {
            "title": f"Biểu đồ thu mua: {scope_label}",
            "subtitle": f"Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}",
            "charts": charts_list
        }
        
        img_buf = await generate_chart_image(chart_data)
        
        await callback_query.message.reply_document(
            document=img_buf,
            caption=f"<b>BIỂU ĐỒ THU MUA MỦ (TÙY CHỌN)</b>\nXưởng: <b>{scope_label}</b>\nThời gian: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            parse_mode=ParseMode.HTML
        )
        await callback_query.message.delete()
        
    except Exception as e:
        LogError(f"Error drawing custom chart: {e}", LogType.SYSTEM_STATUS)
        try:
            await callback_query.message.edit_text("❌ Có lỗi xảy ra trong quá trình xuất dữ liệu tùy chỉnh.")
        except Exception:
            pass
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_payment_of_debt", "tien_nga_thanh_toan_cong_no"]) | filters.regex(r"^@\w+\s+/(tien_nga_payment_of_debt|tien_nga_thanh_toan_cong_no)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
async def tien_nga_payment_of_debt_handler(client, message: Message) -> None:
    args = message.text.split()
    if len(args) < 3:
        await message.reply_text(
            "⚠️ Cú pháp: <code>/tien_nga_thanh_toan_cong_no [Mã] [Số Tiền]</code>\n\n"
            "<i>(Ví dụ: <code>/tien_nga_thanh_toan_cong_no DT001 500000</code>)</i>\n"
            "<i>Hỗ trợ: Đối tác (DT), Khách hàng (KH), Nhân sự (NV), Hộ Dân (HD).</i>",
            parse_mode=ParseMode.HTML
        )
        return

    target_id = args[1].upper()

    try:
        amount_str = args[2].replace(".", "").replace(",", "")
        amount = int(amount_str)
    except ValueError:
        await message.reply_text("⚠️ <b>Số Tiền</b> phải là một số hợp lệ.", parse_mode=ParseMode.HTML)
        return

    if amount <= 0:
        await message.reply_text("⚠️ <b>Số Tiền</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return

    from app.models.business import Customers, Partners, Households, Projects
    from app.models.employee import Employee
    from app.models.telegram import TelegramProjectMember

    db = SessionLocal()
    try:
        # Try to find in Partners, Customers, then Employee
        partner = None
        customer = None
        employee = None
        household = None
        target_type = None  # "partner", "customer", "employee" or "household"

        if target_id.startswith("DT"):
            partner = db.query(Partners).filter(Partners.partner_id == target_id).first()
            if partner:
                target_type = "partner"
                old_debt = partner.total_debt or 0
                name = partner.partner_name or partner.partner_id

        if not target_type and target_id.startswith("HD"):
            household = db.query(Households).filter(Households.household_code == target_id, Households.status == "ACTIVE").first()
            if household:
                target_type = "household"
                old_debt = household.total_debt or 0
                name = household.fullname or household.household_code

        if not target_type:
            customer = db.query(Customers).filter(Customers.hoursehold_id == target_id).first()
            if customer:
                target_type = "customer"
                old_debt = customer.total_debt or 0
                name = customer.fullname or customer.hoursehold_id

        if not target_type:
            employee = db.query(Employee).filter(Employee.id == target_id).first()
            if employee:
                target_type = "employee"
                old_debt = employee.total_debt or 0
                name = f"{employee.last_name or ''} {employee.first_name or ''}".strip() or employee.id

        if not target_type:
            await message.reply_text(
                f"⚠️ Không tìm thấy Đối tác, Khách hàng, Nhân sự hoặc Hộ Dân với mã <b>{target_id}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        # Validate: debt must be positive
        if old_debt <= 0:
            await message.reply_text(
                f"⚠️ <b>Không thể thanh toán!</b>\n\n"
                f"<b>Mã:</b> <code>{target_id}</code>\n"
                f"<b>Tên:</b> {name}\n"
                f"<b>Công nợ hiện tại:</b> <code>{fmt_vn(old_debt)}</code>\n\n"
                f"<i>Đối tượng này không có công nợ hoặc đang được nợ ngược.</i>",
                parse_mode=ParseMode.HTML
            )
            return

        # Validate: amount must not exceed total_debt
        if amount > old_debt:
            await message.reply_text(
                f"⚠️ <b>Số tiền thanh toán vượt quá công nợ!</b>\n\n"
                f"<b>Mã:</b> <code>{target_id}</code>\n"
                f"<b>Tên:</b> {name}\n"
                f"<b>Công nợ hiện tại:</b> <code>{fmt_vn(old_debt)}</code>\n"
                f"<b>Số tiền yêu cầu:</b> <code>{fmt_vn(amount)}</code>\n\n"
                f"<i>Số tiền thanh toán không được lớn hơn công nợ hiện tại.</i>",
                parse_mode=ParseMode.HTML
            )
            return

        # Update debt
        new_debt = old_debt - amount
        if target_type == "partner":
            partner.total_debt = new_debt
        elif target_type == "customer":
            customer.total_debt = new_debt
        elif target_type == "household":
            household.total_debt = new_debt
        else:
            employee.total_debt = new_debt

        db.commit()

        type_label = {"partner": "Đối tác", "customer": "Khách hàng", "employee": "Nhân sự", "household": "Hộ Dân"}.get(target_type, "")
        success_msg = (
            f"✅ <b>THANH TOÁN CÔNG NỢ THÀNH CÔNG</b>\n\n"
            f"<b>Loại:</b> {type_label}\n"
            f"<b>Mã:</b> <code>{target_id}</code>\n"
            f"<b>Tên:</b> {name}\n"
            f"<b>Số Tiền Thanh Toán:</b> <code>{fmt_vn(amount)}</code>\n\n"
            f"<b>Công Nợ Cũ:</b> <code>{fmt_vn(old_debt)}</code>\n"
            f"<b>Công Nợ Mới:</b> <code>{fmt_vn(new_debt)}</code>"
        )
        if new_debt > 0:
            success_msg += f"\n\n<i>Còn nợ <code>{fmt_vn(new_debt)}</code>. Vui lòng thanh toán phần còn lại.</i>"
        elif new_debt == 0:
            success_msg += f"\n\n<i>Đã thanh toán hết công nợ!</i>"
        await message.reply_text(success_msg, parse_mode=ParseMode.HTML)

        LogInfo(f"[TienNga] /tien_nga_payment_of_debt {target_type} {target_id} amount={amount} by {message.from_user.id}", LogType.SYSTEM_STATUS)

        # Notify member group
        if target_type == "partner" and partner.telegram_group:
            project = db.query(Projects).filter(Projects.project_name == "Tiến Nga").first()
            if project:
                member_chat = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.project_id == project.id,
                    TelegramProjectMember.role == "member",
                    TelegramProjectMember.group_name == partner.telegram_group
                ).first()

                if member_chat and member_chat.chat_id:
                    try:
                        await client.send_message(
                            chat_id=int(member_chat.chat_id),
                            text=(
                                f"🔔 <b>THÔNG BÁO CẬP NHẬT CÔNG NỢ</b>\n\n"
                                f"<b>Kính gửi:</b> {name}\n"
                                f"<b>Số Tiền Thanh Toán:</b> <code>{fmt_vn(amount)}</code>\n"
                                f"<b>Công Nợ Hiện Tại:</b> <code>{fmt_vn(new_debt)}</code>\n\n"
                                f"<i>Cảm ơn quý đối tác!</i>"
                            ),
                            parse_mode=ParseMode.HTML
                        )
                    except Exception as e:
                        LogError(f"Error sending debt update to partner {target_id}: {e}", LogType.SYSTEM_STATUS)

        elif target_type == "customer" and customer.username:
            username_clean = customer.username.lstrip('@')
            project = db.query(Projects).filter(Projects.project_name == "Tiến Nga").first()
            if project:
                member_chat = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.project_id == project.id,
                    TelegramProjectMember.role == "member",
                    TelegramProjectMember.user_name == username_clean
                ).first()

                if member_chat and member_chat.chat_id:
                    try:
                        await client.send_message(
                            chat_id=int(member_chat.chat_id),
                            text=(
                                f"🔔 <b>THÔNG BÁO CẬP NHẬT CÔNG NỢ</b>\n\n"
                                f"<b>Kính gửi:</b> {name}\n"
                                f"<b>Số Tiền Thanh Toán:</b> <code>{fmt_vn(amount)}</code>\n"
                                f"<b>Công Nợ Hiện Tại:</b> <code>{fmt_vn(new_debt)}</code>\n\n"
                                f"<i>Cảm ơn quý khách!</i>"
                            ),
                            parse_mode=ParseMode.HTML
                        )
                    except Exception as e:
                        LogError(f"Error sending debt update to {target_id}: {e}", LogType.SYSTEM_STATUS)

        elif target_type == "household" and household.telegram_group:
            project = db.query(Projects).filter(Projects.project_name == "Tiến Nga").first()
            if project:
                member_chat = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.project_id == project.id,
                    TelegramProjectMember.role == "member",
                    TelegramProjectMember.group_name == household.telegram_group
                ).first()

                if member_chat and member_chat.chat_id:
                    try:
                        await client.send_message(
                            chat_id=int(member_chat.chat_id),
                            text=(
                                f"🔔 <b>THÔNG BÁO CẬP NHẬT CÔNG NỢ</b>\n\n"
                                f"<b>Kính gửi:</b> {name}\n"
                                f"<b>Số Tiền Thanh Toán:</b> <code>{fmt_vn(amount)}</code>\n"
                                f"<b>Công Nợ Hiện Tại:</b> <code>{fmt_vn(new_debt)}</code>\n\n"
                                f"<i>Cảm ơn hộ dân!</i>"
                            ),
                            parse_mode=ParseMode.HTML
                        )
                    except Exception as e:
                        LogError(f"Error sending debt update to household {target_id}: {e}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in tien_nga_payment_of_debt: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Hệ thống gặp lỗi khi cập nhật công nợ.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

# =========================================================================================
# TẠO KHOẢN ĐẦU TƯ
# =========================================================================================

@bot.on_message(filters.command(["tien_nga_create_investment", "tien_nga_tao_dau_tu"]) | filters.regex(r"^@\w+\s+/(tien_nga_create_investment|tien_nga_tao_dau_tu)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_FINANCE, CustomTitle.MAIN_SHAREHOLDER)
async def tien_nga_create_investment_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")

    if len(lines) < 2:
        today_str = datetime.now().strftime("%d/%m/%Y")
        form_template = f"""<b>FORM TẠO KHOẢN ĐẦU TƯ</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/tien_nga_create_investment
Mã Đầu Tư: 
Tên Đầu Tư: 
Vốn Ban Đầu: 0
Ngày Bắt Đầu: {today_str}
Ngày Kết Thúc: 
Ghi Chú: </pre>

<i> Lưu ý: Vốn ban đầu phải là 0 VNĐ, sau đó góp vốn vào quỹ sau</i>
<i>(*Các trường bắt buộc: <b>Mã Đầu Tư</b>, <b>Tên Đầu Tư</b>)</i>"""
        await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        return

    # Parse form data
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    inv_code = data.get("Mã Đầu Tư", "").strip()
    name = data.get("Tên Đầu Tư", "").strip()
    capital_str = data.get("Vốn Ban Đầu", "0").strip()
    start_str = data.get("Ngày Bắt Đầu", "").strip()
    end_str = data.get("Ngày Kết Thúc", "").strip()
    notes = data.get("Ghi Chú", "").strip()

    if not inv_code:
        await message.reply_text("⚠️ <b>Mã Đầu Tư</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    if not name:
        await message.reply_text("⚠️ <b>Tên Đầu Tư</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    try:
        capital = float(capital_str.replace(".", "").replace(",", "")) if capital_str else 0.0
    except ValueError:
        await message.reply_text("⚠️ <b>Vốn Ban Đầu</b> phải là một số hợp lệ.", parse_mode=ParseMode.HTML)
        return

    if capital < 0:
        await message.reply_text("⚠️ <b>Vốn Ban Đầu</b> không được là số âm.", parse_mode=ParseMode.HTML)
        return

    # Parse dates
    start_date = None
    if start_str:
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                start_date = datetime.strptime(start_str, fmt).date()
                break
            except ValueError:
                continue

    end_date = None
    if end_str:
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                end_date = datetime.strptime(end_str, fmt).date()
                break
            except ValueError:
                continue

    from app.models.business import Investment
    import uuid as _uuid
    from sqlalchemy.exc import IntegrityError

    db = SessionLocal()
    try:
        # Check for duplicate investment_code
        existing = db.query(Investment).filter(Investment.investment_code == inv_code).first()
        if existing:
            await message.reply_text(
                f"⚠️ Mã Đầu Tư <b>{inv_code}</b> đã tồn tại.\n"
                f"Tên: <b>{existing.name}</b>\n"
                f"Trạng thái: <b>{existing.status}</b>\n\n"
                f"<i>Vui lòng chọn mã khác.</i>",
                parse_mode=ParseMode.HTML
            )
            return

        new_inv = Investment(
            id=_uuid.uuid4(),
            investment_code=inv_code,
            name=name,
            initial_capital=capital,
            start_date=start_date,
            end_date=end_date,
            total_income=capital,
            total_expense=0.0,
            profit=capital,
            notes=notes or None,
            status="ACTIVE",
        )
        db.add(new_inv)
        db.commit()

        LogInfo(f"[TienNga] Created investment '{name}' with code '{inv_code}' by user {message.from_user.id}", LogType.SYSTEM_STATUS)

        success_msg = (
            f"✅ <b>TẠO KHOẢN ĐẦU TƯ THÀNH CÔNG</b>\n\n"
            f"<b>Mã Đầu Tư:</b> <code>{inv_code}</code>\n"
            f"<b>Tên:</b> {name}\n"
            f"<b>Vốn Ban Đầu:</b> <code>{fmt_vn(capital)}</code>\n"
        )
        if start_date:
            success_msg += f"<b>Ngày Bắt Đầu:</b> {start_date.strftime('%d/%m/%Y')}\n"
        if end_date:
            success_msg += f"<b>Ngày Kết Thúc:</b> {end_date.strftime('%d/%m/%Y')}\n"
        if notes:
            success_msg += f"<b>Ghi Chú:</b> {notes}\n"
        success_msg += f"\n<b>Mã:</b> <code>{new_inv.id}</code>"

        await message.reply_text(success_msg, parse_mode=ParseMode.HTML)
    except Exception as e:
        db.rollback()
        LogError(f"Error creating investment: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra khi tạo khoản đầu tư.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

# =========================================================================================
# YÊU CẦU THU / CHI HÀNG NGÀY
# =========================================================================================

@bot.on_message(filters.command(["tien_nga_request_daily_payments", "tien_nga_yeu_cau_thu_chi"]) | filters.regex(r"^@\w+\s+/(tien_nga_request_daily_payments|tien_nga_yeu_cau_thu_chi)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
async def tien_nga_request_daily_payments_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")

    # If multi-line → user is submitting form data
    if len(lines) >= 2:
        await _process_daily_payment_form(client, message, lines)
        return

    # Otherwise → show list of active investments
    from app.models.business import Investment
    db = SessionLocal()
    try:
        investments = db.query(Investment).filter(Investment.status == "ACTIVE").all()
        if not investments:
            await message.reply_text("⚠️ Chưa có khoản đầu tư nào. Vui lòng tạo trước bằng lệnh /tien_nga_create_investment.", parse_mode=ParseMode.HTML)
            return

        buttons = []
        for inv in investments:
            label = f"[{inv.investment_code}] {inv.name}" if inv.investment_code else inv.name
            if not label:
                label = str(inv.id)[:8]
            buttons.append([InlineKeyboardButton(label, callback_data=f"rdp_sel_{inv.id}")])
        buttons.append([InlineKeyboardButton("Hủy", callback_data="rdp_cancel")])

        await message.reply_text(
            "<b>YÊU CẦU THU / CHI HÀNG NGÀY</b>\n\n"
            "Vui lòng chọn khoản đầu tư:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error in tien_nga_request_daily_payments_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# --- Cancel ---
@bot.on_callback_query(filters.regex(r"^rdp_cancel$"))
async def rdp_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("❌ Đã hủy yêu cầu thu/chi.")


# --- Select investment → show Thu/Chi ---
@bot.on_callback_query(filters.regex(r"^rdp_sel_(.+)$"))
async def rdp_sel_callback(client, callback_query: CallbackQuery):
    inv_id = callback_query.matches[0].group(1)

    from app.models.business import Investment
    db = SessionLocal()
    try:
        inv = db.query(Investment).filter(Investment.id == inv_id).first()
        inv_name = inv.name if inv else "N/A"

        buttons = [
            [
                InlineKeyboardButton("Thu", callback_data=f"rdp_inv_thu_{inv_id}"),
                InlineKeyboardButton("Chi", callback_data=f"rdp_inv_chi_{inv_id}"),
            ],
            [
                InlineKeyboardButton("Quay lại", callback_data="rdp_back"),
                InlineKeyboardButton("Hủy", callback_data="rdp_cancel"),
            ],
        ]
        await callback_query.message.edit_text(
            f"<b>KHOẢN ĐẦU TƯ: {inv_name}</b>\n\n"
            "Vui lòng chọn loại giao dịch:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        LogError(f"Error in rdp_sel_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống.", show_alert=True)
    finally:
        db.close()


# --- Back to investment list ---
@bot.on_callback_query(filters.regex(r"^rdp_back$"))
async def rdp_back_callback(client, callback_query: CallbackQuery):
    from app.models.business import Investment
    db = SessionLocal()
    try:
        investments = db.query(Investment).filter(Investment.status == "ACTIVE").all()
        buttons = []
        for inv in investments:
            label = f"[{inv.investment_code}] {inv.name}" if inv.investment_code else inv.name
            if not label:
                label = str(inv.id)[:8]
            buttons.append([InlineKeyboardButton(label, callback_data=f"rdp_sel_{inv.id}")])
        buttons.append([InlineKeyboardButton("Hủy", callback_data="rdp_cancel")])

        await callback_query.message.edit_text(
            "<b>YÊU CẦU THU / CHI HÀNG NGÀY</b>\n\n"
            "Vui lòng chọn khoản đầu tư:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        LogError(f"Error in rdp_back_callback: {e}", LogType.SYSTEM_STATUS)
    finally:
        db.close()


# --- Select Thu/Chi → show form ---
@bot.on_callback_query(filters.regex(r"^rdp_inv_(thu|chi)_(.+)$"))
async def rdp_inv_callback(client, callback_query: CallbackQuery):
    payment_type = callback_query.matches[0].group(1)
    inv_id = callback_query.matches[0].group(2)
    type_label = "THU" if payment_type == "thu" else "CHI"

    from app.models.business import Investment
    db = SessionLocal()
    try:
        inv = db.query(Investment).filter(Investment.id == inv_id).first()
        inv_name = inv.name if inv else "N/A"

        today_str = datetime.now().strftime("%d/%m/%Y")
        form_code = inv.investment_code if inv and inv.investment_code else inv_id

        form_template = f"""<b>FORM {type_label} HÀNG NGÀY</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/tien_nga_request_daily_payments
Mã Đầu Tư: {form_code}
Loại: {payment_type}
Ngày: {today_str}
Số Tiền: 
Người Yêu Cầu: 
Người Thực Hiện: 
Người Nhận: 
Mục Đích: 
Lý Do: 
Ghi Chú: </pre>

<i>(*Đang tạo phiếu {type_label.lower()} cho: <b>{inv_name}</b>)</i>"""

        await callback_query.message.reply_text(form_template, parse_mode=ParseMode.HTML)
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rdp_inv_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống.", show_alert=True)
    finally:
        db.close()


# --- Process submitted form ---
async def _process_daily_payment_form(client, message: Message, lines: list):
    # Parse key:value pairs
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    inv_id = data.get("Mã Đầu Tư", "").strip()
    payment_type = data.get("Loại", "").strip().lower()
    day_str = data.get("Ngày", "").strip()
    amount_str = data.get("Số Tiền (VNĐ)", data.get("Số Tiền", "")).strip()
    requester = data.get("Người Yêu Cầu", "").strip()
    executor = data.get("Người Thực Hiện", "").strip()
    receiver = data.get("Người Nhận", "").strip()
    purpose = data.get("Mục Đích", "").strip()
    reason = data.get("Lý Do", "").strip()
    notes = data.get("Ghi Chú", "").strip()

    # Validate required fields
    if not inv_id or not payment_type or not amount_str:
        await message.reply_text(
            "⚠️ <b>Mã Đầu Tư</b>, <b>Loại</b> và <b>Số Tiền</b> là bắt buộc.",
            parse_mode=ParseMode.HTML
        )
        return

    if payment_type not in ("thu", "chi"):
        await message.reply_text("⚠️ <b>Loại</b> phải là <code>thu</code> hoặc <code>chi</code>.", parse_mode=ParseMode.HTML)
        return

    try:
        amount = float(amount_str.replace(".", "").replace(",", ""))
    except ValueError:
        await message.reply_text("⚠️ <b>Số Tiền</b> phải là một số hợp lệ.", parse_mode=ParseMode.HTML)
        return

    if amount <= 0:
        await message.reply_text("⚠️ <b>Số Tiền</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return

    # Parse date
    day = None
    if day_str:
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                day = datetime.strptime(day_str, fmt).date()
                break
            except ValueError:
                continue
    if not day:
        day = datetime.now().date()

    from app.models.business import Investment, DailyPayment
    import uuid as _uuid

    db = SessionLocal()
    try:
        import uuid as _uuid
        try:
            parsed_id = _uuid.UUID(inv_id)
            inv = db.query(Investment).filter(Investment.id == parsed_id).first()
        except ValueError:
            # Not a UUID, search by investment_code
            inv = db.query(Investment).filter(Investment.investment_code == inv_id).first()

        if not inv:
            await message.reply_text(f"⚠️ Không tìm thấy khoản đầu tư mã <b>{inv_id}</b>.", parse_mode=ParseMode.HTML)
            return

        # Validate expense against current balance
        if payment_type == "chi":
            current_balance = (inv.total_income or 0) - (inv.total_expense or 0)
            if amount > current_balance:
                await message.reply_text(
                    f"⚠️ <b>SỐ TIỀN CHI VƯỢT QUÁ SỐ DƯ!</b>\n\n"
                    f"<b>Khoản Đầu Tư:</b> {inv.name}\n"
                    f"<b>Số dư hiện tại:</b> <code>{fmt_vn(current_balance)}</code>\n"
                    f"<b>Số tiền yêu cầu chi:</b> <code>{fmt_vn(amount)}</code>\n\n"
                    f"<i>Vui lòng nhập số tiền nhỏ hơn hoặc bằng số dư hiện tại.</i>",
                    parse_mode=ParseMode.HTML
                )
                return

        from app.core.config import settings
        max_amount = settings.TienNga.Max_Daily_Payment_Auto_Approve
        requires_approval = False

        if payment_type == "chi" and amount > max_amount:
            requires_approval = True

        # Create DailyPayment record
        new_payment = DailyPayment(
            id=_uuid.uuid4(),
            investment_id=inv.id,
            requester=requester or None,
            executor=executor or None,
            receiver=receiver or None,
            payment_type=payment_type,
            purpose=purpose or None,
            reason=reason or None,
            amount=amount,
            day=day,
            notes=notes or None,
            status="PENDING" if requires_approval else "APPROVED",
        )
        db.add(new_payment)

        # Update Investment totals only if approved
        if not requires_approval:
            if payment_type == "thu":
                inv.total_income = (inv.total_income or 0) + amount
            else:
                inv.total_expense = (inv.total_expense or 0) + amount

            inv.profit = (inv.total_income or 0) - (inv.total_expense or 0)

        db.commit()

        if requires_approval:
            # Attempt to find the dynamic Owner's username
            from app.models.business import Projects
            from app.models.telegram import TelegramProjectMember
            
            owner_str = "@username (Owner)"
            project = db.query(Projects).filter(Projects.project_name == "Tiến Nga").first()
            if project:
                owner_member = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.project_id == project.id,
                    TelegramProjectMember.member_status == "OWNER"
                ).first()
                if owner_member:
                    if owner_member.user_name:
                        owner_str = f"@{owner_member.user_name} (Owner)"
                    elif owner_member.full_name:
                        owner_str = f"{owner_member.full_name} (Owner)"

            approval_msg = (
                f"⚠️ <b>YÊU CẦU DUYỆT PHIẾU CHI SỐ TIỀN LỚN</b> ⚠️\n\n"
                f"<b>Khoản Đầu Tư:</b> {inv.name}\n"
                f"<b>Người Yêu Cầu:</b> {requester or 'N/A'}\n"
                f"<b>Mục Đích:</b> {purpose or 'N/A'}\n"
                f"<b>Số Tiền:</b> <code>{fmt_vn(amount)}</code>\n"
                f"<b>Mã Phiếu:</b> <code>{new_payment.id}</code>\n\n"
                f"📌 <i>Ghi chú: {owner_str} vui lòng <b>Reply</b> tin nhắn này với lệnh <code>/confirm_payment</code> để thông qua hoặc <code>/deny_payment</code> để huỷ báo cáo hạch toán.</i>"
            )
            await message.reply_text(approval_msg, parse_mode=ParseMode.HTML)
            LogInfo(f"[TienNga] daily_payment pending approval for inv {inv.name} amount {amount} by {message.from_user.id}", LogType.SYSTEM_STATUS)
            return

        type_label = "THU" if payment_type == "thu" else "CHI"
        success_msg = (
            f"✅ <b>ĐÃ GHI NHẬN PHIẾU {type_label}</b>\n\n"
            f"<b>Khoản Đầu Tư:</b> {inv.name}\n"
            f"<b>Loại:</b> {type_label}\n"
            f"<b>Ngày:</b> {day.strftime('%d/%m/%Y')}\n"
            f"<b>Số Tiền:</b> <code>{fmt_vn(amount)}</code>\n"
        )
        if requester:
            success_msg += f"<b>Người Yêu Cầu:</b> {requester}\n"
        if executor:
            success_msg += f"<b>Người Thực Hiện:</b> {executor}\n"
        if receiver:
            success_msg += f"<b>Người Nhận:</b> {receiver}\n"
        if purpose:
            success_msg += f"<b>Mục Đích:</b> {purpose}\n"
        if reason:
            success_msg += f"<b>Lý Do:</b> {reason}\n"
        if notes:
            success_msg += f"<b>Ghi Chú:</b> {notes}\n"


        await message.reply_text(success_msg, parse_mode=ParseMode.HTML)
        LogInfo(f"[TienNga] daily_payment {payment_type} {amount} for inv {inv.name} by {message.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in _process_daily_payment_form: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Hệ thống gặp lỗi khi lưu phiếu thu/chi.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

# =========================================================================================
# DUYỆT PHIẾU CHI LỚN BẰNG REPLY
# =========================================================================================
import re

@bot.on_message(filters.command(["confirm_payment"]))
@require_user_type(UserType.OWNER)
@require_project_name("Tiến Nga")
@require_group_role("main")
async def tien_nga_confirm_payment_handler(client, message: Message) -> None:
    if not message.reply_to_message:
        await message.reply_text("⚠️ Vui lòng <b>Reply</b> một tin nhắn yêu cầu duyệt phiếu chi.", parse_mode=ParseMode.HTML)
        return
        
    replied_text = message.reply_to_message.text or ""
    # Extract Payment ID
    match = re.search(r"Mã Phiếu:\s*([a-f0-9\-]{36})", replied_text)
    if not match:
        await message.reply_text("⚠️ Không tìm thấy Mã Phiếu trong tin nhắn đang reply.", parse_mode=ParseMode.HTML)
        return
        
    payment_id = match.group(1)
    
    from app.models.business import Investment, DailyPayment
    db = SessionLocal()
    try:
        payment = db.query(DailyPayment).filter(DailyPayment.id == payment_id).first()
        if not payment:
            await message.reply_text("⚠️ Không tìm thấy phiếu chi trong hệ thống.", parse_mode=ParseMode.HTML)
            return
            
        if payment.status != "PENDING":
            await message.reply_text(f"⚠️ Phiếu chi này đang ở trạng thái: <b>{payment.status}</b>", parse_mode=ParseMode.HTML)
            return
            
        inv = db.query(Investment).filter(Investment.id == payment.investment_id).first()
        if not inv:
            await message.reply_text("⚠️ Không tìm thấy thông tin khoản đầu tư.", parse_mode=ParseMode.HTML)
            return
            
        # Update status and totals
        payment.status = "APPROVED"
        inv.total_expense = (inv.total_expense or 0) + payment.amount
        inv.profit = (inv.total_income or 0) - (inv.total_expense or 0)
        
        db.commit()
        
        success_msg = (
            f"✅ <b>ĐÃ PHÊ DUYỆT PHIẾU CHI</b>\n\n"
            f"<b>Mã Phiếu:</b> <code>{payment.id}</code>\n"
            f"<b>Người duyệt:</b> {message.from_user.first_name}\n"
            f"<b>Số Tiền:</b> <code>{fmt_vn(payment.amount)}</code>\n"
            f"<b>Đã giải ngân hạch toán thành công!</b>"
        )
        await message.reply_text(success_msg, parse_mode=ParseMode.HTML)
        LogInfo(f"[TienNga] Owner {message.from_user.id} approved payment {payment.id}", LogType.SYSTEM_STATUS)
        
    except Exception as e:
        db.rollback()
        LogError(f"Error in confirm_payment: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Hệ thống gặp lỗi khi duyệt phiếu chi.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


@bot.on_message(filters.command(["deny_payment"]))
@require_user_type(UserType.OWNER)
@require_project_name("Tiến Nga")
@require_group_role("main")
async def tien_nga_deny_payment_handler(client, message: Message) -> None:
    if not message.reply_to_message:
        await message.reply_text("⚠️ Vui lòng <b>Reply</b> một tin nhắn yêu cầu duyệt phiếu chi.", parse_mode=ParseMode.HTML)
        return
        
    replied_text = message.reply_to_message.text or ""
    # Extract Payment ID
    match = re.search(r"Mã Phiếu:\s*([a-f0-9\-]{36})", replied_text)
    if not match:
        await message.reply_text("⚠️ Không tìm thấy Mã Phiếu trong tin nhắn đang reply.", parse_mode=ParseMode.HTML)
        return
        
    payment_id = match.group(1)
    
    from app.models.business import DailyPayment
    db = SessionLocal()
    try:
        payment = db.query(DailyPayment).filter(DailyPayment.id == payment_id).first()
        if not payment:
            await message.reply_text("⚠️ Không tìm thấy phiếu chi trong hệ thống.", parse_mode=ParseMode.HTML)
            return
            
        if payment.status != "PENDING":
            await message.reply_text(f"⚠️ Phiếu chi này đang ở trạng thái: <b>{payment.status}</b>", parse_mode=ParseMode.HTML)
            return
            
        payment.status = "REJECTED"
        db.commit()
        
        await message.reply_text(f"❌ <b>ĐÃ TỪ CHỐI PHIẾU CHI</b>\n\nMã Phiếu: <code>{payment.id}</code>", parse_mode=ParseMode.HTML)
        LogInfo(f"[TienNga] Owner {message.from_user.id} rejected payment {payment.id}", LogType.SYSTEM_STATUS)
        
    except Exception as e:
        db.rollback()
        LogError(f"Error in deny_payment: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Hệ thống gặp lỗi khi huỷ phiếu chi.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# =========================================================================================
# XÁC NHẬN THANH TOÁN CÔNG NỢ HÀNG LOẠT
# =========================================================================================

# In-memory toggle selections keyed by message_id
_cpd_selections: dict[int, set[str]] = {}
_cpd_pages: dict[int, int] = {}  # current page per message_id
CPD_PAGE_SIZE = 10


@bot.on_message(filters.command(["tien_nga_confirm_payment_debt", "tien_nga_xn_thanh_toan_cong_no"]) | filters.regex(r"^@\w+\s+/(tien_nga_confirm_payment_debt|tien_nga_xn_thanh_toan_cong_no)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
async def tien_nga_confirm_payment_debt_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_confirm_payment_debt", "tien_nga_xn_thanh_toan_cong_no"])
    if args is None: return

    buttons = [
        [InlineKeyboardButton("Nhân sự", callback_data="cpd_cat_hr")],
        [InlineKeyboardButton("Nhà cung cấp", callback_data="cpd_cat_supplier")],
        [InlineKeyboardButton("Hủy", callback_data="cpd_cancel")],
    ]
    await message.reply_text(
        "<b>XÁC NHẬN THANH TOÁN CÔNG NỢ</b>\n\nVui lòng chọn đối tượng:",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode=ParseMode.HTML
    )


# --- Cancel ---
@bot.on_callback_query(filters.regex(r"^cpd_cancel$"))
async def cpd_cancel_callback(client, callback_query: CallbackQuery):
    msg_id = callback_query.message.id
    _cpd_selections.pop(msg_id, None)
    await callback_query.message.edit_text("❌ Đã hủy xác nhận thanh toán công nợ.")


# --- Back to main menu ---
@bot.on_callback_query(filters.regex(r"^cpd_back$"))
async def cpd_back_callback(client, callback_query: CallbackQuery):
    msg_id = callback_query.message.id
    _cpd_selections.pop(msg_id, None)
    buttons = [
        [InlineKeyboardButton("Nhân sự", callback_data="cpd_cat_hr")],
        [InlineKeyboardButton("Nhà cung cấp", callback_data="cpd_cat_supplier")],
        [InlineKeyboardButton("Hủy", callback_data="cpd_cancel")],
    ]
    await callback_query.message.edit_text(
        "<b>XÁC NHẬN THANH TOÁN CÔNG NỢ</b>\n\nVui lòng chọn đối tượng:",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode=ParseMode.HTML
    )


# --- Category: HR (Nhân sự) ---
@bot.on_callback_query(filters.regex(r"^cpd_cat_hr$"))
async def cpd_cat_hr_callback(client, callback_query: CallbackQuery):
    from app.models.employee import Employee

    db = SessionLocal()
    try:
        employees = db.query(Employee).filter(Employee.status == "ACTIVE").order_by(Employee.id).all()
        if not employees:
            await callback_query.answer("Không có nhân viên nào trong hệ thống.", show_alert=True)
            return

        msg_id = callback_query.message.id
        _cpd_selections[msg_id] = set()

        buttons = []
        row = []
        for emp in employees:
            name = f"{emp.last_name or ''} {emp.first_name or ''}".strip() or emp.id
            debt_str = f" ({fmt_vn(emp.total_debt or 0)})" if (emp.total_debt or 0) > 0 else ""
            label = f"☐ {name}{debt_str}"
            row.append(InlineKeyboardButton(label, callback_data=f"cpd_tgl_hr_{emp.id}"))
            if len(row) == 2:
                buttons.append(row)
                row = []
        if row:
            buttons.append(row)

        buttons.append([
            InlineKeyboardButton("✅ Xác nhận thanh toán", callback_data="cpd_confirm_hr"),
        ])
        buttons.append([
            InlineKeyboardButton("Quay lại", callback_data="cpd_back"),
            InlineKeyboardButton("Hủy", callback_data="cpd_cancel"),
        ])

        await callback_query.message.edit_text(
            "<b>CHỌN NHÂN SỰ CẦN THANH TOÁN CÔNG NỢ</b>\n\n"
            "<i>Nhấn vào tên để chọn/bỏ chọn, sau đó nhấn Xác nhận.</i>",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        LogError(f"Error in cpd_cat_hr_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống.", show_alert=True)
    finally:
        db.close()


# --- Toggle HR ---
@bot.on_callback_query(filters.regex(r"^cpd_tgl_hr_(.+)$"))
async def cpd_tgl_hr_callback(client, callback_query: CallbackQuery):
    emp_id = callback_query.matches[0].group(1)
    msg_id = callback_query.message.id

    if msg_id not in _cpd_selections:
        _cpd_selections[msg_id] = set()

    selected = _cpd_selections[msg_id]
    if emp_id in selected:
        selected.discard(emp_id)
    else:
        selected.add(emp_id)

    # Rebuild buttons from DB
    from app.models.employee import Employee
    db = SessionLocal()
    try:
        employees = db.query(Employee).filter(Employee.status == "ACTIVE").order_by(Employee.id).all()
        buttons = []
        row = []
        for emp in employees:
            name = f"{emp.last_name or ''} {emp.first_name or ''}".strip() or emp.id
            debt_str = f" ({fmt_vn(emp.total_debt or 0)})" if (emp.total_debt or 0) > 0 else ""
            is_selected = emp.id in selected
            icon = "☑" if is_selected else "☐"
            label = f"{icon} {name}{debt_str}"
            row.append(InlineKeyboardButton(label, callback_data=f"cpd_tgl_hr_{emp.id}"))
            if len(row) == 2:
                buttons.append(row)
                row = []
        if row:
            buttons.append(row)

        buttons.append([
            InlineKeyboardButton(f"✅ Xác nhận thanh toán ({len(selected)})", callback_data="cpd_confirm_hr"),
        ])
        buttons.append([
            InlineKeyboardButton("Quay lại", callback_data="cpd_back"),
            InlineKeyboardButton("Hủy", callback_data="cpd_cancel"),
        ])

        await callback_query.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(buttons))
    except Exception as e:
        if "MESSAGE_NOT_MODIFIED" in str(e):
            await callback_query.answer()
        elif "FLOOD_WAIT" in str(e):
            await callback_query.answer("⏳ Vui lòng chờ một chút.", show_alert=False)
        else:
            LogError(f"Error in cpd_tgl_hr_callback: {e}", LogType.SYSTEM_STATUS)
    finally:
        db.close()


# --- Confirm HR ---
@bot.on_callback_query(filters.regex(r"^cpd_confirm_hr$"))
async def cpd_confirm_hr_callback(client, callback_query: CallbackQuery):
    msg_id = callback_query.message.id
    selected = _cpd_selections.pop(msg_id, set())

    if not selected:
        await callback_query.answer("⚠️ Chưa chọn nhân sự nào!", show_alert=True)
        return

    from app.models.employee import Employee
    db = SessionLocal()
    try:
        success_list = []
        skip_list = []

        for emp_id in selected:
            emp = db.query(Employee).filter(Employee.id == emp_id).first()
            if not emp:
                continue
            name = f"{emp.last_name or ''} {emp.first_name or ''}".strip() or emp.id
            old_debt = emp.total_debt or 0
            if old_debt > 0:
                emp.total_debt = 0
                success_list.append(f"• <b>{name}</b> ({emp.id}): <code>{fmt_vn(old_debt)}</code> → <code>0</code>")
            else:
                skip_list.append(f"• <b>{name}</b> ({emp.id}): Công nợ = <code>{fmt_vn(old_debt)}</code>")

        db.commit()

        result = "<b>✅ KẾT QUẢ THANH TOÁN CÔNG NỢ NHÂN SỰ</b>\n\n"
        if success_list:
            result += "<b>Đã thanh toán:</b>\n" + "\n".join(success_list) + "\n\n"
        if skip_list:
            result += "<b>⚠️ Không có công nợ (bỏ qua):</b>\n" + "\n".join(skip_list) + "\n"

        await callback_query.message.edit_text(result, parse_mode=ParseMode.HTML)
        LogInfo(f"[TienNga] confirm_payments_of_debts HR: {len(success_list)} cleared, {len(skip_list)} skipped by {callback_query.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in cpd_confirm_hr_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.message.edit_text("❌ Lỗi hệ thống khi thanh toán công nợ nhân sự.")
    finally:
        db.close()


# --- Category: Supplier (Nhà cung cấp) → Chọn Xưởng ---
@bot.on_callback_query(filters.regex(r"^cpd_cat_supplier$"))
async def cpd_cat_supplier_callback(client, callback_query: CallbackQuery):
    from app.models.business import CollectionPoint

    db = SessionLocal()
    try:
        cps = db.query(CollectionPoint).all()
        if not cps:
            await callback_query.answer("Không có xưởng/điểm thu mua nào.", show_alert=True)
            return

        buttons = []
        for cp in cps:
            if cp.collection_name:
                buttons.append([InlineKeyboardButton(cp.collection_name, callback_data=f"cpd_cp_{cp.id}")])

        buttons.append([
            InlineKeyboardButton("Quay lại", callback_data="cpd_back"),
            InlineKeyboardButton("Hủy", callback_data="cpd_cancel"),
        ])

        await callback_query.message.edit_text(
            "<b>CHỌN XƯỞNG / ĐIỂM THU MUA</b>\n\n"
            "Vui lòng chọn xưởng để xem danh sách hộ dân:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        LogError(f"Error in cpd_cat_supplier_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống.", show_alert=True)
    finally:
        db.close()


# --- Helper: build paginated customer buttons ---
def _build_sup_buttons(customers, selected: set, cp_id: str, page: int):
    """Build paginated inline buttons for customer list."""
    total = len(customers)
    total_pages = max(1, (total + CPD_PAGE_SIZE - 1) // CPD_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))

    start = page * CPD_PAGE_SIZE
    end = start + CPD_PAGE_SIZE
    page_customers = customers[start:end]

    buttons = []
    for cust in page_customers:
        name = cust.fullname or cust.hoursehold_id
        debt_str = f" ({fmt_vn(cust.total_debt or 0)})" if (cust.total_debt or 0) > 0 else ""
        is_selected = cust.hoursehold_id in selected
        icon = "☑" if is_selected else "☐"
        label = f"{icon} {cust.hoursehold_id} - {name}{debt_str}"
        buttons.append([InlineKeyboardButton(label, callback_data=f"cpd_tgl_sup_{cp_id}_{cust.hoursehold_id}")])

    # Pagination nav
    if total_pages > 1:
        nav_row = []
        if page > 0:
            nav_row.append(InlineKeyboardButton("◀️ Trước", callback_data=f"cpd_page_{cp_id}_{page - 1}"))
        nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="cpd_noop"))
        if page < total_pages - 1:
            nav_row.append(InlineKeyboardButton("Sau ▶️", callback_data=f"cpd_page_{cp_id}_{page + 1}"))
        buttons.append(nav_row)

    sel_count = len(selected)
    confirm_label = f"✅ Xác nhận thanh toán ({sel_count})" if sel_count else "✅ Xác nhận thanh toán"
    buttons.append([InlineKeyboardButton(confirm_label, callback_data=f"cpd_confirm_sup_{cp_id}")])
    buttons.append([
        InlineKeyboardButton("Quay lại", callback_data="cpd_back_sup"),
        InlineKeyboardButton("Hủy", callback_data="cpd_cancel"),
    ])
    return buttons


# --- Select Collection Point → Show customers (page 0) ---
@bot.on_callback_query(filters.regex(r"^cpd_cp_(.+)$"))
async def cpd_cp_callback(client, callback_query: CallbackQuery):
    cp_id = callback_query.matches[0].group(1)
    from app.models.business import Customers, CollectionPoint

    db = SessionLocal()
    try:
        cp = db.query(CollectionPoint).filter(CollectionPoint.id == cp_id).first()
        cp_name = cp.collection_name if cp else "N/A"

        customers = db.query(Customers).filter(
            Customers.collection_point_id == cp_id,
            Customers.status == "ACTIVE",
            Customers.total_debt > 0
        ).order_by(Customers.hoursehold_id).all()

        if not customers:
            await callback_query.answer(f"Xưởng {cp_name} không có hộ dân nào có công nợ.", show_alert=True)
            return

        msg_id = callback_query.message.id
        _cpd_selections[msg_id] = set()
        _cpd_pages[msg_id] = 0

        buttons = _build_sup_buttons(customers, set(), cp_id, 0)

        await callback_query.message.edit_text(
            f"<b>CHỌN HỘ DÂN — {cp_name}</b>\n\n"
            "<i>Nhấn vào tên để chọn/bỏ chọn, sau đó nhấn Xác nhận.</i>",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        LogError(f"Error in cpd_cp_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống.", show_alert=True)
    finally:
        db.close()


# --- Page navigation ---
@bot.on_callback_query(filters.regex(r"^cpd_page_(.+)_(\d+)$"))
async def cpd_page_callback(client, callback_query: CallbackQuery):
    cp_id = callback_query.matches[0].group(1)
    page = int(callback_query.matches[0].group(2))
    msg_id = callback_query.message.id

    _cpd_pages[msg_id] = page
    selected = _cpd_selections.get(msg_id, set())

    from app.models.business import Customers, CollectionPoint
    db = SessionLocal()
    try:
        customers = db.query(Customers).filter(
            Customers.collection_point_id == cp_id,
            Customers.status == "ACTIVE",
            Customers.total_debt > 0
        ).order_by(Customers.hoursehold_id).all()

        buttons = _build_sup_buttons(customers, selected, cp_id, page)
        await callback_query.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(buttons))
    except Exception as e:
        if "MESSAGE_NOT_MODIFIED" in str(e):
            await callback_query.answer()
        elif "FLOOD_WAIT" in str(e):
            await callback_query.answer("⏳ Vui lòng chờ một chút.", show_alert=False)
        else:
            LogError(f"Error in cpd_page_callback: {e}", LogType.SYSTEM_STATUS)
    finally:
        db.close()


# --- Noop (page indicator button) ---
@bot.on_callback_query(filters.regex(r"^cpd_noop$"))
async def cpd_noop_callback(client, callback_query: CallbackQuery):
    await callback_query.answer()


# --- Back to supplier list ---
@bot.on_callback_query(filters.regex(r"^cpd_back_sup$"))
async def cpd_back_sup_callback(client, callback_query: CallbackQuery):
    msg_id = callback_query.message.id
    _cpd_selections.pop(msg_id, None)
    _cpd_pages.pop(msg_id, None)

    from app.models.business import CollectionPoint
    db = SessionLocal()
    try:
        cps = db.query(CollectionPoint).all()
        buttons = []
        for cp in cps:
            if cp.collection_name:
                buttons.append([InlineKeyboardButton(cp.collection_name, callback_data=f"cpd_cp_{cp.id}")])
        buttons.append([
            InlineKeyboardButton("Quay lại", callback_data="cpd_back"),
            InlineKeyboardButton("Hủy", callback_data="cpd_cancel"),
        ])
        await callback_query.message.edit_text(
            "<b>CHỌN XƯỞNG / ĐIỂM THU MUA</b>\n\n"
            "Vui lòng chọn xưởng để xem danh sách hộ dân:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        LogError(f"Error in cpd_back_sup_callback: {e}", LogType.SYSTEM_STATUS)
    finally:
        db.close()


# --- Toggle Supplier (Customer) ---
@bot.on_callback_query(filters.regex(r"^cpd_tgl_sup_(.+)_(.+)$"))
async def cpd_tgl_sup_callback(client, callback_query: CallbackQuery):
    cp_id = callback_query.matches[0].group(1)
    cust_id = callback_query.matches[0].group(2)
    msg_id = callback_query.message.id

    if msg_id not in _cpd_selections:
        _cpd_selections[msg_id] = set()

    selected = _cpd_selections[msg_id]
    if cust_id in selected:
        selected.discard(cust_id)
    else:
        selected.add(cust_id)

    page = _cpd_pages.get(msg_id, 0)

    # Rebuild buttons
    from app.models.business import Customers, CollectionPoint
    db = SessionLocal()
    try:
        customers = db.query(Customers).filter(
            Customers.collection_point_id == cp_id,
            Customers.status == "ACTIVE",
            Customers.total_debt > 0
        ).order_by(Customers.hoursehold_id).all()

        buttons = _build_sup_buttons(customers, selected, cp_id, page)
        await callback_query.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(buttons))
    except Exception as e:
        if "MESSAGE_NOT_MODIFIED" in str(e):
            await callback_query.answer()
        elif "FLOOD_WAIT" in str(e):
            await callback_query.answer("⏳ Vui lòng chờ một chút.", show_alert=False)
        else:
            LogError(f"Error in cpd_tgl_sup_callback: {e}", LogType.SYSTEM_STATUS)
    finally:
        db.close()


# --- Confirm Supplier ---
@bot.on_callback_query(filters.regex(r"^cpd_confirm_sup_(.+)$"))
async def cpd_confirm_sup_callback(client, callback_query: CallbackQuery):
    cp_id = callback_query.matches[0].group(1)
    msg_id = callback_query.message.id
    selected = _cpd_selections.pop(msg_id, set())

    if not selected:
        await callback_query.answer("⚠️ Chưa chọn hộ dân nào!", show_alert=True)
        return

    from app.models.business import Customers, CollectionPoint
    db = SessionLocal()
    try:
        cp = db.query(CollectionPoint).filter(CollectionPoint.id == cp_id).first()
        cp_name = cp.collection_name if cp else "N/A"

        success_list = []
        skip_list = []

        for cust_id in selected:
            cust = db.query(Customers).filter(Customers.hoursehold_id == cust_id).first()
            if not cust:
                continue
            name = cust.fullname or cust.hoursehold_id
            old_debt = cust.total_debt or 0
            if old_debt > 0:
                cust.total_debt = 0
                success_list.append(f"• <b>{cust.hoursehold_id}</b> — {name}: <code>{fmt_vn(old_debt)}</code> → <code>0</code>")
            else:
                skip_list.append(f"• <b>{cust.hoursehold_id}</b> — {name}: Công nợ = <code>{fmt_vn(old_debt)}</code>")

        db.commit()

        result = f"<b>✅ KẾT QUẢ THANH TOÁN CÔNG NỢ</b>\n<b>Xưởng:</b> {cp_name}\n\n"
        if success_list:
            result += "<b>Đã thanh toán:</b>\n" + "\n".join(success_list) + "\n\n"
        if skip_list:
            result += "<b>⚠️ Không có công nợ (bỏ qua):</b>\n" + "\n".join(skip_list) + "\n"

        await callback_query.message.edit_text(result, parse_mode=ParseMode.HTML)
        LogInfo(f"[TienNga] confirm_payments_of_debts Supplier ({cp_name}): {len(success_list)} cleared, {len(skip_list)} skipped by {callback_query.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        if "FLOOD_WAIT" in str(e):
            import asyncio
            await asyncio.sleep(5)
            try:
                await callback_query.message.edit_text(result, parse_mode=ParseMode.HTML)
            except Exception:
                pass
        else:
            LogError(f"Error in cpd_confirm_sup_callback: {e}", LogType.SYSTEM_STATUS)
            try:
                await callback_query.message.edit_text("❌ Lỗi hệ thống khi thanh toán công nợ nhà cung cấp.")
            except Exception:
                pass
    finally:
        db.close()


# =========================================================================================
# XUẤT BÁO CÁO THU CHI EXCEL
# =========================================================================================

@bot.on_message(filters.command(["tien_nga_export_daily_payment", "tien_nga_xuat_bao_cao_thu_chi"]) | filters.regex(r"^@\w+\s+/(tien_nga_export_daily_payment|tien_nga_xuat_bao_cao_thu_chi)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
async def tien_nga_export_daily_payment_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_export_daily_payment", "tien_nga_xuat_bao_cao_thu_chi"])
    if args is None: return

    # Check for direct dates: dd/mm/yyyy - dd/mm/yyyy
    if len(args) > 1:
        date_str = " ".join(args[1:])
        try:
            parts = date_str.split("-")
            if len(parts) != 2:
                raise ValueError()
            d1 = datetime.strptime(parts[0].strip(), "%d/%m/%Y").date()
            d2 = datetime.strptime(parts[1].strip(), "%d/%m/%Y").date()
            if d1 > d2:
                d1, d2 = d2, d1
            # Step 2: Show Investments
            await _show_edp_investments(client, message, d1, d2)
            return
        except ValueError:
            await message.reply_text("⚠️ Định dạng khoảng thời gian không hợp lệ. Vui lòng dùng: <code>DD/MM/YYYY - DD/MM/YYYY</code>", parse_mode=ParseMode.HTML)
            return

    # Step 1: Interactive Flow
    buttons = [
        [InlineKeyboardButton("1 ngày", callback_data="edp_time_1d"), InlineKeyboardButton("7 ngày", callback_data="edp_time_7d")],
        [InlineKeyboardButton("14 ngày", callback_data="edp_time_14d"), InlineKeyboardButton("1 tháng", callback_data="edp_time_1m")],
        [InlineKeyboardButton("2 tháng", callback_data="edp_time_2m"), InlineKeyboardButton("3 tháng", callback_data="edp_time_3m")],
        [InlineKeyboardButton("6 tháng", callback_data="edp_time_6m"), InlineKeyboardButton("1 năm", callback_data="edp_time_1y")],
        [InlineKeyboardButton("Năm trước", callback_data="edp_time_lasty"), InlineKeyboardButton("Tất cả", callback_data="edp_time_all")],
        [InlineKeyboardButton("Hủy", callback_data="edp_cancel")]
    ]
    await message.reply_text("<b>XUẤT BÁO CÁO THU CHI</b>\n<code>tien_nga_export_daily_payment [DD/MM/YYYY - DD/MM/YYYY]</code>\nVui lòng chọn khoảng thời gian:", reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)

@bot.on_callback_query(filters.regex(r"^edp_cancel$"))
async def edp_cancel_cb(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("Đã hủy thao tác xuất báo cáo.")

@bot.on_callback_query(filters.regex(r"^edp_time_(.+)$"))
async def edp_time_cb(client, callback_query: CallbackQuery):
    period = callback_query.matches[0].group(1)
    
    today = datetime.now().date()
    end_date = today
    if period == "1d": start_date = today
    elif period == "7d": start_date = today - timedelta(days=7)
    elif period == "14d": start_date = today - timedelta(days=14)
    elif period == "1m": start_date = today - timedelta(days=30)
    elif period == "2m": start_date = today - timedelta(days=60)
    elif period == "3m": start_date = today - timedelta(days=90)
    elif period == "6m": start_date = today - timedelta(days=180)
    elif period == "1y": start_date = today - timedelta(days=365)
    elif period == "lasty":
        start_date = datetime(today.year - 1, 1, 1).date()
        end_date = datetime(today.year - 1, 12, 31).date()
    elif period == "all":
        start_date = datetime(2000, 1, 1).date()
        end_date = datetime(2100, 1, 1).date()
    else:
        start_date = today
        
    await _show_edp_investments(client, callback_query.message, start_date, end_date, edit=True)

async def _show_edp_investments(client, message, start_date, end_date, edit=False):
    from app.models.business import Investment
    db = SessionLocal()
    try:
        investments = db.query(Investment).all()
        buttons = []
        # Support ALL
        buttons.append([InlineKeyboardButton("TỔNG HỢP TẤT CẢ", callback_data=f"edp_inv_ALL_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}")])
        for inv in investments:
            label = f"[{inv.investment_code}] {inv.name}" if inv.investment_code else inv.name
            if not label: label = str(inv.id)[:8]
            buttons.append([InlineKeyboardButton(label, callback_data=f"edp_inv_{inv.id}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}")])
        
        buttons.append([InlineKeyboardButton("Hủy", callback_data="edp_cancel")])
        
        text = f"<b>XUẤT BÁO CÁO TỪ {start_date.strftime('%d/%m/%Y')} ĐẾN {end_date.strftime('%d/%m/%Y')}</b>\n\nVui lòng chọn Quỹ Đầu Tư:"
        if edit:
            await message.edit_text(text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
        else:
            await message.reply_text(text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^edp_inv_(.+)_(.+)_(.+)$"))
async def edp_inv_cb(client, callback_query: CallbackQuery):
    inv_id = callback_query.matches[0].group(1)
    d1_str = callback_query.matches[0].group(2)
    d2_str = callback_query.matches[0].group(3)
    
    start_date = datetime.strptime(d1_str, "%Y%m%d").date()
    end_date = datetime.strptime(d2_str, "%Y%m%d").date()
    
    await callback_query.message.edit_text("⏳ Đang tổng hợp dữ liệu, vui lòng chờ...", parse_mode=ParseMode.HTML)
    
    from app.models.business import DailyPayment, Investment
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import os, tempfile
    
    db = SessionLocal()
    try:
        query = db.query(DailyPayment, Investment).outerjoin(Investment, DailyPayment.investment_id == Investment.id).filter(
            DailyPayment.day >= start_date,
            DailyPayment.day <= end_date
        )
        if inv_id != "ALL":
            import uuid as _uuid
            try:
                parsed_id = _uuid.UUID(inv_id)
                query = query.filter(DailyPayment.investment_id == parsed_id)
            except ValueError:
                pass # skip filtering if bad uuid
            
        payments = query.order_by(DailyPayment.day.asc(), DailyPayment.payment_type.asc()).all()
        
        if not payments:
            await callback_query.message.edit_text("⚠️ Không có dữ liệu trong khoảng thời gian và quỹ này.", parse_mode=ParseMode.HTML)
            return
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Báo Cáo Thu Chi"
        
        headers = ["Ngày", "Mã Phiếu", "Tên Đầu Tư", "Loại", "Số Tiền", "Trạng thái", "Người Yêu Cầu", "Người Thực Hiện", "Người Nhận", "Mục Đích", "Lý Do", "Ghi Chú"]
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5496")
        for col, val in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        total_thu = 0
        total_chi = 0
        
        for row_idx, (p, inv) in enumerate(payments, 2):
            inv_name = f"[{inv.investment_code}] {inv.name}" if inv and getattr(inv, "investment_code", None) else (inv.name if inv else "N/A")
            ptype = p.payment_type.upper() if p.payment_type else ""
            amt = p.amount or 0
            
            if p.status == "APPROVED":
                if ptype == "THU": total_thu += amt
                elif ptype == "CHI": total_chi += amt
            
            row_data = [
                p.day.strftime("%d/%m/%Y") if p.day else "",
                str(p.id)[:8],
                inv_name,
                ptype,
                amt,
                p.status or "APPROVED",
                p.requester or "",
                p.executor or "",
                p.receiver or "",
                p.purpose or "",
                p.reason or "",
                p.notes or ""
            ]
            ws.append(row_data)
            
            # Format currency
            ws.cell(row=row_idx, column=5).number_format = '#,##0'
            
            # Add color for THU / CHI
            if ptype == "THU":
                ws.cell(row=row_idx, column=4).font = Font(color="00B050", bold=True)
            elif ptype == "CHI":
                ws.cell(row=row_idx, column=4).font = Font(color="C00000", bold=True)
                
            if getattr(p, "status", "APPROVED") != "APPROVED":
                ws.cell(row=row_idx, column=6).font = Font(color="ED7D31", bold=False)
                
        # Append sum rows
        r = len(payments) + 2
        ws.cell(row=r, column=4, value="Tổng Thu (Approved):").font = Font(bold=True, color="00B050")
        ws.cell(row=r, column=5, value=total_thu).number_format = '#,##0'
        
        ws.cell(row=r+1, column=4, value="Tổng Chi (Approved):").font = Font(bold=True, color="C00000")
        ws.cell(row=r+1, column=5, value=total_chi).number_format = '#,##0'
        
        ws.cell(row=r+2, column=4, value="Chênh Lệch:").font = Font(bold=True)
        ws.cell(row=r+2, column=5, value=total_thu - total_chi).number_format = '#,##0'
        
        # Format font bold for sum values
        for i in range(3):
            ws.cell(row=r+i, column=5).font = Font(bold=True)
        
        # Auto adjust width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="export_daily_payment_")
        os.close(fd)
        wb.save(temp_path)
        
        # Send file
        file_name = f"export_daily_payment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        await client.send_document(
            chat_id=callback_query.message.chat.id,
            document=temp_path,
            file_name=file_name,
            caption=f"<b>BÁO CÁO THU CHI</b>\nTừ: {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}",
            parse_mode=ParseMode.HTML
        )
        
        os.remove(temp_path)
        await callback_query.message.delete()
        LogInfo(f"[TienNga] {callback_query.from_user.id} exported daily payments.", LogType.SYSTEM_STATUS)
        
    except Exception as e:
        LogError(f"Error exporting daily payments: {e}", LogType.SYSTEM_STATUS)
        await callback_query.message.edit_text("❌ Có lỗi xảy ra khi xuất báo cáo.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# =========================================================================================
# BIỂU ĐỒ THU CHI HÀNG NGÀY
# =========================================================================================

@bot.on_message(filters.command(["tien_nga_chart_daily_payment", "tien_nga_bieu_do_thu_chi"]) | filters.regex(r"^@\w+\s+/(tien_nga_chart_daily_payment|tien_nga_bieu_do_thu_chi)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
async def tien_nga_chart_daily_payment_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["tien_nga_chart_daily_payment", "tien_nga_bieu_do_thu_chi"])
    if args is None: return

    # Check for direct dates: dd/mm/yyyy - dd/mm/yyyy
    if len(args) > 1:
        date_str = " ".join(args[1:])
        try:
            parts = date_str.split("-")
            if len(parts) != 2: raise ValueError()
            d1 = datetime.strptime(parts[0].strip(), "%d/%m/%Y").date()
            d2 = datetime.strptime(parts[1].strip(), "%d/%m/%Y").date()
            if d1 > d2: d1, d2 = d2, d1
            await _show_cdp_investments(client, message, d1, d2)
            return
        except ValueError:
            await message.reply_text("⚠️ Định dạng khoảng thời gian không hợp lệ. Vui lòng dùng: <code>DD/MM/YYYY - DD/MM/YYYY</code>", parse_mode=ParseMode.HTML)
            return

    # Interactive Flow
    buttons = [
        [InlineKeyboardButton("1 tháng", callback_data="cdp_time_1m"), InlineKeyboardButton("3 tháng", callback_data="cdp_time_3m")],
        [InlineKeyboardButton("6 tháng", callback_data="cdp_time_6m"), InlineKeyboardButton("1 năm", callback_data="cdp_time_1y")],
        [InlineKeyboardButton("Hủy", callback_data="cdp_cancel")]
    ]
    await message.reply_text("<b>BIỂU ĐỒ THU CHI HÀNG NGÀY</b>\n<code>/tien_nga_chart_daily_payment [DD/MM/YYYY - DD/MM/YYYY]</code>\nVui lòng chọn khoảng thời gian:", reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)

@bot.on_callback_query(filters.regex(r"^cdp_cancel$"))
async def cdp_cancel_cb(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("Đã hủy thao tác vẽ biểu đồ.")

@bot.on_callback_query(filters.regex(r"^cdp_time_(.+)$"))
async def cdp_time_cb(client, callback_query: CallbackQuery):
    period = callback_query.matches[0].group(1)
    
    today = datetime.now().date()
    end_date = today
    if period == "1m": start_date = today - timedelta(days=30)
    elif period == "3m": start_date = today - timedelta(days=90)
    elif period == "6m": start_date = today - timedelta(days=180)
    elif period == "1y": start_date = today - timedelta(days=365)
    else: start_date = today
        
    await _show_cdp_investments(client, callback_query.message, start_date, end_date, edit=True)

async def _show_cdp_investments(client, message, start_date, end_date, edit=False):
    from app.models.business import Investment
    db = SessionLocal()
    try:
        investments = db.query(Investment).filter(Investment.status == "ACTIVE").all()
        buttons = []
        for inv in investments:
            label = f"[{inv.investment_code}] {inv.name}" if inv.investment_code else inv.name
            if not label: label = str(inv.id)[:8]
            buttons.append([InlineKeyboardButton(label, callback_data=f"cdp_inv_{inv.id}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}")])
        
        buttons.append([InlineKeyboardButton("Hủy", callback_data="cdp_cancel")])
        
        text = f"<b>BIỂU ĐỒ TỪ {start_date.strftime('%d/%m/%Y')} ĐẾN {end_date.strftime('%d/%m/%Y')}</b>\n\nVui lòng chọn Quỹ Đầu Tư:"
        if edit:
            await message.edit_text(text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
        else:
            await message.reply_text(text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^cdp_inv_(.+)_(.+)_(.+)$"))
async def cdp_inv_cb(client, callback_query: CallbackQuery):
    inv_id = callback_query.matches[0].group(1)
    d1_str = callback_query.matches[0].group(2)
    d2_str = callback_query.matches[0].group(3)
    
    start_date = datetime.strptime(d1_str, "%Y%m%d").date()
    end_date = datetime.strptime(d2_str, "%Y%m%d").date()
    
    await callback_query.message.edit_text("⏳ Đang vẽ biểu đồ, vui lòng chờ...", parse_mode=ParseMode.HTML)
    
    from app.models.business import DailyPayment, Investment
    from sqlalchemy import func, extract
    from collections import defaultdict
    import calendar
    import bot.utils.daily_payment_chart_generator as chart_gen
    
    db = SessionLocal()
    try:
        inv = db.query(Investment).filter(Investment.id == inv_id).first()
        inv_name = f"[{inv.investment_code}] {inv.name}" if inv and getattr(inv, "investment_code", None) else (inv.name if inv else "N/A")
        
        # Determine all months in range
        months = []
        current = datetime(start_date.year, start_date.month, 1).date()
        end_month_bound = datetime(end_date.year, end_date.month, 1).date()
        while current <= end_month_bound:
            months.append((current.year, current.month))
            next_month = current.month + 1 if current.month < 12 else 1
            next_year = current.year if current.month < 12 else current.year + 1
            current = datetime(next_year, next_month, 1).date()
            
        data = {
            "title": "BIỂU ĐỒ THU CHI HÀNG NGÀY",
            "subtitle": f"Quỹ Đầu Tư: {inv_name}\nThời gian: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            "charts": []
        }
        
        for y, m in months:
            # Query db for this specific month
            q_start = datetime(y, m, 1).date()
            _, last_day = calendar.monthrange(y, m)
            q_end = datetime(y, m, last_day).date()
            
            # Clip by absolute bounds
            if q_start < start_date: q_start = start_date
            if q_end > end_date: q_end = end_date
            
            records = db.query(
                DailyPayment.day,
                DailyPayment.payment_type,
                func.sum(DailyPayment.amount).label("total")
            ).filter(
                DailyPayment.investment_id == inv_id,
                DailyPayment.status == "APPROVED",
                DailyPayment.day >= q_start,
                DailyPayment.day <= q_end
            ).group_by(DailyPayment.day, DailyPayment.payment_type).all()
            
            # If no data and it's not the only month, we might still draw an empty chart or skip.
            # We'll draw anyway if it's within bounds so user sees it's zero.
            day_dict = defaultdict(lambda: {"thu": 0, "chi": 0})
            for r in records:
                day_val = r.day
                type_val = r.payment_type.upper()
                amount_val = float(r.total)
                if type_val == "THU": day_dict[day_val]["thu"] += amount_val
                elif type_val == "CHI": day_dict[day_val]["chi"] += amount_val
                
            labels = []
            thu_data = []
            chi_data = []
            
            curr_d = q_start
            while curr_d <= q_end:
                labels.append(curr_d.strftime("%d/%m"))
                thu_data.append(day_dict[curr_d]["thu"])
                chi_data.append(day_dict[curr_d]["chi"])
                curr_d += timedelta(days=1)
                
            data["charts"].append({
                "month_label": f"Tháng {m:02d}/{y}",
                "labels": labels,
                "thu": thu_data,
                "chi": chi_data
            })
            
        buf = await chart_gen.generate_daily_payment_chart_image(data)
        
        await client.send_photo(
            chat_id=callback_query.message.chat.id,
            photo=buf,
            caption=f"<b>BÁO CÁO BIỂU ĐỒ</b>\nQuỹ: {inv_name}\nTừ ngày {start_date.strftime('%d/%m/%Y')} tới {end_date.strftime('%d/%m/%Y')}",
            parse_mode=ParseMode.HTML
        )
        
        await callback_query.message.delete()
        LogInfo(f"[TienNga] {callback_query.from_user.id} generated daily payment chart for inv {inv_id}", LogType.SYSTEM_STATUS)
        
    except Exception as e:
        LogError(f"Error generating daily payment chart: {e}", LogType.SYSTEM_STATUS)
        await callback_query.message.edit_text("❌ Có lỗi hệ thống khi vẽ biểu đồ.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# =========================================================================================
# KIỂM TRA CHI TIẾT QUỸ ĐẦU TƯ
# =========================================================================================

@bot.on_message(filters.command(["tien_nga_check_investments", "tien_nga_kiem_tra_quy_dau_tu"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_investments|tien_nga_kiem_tra_quy_dau_tu)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN, UserType.MEMBER)
@require_project_name("Tiến Nga")
async def tien_nga_check_investments_handler(client, message: Message) -> None:
    from app.models.business import Investment
    from app.models.telegram import TelegramProjectMember

    chat_id = str(message.chat.id)

    db = SessionLocal()
    try:
        # Determine group role
        group_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()
        group_role = group_member.role if group_member else None
        group_name = group_member.group_name if group_member else None

        investments = db.query(Investment).all()
        if not investments:
            await message.reply_text("⚠️ Chưa có khoản đầu tư nào trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        # For member groups, filter investments that have a shareholder matching group_name
        if group_role == "member" and group_name:
            from app.models.business import Shareholder
            filtered_investments = []
            for inv in investments:
                sh = db.query(Shareholder).filter(
                    Shareholder.investment_id == inv.id,
                    Shareholder.telegram_group == group_name
                ).first()
                if sh:
                    filtered_investments.append(inv)
            investments = filtered_investments

            if not investments:
                await message.reply_text(
                    f"⚠️ Không tìm thấy Quỹ Đầu Tư nào liên kết với nhóm <b>{group_name}</b>.",
                    parse_mode=ParseMode.HTML
                )
                return

        buttons = []
        for inv in investments:
            label = f"[{inv.investment_code}] {inv.name}" if getattr(inv, "investment_code", None) else inv.name
            if not label: label = str(inv.id)[:8]
            
            if inv.status != "ACTIVE":
                label = f"{label} (ĐÃ ĐÓNG)"
                
            buttons.append([InlineKeyboardButton(label, callback_data=f"chk_inv_id_{inv.id}")])
            
        buttons.append([InlineKeyboardButton("Hủy", callback_data="chk_inv_cancel")])

        await message.reply_text(
            "<b>KIỂM TRA QUỸ ĐẦU TƯ</b>\n\nVui lòng chọn Quỹ để xem chi tiết:", 
            reply_markup=InlineKeyboardMarkup(buttons), 
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error checking inv list: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^chk_inv_cancel$"))
async def chk_inv_cancel_cb(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("Đã hủy thao tác tra cứu Quỹ Đầu Tư.")


@bot.on_callback_query(filters.regex(r"^chk_inv_id_(.+)$"))
async def chk_inv_cb(client, callback_query: CallbackQuery):
    raw_id = callback_query.matches[0].group(1)
    show_sh = False
    if raw_id.endswith("_showsh"):
        show_sh = True
        inv_id = raw_id.replace("_showsh", "")
    else:
        inv_id = raw_id
    
    from app.models.business import Investment, Shareholder
    from app.models.telegram import TelegramProjectMember

    chat_id = str(callback_query.message.chat.id)

    db = SessionLocal()
    try:
        if inv_id == "cancel": return # safety check
        
        # Determine group role
        group_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()
        group_role = group_member.role if group_member else None
        group_name = group_member.group_name if group_member else None
        is_member_group = (group_role == "member")

        import uuid as _uuid
        try:
            parsed_id = _uuid.UUID(inv_id)
            inv = db.query(Investment).filter(Investment.id == parsed_id).first()
        except ValueError:
            inv = db.query(Investment).filter(Investment.investment_code == inv_id).first()
            
        if not inv:
            await callback_query.answer("⚠️ Không tìm thấy thông tin Quỹ.", show_alert=True)
            return

        # For member groups, verify the shareholder exists for this group
        matched_shareholder = None
        if is_member_group and group_name:
            matched_shareholder = db.query(Shareholder).filter(
                Shareholder.investment_id == inv.id,
                Shareholder.telegram_group == group_name
            ).first()
            if not matched_shareholder:
                await callback_query.answer(
                    f"⚠️ Nhóm '{group_name}' không có cổ đông trong Quỹ này.", show_alert=True
                )
                return

        inv_code = getattr(inv, "investment_code", "") or "N/A"
        date_start_str = inv.start_date.strftime("%d/%m/%Y") if inv.start_date else "N/A"
        date_end_str = inv.end_date.strftime("%d/%m/%Y") if inv.end_date else "Chưa xác định"
        
        status_str = "ĐANG HOẠT ĐỘNG (ACTIVE)" if inv.status == "ACTIVE" else "ĐÃ ĐÓNG (CLOSED)"
        
        msg = (
            f"<b>THÔNG TIN QUỸ ĐẦU TƯ</b>\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"<b>Mã Quỹ:</b> <code>{inv_code}</code>\n"
            f"<b>Tên Quỹ:</b> {inv.name or 'N/A'}\n"
            f"<b>Trạng thái:</b> {status_str}\n\n"
            f"<b>Ngày Bắt Đầu:</b> {date_start_str}\n"
            f"<b>Ngày Kết Thúc:</b> {date_end_str}\n\n"
        )

        # Show financial info (both main and member groups)
        msg += (
            f"<b>Vốn Ban Đầu:</b> <code>{fmt_vn(inv.initial_capital or 0)}</code>\n"
            f"<b>Tổng Thu:</b> <code>{fmt_vn(inv.total_income or 0)}</code>\n"
            f"<b>Tổng Chi:</b> <code>{fmt_vn(inv.total_expense or 0)}</code>\n"
            f"<b>Lợi Nhuận / Số Dư:</b> <code>{fmt_vn(inv.profit or 0)}</code>\n\n"
            f"<b>Ghi chú:</b> {inv.notes or 'Không có'}\n"
        )

        # Show all shareholders
        shareholders = db.query(Shareholder).filter(Shareholder.investment_id == inv.id).order_by(Shareholder.shareholder_code).all()
        count_sh = len(shareholders)
        
        if show_sh:
            if shareholders:
                total_shareholder_amount = sum(sh.investment_amount or 0 for sh in shareholders)
                msg += f"\n━━━━━━━━━━━━━━━━━━\n"
                msg += f"<b>DANH SÁCH CỔ ĐÔNG ({count_sh})</b>\n\n"
                for idx, sh in enumerate(shareholders, 1):
                    start_str = sh.start_date.strftime('%d/%m/%Y') if sh.start_date else '—'
                    sh_amount = sh.investment_amount or 0
                    share_pct = (sh_amount / total_shareholder_amount * 100) if total_shareholder_amount > 0 else 0
                    msg += (
                        f"<b>{idx}. {sh.fullname}</b> (<code>{sh.shareholder_code}</code>)\n"
                        f"   Góp vốn: <code>{fmt_vn(sh_amount)}</code>\n"
                        f"   Cổ phần: <b>{share_pct:.1f}%</b>\n"
                        f"   Ngày BĐ: {start_str}\n"
                    )
                    if sh.username:
                        msg += f"   TG: @{sh.username}\n"
                    if sh.notes:
                        msg += f"   {sh.notes}\n"
                    msg += "\n"
            else:
                msg += f"\n━━━━━━━━━━━━━━━━━━\n"
                msg += "<i>Chưa có cổ đông nào.</i>\n"

        # In member group → also show transaction summary for the matched shareholder
        if is_member_group and matched_shareholder:
            from app.models.business import DailyPayment
            from sqlalchemy import or_, func

            sh_code = matched_shareholder.shareholder_code

            total_thu = db.query(func.coalesce(func.sum(DailyPayment.amount), 0)).filter(
                DailyPayment.investment_id == inv.id,
                DailyPayment.status == "APPROVED",
                DailyPayment.payment_type == "thu",
                or_(
                    DailyPayment.executor.ilike(f"%{sh_code}%"),
                    DailyPayment.receiver.ilike(f"%{sh_code}%"),
                    DailyPayment.requester.ilike(f"%{sh_code}%"),
                )
            ).scalar() or 0

            total_chi = db.query(func.coalesce(func.sum(DailyPayment.amount), 0)).filter(
                DailyPayment.investment_id == inv.id,
                DailyPayment.status == "APPROVED",
                DailyPayment.payment_type == "chi",
                or_(
                    DailyPayment.executor.ilike(f"%{sh_code}%"),
                    DailyPayment.receiver.ilike(f"%{sh_code}%"),
                    DailyPayment.requester.ilike(f"%{sh_code}%"),
                )
            ).scalar() or 0

            msg += (
                f"━━━━━━━━━━━━━━━━━━\n"
                f"<b>TỔNG KẾT GIAO DỊCH ({matched_shareholder.fullname})</b>\n\n"
                f"Tổng Thu: <code>{fmt_vn(total_thu)}</code>\n"
                f"Tổng Chi: <code>{fmt_vn(total_chi)}</code>\n"
                f"Chênh Lệch: <code>{fmt_vn(total_thu - total_chi)}</code>\n"
            )
        
        # Add buttons
        buttons = []
        if not show_sh and count_sh > 0:
            buttons.append([InlineKeyboardButton(f"Hiển thị Cổ đông ({count_sh})", callback_data=f"chk_inv_id_{inv.id}_showsh")])
        elif show_sh and count_sh > 0:
            buttons.append([InlineKeyboardButton("Ẩn Cổ đông", callback_data=f"chk_inv_id_{inv.id}")])
            
        buttons.append([
            InlineKeyboardButton("Quay Lại Danh Sách", callback_data="chk_inv_back"),
            InlineKeyboardButton("Hủy", callback_data="chk_inv_cancel")
        ])
        
        await callback_query.message.edit_text(msg, parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(buttons))
    except Exception as e:
        LogError(f"Error showing specific inv check: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi tải dữ liệu.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^chk_inv_back$"))
async def chk_inv_back_cb(client, callback_query: CallbackQuery):
    from app.models.business import Investment
    from app.models.telegram import TelegramProjectMember

    chat_id = str(callback_query.message.chat.id)

    db = SessionLocal()
    try:
        # Determine group role for filtering
        group_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()
        group_role = group_member.role if group_member else None
        group_name = group_member.group_name if group_member else None

        investments = db.query(Investment).all()

        # Filter for member groups
        if group_role == "member" and group_name:
            from app.models.business import Shareholder
            filtered = []
            for inv in investments:
                sh = db.query(Shareholder).filter(
                    Shareholder.investment_id == inv.id,
                    Shareholder.telegram_group == group_name
                ).first()
                if sh:
                    filtered.append(inv)
            investments = filtered

        buttons = []
        for inv in investments:
            label = f"[{inv.investment_code}] {inv.name}" if getattr(inv, "investment_code", None) else inv.name
            if not label: label = str(inv.id)[:8]
            if inv.status != "ACTIVE": label = f"{label} (ĐÃ ĐÓNG)"
            buttons.append([InlineKeyboardButton(label, callback_data=f"chk_inv_id_{inv.id}")])
            
        buttons.append([InlineKeyboardButton("Hủy", callback_data="chk_inv_cancel")])
        await callback_query.message.edit_text(
            "<b>KIỂM TRA QUỸ ĐẦU TƯ</b>\n\nVui lòng chọn Quỹ để xem chi tiết:", 
            reply_markup=InlineKeyboardMarkup(buttons), 
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error back inv check: {e}", LogType.SYSTEM_STATUS)
    finally:
        db.close()


# ==============================================================================
# HỆ THỐNG QUẢN LÝ KHO (INVENTORY MANAGEMENT)
# ==============================================================================

# --- TẠO KHO MỚI ---
@bot.on_message(filters.command(["tien_nga_create_inventory", "tien_nga_tao_kho"]) | filters.regex(r"^@\w+\s+/(tien_nga_create_inventory|tien_nga_tao_kho)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_INVENTORY, CustomTitle.MAIN_PRODUCT)
async def tien_nga_create_inventory_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        form = (
            "<b>FORM TẠO KHO MỚI</b>\n"
            "Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:\n\n"
            "<pre>/tien_nga_create_inventory\n"
            "Tên Nguyên Liệu: \n"
            "Tên Kho: \n"
            "Số Lượng Ban Đầu: 0\n"
            "Địa Chỉ Lưu Trữ: \n"
            "Sức Chứa: 0</pre>"
        )
        await message.reply_text(form, parse_mode=ParseMode.HTML)
        return

    data = {}
    for line in lines[1:]:
        if ":" in line:
            k, v = line.split(":", 1)
            data[k.strip()] = v.strip()

    material_name = data.get("Tên Nguyên Liệu", "").strip()
    storage_name = data.get("Tên Kho", "").strip()
    storage_loc = data.get("Địa Chỉ Lưu Trữ", "").strip()

    def parse_float_val(val_str):
        if not val_str: return 0.0
        val_str = val_str.replace(".", "").replace(",", ".").replace(" ", "")
        try: return float(val_str)
        except: return 0.0

    qty = parse_float_val(data.get("Số Lượng Ban Đầu", "0"))
    cap = parse_float_val(data.get("Sức Chứa", "0"))

    if not material_name:
        await message.reply_text("⚠️ Tên Nguyên Liệu là bắt buộc.")
        return

    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    db = SessionLocal()
    try:
        new_inv = Inventory(material_name=material_name, quantity=qty, storage_name=storage_name, storage_location=storage_loc, capacity=cap)
        db.add(new_inv)
        db.commit()
        await message.reply_text(f"✅ Đã tạo kho <b>{material_name}</b> thành công!", parse_mode=ParseMode.HTML)
    except Exception as e:
        db.rollback()
        await message.reply_text("❌ Lỗi khi thêm vào DB.")
    finally:
        db.close()

# --- DANH SÁCH KHO ---
@bot.on_message(filters.command(["tien_nga_list_inventory", "tien_nga_danh_sach_kho"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_inventory|tien_nga_danh_sach_kho)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_INVENTORY, CustomTitle.MAIN_PRODUCT)
async def tien_nga_list_inventory_handler(client, message: Message) -> None:
    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    db = SessionLocal()
    try:
        invs = db.query(Inventory).order_by(Inventory.material_name).all()
        if not invs:
            await message.reply_text("⚠️ Chưa có kho chứa nào.")
            return
        
        text = "<b>📦 DANH SÁCH HÀNG TỒN KHO</b>\n\n"
        for idx, inv in enumerate(invs, 1):
            text += f"<b>{idx}. {inv.material_name}</b>\n"
            text += f"   Tên kho: {inv.storage_name or '—'}\n"
            text += f"   Khối lượng hiện tại: {inv.quantity:,.0f} kg <i>(Sức chứa: {inv.capacity:,.0f} kg)</i>\n"
            text += f"   Địa chỉ: {inv.storage_location or '—'}\n\n"
        await message.reply_text(text, parse_mode=ParseMode.HTML)
    except Exception as e:
        await message.reply_text("❌ Lỗi hệ thống.")
    finally:
        db.close()

# --- CẬP NHẬT KHO ---
@bot.on_message(filters.command(["tien_nga_update_inventory", "tien_nga_cap_nhat_ton_kho"]) | filters.regex(r"^@\w+\s+/(tien_nga_update_inventory|tien_nga_cap_nhat_ton_kho)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_INVENTORY, CustomTitle.MAIN_PRODUCT)
async def tien_nga_update_inventory_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        from app.db.session import SessionLocal
        from app.models.inventory import Inventory
        db = SessionLocal()
        try:
            invs = db.query(Inventory).all()
            if not invs:
                await message.reply_text("⚠️ Chưa có hàng tồn kho nào.")
                return
            buttons = []
            for inv in invs:
                btn_text = f"{inv.material_name} ({inv.storage_name})" if inv.storage_name else inv.material_name
                buttons.append([InlineKeyboardButton(btn_text, callback_data=f"tn_selinv_{inv.id}")])
            buttons.append([InlineKeyboardButton("Hủy", callback_data="tn_selinv_cancel")])
            await message.reply_text("<b>Vui lòng chọn Hàng Tồn Kho cần cập nhật:</b>", reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    data = {}
    for line in lines[1:]:
        if ":" in line:
            k, v = line.split(":", 1)
            data[k.strip()] = v.strip()

    material_name = data.get("Tên Nguyên Liệu", "").strip()
    storage_name = data.get("Tên Kho", "").strip()
    storage_loc = data.get("Địa Chỉ Lưu Trữ", "").strip()

    def parse_float_val(val_str):
        if not val_str: return 0.0
        val_str = val_str.replace(".", "").replace(",", ".").replace(" ", "")
        try: return float(val_str)
        except: return 0.0

    qty = parse_float_val(data.get("Số Lượng", "0"))
    cap = parse_float_val(data.get("Sức Chứa", "0"))

    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    db = SessionLocal()
    try:
        inv = db.query(Inventory).filter(Inventory.material_name == material_name).first()
        if not inv:
            await message.reply_text("⚠️ Không tìm thấy nguyên liệu này trong hệ thống.")
            return
        if storage_name: inv.storage_name = storage_name
        if storage_loc: inv.storage_location = storage_loc
        inv.quantity = qty
        inv.capacity = cap
        db.commit()
        await message.reply_text(f"✅ Đã cập nhật kho <b>{material_name}</b> thành công!", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^tn_selinv_(.+)$"))
async def _sel_inv_cb(client, callback_query):
    inv_id = callback_query.matches[0].group(1)
    if inv_id == "cancel":
        await callback_query.message.edit_text("❌ <b>Đã hủy thao tác.</b>", parse_mode=ParseMode.HTML)
        return
    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    import uuid
    db = SessionLocal()
    try:
        inv = db.query(Inventory).filter(Inventory.id == uuid.UUID(inv_id)).first()
        if not inv:
            await callback_query.answer("Lỗi", show_alert=True)
            return
        form = (
            "<b>FORM CẬP NHẬT KHO</b>\n\n"
            "<pre>/tien_nga_update_inventory\n"
            f"Tên Nguyên Liệu: {inv.material_name}\n"
            f"Tên Kho: {inv.storage_name or ''}\n"
            f"Số Lượng: {inv.quantity:,.0f}\n"
            f"Địa Chỉ Lưu Trữ: {inv.storage_location or ''}\n"
            f"Sức Chứa: {inv.capacity:,.0f}</pre>"
        )
        await callback_query.message.reply_text(form.replace(",", "."), parse_mode=ParseMode.HTML)
        await callback_query.answer()
    finally:
        db.close()

# --- KIỂM TRA KHO ---
@bot.on_message(filters.command(["tien_nga_check_inventory", "tien_nga_kiem_tra_kho"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_inventory|tien_nga_kiem_tra_kho)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_INVENTORY, CustomTitle.MAIN_PRODUCT, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_check_inventory_handler(client, message: Message) -> None:
    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    db = SessionLocal()
    try:
        invs = db.query(Inventory).all()
        if not invs:
            await message.reply_text("⚠️ Chưa có kho nào trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        buttons = []
        for inv in invs:
            btn_text = f"{inv.material_name} ({inv.storage_name})" if inv.storage_name else inv.material_name
            buttons.append([InlineKeyboardButton(btn_text, callback_data=f"tn_chkinv_{inv.id}")])
        buttons.append([InlineKeyboardButton("Hủy", callback_data="tn_chkinv_cancel")])

        await message.reply_text(
            "<b>📦 KIỂM TRA KHO</b>\n\nVui lòng chọn kho để xem thông tin:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        from bot.utils.logger import LogError
        LogError(f"Error in tien_nga_check_inventory_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^tn_chkinv_(.+)$"))
async def _check_inv_cb(client, callback_query):
    inv_id = callback_query.matches[0].group(1)
    if inv_id == "cancel":
        await callback_query.message.edit_text("❌ <b>Đã hủy.</b>", parse_mode=ParseMode.HTML)
        return
    if inv_id == "back":
        from app.db.session import SessionLocal
        from app.models.inventory import Inventory
        db_back = SessionLocal()
        try:
            invs = db_back.query(Inventory).all()
            if not invs:
                await callback_query.message.edit_text("⚠️ Chưa có kho nào.", parse_mode=ParseMode.HTML)
                return
            buttons = []
            for inv in invs:
                btn_text = f"{inv.material_name} ({inv.storage_name})" if inv.storage_name else inv.material_name
                buttons.append([InlineKeyboardButton(btn_text, callback_data=f"tn_chkinv_{inv.id}")])
            buttons.append([InlineKeyboardButton("Hủy", callback_data="tn_chkinv_cancel")])
            await callback_query.message.edit_text(
                "<b>KIỂM TRA KHO</b>\n\nVui lòng chọn kho để xem thông tin:",
                reply_markup=InlineKeyboardMarkup(buttons),
                parse_mode=ParseMode.HTML
            )
        finally:
            db_back.close()
        return

    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    import uuid
    db = SessionLocal()
    try:
        inv = db.query(Inventory).filter(Inventory.id == uuid.UUID(inv_id)).first()
        if not inv:
            await callback_query.answer("⚠️ Không tìm thấy kho.", show_alert=True)
            return

        quantity = inv.quantity or 0
        capacity = inv.capacity or 0
        usage_pct = (quantity / capacity * 100) if capacity > 0 else 0

        # Progress bar
        filled = int(usage_pct / 5)
        bar = "█" * filled + "░" * (20 - filled)

        msg = (
            f"<b>THÔNG TIN KHO</b>\n\n"
            f"<b>Tên Kho:</b> {inv.storage_name or '—'}\n"
            f"<b>Nguyên Liệu:</b> {inv.material_name or '—'}\n"
            f"<b>Địa Chỉ:</b> {inv.storage_location or '—'}\n\n"
            f"<b>Tồn Kho:</b> <code>{quantity:,.0f}</code> kg\n"
            f"<b>Sức Chứa:</b> <code>{capacity:,.0f}</code> kg\n"
            f"<b>Sử Dụng:</b> <code>{usage_pct:.1f}%</code>\n"
            f"<code>{bar}</code>\n\n"
            f"🆔 <code>{inv.id}</code>"
        )

        buttons = [
            [InlineKeyboardButton("⬅️ Quay lại", callback_data="tn_chkinv_back")],
            [InlineKeyboardButton("Đóng", callback_data="tn_chkinv_cancel")]
        ]
        await callback_query.message.edit_text(
            msg,
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        from bot.utils.logger import LogError
        LogError(f"Error in _check_inv_cb: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^tn_chkinv_back$"))
async def _check_inv_back_cb(client, callback_query):
    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    db = SessionLocal()
    try:
        invs = db.query(Inventory).all()
        if not invs:
            await callback_query.message.edit_text("⚠️ Chưa có kho nào.", parse_mode=ParseMode.HTML)
            return

        buttons = []
        for inv in invs:
            btn_text = f"{inv.material_name} ({inv.storage_name})" if inv.storage_name else inv.material_name
            buttons.append([InlineKeyboardButton(btn_text, callback_data=f"tn_chkinv_{inv.id}")])
        buttons.append([InlineKeyboardButton("Hủy", callback_data="tn_chkinv_cancel")])

        await callback_query.message.edit_text(
            "<b>KIỂM TRA KHO</b>\n\nVui lòng chọn kho để xem thông tin:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        from bot.utils.logger import LogError
        LogError(f"Error in _check_inv_back_cb: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()

# --- THU MUA ---
@bot.on_message(filters.command(["tien_nga_material_purchase", "tien_nga_thu_mua_nguyen_lieu"]) | filters.regex(r"^@\w+\s+/(tien_nga_material_purchase|tien_nga_thu_mua_nguyen_lieu)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_INVENTORY, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_material_purchase_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        from app.db.session import SessionLocal
        from app.models.inventory import Inventory
        db = SessionLocal()
        try:
            invs = db.query(Inventory).all()
            if not invs:
                await message.reply_text("⚠️ Chưa có hàng tồn kho nào.")
                return
            buttons = []
            for inv in invs:
                btn_text = f"{inv.material_name} ({inv.storage_name})" if inv.storage_name else inv.material_name
                buttons.append([InlineKeyboardButton(btn_text, callback_data=f"tn_purmat_{inv.id}")])
            buttons.append([InlineKeyboardButton("Hủy", callback_data="tn_purmat_cancel")])
            await message.reply_text("<b>Vui lòng chọn Loại Nguyên Liệu cần nhập mua:</b>", reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    data = {}
    for line in lines[1:]:
        if ":" in line:
            k, v = line.split(":", 1)
            data[k.strip()] = v.strip()

    material_type = data.get("Loại Nguyên Liệu", "").strip()
    storage_name = data.get("Tên Kho", "").strip()
    transaction_date_str = data.get("Ngày Giao Dịch", "").strip()
    customer_id = data.get("Mã Khách Hàng", "").strip()
    notes = data.get("Ghi Chú", "").strip()

    def parse_float_val(val_str):
        if not val_str: return 0.0
        val_str = val_str.replace(".", "").replace(",", ".").replace(" ", "")
        try: return float(val_str)
        except: return 0.0

    trip_count = int(parse_float_val(data.get("Số Chuyến", "1")))
    weight = parse_float_val(data.get("Khối Lượng (Kg)", data.get("Khối Lượng", "0")))
    unit_price = parse_float_val(data.get("Đơn Giá (VNĐ)", data.get("Đơn Giá", "0")))
    total_amount = weight * unit_price
    advance_payment = parse_float_val(data.get("Tạm Ứng (VNĐ)", data.get("Tạm Ứng", "0")))
    parsed_debt = parse_float_val(data.get("Công Nợ (VNĐ)", data.get("Công Nợ", "0")))
    debt = parsed_debt if parsed_debt != 0 else (total_amount - advance_payment)

    from datetime import datetime
    try:
        transaction_date = datetime.strptime(transaction_date_str, "%d/%m/%Y").date() if transaction_date_str else datetime.now().date()
    except:
        transaction_date = datetime.now().date()

    from app.db.session import SessionLocal
    from app.models.inventory import Inventory, MaterialPurchase
    from app.models.business import Customers
    import uuid
    db = SessionLocal()
    try:
        inv = db.query(Inventory).filter(Inventory.material_name == material_type).first()
        if not inv:
            await message.reply_text("⚠️ Không tìm thấy nguyên liệu này.")
            return

        purchase = MaterialPurchase(
            id=uuid.uuid4(), transaction_date=transaction_date, customer_id=customer_id,
            material_type=material_type, storage_name=storage_name, trip_count=trip_count,
            weight=weight, unit_price=unit_price, total_amount=total_amount,
            advance_payment=advance_payment, debt=debt, notes=notes
        )
        db.add(purchase)
        
        old_quantity = inv.quantity
        inv.quantity = old_quantity + weight

        customer_display = customer_id or "—"
        new_customer_debt = None
        if customer_id:
            customer = db.query(Customers).filter(Customers.hoursehold_id == customer_id).first()
            if customer:
                if customer.total_debt is None: customer.total_debt = 0
                customer.total_debt += int(debt)
                new_customer_debt = customer.total_debt
                customer_display = f"{customer.hoursehold_id} - {customer.fullname}"
            else:
                customer_display = f"{customer_id} (Khách mới/Không có trên HT)"

        db.commit()

        def fmt_vn(val): return f"{val:,.0f}".replace(",", ".")
        msg = (
            f"<b>✅ NHẬP THU MUA THÀNH CÔNG</b>\n\n"
            f"<b>Nguyên Liệu:</b> {material_type}\n"
            f"<b>Tên Kho:</b> {storage_name or '—'}\n"
            f"<b>Ngày GD:</b> {transaction_date.strftime('%d/%m/%Y')}\n"
            f"<b>Khách Hàng:</b> {customer_display}\n"
            f"<b>Khối Lượng Nhập:</b> {weight:,.0f} kg\n"
            f"<b>Đơn Giá:</b> {fmt_vn(unit_price)} đ\n"
            f"<b>Thành Tiền:</b> {fmt_vn(total_amount)} đ\n"
            f"<b>Công Nợ Đợt Này:</b> {fmt_vn(debt)} đ\n"
        )
        if new_customer_debt is not None:
            msg += f"<b>Tổng Công Nợ KH:</b> {fmt_vn(new_customer_debt)} đ\n\n"
        else:
            msg += "\n"
            
        msg += f"📦 <b>Số dư kho hiện tại:</b> {inv.quantity:,.0f} kg <i>(+{weight:,.0f} kg)</i>"
        await message.reply_text(msg, parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^tn_purmat_(.+)$"))
async def _purmat_cb(client, callback_query):
    inv_id = callback_query.matches[0].group(1)
    if inv_id == "cancel":
        await callback_query.message.edit_text("❌ <b>Đã hủy.</b>", parse_mode=ParseMode.HTML)
        return
    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    import uuid
    from datetime import datetime
    db = SessionLocal()
    try:
        inv = db.query(Inventory).filter(Inventory.id == uuid.UUID(inv_id)).first()
        if not inv: return
        today_str = datetime.now().strftime("%d/%m/%Y")
        form = (
            "<b>FORM NHẬP THU MUA</b>\n\n"
            "<pre>/tien_nga_material_purchase\n"
            f"Loại Nguyên Liệu: {inv.material_name}\n"
            f"Tên Kho: {inv.storage_name or ''}\n"
            f"Ngày Giao Dịch: {today_str}\n"
            "Mã Khách Hàng: \n"
            "Số Chuyến: 1\n"
            "Khối Lượng (Kg): 0\n"
            "Đơn Giá (VNĐ): 0\n"
            "Ghi Chú: </pre>\n"
            "<i>Lưu ý: Thành tiền = Khối lượng x Đơn giá</i>\n"
            "<i>Lưu ý: Thành tiền, Tạm ứng và Công nợ sẽ được hệ thống xử lý ngầm, vui lòng không điền thêm.</i>"
        )
        await callback_query.message.reply_text(form, parse_mode=ParseMode.HTML)
        await callback_query.answer()
    finally:
        db.close()

# --- XUẤT KHO ---
@bot.on_message(filters.command(["tien_nga_export_inventory", "tien_nga_xuat_kho"]) | filters.regex(r"^@\w+\s+/(tien_nga_export_inventory|tien_nga_xuat_kho)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_INVENTORY, CustomTitle.MAIN_SUPPLIER)
async def tien_nga_export_inventory_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        from app.db.session import SessionLocal
        from app.models.inventory import Inventory
        db = SessionLocal()
        try:
            invs = db.query(Inventory).all()
            if not invs:
                await message.reply_text("⚠️ Chưa có hàng tồn kho nào.")
                return
            buttons = []
            for inv in invs:
                btn_text = f"{inv.material_name} ({inv.storage_name})" if inv.storage_name else inv.material_name
                buttons.append([InlineKeyboardButton(btn_text, callback_data=f"tn_expinv_{inv.id}")])
            buttons.append([InlineKeyboardButton("Hủy", callback_data="tn_expinv_cancel")])
            await message.reply_text("<b>Vui lòng chọn Hàng Tồn Kho cần xuất đi:</b>", reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    data = {}
    for line in lines[1:]:
        if ":" in line:
            k, v = line.split(":", 1)
            data[k.strip()] = v.strip()

    material_type = data.get("Loại Nguyên Liệu", "").strip()
    storage_name = data.get("Tên Kho", "").strip()
    export_date_str = data.get("Ngày Xuất Kho", "").strip()
    performer_name = data.get("Người Thực Hiện", "").strip()
    notes = data.get("Ghi Chú", "").strip()

    def parse_float_val(val_str):
        if not val_str: return 0.0
        val_str = val_str.replace(".", "").replace(",", ".").replace(" ", "")
        try: return float(val_str)
        except: return 0.0

    export_weight = parse_float_val(data.get("Khối Lượng Xuất (Kg)", data.get("Khối Lượng Xuất", "0")))

    from datetime import datetime
    try:
        export_date = datetime.strptime(export_date_str, "%d/%m/%Y").date() if export_date_str else datetime.now().date()
    except:
        export_date = datetime.now().date()

    from app.db.session import SessionLocal
    from app.models.inventory import Inventory, InventoryExport
    import uuid
    db = SessionLocal()
    try:
        inv = db.query(Inventory).filter(Inventory.material_name == material_type).first()
        if not inv:
            await message.reply_text("⚠️ Không tìm thấy nguyên liệu này.")
            return

        if inv.quantity < export_weight:
            await message.reply_text(f"⚠️ Lỗi: Xuất ({export_weight} kg) quá tồn kho ({inv.quantity} kg).")
            return

        rem_w = inv.quantity - export_weight
        export_rec = InventoryExport(
            id=uuid.uuid4(), export_date=export_date, performer_name=performer_name,
            material_type=material_type, storage_name=storage_name, export_weight=export_weight,
            remaining_weight=rem_w, notes=notes
        )
        db.add(export_rec)
        inv.quantity = rem_w
        db.commit()

        await message.reply_text(
            f"<b>✅ XUẤT KHO THÀNH CÔNG</b>\n\n"
            f"<b>Nguyên Liệu:</b> {material_type}\n"
            f"<b>Khối Lượng Xuất (Kg):</b> {export_weight:,.0f} kg\n"
            f"<b>Tồn Kho Còn Lại (Kg):</b> {rem_w:,.0f} kg\n",
            parse_mode=ParseMode.HTML
        )

        from app.core.config import settings
        config_key_map = {"Củi": "Firewood"}
        config_key = config_key_map.get(material_type, material_type)
        threshold_pct = settings.TienNga.Inventory_Warning_Thresholds.get(config_key)
        
        if threshold_pct and inv.capacity > 0:
            threshold_w = (threshold_pct / 100.0) * inv.capacity
            if rem_w < threshold_w:
                msg = (
                    f"⚠️ <b>CẢNH BÁO TỒN KHO: {material_type.upper()}</b> ⚠️\n\n"
                    f"Kho <b>{inv.storage_name or '—'}</b> hiện tại còn <b>{rem_w:,.0f} kg</b>.\n"
                    f"<i>(Đã xuống dưới ngưỡng an toàn {threshold_pct}% sức chứa: {threshold_w:,.0f} kg)</i>.\n"
                    f"👉 Đề nghị lên kế hoạch nhập thêm nguyên liệu!"
                )
                await client.send_message(chat_id=message.chat.id, text=msg, parse_mode=ParseMode.HTML)

    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^tn_expinv_(.+)$"))
async def _expinv_cb(client, callback_query):
    inv_id = callback_query.matches[0].group(1)
    if inv_id == "cancel":
        await callback_query.message.edit_text("❌ <b>Đã hủy.</b>", parse_mode=ParseMode.HTML)
        return
    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    import uuid
    from datetime import datetime
    db = SessionLocal()
    try:
        inv = db.query(Inventory).filter(Inventory.id == uuid.UUID(inv_id)).first()
        if not inv: return
        today_str = datetime.now().strftime("%d/%m/%Y")
        name = (callback_query.from_user.first_name or '') + ' ' + (callback_query.from_user.last_name or '')
        form = (
            "<b>FORM XUẤT KHO</b>\n\n"
            "<pre>/tien_nga_export_inventory\n"
            f"Loại Nguyên Liệu: {inv.material_name}\n"
            f"Tên Kho: {inv.storage_name or ''}\n"
            f"Ngày Xuất Kho: {today_str}\n"
            f"Người Thực Hiện: {name.strip()}\n"
            "Khối Lượng Xuất (Kg): 0\n"
            "Ghi Chú: </pre>"
        )
        await callback_query.message.reply_text(form, parse_mode=ParseMode.HTML)
        await callback_query.answer()
    finally:
        db.close()



# =====================================================================
# PRODUCT TRANSACTIONS
# =====================================================================

@bot.on_message(filters.command(["tien_nga_product_transaction","tien_nga_giao_dich_san_pham"]) | filters.regex(r"^@\w+\s+/(tien_nga_product_transaction|tien_nga_giao_dich_san_pham)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_PRODUCT)
async def tien_nga_product_transaction_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        from app.db.session import SessionLocal
        from app.models.inventory import Inventory
        db = SessionLocal()
        try:
            invs = db.query(Inventory).all()
            if not invs:
                await message.reply_text("⚠️ Chưa có kho nào. Vui lòng tạo kho trước.")
                return
            buttons = []
            for inv in invs:
                buttons.append([InlineKeyboardButton(f"{inv.material_name} ({inv.storage_name})", callback_data=f"tn_selprodinv_{inv.id}")])
            buttons.append([InlineKeyboardButton("Hủy thao tác", callback_data="tn_selprodinv_cancel")])
            markup = InlineKeyboardMarkup(buttons)
            await message.reply_text("<b>Vui lòng chọn Kho (Sản phẩm) để giao dịch:</b>", reply_markup=markup, parse_mode=ParseMode.HTML)
        except Exception as e:
            await message.reply_text("❌ Lỗi hệ thống.")
        finally:
            db.close()
        return

    # Parse form data
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    storage_id = data.get("Mã Kho", "").strip()
    storage_name = data.get("Tên Kho", "").strip()
    txn_type = data.get("Loại Giao Dịch", "").strip()
    material_name = data.get("Sản Phẩm", "").strip()
    customer_id = data.get("Mã Khách Hàng", "").strip()
    txn_date_str = data.get("Ngày Giao Dịch", "").strip()
    note = data.get("Ghi Chú", "").strip()
    
    # Parse numbers (support both old "Số Lượng" and new "Số Lượng (Kg)" keys)
    qty = parse_float_vn(data.get("Số Lượng (Kg)", data.get("Số Lượng", "0")))
    price = parse_float_vn(data.get("Đơn Giá (VNĐ)", data.get("Đơn Giá", "0")))

    # Auto-calculate total and debt (removed from form)
    total = qty * price if qty > 0 and price > 0 else 0

    if total > 0:
        if txn_type.lower() == "xuất":
            debt = -total
        elif txn_type.lower() == "nhập":
            debt = total
        else:
            debt = 0
    else:
        debt = 0

    if not storage_id or not txn_type.lower() in ['nhập', 'xuất']:
        await message.reply_text("⚠️ <b>Mã Kho</b> và <b>Loại Giao Dịch</b> (Nhập/Xuất) là bắt buộc.", parse_mode=ParseMode.HTML)
        return
        
    if qty <= 0:
        await message.reply_text("⚠️ <b>Số Lượng</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return

    from datetime import datetime
    try:
        if txn_date_str:
            txn_date = datetime.strptime(txn_date_str.split()[0], "%d/%m/%Y").date()
        else:
            txn_date = datetime.now().date()
    except Exception:
        txn_date = datetime.now().date()

    from app.db.session import SessionLocal
    from app.models.inventory import Inventory, ProductTransaction
    from app.models.business import Customers
    import uuid
    
    db = SessionLocal()
    try:
        # Validate Customer if provided
        if customer_id:
            cust = db.query(Customers).filter(Customers.hoursehold_id == customer_id).first()
            if not cust:
                await message.reply_text(f"⚠️ Không tìm thấy Khách Hàng mã <b>{customer_id}</b>.", parse_mode=ParseMode.HTML)
                return
            
            if debt != 0:
                if cust.amount_of_debt is None: cust.amount_of_debt = 0
                if cust.total_debt is None: cust.total_debt = 0
                cust.amount_of_debt += int(debt)
                cust.total_debt += int(debt)

        # Validate Inventory
        inv = db.query(Inventory).filter(Inventory.id == storage_id).first()
        if not inv:
            await message.reply_text(f"⚠️ Không tìm thấy Kho mã <b>{storage_id}</b>.", parse_mode=ParseMode.HTML)
            return

        # Check Inventory Capacity / Balance
        if txn_type.lower() == "nhập":
            inv.quantity += qty
        elif txn_type.lower() == "xuất":
            if inv.quantity < qty:
                await message.reply_text(f"⚠️ Trữ lượng trong kho (<b>{inv.quantity:,.0f}</b>) không đủ để xuất <b>{qty:,.0f}</b>.", parse_mode=ParseMode.HTML)
                return
            inv.quantity -= qty

        # Create Transaction Log
        new_txn = ProductTransaction(
            transaction_date=txn_date,
            customer_id=customer_id if customer_id else None,
            transaction_type=txn_type.capitalize(),
            material_type=material_name or inv.material_name,
            storage_id=inv.id,
            storage_name=storage_name or inv.storage_name,
            quantity=qty,
            unit_price=price,
            total_amount=total,
            debt=debt,
            note=note
        )
        db.add(new_txn)
        db.commit()

        # Reply Success
        await message.reply_text(
            f"✅ <b>GIAO DỊCH {txn_type.upper()} THÀNH CÔNG</b>\n\n"
            f"<b>Kho:</b> {inv.material_name} ({inv.storage_name})\n"
            f"<b>Sản Phẩm:</b> {new_txn.material_type}\n"
            f"<b>Mã Khách Hàng:</b> {new_txn.customer_id or '—'}\n"
            f"<b>Ngày:</b> {txn_date.strftime('%d/%m/%Y')}\n"
            f"<b>Số Lượng {txn_type.capitalize()}:</b> <code>{qty:,.0f} kg</code>\n"
            f"<b>Đơn Giá:</b> <code>{price:,.0f} VNĐ</code>\n"
            f"<b>Thành Tiền:</b> <code>{total:,.0f} VNĐ</code>\n"
            f"<b>Công Nợ:</b> <code>{debt:,.0f} VNĐ</code>\n"
            f"<b>Ghi Chú:</b> {note or '—'}\n"
            f"<b>Tồn Kho Hiện Tại:</b> <code>{inv.quantity:,.0f} kg</code>",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        db.rollback()
        LogError(f"Error handling product transaction: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Lỗi hệ thống khi cập nhật cơ sở dữ liệu.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^tn_selprodinv_(.+)$"))
async def tien_nga_select_prod_inv_callback(client, callback_query):
    action = callback_query.matches[0].group(1)
    if action == "cancel":
        await callback_query.message.edit_text("🚫 Giao dịch đã bị hủy.")
        await callback_query.answer()
        return

    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    import uuid
    db = SessionLocal()
    try:
        inv = db.query(Inventory).filter(Inventory.id == uuid.UUID(action)).first()
        if not inv:
            await callback_query.answer("⚠️ Không tìm thấy Kho", show_alert=True)
            return

        buttons = [
            [InlineKeyboardButton("Nhập Kho", callback_data=f"tn_selprodttype_{action}_Nhập")],
            [InlineKeyboardButton("Xuất Kho", callback_data=f"tn_selprodttype_{action}_Xuất")],
            [InlineKeyboardButton("Hủy Thao Tác", callback_data="tn_selprodttype_cancel")]
        ]
        markup = InlineKeyboardMarkup(buttons)
        await callback_query.message.edit_text(
            f"<b>Kho:</b> {inv.material_name} ({inv.storage_name})\n"
            f"<b>Tồn hiện tại:</b> {inv.quantity:,.0f}\n\n"
            f"Vui lòng chọn <b>Loại Giao Dịch</b>:",
            reply_markup=markup, parse_mode=ParseMode.HTML
        )
        await callback_query.answer()
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^tn_selprodttype_(.+)$"))
async def tien_nga_select_prod_ttype_callback(client, callback_query):
    action = callback_query.matches[0].group(1)
    if action == "cancel":
        await callback_query.message.edit_text("🚫 Giao dịch đã bị hủy.")
        await callback_query.answer()
        return

    parts = action.split("_", 1)
    if len(parts) < 2: return
    inv_id, txn_type = parts[0], parts[1]

    from app.db.session import SessionLocal
    from app.models.inventory import Inventory
    from datetime import datetime
    import uuid
    db = SessionLocal()
    try:
        inv = db.query(Inventory).filter(Inventory.id == uuid.UUID(inv_id)).first()
        if not inv: return
        today_str = datetime.now().strftime("%d/%m/%Y")
        
        form = (
            f"<b>FORM GIAO DỊCH {txn_type.upper()} SẢN PHẨM</b>\n\n"
            "<pre>/tien_nga_giao_dich_san_pham\n"
            f"Mã Kho: {inv.id}\n"
            f"Tên Kho: {inv.storage_name or ''}\n"
            f"Loại Giao Dịch: {txn_type}\n"
            f"Sản Phẩm: {inv.material_name}\n"
            "Mã Khách Hàng: \n"
            f"Ngày Giao Dịch: {today_str}\n"
            "Số Lượng (Kg): 0\n"
            "Đơn Giá (VNĐ): 0\n"
            "Ghi Chú: </pre>\n\n"
            "<i>(Mã Khách Hàng có thể để trống nếu không xác định)</i>"
        )
        await callback_query.message.reply_text(form, parse_mode=ParseMode.HTML)
        await callback_query.answer()
    finally:
        db.close()


# Pending shareholder creation data (in-memory store)
_pending_shareholder_data = {}

@bot.on_message(filters.command(["tien_nga_create_shareholder", "tien_nga_tao_co_dong"]) | filters.regex(r"^@\w+\s+/(tien_nga_create_shareholder|tien_nga_tao_co_dong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_FINANCE, CustomTitle.MAIN_SHAREHOLDER)
async def tien_nga_create_shareholder_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    
    if len(lines) < 2:
        from app.models.business import Investment
        db = SessionLocal()
        try:
            investments = db.query(Investment).filter(Investment.status == "ACTIVE").all()
            if not investments:
                await message.reply_text("⚠️ Không có Quỹ Đầu Tư nào đang hoạt động.", parse_mode=ParseMode.HTML)
                return
            
            buttons = []
            for inv in investments:
                buttons.append([InlineKeyboardButton(inv.name, callback_data=f"tn_shrinv_{inv.id}")])
                
            buttons.append([InlineKeyboardButton("Hủy", callback_data="cancel_action")])
            reply_markup = InlineKeyboardMarkup(buttons)
            
            await message.reply_text(
                "<b>Vui lòng chọn Quỹ Đầu Tư để thêm Cổ Đông:</b>",
                reply_markup=reply_markup,
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            LogError(f"Error fetching investments: {e}", LogType.SYSTEM_STATUS)
            await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    # Parse form data
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    inv_id = data.get("Mã Quỹ Đầu Tư", "").strip()
    shareholder_code = data.get("Mã Cổ Đông", "").strip()
    fullname = data.get("Tên Cổ Đông", "").strip()
    username = data.get("Username TG", "").strip()
    telegram_group = data.get("Nhóm Telegram", "").strip()
    notes = data.get("Ghi Chú", "").strip()
    amount = parse_float_vn(data.get("Số Tiền Đầu Tư", "0"))
    start_date_str = data.get("Ngày Bắt Đầu", "").strip()

    if amount <= 0:
        await message.reply_text("⚠️ <b>Số Tiền Đầu Tư</b> phải là số dương.", parse_mode=ParseMode.HTML)
        return
        
    if not inv_id or not shareholder_code or not fullname:
        await message.reply_text("⚠️ Vui lòng điền đủ Mã Quỹ Đầu Tư, Mã Cổ Đông và Tên Cổ Đông.", parse_mode=ParseMode.HTML)
        return

    start_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%d/%m/%Y").date()
        except ValueError:
            await message.reply_text("⚠️ <b>Ngày Bắt Đầu</b> sai định dạng (dd/mm/yyyy).", parse_mode=ParseMode.HTML)
            return

    from app.models.business import Shareholder, Investment

    db = SessionLocal()
    try:
        investment = db.query(Investment).filter(Investment.id == inv_id).first()
        if not investment:
            await message.reply_text("⚠️ Quỹ Đầu Tư không tồn tại.", parse_mode=ParseMode.HTML)
            return

        existing = db.query(Shareholder).filter(Shareholder.shareholder_code == shareholder_code).first()
        is_existing = existing is not None

        # Determine action label
        if is_existing:
            action_label = "GÓP VỐN THÊM"
            old_amount = existing.investment_amount or 0
            new_total = old_amount + amount
        else:
            action_label = "TẠO CỔ ĐÔNG MỚI"
            old_amount = 0
            new_total = amount

        today_str = datetime.now().strftime("%d/%m/%Y")
        start_date_display = start_date.strftime("%d/%m/%Y") if start_date else today_str

        # Generate a unique pending key
        import uuid as _uuid
        pending_key = str(_uuid.uuid4())[:8]

        # Store pending data
        _pending_shareholder_data[pending_key] = {
            "inv_id": inv_id,
            "inv_name": investment.name,
            "inv_code": investment.investment_code or "N/A",
            "shareholder_code": shareholder_code,
            "fullname": fullname,
            "username": username,
            "telegram_group": telegram_group,
            "notes": notes,
            "amount": amount,
            "start_date": start_date.isoformat() if start_date else None,
            "is_existing": is_existing,
            "old_amount": old_amount,
            "new_total": new_total,
            "user_id": message.from_user.id,
        }

        # Build preview message
        preview = (
            f"📋 <b>XÁC NHẬN {action_label}</b>\n"
            f"━━━━━━━━━━━━━━━━━━\n\n"
            f"<b>📌 THÔNG TIN CỔ ĐÔNG</b>\n"
            f"<b>Quỹ Đầu Tư:</b> {investment.name} (<code>{investment.investment_code or 'N/A'}</code>)\n"
            f"<b>Mã Cổ Đông:</b> <code>{shareholder_code}</code>\n"
            f"<b>Tên Cổ Đông:</b> {fullname}\n"
        )
        if username:
            preview += f"<b>Username TG:</b> @{username}\n"
        if telegram_group:
            preview += f"<b>Nhóm Telegram:</b> {telegram_group}\n"
        preview += f"<b>Ngày Bắt Đầu:</b> {start_date_display}\n"
        if notes:
            preview += f"<b>Ghi Chú:</b> {notes}\n"

        if is_existing:
            preview += (
                f"\n<b>Vốn góp hiện tại:</b> <code>{fmt_vn(old_amount)}</code>\n"
                f"<b>Số tiền góp thêm:</b> <code>{fmt_vn(amount)}</code>\n"
                f"<b>Tổng vốn sau góp:</b> <code>{fmt_vn(new_total)}</code>\n"
            )
        else:
            preview += f"<b>Số Tiền Đầu Tư:</b> <code>{fmt_vn(amount)}</code>\n"

        preview += (
            f"\n━━━━━━━━━━━━━━━━━━\n"
            f"<b>💰 PHIẾU THU - DAILY PAYMENT</b>\n"
            f"<b>Loại:</b> THU\n"
            f"<b>Người Thực Hiện:</b> {shareholder_code}\n"
            f"<b>Mục Đích:</b> Góp vốn Quỹ {investment.name}\n"
            f"<b>Số Tiền:</b> <code>{fmt_vn(amount)}</code>\n"
            f"<b>Ngày:</b> {start_date_display}\n"
            f"<b>Trạng Thái:</b> APPROVED\n"
            f"━━━━━━━━━━━━━━━━━━\n\n"
            f"<i>Vui lòng kiểm tra thông tin và nhấn nút bên dưới để xác nhận.</i>"
        )

        buttons = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Xác Nhận", callback_data=f"shcf_y_{pending_key}"),
                InlineKeyboardButton("❌ Hủy", callback_data=f"shcf_n_{pending_key}"),
            ]
        ])

        await message.reply_text(preview, reply_markup=buttons, parse_mode=ParseMode.HTML)

    except Exception as e:
        LogError(f"Error preparing shareholder preview: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# --- Confirm / Cancel shareholder creation ---
@bot.on_callback_query(filters.regex(r"^shcf_(y|n)_(.+)$"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
async def tien_nga_confirm_shareholder_callback(client, callback_query):
    action = callback_query.matches[0].group(1)  # "y" or "n"
    pending_key = callback_query.matches[0].group(2)

    pending = _pending_shareholder_data.pop(pending_key, None)
    if not pending:
        await callback_query.answer("⚠️ Dữ liệu đã hết hạn hoặc đã được xử lý.", show_alert=True)
        try:
            await callback_query.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
        return

    # Cancel
    if action == "n":
        await callback_query.message.edit_text(
            "❌ <b>ĐÃ HỦY</b>\n\n<i>Thao tác tạo/góp vốn cổ đông đã bị hủy.</i>",
            parse_mode=ParseMode.HTML
        )
        await callback_query.answer("Đã hủy thao tác.")
        return

    # Confirm → save both Shareholder + DailyPayment
    from app.models.business import Shareholder, Investment, DailyPayment
    import uuid as _uuid

    db = SessionLocal()
    try:
        inv_id = pending["inv_id"]
        investment = db.query(Investment).filter(Investment.id == inv_id).first()
        if not investment:
            await callback_query.answer("⚠️ Quỹ Đầu Tư không tồn tại.", show_alert=True)
            return

        shareholder_code = pending["shareholder_code"]
        fullname = pending["fullname"]
        amount = pending["amount"]
        notes = pending["notes"]
        username = pending["username"]
        telegram_group = pending["telegram_group"]
        is_existing = pending["is_existing"]

        start_date = None
        if pending["start_date"]:
            start_date = datetime.strptime(pending["start_date"], "%Y-%m-%d").date()
        payment_day = start_date or datetime.now().date()

        existing = db.query(Shareholder).filter(Shareholder.shareholder_code == shareholder_code).first()

        if is_existing and existing:
            # Cộng dồn vốn góp
            old_amount = existing.investment_amount or 0
            existing.investment_amount = old_amount + amount
            if notes:
                existing.notes = (existing.notes or "") + f"\n[Góp thêm] {notes}" if existing.notes else notes
            action_label = "GÓP VỐN THÊM"
        else:
            # Tạo mới cổ đông
            new_shareholder = Shareholder(
                id=_uuid.uuid4(),
                shareholder_code=shareholder_code,
                fullname=fullname,
                investment_id=investment.id,
                investment_amount=amount,
                start_date=start_date,
                username=username if username else None,
                telegram_group=telegram_group if telegram_group else None,
                notes=notes if notes else None
            )
            db.add(new_shareholder)
            action_label = "TẠO CỔ ĐÔNG"

        # Cập nhật Quỹ ban đầu và Tổng thu
        if investment.initial_capital is None:
            investment.initial_capital = 0
        if investment.total_income is None:
            investment.total_income = 0
        investment.initial_capital += amount
        investment.total_income += amount

        # Tạo DailyPayment record cho khoản góp vốn
        purpose_text = f"Góp vốn Quỹ {investment.name}"
        if is_existing:
            purpose_text = f"Góp vốn thêm Quỹ {investment.name}"

        new_payment = DailyPayment(
            id=_uuid.uuid4(),
            investment_id=investment.id,
            requester=fullname,
            executor=shareholder_code,
            receiver=investment.name,
            payment_type="thu",
            purpose=purpose_text,
            reason=f"Cổ đông {shareholder_code} góp vốn",
            amount=amount,
            day=payment_day,
            notes=notes if notes else None,
            status="APPROVED",
        )
        db.add(new_payment)

        db.commit()

        LogInfo(f"[TienNga] {action_label} '{shareholder_code}' amount {fmt_vn(amount)} for Investment '{investment.name}' + DailyPayment created", LogType.SYSTEM_STATUS)

        # Build success message
        if is_existing and existing:
            success_msg = (
                f"✅ <b>{action_label} THÀNH CÔNG</b>\n"
                f"━━━━━━━━━━━━━━━━━━\n\n"
                f"<b>📌 THÔNG TIN CỔ ĐÔNG</b>\n"
                f"<b>Quỹ Đầu Tư:</b> {investment.name}\n"
                f"<b>Mã Cổ Đông:</b> <code>{shareholder_code}</code>\n"
                f"<b>Tên Cổ Đông:</b> {existing.fullname}\n"
                f"<b>Số Tiền Góp Thêm:</b> <code>{fmt_vn(amount)}</code>\n"
                f"<b>Tổng Tiền Đã Góp:</b> <code>{fmt_vn(existing.investment_amount)}</code>\n\n"
            )
        else:
            success_msg = (
                f"✅ <b>{action_label} THÀNH CÔNG</b>\n"
                f"━━━━━━━━━━━━━━━━━━\n\n"
                f"<b>📌 THÔNG TIN CỔ ĐÔNG</b>\n"
                f"<b>Quỹ Đầu Tư:</b> {investment.name}\n"
                f"<b>Mã Cổ Đông:</b> <code>{shareholder_code}</code>\n"
                f"<b>Tên Cổ Đông:</b> {fullname}\n"
                f"<b>Số Tiền Đầu Tư:</b> <code>{fmt_vn(amount)}</code>\n\n"
            )

        success_msg += (
            f"<b>💰 PHIẾU THU ĐÃ GHI NHẬN</b>\n"
            f"<b>Loại:</b> THU\n"
            f"<b>Người Thực Hiện:</b> {shareholder_code}\n"
            f"<b>Mục Đích:</b> {purpose_text}\n"
            f"<b>Số Tiền:</b> <code>{fmt_vn(amount)}</code>\n"
            f"<b>Ngày:</b> {payment_day.strftime('%d/%m/%Y')}\n"
            f"<b>Trạng Thái:</b> APPROVED\n\n"
            f"<i>Quỹ ban đầu của '{investment.name}' đã tăng thêm {fmt_vn(amount)}.</i>"
        )

        await callback_query.message.edit_text(success_msg, parse_mode=ParseMode.HTML)
        await callback_query.answer("✅ Đã xác nhận thành công!")

        # Gửi thông báo xuống nhóm member của cổ đông
        if telegram_group:
            try:
                from app.models.telegram import TelegramProjectMember
                member_group = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.group_name == telegram_group,
                    TelegramProjectMember.role == "member"
                ).first()

                if member_group and member_group.chat_id:
                    notify_msg = (
                        f"<b>THÔNG BÁO GÓP VỐN CỔ ĐÔNG</b>\n"
                        f"━━━━━━━━━━━━━━━━━━\n\n"
                        f"<b>Quỹ Đầu Tư:</b> {investment.name}\n"
                        f"<b>Mã Cổ Đông:</b> <code>{shareholder_code}</code>\n"
                        f"<b>Tên Cổ Đông:</b> {fullname}\n"
                        f"<b>Loại:</b> {action_label}\n"
                        f"<b>Số Tiền:</b> <code>{fmt_vn(amount)}</code>\n"
                        f"<b>Ngày:</b> {payment_day.strftime('%d/%m/%Y')}\n"
                    )
                    if is_existing and existing:
                        notify_msg += f"<b>Tổng Vốn Góp:</b> <code>{fmt_vn(existing.investment_amount)}</code>\n"
                    notify_msg += (
                        f"\n<b>Trạng Thái:</b> ĐÃ XÁC NHẬN ✅\n\n"
                        f"<i>Phiếu thu đã được ghi nhận vào hệ thống.</i>"
                    )
                    await client.send_message(
                        chat_id=int(member_group.chat_id),
                        text=notify_msg,
                        parse_mode=ParseMode.HTML
                    )
                    LogInfo(f"[TienNga] Sent shareholder notification to member group '{telegram_group}' (chat_id: {member_group.chat_id})", LogType.SYSTEM_STATUS)
            except Exception as notify_err:
                LogError(f"[TienNga] Failed to send notification to member group '{telegram_group}': {notify_err}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error confirming shareholder creation: {e}\n{traceback.format_exc()}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi lưu dữ liệu.", show_alert=True)
        await callback_query.message.edit_text(
            f"❌ <b>LỖI</b>\n\nCó lỗi xảy ra khi lưu vào database: {e}",
            parse_mode=ParseMode.HTML
        )
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^tn_shrinv_(.+)$"))
async def tien_nga_select_investment_callback(client, callback_query):
    inv_id = callback_query.matches[0].group(1)
    from app.db.session import SessionLocal
    from app.models.business import Investment
    
    db = SessionLocal()
    try:
        inv = db.query(Investment).filter(Investment.id == inv_id).first()
        if not inv:
            await callback_query.answer("⚠️ Không tìm thấy Quỹ Đầu Tư", show_alert=True)
            return
            
        import datetime
        today_str = datetime.datetime.now().strftime("%d/%m/%Y")
        form_template = f"""<b>FORM TẠO CỔ ĐÔNG MỚI</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/tien_nga_tao_co_dong
Mã Quỹ Đầu Tư: {inv_id}
Mã Cổ Đông: 
Tên Cổ Đông: 
Số Tiền Đầu Tư: 
Ngày Bắt Đầu: {today_str}
Username TG: 
Nhóm Telegram: 
Ghi Chú: </pre>

<i>(*Ghi chú: Đang thêm cổ đông cho Quỹ: <b>{inv.name}</b>)</i>"""
        await callback_query.message.reply_text(form_template, parse_mode=ParseMode.HTML)
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in select investment point: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống", show_alert=True)
    finally:
        db.close()

# --- CHIA CỔ TỨC ---
@bot.on_message(filters.command(["tien_nga_dividend_distribution","tien_nga_chia_co_tuc"]) | filters.regex(r"^@\w+\s+/(tien_nga_dividend_distribution|tien_nga_chia_co_tuc)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_FINANCE, CustomTitle.MAIN_SHAREHOLDER)
async def tien_nga_chia_co_tuc_handler(client, message: Message) -> None:
    from app.db.session import SessionLocal
    from app.models.business import Investment
    db = SessionLocal()
    try:
        investments = db.query(Investment).filter(Investment.status == "ACTIVE").all()
        if not investments:
            await message.reply_text("⚠️ Không có Quỹ Đầu Tư nào đang hoạt động.", parse_mode=ParseMode.HTML)
            return
        
        buttons = []
        for inv in investments:
            label = f"[{inv.investment_code}] {inv.name}" if inv.investment_code else inv.name
            buttons.append([InlineKeyboardButton(label, callback_data=f"tn_divid_{inv.id}")])
        buttons.append([InlineKeyboardButton("Hủy", callback_data="cancel_action")])
        
        await message.reply_text(
            "<b>CHIA CỔ TỨC</b>\n\nVui lòng chọn Quỹ Đầu Tư để chia cổ tức:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error in chia co tuc: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# Pending dividend distribution data (in-memory store)
_pending_dividend_data = {}

@bot.on_callback_query(filters.regex(r"^tn_divid_(.+)$"))
async def tien_nga_dividend_callback(client, callback_query):
    inv_id = callback_query.matches[0].group(1)
    from app.models.business import Investment, Shareholder
    
    db = SessionLocal()
    try:
        inv = db.query(Investment).filter(Investment.id == inv_id).first()
        if not inv:
            await callback_query.answer("⚠️ Không tìm thấy Quỹ Đầu Tư", show_alert=True)
            return
        
        shareholders = db.query(Shareholder).filter(Shareholder.investment_id == inv.id).order_by(Shareholder.shareholder_code).all()
        if not shareholders:
            await callback_query.answer("⚠️ Chưa có cổ đông nào trong quỹ này.", show_alert=True)
            return
        
        total_income = inv.total_income or 0
        total_expense = inv.total_expense or 0
        
        total_shareholder_amount = sum(sh.investment_amount or 0 for sh in shareholders)
        
        # Tiền lời thực tế = Tổng Thu - Tổng Chi - Tổng Vốn Góp
        # (vì tổng thu đã bao gồm tiền góp vốn)
        distributable_profit = total_income - total_expense - total_shareholder_amount

        # Calculate dividend for each shareholder
        dividend_details = []
        for sh in shareholders:
            sh_amount = sh.investment_amount or 0
            share_pct = (sh_amount / total_shareholder_amount * 100) if total_shareholder_amount > 0 else 0
            dividend = distributable_profit * (sh_amount / total_shareholder_amount) if total_shareholder_amount > 0 and distributable_profit > 0 else 0
            dividend_details.append({
                "shareholder_code": sh.shareholder_code,
                "fullname": sh.fullname,
                "investment_amount": sh_amount,
                "share_pct": share_pct,
                "dividend": dividend,
            })

        # Store pending data
        import uuid as _uuid
        pending_key = str(_uuid.uuid4())[:8]
        today_str = datetime.now().strftime("%d/%m/%Y")

        _pending_dividend_data[pending_key] = {
            "inv_id": str(inv.id),
            "inv_name": inv.name,
            "inv_code": inv.investment_code or "N/A",
            "total_income": total_income,
            "total_expense": total_expense,
            "distributable_profit": distributable_profit,
            "total_shareholder_amount": total_shareholder_amount,
            "dividend_details": dividend_details,
            "user_id": callback_query.from_user.id,
        }

        # Build preview message
        msg = (
            f"<b>XÁC NHẬN CHIA CỔ TỨC</b>\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"<b>Quỹ:</b> {inv.name}\n"
            f"<b>Mã Quỹ:</b> <code>{inv.investment_code or 'N/A'}</code>\n\n"
            f"<b>Tổng Thu:</b> <code>{fmt_vn(total_income)}</code>\n"
            f"<b>Tổng Chi:</b> <code>{fmt_vn(total_expense)}</code>\n"
            f"<b>Tổng Vốn Góp:</b> <code>{fmt_vn(total_shareholder_amount)}</code>\n"
            f"<b>Lợi Nhuận Chia:</b> <code>{fmt_vn(distributable_profit)}</code>\n"
            f"<i>(= Tổng Thu - Tổng Chi - Tổng Vốn Góp)</i>\n\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"<b>CHI TIẾT CHIA CỔ TỨC</b>\n\n"
        )

        
        total_dividend = 0
        for idx, d in enumerate(dividend_details, 1):
            msg += (
                f"<b>{idx}. {d['fullname']}</b> (<code>{d['shareholder_code']}</code>)\n"
                f"   Vốn góp: <code>{fmt_vn(d['investment_amount'])}</code>\n"
                f"   Cổ phần: <b>{d['share_pct']:.1f}%</b>\n"
                f"   Cổ tức nhận được: <code>{fmt_vn(d['dividend'])}</code>\n\n"
            )
            total_dividend += d['dividend']
        
        msg += (
            f"━━━━━━━━━━━━━━━━━━\n"
            f"<b>PHIẾU CHI - DAILY PAYMENT ({len(dividend_details)} phiếu)</b>\n\n"
        )

        for idx, d in enumerate(dividend_details, 1):
            msg += (
                f"<b>{idx}. Phiếu Chi</b>\n"
                f"   <b>Loại:</b> CHI\n"
                f"   <b>Người Thực Hiện:</b> {d['shareholder_code']}\n"
                f"   <b>Người Nhận:</b> {d['fullname']}\n"
                f"   <b>Mục Đích:</b> Nhận cổ tức Quỹ {inv.name}\n"
                f"   <b>Số Tiền:</b> <code>{fmt_vn(d['dividend'])}</code>\n"
                f"   <b>Ngày:</b> {today_str}\n"
                f"   <b>Trạng Thái:</b> APPROVED\n\n"
            )

        msg += (
            f"<b>Tổng Số Tiền Chi:</b> <code>{fmt_vn(total_dividend)}</code>\n"
            f"━━━━━━━━━━━━━━━━━━\n\n"
        )


        if distributable_profit <= 0:
            msg += "<i>⚠️ Quỹ chưa có lợi nhuận hoặc đang lỗ, không nên chia cổ tức.</i>\n\n"

        msg += "<i>Vui lòng kiểm tra thông tin và nhấn nút bên dưới để xác nhận.</i>"

        buttons = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Xác Nhận Chia Cổ Tức", callback_data=f"divcf_y_{pending_key}"),
                InlineKeyboardButton("❌ Hủy", callback_data=f"divcf_n_{pending_key}"),
            ]
        ])

        await callback_query.message.reply_text(msg, reply_markup=buttons, parse_mode=ParseMode.HTML)
        await callback_query.answer()
        
    except Exception as e:
        LogError(f"Error in dividend callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống", show_alert=True)
    finally:
        db.close()


# --- Confirm / Cancel dividend distribution ---
@bot.on_callback_query(filters.regex(r"^divcf_(y|n)_(.+)$"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
async def tien_nga_confirm_dividend_callback(client, callback_query):
    action = callback_query.matches[0].group(1)  # "y" or "n"
    pending_key = callback_query.matches[0].group(2)

    pending = _pending_dividend_data.pop(pending_key, None)
    if not pending:
        await callback_query.answer("⚠️ Dữ liệu đã hết hạn hoặc đã được xử lý.", show_alert=True)
        try:
            await callback_query.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
        return

    # Cancel
    if action == "n":
        await callback_query.message.edit_text(
            "❌ <b>ĐÃ HỦY</b>\n\n<i>Thao tác chia cổ tức đã bị hủy.</i>",
            parse_mode=ParseMode.HTML
        )
        await callback_query.answer("Đã hủy thao tác.")
        return

    # Confirm → create DailyPayment for each shareholder
    from app.models.business import Investment, DailyPayment
    import uuid as _uuid

    db = SessionLocal()
    try:
        inv_id = pending["inv_id"]
        investment = db.query(Investment).filter(Investment.id == inv_id).first()
        if not investment:
            await callback_query.answer("⚠️ Quỹ Đầu Tư không tồn tại.", show_alert=True)
            return

        dividend_details = pending["dividend_details"]
        distributable_profit = pending["distributable_profit"]
        today = datetime.now().date()

        total_dividend_paid = 0

        for d in dividend_details:
            dividend_amount = d["dividend"]
            if dividend_amount <= 0:
                continue

            new_payment = DailyPayment(
                id=_uuid.uuid4(),
                investment_id=investment.id,
                requester=investment.name,
                executor=d["shareholder_code"],
                receiver=d["fullname"],
                payment_type="chi",
                purpose=f"Nhận cổ tức Quỹ {investment.name}",
                reason=f"Cổ đông {d['shareholder_code']} nhận cổ tức ({d['share_pct']:.1f}%)",
                amount=dividend_amount,
                day=today,
                notes=f"Vốn góp: {fmt_vn(d['investment_amount'])} | Cổ phần: {d['share_pct']:.1f}%",
                status="APPROVED",
            )
            db.add(new_payment)
            total_dividend_paid += dividend_amount

        # Cập nhật tổng chi của Investment
        if investment.total_expense is None:
            investment.total_expense = 0
        investment.total_expense += total_dividend_paid
        investment.profit = (investment.total_income or 0) - investment.total_expense

        db.commit()

        LogInfo(f"[TienNga] Dividend distribution confirmed for '{investment.name}', total paid: {fmt_vn(total_dividend_paid)} by @{callback_query.from_user.username}", LogType.SYSTEM_STATUS)

        # Build success message
        success_msg = (
            f"✅ <b>CHIA CỔ TỨC THÀNH CÔNG</b>\n"
            f"━━━━━━━━━━━━━━━━━━\n\n"
            f"<b>Quỹ:</b> {investment.name} (<code>{investment.investment_code or 'N/A'}</code>)\n"
            f"<b>Ngày:</b> {today.strftime('%d/%m/%Y')}\n\n"
        )

        for idx, d in enumerate(dividend_details, 1):
            if d["dividend"] <= 0:
                continue
            success_msg += (
                f"<b>{idx}. {d['fullname']}</b> (<code>{d['shareholder_code']}</code>)\n"
                f"   Cổ tức: <code>{fmt_vn(d['dividend'])}</code>\n"
            )

        success_msg += (
            f"\n━━━━━━━━━━━━━━━━━━\n"
            f"<b>TỔNG CHI CỔ TỨC:</b> <code>{fmt_vn(total_dividend_paid)}</code>\n"
            f"<b>Số phiếu chi đã tạo:</b> {sum(1 for d in dividend_details if d['dividend'] > 0)}\n"
            f"<b>Trạng Thái:</b> APPROVED\n\n"
            f"<i>Tất cả phiếu chi đã được ghi nhận vào hệ thống.</i>"
        )

        await callback_query.message.edit_text(success_msg, parse_mode=ParseMode.HTML)
        await callback_query.answer("✅ Chia cổ tức thành công!")

        # Gửi thông báo xuống nhóm member của từng cổ đông
        try:
            from app.models.business import Shareholder
            from app.models.telegram import TelegramProjectMember

            for d in dividend_details:
                if d["dividend"] <= 0:
                    continue
                sh = db.query(Shareholder).filter(
                    Shareholder.shareholder_code == d["shareholder_code"],
                    Shareholder.investment_id == investment.id
                ).first()
                if not sh or not sh.telegram_group:
                    continue

                member_group = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.group_name == sh.telegram_group,
                    TelegramProjectMember.role == "member"
                ).first()
                if not member_group or not member_group.chat_id:
                    continue

                notify_msg = (
                    f"<b>THÔNG BÁO CHIA CỔ TỨC</b>\n"
                    f"━━━━━━━━━━━━━━━━━━\n\n"
                    f"<b>Quỹ Đầu Tư:</b> {investment.name}\n"
                    f"<b>Mã Cổ Đông:</b> <code>{d['shareholder_code']}</code>\n"
                    f"<b>Tên Cổ Đông:</b> {d['fullname']}\n"
                    f"<b>Cổ Phần:</b> {d['share_pct']:.1f}%\n"
                    f"<b>Cổ Tức Nhận Được:</b> <code>{fmt_vn(d['dividend'])}</code>\n"
                    f"<b>Ngày:</b> {today.strftime('%d/%m/%Y')}\n\n"
                    f"<b>Trạng Thái:</b> ĐÃ XÁC NHẬN ✅\n\n"
                    f"<i>Phiếu chi đã được ghi nhận vào hệ thống.</i>"
                )
                try:
                    await client.send_message(
                        chat_id=int(member_group.chat_id),
                        text=notify_msg,
                        parse_mode=ParseMode.HTML
                    )
                    LogInfo(f"[TienNga] Sent dividend notification to '{sh.telegram_group}' for {d['shareholder_code']}", LogType.SYSTEM_STATUS)
                except Exception as send_err:
                    LogError(f"[TienNga] Failed to send dividend notification to '{sh.telegram_group}': {send_err}", LogType.SYSTEM_STATUS)
        except Exception as notify_err:
            LogError(f"[TienNga] Error sending dividend notifications: {notify_err}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error confirming dividend distribution: {e}\n{traceback.format_exc()}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi lưu dữ liệu.", show_alert=True)
        await callback_query.message.edit_text(
            f"❌ <b>LỖI</b>\n\nCó lỗi xảy ra khi lưu vào database: {e}",
            parse_mode=ParseMode.HTML
        )
    finally:
        db.close()



# --- THANH TOÁN QUỸ CỔ ĐÔNG ---
@bot.on_message(filters.command(["tien_nga_payment_shareholder","tien_nga_thanh_toan_co_dong"]) | filters.regex(r"^@\w+\s+/tien_nga_payment_shareholder|tien_nga_thanh_toan_co_dong\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_FINANCE, CustomTitle.MAIN_SHAREHOLDER)
async def tien_nga_payment_shareholder_handler(client, message: Message) -> None:
    from app.db.session import SessionLocal
    from app.models.business import Investment
    db = SessionLocal()
    try:
        investments = db.query(Investment).filter(Investment.status == "ACTIVE").all()
        if not investments:
            await message.reply_text("⚠️ Không có Quỹ Đầu Tư nào đang hoạt động.", parse_mode=ParseMode.HTML)
            return
        
        buttons = []
        for inv in investments:
            label = f"[{inv.investment_code}] {inv.name}" if inv.investment_code else inv.name
            buttons.append([InlineKeyboardButton(label, callback_data=f"tn_paysh_{inv.id}")])
        buttons.append([InlineKeyboardButton("Hủy", callback_data="cancel_action")])
        
        await message.reply_text(
            "<b>THANH TOÁN QUỸ CỔ ĐÔNG</b>\n\nVui lòng chọn Quỹ Đầu Tư để thanh toán:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error in thanh toan quy co dong: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# Pending payment shareholder data (in-memory store)
_pending_payment_sh_data = {}

@bot.on_callback_query(filters.regex(r"^tn_paysh_(.+)$"))
async def tien_nga_payment_shareholder_callback(client, callback_query):
    inv_id = callback_query.matches[0].group(1)
    from app.models.business import Investment, Shareholder
    
    db = SessionLocal()
    try:
        inv = db.query(Investment).filter(Investment.id == inv_id).first()
        if not inv:
            await callback_query.answer("⚠️ Không tìm thấy Quỹ Đầu Tư", show_alert=True)
            return
        
        shareholders = db.query(Shareholder).filter(Shareholder.investment_id == inv.id).order_by(Shareholder.shareholder_code).all()
        if not shareholders:
            await callback_query.answer("⚠️ Chưa có cổ đông nào trong quỹ này.", show_alert=True)
            return
        
        total_income = inv.total_income or 0
        total_expense = inv.total_expense or 0
        initial_capital = inv.initial_capital or 0
        profit = total_income - total_expense
        
        total_shareholder_amount = sum(sh.investment_amount or 0 for sh in shareholders)

        # Xác định quỹ thanh toán
        if profit > initial_capital:
            payout_pool = initial_capital
            payout_label = "Quỹ Ban Đầu (Lợi nhuận >= Vốn)"
        else:
            payout_pool = profit
            payout_label = "Lợi Nhuận (Lợi nhuận < Vốn)"

        # Calculate payment for each shareholder
        payment_details = []
        for sh in shareholders:
            sh_amount = sh.investment_amount or 0
            share_pct = (sh_amount / total_shareholder_amount * 100) if total_shareholder_amount > 0 else 0
            payment = payout_pool * (sh_amount / total_shareholder_amount) if total_shareholder_amount > 0 and payout_pool > 0 else 0
            payment_details.append({
                "shareholder_code": sh.shareholder_code,
                "fullname": sh.fullname,
                "investment_amount": sh_amount,
                "share_pct": share_pct,
                "payment": payment,
            })

        # Store pending data
        import uuid as _uuid
        pending_key = str(_uuid.uuid4())[:8]
        today_str = datetime.now().strftime("%d/%m/%Y")

        _pending_payment_sh_data[pending_key] = {
            "inv_id": str(inv.id),
            "inv_name": inv.name,
            "inv_code": inv.investment_code or "N/A",
            "total_income": total_income,
            "total_expense": total_expense,
            "initial_capital": initial_capital,
            "profit": profit,
            "payout_pool": payout_pool,
            "payout_label": payout_label,
            "total_shareholder_amount": total_shareholder_amount,
            "payment_details": payment_details,
            "user_id": callback_query.from_user.id,
        }

        # Build preview message
        msg = (
            f"<b>XÁC NHẬN THANH TOÁN QUỸ CỔ ĐÔNG</b>\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"<b>Quỹ:</b> {inv.name}\n"
            f"<b>Mã Quỹ:</b> <code>{inv.investment_code or 'N/A'}</code>\n\n"
            f"<b>Vốn Ban Đầu:</b> <code>{fmt_vn(initial_capital)}</code>\n"
            f"<b>Tổng Thu:</b> <code>{fmt_vn(total_income)}</code>\n"
            f"<b>Tổng Chi:</b> <code>{fmt_vn(total_expense)}</code>\n"
            f"<b>Lợi Nhuận:</b> <code>{fmt_vn(profit)}</code>\n\n"
            f"<b>Căn cứ thanh toán:</b> {payout_label}\n"
            f"<b>Quỹ Thanh Toán:</b> <code>{fmt_vn(payout_pool)}</code>\n"
            f"<b>Tổng Vốn Góp:</b> <code>{fmt_vn(total_shareholder_amount)}</code>\n\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"<b>CHI TIẾT THANH TOÁN CHO CỔ ĐÔNG</b>\n\n"
        )
        
        for idx, d in enumerate(payment_details, 1):
            msg += (
                f"<b>{idx}. {d['fullname']}</b> (<code>{d['shareholder_code']}</code>)\n"
                f"   Vốn góp: <code>{fmt_vn(d['investment_amount'])}</code>\n"
                f"   Cổ phần: <b>{d['share_pct']:.1f}%</b>\n"
                f"   Thanh toán: <code>{fmt_vn(d['payment'])}</code>\n\n"
            )

        # Daily Payment details
        total_payout = sum(d['payment'] for d in payment_details)
        msg += (
            f"━━━━━━━━━━━━━━━━━━\n"
            f"<b>PHIẾU CHI - DAILY PAYMENT ({len(payment_details)} phiếu)</b>\n\n"
        )

        for idx, d in enumerate(payment_details, 1):
            msg += (
                f"<b>{idx}. Phiếu Chi</b>\n"
                f"   <b>Loại:</b> CHI\n"
                f"   <b>Người Thực Hiện:</b> {d['shareholder_code']}\n"
                f"   <b>Người Nhận:</b> {d['fullname']}\n"
                f"   <b>Mục Đích:</b> Thanh toán Quỹ cổ đông {inv.name}\n"
                f"   <b>Số Tiền:</b> <code>{fmt_vn(d['payment'])}</code>\n"
                f"   <b>Ngày:</b> {today_str}\n"
                f"   <b>Trạng Thái:</b> APPROVED\n\n"
            )

        msg += (
            f"<b>Tổng Số Tiền Chi:</b> <code>{fmt_vn(total_payout)}</code>\n"
            f"━━━━━━━━━━━━━━━━━━\n\n"
        )

        if profit <= 0:
            msg += "<i>⚠️ Quỹ đang lỗ, không thể thanh toán cho cổ đông.</i>\n\n"

        msg += "<i>Vui lòng kiểm tra thông tin và nhấn nút bên dưới để xác nhận.</i>"

        buttons = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Xác Nhận Thanh Toán", callback_data=f"payshcf_y_{pending_key}"),
                InlineKeyboardButton("❌ Hủy", callback_data=f"payshcf_n_{pending_key}"),
            ]
        ])

        await callback_query.message.reply_text(msg, reply_markup=buttons, parse_mode=ParseMode.HTML)
        await callback_query.answer()
        
    except Exception as e:
        LogError(f"Error in payment shareholder callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống", show_alert=True)
    finally:
        db.close()


# --- Confirm / Cancel payment shareholder ---
@bot.on_callback_query(filters.regex(r"^payshcf_(y|n)_(.+)$"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
async def tien_nga_confirm_payment_sh_callback(client, callback_query):
    action = callback_query.matches[0].group(1)  # "y" or "n"
    pending_key = callback_query.matches[0].group(2)

    pending = _pending_payment_sh_data.pop(pending_key, None)
    if not pending:
        await callback_query.answer("⚠️ Dữ liệu đã hết hạn hoặc đã được xử lý.", show_alert=True)
        try:
            await callback_query.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
        return

    # Cancel
    if action == "n":
        await callback_query.message.edit_text(
            "❌ <b>ĐÃ HỦY</b>\n\n<i>Thao tác thanh toán quỹ cổ đông đã bị hủy.</i>",
            parse_mode=ParseMode.HTML
        )
        await callback_query.answer("Đã hủy thao tác.")
        return

    # Confirm → create DailyPayment for each shareholder
    from app.models.business import Investment, DailyPayment
    import uuid as _uuid

    db = SessionLocal()
    try:
        inv_id = pending["inv_id"]
        investment = db.query(Investment).filter(Investment.id == inv_id).first()
        if not investment:
            await callback_query.answer("⚠️ Quỹ Đầu Tư không tồn tại.", show_alert=True)
            return

        payment_details = pending["payment_details"]
        today = datetime.now().date()

        total_paid = 0

        for d in payment_details:
            pay_amount = d["payment"]
            if pay_amount <= 0:
                continue

            new_payment = DailyPayment(
                id=_uuid.uuid4(),
                investment_id=investment.id,
                requester=investment.name,
                executor=d["shareholder_code"],
                receiver=d["fullname"],
                payment_type="chi",
                purpose=f"Thanh toán Quỹ cổ đông {investment.name}",
                reason=f"Cổ đông {d['shareholder_code']} nhận thanh toán ({d['share_pct']:.1f}%)",
                amount=pay_amount,
                day=today,
                notes=f"Vốn góp: {fmt_vn(d['investment_amount'])} | Cổ phần: {d['share_pct']:.1f}%",
                status="APPROVED",
            )
            db.add(new_payment)
            total_paid += pay_amount

        # Cập nhật tổng chi của Investment
        if investment.total_expense is None:
            investment.total_expense = 0
        investment.total_expense += total_paid
        investment.profit = (investment.total_income or 0) - investment.total_expense

        # Quỹ hết hoạt động sau khi thanh toán
        investment.status = "CLOSED"
        investment.end_date = today

        db.commit()

        LogInfo(f"[TienNga] Payment shareholder confirmed for '{investment.name}', total paid: {fmt_vn(total_paid)} by @{callback_query.from_user.username}", LogType.SYSTEM_STATUS)

        # Build success message
        success_msg = (
            f"✅ <b>THANH TOÁN QUỸ CỔ ĐÔNG THÀNH CÔNG</b>\n"
            f"━━━━━━━━━━━━━━━━━━\n\n"
            f"<b>Quỹ:</b> {investment.name} (<code>{investment.investment_code or 'N/A'}</code>)\n"
            f"<b>Ngày:</b> {today.strftime('%d/%m/%Y')}\n\n"
        )

        for idx, d in enumerate(payment_details, 1):
            if d["payment"] <= 0:
                continue
            success_msg += (
                f"<b>{idx}. {d['fullname']}</b> (<code>{d['shareholder_code']}</code>)\n"
                f"   Thanh toán: <code>{fmt_vn(d['payment'])}</code>\n"
            )

        success_msg += (
            f"\n━━━━━━━━━━━━━━━━━━\n"
            f"<b>TỔNG CHI THANH TOÁN:</b> <code>{fmt_vn(total_paid)}</code>\n"
            f"<b>Số phiếu chi đã tạo:</b> {sum(1 for d in payment_details if d['payment'] > 0)}\n"
            f"<b>Trạng Thái Phiếu:</b> APPROVED\n\n"
            f"<b>📌 Quỹ '{investment.name}' đã được ĐÓNG.</b>\n"
            f"<b>Ngày Kết Thúc:</b> {today.strftime('%d/%m/%Y')}\n\n"
            f"<i>Tất cả phiếu chi đã được ghi nhận vào hệ thống.</i>"
        )

        await callback_query.message.edit_text(success_msg, parse_mode=ParseMode.HTML)
        await callback_query.answer("✅ Thanh toán thành công!")

        # Gửi thông báo xuống nhóm member của từng cổ đông
        try:
            from app.models.business import Shareholder
            from app.models.telegram import TelegramProjectMember

            for d in payment_details:
                if d["payment"] <= 0:
                    continue
                sh = db.query(Shareholder).filter(
                    Shareholder.shareholder_code == d["shareholder_code"],
                    Shareholder.investment_id == investment.id
                ).first()
                if not sh or not sh.telegram_group:
                    continue

                member_group = db.query(TelegramProjectMember).filter(
                    TelegramProjectMember.group_name == sh.telegram_group,
                    TelegramProjectMember.role == "member"
                ).first()
                if not member_group or not member_group.chat_id:
                    continue

                notify_msg = (
                    f"<b>THÔNG BÁO THANH TOÁN QUỸ CỔ ĐÔNG</b>\n"
                    f"━━━━━━━━━━━━━━━━━━\n\n"
                    f"<b>Quỹ Đầu Tư:</b> {investment.name}\n"
                    f"<b>Mã Cổ Đông:</b> <code>{d['shareholder_code']}</code>\n"
                    f"<b>Tên Cổ Đông:</b> {d['fullname']}\n"
                    f"<b>Cổ Phần:</b> {d['share_pct']:.1f}%\n"
                    f"<b>Số Tiền Thanh Toán:</b> <code>{fmt_vn(d['payment'])}</code>\n"
                    f"<b>Ngày:</b> {today.strftime('%d/%m/%Y')}\n\n"
                    f"<b>Trạng Thái:</b> ĐÃ XÁC NHẬN ✅\n\n"
                    f"<i>Phiếu chi đã được ghi nhận vào hệ thống.</i>"
                )
                try:
                    await client.send_message(
                        chat_id=int(member_group.chat_id),
                        text=notify_msg,
                        parse_mode=ParseMode.HTML
                    )
                    LogInfo(f"[TienNga] Sent payment notification to '{sh.telegram_group}' for {d['shareholder_code']}", LogType.SYSTEM_STATUS)
                except Exception as send_err:
                    LogError(f"[TienNga] Failed to send payment notification to '{sh.telegram_group}': {send_err}", LogType.SYSTEM_STATUS)
        except Exception as notify_err:
            LogError(f"[TienNga] Error sending payment notifications: {notify_err}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error confirming payment shareholder: {e}\n{traceback.format_exc()}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi lưu dữ liệu.", show_alert=True)
        await callback_query.message.edit_text(
            f"❌ <b>LỖI</b>\n\nCó lỗi xảy ra khi lưu vào database: {e}",
            parse_mode=ParseMode.HTML
        )
    finally:
        db.close()



# --- XUẤT BÁO CÁO SẢN PHẨM HÀNG NGÀY ---
@bot.on_message(filters.command(["tien_nga_export_daily_product", "tien_nga_xuat_bao_cao_san_pham"]) | filters.regex(r"^@\w+\s+/(tien_nga_export_daily_product|tien_nga_xuat_bao_cao_san_pham)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_PRODUCT)
async def tien_nga_export_daily_product_handler(client, message: Message) -> None:
    from datetime import datetime, timedelta
    text = message.text.strip()
    parts = text.split(maxsplit=1)
    
    if len(parts) < 2:
        # Không có tham số → hiển thị button chọn khoảng thời gian
        buttons = [
            [InlineKeyboardButton("1 Ngày", callback_data="tn_expprod_1d"),
             InlineKeyboardButton("7 Ngày", callback_data="tn_expprod_7d")],
            [InlineKeyboardButton("14 Ngày", callback_data="tn_expprod_14d"),
             InlineKeyboardButton("1 Tháng", callback_data="tn_expprod_1m")],
            [InlineKeyboardButton("3 Tháng", callback_data="tn_expprod_3m"),
             InlineKeyboardButton("6 Tháng", callback_data="tn_expprod_6m")],
            [InlineKeyboardButton("1 Năm", callback_data="tn_expprod_1y"),
             InlineKeyboardButton("Năm Trước", callback_data="tn_expprod_ly")],
            [InlineKeyboardButton("Hủy", callback_data="cancel_action")]
        ]
        await message.reply_text(
            "<b>XUẤT BÁO CÁO SẢN PHẨM</b>\n\nChọn khoảng thời gian:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.HTML
        )
        return
    
    # Có tham số ngày
    date_str = parts[1].strip()
    try:
        if " - " in date_str:
            d1, d2 = date_str.split(" - ", 1)
            start_date = datetime.strptime(d1.strip(), "%d/%m/%Y").date()
            end_date = datetime.strptime(d2.strip(), "%d/%m/%Y").date()
        else:
            start_date = datetime.strptime(date_str.strip(), "%d/%m/%Y").date()
            end_date = start_date
    except ValueError:
        await message.reply_text(
            "⚠️ Sai định dạng ngày.\n\n"
            "<i>VD: <code>/tien_nga_export_daily_product 22/04/2026</code>\n"
            "<code>/tien_nga_export_daily_product 01/04/2026 - 22/04/2026</code></i>",
            parse_mode=ParseMode.HTML
        )
        return
    
    await _generate_daily_product_report(message, start_date, end_date)


@bot.on_callback_query(filters.regex(r"^tn_expprod_(.+)$"))
async def tien_nga_export_daily_product_callback(client, callback_query):
    from datetime import datetime, timedelta
    period = callback_query.matches[0].group(1)
    today = datetime.now().date()
    
    if period == "1d":
        start_date = today
        end_date = today
    elif period == "7d":
        start_date = today - timedelta(days=6)
        end_date = today
    elif period == "14d":
        start_date = today - timedelta(days=13)
        end_date = today
    elif period == "1m":
        start_date = today - timedelta(days=29)
        end_date = today
    elif period == "3m":
        start_date = today - timedelta(days=89)
        end_date = today
    elif period == "6m":
        start_date = today - timedelta(days=179)
        end_date = today
    elif period == "1y":
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif period == "ly":
        start_date = today.replace(year=today.year - 1, month=1, day=1)
        end_date = today.replace(year=today.year - 1, month=12, day=31)
    else:
        await callback_query.answer("❌ Không hợp lệ", show_alert=True)
        return
    
    await callback_query.answer("⏳ Đang tạo báo cáo...")
    await _generate_daily_product_report(callback_query.message, start_date, end_date)


async def _generate_daily_product_report(message, start_date, end_date):
    from app.db.session import SessionLocal
    from app.models.inventory import ProductTransaction, Inventory
    from datetime import datetime, timedelta
    import tempfile
    import os
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    db = SessionLocal()
    try:
        # Query transactions chỉ liên quan Mủ thành phẩm
        transactions = db.query(ProductTransaction).filter(
            ProductTransaction.transaction_date >= start_date,
            ProductTransaction.transaction_date <= end_date,
            ProductTransaction.material_type.ilike("%Mủ%")
        ).order_by(ProductTransaction.transaction_date).all()
        
        # Query current inventory - chỉ kho Mủ thành phẩm
        inventories = db.query(Inventory).filter(Inventory.material_name.ilike("%Mủ%")).all()
        inv_map = {str(inv.id): inv for inv in inventories}
        
        # Get unique storage names from filtered transactions
        storage_names = sorted(set(t.storage_name for t in transactions if t.storage_name))
        if not storage_names and not transactions:
            await message.reply_text("⚠️ Không có dữ liệu giao dịch Mủ thành phẩm trong khoảng thời gian này.", parse_mode=ParseMode.HTML)
            return
        
        # Also add storages from Mủ inventory that may not have transactions
        for inv in inventories:
            if inv.storage_name and inv.storage_name not in storage_names:
                storage_names.append(inv.storage_name)
        storage_names = sorted(set(storage_names))
        
        # Build date list
        date_list = []
        d = start_date
        while d <= end_date:
            date_list.append(d)
            d += timedelta(days=1)
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        sub_header_fill = PatternFill("solid", fgColor="4472C4")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        money_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        alt_fill = PatternFill("solid", fgColor="D9E2F3")
        total_fill = PatternFill("solid", fgColor="FFC000")
        total_font = Font(bold=True, size=11)
        inventory_fill = PatternFill("solid", fgColor="92D050")
        
        wb = openpyxl.Workbook()
        
        def fmt_money(val):
            if val is None or val == 0:
                return "0"
            return f"{int(val):,}".replace(",", ".")
        
        def write_storage_sheet(ws, sheet_storage_name, txns, is_total=False):
            """Write a sheet for one storage or total"""
            ws.title = sheet_storage_name[:31]  # Excel sheet name limit
            
            headers = ["Ngày", "SL Nhập (kg)", "SL Xuất (kg)", "Thành Tiền Nhập", "Thành Tiền Xuất"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            grand_total_import_qty = 0
            grand_total_export_qty = 0
            grand_total_import_amount = 0
            grand_total_export_amount = 0
            
            row = 2
            for idx, d in enumerate(date_list):
                day_txns = [t for t in txns if t.transaction_date == d]
                
                import_qty = sum(t.quantity or 0 for t in day_txns if t.transaction_type and t.transaction_type.lower() in ("nhập", "nhap", "import"))
                export_qty = sum(t.quantity or 0 for t in day_txns if t.transaction_type and t.transaction_type.lower() in ("xuất", "xuat", "export"))
                import_amount = sum(t.total_amount or 0 for t in day_txns if t.transaction_type and t.transaction_type.lower() in ("nhập", "nhap", "import"))
                export_amount = sum(t.total_amount or 0 for t in day_txns if t.transaction_type and t.transaction_type.lower() in ("xuất", "xuat", "export"))
                
                grand_total_import_qty += import_qty
                grand_total_export_qty += export_qty
                grand_total_import_amount += import_amount
                grand_total_export_amount += export_amount
                
                row_fill = alt_fill if idx % 2 == 0 else None
                values = [
                    d.strftime("%d/%m/%Y"),
                    f"{import_qty:,.1f}".replace(",", ".") if import_qty else "0",
                    f"{export_qty:,.1f}".replace(",", ".") if export_qty else "0",
                    fmt_money(import_amount),
                    fmt_money(export_amount)
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx == 1:
                        cell.alignment = center_align
                    elif col_idx in (4, 5):
                        cell.alignment = money_align
                    else:
                        cell.alignment = center_align
                    if row_fill:
                        cell.fill = row_fill
                row += 1
            
            # Dòng TỔNG
            total_values = [
                "TỔNG CỘNG",
                f"{grand_total_import_qty:,.1f}".replace(",", "."),
                f"{grand_total_export_qty:,.1f}".replace(",", "."),
                fmt_money(grand_total_import_amount),
                fmt_money(grand_total_export_amount)
            ]
            for col_idx, val in enumerate(total_values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = center_align
                elif col_idx in (4, 5):
                    cell.alignment = money_align
                else:
                    cell.alignment = center_align
            row += 1
            
            # Dòng TỒN KHO
            if is_total:
                total_inv_qty = sum((inv.quantity or 0) for inv in inventories)
                inv_label = "Tồn kho (Tất cả kho)"
            else:
                matched_inv = [inv for inv in inventories if inv.storage_name == sheet_storage_name]
                total_inv_qty = sum((inv.quantity or 0) for inv in matched_inv)
                inv_label = f"Tồn kho ({sheet_storage_name})"
            
            cell = ws.cell(row=row, column=1, value=inv_label)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = inventory_fill
            cell.alignment = center_align
            cell.border = thin_border
            
            cell = ws.cell(row=row, column=2, value=f"{total_inv_qty:,.1f}".replace(",", "."))
            cell.font = Font(bold=True, size=11)
            cell.fill = inventory_fill
            cell.alignment = center_align
            cell.border = thin_border
            
            for col_idx in range(3, 6):
                cell = ws.cell(row=row, column=col_idx, value="")
                cell.fill = inventory_fill
                cell.border = thin_border
            
            # Column widths
            col_widths = [15, 15, 15, 20, 20]
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            ws.freeze_panes = "A2"
        
        # Sheet cho từng kho
        first_sheet = True
        for sn in storage_names:
            if first_sheet:
                ws = wb.active
                first_sheet = False
            else:
                ws = wb.create_sheet()
            
            storage_txns = [t for t in transactions if t.storage_name == sn]
            write_storage_sheet(ws, sn, storage_txns, is_total=False)
        
        # Sheet TỔNG HỢP
        ws_total = wb.create_sheet() if not first_sheet else wb.active
        write_storage_sheet(ws_total, "TỔNG HỢP", transactions, is_total=True)
        
        # Save and send
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)
        
        now_str = datetime.now().strftime("%Y%m%d_%H%M")
        period_str = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        
        await message.reply_document(
            document=tmp_path,
            file_name=f"bao_cao_san_pham_{now_str}.xlsx",
            caption=(
                f"<b>BÁO CÁO XUẤT NHẬP KHO MỦ THÀNH PHẨM</b>\n\n"
                f"<b>Khoảng thời gian:</b> {period_str}\n"
                f"<b>Tổng giao dịch:</b> {len(transactions)}\n"
                f"<b>Số kho:</b> {len(storage_names)}\n\n"
                f"<i>Xuất lúc: {datetime.now().strftime('%H:%M %d/%m/%Y')}</i>"
            ),
            parse_mode=ParseMode.HTML,
        )
        
        os.remove(tmp_path)
        LogInfo(f"[TienNga] Exported daily product report ({period_str}), {len(transactions)} txns", LogType.SYSTEM_STATUS)
        
    except Exception as e:
        LogError(f"Error in export daily product: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(f"❌ Có lỗi xảy ra: {e}", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# --- KIỂM TRA LỊCH SỬ GIAO DỊCH CỔ ĐÔNG ---
@bot.on_message(filters.command(["tien_nga_check_history_transaction", "tien_nga_lich_su_giao_dich_co_dong", "tien_nga_lich_su_gd"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_history_transaction|tien_nga_lich_su_giao_dich_co_dong|tien_nga_lich_su_gd)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_FINANCE, CustomTitle.MAIN_SHAREHOLDER)
async def tien_nga_check_history_transaction_handler(client, message: Message) -> None:
    """
    /tien_nga_check_history_transaction [mã cổ đông] [dd/mm/yyyy - dd/mm/yyyy]
    Lấy lịch sử giao dịch của cổ đông: góp vốn, rút vốn, nhận cổ tức, thanh toán quỹ cổ đông, ...
    """
    args_text = (message.text or "").strip()
    # Remove command prefix
    for cmd in ["tien_nga_check_history_transaction", "tien_nga_lich_su_giao_dich_co_dong", "tien_nga_lich_su_gd"]:
        args_text = re.sub(rf"^/?{cmd}\s*", "", args_text, flags=re.IGNORECASE).strip()
    # Also handle @bot prefix
    args_text = re.sub(r"^@\w+\s+/?(?:tien_nga_check_history_transaction|tien_nga_lich_su_giao_dich_co_dong|tien_nga_lich_su_gd)\s*", "", args_text, flags=re.IGNORECASE).strip()

    if not args_text:
        await message.reply_text(
            "<b>KIỂM TRA LỊCH SỬ GIAO DỊCH CỔ ĐÔNG</b>\n\n"
            "Cú pháp: <code>/tien_nga_check_history_transaction [mã cổ đông] [dd/mm/yyyy - dd/mm/yyyy]</code>\n\n"
            "<i>Ví dụ:</i>\n"
            "<code>/tien_nga_check_history_transaction CD001 01/01/2025 - 31/12/2025</code>\n\n"
            "<i>Nếu không nhập ngày, hệ thống sẽ lấy toàn bộ lịch sử giao dịch.</i>",
            parse_mode=ParseMode.HTML
        )
        return

    # Parse: [mã cổ đông] [dd/mm/yyyy - dd/mm/yyyy]
    date_range_match = re.search(r"(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})", args_text)

    start_date = None
    end_date = None
    if date_range_match:
        try:
            start_date = datetime.strptime(date_range_match.group(1), "%d/%m/%Y").date()
            end_date = datetime.strptime(date_range_match.group(2), "%d/%m/%Y").date()
        except ValueError:
            await message.reply_text(
                "⚠️ Định dạng ngày không hợp lệ. Vui lòng dùng <code>dd/mm/yyyy - dd/mm/yyyy</code>.",
                parse_mode=ParseMode.HTML
            )
            return

        if start_date > end_date:
            start_date, end_date = end_date, start_date

        # Shareholder code is everything before the date range
        shareholder_code = args_text[:date_range_match.start()].strip()
    else:
        # No date range: entire string is the shareholder code
        shareholder_code = args_text.strip()

    if not shareholder_code:
        await message.reply_text(
            "⚠️ Vui lòng nhập <b>Mã Cổ Đông</b>.",
            parse_mode=ParseMode.HTML
        )
        return

    from app.models.business import Shareholder, Investment, DailyPayment
    from sqlalchemy import or_

    db = SessionLocal()
    try:
        # Find shareholder
        shareholder = db.query(Shareholder).filter(Shareholder.shareholder_code == shareholder_code).first()
        if not shareholder:
            await message.reply_text(
                f"⚠️ Không tìm thấy cổ đông với mã <b>{shareholder_code}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        # Find associated investment
        investment = db.query(Investment).filter(Investment.id == shareholder.investment_id).first()
        inv_name = investment.name if investment else "N/A"
        inv_code = (investment.investment_code if investment and investment.investment_code else "N/A")

        # Build date range label
        if start_date and end_date:
            period_label = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        else:
            period_label = "Toàn bộ"

        # ------------------------------------------------------------------
        # Collect transactions from DailyPayment that relate to this shareholder
        # Match by: receiver, requester, executor, purpose, or notes containing shareholder_code or shareholder fullname
        # ------------------------------------------------------------------
        sh_code = shareholder.shareholder_code
        sh_name = shareholder.fullname or ""

        query = db.query(DailyPayment).filter(
            DailyPayment.investment_id == shareholder.investment_id,
            DailyPayment.status == "APPROVED",
            or_(
                DailyPayment.receiver.ilike(f"%{sh_code}%"),
                DailyPayment.requester.ilike(f"%{sh_code}%"),
                DailyPayment.executor.ilike(f"%{sh_code}%"),
                DailyPayment.purpose.ilike(f"%{sh_code}%"),
                DailyPayment.notes.ilike(f"%{sh_code}%"),
                DailyPayment.receiver.ilike(f"%{sh_name}%"),
                DailyPayment.requester.ilike(f"%{sh_name}%"),
                DailyPayment.executor.ilike(f"%{sh_name}%"),
                DailyPayment.purpose.ilike(f"%{sh_name}%"),
                DailyPayment.notes.ilike(f"%{sh_name}%"),
            )
        )

        if start_date:
            query = query.filter(DailyPayment.day >= start_date)
        if end_date:
            query = query.filter(DailyPayment.day <= end_date)

        payments = query.order_by(DailyPayment.day.asc()).all()

        # ------------------------------------------------------------------
        # Aggregate totals by shareholders portion vs investment fund
        # ------------------------------------------------------------------
        all_shareholders = db.query(Shareholder).filter(
            Shareholder.investment_id == shareholder.investment_id
        ).all()
        total_invest_amount = sum(sh.investment_amount or 0 for sh in all_shareholders)
        sh_invest_amount = shareholder.investment_amount or 0
        share_pct = (sh_invest_amount / total_invest_amount * 100) if total_invest_amount > 0 else 0

        # ------------------------------------------------------------------
        # Build the report message
        # ------------------------------------------------------------------
        msg = (
            f"<b>LỊCH SỬ GIAO DỊCH CỔ ĐÔNG</b>\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"<b>Mã Cổ Đông:</b> <code>{sh_code}</code>\n"
            f"<b>Tên Cổ Đông:</b> {sh_name}\n"
            f"<b>Quỹ Đầu Tư:</b> {inv_name} (<code>{inv_code}</code>)\n"
            f"<b>Vốn Góp:</b> <code>{fmt_vn(sh_invest_amount)}</code>\n"
            f"<b>Cổ Phần:</b> <b>{share_pct:.1f}%</b>\n"
            f"<b>Khoảng thời gian:</b> {period_label}\n"
            f"━━━━━━━━━━━━━━━━━━\n\n"
        )

        if not payments:
            msg += "<i>Không có giao dịch thu/chi nào trong khoảng thời gian này.</i>\n"
        else:
            total_thu = 0
            total_chi = 0

            msg += f"<b>CHI TIẾT GIAO DỊCH ({len(payments)})</b>\n\n"

            for idx, p in enumerate(payments, 1):
                p_type = "THU" if p.payment_type == "thu" else "CHI"
                p_icon = "📈" if p.payment_type == "thu" else "📉"
                p_day = p.day.strftime("%d/%m/%Y") if p.day else "—"
                p_amount = p.amount or 0

                if p.payment_type == "thu":
                    total_thu += p_amount
                else:
                    total_chi += p_amount

                msg += (
                    f"{p_icon} <b>{idx}. [{p_type}] {p.purpose or 'Không rõ'}</b>\n"
                    f"   Ngày: {p_day}\n"
                    f"   Số tiền: <code>{fmt_vn(p_amount)}</code>\n"
                )
                if p.executor:
                    msg += f"   Người TH: {p.executor}\n"
                if p.receiver:
                    msg += f"   Người nhận: {p.receiver}\n"
                if p.requester:
                    msg += f"   Người YC: {p.requester}\n"
                if p.reason:
                    msg += f"   Lý do: {p.reason}\n"
                if p.notes:
                    msg += f"   Ghi chú: {p.notes}\n"
                msg += "\n"

            # Summary
            msg += (
                f"━━━━━━━━━━━━━━━━━━\n"
                f"<b>TỔNG KẾT</b>\n\n"
                f"Tổng Thu: <code>{fmt_vn(total_thu)}</code>\n"
                f"Tổng Chi: <code>{fmt_vn(total_chi)}</code>\n"
                f"Chênh Lệch: <code>{fmt_vn(total_thu - total_chi)}</code>\n"
            )

        # If message too long, split into parts
        if len(msg) > 4000:
            parts = []
            while msg:
                if len(msg) <= 4000:
                    parts.append(msg)
                    break
                split_idx = msg.rfind("\n", 0, 4000)
                if split_idx == -1:
                    split_idx = 4000
                parts.append(msg[:split_idx])
                msg = msg[split_idx:]

            for part in parts:
                await message.reply_text(part, parse_mode=ParseMode.HTML)
        else:
            await message.reply_text(msg, parse_mode=ParseMode.HTML)

        LogInfo(f"[TienNga] Check history transaction for shareholder '{sh_code}' ({period_label}) by @{message.from_user.username or message.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error checking shareholder transaction history: {e}\n{traceback.format_exc()}", LogType.SYSTEM_STATUS)
        await message.reply_text(f"❌ Có lỗi xảy ra: {e}", parse_mode=ParseMode.HTML)
    finally:
        db.close()

# =========================================================================================
# QUẢN LÝ ĐẤT TRỒNG TRỌT (THU HOẠCH)
# =========================================================================================

@bot.on_message(filters.command(["tien_nga_create_agricultural_land", "tien_nga_tao_dat_trong_trot"]) | filters.regex(r"^@\w+\s+/(tien_nga_create_agricultural_land|tien_nga_tao_dat_trong_trot)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_create_agricultural_land_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        form = """<b>FORM TẠO ĐẤT TRỒNG TRỌT</b>

Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/tien_nga_tao_dat_trong_trot
Mã Đất: 
Tên Đất: 
Địa Chỉ: 
Diện Tích (ha): 0
DT Khai Thác Cao Su (ha): 0
DT Trống (ha): 0
DT Đang Trồng (ha): 0
SL Cây Thu Hoạch: 0
SL Cây Đang Trồng: 0</pre>

<i>Ghi chú: Mã đất nên viết liền không dấu (VD: DCS001).</i>"""
        await message.reply_text(form, parse_mode=ParseMode.HTML)
        return

    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    land_code = data.get("Mã Đất", "").strip().upper()
    land_name = data.get("Tên Đất", "").strip()
    address = data.get("Địa Chỉ", "").strip()
    total_area = parse_float_vn(data.get("Diện Tích (ha)", "0"))
    rubber_area = parse_float_vn(data.get("DT Khai Thác Cao Su (ha)", "0"))
    empty_area = parse_float_vn(data.get("DT Trống (ha)", "0"))
    planting_area = parse_float_vn(data.get("DT Đang Trồng (ha)", "0"))
    harvesting_trees = int(parse_float_vn(data.get("SL Cây Thu Hoạch", "0")))
    planting_trees = int(parse_float_vn(data.get("SL Cây Đang Trồng", "0")))

    if not land_code:
        await message.reply_text("⚠️ <b>Mã Đất</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    from app.models.business import AgriculturalLand
    import uuid as uuid_lib

    db = SessionLocal()
    try:
        existing = db.query(AgriculturalLand).filter(AgriculturalLand.land_code == land_code).first()
        if existing:
            await message.reply_text(f"⚠️ Mã đất <b>{land_code}</b> đã tồn tại.", parse_mode=ParseMode.HTML)
            return

        new_land = AgriculturalLand(
            id=uuid_lib.uuid4(), land_code=land_code, land_name=land_name or None,
            address=address or None, total_area=total_area, rubber_area=rubber_area,
            empty_area=empty_area, planting_area=planting_area,
            harvesting_trees=harvesting_trees, planting_trees=planting_trees, status="ACTIVE"
        )
        db.add(new_land)
        db.commit()

        await message.reply_text(
            f"✅ <b>Tạo Đất Trồng Trọt Thành Công!</b>\n\n"
            f"<b>Mã Đất:</b> <code>{land_code}</code>\n"
            f"<b>Tên Đất:</b> {land_name or '—'}\n"
            f"<b>Địa Chỉ:</b> {address or '—'}\n"
            f"<b>Diện Tích:</b> {total_area} ha\n"
            f"<b>DT Cao Su:</b> {rubber_area} ha\n"
            f"<b>DT Trống:</b> {empty_area} ha\n"
            f"<b>DT Đang Trồng:</b> {planting_area} ha\n"
            f"<b>SL Cây Thu Hoạch:</b> {harvesting_trees}\n"
            f"<b>SL Cây Đang Trồng:</b> {planting_trees}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        db.rollback()
        LogError(f"Error creating agricultural land: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_update_agricultural_land", "tien_nga_cap_nhat_dat_trong_trot"]) | filters.regex(r"^@\w+\s+/(tien_nga_update_agricultural_land|tien_nga_cap_nhat_dat_trong_trot)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_update_agricultural_land_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    from app.models.business import AgriculturalLand

    if len(lines) < 2:
        args = lines[0].split()
        if len(args) < 2:
            await message.reply_text("⚠️ Cú pháp: <code>/tien_nga_cap_nhat_dat_trong_trot [Mã Đất]</code>", parse_mode=ParseMode.HTML)
            return
        land_code = args[1].upper()
        db = SessionLocal()
        try:
            land = db.query(AgriculturalLand).filter(AgriculturalLand.land_code == land_code, AgriculturalLand.status == "ACTIVE").first()
            if not land:
                await message.reply_text(f"⚠️ Không tìm thấy đất với mã <b>{land_code}</b>.", parse_mode=ParseMode.HTML)
                return
            form = f"""<b>CẬP NHẬT ĐẤT TRỒNG TRỌT</b>

Sao chép form dưới, chỉnh sửa và gửi lại:

<pre>/tien_nga_cap_nhat_dat_trong_trot
Mã Đất: {land.land_code}
Tên Đất: {land.land_name or ''}
Địa Chỉ: {land.address or ''}
Diện Tích (ha): {land.total_area}
DT Khai Thác Cao Su (ha): {land.rubber_area}
DT Trống (ha): {land.empty_area}
DT Đang Trồng (ha): {land.planting_area}
SL Cây Thu Hoạch: {land.harvesting_trees or 0}
SL Cây Đang Trồng: {land.planting_trees or 0}</pre>"""
            await message.reply_text(form, parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    land_code = data.get("Mã Đất", "").strip().upper()
    if not land_code:
        await message.reply_text("⚠️ <b>Mã Đất</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        land = db.query(AgriculturalLand).filter(AgriculturalLand.land_code == land_code, AgriculturalLand.status == "ACTIVE").first()
        if not land:
            await message.reply_text(f"⚠️ Không tìm thấy đất với mã <b>{land_code}</b>.", parse_mode=ParseMode.HTML)
            return

        if "Tên Đất" in data: land.land_name = data["Tên Đất"] or None
        if "Địa Chỉ" in data: land.address = data["Địa Chỉ"] or None
        if "Diện Tích (ha)" in data: land.total_area = parse_float_vn(data["Diện Tích (ha)"])
        if "DT Khai Thác Cao Su (ha)" in data: land.rubber_area = parse_float_vn(data["DT Khai Thác Cao Su (ha)"])
        if "DT Trống (ha)" in data: land.empty_area = parse_float_vn(data["DT Trống (ha)"])
        if "DT Đang Trồng (ha)" in data: land.planting_area = parse_float_vn(data["DT Đang Trồng (ha)"])
        if "SL Cây Thu Hoạch" in data: land.harvesting_trees = int(parse_float_vn(data["SL Cây Thu Hoạch"]))
        if "SL Cây Đang Trồng" in data: land.planting_trees = int(parse_float_vn(data["SL Cây Đang Trồng"]))
        db.commit()

        await message.reply_text(
            f"✅ <b>Cập Nhật Đất Trồng Trọt Thành Công!</b>\n\n"
            f"<b>Mã Đất:</b> <code>{land.land_code}</code>\n"
            f"<b>Tên Đất:</b> {land.land_name or '—'}\n"
            f"<b>Địa Chỉ:</b> {land.address or '—'}\n"
            f"<b>Diện Tích:</b> {land.total_area} ha\n"
            f"<b>DT Cao Su:</b> {land.rubber_area} ha\n"
            f"<b>DT Trống:</b> {land.empty_area} ha\n"
            f"<b>DT Đang Trồng:</b> {land.planting_area} ha\n"
            f"<b>SL Cây Thu Hoạch:</b> {land.harvesting_trees or 0}\n"
            f"<b>SL Cây Đang Trồng:</b> {land.planting_trees or 0}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        db.rollback()
        LogError(f"Error updating agricultural land: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_delete_agricultural_land", "tien_nga_xoa_dat_trong_trot"]) | filters.regex(r"^@\w+\s+/(tien_nga_delete_agricultural_land|tien_nga_xoa_dat_trong_trot)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_delete_agricultural_land_handler(client, message: Message) -> None:
    args = message.text.split()
    if len(args) < 2:
        await message.reply_text("⚠️ Cú pháp: <code>/tien_nga_xoa_dat_trong_trot [Mã Đất]</code>", parse_mode=ParseMode.HTML)
        return

    land_code = args[1].upper()
    from app.models.business import AgriculturalLand

    db = SessionLocal()
    try:
        land = db.query(AgriculturalLand).filter(AgriculturalLand.land_code == land_code, AgriculturalLand.status == "ACTIVE").first()
        if not land:
            await message.reply_text(f"⚠️ Không tìm thấy đất với mã <b>{land_code}</b>.", parse_mode=ParseMode.HTML)
            return

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Xóa", callback_data=f"del_land_confirm:{land_code}"),
             InlineKeyboardButton("Hủy", callback_data="del_land_cancel")]
        ])
        await message.reply_text(
            f"⚠️ <b>Xác nhận xóa đất trồng trọt?</b>\n\n"
            f"<b>Mã Đất:</b> <code>{land_code}</code>\n"
            f"<b>Địa Chỉ:</b> {land.address or '—'}\n"
            f"<b>Diện Tích:</b> {land.total_area} ha",
            reply_markup=keyboard, parse_mode=ParseMode.HTML
        )
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^del_land_(confirm|cancel)"))
async def del_land_callback(client, callback_query) -> None:
    data = callback_query.data
    if data == "del_land_cancel":
        await callback_query.message.delete()
        return

    land_code = data.split(":")[1]
    from app.models.business import AgriculturalLand

    db = SessionLocal()
    try:
        land = db.query(AgriculturalLand).filter(AgriculturalLand.land_code == land_code, AgriculturalLand.status == "ACTIVE").first()
        if land:
            land.status = "INACTIVE"
            db.commit()
            await callback_query.message.edit_text(f"✅ Đã ngưng hoạt động đất <b>{land_code}</b> thành công.", parse_mode=ParseMode.HTML)
        else:
            await callback_query.answer("⚠️ Không tìm thấy.", show_alert=True)
    except Exception as e:
        db.rollback()
        LogError(f"Error deleting land: {e}", LogType.SYSTEM_STATUS)
        await callback_query.message.edit_text("❌ Lỗi hệ thống.")
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_list_agricultural_land", "tien_nga_danh_sach_dat_trong_trot", "tien_nga_ds_dat_trong_trot"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_agricultural_land|tien_nga_danh_sach_dat_trong_trot|tien_nga_ds_dat_trong_trot)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_list_agricultural_land_handler(client, message: Message) -> None:
    from app.models.business import AgriculturalLand

    db = SessionLocal()
    try:
        lands = db.query(AgriculturalLand).filter(AgriculturalLand.status == "ACTIVE").order_by(AgriculturalLand.land_code).all()
        if not lands:
            await message.reply_text("📋 Chưa có đất trồng trọt nào trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        if len(lands) <= 10:
            text = f"<b>DANH SÁCH ĐẤT TRỒNG TRỌT ({len(lands)})</b>\n\n"
            for idx, l in enumerate(lands, 1):
                text += (
                    f"{idx}. <b>{l.land_name or l.land_code}</b>\n"
                    f"   Mã: <code>{l.land_code}</code>\n"
                    f"   Tên Đất: {l.land_name or '—'}\n"
                    f"   Địa Chỉ: {l.address or '—'}\n"
                    f"   DT Tổng: {l.total_area} ha | Cao Su: {l.rubber_area} ha\n"
                    f"   Trống: {l.empty_area} ha | Đang Trồng: {l.planting_area} ha\n"
                    f"   Cây Thu Hoạch: {l.harvesting_trees or 0} | Cây Đang Trồng: {l.planting_trees or 0}\n\n"
                )
            await message.reply_text(text, parse_mode=ParseMode.HTML)
        else:
            import openpyxl, tempfile, os
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Đất Trồng Trọt"
            headers = ["STT", "Mã Đất", "Địa Chỉ", "DT Tổng (ha)", "DT Cao Su (ha)", "DT Trống (ha)", "DT Đang Trồng (ha)", "Cây Thu Hoạch", "Cây Đang Trồng"]
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            for row_idx, l in enumerate(lands, 2):
                vals = [row_idx - 1, l.land_code, l.address or "", l.total_area, l.rubber_area, l.empty_area, l.planting_area, l.harvesting_trees or 0, l.planting_trees or 0]
                for col_idx, v in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=v)
                    cell.border = thin_border
            col_widths = [6, 12, 30, 15, 18, 15, 20, 15, 18]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp_path = tmp.name
            wb.save(tmp_path)
            await message.reply_document(document=tmp_path, file_name=f"dat_trong_trot_{len(lands)}.xlsx",
                caption=f"<b>DANH SÁCH ĐẤT TRỒNG TRỌT ({len(lands)})</b>", parse_mode=ParseMode.HTML)
            os.remove(tmp_path)
    except Exception as e:
        LogError(f"Error listing agricultural land: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

# =========================================================================================
# QUẢN LÝ HỘ DÂN (THU HOẠCH)
# =========================================================================================

@bot.on_message(filters.command(["tien_nga_list_household", "tien_nga_ds_ho_dan"]) | filters.regex(r"^@\w+\s+/(tien_nga_list_household|tien_nga_ds_ho_dan)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_list_household_handler(client, message: Message) -> None:
    from app.models.business import Households

    db = SessionLocal()
    try:
        households = db.query(Households).filter(Households.status == "ACTIVE").order_by(Households.household_code).all()
        if not households:
            await message.reply_text("📋 Chưa có hộ dân nào trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        def fmt_money(val):
            if val is None or val == 0: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        if len(households) <= 10:
            text = f"<b>DANH SÁCH HỘ DÂN ({len(households)})</b>\n\n"
            for idx, h in enumerate(households, 1):
                text += (
                    f"{idx}. <b>{h.fullname or h.household_code}</b>\n"
                    f"   Mã HD: <code>{h.household_code}</code>\n"
                    f"   Mã TM: <code>{h.purchase_code or '—'}</code>\n"
                    f"   Mã Đất: <code>{h.land_code or '—'}</code>\n"
                    f"   Username: @{h.username or '—'}\n"
                    f"   Nhóm TG: {h.telegram_group or '—'}\n"
                    f"   SĐT: {h.phone or '—'}\n"
                    f"   Địa Chỉ: {h.address or '—'}\n"
                    f"   Công Nợ: <code>{fmt_money(h.total_debt)}</code>\n"
                    f"   Đơn Giá Cạo Mủ: <code>{fmt_money(h.tapping_price)}</code>\n"
                    f"   Ngân Hàng: {h.bank_name or '—'}\n"
                    f"   Số TK: <code>{h.bank_account or '—'}</code>\n\n"
                )
            await message.reply_text(text, parse_mode=ParseMode.HTML)
        else:
            import openpyxl, tempfile, os
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Hộ Dân"
            headers = ["STT", "Mã Hộ Dân", "Mã Thu Mua", "Mã Đất", "Họ Tên", "Username", "Nhóm TG", "SĐT", "Địa Chỉ", "Công Nợ", "ĐG Cạo Mủ", "Ngân Hàng", "Số TK"]
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col_idx, hd in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=hd)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            for row_idx, h in enumerate(households, 2):
                vals = [row_idx - 1, h.household_code, h.purchase_code or "", h.land_code or "",
                        h.fullname or "", h.username or "", h.telegram_group or "",
                        h.phone or "", h.address or "",
                        h.total_debt or 0, h.tapping_price or 0, h.bank_name or "", h.bank_account or ""]
                for col_idx, v in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=v)
                    cell.border = thin_border
            col_widths = [6, 14, 14, 12, 20, 16, 16, 14, 25, 15, 14, 15, 18]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp_path = tmp.name
            wb.save(tmp_path)
            await message.reply_document(document=tmp_path, file_name=f"ho_dan_{len(households)}.xlsx",
                caption=f"<b>DANH SÁCH HỘ DÂN ({len(households)})</b>", parse_mode=ParseMode.HTML)
            os.remove(tmp_path)
    except Exception as e:
        LogError(f"Error listing households: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_check_harvest", "tien_nga_kiem_tra_thu_hoach"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_harvest|tien_nga_kiem_tra_thu_hoach)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_check_harvest_handler(client, message: Message) -> None:
    from datetime import datetime, timedelta
    from app.models.business import Households, DailyPurchases

    args = message.text.strip().split(maxsplit=1)
    if len(args) < 2:
        today = datetime.now()
        first_day = today.replace(day=1)
        await message.reply_text(
            "<b>KIỂM TRA THU HOẠCH</b>\n\n"
            "Cú pháp: <code>/tien_nga_kiem_tra_thu_hoach dd/mm/yyyy - dd/mm/yyyy</code>\n\n"
            f"VD: <code>/tien_nga_kiem_tra_thu_hoach {first_day.strftime('%d/%m/%Y')} - {today.strftime('%d/%m/%Y')}</code>",
            parse_mode=ParseMode.HTML
        )
        return

    date_str = args[1].strip()
    try:
        parts = date_str.split("-")
        if len(parts) != 2:
            raise ValueError("Thiếu dấu -")
        start_date = datetime.strptime(parts[0].strip(), "%d/%m/%Y").date()
        end_date = datetime.strptime(parts[1].strip(), "%d/%m/%Y").date()
    except Exception:
        await message.reply_text("⚠️ Định dạng ngày không hợp lệ.\nVD: <code>/tien_nga_kiem_tra_thu_hoach 01/04/2026 - 30/04/2026</code>", parse_mode=ParseMode.HTML)
        return

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    db = SessionLocal()
    try:
        # Lấy tất cả hộ dân active
        households = db.query(Households).filter(Households.status == "ACTIVE").order_by(Households.household_code).all()
        if not households:
            await message.reply_text("📋 Chưa có hộ dân nào trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        import openpyxl, tempfile, os
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal="center", vertical="center")

        has_data = False

        for hh in households:
            # Tìm purchases bằng purchase_code (hoursehold_id trong daily_purchases)
            purchases = db.query(DailyPurchases).filter(
                DailyPurchases.hoursehold_id == hh.purchase_code,
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date
            ).order_by(DailyPurchases.day).all()

            if not purchases:
                continue

            has_data = True
            # Sheet name max 31 chars
            sheet_name = f"{hh.household_code} - {hh.fullname or ''}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Title
            ws.merge_cells("A1:N1")
            title_cell = ws.cell(row=1, column=1,
                value=f"THU HOẠCH - {hh.fullname or hh.household_code} ({hh.purchase_code})")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:N2")
            info_cell = ws.cell(row=2, column=1,
                value=f"Mã HD: {hh.household_code} | Mã TM: {hh.purchase_code} | Mã Đất: {hh.land_code or '—'} | Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}")
            info_cell.alignment = Alignment(horizontal="center")

            # Headers
            headers = [
                "STT", "Tuần", "Ngày", "Trợ Giá",
                "KL (kg)", "Trừ Bì (kg)", "KL Mủ TT (kg)",
                "Số Độ (%)", "Mủ Khô", "Đơn Giá",
                "Giá HT", "Thành Tiền", "Đã TT", "Lưu Sổ"
            ]
            for col_idx, hd in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=hd)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = thin_border
                cell.alignment = center

            # Data rows
            for row_idx, p in enumerate(purchases, 5):
                vals = [
                    row_idx - 4,
                    p.week or "",
                    p.day.strftime("%d/%m/%Y") if p.day else "",
                    p.is_subsidized or 0,
                    p.weight or 0,
                    p.tare_weight or 0,
                    p.actual_weight or 0,
                    p.degree or 0,
                    p.dry_rubber or 0,
                    p.unit_price or 0,
                    p.subsidy_price or 0,
                    p.total_amount or 0,
                    p.paid_amount or 0,
                    p.saved_amount or 0,
                ]
                for col_idx, v in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=v)
                    cell.border = thin_border
                    if col_idx >= 5:
                        cell.number_format = '#,##0'

            # Totals row
            total_row = len(purchases) + 5
            ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = total_font
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
            ws.cell(row=total_row, column=1).fill = total_fill
            ws.cell(row=total_row, column=1).border = thin_border
            ws.cell(row=total_row, column=1).alignment = center

            # Sum columns E(5) to N(14)
            sum_cols = {
                5: sum(p.weight or 0 for p in purchases),
                6: sum(p.tare_weight or 0 for p in purchases),
                7: sum(p.actual_weight or 0 for p in purchases),
                9: sum(p.dry_rubber or 0 for p in purchases),
                12: sum(p.total_amount or 0 for p in purchases),
                13: sum(p.paid_amount or 0 for p in purchases),
                14: sum(p.saved_amount or 0 for p in purchases),
            }
            for col_idx in range(4, 15):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.fill = total_fill
                cell.border = thin_border
                cell.font = total_font
                if col_idx in sum_cols:
                    cell.value = sum_cols[col_idx]
                    cell.number_format = '#,##0'

            # Column widths
            col_widths = [6, 8, 12, 10, 12, 12, 14, 10, 10, 12, 12, 15, 12, 12]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        if not has_data:
            await message.reply_text(
                f"📋 Không có dữ liệu thu hoạch từ <b>{start_date.strftime('%d/%m/%Y')}</b> đến <b>{end_date.strftime('%d/%m/%Y')}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        await message.reply_document(
            document=tmp_path,
            file_name=f"thu_hoach_{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}.xlsx",
            caption=(
                f"<b>BÁO CÁO THU HOẠCH</b>\n"
                f"Từ <b>{start_date.strftime('%d/%m/%Y')}</b> đến <b>{end_date.strftime('%d/%m/%Y')}</b>\n"
                f"Tổng hộ dân có dữ liệu: <b>{len(wb.sheetnames)}</b>"
            ),
            parse_mode=ParseMode.HTML
        )
        os.remove(tmp_path)
    except Exception as e:
        LogError(f"Error in check_harvest: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_compare_harvest", "tien_nga_so_sanh_thu_hoach"]) | filters.regex(r"^@\w+\s+/(tien_nga_compare_harvest|tien_nga_so_sanh_thu_hoach)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_compare_harvest_handler(client, message: Message) -> None:
    from datetime import datetime
    from app.models.business import Households, DailyPurchases, AgriculturalLand
    from sqlalchemy import func

    args = message.text.strip().split(maxsplit=1)
    if len(args) < 2:
        today = datetime.now()
        first_day = today.replace(day=1)
        await message.reply_text(
            "<b>SO SÁNH THU HOẠCH</b>\n\n"
            "Cú pháp: <code>/tien_nga_so_sanh_thu_hoach dd/mm/yyyy - dd/mm/yyyy</code>\n\n"
            f"VD: <code>/tien_nga_so_sanh_thu_hoach {first_day.strftime('%d/%m/%Y')} - {today.strftime('%d/%m/%Y')}</code>",
            parse_mode=ParseMode.HTML
        )
        return

    date_str = args[1].strip()
    try:
        parts = date_str.split("-")
        if len(parts) != 2:
            raise ValueError("Thiếu dấu -")
        start_date = datetime.strptime(parts[0].strip(), "%d/%m/%Y").date()
        end_date = datetime.strptime(parts[1].strip(), "%d/%m/%Y").date()
    except Exception:
        await message.reply_text("⚠️ Định dạng ngày không hợp lệ.\nVD: <code>/tien_nga_so_sanh_thu_hoach 01/04/2026 - 30/04/2026</code>", parse_mode=ParseMode.HTML)
        return

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    db = SessionLocal()
    try:
        from datetime import timedelta
        # Build date list
        delta = end_date - start_date
        date_list = [start_date + timedelta(days=i) for i in range(delta.days + 1)]

        households = db.query(Households).filter(Households.status == "ACTIVE").order_by(Households.household_code).all()
        if not households:
            await message.reply_text("📋 Chưa có hộ dân nào trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        # Cache land data
        lands = db.query(AgriculturalLand).filter(AgriculturalLand.status == "ACTIVE").all()
        land_map = {l.land_code: l for l in lands}

        rows = []
        for hh in households:
            # Lấy đất
            land = land_map.get(hh.land_code) if hh.land_code else None
            rubber_area = land.rubber_area if land else 0

            # Lấy purchases trong khoảng thời gian
            purchases = db.query(DailyPurchases).filter(
                DailyPurchases.hoursehold_id == hh.purchase_code,
                DailyPurchases.day >= start_date,
                DailyPurchases.day <= end_date
            ).all()

            if not purchases:
                continue

            # Lấy data từng ngày
            purchases = sorted(purchases, key=lambda x: x.day if x.day else datetime.min.date())
            
            total_actual = sum(p.actual_weight or 0 for p in purchases)
            total_dry = sum(p.dry_rubber or 0 for p in purchases)
            total_amount = sum(p.total_amount or 0 for p in purchases)
            total_paid = sum(p.paid_amount or 0 for p in purchases)
            total_saved = sum(p.saved_amount or 0 for p in purchases)

            rows.append({
                "hh": hh,
                "land": land,
                "rubber_area": rubber_area,
                "purchases": purchases,
                "total_actual_weight": total_actual,
                "total_dry_rubber": total_dry,
                "total_amount": total_amount,
                "total_paid": total_paid,
                "total_saved": total_saved,
                "avg_per_ha": total_actual / rubber_area if rubber_area > 0 else 0,
            })

        if not rows:
            await message.reply_text(
                f"📋 Không có dữ liệu thu hoạch từ <b>{start_date.strftime('%d/%m/%Y')}</b> đến <b>{end_date.strftime('%d/%m/%Y')}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        import openpyxl, tempfile, os
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Tổng Hợp"

        # --- SUMMARY TAB ---
        sum_headers = ["STT", "Mã HD", "Họ Tên", "Mã Đất", "DT Cao Su (ha)", "KL Mủ TT (kg)", "Mủ Khô (kg)", "Thành Tiền", "Đã TT", "Lưu Sổ", "TB KL/ha"]
        
        ws_summary.merge_cells("A1:K1")
        title_cell = ws_summary.cell(row=1, column=1, value=f"TỔNG HỢP SO SÁNH THU HOẠCH - Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal="center", vertical="center")

        for col_idx, hd in enumerate(sum_headers, 1):
            cell = ws_summary.cell(row=3, column=col_idx, value=hd)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = thin_border
            cell.alignment = center

        for row_idx, r in enumerate(rows, 1):
            hh = r["hh"]
            vals = [
                row_idx,
                hh.household_code,
                hh.fullname or "",
                hh.land_code or "—",
                r["rubber_area"],
                round(r["total_actual_weight"], 2),
                round(r["total_dry_rubber"], 2),
                round(r["total_amount"], 0),
                round(r["total_paid"], 0),
                round(r["total_saved"], 0),
                round(r["avg_per_ha"], 2)
            ]
            for col_idx, v in enumerate(vals, 1):
                cell = ws_summary.cell(row=row_idx+3, column=col_idx, value=v)
                cell.border = thin_border
                if col_idx >= 6:
                    cell.number_format = '#,##0.00' if col_idx in [6,7,11] else '#,##0'

        sum_row = len(rows) + 4
        ws_summary.cell(row=sum_row, column=1, value="TỔNG CỘNG").font = total_font
        ws_summary.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=4)
        
        g_rubber = sum(r["rubber_area"] for r in rows)
        g_act = sum(r["total_actual_weight"] for r in rows)
        g_dry = sum(r["total_dry_rubber"] for r in rows)
        g_amt = sum(r["total_amount"] for r in rows)
        g_paid = sum(r["total_paid"] for r in rows)
        g_sav = sum(r["total_saved"] for r in rows)
        g_avg = g_act / g_rubber if g_rubber > 0 else 0

        g_vals = [g_rubber, g_act, g_dry, g_amt, g_paid, g_sav, g_avg]
        
        for i in range(1, 12):
            c = ws_summary.cell(row=sum_row, column=i)
            c.border = thin_border
            c.fill = total_fill
            c.font = total_font
            
        for i, v in enumerate(g_vals):
            c = ws_summary.cell(row=sum_row, column=i+5, value=round(v, 2))
            c.number_format = '#,##0.00' if i in [0,1,2,6] else '#,##0'

        col_w = [6, 10, 20, 10, 14, 15, 15, 14, 14, 14, 14]
        for i, w in enumerate(col_w, 1):
            ws_summary.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # --- INDIVIDUAL TABS ---
        hh_headers = ["STT", "Ngày", "KL Mủ TT (kg)", "Mủ Khô (kg)", "Thành Tiền", "Đã TT", "Lưu Sổ"]
        for r in rows:
            hh = r["hh"]
            purchases = r["purchases"]
            
            sheet_name = f"{hh.household_code} - {hh.fullname or ''}"[:31]
            ws_hh = wb.create_sheet(title=sheet_name)
            
            ws_hh.merge_cells("A1:G1")
            title_cell = ws_hh.cell(row=1, column=1, value=f"CHI TIẾT: {hh.household_code} - {hh.fullname or ''}")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")
            
            ws_hh.merge_cells("A2:G2")
            info_cell = ws_hh.cell(row=2, column=1, value=f"Mã Đất: {hh.land_code or '—'} | DT Cao Su: {r['rubber_area']} ha | TB KL/ha: {round(r['avg_per_ha'], 2)} kg/ha")
            info_cell.alignment = Alignment(horizontal="center")

            for col_idx, hd in enumerate(hh_headers, 1):
                cell = ws_hh.cell(row=4, column=col_idx, value=hd)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = thin_border
                cell.alignment = center
                
            c_row = 5
            for i, p in enumerate(purchases, 1):
                vals = [
                    i,
                    p.day.strftime("%d/%m/%Y") if p.day else "",
                    round(p.actual_weight or 0, 2),
                    round(p.dry_rubber or 0, 2),
                    round(p.total_amount or 0, 0),
                    round(p.paid_amount or 0, 0),
                    round(p.saved_amount or 0, 0)
                ]
                for col_idx, v in enumerate(vals, 1):
                    cell = ws_hh.cell(row=c_row, column=col_idx, value=v)
                    cell.border = thin_border
                    if col_idx >= 3:
                        cell.number_format = '#,##0.00' if col_idx in [3,4] else '#,##0'
                c_row += 1

            # Total row for household
            ws_hh.cell(row=c_row, column=1, value="TỔNG CỘNG").font = total_font
            ws_hh.merge_cells(start_row=c_row, start_column=1, end_row=c_row, end_column=2)
            
            t_vals = [
                round(r["total_actual_weight"], 2),
                round(r["total_dry_rubber"], 2),
                round(r["total_amount"], 0),
                round(r["total_paid"], 0),
                round(r["total_saved"], 0)
            ]
            for col_idx in range(1, 8):
                cell = ws_hh.cell(row=c_row, column=col_idx)
                cell.fill = total_fill
                cell.border = thin_border
                cell.font = total_font
                
            for i, v in enumerate(t_vals):
                cell = ws_hh.cell(row=c_row, column=i+3, value=v)
                cell.number_format = '#,##0.00' if i in [0,1] else '#,##0'
                
            col_w2 = [6, 12, 15, 15, 15, 15, 15]
            for i, w in enumerate(col_w2, 1):
                ws_hh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        await message.reply_document(
            document=tmp_path,
            file_name=f"so_sanh_thu_hoach_{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}.xlsx",
            caption=(
                f"<b>SO SÁNH THU HOẠCH</b>\n"
                f"Từ <b>{start_date.strftime('%d/%m/%Y')}</b> đến <b>{end_date.strftime('%d/%m/%Y')}</b>\n"
                f"Tổng hộ dân: <b>{len(rows)}</b> | Tổng DT Cao Su: <b>{round(g_rubber, 2)} ha</b>"
            ),
            parse_mode=ParseMode.HTML
        )
        os.remove(tmp_path)
    except Exception as e:
        LogError(f"Error in compare_harvest: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_daily_harvest", "tien_nga_thu_hoach_hang_ngay"]) | filters.regex(r"^@\w+\s+/(tien_nga_daily_harvest|tien_nga_thu_hoach_hang_ngay)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN, UserType.MEMBER)
@require_project_name("Tiến Nga")
@require_group_role("member")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST, CustomTitle.MEMBER_HARVEST)
async def tien_nga_daily_harvest_handler(client, message: Message) -> None:
    from datetime import datetime
    from app.models.business import Households, AgriculturalLand, DailyHarvest
    import uuid as uuid_lib

    lines = message.text.strip().split("\n")
    args = lines[0].strip().split()

    if len(lines) < 2:
        # Hiện form — cần mã hộ dân
        if len(args) < 2:
            await message.reply_text(
                "<b>THU HOẠCH HÀNG NGÀY</b>\n\n"
                "Cú pháp: <code>/tien_nga_thu_hoach_hang_ngay [Mã Hộ Dân]</code>\n\n"
                "VD: <code>/tien_nga_thu_hoach_hang_ngay HD001</code>",
                parse_mode=ParseMode.HTML
            )
            return

        hh_code = args[1].upper()
        db = SessionLocal()
        try:
            hh = db.query(Households).filter(Households.household_code == hh_code, Households.status == "ACTIVE").first()
            if not hh:
                await message.reply_text(f"⚠️ Không tìm thấy hộ dân với mã <b>{hh_code}</b>.", parse_mode=ParseMode.HTML)
                return

            today = datetime.now().strftime("%d/%m/%Y")
            form = f"""<b>FORM THU HOẠCH HÀNG NGÀY</b>

Hộ dân: <b>{hh.fullname or hh_code}</b> (<code>{hh_code}</code>)
Mã đất: <code>{hh.land_code or '—'}</code>

Sao chép form dưới, điền thông tin và gửi lại:

<pre>/tien_nga_thu_hoach_hang_ngay
Mã Hộ Dân: {hh_code}
Mã Đất: {hh.land_code or ''}
Ngày: {today}
Số Lượng Cây: 0</pre>

<i>Ghi chú: Đơn giá được tự động lấy từ thông tin hộ dân.</i>"""
            await message.reply_text(form, parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    # Xử lý form submit
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    hh_code = data.get("Mã Hộ Dân", "").strip().upper()
    land_code = data.get("Mã Đất", "").strip().upper()
    day_str = data.get("Ngày", "").strip()
    tree_count_str = data.get("Số Lượng Cây", "0").strip()

    if not hh_code:
        await message.reply_text("⚠️ <b>Mã Hộ Dân</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    try:
        day = datetime.strptime(day_str, "%d/%m/%Y").date()
    except Exception:
        await message.reply_text("⚠️ Định dạng ngày không hợp lệ. VD: <code>27/04/2026</code>", parse_mode=ParseMode.HTML)
        return

    tree_count = int(parse_float_vn(tree_count_str))
    if tree_count <= 0:
        await message.reply_text("⚠️ <b>Số Lượng Cây</b> phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        hh = db.query(Households).filter(Households.household_code == hh_code, Households.status == "ACTIVE").first()
        if not hh:
            await message.reply_text(f"⚠️ Không tìm thấy hộ dân với mã <b>{hh_code}</b>.", parse_mode=ParseMode.HTML)
            return

        unit_price = hh.tapping_price or 0
        total_amount = tree_count * unit_price

        # Lấy thông tin đất
        land = None
        use_land_code = land_code or hh.land_code or ""
        if use_land_code:
            land = db.query(AgriculturalLand).filter(AgriculturalLand.land_code == use_land_code, AgriculturalLand.status == "ACTIVE").first()

        def fmt_money(val):
            if val is None or val == 0: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        # Tìm nhóm main để gửi yêu cầu xác nhận
        from app.models.business import Projects
        from app.models.telegram import TelegramProjectMember

        project = db.query(Projects).filter(Projects.project_name == "Tiến Nga").first()
        if not project:
            await message.reply_text("⚠️ Không tìm thấy dự án Tiến Nga.", parse_mode=ParseMode.HTML)
            return

        main_chat = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project.id,
            TelegramProjectMember.role == "main"
        ).first()

        if not main_chat or not main_chat.chat_id:
            await message.reply_text("⚠️ Không tìm thấy nhóm chính để gửi yêu cầu.", parse_mode=ParseMode.HTML)
            return

        user_display = message.from_user.first_name or message.from_user.username or str(message.from_user.id)
        member_chat_id = message.chat.id

        # Gửi yêu cầu xác nhận lên nhóm main
        cb_data = f"harvest_confirm:{hh_code}:{use_land_code}:{day.strftime('%d%m%Y')}:{tree_count}:{int(unit_price)}:{member_chat_id}"
        cb_deny = f"harvest_deny:{hh_code}:{member_chat_id}"

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Xác Nhận", callback_data=cb_data)],
            [InlineKeyboardButton("Từ Chối", callback_data=cb_deny)]
        ])

        land_display = f"<code>{use_land_code or '—'}</code>"
        if land and land.land_name:
            land_display += f" ({land.land_name})"

        await client.send_message(
            chat_id=int(main_chat.chat_id),
            text=(
                f"🔔 <b>YÊU CẦU XÁC NHẬN THU HOẠCH</b>\n\n"
                f"<b>Người gửi:</b> {user_display}\n"
                f"<b>Hộ Dân:</b> {hh.fullname or hh_code} (<code>{hh_code}</code>)\n"
                f"<b>Mã Đất:</b> {land_display}\n"
                f"<b>Ngày:</b> {day.strftime('%d/%m/%Y')}\n"
                f"<b>Số Lượng Cây:</b> {tree_count:,}\n"
                f"<b>Đơn Giá:</b> <code>{fmt_money(unit_price)}</code>\n"
                f"<b>Thành Tiền:</b> <code>{fmt_money(total_amount)}</code>\n\n"
                f"<i>Vui lòng xác nhận hoặc từ chối yêu cầu này.</i>"
            ),
            reply_markup=keyboard,
            parse_mode=ParseMode.HTML
        )

        await message.reply_text(
            f"<b>Đã gửi yêu cầu thu hoạch lên nhóm chính!</b>\n\n"
            f"<b>Hộ Dân:</b> {hh.fullname or hh_code}\n"
            f"<b>Số Lượng Cây:</b> {tree_count:,}\n"
            f"<b>Thành Tiền:</b> <code>{fmt_money(total_amount)}</code>\n\n"
            f"<i>Vui lòng chờ Admin/Owner xác nhận.</i>",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        db.rollback()
        LogError(f"Error creating daily harvest request: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

# Callback: Xác nhận thu hoạch
@bot.on_callback_query(filters.regex(r"^harvest_confirm:"))
async def harvest_confirm_callback(client, callback_query):
    from datetime import datetime
    from app.models.business import Households, AgriculturalLand, DailyHarvest
    import uuid as uuid_lib

    parts = callback_query.data.split(":")
    # harvest_confirm:hh_code:land_code:day:tree_count:unit_price:member_chat_id
    if len(parts) < 7:
        await callback_query.answer("⚠️ Dữ liệu không hợp lệ.", show_alert=True)
        return

    hh_code = parts[1]
    land_code_val = parts[2]
    day_str = parts[3]
    tree_count = int(parts[4])
    unit_price = int(parts[5])
    member_chat_id = parts[6]

    total_amount = tree_count * unit_price

    try:
        day = datetime.strptime(day_str, "%d%m%Y").date()
    except Exception:
        await callback_query.answer("⚠️ Ngày không hợp lệ.", show_alert=True)
        return

    db = SessionLocal()
    try:
        hh = db.query(Households).filter(Households.household_code == hh_code, Households.status == "ACTIVE").first()
        if not hh:
            await callback_query.answer(f"⚠️ Không tìm thấy hộ dân {hh_code}.", show_alert=True)
            return

        land = None
        if land_code_val:
            land = db.query(AgriculturalLand).filter(AgriculturalLand.land_code == land_code_val, AgriculturalLand.status == "ACTIVE").first()

        # Lưu vào DB
        new_harvest = DailyHarvest(
            id=uuid_lib.uuid4(),
            day=day,
            household_code=hh_code,
            land_code=land_code_val or None,
            tree_count=tree_count,
            unit_price=unit_price,
            total_amount=total_amount
        )
        db.add(new_harvest)

        # Cộng thành tiền vào công nợ
        hh.total_debt = (hh.total_debt or 0) + total_amount
        db.commit()

        def fmt_money(val):
            if val is None or val == 0: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        approver = callback_query.from_user.first_name or callback_query.from_user.username or str(callback_query.from_user.id)

        land_display = f"<code>{land_code_val or '—'}</code>"
        if land and land.land_name:
            land_display += f" ({land.land_name})"

        # Cập nhật message trên nhóm main
        await callback_query.message.edit_text(
            f"✅ <b>ĐÃ XÁC NHẬN THU HOẠCH</b>\n\n"
            f"<b>Hộ Dân:</b> {hh.fullname or hh_code} (<code>{hh_code}</code>)\n"
            f"<b>Mã Đất:</b> {land_display}\n"
            f"<b>Ngày:</b> {day.strftime('%d/%m/%Y')}\n"
            f"<b>Số Lượng Cây:</b> {tree_count:,}\n"
            f"<b>Đơn Giá:</b> <code>{fmt_money(unit_price)}</code>\n"
            f"<b>Thành Tiền:</b> <code>{fmt_money(total_amount)}</code>\n"
            f"<b>Công Nợ Mới:</b> <code>{fmt_money(hh.total_debt)}</code>\n\n"
            f"<b>Xác nhận bởi:</b> {approver}",
            parse_mode=ParseMode.HTML
        )

        # Gửi thông báo về nhóm member
        try:
            await client.send_message(
                chat_id=int(member_chat_id),
                text=(
                    f"✅ <b>Thu Hoạch Đã Được Xác Nhận!</b>\n\n"
                    f"<b>Hộ Dân:</b> {hh.fullname or hh_code} (<code>{hh_code}</code>)\n"
                    f"<b>Ngày:</b> {day.strftime('%d/%m/%Y')}\n"
                    f"<b>Số Lượng Cây:</b> {tree_count:,}\n"
                    f"<b>Đơn Giá:</b> <code>{fmt_money(unit_price)}</code>\n"
                    f"<b>Thành Tiền:</b> <code>{fmt_money(total_amount)}</code>\n"
                    f"<b>Công Nợ Hiện Tại:</b> <code>{fmt_money(hh.total_debt)}</code>\n\n"
                    f"<b>Xác nhận bởi:</b> {approver}"
                ),
                parse_mode=ParseMode.HTML
            )
        except Exception:
            pass

        await callback_query.answer("✅ Đã xác nhận thu hoạch.", show_alert=False)
    except Exception as e:
        db.rollback()
        LogError(f"Error confirming harvest: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Lỗi hệ thống.", show_alert=True)
    finally:
        db.close()

# Callback: Từ chối thu hoạch
@bot.on_callback_query(filters.regex(r"^harvest_deny:"))
async def harvest_deny_callback(client, callback_query):
    parts = callback_query.data.split(":")
    # harvest_deny:hh_code:member_chat_id
    if len(parts) < 3:
        await callback_query.answer("⚠️ Dữ liệu không hợp lệ.", show_alert=True)
        return

    hh_code = parts[1]
    member_chat_id = parts[2]

    denier = callback_query.from_user.first_name or callback_query.from_user.username or str(callback_query.from_user.id)

    await callback_query.message.edit_text(
        f"❌ <b>ĐÃ TỪ CHỐI THU HOẠCH</b>\n\n"
        f"<b>Hộ Dân:</b> <code>{hh_code}</code>\n"
        f"<b>Từ chối bởi:</b> {denier}",
        parse_mode=ParseMode.HTML
    )

    # Gửi thông báo về nhóm member
    try:
        await client.send_message(
            chat_id=int(member_chat_id),
            text=(
                f"❌ <b>Yêu Cầu Thu Hoạch Bị Từ Chối</b>\n\n"
                f"<b>Hộ Dân:</b> <code>{hh_code}</code>\n"
                f"<b>Từ chối bởi:</b> {denier}\n\n"
                f"<i>Vui lòng kiểm tra lại và gửi yêu cầu mới nếu cần.</i>"
            ),
            parse_mode=ParseMode.HTML
        )
    except Exception:
        pass

    await callback_query.answer("❌ Đã từ chối.", show_alert=False)

@bot.on_message(filters.command(["tien_nga_check_household", "tien_nga_kiem_tra_ho_dan"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_household|tien_nga_kiem_tra_ho_dan)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN, UserType.MEMBER)
@require_project_name("Tiến Nga")
@require_group_role("member")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST, CustomTitle.MEMBER_HARVEST)
async def tien_nga_check_household_handler(client, message: Message) -> None:
    from app.models.business import Households, AgriculturalLand

    caller_username = message.from_user.username
    if not caller_username:
        await message.reply_text("⚠️ Bạn chưa cài đặt username trên Telegram.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        hh = db.query(Households).filter(
            Households.username == caller_username,
            Households.status == "ACTIVE"
        ).first()

        if not hh:
            await message.reply_text(
                f"⚠️ Không tìm thấy hộ dân nào liên kết với username <b>@{caller_username}</b>.\n\n"
                f"<i>Vui lòng liên hệ Admin để được cập nhật username vào hồ sơ hộ dân.</i>",
                parse_mode=ParseMode.HTML
            )
            return

        # Lấy thông tin đất
        land = None
        if hh.land_code:
            land = db.query(AgriculturalLand).filter(
                AgriculturalLand.land_code == hh.land_code,
                AgriculturalLand.status == "ACTIVE"
            ).first()

        def fmt_money(val):
            if val is None or val == 0: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        land_info = ""
        if land:
            land_info = (
                f"\n\n<b>THÔNG TIN ĐẤT</b>\n"
                f"<b>Mã Đất:</b> <code>{land.land_code}</code>\n"
                f"<b>Tên Đất:</b> {land.land_name or '—'}\n"
                f"<b>Địa Chỉ:</b> {land.address or '—'}\n"
                f"<b>DT Tổng:</b> {land.total_area} ha\n"
                f"<b>DT Cao Su:</b> {land.rubber_area} ha\n"
                f"<b>Cây Thu Hoạch:</b> {land.harvesting_trees or 0}\n"
                f"<b>Cây Đang Trồng:</b> {land.planting_trees or 0}"
            )

        await message.reply_text(
            f"<b>THÔNG TIN HỘ DÂN</b>\n\n"
            f"<b>Mã Hộ Dân:</b> <code>{hh.household_code}</code>\n"
            f"<b>Mã Thu Mua:</b> <code>{hh.purchase_code or '—'}</code>\n"
            f"<b>Họ Tên:</b> {hh.fullname or '—'}\n"
            f"<b>Username:</b> @{hh.username or '—'}\n"
            f"<b>Nhóm Telegram:</b> {hh.telegram_group or '—'}\n"
            f"<b>SĐT:</b> {hh.phone or '—'}\n"
            f"<b>Địa Chỉ:</b> {hh.address or '—'}\n"
            f"<b>Mã Đất:</b> <code>{hh.land_code or '—'}</code>\n"
            f"<b>Đơn Giá Cạo Mủ:</b> <code>{fmt_money(hh.tapping_price)}</code>\n"
            f"<b>Công Nợ:</b> <code>{fmt_money(hh.total_debt)}</code>\n"
            f"<b>Số TK:</b> <code>{hh.bank_account or '—'}</code>\n"
            f"<b>Ngân Hàng:</b> {hh.bank_name or '—'}"
            + land_info,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error checking household: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_check_daily_harvest", "tien_nga_kt_thu_hoach_hang_ngay"]) | filters.regex(r"^@\w+\s+/(tien_nga_check_daily_harvest|tien_nga_kt_thu_hoach_hang_ngay)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_check_daily_harvest_handler(client, message: Message) -> None:
    from datetime import datetime
    from app.models.business import Households, AgriculturalLand, DailyHarvest

    text = message.text.strip()
    # Tách command ra
    parts = text.split(None, 1)
    if len(parts) < 2:
        await message.reply_text(
            "<b>KIỂM TRA THU HOẠCH HÀNG NGÀY</b>\n\n"
            "Cú pháp:\n"
            "• <code>/tien_nga_kt_thu_hoach_hang_ngay [Mã Hộ] dd/mm/yyyy - dd/mm/yyyy</code>\n"
            "• <code>/tien_nga_kt_thu_hoach_hang_ngay dd/mm/yyyy - dd/mm/yyyy</code>\n\n"
            "<i>Nếu không có Mã Hộ, hệ thống sẽ tổng hợp tất cả hộ dân.</i>",
            parse_mode=ParseMode.HTML
        )
        return

    args_text = parts[1].strip()

    # Parse: có thể là [mã hộ] dd/mm/yyyy - dd/mm/yyyy hoặc dd/mm/yyyy - dd/mm/yyyy
    import re
    date_pattern = r"(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})"
    date_match = re.search(date_pattern, args_text)

    if not date_match:
        await message.reply_text("⚠️ Định dạng ngày không hợp lệ. VD: <code>01/04/2026 - 27/04/2026</code>", parse_mode=ParseMode.HTML)
        return

    try:
        start_date = datetime.strptime(date_match.group(1), "%d/%m/%Y").date()
        end_date = datetime.strptime(date_match.group(2), "%d/%m/%Y").date()
    except Exception:
        await message.reply_text("⚠️ Định dạng ngày không hợp lệ.", parse_mode=ParseMode.HTML)
        return

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    # Lấy mã hộ (nếu có) — phần trước date_match
    hh_code_filter = args_text[:date_match.start()].strip().upper() or None

    db = SessionLocal()
    try:
        # Lấy danh sách hộ dân
        if hh_code_filter:
            households = db.query(Households).filter(
                Households.household_code == hh_code_filter,
                Households.status == "ACTIVE"
            ).all()
            if not households:
                await message.reply_text(f"⚠️ Không tìm thấy hộ dân <b>{hh_code_filter}</b>.", parse_mode=ParseMode.HTML)
                return
        else:
            households = db.query(Households).filter(Households.status == "ACTIVE").order_by(Households.household_code).all()
            if not households:
                await message.reply_text("📋 Chưa có hộ dân nào.", parse_mode=ParseMode.HTML)
                return

        # Lấy data thu hoạch
        hh_codes = [h.household_code for h in households]
        harvests = db.query(DailyHarvest).filter(
            DailyHarvest.household_code.in_(hh_codes),
            DailyHarvest.day >= start_date,
            DailyHarvest.day <= end_date
        ).order_by(DailyHarvest.day).all()

        if not harvests:
            await message.reply_text(
                f"📋 Không có dữ liệu thu hoạch từ <b>{start_date.strftime('%d/%m/%Y')}</b> đến <b>{end_date.strftime('%d/%m/%Y')}</b>.",
                parse_mode=ParseMode.HTML
            )
            return

        # Nhóm theo household_code
        harvest_map = {}
        for hv in harvests:
            harvest_map.setdefault(hv.household_code, []).append(hv)

        import openpyxl, tempfile, os
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal="center", vertical="center")

        headers = ["STT", "Ngày", "Mã Đất", "Số Lượng Cây", "Đơn Giá", "Thành Tiền"]
        col_widths = [6, 14, 12, 16, 14, 16]

        hh_map = {h.household_code: h for h in households}
        grand_total = 0
        total_hh_count = 0

        for hh_code, hv_list in harvest_map.items():
            hh = hh_map.get(hh_code)
            if not hh:
                continue
            total_hh_count += 1

            sheet_name = f"{hh_code} - {hh.fullname or ''}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Title
            ws.merge_cells("A1:F1")
            title_cell = ws.cell(row=1, column=1,
                value=f"THU HOẠCH: {hh_code} - {hh.fullname or ''}")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:F2")
            info_cell = ws.cell(row=2, column=1,
                value=f"Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')} | Mã Đất: {hh.land_code or '—'} | ĐG Cạo Mủ: {int(hh.tapping_price or 0):,}")
            info_cell.alignment = Alignment(horizontal="center")

            for col_idx, hd in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=hd)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = thin_border
                cell.alignment = center

            c_row = 5
            sheet_total = 0
            for i, hv in enumerate(hv_list, 1):
                vals = [
                    i,
                    hv.day.strftime("%d/%m/%Y") if hv.day else "",
                    hv.land_code or "—",
                    hv.tree_count or 0,
                    round(hv.unit_price or 0, 0),
                    round(hv.total_amount or 0, 0)
                ]
                for col_idx, v in enumerate(vals, 1):
                    cell = ws.cell(row=c_row, column=col_idx, value=v)
                    cell.border = thin_border
                    if col_idx >= 4:
                        cell.number_format = '#,##0'
                sheet_total += (hv.total_amount or 0)
                c_row += 1

            # Dòng tổng
            ws.cell(row=c_row, column=1, value="TỔNG CỘNG").font = total_font
            ws.merge_cells(start_row=c_row, start_column=1, end_row=c_row, end_column=3)
            
            sum_trees = sum(hv.tree_count or 0 for hv in hv_list)
            
            for col_idx in range(1, 7):
                cell = ws.cell(row=c_row, column=col_idx)
                cell.fill = total_fill
                cell.border = thin_border
                cell.font = total_font

            ws.cell(row=c_row, column=4, value=sum_trees).number_format = '#,##0'
            ws.cell(row=c_row, column=6, value=round(sheet_total, 0)).number_format = '#,##0'

            grand_total += sheet_total

            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Nếu nhiều hộ → thêm tab Tổng Hợp ở đầu
        if total_hh_count > 1:
            ws_sum = wb.create_sheet(title="Tổng Hợp", index=0)
            ws_sum.merge_cells("A1:E1")
            title_cell = ws_sum.cell(row=1, column=1,
                value=f"TỔNG HỢP THU HOẠCH - Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")

            sum_headers = ["STT", "Mã HD", "Họ Tên", "Tổng Cây", "Tổng Thành Tiền"]
            for col_idx, hd in enumerate(sum_headers, 1):
                cell = ws_sum.cell(row=3, column=col_idx, value=hd)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = thin_border
                cell.alignment = center

            s_row = 4
            g_trees = 0
            g_amount = 0
            for idx, (hh_code, hv_list) in enumerate(harvest_map.items(), 1):
                hh = hh_map.get(hh_code)
                if not hh:
                    continue
                s_trees = sum(hv.tree_count or 0 for hv in hv_list)
                s_amount = sum(hv.total_amount or 0 for hv in hv_list)
                g_trees += s_trees
                g_amount += s_amount

                vals = [idx, hh_code, hh.fullname or "", s_trees, round(s_amount, 0)]
                for col_idx, v in enumerate(vals, 1):
                    cell = ws_sum.cell(row=s_row, column=col_idx, value=v)
                    cell.border = thin_border
                    if col_idx >= 4:
                        cell.number_format = '#,##0'
                s_row += 1

            # Tổng cuối
            ws_sum.cell(row=s_row, column=1, value="TỔNG CỘNG").font = total_font
            ws_sum.merge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=3)
            for col_idx in range(1, 6):
                cell = ws_sum.cell(row=s_row, column=col_idx)
                cell.fill = total_fill
                cell.border = thin_border
                cell.font = total_font
            ws_sum.cell(row=s_row, column=4, value=g_trees).number_format = '#,##0'
            ws_sum.cell(row=s_row, column=5, value=round(g_amount, 0)).number_format = '#,##0'

            sum_widths = [6, 12, 22, 14, 18]
            for i, w in enumerate(sum_widths, 1):
                ws_sum.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        def fmt_money(val):
            if val is None or val == 0: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        caption = (
            f"<b>KIỂM TRA THU HOẠCH HÀNG NGÀY</b>\n"
            f"Từ <b>{start_date.strftime('%d/%m/%Y')}</b> đến <b>{end_date.strftime('%d/%m/%Y')}</b>\n"
            f"Số hộ dân: <b>{total_hh_count}</b>\n"
            f"Tổng thành tiền: <b>{fmt_money(grand_total)}</b>"
        )

        await message.reply_document(
            document=tmp_path,
            file_name=f"kt_thu_hoach_{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}.xlsx",
            caption=caption,
            parse_mode=ParseMode.HTML
        )
        os.remove(tmp_path)
    except Exception as e:
        LogError(f"Error in check daily harvest: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_create_household", "tien_nga_tao_ho_dan"]) | filters.regex(r"^@\w+\s+/(tien_nga_create_household|tien_nga_tao_ho_dan)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_create_household_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    if len(lines) < 2:
        form = """<b>FORM TẠO HỘ DÂN</b>

Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/tien_nga_tao_ho_dan
Mã Hộ Dân: 
Mã Hộ Thu Mua: 
Mã Đất: 
Họ Và Tên: 
Username: 
Nhóm Telegram: 
SĐT: 
Địa Chỉ: 
Công Nợ: 0
Đơn Giá Cạo Mủ: 0
Số TK: 
Ngân Hàng: </pre>

<i>Ghi chú: VD Mã Hộ Dân: HD001, Mã Thu Mua: TM001, Mã Đất: DCS001.</i>"""
        await message.reply_text(form, parse_mode=ParseMode.HTML)
        return

    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    household_code = data.get("Mã Hộ Dân", "").strip().upper()
    purchase_code = data.get("Mã Hộ Thu Mua", "").strip().upper()
    land_code = data.get("Mã Đất", "").strip().upper()
    fullname = data.get("Họ Và Tên", "").strip()
    username = data.get("Username", "").strip().lstrip("@")
    telegram_group = data.get("Nhóm Telegram", "").strip()
    phone = data.get("SĐT", "").strip()
    address = data.get("Địa Chỉ", "").strip()
    total_debt = parse_float_vn(data.get("Công Nợ", "0"))
    tapping_price = parse_float_vn(data.get("Đơn Giá Cạo Mủ", "0"))
    bank_account = data.get("Số TK", "").strip()
    bank_name = data.get("Ngân Hàng", "").strip()

    if not household_code:
        await message.reply_text("⚠️ <b>Mã Hộ Dân</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not purchase_code:
        await message.reply_text("⚠️ <b>Mã Hộ Thu Mua</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return
    if not fullname:
        await message.reply_text("⚠️ <b>Họ Và Tên</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    from app.models.business import Households
    import uuid as uuid_lib

    db = SessionLocal()
    try:
        if db.query(Households).filter(Households.household_code == household_code).first():
            await message.reply_text(f"⚠️ Mã hộ dân <b>{household_code}</b> đã tồn tại.", parse_mode=ParseMode.HTML)
            return
        if db.query(Households).filter(Households.purchase_code == purchase_code).first():
            await message.reply_text(f"⚠️ Mã hộ thu mua <b>{purchase_code}</b> đã tồn tại.", parse_mode=ParseMode.HTML)
            return

        new_hh = Households(
            id=uuid_lib.uuid4(), household_code=household_code, purchase_code=purchase_code,
            land_code=land_code or None, fullname=fullname, username=username or None,
            telegram_group=telegram_group or None, phone=phone or None,
            address=address or None, total_debt=total_debt, tapping_price=tapping_price,
            bank_account=bank_account or None, bank_name=bank_name or None, status="ACTIVE"
        )
        db.add(new_hh)
        db.commit()

        def fmt_money(val):
            if val is None or val == 0: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        await message.reply_text(
            f"✅ <b>Tạo Hộ Dân Thành Công!</b>\n\n"
            f"<b>Mã Hộ Dân:</b> <code>{household_code}</code>\n"
            f"<b>Mã Thu Mua:</b> <code>{purchase_code}</code>\n"
            f"<b>Mã Đất:</b> <code>{land_code or '—'}</code>\n"
            f"<b>Họ Tên:</b> {fullname}\n"
            f"<b>Username:</b> @{username or '—'}\n"
            f"<b>Nhóm Telegram:</b> {telegram_group or '—'}\n"
            f"<b>SĐT:</b> {phone or '—'}\n"
            f"<b>Địa Chỉ:</b> {address or '—'}\n"
            f"<b>Công Nợ:</b> <code>{fmt_money(total_debt)}</code>\n"
            f"<b>Đơn Giá Cạo Mủ:</b> <code>{fmt_money(tapping_price)}</code>\n"
            f"<b>Số TK:</b> <code>{bank_account or '—'}</code>\n"
            f"<b>Ngân Hàng:</b> {bank_name or '—'}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        db.rollback()
        LogError(f"Error creating household: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_update_household", "tien_nga_cap_nhat_ho_dan"]) | filters.regex(r"^@\w+\s+/(tien_nga_update_household|tien_nga_cap_nhat_ho_dan)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_update_household_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")
    from app.models.business import Households

    if len(lines) < 2:
        args = lines[0].split()
        if len(args) < 2:
            await message.reply_text("⚠️ Cú pháp: <code>/tien_nga_cap_nhat_ho_dan [Mã Hộ Dân]</code>", parse_mode=ParseMode.HTML)
            return
        hh_code = args[1].upper()
        db = SessionLocal()
        try:
            hh = db.query(Households).filter(Households.household_code == hh_code, Households.status == "ACTIVE").first()
            if not hh:
                await message.reply_text(f"⚠️ Không tìm thấy hộ dân với mã <b>{hh_code}</b>.", parse_mode=ParseMode.HTML)
                return

            def fmt_money(val):
                if val is None or val == 0: return "0"
                return str(int(val))

            form = f"""<b>CẬP NHẬT HỘ DÂN</b>

Sao chép form dưới, chỉnh sửa và gửi lại:

<pre>/tien_nga_cap_nhat_ho_dan
Mã Hộ Dân: {hh.household_code}
Mã Hộ Thu Mua: {hh.purchase_code or ''}
Mã Đất: {hh.land_code or ''}
Họ Và Tên: {hh.fullname or ''}
Username: {hh.username or ''}
Nhóm Telegram: {hh.telegram_group or ''}
SĐT: {hh.phone or ''}
Địa Chỉ: {hh.address or ''}
Công Nợ: {fmt_money(hh.total_debt)}
Đơn Giá Cạo Mủ: {fmt_money(hh.tapping_price)}
Số TK: {hh.bank_account or ''}
Ngân Hàng: {hh.bank_name or ''}</pre>"""
            await message.reply_text(form, parse_mode=ParseMode.HTML)
        finally:
            db.close()
        return

    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    hh_code = data.get("Mã Hộ Dân", "").strip().upper()
    if not hh_code:
        await message.reply_text("⚠️ <b>Mã Hộ Dân</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        hh = db.query(Households).filter(Households.household_code == hh_code, Households.status == "ACTIVE").first()
        if not hh:
            await message.reply_text(f"⚠️ Không tìm thấy hộ dân với mã <b>{hh_code}</b>.", parse_mode=ParseMode.HTML)
            return

        if "Mã Hộ Thu Mua" in data and data["Mã Hộ Thu Mua"]:
            new_pc = data["Mã Hộ Thu Mua"].upper()
            dup = db.query(Households).filter(Households.purchase_code == new_pc, Households.household_code != hh_code).first()
            if dup:
                await message.reply_text(f"⚠️ Mã thu mua <b>{new_pc}</b> đã được sử dụng.", parse_mode=ParseMode.HTML)
                return
            hh.purchase_code = new_pc
        if "Mã Đất" in data: hh.land_code = data["Mã Đất"].upper() or None
        if "Họ Và Tên" in data: hh.fullname = data["Họ Và Tên"] or None
        if "Username" in data: hh.username = data["Username"].lstrip("@") or None
        if "Nhóm Telegram" in data: hh.telegram_group = data["Nhóm Telegram"] or None
        if "SĐT" in data: hh.phone = data["SĐT"] or None
        if "Địa Chỉ" in data: hh.address = data["Địa Chỉ"] or None
        if "Công Nợ" in data: hh.total_debt = parse_float_vn(data["Công Nợ"])
        if "Đơn Giá Cạo Mủ" in data: hh.tapping_price = parse_float_vn(data["Đơn Giá Cạo Mủ"])
        if "Số TK" in data: hh.bank_account = data["Số TK"] or None
        if "Ngân Hàng" in data: hh.bank_name = data["Ngân Hàng"] or None
        db.commit()

        def fmt_money(val):
            if val is None or val == 0: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        await message.reply_text(
            f"✅ <b>Cập Nhật Hộ Dân Thành Công!</b>\n\n"
            f"<b>Mã Hộ Dân:</b> <code>{hh.household_code}</code>\n"
            f"<b>Mã Thu Mua:</b> <code>{hh.purchase_code or '—'}</code>\n"
            f"<b>Mã Đất:</b> <code>{hh.land_code or '—'}</code>\n"
            f"<b>Họ Tên:</b> {hh.fullname or '—'}\n"
            f"<b>Username:</b> @{hh.username or '—'}\n"
            f"<b>Nhóm Telegram:</b> {hh.telegram_group or '—'}\n"
            f"<b>SĐT:</b> {hh.phone or '—'}\n"
            f"<b>Địa Chỉ:</b> {hh.address or '—'}\n"
            f"<b>Công Nợ:</b> <code>{fmt_money(hh.total_debt)}</code>\n"
            f"<b>Đơn Giá Cạo Mủ:</b> <code>{fmt_money(hh.tapping_price)}</code>\n"
            f"<b>Số TK:</b> <code>{hh.bank_account or '—'}</code>\n"
            f"<b>Ngân Hàng:</b> {hh.bank_name or '—'}",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        db.rollback()
        LogError(f"Error updating household: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi hệ thống.", parse_mode=ParseMode.HTML)
    finally:
        db.close()

@bot.on_message(filters.command(["tien_nga_delete_household", "tien_nga_xoa_ho_dan"]) | filters.regex(r"^@\w+\s+/(tien_nga_delete_household|tien_nga_xoa_ho_dan)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Tiến Nga")
@require_group_role("main")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HARVEST)
async def tien_nga_delete_household_handler(client, message: Message) -> None:
    args = message.text.split()
    if len(args) < 2:
        await message.reply_text("⚠️ Cú pháp: <code>/tien_nga_xoa_ho_dan [Mã Hộ Dân]</code>", parse_mode=ParseMode.HTML)
        return

    hh_code = args[1].upper()
    from app.models.business import Households

    db = SessionLocal()
    try:
        hh = db.query(Households).filter(Households.household_code == hh_code, Households.status == "ACTIVE").first()
        if not hh:
            await message.reply_text(f"⚠️ Không tìm thấy hộ dân với mã <b>{hh_code}</b>.", parse_mode=ParseMode.HTML)
            return

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Xóa", callback_data=f"del_hh_confirm:{hh_code}"),
             InlineKeyboardButton("Hủy", callback_data="del_hh_cancel")]
        ])
        await message.reply_text(
            f"⚠️ <b>Xác nhận xóa hộ dân?</b>\n\n"
            f"<b>Mã Hộ Dân:</b> <code>{hh_code}</code>\n"
            f"<b>Họ Tên:</b> {hh.fullname or '—'}\n"
            f"<b>Mã Thu Mua:</b> <code>{hh.purchase_code or '—'}</code>",
            reply_markup=keyboard, parse_mode=ParseMode.HTML
        )
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^del_hh_(confirm|cancel)"))
async def del_hh_callback(client, callback_query) -> None:
    data = callback_query.data
    if data == "del_hh_cancel":
        await callback_query.message.delete()
        return

    hh_code = data.split(":")[1]
    from app.models.business import Households

    db = SessionLocal()
    try:
        hh = db.query(Households).filter(Households.household_code == hh_code).first()
        if hh:
            hh.status = "DELETED"
            db.commit()
            await callback_query.message.edit_text(f"✅ Đã xóa hộ dân <b>{hh_code}</b> thành công.", parse_mode=ParseMode.HTML)
        else:
            await callback_query.answer("⚠️ Không tìm thấy.", show_alert=True)
    except Exception as e:
        db.rollback()
        LogError(f"Error deleting household: {e}", LogType.SYSTEM_STATUS)
        await callback_query.message.edit_text("❌ Lỗi hệ thống.")
    finally:
        db.close()

from pyrogram import filters
from pyrogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from pyrogram.enums import ParseMode, ChatMemberStatus
from bot.utils.bot import bot
from bot.utils.utils import check_command_target, require_user_type, require_group_role
from bot.utils.enums import UserType
from bot.utils.logger import LogInfo, LogError, LogType
from app.db.session import SessionLocal
from app.models.business import Projects
from app.models.telegram import TelegramProjectMember
import datetime

# --- Step 1: Request Project Selection ---
@bot.on_message(filters.command("syncchat") | filters.regex(r"^@\w+\s+/syncchat\b"))
async def syncchat_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, "syncchat")
    if args is None: return

    # Check permission directly from Telegram (not DB) since this is the bootstrap command
    try:
        member = await client.get_chat_member(message.chat.id, message.from_user.id)
        if member.status not in (ChatMemberStatus.OWNER, ChatMemberStatus.ADMINISTRATOR):
            await message.reply_text("⚠️ Chỉ <b>Owner</b> và <b>Admin</b> mới được sử dụng lệnh này.", parse_mode=ParseMode.HTML)
            return
    except Exception as e:
        LogError(f"[SyncChat] Error checking permission: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Không thể kiểm tra quyền hạn của bạn.")
        return

    LogInfo(f"[SyncChat] Started by @{message.from_user.username or message.from_user.id} in {message.chat.title}", LogType.SYSTEM_STATUS)
    
    db = SessionLocal()
    try:
        projects = db.query(Projects).all()
        if not projects:
            await message.reply_text("⚠️ Không tìm thấy dự án nào trong hệ thống.")
            return

        buttons = []
        for p in projects:
            buttons.append([InlineKeyboardButton(p.project_name, callback_data=f"sync_proj_{p.id}")])
        buttons.append([InlineKeyboardButton("Hủy", callback_data="sync_cancel")])
        
        reply_markup = InlineKeyboardMarkup(buttons)
        await message.reply_text(
            "<b>ĐỒNG BỘ NHÓM</b>\n\nVui lòng chọn dự án để đồng bộ nhóm này vào:",
            reply_markup=reply_markup,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error in syncchat_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra khi lấy danh sách dự án.")
    finally:
        db.close()

# --- Step 2: Request Group Role Selection ---
@bot.on_callback_query(filters.regex(r"^sync_proj_(.*)$"))
async def sync_proj_callback(client, callback_query: CallbackQuery):
    project_id = callback_query.matches[0].group(1)

    # Check if this is a special project with sub-menus
    db = SessionLocal()
    try:
        project = db.query(Projects).filter(Projects.id == project_id).first()
        
        # --- Other project sub-menu ---
        if project and project.project_name.lower() == "other":
            buttons = [
                [InlineKeyboardButton("Quản lý thiết bị", callback_data=f"sync_other_{project_id}_device")],
                [InlineKeyboardButton("Quản lý phương tiện", callback_data=f"sync_other_{project_id}_vehicle")],
                [InlineKeyboardButton("Quản lý hình ảnh, giấy tờ", callback_data=f"sync_other_{project_id}_image")],
                [
                    InlineKeyboardButton("Quay lại", callback_data="sync_back_to_proj"),
                    InlineKeyboardButton("Hủy", callback_data="sync_cancel")
                ]
            ]
            reply_markup = InlineKeyboardMarkup(buttons)
            await callback_query.message.edit_text(
                "<b>DỰ ÁN OTHER</b>\n\nVui lòng chọn loại nhóm:",
                reply_markup=reply_markup,
                parse_mode=ParseMode.HTML
            )
            return

        # --- Tiến Nga project sub-menu ---
        if project and "tiến nga" in project.project_name.lower():
            buttons = [
                [InlineKeyboardButton("Nhóm tổng", callback_data=f"sync_tn_{project_id}_tong")],
                [InlineKeyboardButton("Nhà cung cấp", callback_data=f"sync_tn_{project_id}_supplier")],
                [InlineKeyboardButton("Kinh doanh", callback_data=f"sync_tn_{project_id}_sales")],
                [InlineKeyboardButton("Nhân sự", callback_data=f"sync_tn_{project_id}_hr")],
                [InlineKeyboardButton("Tài chính", callback_data=f"sync_tn_{project_id}_finance")],
                [InlineKeyboardButton("Quản lý hàng thành phẩm", callback_data=f"sync_tn_{project_id}_product")],
                [InlineKeyboardButton("Quản lý đối tác", callback_data=f"sync_tn_{project_id}_partner")],
                [InlineKeyboardButton("Quản lý kho", callback_data=f"sync_tn_{project_id}_inventory")],
                [InlineKeyboardButton("Quản lý Cổ Đông", callback_data=f"sync_tn_{project_id}_sh")],
                [InlineKeyboardButton("Thu hoạch cao su", callback_data=f"sync_tn_{project_id}_harvest")],
                [
                    InlineKeyboardButton("Trở lại", callback_data="sync_back_to_proj"),
                    InlineKeyboardButton("Hủy", callback_data="sync_cancel")
                ]
            ]
            reply_markup = InlineKeyboardMarkup(buttons)
            await callback_query.message.edit_text(
                "<b>DỰ ÁN TIẾN NGA</b>\n\nVui lòng chọn phòng ban / loại nhóm:",
                reply_markup=reply_markup,
                parse_mode=ParseMode.HTML
            )
            return

    finally:
        db.close()

    # Normal project flow
    buttons = [
        [
            InlineKeyboardButton("Nhóm Chính (Main)", callback_data=f"sync_role_{project_id}_main_none"),
            InlineKeyboardButton("Nhóm Thành viên (Member)", callback_data=f"sync_role_{project_id}_member_none")
        ],
        [
            InlineKeyboardButton("Quay lại", callback_data="sync_back_to_proj"),
            InlineKeyboardButton("Hủy", callback_data="sync_cancel")
        ]
    ]
    
    reply_markup = InlineKeyboardMarkup(buttons)
    await callback_query.message.edit_text(
        "<b>Chọn vai trò của nhóm này trong dự án:</b>",
        reply_markup=reply_markup,
        parse_mode=ParseMode.HTML
    )


# --- Step 2b-TN: Tiến Nga - Select Role after department ---
TN_DEPT_LABELS = {
    "tong": "Nhóm Tổng",
    "supplier": "Nhà cung cấp",
    "sales": "Kinh doanh",
    "hr": "Nhân sự",
    "finance": "Tài chính",
    "product": "Quản lý hàng thành phẩm",
    "inventory": "Quản lý kho",
    "partner": "Quản lý đối tác",
    "sh": "Quản lý Cổ Đông",
    "harvest": "Thu hoạch cao su",
}

@bot.on_callback_query(filters.regex(r"^sync_tn_(.+)_(tong|supplier|sales|hr|finance|product|inventory|partner|sh|harvest)$"))
async def sync_tn_dept_callback(client, callback_query: CallbackQuery):
    project_id = callback_query.matches[0].group(1)
    dept = callback_query.matches[0].group(2)

    label = TN_DEPT_LABELS.get(dept, dept)

    # Special case: "Nhóm Tổng" -> auto-set role=main, custom_title=super_main
    if dept == "tong":
        buttons = [
            [
                InlineKeyboardButton("Nhóm Main", callback_data=f"sync_role_{project_id}_main_super_main"),
                InlineKeyboardButton("Nhóm Member", callback_data=f"sync_role_{project_id}_member_{dept}")
            ],
            [
                InlineKeyboardButton("Quay lại", callback_data=f"sync_proj_{project_id}"),
                InlineKeyboardButton("Hủy", callback_data="sync_cancel")
            ]
        ]
    else:
        buttons = [
            [
                InlineKeyboardButton("Nhóm Main", callback_data=f"sync_role_{project_id}_main_{dept}"),
                InlineKeyboardButton("Nhóm Member", callback_data=f"sync_role_{project_id}_member_{dept}")
            ],
            [
                InlineKeyboardButton("Quay lại", callback_data=f"sync_proj_{project_id}"),
                InlineKeyboardButton("Hủy", callback_data="sync_cancel")
            ]
        ]

    reply_markup = InlineKeyboardMarkup(buttons)
    await callback_query.message.edit_text(
        f"<b>{label}</b>\n\nChọn vai trò của nhóm này:",
        reply_markup=reply_markup,
        parse_mode=ParseMode.HTML
    )

# --- Step 2b: Other project - Select Role after sub-category ---
@bot.on_callback_query(filters.regex(r"^sync_other_(.+)_(device|vehicle|image)$"))
async def sync_other_subcategory_callback(client, callback_query: CallbackQuery):
    project_id = callback_query.matches[0].group(1)
    subcategory = callback_query.matches[0].group(2)

    label_map = {
        "device": "Quản lý thiết bị",
        "vehicle": "Quản lý phương tiện",
        "image": "Quản lý hình ảnh, giấy tờ",
    }
    label = label_map.get(subcategory, subcategory)

    buttons = [
        [
            InlineKeyboardButton("Nhóm Chính (Main)", callback_data=f"sync_role_{project_id}_main_{subcategory}"),
            InlineKeyboardButton("Nhóm Thành viên (Member)", callback_data=f"sync_role_{project_id}_member_{subcategory}")
        ],
        [
            InlineKeyboardButton("Quay lại", callback_data=f"sync_proj_{project_id}"),
            InlineKeyboardButton("Hủy", callback_data="sync_cancel")
        ]
    ]

    reply_markup = InlineKeyboardMarkup(buttons)
    await callback_query.message.edit_text(
        f"<b>{label}</b>\n\nChọn vai trò của nhóm này:",
        reply_markup=reply_markup,
        parse_mode=ParseMode.HTML
    )

# --- Step 3: Run Sync Logic ---
@bot.on_callback_query(filters.regex(r"^sync_role_(.+)_(main|member)_(.+)$"))
async def sync_role_callback(client, callback_query: CallbackQuery):
    project_id = callback_query.matches[0].group(1)
    group_role = callback_query.matches[0].group(2)
    subcategory = callback_query.matches[0].group(3)  # "none" for normal, "device"/"vehicle"/"image" for Other
    chat_id = callback_query.message.chat.id
    chat_title = callback_query.message.chat.title or ""

    # Determine custom_title for project sub-categories
    custom_title = None
    if subcategory != "none":
        from bot.utils.enums import CustomTitle
        # Special case: "super_main" is used directly as custom_title (Tiến Nga - Nhóm Tổng)
        if subcategory == "super_main":
            custom_title = CustomTitle.SUPER_MAIN.value
        else:
            # We map it to matching Enum value if possible, else fallback to string.
            # But the requirement says these are explicitly the values like "main_device", "member_vehicle"
            # Expand abbreviated callback keys to full subcategory names
            SUBCATEGORY_MAP = {"sh": "shareholder"}
            resolved = SUBCATEGORY_MAP.get(subcategory, subcategory)
            custom_title = f"{group_role}_{resolved}"  # e.g. CustomTitle.MAIN_DEVICE.value
            try:
                custom_title = CustomTitle(custom_title).value
            except ValueError:
                pass

    display_role = group_role.upper()
    extra_info = ""
    if custom_title:
        extra_info = f"\nCustom title: <b>{custom_title}</b>"
    
    await callback_query.message.edit_text(
        f"🔄 Đang bắt đầu đồng bộ nhóm với vai trò <b>{display_role}</b>...{extra_info}",
        parse_mode=ParseMode.HTML
    )
    
    db = SessionLocal()
    try:
        # Get project name for logging
        project = db.query(Projects).filter(Projects.id == project_id).first()
        project_name = project.project_name if project else "N/A"
        
        LogInfo(f"[SyncChat] Syncing Project: {project_name}, Role: {group_role}, CustomTitle: {custom_title}, Chat: {chat_id}", LogType.SYSTEM_STATUS)
        
        # Clean up ALL old records for this chat_id before re-syncing
        # This prevents stale data when a group is re-synced with a different project/role/custom_title
        old_records = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == str(chat_id)
        ).all()
        old_record_count = len(old_records)
        for old_record in old_records:
            db.delete(old_record)
        if old_record_count > 0:
            LogInfo(f"[SyncChat] Cleaned up {old_record_count} old records for chat {chat_id} before re-syncing.", LogType.SYSTEM_STATUS)
        
        synced_count = 0
        admin_count = 0
        member_count = 0
        
        owner_names = []
        admin_names = []
        member_names = []
        
        # Get all members
        async for member in client.get_chat_members(chat_id):
            if not member.user or member.user.is_bot:
                continue
            
            user_id = str(member.user.id)
            username = member.user.username
            full_name = f"{member.user.first_name or ''} {member.user.last_name or ''}".strip()
            
            # Determine role and slot_name
            status = member.status
            role = "MEMBER"
            slot_name = ""
            
            if status == ChatMemberStatus.OWNER:
                role = "OWNER"
                slot_name = "owner"
                owner_names.append(f"@{username}" if username else f"ID:{user_id}")
            elif status == ChatMemberStatus.ADMINISTRATOR:
                role = "ADMIN"
                admin_count += 1
                slot_name = f"admin_{admin_count:02d}"
                admin_names.append(f"@{username}" if username else f"ID:{user_id}")
            else:
                role = "MEMBER"
                member_count += 1
                slot_name = f"member_{member_count:02d}"
                member_names.append(f"@{username}" if username else f"ID:{user_id}")
            
            # Insert fresh record (old records were already cleaned up)
            new_member = TelegramProjectMember(
                project_id=project_id,
                chat_id=str(chat_id),
                group_name=chat_title,
                user_id=user_id,
                user_name=username,
                full_name=full_name,
                role=group_role,
                slot_name=slot_name,
                member_status=status.name if hasattr(status, 'name') else str(status),
                is_bot=False,
                custom_title=custom_title,
                first_seen_at=datetime.datetime.now(),
                last_seen_at=datetime.datetime.now()
            )
            db.add(new_member)
            
            synced_count += 1
            if synced_count % 10 == 0:
                LogInfo(f"[SyncChat] Progress: {synced_count} members processed...", LogType.SYSTEM_STATUS)

        db.commit()
        
        custom_title_line = ""
        if custom_title:
            custom_title_line = f"Custom title: <b>{custom_title}</b>\n"

        result_text = (
            f"✅ <b>Đồng bộ hoàn tất!</b>\n\n"
            f"Dự án: <b>{project_name}</b>\n"
            f"Vai trò nhóm: <b>{group_role.upper()}</b>\n"
            f"{custom_title_line}"
            f"Tổng số: <b>{synced_count}</b> hội viên.\n"
            f"- Chủ nhóm: {len(owner_names)} ({', '.join(owner_names)})\n"
            f"- Quản trị viên: {admin_count} ({', '.join(admin_names)})\n"
            f"- Thành viên: {member_count} ({', '.join(member_names)})\n"
            f"🗑 <i>Đã xóa <b>{old_record_count}</b> bản ghi cũ trước khi đồng bộ lại.</i>"
        )
        await callback_query.message.edit_text(result_text, parse_mode=ParseMode.HTML)
        LogInfo(f"[SyncChat] Successfully synced {synced_count} members for project {project_name}", LogType.SYSTEM_STATUS)
        
    except Exception as e:
        db.rollback()
        LogError(f"Error in sync_role_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.message.edit_text(f"❌ Có lỗi xảy ra trong quá trình đồng bộ: {str(e)}")
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^sync_back_to_proj$"))
async def sync_back_to_proj_callback(client, callback_query: CallbackQuery):
    db = SessionLocal()
    try:
        projects = db.query(Projects).all()
        if not projects:
            await callback_query.message.edit_text("⚠️ Không tìm thấy dự án nào trong hệ thống.")
            return

        buttons = []
        for p in projects:
            buttons.append([InlineKeyboardButton(p.project_name, callback_data=f"sync_proj_{p.id}")])
        buttons.append([InlineKeyboardButton("Hủy", callback_data="sync_cancel")])
        
        reply_markup = InlineKeyboardMarkup(buttons)
        await callback_query.message.edit_text(
            "<b>ĐỒNG BỘ NHÓM</b>\n\nVui lòng chọn dự án để đồng bộ nhóm này vào:",
            reply_markup=reply_markup,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error in sync_back_to_proj_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi lấy danh sách dự án.", show_alert=True)
    finally:
        db.close()

@bot.on_callback_query(filters.regex(r"^sync_cancel$"))
async def sync_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("❌ Đã hủy quá trình đồng bộ.")

# --- Get Info Customer ---
@bot.on_message(filters.command("get_info_customer") | filters.regex(r"^@\w+\s+/get_info_customer\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
async def get_info_customer_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, "get_info_customer")
    if args is None: return

    chat_id = str(message.chat.id)
    db = SessionLocal()
    try:
        # Check which project this chat belongs to
        current_project_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()

        if not current_project_member:
            await message.reply_text("⚠️ Nhóm này chưa được đồng bộ vào dự án nào. Vui lòng chạy lệnh /syncchat trước.")
            return

        project_id = current_project_member.project_id
        project = db.query(Projects).filter(Projects.id == project_id).first()
        project_name = project.project_name if project else "Không xác định"

        # Get all distinct member groups for this project
        member_groups = db.query(TelegramProjectMember.chat_id).filter(
            TelegramProjectMember.project_id == project_id,
            TelegramProjectMember.role == "member" # user specified that the group role is member
        ).distinct().all()

        if not member_groups:
            await message.reply_text(f"⚠️ Không tìm thấy nhóm nào có vai trò 'Member' trong dự án <b>{project_name}</b>.", parse_mode=ParseMode.HTML)
            return

        buttons = []
        for g_chat_id_tuple in member_groups:
            g_chat_id = g_chat_id_tuple[0]
            try:
                chat_info = await client.get_chat(int(g_chat_id))
                chat_title = chat_info.title if chat_info.title else g_chat_id
            except Exception:
                chat_title = f"Nhóm {g_chat_id}"
                
            buttons.append([InlineKeyboardButton(chat_title, callback_data=f"gic_grp_{g_chat_id}_{project_id}")])
        
        buttons.append([InlineKeyboardButton("Đóng", callback_data="gic_cancel")])
        
        reply_markup = InlineKeyboardMarkup(buttons)
        await message.reply_text(
            f"<b>TRUY XUẤT THÔNG TIN KHÁCH HÀNG</b>\n\nDự án: <b>{project_name}</b>\nVui lòng chọn nhóm để xem thành viên:",
            reply_markup=reply_markup,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error in get_info_customer_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra khi lấy danh sách nhóm.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^gic_grp_(.*)_(.*)$"))
async def gic_grp_callback(client, callback_query: CallbackQuery):
    target_chat_id = callback_query.matches[0].group(1)
    project_id = callback_query.matches[0].group(2)
    
    db = SessionLocal()
    try:
        # Fetch members in that group where slot_name starts with 'member'
        members = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == target_chat_id,
            TelegramProjectMember.slot_name.like("member%")
        ).order_by(TelegramProjectMember.slot_name).all()
        
        if not members:
            await callback_query.answer("⚠️ Nhóm này không có thành viên (khách hàng) nào.", show_alert=True)
            return
            
        try:
            chat_info = await client.get_chat(int(target_chat_id))
            chat_title = chat_info.title if chat_info.title else target_chat_id
        except Exception:
            chat_title = f"Nhóm {target_chat_id}"

        text = f"<b>Danh sách khách hàng - {chat_title}</b>\n\n"
        for idx, m in enumerate(members, 1):
            username_str = f"@{m.user_name}" if m.user_name else "Không có"
            full_name = m.full_name or "N/A"
            text += f"{idx}. <b>Fullname:</b> {full_name}\n"
            text += f"   - <b>Username:</b> {username_str}\n"
            text += f"   - <b>User ID:</b> <code>{m.user_id}</code>\n"
            text += f"   - <b>Slot:</b> {m.slot_name}\n\n"
            
        buttons = [
            [InlineKeyboardButton("Quay lại", callback_data=f"gic_back_{project_id}")],
            [InlineKeyboardButton("Đóng", callback_data="gic_cancel")]
        ]
        
        reply_markup = InlineKeyboardMarkup(buttons)
        await callback_query.message.edit_text(text, reply_markup=reply_markup, parse_mode=ParseMode.HTML)
    except Exception as e:
        LogError(f"Error in gic_grp_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi lấy danh sách khách hàng.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^gic_back_(.*)$"))
async def gic_back_callback(client, callback_query: CallbackQuery):
    project_id = callback_query.matches[0].group(1)
    
    db = SessionLocal()
    try:
        project = db.query(Projects).filter(Projects.id == project_id).first()
        project_name = project.project_name if project else "Không xác định"

        member_groups = db.query(TelegramProjectMember.chat_id).filter(
            TelegramProjectMember.project_id == project_id,
            TelegramProjectMember.role == "member"
        ).distinct().all()

        if not member_groups:
            await callback_query.answer("⚠️ Không tìm thấy nhóm nào.", show_alert=True)
            return

        buttons = []
        for g_chat_id_tuple in member_groups:
            g_chat_id = g_chat_id_tuple[0]
            try:
                chat_info = await client.get_chat(int(g_chat_id))
                chat_title = chat_info.title if chat_info.title else g_chat_id
            except Exception:
                chat_title = f"Nhóm {g_chat_id}"
                
            buttons.append([InlineKeyboardButton(chat_title, callback_data=f"gic_grp_{g_chat_id}_{project_id}")])
        
        buttons.append([InlineKeyboardButton("❌ Đóng", callback_data="gic_cancel")])
        
        reply_markup = InlineKeyboardMarkup(buttons)
        await callback_query.message.edit_text(
            f"<b>TRUY XUẤT THÔNG TIN KHÁCH HÀNG</b>\n\nDự án: <b>{project_name}</b>\nVui lòng chọn nhóm để xem thành viên:",
            reply_markup=reply_markup,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error in gic_back_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^gic_cancel$"))
async def gic_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.edit_text("❌ Đã đóng danh sách khách hàng.")


# --- Export Syncchat ---
@bot.on_message(filters.command(["export_syncchat"]) | filters.regex(r"^@\w+\s+/export_syncchat\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
async def export_syncchat_handler(client, message: Message) -> None:
    db = SessionLocal()
    try:
        projects = db.query(Projects).all()
        if not projects:
            await message.reply_text("\u26a0\ufe0f Kh\u00f4ng c\u00f3 d\u1ef1 \u00e1n n\u00e0o trong h\u1ec7 th\u1ed1ng.", parse_mode=ParseMode.HTML)
            return

        import tempfile
        import os
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F4F8F")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for project in projects:
            all_members = db.query(TelegramProjectMember).filter(
                TelegramProjectMember.project_id == project.id
            ).all()

            # Tab name: max 31 chars (Excel sheet name limit)
            sheet_name = (project.project_name or str(project.id))[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Header row
            headers = ["T\u00ean Nh\u00f3m", "Chat ID", "Role c\u1ee7a Nh\u00f3m", "Owner", "Admins", "Th\u00e0nh vi\u00ean"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # Group members by chat_id
            groups_data = {}
            for m in all_members:
                if m.chat_id not in groups_data:
                    groups_data[m.chat_id] = {
                        "group_name": m.group_name or f"Nh\u00f3m {m.chat_id}",
                        "role": m.role or "Kh\u00f4ng x\u00e1c \u0111\u1ecbnh",
                        "owners": [], "admins": [], "members": []
                    }
                contact = f"@{m.user_name}" if m.user_name else str(m.user_id)
                name = f"{m.full_name} ({contact})" if m.full_name else contact
                if m.member_status == "OWNER":
                    groups_data[m.chat_id]["owners"].append(name)
                elif m.member_status == "ADMINISTRATOR":
                    groups_data[m.chat_id]["admins"].append(name)
                elif m.member_status == "MEMBER":
                    groups_data[m.chat_id]["members"].append(name)

            # Data rows
            for row_idx, (cid, data) in enumerate(groups_data.items(), 2):
                ws.cell(row=row_idx, column=1, value=data["group_name"])
                ws.cell(row=row_idx, column=2, value=cid)
                ws.cell(row=row_idx, column=3, value=data["role"])
                ws.cell(row=row_idx, column=4, value=", ".join(data["owners"]) if data["owners"] else "Kh\u00f4ng c\u00f3")
                ws.cell(row=row_idx, column=5, value=", ".join(data["admins"]) if data["admins"] else "Kh\u00f4ng c\u00f3")
                ws.cell(row=row_idx, column=6, value=", ".join(data["members"]) if data["members"] else "Kh\u00f4ng c\u00f3")

            # Auto column width
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            if not groups_data:
                ws.cell(row=2, column=1, value="Ch\u01b0a c\u00f3 d\u1eef li\u1ec7u \u0111\u1ed3ng b\u1ed9")

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        await message.reply_document(
            document=tmp_path,
            file_name=f"export_syncchat_{now_str}.xlsx",
            caption=(
                f"<b>B\u00c1O C\u00c1O \u0110\u1ed2NG B\u1ed8 D\u1ef0 \u00c1N</b>\n"
                f"{len(projects)} d\u1ef1 \u00e1n | Xu\u1ea5t l\u00fac: {datetime.datetime.now().strftime('%H:%M %d/%m/%Y')}"
            ),
            parse_mode=ParseMode.HTML
        )
        os.remove(tmp_path)
        LogInfo(f"[ExportSyncchat] Exported {len(projects)} projects by @{message.from_user.username or message.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error in export_syncchat_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(f"\u274c C\u00f3 l\u1ed7i x\u1ea3y ra khi xu\u1ea5t b\u00e1o c\u00e1o: {e}", parse_mode=ParseMode.HTML)
    finally:
        db.close()

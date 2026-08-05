from pyrogram import filters
from pyrogram.types import Message, InlineKeyboardButton, InlineKeyboardMarkup, CallbackQuery
from pyrogram.enums import ParseMode
from bot.utils.bot import bot
from bot.utils.utils import check_command_target, require_user_type, require_project_name, require_group_role, command_timeout
from bot.utils.enums import UserType
from bot.utils.logger import LogInfo, LogError, LogType
from app.db.session import SessionLocal
from bot.utils.states import form_tracker
import datetime

# --- Rental: Create Customer ---
@bot.on_message(filters.command(["rental_create_customer", "rental_tao_khach_hang"]) | filters.regex(r"^@\w+\s+/(rental_create_customer|rental_tao_khach_hang)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(timeout_seconds=600, auto_delete_cmd=True)  # Form nhiều trường -> cần thời gian điền
async def rental_create_customer_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_create_customer", "rental_tao_khach_hang"])
    if args is None: return

    lines = message.text.strip().split("\n")
    if len(lines) < 3:
        form_template = """<b>FORM TẠO KHÁCH HÀNG CHO THUÊ</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/rental_create_customer
Mã Khách Hàng:
Tên Nhóm:
Tên Khách Hàng:
Liên Hệ Khách Hàng:
Số Điện Thoại:
Chat ID (Telegram):
</pre>

<i>Mã Khách Hàng là mã duy nhất để định danh khách hàng (ví dụ: KH001)
Chat ID (Telegram) là Chat ID nhóm member của khách hàng (ví dụ: -1001234567890). Nếu để trống, bot sẽ tự lấy theo Tên Nhóm đã đồng bộ.</i>"""
        form_msg = await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        form_tracker.track(message.chat.id, "rental_create_customer", "create", form_msg.id)
        return

    # Parse Form
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    customer_id_str = data.get("Mã Khách Hàng", "")
    group_name = data.get("Tên Nhóm", "")
    customer_name = data.get("Tên Khách Hàng", "")
    contact_info = data.get("Liên Hệ Khách Hàng", "")
    number_phone = data.get("Số Điện Thoại", "")
    input_chat_id = data.get("Chat ID (Telegram)", "").strip()

    if not customer_id_str:
        await message.reply_text("⚠️ <b>Mã Khách Hàng</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    if not customer_name:
        await message.reply_text("⚠️ <b>Tên Khách Hàng</b> là bắt buộc.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        from app.models.telegram import TelegramProjectMember
        from app.models.rental import RentalCustomer

        chat_id = str(message.chat.id)
        current_project_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()

        if not current_project_member:
            await message.reply_text("⚠️ Nhóm này chưa được đồng bộ vào dự án nào. Vui lòng sử dụng lệnh /syncchat trước.")
            return

        project_id = current_project_member.project_id

        # Check if contact is in the valid member list
        valid_members = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project_id,
            TelegramProjectMember.role == "member"
        ).all()

        valid_groups = []
        group_chat_ids = {}
        for m in valid_members:
            if m.group_name:
                valid_groups.append(m.group_name)
                if m.chat_id and m.group_name not in group_chat_ids:
                    group_chat_ids[m.group_name] = str(m.chat_id)

        if not group_name or group_name not in valid_groups:
            await message.reply_text(f"⚠️ Tên Nhóm <b>{group_name}</b> không hợp lệ hoặc chưa có mặt trong dự án. Vui lòng kiểm tra lại Tên Nhóm và dùng lệnh /syncchat để đồng bộ trước.", parse_mode=ParseMode.HTML)
            return

        # Chat ID nhóm member Telegram của khách hàng: ưu tiên giá trị nhập tay, nếu trống thì lấy theo Tên Nhóm
        if input_chat_id:
            valid_chat_ids = {str(m.chat_id) for m in valid_members if m.chat_id}
            if input_chat_id not in valid_chat_ids:
                await message.reply_text(f"⚠️ Chat ID <b>{input_chat_id}</b> không thuộc nhóm member nào của dự án. Vui lòng kiểm tra lại hoặc để trống để bot tự lấy theo Tên Nhóm.", parse_mode=ParseMode.HTML)
                return
            customer_chat_id = input_chat_id
        else:
            customer_chat_id = group_chat_ids.get(group_name)
            if not customer_chat_id:
                await message.reply_text(f"⚠️ Không xác định được Chat ID của nhóm <b>{group_name}</b>. Vui lòng điền trực tiếp <b>Chat ID (Telegram)</b> trên Form, hoặc dùng lệnh /syncchat trong nhóm đó trước.", parse_mode=ParseMode.HTML)
                return

        # Check duplicate by customer_id
        existing = db.query(RentalCustomer).filter(RentalCustomer.customer_id == customer_id_str).first()
        if existing:
            await message.reply_text(f"⚠️ Mã Khách Hàng <b>{customer_id_str}</b> đã tồn tại trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        new_customer = RentalCustomer(
            customer_id=customer_id_str,
            group_name=group_name,
            customer_name=customer_name,
            contact_info=contact_info,
            number_phone=number_phone,
            chat_id=customer_chat_id
        )
        db.add(new_customer)
        db.commit()

        await message.reply_text(
            f"✅ Đã tạo khách hàng cho thuê <b>{customer_name}</b> (Mã: {customer_id_str}) thành công!\n"
            f"Chat ID (Telegram): <code>{customer_chat_id}</code>",
            parse_mode=ParseMode.HTML
        )
        LogInfo(f"[RentalCreateCustomer] Created rental customer {customer_name} (ID: {customer_id_str}) chat_id={customer_chat_id} by {message.from_user.id}", LogType.SYSTEM_STATUS)

        # Delete the form template message after successful creation
        form_msg_id = form_tracker.pop(message.chat.id, "rental_create_customer", "create")
        if form_msg_id:
            try:
                await client.delete_messages(chat_id=message.chat.id, message_ids=form_msg_id)
            except Exception as del_err:
                LogError(f"Failed to delete rental create customer form: {del_err}", LogType.SYSTEM_STATUS)
    except Exception as e:
        db.rollback()
        LogError(f"Error in rental_create_customer_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình xử lý lưu khách hàng.")
    finally:
        db.close()


import re

def _parse_float_rental(val_str: str) -> float:
    if not val_str: return 0.0
    v = re.sub(r'[^\d\.,-]', '', str(val_str))
    if not v: return 0.0
    if "," in v and "." in v:
        if v.rfind(",") > v.rfind("."):
            v = v.replace(".", "").replace(",", ".")
        else:
            v = v.replace(",", "")
    elif "," in v:
        if len(v.split(",")[-1]) == 3:
            v = v.replace(",", "")
        else:
            v = v.replace(",", ".")
    elif "." in v:
        if len(v.split(".")[-1]) == 3:
            v = v.replace(".", "")
    if v.count('.') > 1:
        parts = v.split('.')
        v = "".join(parts[:-1]) + "." + parts[-1]
    return float(v)


# --- Rental: Create Contract (Interactive) ---
def _build_rental_create_contract_form(customer) -> str:
    """Form tạo hợp đồng cho thuê đã điền sẵn thông tin khách hàng."""
    return f"""<b>FORM TẠO HỢP ĐỒNG CHO THUÊ</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/rental_create_contract {customer.customer_id}
Mã Khách Hàng: {customer.customer_id or ""}
Tên Nhóm: {customer.group_name or ""}
Tên Khách Hàng: {customer.customer_name or ""}
Liên Hệ Khách Hàng: {customer.contact_info or ""}
Mã Hợp Đồng:
Mã Bất Động Sản:
Loại Hợp Đồng:
Ngày Bắt Đầu Thuê (dd/mm/yyyy):
Ngày Kết Thúc Thuê (dd/mm/yyyy):
Tiền Cọc:
Tiền Thuê / Tháng:
Số tiền nợ của khách: 0
</pre>

<i>Ví dụ mã hợp đồng: LMD-HD220101. Loại hợp đồng gồm (Nhà nguyên căn, Phòng trọ, Căn hộ, Mặt bằng, v.v.)</i>"""


def _rth_customer_list_keyboard(customers, page):
    """Bàn phím chọn khách hàng để tạo hợp đồng: tối đa 10 nút/trang, có Trước/Sau và Hủy."""
    PAGE_SIZE = 10
    total = len(customers)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_customers = customers[start:min(start + PAGE_SIZE, total)]

    buttons = [
        [InlineKeyboardButton(
            f"{c.customer_id} - {c.customer_name}",
            callback_data=f"rth_s|{_rsid(c.id)}|0"
        )]
        for c in page_customers
    ]

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rth_p|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rth_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rth_p|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rth_x")])
    return InlineKeyboardMarkup(buttons)


def _build_rth_contract_list_text(db, customer, page):
    """Nội dung + bàn phím danh sách hợp đồng hiện có của khách hàng (10 HĐ / trang)."""
    from app.models.rental import Rental

    PAGE_SIZE = 10
    contracts = db.query(Rental).filter(Rental.customer_id == customer.id).all()
    total = len(contracts)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_contracts = contracts[start:min(start + PAGE_SIZE, total)]

    lines = [
        f"<b>TẠO HỢP ĐỒNG CHO THUÊ</b>",
        f"Khách hàng: <b>{customer.customer_name}</b> ({customer.customer_id or 'N/A'})",
        f"Nhóm: {customer.group_name or 'N/A'} | SĐT: {customer.number_phone or 'N/A'}",
        f"{'━' * 15}",
        f"<b>DANH SÁCH HỢP ĐỒNG HIỆN CÓ ({total})</b>",
    ]

    if not contracts:
        lines.append("<i>Khách hàng chưa có hợp đồng cho thuê nào.</i>")
    else:
        for idx, r in enumerate(page_contracts, start + 1):
            lines.append(
                f"{idx}. {_rental_status_label(r.status)} <b>{r.contract_id}</b>"
                f" ({r.type_contract or 'N/A'})"
                f"\n        BĐS: {r.real_estate_id or 'N/A'}"
                f"\n        Thời gian thuê: {_rental_fmt_dt(r.start_rental)} - {_rental_fmt_dt(r.end_rental)}"
                f"\n        Tiền thuê: <b>{_rental_fmt_num(r.monthly_rental):,}</b>/tháng"
            )

    lines.append("")
    lines.append("Nhấn <b>Thêm hợp đồng</b> để tạo hợp đồng mới cho khách hàng này.")

    customer_hex = _rsid(customer.id)
    buttons = []
    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rth_s|{customer_hex}|{page - 1}"))
    if total_pages > 1:
        nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rth_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rth_s|{customer_hex}|{page + 1}"))
    if nav_row:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Thêm hợp đồng", callback_data=f"rth_add|{customer_hex}")])
    buttons.append([InlineKeyboardButton("Hủy", callback_data="rth_x")])

    return "\n".join(lines), InlineKeyboardMarkup(buttons)


@bot.on_callback_query(filters.regex(r"^rth_noop$"))
@command_timeout(auto_delete_cmd=True)
async def rth_noop_callback(client, callback_query: CallbackQuery):
    await callback_query.answer()


@bot.on_callback_query(filters.regex(r"^rth_x$"))
@command_timeout(auto_delete_cmd=True)
async def rth_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


@bot.on_callback_query(filters.regex(r"^rth_p\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rth_page_callback(client, callback_query: CallbackQuery):
    """Phân trang danh sách khách hàng khi tạo hợp đồng."""
    page = int(callback_query.matches[0].group(1))
    db = SessionLocal()
    try:
        customers = _rental_customers_of_chat(db, callback_query.message.chat.id)
        if not customers:
            await _rk_safe_edit(callback_query, "ℹ️ Chưa có khách hàng cho thuê nào trong dự án này.")
            return

        await _rk_safe_edit(
            callback_query,
            "<b>TẠO HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng cần tạo hợp đồng:",
            _rth_customer_list_keyboard(customers, page),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rth_page_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rth_s\|([a-f0-9]{32})\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rth_select_customer_callback(client, callback_query: CallbackQuery):
    """Chọn khách hàng -> hiển thị danh sách hợp đồng hiện có + nút Thêm hợp đồng."""
    customer_uuid = _ruid(callback_query.matches[0].group(1))
    page = int(callback_query.matches[0].group(2))
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == customer_uuid).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng.", show_alert=True)
            return

        text, keyboard = _build_rth_contract_list_text(db, customer, page)
        await _rk_safe_edit(callback_query, text, keyboard)
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rth_select_customer_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rth_add\|([a-f0-9]{32})$"))
@command_timeout(timeout_seconds=600, auto_delete_cmd=True)  # Form nhiều trường -> cần thời gian điền
async def rth_add_contract_callback(client, callback_query: CallbackQuery):
    """Hiển thị form tạo hợp đồng mới cho khách hàng đã chọn."""
    customer_uuid = _ruid(callback_query.matches[0].group(1))
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == customer_uuid).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng.", show_alert=True)
            return

        form_msg = await callback_query.message.reply_text(
            _build_rental_create_contract_form(customer), parse_mode=ParseMode.HTML
        )
        form_tracker.track(callback_query.message.chat.id, "rental_create_contract", customer.customer_id, form_msg.id)
        await callback_query.message.delete()
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rth_add_contract_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


# --- Rental: Create Contract ---
@bot.on_message(filters.command(["rental_create_contract", "rental_tao_hop_dong"]) | filters.regex(r"^@\w+\s+/(rental_create_contract|rental_tao_hop_dong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(timeout_seconds=600, auto_delete_cmd=True)  # Form nhiều trường -> cần thời gian điền
async def rental_create_contract_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_create_contract", "rental_tao_hop_dong"])
    if args is None: return

    chat_id = str(message.chat.id)
    db = SessionLocal()
    try:
        from app.models.telegram import TelegramProjectMember
        from app.models.rental import RentalCustomer, Rental, RentalStatus

        lines = message.text.strip().split("\n")
        if len(lines) < 3:
            if len(args) < 2:
                customers = _rental_customers_of_chat(db, message.chat.id)
                if not customers:
                    await message.reply_text("ℹ️ Chưa có khách hàng cho thuê nào trong dự án này. Vui lòng tạo khách hàng bằng lệnh /rental_create_customer trước.", parse_mode=ParseMode.HTML)
                    return

                await message.reply_text(
                    "<b>TẠO HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng cần tạo hợp đồng:",
                    reply_markup=_rth_customer_list_keyboard(customers, 0),
                    parse_mode=ParseMode.HTML
                )
                return

            target = args[1]
            from sqlalchemy import or_
            customer = db.query(RentalCustomer).filter(
                or_(RentalCustomer.customer_id == target, RentalCustomer.group_name == target)
            ).first()
            if not customer:
                await message.reply_text(f"⚠️ Khách hàng <b>{target}</b> chưa tồn tại trong hệ thống cho thuê. Vui lòng tạo khách hàng bằng lệnh /rental_create_customer trước.", parse_mode=ParseMode.HTML)
                return

            form_msg = await message.reply_text(
                _build_rental_create_contract_form(customer), parse_mode=ParseMode.HTML
            )
            form_tracker.track(message.chat.id, "rental_create_contract", customer.customer_id, form_msg.id)
            return

        # If user actually submitted data in the form format
        data = {}
        for line in lines[1:]:
            if ":" in line:
                key, val = line.split(":", 1)
                data[key.strip()] = val.strip()

        customer_id_input = data.get("Mã Khách Hàng", "")
        group_name = data.get("Tên Nhóm", "")

        if customer_id_input:
            customer = db.query(RentalCustomer).filter(RentalCustomer.customer_id == customer_id_input).first()
        elif group_name:
            customer = db.query(RentalCustomer).filter(RentalCustomer.group_name == group_name).first()
        elif len(args) > 1:
            target = args[1]
            from sqlalchemy import or_
            customer = db.query(RentalCustomer).filter(
                or_(RentalCustomer.customer_id == target, RentalCustomer.group_name == target)
            ).first()
        else:
            customer = None

        if not customer:
            await message.reply_text(f"⚠️ Khách hàng chưa tồn tại trong hệ thống cho thuê. Vui lòng tạo khách trước.", parse_mode=ParseMode.HTML)
            return

        parse_float = _parse_float_rental

        def parse_date(date_str: str):
            if not date_str: return None
            try:
                import datetime
                return datetime.datetime.strptime(date_str, "%d/%m/%Y").date()
            except:
                return None

        contract_id = data.get("Mã Hợp Đồng", "")
        if not contract_id:
            await message.reply_text("⚠️ <b>Mã Hợp Đồng</b> là bắt buộc.", parse_mode=ParseMode.HTML)
            return

        existing_contract = db.query(Rental).filter(Rental.contract_id == contract_id).first()
        if existing_contract:
            await message.reply_text(f"⚠️ Hợp đồng cho thuê mã <b>{contract_id}</b> đã tồn tại.", parse_mode=ParseMode.HTML)
            return

        real_estate_id = data.get("Mã Bất Động Sản", "")
        type_contract = data.get("Loại Hợp Đồng", "")
        start_rental = parse_date(data.get("Ngày Bắt Đầu Thuê (dd/mm/yyyy)", ""))
        end_rental = parse_date(data.get("Ngày Kết Thúc Thuê (dd/mm/yyyy)", ""))
        deposit = parse_float(data.get("Tiền Cọc", "0"))
        monthly_rental = parse_float(data.get("Tiền Thuê / Tháng", "0"))
        rental_debt = parse_float(data.get("Số tiền nợ của khách", "0"))

        from app.schemas.rental import RentalCreate
        from app.crud.rental import create_rental

        new_contract = RentalCreate(
            customer_id=customer.id,
            contract_id=contract_id,
            real_estate_id=real_estate_id,
            type_contract=type_contract,
            start_rental=start_rental,
            end_rental=end_rental,
            deposit=deposit,
            monthly_rental=monthly_rental,
            rental_debt=rental_debt,
            status=RentalStatus.ACTIVE.value
        )

        create_rental(db, obj_in=new_contract)
        await message.reply_text(f"✅ Đã tạo hợp đồng cho thuê <b>{contract_id}</b> cho khách hàng <b>{customer.customer_name}</b> thành công!", parse_mode=ParseMode.HTML)
        LogInfo(f"[RentalCreateContract] Created rental contract {contract_id} for {customer.customer_name} by {message.from_user.id}", LogType.SYSTEM_STATUS)

        # Delete the form template message after successful creation
        form_msg_id = form_tracker.pop(message.chat.id, "rental_create_contract", customer.customer_id)
        if form_msg_id:
            try:
                await client.delete_messages(chat_id=message.chat.id, message_ids=form_msg_id)
            except Exception as del_err:
                LogError(f"Failed to delete rental create contract form: {del_err}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in rental_create_contract_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình xử lý tạo hợp đồng cho thuê.")
    finally:
        db.close()


# --- Rental: Check Customer (Interactive) ---
def _rsid(uuid_val):
    """Rút gọn UUID thành chuỗi 32 ký tự hex (bỏ dấu gạch) để nhét vào callback_data."""
    return str(uuid_val).replace("-", "")


def _ruid(short_hex):
    """Khôi phục UUID từ chuỗi 32 ký tự hex."""
    import uuid as _uuid
    return str(_uuid.UUID(short_hex))


def _rental_fmt_num(val):
    if val is None: return 0
    return int(val) if val == int(val) else val


def _rental_fmt_dt(dt):
    if not dt: return "N/A"
    return dt.strftime('%d/%m/%Y')


def _rental_status_label(status):
    from app.models.rental import RentalStatus
    if status == RentalStatus.ACTIVE.value:
        return "Đang thuê"
    if status == RentalStatus.EXPIRED.value:
        return "Hết hạn"
    if status == RentalStatus.CANCELLED.value:
        return "Đã hủy"
    if status == RentalStatus.BAD_DEBT.value:
        return "Nợ xấu"
    return "Không rõ"


def _rental_valid_groups(db, chat_id):
    """Danh sách Tên Nhóm hợp lệ (role member) thuộc dự án của group hiện tại."""
    from app.models.telegram import TelegramProjectMember
    current_group = db.query(TelegramProjectMember).filter(
        TelegramProjectMember.chat_id == str(chat_id)
    ).first()
    if not current_group:
        return []

    valid_members = db.query(TelegramProjectMember).filter(
        TelegramProjectMember.project_id == current_group.project_id,
        TelegramProjectMember.role == "member"
    ).all()
    return [m.group_name for m in valid_members if m.group_name]


def _rental_customers_of_chat(db, chat_id):
    """Khách hàng cho thuê thuộc các nhóm hợp lệ của dự án hiện tại."""
    from app.models.rental import RentalCustomer
    valid_groups = _rental_valid_groups(db, chat_id)
    if not valid_groups:
        return []
    return db.query(RentalCustomer).filter(
        RentalCustomer.group_name.in_(valid_groups)
    ).order_by(RentalCustomer.customer_name).all()


def _build_rental_customer_info_text(db, customer) -> str:
    """Nội dung thông tin khách hàng + danh sách hợp đồng cho thuê."""
    from app.models.rental import Rental

    reply_lines = [
        f"<b>THÔNG TIN KHÁCH HÀNG CHO THUÊ</b>",
        f"Mã Khách Hàng: <b>{customer.customer_id or 'N/A'}</b>",
        f"Tên Nhóm: <b>{customer.group_name or 'N/A'}</b>",
        f"Tên Khách Hàng: <b>{customer.customer_name}</b>",
        f"Liên Hệ: <b>{customer.contact_info}</b>",
        f"Số Điện Thoại: <b>{customer.number_phone or 'N/A'}</b>",
        f"Chat ID (Telegram): <code>{customer.chat_id or 'Chưa có'}</code>",
        f"{'━' * 15}",
        f"<b>DANH SÁCH HỢP ĐỒNG CHO THUÊ</b>"
    ]

    rentals = db.query(Rental).filter(Rental.customer_id == customer.id).all()
    if not rentals:
        reply_lines.append("<i>Khách hàng chưa có hợp đồng cho thuê nào.</i>")
    else:
        for idx, r in enumerate(rentals, 1):
            reply_lines.append(
                f"{idx}. {_rental_status_label(r.status)} <b>{r.contract_id}</b>"
                f" ({r.type_contract or 'N/A'})"
                f"\n        BĐS: {r.real_estate_id or 'N/A'}"
                f"\n        Thời gian thuê: {_rental_fmt_dt(r.start_rental)} - {_rental_fmt_dt(r.end_rental)}"
                f"\n        Tiền thuê: <b>{_rental_fmt_num(r.monthly_rental):,}</b>/tháng"
                f"\n        Cọc: {_rental_fmt_num(r.deposit):,}"
            )

    return "\n".join(reply_lines)


def _build_rk_customer_list_keyboard(customers, page, selected_customer_id=None):
    """Bàn phím danh sách khách hàng (10 KH / trang) + nút chọn thao tác."""
    PAGE_SIZE = 10
    total = len(customers)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    end = min(start + PAGE_SIZE, total)
    page_customers = customers[start:end]

    sel_hex = _rsid(selected_customer_id) if selected_customer_id else None

    buttons = []
    for c in page_customers:
        c_hex = _rsid(c.id)
        is_selected = (c_hex == sel_hex) if sel_hex else False
        prefix = "✅" if is_selected else "⬜"
        label = f"{prefix} {c.customer_id} - {c.customer_name}"
        # rk_s|<32hex>|<page>
        buttons.append([InlineKeyboardButton(label, callback_data=f"rk_s|{c_hex}|{page}")])

    # Hàng phân trang
    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rk_p|{page - 1}|{sel_hex or ''}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rk_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rk_p|{page + 1}|{sel_hex or ''}"))
    buttons.append(nav_row)

    # Hàng thao tác (chỉ hiện khi đã chọn khách hàng)
    if sel_hex:
        buttons.append([
            InlineKeyboardButton("Thông tin khách hàng", callback_data=f"rk_info|{sel_hex}"),
        ])
        buttons.append([
            InlineKeyboardButton("Thông tin hợp đồng", callback_data=f"rk_hd|{sel_hex}"),
        ])

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rk_x")])

    buttons = [row for row in buttons if row]
    return InlineKeyboardMarkup(buttons)


def _build_rk_contract_list_keyboard(contracts, page, customer_hex):
    """Bàn phím danh sách hợp đồng cho thuê (10 HĐ / trang)."""
    PAGE_SIZE = 10
    total = len(contracts)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    end = min(start + PAGE_SIZE, total)
    page_contracts = contracts[start:end]

    buttons = []
    for c in page_contracts:
        label = f"{c.contract_id} ({_rental_status_label(c.status)})"
        # rk_sc|<32hex>
        buttons.append([InlineKeyboardButton(label, callback_data=f"rk_sc|{_rsid(c.id)}")])

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rk_cp|{page - 1}|{customer_hex}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rk_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rk_cp|{page + 1}|{customer_hex}"))
    buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rk_x")])

    buttons = [row for row in buttons if row]
    return InlineKeyboardMarkup(buttons)


def _build_rk_customer_list_header(customers, selected_uuid=None):
    header = "<b>DANH SÁCH KHÁCH HÀNG CHO THUÊ</b>\n\n"
    if selected_uuid:
        selected_cust = None
        for c in customers:
            if str(c.id) == selected_uuid:
                selected_cust = c
                break
        if selected_cust:
            header += (
                f"<b>Đã chọn:</b> {selected_cust.customer_name} ({selected_cust.customer_id})\n"
                f"Nhóm: {selected_cust.group_name or 'N/A'} | "
                f"SĐT: {selected_cust.number_phone or 'N/A'}\n\n"
            )
    header += "Chọn khách hàng để xem thông tin:"
    return header


@bot.on_message(filters.command(["rental_check_customer", "rental_kiem_tra_khach_hang"]) | filters.regex(r"^@\w+\s+/(rental_check_customer|rental_kiem_tra_khach_hang)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(auto_delete_cmd=True)
async def rental_check_customer_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_check_customer", "rental_kiem_tra_khach_hang"])
    if args is None: return

    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Danh sách khách hàng", callback_data="rk_list|0")],
        [InlineKeyboardButton("Hủy", callback_data="rk_x")]
    ])
    await message.reply_text(
        "<b>KIỂM TRA KHÁCH HÀNG CHO THUÊ</b>\n\nVui lòng chọn thao tác:",
        reply_markup=keyboard,
        parse_mode=ParseMode.HTML
    )


@bot.on_callback_query(filters.regex(r"^rk_noop$"))
@command_timeout(auto_delete_cmd=True)
async def rk_noop_callback(client, callback_query: CallbackQuery):
    await callback_query.answer()


@bot.on_callback_query(filters.regex(r"^rk_x$"))
@command_timeout(auto_delete_cmd=True)
async def rk_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


async def _rk_safe_edit(callback_query: CallbackQuery, text: str, reply_markup=None) -> None:
    """edit_text nhưng bỏ qua lỗi MESSAGE_NOT_MODIFIED (bấm lại đúng nút đang chọn)."""
    from pyrogram.errors import MessageNotModified
    try:
        await callback_query.message.edit_text(
            text, reply_markup=reply_markup, parse_mode=ParseMode.HTML
        )
    except MessageNotModified:
        pass


@bot.on_callback_query(filters.regex(r"^rk_list\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rk_show_list_callback(client, callback_query: CallbackQuery):
    page = int(callback_query.matches[0].group(1))
    db = SessionLocal()
    try:
        customers = _rental_customers_of_chat(db, callback_query.message.chat.id)
        if not customers:
            await _rk_safe_edit(callback_query, "ℹ️ Không có khách hàng cho thuê nào trong dự án này.")
            return

        await _rk_safe_edit(
            callback_query,
            _build_rk_customer_list_header(customers),
            _build_rk_customer_list_keyboard(customers, page),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rk_show_list_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rk_s\|([a-f0-9]{32})\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rk_sel_callback(client, callback_query: CallbackQuery):
    """Chọn 1 khách hàng -> hiện thêm nút Thông tin khách hàng / Thông tin hợp đồng."""
    selected_hex = callback_query.matches[0].group(1)
    page = int(callback_query.matches[0].group(2))
    selected_uuid = _ruid(selected_hex)
    db = SessionLocal()
    try:
        customers = _rental_customers_of_chat(db, callback_query.message.chat.id)
        if not customers:
            await _rk_safe_edit(callback_query, "ℹ️ Không có khách hàng cho thuê nào trong dự án này.")
            return

        await _rk_safe_edit(
            callback_query,
            _build_rk_customer_list_header(customers, selected_uuid),
            _build_rk_customer_list_keyboard(customers, page, selected_uuid),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rk_sel_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rk_p\|(\d+)\|(.*)$"))
@command_timeout(auto_delete_cmd=True)
async def rk_page_callback(client, callback_query: CallbackQuery):
    """Phân trang danh sách khách hàng, giữ nguyên khách hàng đang chọn."""
    page = int(callback_query.matches[0].group(1))
    selected_hex = callback_query.matches[0].group(2) or None
    selected_uuid = _ruid(selected_hex) if selected_hex else None
    db = SessionLocal()
    try:
        customers = _rental_customers_of_chat(db, callback_query.message.chat.id)
        if not customers:
            await _rk_safe_edit(callback_query, "ℹ️ Không có khách hàng cho thuê nào trong dự án này.")
            return

        await _rk_safe_edit(
            callback_query,
            _build_rk_customer_list_header(customers, selected_uuid),
            _build_rk_customer_list_keyboard(customers, page, selected_uuid),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rk_page_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rk_info\|([a-f0-9]{32})$"))
@command_timeout(auto_delete_cmd=True)
async def rk_customer_info_callback(client, callback_query: CallbackQuery):
    """Hiển thị thông tin khách hàng đã chọn."""
    customer_hex = callback_query.matches[0].group(1)
    customer_uuid = _ruid(customer_hex)
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == customer_uuid).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng.", show_alert=True)
            return

        await callback_query.message.reply_text(
            _build_rental_customer_info_text(db, customer), parse_mode=ParseMode.HTML
        )
        await callback_query.message.delete()
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rk_customer_info_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rk_hd\|([a-f0-9]{32})$"))
@command_timeout(auto_delete_cmd=True)
async def rk_show_contract_list_callback(client, callback_query: CallbackQuery):
    """Hiển thị danh sách hợp đồng của khách hàng đã chọn."""
    customer_hex = callback_query.matches[0].group(1)
    customer_uuid = _ruid(customer_hex)
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == customer_uuid).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng.", show_alert=True)
            return

        contracts = db.query(Rental).filter(Rental.customer_id == customer.id).all()
        if not contracts:
            await callback_query.answer("ℹ️ Khách hàng chưa có hợp đồng cho thuê nào.", show_alert=True)
            return

        await _rk_safe_edit(
            callback_query,
            f"<b>DANH SÁCH HỢP ĐỒNG CHO THUÊ</b>\n"
            f"Khách hàng: <b>{customer.customer_name}</b> ({customer.customer_id})\n\n"
            f"Chọn hợp đồng để xem thông tin:",
            _build_rk_contract_list_keyboard(contracts, 0, customer_hex),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rk_show_contract_list_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rk_cp\|(\d+)\|([a-f0-9]{32})$"))
@command_timeout(auto_delete_cmd=True)
async def rk_contract_page_callback(client, callback_query: CallbackQuery):
    """Phân trang danh sách hợp đồng."""
    page = int(callback_query.matches[0].group(1))
    customer_hex = callback_query.matches[0].group(2)
    customer_uuid = _ruid(customer_hex)
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == customer_uuid).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng.", show_alert=True)
            return

        contracts = db.query(Rental).filter(Rental.customer_id == customer.id).all()
        if not contracts:
            await callback_query.answer("ℹ️ Khách hàng chưa có hợp đồng cho thuê nào.", show_alert=True)
            return

        await _rk_safe_edit(
            callback_query,
            f"<b>DANH SÁCH HỢP ĐỒNG CHO THUÊ</b>\n"
            f"Khách hàng: <b>{customer.customer_name}</b> ({customer.customer_id})\n\n"
            f"Chọn hợp đồng để xem thông tin:",
            _build_rk_contract_list_keyboard(contracts, page, customer_hex),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rk_contract_page_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rk_sc\|([a-f0-9]{32})$"))
@command_timeout(auto_delete_cmd=True)
async def rk_select_contract_callback(client, callback_query: CallbackQuery):
    """Hiển thị thông tin hợp đồng cho thuê đã chọn."""
    contract_hex = callback_query.matches[0].group(1)
    contract_uuid = _ruid(contract_hex)
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental
        contract = db.query(Rental).filter(Rental.id == contract_uuid).first()
        if not contract:
            await callback_query.answer("⚠️ Không tìm thấy hợp đồng.", show_alert=True)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng của hợp đồng này.", show_alert=True)
            return

        await _send_rental_contract_detail(callback_query.message, db, contract, customer)
        await callback_query.message.delete()
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rk_select_contract_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


async def _send_rental_contract_detail(message, db, contract, customer) -> None:
    """Gửi thông tin chi tiết 1 hợp đồng cho thuê (kèm file Excel nếu > 12 tháng)."""
    from app.models.rental import RentalStatus

    # Get start date and determine months to display
    import datetime
    from app.models.rental import RentalPayment
    
    # Load all payments for this contract
    payments = db.query(RentalPayment).filter(
        RentalPayment.contract_id == contract.contract_id
    ).order_by(RentalPayment.payment_date.asc()).all()
    
    # Calculate total paid
    total_paid = sum([p.payment_amount or 0.0 for p in payments])
    
    # Format helpers
    def fmt_num(val):
        if val is None: return 0
        return int(val) if val == int(val) else val

    def fmt_dt(dt):
        if not dt: return "N/A"
        return dt.strftime('%d/%m/%Y')

    # Group payments by month (year, month)
    monthly_payments = {}
    for p in payments:
        if p.payment_date:
            ym = (p.payment_date.year, p.payment_date.month)
            monthly_payments.setdefault(ym, []).append(p)
            
    # Generate the list of lease months from start_rental's month up to:
    # - end_rental's month (if end_rental is in the past)
    # - current month (otherwise)
    months_list = []
    if contract.start_rental:
        start_rental_date = contract.start_rental
        end_limit_date = datetime.date.today()
        if contract.end_rental and contract.end_rental <= datetime.date.today():
            end_limit_date = contract.end_rental
            
        curr = datetime.date(start_rental_date.year, start_rental_date.month, 1)
        end_limit = datetime.date(end_limit_date.year, end_limit_date.month, 1)
        while curr <= end_limit:
            months_list.append((curr.year, curr.month))
            if curr.month == 12:
                curr = datetime.date(curr.year + 1, 1, 1)
            else:
                curr = datetime.date(curr.year, curr.month + 1, 1)
    else:
        # Fallback to payments keys only
        months_list = list(monthly_payments.keys())
                
    # Ensure any month with an actual payment is included, even if outside of standard range
    all_keys = set(months_list).union(monthly_payments.keys())
    sorted_keys = sorted(list(all_keys), key=lambda x: (x[0], x[1]))

    if contract.status == RentalStatus.ACTIVE.value:
        status_label = "Đang thuê"
    elif contract.status == RentalStatus.EXPIRED.value:
        status_label = "Hết hạn"
    elif contract.status == RentalStatus.CANCELLED.value:
        status_label = "Đã hủy"
    elif contract.status == RentalStatus.BAD_DEBT.value:
        status_label = "Nợ xấu"
    else:
        status_label = "Không rõ"

    reply_lines = [
        f"<b>THÔNG TIN HỢP ĐỒNG CHO THUÊ: {contract.contract_id}</b>",
        f"Trạng thái: <b>{status_label}</b>",
        f"Loại hợp đồng: <b>{contract.type_contract or 'N/A'}</b>",
        f"Mã Khách Hàng: <b>{customer.customer_id or 'N/A'}</b>",
        f"Tên: <b>{customer.customer_name}</b>",
        f"Liên hệ: <b>{customer.contact_info or 'N/A'}</b>",
        f"Số Điện Thoại: <b>{customer.number_phone or 'N/A'}</b>",
        f"Mã BĐS: <b>{contract.real_estate_id or 'N/A'}</b>",
        f"Tiền Cọc: <b>{fmt_num(contract.deposit):,} VNĐ</b>",
        f"Tiền Thuê / Tháng: <b>{fmt_num(contract.monthly_rental):,} VNĐ</b>",
        f"Tổng giá trị hợp đồng: <b>{fmt_num(total_paid):,} VNĐ</b>",
        f"Ngày Bắt Đầu Thuê: <b>{fmt_dt(contract.start_rental)}</b>",
        f"Ngày Kết Thúc Thuê: <b>{fmt_dt(contract.end_rental)}</b>",
    ]

    if len(sorted_keys) > 12:
        import tempfile
        import os
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lịch sử thanh toán"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        money_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        alt_fill = PatternFill("solid", fgColor="D9E2F3")

        headers = ["STT", "Tháng", "Số tiền đã đóng (VNĐ)", "Ngày đóng", "Trạng thái"]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        monthly_rental = contract.monthly_rental or 0.0

        for idx, ym in enumerate(sorted_keys, 1):
            row = idx + 1
            row_fill = alt_fill if idx % 2 == 0 else None
            
            month_str = f"{ym[1]:02d}/{ym[0]}"
            
            p_list = monthly_payments.get(ym, [])
            total_amt = sum([p.payment_amount or 0 for p in p_list])
            
            if p_list:
                dates_str = ", ".join([p.payment_date.strftime('%d/%m/%Y') for p in p_list if p.payment_date])
                if total_amt >= monthly_rental:
                    status_str = "Đã đóng đủ"
                else:
                    status_str = "Đóng thiếu"
            else:
                dates_str = "—"
                status_str = "Chưa đóng"

            values = [
                idx,
                month_str,
                f"{int(total_amt):,}".replace(",", ".") if total_amt > 0 else "0",
                dates_str,
                status_str
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                
                if col_idx in [1, 2, 4]:
                    cell.alignment = center_align
                elif col_idx == 3:
                    cell.alignment = money_align
                else:
                    cell.alignment = left_align
                    
                if row_fill:
                    cell.fill = row_fill

        col_widths = [8, 15, 25, 25, 18]
        for col_idx, width in enumerate(col_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = "A2"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)

        now = datetime.datetime.now()
        file_name = f"rental_payments_{now.strftime('%Y_%m_%d_%H_%M_%S')}.xlsx"
        
        # Send contract info and note in text
        reply_lines.append("")
        reply_lines.append("<i>⚠️ Chi tiết đóng tiền từng tháng (>12 tháng) được xuất ra file Excel đính kèm dưới đây.</i>")
        
        await message.reply_document(
            document=tmp_path,
            file_name=file_name,
            caption="\n".join(reply_lines),
            parse_mode=ParseMode.HTML
        )
        
        # Clean up temp file safely
        try:
            os.remove(tmp_path)
        except Exception as clean_err:
            LogError(f"Error cleaning up temp file {tmp_path}: {clean_err}", LogType.SYSTEM_STATUS)
            
    else:
        # 12 months or fewer: append to reply message directly
        reply_lines.append("")
        reply_lines.append("<b>Chi tiết đóng tiền từng tháng:</b>")
        for ym in sorted_keys:
            month_str = f"{ym[1]:02d}/{ym[0]}"
            p_list = monthly_payments.get(ym, [])
            total_amt = sum([p.payment_amount or 0.0 for p in p_list])
            if p_list:
                reply_lines.append(f"- Tháng {month_str}: Đã đóng <b>{fmt_num(total_amt):,} VNĐ</b>")
            else:
                reply_lines.append(f"- Tháng {month_str}: <b>Chưa đóng</b>")
                
        await message.reply_text("\n".join(reply_lines), parse_mode=ParseMode.HTML)


# --- Rental: Check Contract (Interactive) ---
def _build_rental_cancel_confirm_text(contract, customer) -> str:
    """Nội dung xác nhận trước khi hủy 1 hợp đồng cho thuê."""
    return (
        f"<b>THÔNG TIN HỢP ĐỒNG CHO THUÊ</b>\n\n"
        f"- Tên Khách Hàng: <b>{customer.customer_name if customer else 'N/A'}</b>\n"
        f"- Mã Hợp Đồng: <b>{contract.contract_id}</b>\n"
        f"- Trạng Thái: <b>{_rental_status_label(contract.status)}</b>\n"
        f"- Loại Hợp Đồng: <b>{contract.type_contract or 'N/A'}</b>\n"
        f"- Tiền Thuê / Tháng: <b>{_rental_fmt_num(contract.monthly_rental):,} VNĐ</b>\n"
        f"- Tiền Cọc: <b>{_rental_fmt_num(contract.deposit):,} VNĐ</b>\n"
        f"- Thuê: <b>{_rental_fmt_dt(contract.start_rental)} - {_rental_fmt_dt(contract.end_rental)}</b>\n\n"
        f"Bạn có chắc chắn muốn hủy hợp đồng này không?"
    )


def _rental_cancel_confirm_keyboard(contract):
    """Nút xác nhận / thoát cho thao tác hủy hợp đồng."""
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("Xác nhận hủy", callback_data=f"rcc_confirm_{contract.id}"),
        InlineKeyboardButton("Thoát", callback_data="rcc_exit")
    ]])


def _rkh_customer_list_keyboard(customers, page):
    """Bàn phím chọn khách hàng để kiểm tra hợp đồng: tối đa 10 nút/trang."""
    PAGE_SIZE = 10
    total = len(customers)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_customers = customers[start:min(start + PAGE_SIZE, total)]

    buttons = [
        [InlineKeyboardButton(
            f"{c.customer_id} - {c.customer_name}",
            callback_data=f"rkh_c|{_rsid(c.id)}|0"
        )]
        for c in page_customers
    ]

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rkh_cp|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rkh_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rkh_cp|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rkh_x")])
    return InlineKeyboardMarkup(buttons)


def _rkh_contract_list_keyboard(contracts, page, customer_hex, selected_contract_id=None, allow_admin_actions=True):
    """Bàn phím danh sách hợp đồng; chọn 1 HĐ thì hiện thêm nút thao tác.

    `allow_admin_actions=False` (nhóm member) chỉ cho xem Thông tin hợp đồng,
    không hiện Cập nhật / Hủy hợp đồng.
    """
    PAGE_SIZE = 10
    total = len(contracts)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_contracts = contracts[start:min(start + PAGE_SIZE, total)]

    sel_hex = _rsid(selected_contract_id) if selected_contract_id else None

    buttons = []
    for c in page_contracts:
        c_hex = _rsid(c.id)
        prefix = "✅" if (sel_hex and c_hex == sel_hex) else "⬜"
        buttons.append([InlineKeyboardButton(
            f"{prefix} {c.contract_id} ({_rental_status_label(c.status)})",
            callback_data=f"rkh_s|{c_hex}|{page}"
        )])

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rkh_c|{customer_hex}|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rkh_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rkh_c|{customer_hex}|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    if sel_hex:
        buttons.append([InlineKeyboardButton("Thông tin hợp đồng", callback_data=f"rkh_info|{sel_hex}")])
        if allow_admin_actions:
            buttons.append([InlineKeyboardButton("Cập nhật hợp đồng", callback_data=f"rkh_upd|{sel_hex}")])
            buttons.append([InlineKeyboardButton("Hủy hợp đồng", callback_data=f"rkh_del|{sel_hex}")])

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rkh_x")])
    return InlineKeyboardMarkup(buttons)


def _rkh_contract_list_header(customer, selected_contract=None) -> str:
    header = (
        f"<b>KIỂM TRA HỢP ĐỒNG CHO THUÊ</b>\n"
        f"Khách hàng: <b>{customer.customer_name}</b> ({customer.customer_id or 'N/A'})\n\n"
    )
    if selected_contract:
        header += (
            f"<b>Đã chọn:</b> {selected_contract.contract_id} "
            f"({_rental_status_label(selected_contract.status)})\n"
            f"BĐS: {selected_contract.real_estate_id or 'N/A'} | "
            f"Tiền thuê: {_rental_fmt_num(selected_contract.monthly_rental):,}/tháng\n\n"
        )
    header += "Chọn hợp đồng để xem / thao tác:"
    return header


def _rkh_contracts_of(db, customer):
    from app.models.rental import Rental
    return db.query(Rental).filter(Rental.customer_id == customer.id).all()


def _rental_is_main_group(db, chat_id) -> bool:
    """Nhóm hiện tại có phải nhóm main (điều hành) của dự án không."""
    from app.models.telegram import TelegramProjectMember
    return db.query(TelegramProjectMember).filter(
        TelegramProjectMember.chat_id == str(chat_id),
        TelegramProjectMember.role == "main"
    ).first() is not None


def _rkh_customers_for_chat(db, chat_id):
    """Nhóm main: toàn dự án. Nhóm member: chỉ khách hàng gắn với chính nhóm đó."""
    from app.models.rental import RentalCustomer
    if _rental_is_main_group(db, chat_id):
        return _rental_customers_of_chat(db, chat_id)

    # Nhóm member -> tuyệt đối không rơi về danh sách toàn dự án (lộ khách hàng nhóm khác)
    return db.query(RentalCustomer).filter(
        RentalCustomer.chat_id == str(chat_id)
    ).order_by(RentalCustomer.customer_name).all()


@bot.on_callback_query(filters.regex(r"^rkh_noop$"))
@command_timeout(auto_delete_cmd=True)
async def rkh_noop_callback(client, callback_query: CallbackQuery):
    await callback_query.answer()


@bot.on_callback_query(filters.regex(r"^rkh_x$"))
@command_timeout(auto_delete_cmd=True)
async def rkh_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


@bot.on_callback_query(filters.regex(r"^rkh_cp\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rkh_customer_page_callback(client, callback_query: CallbackQuery):
    """Phân trang danh sách khách hàng khi kiểm tra hợp đồng."""
    page = int(callback_query.matches[0].group(1))
    db = SessionLocal()
    try:
        customers = _rkh_customers_for_chat(db, callback_query.message.chat.id)
        if not customers:
            await _rk_safe_edit(callback_query, "ℹ️ Chưa có khách hàng cho thuê nào trong dự án này.")
            return

        await _rk_safe_edit(
            callback_query,
            "<b>KIỂM TRA HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng cần kiểm tra:",
            _rkh_customer_list_keyboard(customers, page),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rkh_customer_page_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rkh_c\|([a-f0-9]{32})\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rkh_contract_list_callback(client, callback_query: CallbackQuery):
    """Chọn khách hàng -> hiển thị danh sách hợp đồng."""
    customer_hex = callback_query.matches[0].group(1)
    page = int(callback_query.matches[0].group(2))
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == _ruid(customer_hex)).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng.", show_alert=True)
            return

        contracts = _rkh_contracts_of(db, customer)
        if not contracts:
            await callback_query.answer("ℹ️ Khách hàng chưa có hợp đồng cho thuê nào.", show_alert=True)
            return

        await _rk_safe_edit(
            callback_query,
            _rkh_contract_list_header(customer),
            _rkh_contract_list_keyboard(
                contracts, page, customer_hex,
                allow_admin_actions=_rental_is_main_group(db, callback_query.message.chat.id)
            ),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rkh_contract_list_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rkh_s\|([a-f0-9]{32})\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rkh_select_contract_callback(client, callback_query: CallbackQuery):
    """Chọn 1 hợp đồng -> hiện thêm 3 nút Thông tin / Cập nhật / Hủy hợp đồng."""
    contract_hex = callback_query.matches[0].group(1)
    page = int(callback_query.matches[0].group(2))
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental
        contract = db.query(Rental).filter(Rental.id == _ruid(contract_hex)).first()
        if not contract:
            await callback_query.answer("⚠️ Không tìm thấy hợp đồng.", show_alert=True)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng của hợp đồng này.", show_alert=True)
            return

        contracts = _rkh_contracts_of(db, customer)
        await _rk_safe_edit(
            callback_query,
            _rkh_contract_list_header(customer, contract),
            _rkh_contract_list_keyboard(
                contracts, page, _rsid(customer.id), contract.id,
                allow_admin_actions=_rental_is_main_group(db, callback_query.message.chat.id)
            ),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rkh_select_contract_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rkh_info\|([a-f0-9]{32})$"))
@command_timeout(auto_delete_cmd=True)
async def rkh_contract_info_callback(client, callback_query: CallbackQuery):
    """Hiển thị thông tin chi tiết hợp đồng đã chọn."""
    contract_hex = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental
        contract = db.query(Rental).filter(Rental.id == _ruid(contract_hex)).first()
        if not contract:
            await callback_query.answer("⚠️ Không tìm thấy hợp đồng.", show_alert=True)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng của hợp đồng này.", show_alert=True)
            return

        await _send_rental_contract_detail(callback_query.message, db, contract, customer)
        await callback_query.message.delete()
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rkh_contract_info_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rkh_upd\|([a-f0-9]{32})$"))
@command_timeout(timeout_seconds=600, auto_delete_cmd=True)  # Form nhiều trường -> cần thời gian điền
async def rkh_update_contract_callback(client, callback_query: CallbackQuery):
    """Hiển thị form cập nhật cho hợp đồng đã chọn."""
    contract_hex = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental, RentalStatus

        # Cập nhật hợp đồng chỉ thực hiện từ nhóm main (giống /rental_cap_nhat_hop_dong)
        if not _rental_is_main_group(db, callback_query.message.chat.id):
            await callback_query.answer("⚠️ Thao tác này chỉ thực hiện được trong nhóm điều hành.", show_alert=True)
            return

        contract = db.query(Rental).filter(Rental.id == _ruid(contract_hex)).first()
        if not contract:
            await callback_query.answer("⚠️ Không tìm thấy hợp đồng.", show_alert=True)
            return

        if contract.status == RentalStatus.CANCELLED.value:
            await callback_query.answer("⚠️ Hợp đồng đã bị hủy, không thể cập nhật.", show_alert=True)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng của hợp đồng này.", show_alert=True)
            return

        form_msg = await callback_query.message.reply_text(
            _build_rental_update_contract_form(contract, customer), parse_mode=ParseMode.HTML
        )
        form_tracker.track(callback_query.message.chat.id, "rental_update_contract", contract.contract_id, form_msg.id)
        await callback_query.message.delete()
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rkh_update_contract_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rkh_del\|([a-f0-9]{32})$"))
@command_timeout(auto_delete_cmd=True)
async def rkh_cancel_contract_callback(client, callback_query: CallbackQuery):
    """Hiển thị xác nhận hủy hợp đồng (dùng chung nút rcc_confirm / rcc_exit)."""
    contract_hex = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental, RentalStatus

        # Hủy hợp đồng chỉ thực hiện từ nhóm main (giống /rental_huy_hop_dong)
        if not _rental_is_main_group(db, callback_query.message.chat.id):
            await callback_query.answer("⚠️ Thao tác này chỉ thực hiện được trong nhóm điều hành.", show_alert=True)
            return

        contract = db.query(Rental).filter(Rental.id == _ruid(contract_hex)).first()
        if not contract:
            await callback_query.answer("⚠️ Không tìm thấy hợp đồng.", show_alert=True)
            return

        if contract.status == RentalStatus.CANCELLED.value:
            await callback_query.answer("⚠️ Hợp đồng đã bị hủy từ trước.", show_alert=True)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        await _rk_safe_edit(
            callback_query,
            _build_rental_cancel_confirm_text(contract, customer),
            _rental_cancel_confirm_keyboard(contract),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rkh_cancel_contract_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


# --- Rental: Check Contract ---
@bot.on_message(filters.command(["rental_check_contract", "rental_kiem_tra_hop_dong"]) | filters.regex(r"^@\w+\s+/(rental_check_contract|rental_kiem_tra_hop_dong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(auto_delete_cmd=True)
async def rental_check_contract_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_check_contract", "rental_kiem_tra_hop_dong"])
    if args is None: return

    if len(args) < 2:
        db = SessionLocal()
        try:
            customers = _rkh_customers_for_chat(db, message.chat.id)
            if not customers:
                await message.reply_text("ℹ️ Chưa có khách hàng cho thuê nào trong dự án này.", parse_mode=ParseMode.HTML)
                return

            await message.reply_text(
                "<b>KIỂM TRA HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng cần kiểm tra:",
                reply_markup=_rkh_customer_list_keyboard(customers, 0),
                parse_mode=ParseMode.HTML
            )
        finally:
            db.close()
        return

    contract_code = args[1]
    db = SessionLocal()
    try:
        from app.models.telegram import TelegramProjectMember
        from app.models.rental import RentalCustomer, Rental, RentalStatus

        # Check if chat is a "main" group
        chat_id = str(message.chat.id)
        current_group = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()
        project_id = current_group.project_id

        # Get valid contacts for project isolation
        valid_members = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project_id,
            TelegramProjectMember.role == "member"
        ).all()

        valid_groups = []
        for m in valid_members:
            if m.group_name:
                valid_groups.append(m.group_name)

        contract = db.query(Rental).filter(Rental.contract_id == contract_code).first()
        if not contract:
            await message.reply_text(f"⚠️ Không tìm thấy hợp đồng cho thuê <b>{contract_code}</b>.", parse_mode=ParseMode.HTML)
            return

        # Find customer for this contract
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        if not customer or not customer.group_name or customer.group_name not in valid_groups:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> không thuộc về nhóm hợp lệ trong dự án hiện tại (Nhóm: {customer.group_name if customer else 'N/A'}).", parse_mode=ParseMode.HTML)
            return

        await _send_rental_contract_detail(message, db, contract, customer)

    except Exception as e:
        LogError(f"Error in rental_check_contract_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình truy xuất hợp đồng cho thuê.")
    finally:
        db.close()


# --- Rental Member: Xem Thông Tin Khách Hàng / Hợp Đồng (nhóm member) ---
_RMV_PAGE_SIZE = 10


def _rmv_customers(db, chat_id):
    """Khách hàng cho thuê gắn với nhóm member theo chat_id."""
    from app.models.rental import RentalCustomer
    return db.query(RentalCustomer).filter(
        RentalCustomer.chat_id == str(chat_id)
    ).order_by(RentalCustomer.customer_name).all()


def _rmv_owns(db, chat_id, customer) -> bool:
    """Khách hàng này có thuộc nhóm đang thao tác không."""
    if customer is None:
        return False
    return any(str(c.id) == str(customer.id) for c in _rmv_customers(db, chat_id))


def _rmv_contract_text(contract, customer) -> str:
    """Thông tin 1 hợp đồng cho thuê ở góc nhìn khách thuê."""
    lines = [
        f"<b>THÔNG TIN HỢP ĐỒNG CHO THUÊ</b>",
        f"Mã hợp đồng: <b>{contract.contract_id}</b>",
        f"Khách hàng: <b>{customer.customer_name if customer else 'N/A'}</b>",
        f"Trạng thái: <b>{_rental_status_label(contract.status)}</b>",
        f"Loại hợp đồng: {contract.type_contract or 'N/A'}",
        f"Mã bất động sản: {contract.real_estate_id or 'N/A'}",
        f"{'━' * 15}",
        f"Tiền thuê / tháng: <b>{_rental_fmt_num(contract.monthly_rental):,}</b>",
        f"Tiền cọc: <b>{_rental_fmt_num(contract.deposit):,}</b>",
        f"Tổng tiền cần thanh toán: <b>{_rental_fmt_num(contract.rental_debt or 0):,}</b>",
        f"{'━' * 15}",
        f"Ngày bắt đầu thuê: {_rental_fmt_dt(contract.start_rental)}",
        f"Ngày kết thúc thuê: {_rental_fmt_dt(contract.end_rental)}",
    ]
    return "\n".join(lines)


def _rmv_customer_pick_keyboard(customers, action_prefix):
    """Bàn phím chọn khách hàng khi một nhóm gắn với nhiều khách hàng."""
    buttons = [
        [InlineKeyboardButton(
            f"{c.customer_id} - {c.customer_name}",
            callback_data=f"{action_prefix}|{_rsid(c.id)}|0"
        )]
        for c in customers
    ]
    buttons.append([InlineKeyboardButton("Hủy", callback_data="rmv_x")])
    return InlineKeyboardMarkup(buttons)


def _rmv_contract_list_keyboard(contracts, page, customer_hex, page_prefix, item_prefix):
    """Bàn phím danh sách hợp đồng dạng nút, có phân trang và nút Hủy."""
    total = len(contracts)
    total_pages = max(1, (total + _RMV_PAGE_SIZE - 1) // _RMV_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * _RMV_PAGE_SIZE
    page_contracts = contracts[start:min(start + _RMV_PAGE_SIZE, total)]

    buttons = [
        [InlineKeyboardButton(
            f"{c.contract_id} ({_rental_status_label(c.status)})",
            callback_data=f"{item_prefix}|{_rsid(c.id)}"
        )]
        for c in page_contracts
    ]

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"{page_prefix}|{customer_hex}|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rmv_n"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"{page_prefix}|{customer_hex}|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rmv_x")])
    return InlineKeyboardMarkup(buttons)


def _rmv_contracts_of(db, customer):
    from app.models.rental import Rental
    return db.query(Rental).filter(Rental.customer_id == customer.id).order_by(Rental.contract_id).all()


async def _rmv_send_customer_info(db, customer, send):
    """Hiển thị thông tin khách hàng kèm 2 nút: Xem hợp đồng, Hủy."""
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Xem hợp đồng", callback_data=f"rmv_hd|{_rsid(customer.id)}|0")],
        [InlineKeyboardButton("Hủy", callback_data="rmv_x")]
    ])
    await send(_build_rental_customer_info_text(db, customer), reply_markup=keyboard, parse_mode=ParseMode.HTML)


@bot.on_callback_query(filters.regex(r"^rmv_n$"))
@command_timeout(auto_delete_cmd=True)
async def rmv_noop_callback(client, callback_query: CallbackQuery):
    await callback_query.answer()


@bot.on_callback_query(filters.regex(r"^rmv_x$"))
@command_timeout(auto_delete_cmd=True)
async def rmv_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


@bot.on_message(filters.command(["rental_member_check_customer", "rental_kiem_tra_tt_khach_hang"]) | filters.regex(r"^@\w+\s+/(rental_member_check_customer|rental_kiem_tra_tt_khach_hang)\b"))
@require_user_type(UserType.MEMBER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("member")
@command_timeout(auto_delete_cmd=True)
async def rental_member_check_customer_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_member_check_customer", "rental_kiem_tra_tt_khach_hang"])
    if args is None: return

    db = SessionLocal()
    try:
        customers = _rmv_customers(db, message.chat.id)
        if not customers:
            await message.reply_text("ℹ️ Nhóm này chưa được gắn với khách hàng cho thuê nào.")
            return

        if len(customers) > 1:
            await message.reply_text(
                "<b>THÔNG TIN KHÁCH HÀNG CHO THUÊ</b>\n\nChọn khách hàng để xem thông tin:",
                reply_markup=_rmv_customer_pick_keyboard(customers, "rmv_kh"),
                parse_mode=ParseMode.HTML
            )
            return

        await _rmv_send_customer_info(db, customers[0], message.reply_text)
    except Exception as e:
        LogError(f"Error in rental_member_check_customer_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra khi truy xuất thông tin khách hàng.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rmv_kh\|([a-f0-9]{32})\|(\d+)$"))
@require_group_role("member")
@command_timeout(auto_delete_cmd=True)
async def rmv_customer_info_callback(client, callback_query: CallbackQuery):
    """Hiển thị thông tin khách hàng đã chọn."""
    customer_uuid = _ruid(callback_query.matches[0].group(1))
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == customer_uuid).first()
        if not _rmv_owns(db, callback_query.message.chat.id, customer):
            await callback_query.answer("⚠️ Không tìm thấy khách hàng của nhóm này.", show_alert=True)
            return

        await _rmv_send_customer_info(db, customer, callback_query.message.edit_text)
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rmv_customer_info_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_message(filters.command(["rental_member_check_contract", "rental_kiem_tra_tt_hop_dong"]) | filters.regex(r"^@\w+\s+/(rental_member_check_contract|rental_kiem_tra_tt_hop_dong)\b"))
@require_user_type(UserType.MEMBER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("member")
@command_timeout(auto_delete_cmd=True)
async def rental_member_check_contract_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_member_check_contract", "rental_kiem_tra_tt_hop_dong"])
    if args is None: return

    db = SessionLocal()
    try:
        customers = _rmv_customers(db, message.chat.id)
        if not customers:
            await message.reply_text("ℹ️ Nhóm này chưa được gắn với khách hàng cho thuê nào.")
            return

        if len(customers) > 1:
            await message.reply_text(
                "<b>DANH SÁCH HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng để xem danh sách hợp đồng:",
                reply_markup=_rmv_customer_pick_keyboard(customers, "rmv_ds"),
                parse_mode=ParseMode.HTML
            )
            return

        customer = customers[0]
        contracts = _rmv_contracts_of(db, customer)
        if not contracts:
            await message.reply_text("ℹ️ Khách hàng chưa có hợp đồng cho thuê nào.")
            return

        await message.reply_text(
            f"<b>DANH SÁCH HỢP ĐỒNG CHO THUÊ</b>\n"
            f"Khách hàng: <b>{customer.customer_name}</b> ({customer.customer_id or 'N/A'})\n\n"
            f"Chọn hợp đồng để xem thông tin:",
            reply_markup=_rmv_contract_list_keyboard(contracts, 0, _rsid(customer.id), "rmv_ds", "rmv_dsc"),
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        LogError(f"Error in rental_member_check_contract_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra khi truy xuất danh sách hợp đồng.")
    finally:
        db.close()


async def _rmv_show_contract_list(callback_query, page_prefix, item_prefix):
    """Danh sách hợp đồng của khách hàng (dùng chung cho 2 luồng rmv_hd / rmv_ds)."""
    customer_uuid = _ruid(callback_query.matches[0].group(1))
    page = int(callback_query.matches[0].group(2))
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == customer_uuid).first()
        if not _rmv_owns(db, callback_query.message.chat.id, customer):
            await callback_query.answer("⚠️ Không tìm thấy khách hàng của nhóm này.", show_alert=True)
            return

        contracts = _rmv_contracts_of(db, customer)
        if not contracts:
            await callback_query.answer("ℹ️ Khách hàng chưa có hợp đồng cho thuê nào.", show_alert=True)
            return

        await _rk_safe_edit(
            callback_query,
            f"<b>DANH SÁCH HỢP ĐỒNG CHO THUÊ</b>\n"
            f"Khách hàng: <b>{customer.customer_name}</b> ({customer.customer_id or 'N/A'})\n\n"
            f"Chọn hợp đồng để xem thông tin:",
            _rmv_contract_list_keyboard(contracts, page, _rsid(customer.id), page_prefix, item_prefix),
        )
        await callback_query.answer()
    finally:
        db.close()


async def _rmv_show_contract_detail(callback_query):
    """Thông tin chi tiết 1 hợp đồng (dùng chung cho rmv_hdc / rmv_dsc)."""
    contract_uuid = _ruid(callback_query.matches[0].group(1))
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental
        contract = db.query(Rental).filter(Rental.id == contract_uuid).first()
        customer = db.query(RentalCustomer).filter(
            RentalCustomer.id == contract.customer_id
        ).first() if contract else None

        if not _rmv_owns(db, callback_query.message.chat.id, customer):
            await callback_query.answer("⚠️ Không tìm thấy hợp đồng của nhóm này.", show_alert=True)
            return

        await _rk_safe_edit(
            callback_query,
            _rmv_contract_text(contract, customer),
            InlineKeyboardMarkup([[InlineKeyboardButton("Hủy", callback_data="rmv_x")]]),
        )
        await callback_query.answer()
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rmv_hd\|([a-f0-9]{32})\|(\d+)$"))
@require_group_role("member")
@command_timeout(auto_delete_cmd=True)
async def rmv_contract_list_callback(client, callback_query: CallbackQuery):
    """Nút 'Xem hợp đồng' trong luồng xem thông tin khách hàng."""
    try:
        await _rmv_show_contract_list(callback_query, "rmv_hd", "rmv_hdc")
    except Exception as e:
        LogError(f"Error in rmv_contract_list_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)


@bot.on_callback_query(filters.regex(r"^rmv_hdc\|([a-f0-9]{32})$"))
@require_group_role("member")
@command_timeout(auto_delete_cmd=True)
async def rmv_contract_detail_callback(client, callback_query: CallbackQuery):
    """Thông tin hợp đồng (luồng xem thông tin khách hàng)."""
    try:
        await _rmv_show_contract_detail(callback_query)
    except Exception as e:
        LogError(f"Error in rmv_contract_detail_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)


@bot.on_callback_query(filters.regex(r"^rmv_ds\|([a-f0-9]{32})\|(\d+)$"))
@require_group_role("member")
@command_timeout(auto_delete_cmd=True)
async def rmv_contract_list_ds_callback(client, callback_query: CallbackQuery):
    """Danh sách hợp đồng trong luồng /rental_kiem_tra_tt_hop_dong."""
    try:
        await _rmv_show_contract_list(callback_query, "rmv_ds", "rmv_dsc")
    except Exception as e:
        LogError(f"Error in rmv_contract_list_ds_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)


@bot.on_callback_query(filters.regex(r"^rmv_dsc\|([a-f0-9]{32})$"))
@require_group_role("member")
@command_timeout(auto_delete_cmd=True)
async def rmv_contract_detail_ds_callback(client, callback_query: CallbackQuery):
    """Thông tin hợp đồng (luồng /rental_kiem_tra_tt_hop_dong)."""
    try:
        await _rmv_show_contract_detail(callback_query)
    except Exception as e:
        LogError(f"Error in rmv_contract_detail_ds_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)


# --- Rental: Check Debt (Xem Tổng Tiền Thanh Toán HD Hiện Tại) ---
@bot.on_message(filters.command(["rental_check_debt", "rental_xem_cong_no"]) | filters.regex(r"^@\w+\s+/(rental_check_debt|rental_xem_cong_no)\b"))
@require_user_type(UserType.MEMBER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("member")
@command_timeout(auto_delete_cmd=True)
async def rental_check_debt_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_check_debt", "rental_xem_cong_no"])
    if args is None: return

    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental, RentalStatus

        if len(args) >= 2:
            # Tìm khách hàng theo Mã Khách Hàng được nhập
            customer_id_input = args[1]
            customer = db.query(RentalCustomer).filter(RentalCustomer.customer_id == customer_id_input).first()
            if not customer:
                await message.reply_text(f"ℹ️ Không tìm thấy thông tin khách hàng cho Mã <b>{customer_id_input}</b>.", parse_mode=ParseMode.HTML)
                return
        else:
            # Không nhập mã -> xác định khách hàng theo chat_id của chính nhóm gửi lệnh
            customers = db.query(RentalCustomer).filter(
                RentalCustomer.chat_id == str(message.chat.id)
            ).order_by(RentalCustomer.customer_name).all()

            if not customers:
                await message.reply_text("⚠️ Nhóm này chưa được gắn với khách hàng cho thuê nào. Vui lòng cung cấp Mã Khách Hàng. Lệnh ví dụ: <code>/rental_check_debt KH001</code>", parse_mode=ParseMode.HTML)
                return

            if len(customers) > 1:
                ds = "\n".join([f"- <code>{c.customer_id}</code>: {c.customer_name}" for c in customers])
                await message.reply_text(
                    f"⚠️ Nhóm này đang gắn với nhiều khách hàng. Vui lòng cung cấp Mã Khách Hàng:\n{ds}",
                    parse_mode=ParseMode.HTML
                )
                return

            customer = customers[0]

        # Lấy các hợp đồng đang thuê / hết hạn (chưa hủy)
        active_rentals = db.query(Rental).filter(
            Rental.customer_id == customer.id,
            Rental.status.in_([RentalStatus.ACTIVE.value, RentalStatus.EXPIRED.value])
        ).all()

        if not active_rentals:
            await message.reply_text(f"ℹ️ <b>{customer.customer_name}</b>, bạn hiện không có hợp đồng cho thuê nào.", parse_mode=ParseMode.HTML)
            return

        def fmt_num(val):
            if val is None: return 0
            return int(val) if val == int(val) else val

        def fmt_dt(dt):
            if not dt: return "N/A"
            return dt.strftime('%d/%m/%Y')

        total_debt = sum([(r.rental_debt or 0) for r in active_rentals])
        total_monthly = sum([(r.monthly_rental or 0) for r in active_rentals])

        contract_lines = []
        for idx, r in enumerate(active_rentals, 1):
            status_label = "Đang thuê" if r.status == RentalStatus.ACTIVE.value else "Hết hạn"
            contract_lines.append(
                f"{idx}. <b>{r.contract_id}</b> ({r.type_contract or 'N/A'}) - {status_label}\n"
                f"   Tiền thuê: <b>{fmt_num(r.monthly_rental):,}</b>/tháng\n"
                f"   Tổng tiền cần thanh toán: <b>{fmt_num(r.rental_debt or 0):,}</b>\n"
                f"   BĐS: {r.real_estate_id or 'N/A'}\n"
                f"   Thuê: {fmt_dt(r.start_rental)} - {fmt_dt(r.end_rental)}"
            )

        reply_lines = [
            f"<b>TỔNG TIỀN THANH TOÁN HD HIỆN TẠI</b>",
            f"Khách hàng: <b>{customer.customer_name}</b> (Mã: {customer.customer_id})",
            f"{'━' * 15}",
            f"Tổng hợp đồng: <b>{len(active_rentals)}</b>",
            f"Tổng tiền thuê: <b>{fmt_num(total_monthly):,}</b>/tháng",
            f"Tổng tiền cần thanh toán: <b>{fmt_num(total_debt):,}</b>",
            f"{'━' * 15}",
        ] + contract_lines

        await message.reply_text("\n".join(reply_lines), parse_mode=ParseMode.HTML)
        LogInfo(f"[RentalCheckDebt] {customer.customer_id} checked their debt: {len(active_rentals)} contracts, total: {total_debt}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error in rental_check_debt_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra khi truy xuất thông tin.")
    finally:
        db.close()


# --- Rental: Extend Contract (Interactive) ---
def _build_rental_extend_contract_form(contract, customer) -> str:
    """Form gia hạn hợp đồng cho thuê."""
    return f"""<b>FORM GIA HẠN HỢP ĐỒNG CHO THUÊ</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/rental_extend_contract {contract.contract_id}
Mã Hợp Đồng: {contract.contract_id or ""}
Tên Khách Hàng: {customer.customer_name if customer else ""}
Trạng Thái: {_rental_status_label(contract.status)}
Loại Hợp Đồng: {contract.type_contract or ""}
Ngày Kết Thúc Thuê Hiện Tại: {_rental_fmt_dt(contract.end_rental)}
Số Tháng Gia Hạn: 1
</pre>

<i>Số Tháng Gia Hạn là số nguyên từ 1 đến 60.
Nếu hợp đồng đang Hết hạn, bot sẽ tự động đưa về Đang thuê sau khi gia hạn.</i>"""


def _rgh_customer_list_keyboard(customers, page):
    """Bàn phím chọn khách hàng để gia hạn hợp đồng: tối đa 10 nút/trang."""
    PAGE_SIZE = 10
    total = len(customers)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_customers = customers[start:min(start + PAGE_SIZE, total)]

    buttons = [
        [InlineKeyboardButton(
            f"{c.customer_id} - {c.customer_name}",
            callback_data=f"rgh_c|{_rsid(c.id)}|0"
        )]
        for c in page_customers
    ]

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rgh_cp|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rgh_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rgh_cp|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rgh_x")])
    return InlineKeyboardMarkup(buttons)


def _rgh_contract_list_keyboard(contracts, page, customer_hex):
    """Bàn phím chọn hợp đồng cần gia hạn: tối đa 10 nút/trang."""
    PAGE_SIZE = 10
    total = len(contracts)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_contracts = contracts[start:min(start + PAGE_SIZE, total)]

    buttons = [
        [InlineKeyboardButton(
            f"{c.contract_id} ({_rental_status_label(c.status)})",
            callback_data=f"rgh_s|{_rsid(c.id)}"
        )]
        for c in page_contracts
    ]

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rgh_c|{customer_hex}|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rgh_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rgh_c|{customer_hex}|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rgh_x")])
    return InlineKeyboardMarkup(buttons)


@bot.on_callback_query(filters.regex(r"^rgh_noop$"))
@command_timeout(auto_delete_cmd=True)
async def rgh_noop_callback(client, callback_query: CallbackQuery):
    await callback_query.answer()


@bot.on_callback_query(filters.regex(r"^rgh_x$"))
@command_timeout(auto_delete_cmd=True)
async def rgh_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


@bot.on_callback_query(filters.regex(r"^rgh_cp\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rgh_customer_page_callback(client, callback_query: CallbackQuery):
    """Phân trang danh sách khách hàng khi gia hạn hợp đồng."""
    page = int(callback_query.matches[0].group(1))
    db = SessionLocal()
    try:
        customers = _rental_customers_of_chat(db, callback_query.message.chat.id)
        if not customers:
            await _rk_safe_edit(callback_query, "ℹ️ Chưa có khách hàng cho thuê nào trong dự án này.")
            return

        await _rk_safe_edit(
            callback_query,
            "<b>GIA HẠN HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng cần gia hạn hợp đồng:",
            _rgh_customer_list_keyboard(customers, page),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rgh_customer_page_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rgh_c\|([a-f0-9]{32})\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rgh_contract_list_callback(client, callback_query: CallbackQuery):
    """Chọn khách hàng -> hiển thị danh sách hợp đồng có thể gia hạn."""
    customer_hex = callback_query.matches[0].group(1)
    page = int(callback_query.matches[0].group(2))
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental, RentalStatus
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == _ruid(customer_hex)).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng.", show_alert=True)
            return

        # Hợp đồng đã hủy không gia hạn được -> không đưa vào danh sách
        contracts = db.query(Rental).filter(
            Rental.customer_id == customer.id,
            Rental.status != RentalStatus.CANCELLED.value
        ).all()
        if not contracts:
            await callback_query.answer("ℹ️ Khách hàng chưa có hợp đồng nào có thể gia hạn.", show_alert=True)
            return

        await _rk_safe_edit(
            callback_query,
            f"<b>GIA HẠN HỢP ĐỒNG CHO THUÊ</b>\n"
            f"Khách hàng: <b>{customer.customer_name}</b> ({customer.customer_id or 'N/A'})\n\n"
            f"Chọn hợp đồng cần gia hạn:",
            _rgh_contract_list_keyboard(contracts, page, customer_hex),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rgh_contract_list_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rgh_s\|([a-f0-9]{32})$"))
@command_timeout(timeout_seconds=600, auto_delete_cmd=True)  # Form cần thời gian điền
async def rgh_select_contract_callback(client, callback_query: CallbackQuery):
    """Chọn hợp đồng -> hiển thị form gia hạn."""
    contract_hex = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental, RentalStatus
        contract = db.query(Rental).filter(Rental.id == _ruid(contract_hex)).first()
        if not contract:
            await callback_query.answer("⚠️ Không tìm thấy hợp đồng.", show_alert=True)
            return

        if contract.status == RentalStatus.CANCELLED.value:
            await callback_query.answer("⚠️ Hợp đồng đã bị hủy, không thể gia hạn.", show_alert=True)
            return

        if not contract.end_rental:
            await callback_query.answer("⚠️ Hợp đồng chưa có Ngày Kết Thúc Thuê nên không thể gia hạn.", show_alert=True)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()

        form_msg = await callback_query.message.reply_text(
            _build_rental_extend_contract_form(contract, customer), parse_mode=ParseMode.HTML
        )
        form_tracker.track(callback_query.message.chat.id, "rental_extend_contract", contract.contract_id, form_msg.id)
        await callback_query.message.delete()
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rgh_select_contract_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


# --- Rental: Extend Contract ---
@bot.on_message(filters.command(["rental_extend_contract", "rental_gia_han_hop_dong"]) | filters.regex(r"^@\w+\s+/(rental_extend_contract|rental_gia_han_hop_dong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(auto_delete_cmd=True)
async def rental_extend_contract_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_extend_contract", "rental_gia_han_hop_dong"])
    if args is None: return

    lines = message.text.strip().split("\n")
    is_form = len(lines) > 1

    # Không kèm Mã Hợp Đồng -> hiển thị danh sách khách hàng để chọn
    if not is_form and len(args) < 2:
        db = SessionLocal()
        try:
            customers = _rental_customers_of_chat(db, message.chat.id)
            if not customers:
                await message.reply_text("ℹ️ Chưa có khách hàng cho thuê nào trong dự án này.", parse_mode=ParseMode.HTML)
                return

            await message.reply_text(
                "<b>GIA HẠN HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng cần gia hạn hợp đồng:",
                reply_markup=_rgh_customer_list_keyboard(customers, 0),
                parse_mode=ParseMode.HTML
            )
        finally:
            db.close()
        return

    if is_form:
        # Người dùng gửi lại Form gia hạn
        data = {}
        for line in lines[1:]:
            if ":" in line:
                key, val = line.split(":", 1)
                data[key.strip()] = val.strip()

        # Mã HĐ lấy từ chính dòng lệnh (dòng đầu), không có mới lấy trong Form
        first_line_parts = lines[0].split()
        if first_line_parts and first_line_parts[0].startswith("@"):
            first_line_parts = first_line_parts[1:]
        code_from_cmd = first_line_parts[1].strip() if len(first_line_parts) > 1 else ""

        contract_code = (code_from_cmd or data.get("Mã Hợp Đồng", "")).strip()
        months_str = data.get("Số Tháng Gia Hạn", "1").strip()
    else:
        contract_code = args[1].strip()
        months_str = args[2].strip() if len(args) >= 3 else "1"

    if not contract_code:
        await message.reply_text("⚠️ Vui lòng cung cấp <b>Mã Hợp Đồng</b>.\nCú pháp: <code>/rental_extend_contract [Mã HĐ] [Số tháng]</code>", parse_mode=ParseMode.HTML)
        return

    try:
        months_to_add = int(months_str)
        if months_to_add <= 0 or months_to_add > 60:
            raise ValueError()
    except ValueError:
        await message.reply_text("⚠️ Số tháng gia hạn không hợp lệ (phải là số nguyên từ 1 đến 60).", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        from app.models.telegram import TelegramProjectMember
        from app.models.rental import RentalCustomer, Rental, RentalStatus


        contract = db.query(Rental).filter(Rental.contract_id == contract_code).first()
        if not contract:
            await message.reply_text(f"⚠️ Không tìm thấy hợp đồng cho thuê nào có mã <b>{contract_code}</b>.", parse_mode=ParseMode.HTML)
            return

        if contract.status == RentalStatus.CANCELLED.value:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> đã bị hủy, không thể gia hạn.", parse_mode=ParseMode.HTML)
            return

        if not contract.end_rental:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> chưa có ngày kết thúc thuê, không thể gia hạn.", parse_mode=ParseMode.HTML)
            return

        def add_months_to_date(source_date, months):
            import calendar
            month = source_date.month - 1 + months
            year = source_date.year + month // 12
            month = month % 12 + 1
            day = min(source_date.day, calendar.monthrange(year, month)[1])
            return datetime.date(year, month, day)

        new_end_rental = add_months_to_date(contract.end_rental, months_to_add)

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()

        text = (
            f"<b>THÔNG TIN GIA HẠN HỢP ĐỒNG CHO THUÊ</b>\n\n"
            f"- Tên Khách Hàng: <b>{customer.customer_name if customer else 'N/A'}</b>\n"
            f"- Mã Hợp Đồng: <b>{contract.contract_id}</b>\n"
            f"- Loại Hợp Đồng: <b>{contract.type_contract or 'N/A'}</b>\n"
            f"- Thời gian gia hạn thêm: <b>{months_to_add} Tháng</b>\n\n"
            f"- Ngày kết thúc thuê: <b>{contract.end_rental.strftime('%d/%m/%Y')}</b>\n"
            f"- Ngày kết thúc thuê mới: <b>{new_end_rental.strftime('%d/%m/%Y')}</b>\n\n"
            f"<i>Lưu ý: Nếu hợp đồng đang Hết hạn, bot sẽ tự động đưa về Đang thuê.</i>"
        )

        buttons = [
            [
                InlineKeyboardButton("Xác nhận gia hạn", callback_data=f"rec_confirm_{contract.id}_{months_to_add}"),
                InlineKeyboardButton("Hủy", callback_data="rec_cancel")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(buttons)
        await message.reply_text(text, reply_markup=reply_markup, parse_mode=ParseMode.HTML)

        # Xóa Form mẫu sau khi người dùng đã gửi lại thông tin
        form_msg_id = form_tracker.pop(message.chat.id, "rental_extend_contract", contract.contract_id)
        if form_msg_id:
            try:
                await client.delete_messages(chat_id=message.chat.id, message_ids=form_msg_id)
            except Exception as del_err:
                LogError(f"Failed to delete rental extend contract form: {del_err}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error in rental_extend_contract_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình xử lý.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rec_confirm_([^_]+)_(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rental_extend_contract_confirm_callback(client, callback_query: CallbackQuery):
    contract_uuid = callback_query.matches[0].group(1)
    months_to_add = int(callback_query.matches[0].group(2))
    db = SessionLocal()
    try:
        from app.models.rental import Rental, RentalCustomer, RentalStatus
        from app.models.telegram import TelegramProjectMember

        contract = db.query(Rental).filter(Rental.id == contract_uuid).first()
        if not contract:
            await callback_query.message.edit_text("⚠️ Hợp đồng không tồn tại hoặc đã bị xóa.")
            return

        def add_months_to_date(source_date, months):
            import calendar
            month = source_date.month - 1 + months
            year = source_date.year + month // 12
            month = month % 12 + 1
            day = min(source_date.day, calendar.monthrange(year, month)[1])
            return datetime.date(year, month, day)

        new_end_rental = add_months_to_date(contract.end_rental, months_to_add)
        old_end_rental_str = contract.end_rental.strftime('%d/%m/%Y')
        new_end_rental_str = new_end_rental.strftime('%d/%m/%Y')

        contract.end_rental = new_end_rental

        # Reset expired back to active
        if contract.status == RentalStatus.EXPIRED.value:
            contract.status = RentalStatus.ACTIVE.value

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()

        # Find member group for notification (ưu tiên chat_id của khách hàng, sau đó tới Tên Nhóm)
        member_chat_id = None

        if customer:
            from app.crud.rental import match_member_link
            customer_links = db.query(TelegramProjectMember).filter(
                TelegramProjectMember.role == "member"
            ).all()
            matched_link = match_member_link(customer, customer_links)
            if matched_link:
                member_chat_id = matched_link.chat_id

        db.commit()

        success_text = (
            f"✅ <b>Gia hạn thành công hợp đồng cho thuê {contract.contract_id}</b>\n\n"
            f"- Khách hàng: <b>{customer.customer_name if customer else 'N/A'}</b>\n"
            f"- Phương thức: <b>Gia hạn {months_to_add} tháng</b>\n"
            f"- Ngày kết thúc thuê mới: <b>{new_end_rental_str}</b> (cũ: {old_end_rental_str})\n"
        )

        await callback_query.message.edit_text(success_text, parse_mode=ParseMode.HTML)
        LogInfo(f"[RentalExtendContract] User {callback_query.from_user.id} extended rental contract {contract.contract_id}", LogType.SYSTEM_STATUS)

        # Send Member Notification
        if member_chat_id:
            member_alert = (
                f"🔔 <b>THÔNG BÁO TỪ QUẢN TRỊ VIÊN</b>\n\n"
                f"Hợp đồng cho thuê mã <b>{contract.contract_id}</b> của Quý khách đã được gia hạn thành công thêm <b>{months_to_add} tháng</b>.\n\n"
                f"<b>Ngày kết thúc thuê cập nhật mới nhất: {new_end_rental_str}</b>\n"
                f"Quý khách vui lòng lưu ý thời gian kết thúc mới. Xin cảm ơn!"
            )
            try:
                await client.send_message(
                    chat_id=int(member_chat_id),
                    text=member_alert,
                    parse_mode=ParseMode.HTML
                )
            except Exception as e:
                LogError(f"Failed to send rental extend alert to member group {member_chat_id}: {e}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in rec_confirm callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi xác nhận gia hạn.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rec_cancel$"))
@command_timeout(auto_delete_cmd=True)
async def rental_extend_contract_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


# --- Rental: Cancel Contract (Interactive) ---
def _rhh_customer_list_keyboard(customers, page):
    """Bàn phím chọn khách hàng để hủy hợp đồng: tối đa 10 nút/trang."""
    PAGE_SIZE = 10
    total = len(customers)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_customers = customers[start:min(start + PAGE_SIZE, total)]

    buttons = [
        [InlineKeyboardButton(
            f"{c.customer_id} - {c.customer_name}",
            callback_data=f"rhh_c|{_rsid(c.id)}|0"
        )]
        for c in page_customers
    ]

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rhh_cp|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rhh_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rhh_cp|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rhh_x")])
    return InlineKeyboardMarkup(buttons)


def _rhh_contract_list_keyboard(contracts, page, customer_hex):
    """Bàn phím chọn hợp đồng cần hủy: tối đa 10 nút/trang."""
    PAGE_SIZE = 10
    total = len(contracts)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_contracts = contracts[start:min(start + PAGE_SIZE, total)]

    buttons = [
        [InlineKeyboardButton(
            f"{c.contract_id} ({_rental_status_label(c.status)})",
            callback_data=f"rhh_s|{_rsid(c.id)}"
        )]
        for c in page_contracts
    ]

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rhh_c|{customer_hex}|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rhh_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rhh_c|{customer_hex}|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rhh_x")])
    return InlineKeyboardMarkup(buttons)


@bot.on_callback_query(filters.regex(r"^rhh_noop$"))
@command_timeout(auto_delete_cmd=True)
async def rhh_noop_callback(client, callback_query: CallbackQuery):
    await callback_query.answer()


@bot.on_callback_query(filters.regex(r"^rhh_x$"))
@command_timeout(auto_delete_cmd=True)
async def rhh_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


@bot.on_callback_query(filters.regex(r"^rhh_cp\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rhh_customer_page_callback(client, callback_query: CallbackQuery):
    """Phân trang danh sách khách hàng khi hủy hợp đồng."""
    page = int(callback_query.matches[0].group(1))
    db = SessionLocal()
    try:
        customers = _rental_customers_of_chat(db, callback_query.message.chat.id)
        if not customers:
            await _rk_safe_edit(callback_query, "ℹ️ Chưa có khách hàng cho thuê nào trong dự án này.")
            return

        await _rk_safe_edit(
            callback_query,
            "<b>HỦY HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng có hợp đồng cần hủy:",
            _rhh_customer_list_keyboard(customers, page),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rhh_customer_page_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rhh_c\|([a-f0-9]{32})\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rhh_contract_list_callback(client, callback_query: CallbackQuery):
    """Chọn khách hàng -> hiển thị danh sách hợp đồng có thể hủy."""
    customer_hex = callback_query.matches[0].group(1)
    page = int(callback_query.matches[0].group(2))
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental, RentalStatus
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == _ruid(customer_hex)).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng.", show_alert=True)
            return

        # Hợp đồng đã hủy thì không hiện lại trong danh sách
        contracts = db.query(Rental).filter(
            Rental.customer_id == customer.id,
            Rental.status != RentalStatus.CANCELLED.value
        ).all()
        if not contracts:
            await callback_query.answer("ℹ️ Khách hàng không có hợp đồng nào cần hủy.", show_alert=True)
            return

        await _rk_safe_edit(
            callback_query,
            f"<b>HỦY HỢP ĐỒNG CHO THUÊ</b>\n"
            f"Khách hàng: <b>{customer.customer_name}</b> ({customer.customer_id or 'N/A'})\n\n"
            f"Chọn hợp đồng cần hủy:",
            _rhh_contract_list_keyboard(contracts, page, customer_hex),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rhh_contract_list_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rhh_s\|([a-f0-9]{32})$"))
@command_timeout(auto_delete_cmd=True)
async def rhh_select_contract_callback(client, callback_query: CallbackQuery):
    """Chọn hợp đồng -> hiển thị thông báo xác nhận hủy."""
    contract_hex = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental, RentalStatus
        contract = db.query(Rental).filter(Rental.id == _ruid(contract_hex)).first()
        if not contract:
            await callback_query.answer("⚠️ Không tìm thấy hợp đồng.", show_alert=True)
            return

        if contract.status == RentalStatus.CANCELLED.value:
            await callback_query.answer("⚠️ Hợp đồng đã bị hủy từ trước.", show_alert=True)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        await _rk_safe_edit(
            callback_query,
            _build_rental_cancel_confirm_text(contract, customer),
            _rental_cancel_confirm_keyboard(contract),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rhh_select_contract_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


# --- Rental: Cancel Contract ---
@bot.on_message(filters.command(["rental_cancel_contract", "rental_huy_hop_dong"]) | filters.regex(r"^@\w+\s+/(rental_cancel_contract|rental_huy_hop_dong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(auto_delete_cmd=True)
async def rental_cancel_contract_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_cancel_contract", "rental_huy_hop_dong"])
    if args is None: return

    if len(args) < 2:
        db = SessionLocal()
        try:
            customers = _rental_customers_of_chat(db, message.chat.id)
            if not customers:
                await message.reply_text("ℹ️ Chưa có khách hàng cho thuê nào trong dự án này.", parse_mode=ParseMode.HTML)
                return

            await message.reply_text(
                "<b>HỦY HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng có hợp đồng cần hủy:",
                reply_markup=_rhh_customer_list_keyboard(customers, 0),
                parse_mode=ParseMode.HTML
            )
        finally:
            db.close()
        return

    contract_code = args[1]
    chat_id = str(message.chat.id)
    db = SessionLocal()
    try:
        from app.models.telegram import TelegramProjectMember
        from app.models.rental import RentalCustomer, Rental, RentalStatus

        current_group = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()
        project_id = current_group.project_id

        # Get valid contacts for project isolation
        valid_members = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project_id,
            TelegramProjectMember.role == "member"
        ).all()

        valid_groups = []
        for m in valid_members:
            if m.group_name:
                valid_groups.append(m.group_name)

        contract = db.query(Rental).filter(Rental.contract_id == contract_code).first()
        if not contract:
            await message.reply_text(f"⚠️ Không tìm thấy hợp đồng cho thuê nào có mã <b>{contract_code}</b>.", parse_mode=ParseMode.HTML)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        if not customer or not customer.group_name or customer.group_name not in valid_groups:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> không thuộc về nhóm hợp lệ nào trong dự án hiện tại.", parse_mode=ParseMode.HTML)
            return

        if contract.status == RentalStatus.CANCELLED.value:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> đã bị hủy từ trước.", parse_mode=ParseMode.HTML)
            return

        await message.reply_text(
            _build_rental_cancel_confirm_text(contract, customer),
            reply_markup=_rental_cancel_confirm_keyboard(contract),
            parse_mode=ParseMode.HTML
        )

    except Exception as e:
        LogError(f"Error in rental_cancel_contract_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình xử lý.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rcc_confirm_(.*)$"))
@command_timeout(auto_delete_cmd=True)
async def rental_cancel_contract_confirm_callback(client, callback_query: CallbackQuery):
    contract_uuid = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import Rental, RentalStatus

        # Chốt lại 1 lần nữa ở bước thực thi: chỉ nhóm main mới được hủy hợp đồng
        if not _rental_is_main_group(db, callback_query.message.chat.id):
            await callback_query.answer("⚠️ Thao tác này chỉ thực hiện được trong nhóm điều hành.", show_alert=True)
            return

        contract = db.query(Rental).filter(Rental.id == contract_uuid).first()
        if not contract:
            await callback_query.message.edit_text("⚠️ Hợp đồng không tồn tại hoặc đã bị xóa.")
            return

        if contract.status == RentalStatus.CANCELLED.value:
            await callback_query.message.edit_text(f"⚠️ Hợp đồng <b>{contract.contract_id}</b> đã bị hủy từ trước.", parse_mode=ParseMode.HTML)
            return

        contract.status = RentalStatus.CANCELLED.value
        db.commit()

        await callback_query.message.edit_text(f"✅ Đã hủy hợp đồng cho thuê <b>{contract.contract_id}</b> thành công.", parse_mode=ParseMode.HTML)
        LogInfo(f"[RentalCancelContract] Cancelled rental contract {contract.contract_id} by user {callback_query.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in rcc_confirm callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi hủy hợp đồng.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rcc_exit$"))
@command_timeout(auto_delete_cmd=True)
async def rental_cancel_contract_exit_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


# --- Rental: Nút thao tác trên THÔNG BÁO ĐÓNG TIỀN THUÊ ---
def _rental_skip_tag(year: int, month: int) -> str:
    """Cờ đánh dấu bỏ nhắc đóng tiền thuê của 1 tháng (nút 'Lưu sổ')."""
    return f"[SKIP_RENTAL: {month:02d}/{year}]"


def _rental_payment_notify_keyboard(contract_id: str):
    """4 nút thao tác gắn kèm THÔNG BÁO ĐÓNG TIỀN THUÊ."""
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("Đã nhận TT đủ", callback_data=f"rnt_full_pay|{contract_id}"),
            InlineKeyboardButton("Đã nhận TT", callback_data=f"rnt_pay|{contract_id}")
        ],
        [
            InlineKeyboardButton("Lưu sổ", callback_data=f"rnt_remind|{contract_id}"),
            InlineKeyboardButton("Nợ Xấu", callback_data=f"rnt_bad|{contract_id}")
        ]
    ])


async def _check_admin_or_owner_rental(callback_query: CallbackQuery) -> bool:
    """Chỉ Admin / Owner của nhóm mới được thao tác trên thông báo (dùng chung với Credit)."""
    from bot.handlers.credit import _check_admin_or_owner_credit
    return await _check_admin_or_owner_credit(callback_query)


@bot.on_callback_query(filters.regex(r"^rnt_full_pay\|([^|]+)$"))
@command_timeout(auto_delete_cmd=False, auto_expire_menu=False)  # Không được tự xóa tin THÔNG BÁO ĐÓNG TIỀN THUÊ
async def rnt_full_pay_callback(client, callback_query: CallbackQuery):
    """Nút 1: 'Đã nhận TT đủ' -> Ghi nhận thanh toán toàn bộ tiền thuê đang nợ."""
    if not await _check_admin_or_owner_rental(callback_query):
        await callback_query.answer("⚠️ Thao tác này chỉ dành cho Admin và Owner!", show_alert=True)
        return

    contract_code = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import Rental, RentalPayment, RentalStatus

        contract = db.query(Rental).filter(Rental.contract_id == contract_code).first()
        if not contract:
            await callback_query.answer(f"❌ Không tìm thấy hợp đồng {contract_code}.", show_alert=True)
            return

        paid_amount = contract.rental_debt or 0.0
        if paid_amount <= 0:
            await callback_query.answer("⚠️ Hợp đồng này hiện không có tiền thuê cần thanh toán.", show_alert=True)
            return

        # Chỉ về 0 khi công nợ vẫn đúng bằng số vừa đọc -> bấm 2 lần / 2 admin bấm
        # cùng lúc chỉ ghi nhận được 1 lần, không tạo 2 phiếu thu trùng.
        cleared = db.query(Rental).filter(
            Rental.id == contract.id,
            Rental.rental_debt == paid_amount
        ).update({"rental_debt": 0.0}, synchronize_session=False)

        if not cleared:
            db.rollback()
            await callback_query.answer(
                "Công nợ của hợp đồng vừa được cập nhật bởi thao tác khác. Vui lòng kiểm tra lại.",
                show_alert=True
            )
            return

        now = datetime.datetime.now()
        db.add(RentalPayment(
            contract_id=contract.contract_id,
            payment_date=now.date(),
            payment_time=now,
            payment_amount=paid_amount
        ))

        # Đang nợ xấu mà đã trả đủ -> đưa về Đang thuê, gỡ cờ [BLACKLIST]
        if contract.status == RentalStatus.BAD_DEBT.value:
            db.query(Rental).filter(Rental.id == contract.id).update(
                {
                    "status": RentalStatus.ACTIVE.value,
                    "notes": (contract.notes or "").replace("[BLACKLIST]", "").strip() or None,
                },
                synchronize_session=False
            )

        db.commit()

        await callback_query.message.edit_text(
            f"✅ <b>XÁC NHẬN ĐÃ NHẬN THANH TOÁN ĐỦ TIỀN THUÊ</b>\n\n"
            f"- Mã hợp đồng: <code>{contract_code}</code>\n"
            f"- Số tiền đã thu: <b>{_rental_fmt_num(paid_amount):,} VNĐ</b>\n"
            f"- Còn phải thu: <b>0 VNĐ</b>\n"
            f"- Thời gian: <b>{now.strftime('%d/%m/%Y %H:%M')}</b>",
            parse_mode=ParseMode.HTML
        )
        LogInfo(f"[RntFullPay] Contract {contract_code} paid full rent {paid_amount} by {callback_query.from_user.id}", LogType.SYSTEM_STATUS)
    except Exception as e:
        db.rollback()
        LogError(f"Error in rnt_full_pay_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi cập nhật thanh toán.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rnt_pay\|([^|]+)$"))
@command_timeout(auto_delete_cmd=False, auto_expire_menu=False)  # Không được tự xóa tin THÔNG BÁO ĐÓNG TIỀN THUÊ
async def rnt_pay_callback(client, callback_query: CallbackQuery):
    """Nút 2: 'Đã nhận TT' -> Hướng dẫn & mẫu lệnh /rental_payment_confirmed."""
    if not await _check_admin_or_owner_rental(callback_query):
        await callback_query.answer("⚠️ Thao tác này chỉ dành cho Admin và Owner!", show_alert=True)
        return

    contract_code = callback_query.matches[0].group(1)

    await callback_query.message.reply_text(
        f"💡 <b>XÁC NHẬN THANH TOÁN TIỀN THUÊ</b>\n\n"
        f"Mã hợp đồng: <code>{contract_code}</code>\n\n"
        f"Vui lòng <b>Reply</b> vào tin nhắn THÔNG BÁO ĐÓNG TIỀN THUÊ và nhập số tiền đã nhận:\n"
        f"<pre>/rental_payment_confirmed [Số_Tiền_Thanh_Toán]</pre>\n"
        f"<i>(Ví dụ: <code>/rental_payment_confirmed 5000000</code>)</i>",
        parse_mode=ParseMode.HTML
    )
    await callback_query.answer()


@bot.on_callback_query(filters.regex(r"^rnt_remind\|([^|]+)$"))
@command_timeout(auto_delete_cmd=False, auto_expire_menu=False)  # Không được tự xóa tin THÔNG BÁO ĐÓNG TIỀN THUÊ
async def rnt_remind_callback(client, callback_query: CallbackQuery):
    """Nút 3: 'Lưu sổ' -> Dừng nhắc tháng này, tiền thuê vẫn giữ trong công nợ."""
    if not await _check_admin_or_owner_rental(callback_query):
        await callback_query.answer("⚠️ Thao tác này chỉ dành cho Admin và Owner!", show_alert=True)
        return

    contract_code = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import Rental

        contract = db.query(Rental).filter(Rental.contract_id == contract_code).first()
        if not contract:
            await callback_query.answer(f"❌ Không tìm thấy hợp đồng {contract_code}.", show_alert=True)
            return

        # Kỳ đang nhắc: mốc đến hạn gần nhất tính theo ngày bắt đầu thuê
        today = datetime.date.today()
        rental_day = contract.start_rental.day if contract.start_rental else today.day
        if today.day >= rental_day:
            due_year, due_month = today.year, today.month
        else:
            due_year, due_month = (today.year, today.month - 1) if today.month > 1 else (today.year - 1, 12)

        skip_tag = _rental_skip_tag(due_year, due_month)
        if contract.notes and skip_tag in contract.notes:
            await callback_query.answer("ℹ️ Kỳ này đã được lưu sổ từ trước.", show_alert=True)
            return

        contract.notes = f"{contract.notes} {skip_tag}".strip() if contract.notes else skip_tag
        db.commit()

        await callback_query.message.edit_text(
            f"📌 <b>ĐÃ LƯU SỔ KỲ {due_month:02d}/{due_year}</b>\n\n"
            f"- Mã hợp đồng: <code>{contract_code}</code>\n"
            f"- Tiền thuê cần thu: <b>{_rental_fmt_num(contract.rental_debt or 0):,} VNĐ</b> (vẫn giữ trong công nợ)\n"
            f"{'━' * 15}\n"
            f"<i>Bot sẽ ngừng nhắc kỳ này và nhắc lại vào kỳ sau.</i>",
            parse_mode=ParseMode.HTML
        )
        LogInfo(f"[RntRemind] Contract {contract_code} skipped reminder for {due_month:02d}/{due_year} by {callback_query.from_user.id}", LogType.SYSTEM_STATUS)
    except Exception as e:
        db.rollback()
        LogError(f"Error in rnt_remind_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi lưu sổ.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rnt_bad\|([^|]+)$"))
@command_timeout(auto_delete_cmd=False, auto_expire_menu=False)  # Không được tự xóa tin THÔNG BÁO ĐÓNG TIỀN THUÊ
async def rnt_bad_callback(client, callback_query: CallbackQuery):
    """Nút 4: 'Nợ Xấu' -> Chuyển hợp đồng sang Nợ Xấu (BAD_DEBT)."""
    if not await _check_admin_or_owner_rental(callback_query):
        await callback_query.answer("⚠️ Thao tác này chỉ dành cho Admin và Owner!", show_alert=True)
        return

    contract_code = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import Rental, RentalStatus

        contract = db.query(Rental).filter(Rental.contract_id == contract_code).first()
        if not contract:
            await callback_query.answer(f"❌ Không tìm thấy hợp đồng {contract_code}.", show_alert=True)
            return

        if contract.status == RentalStatus.BAD_DEBT.value:
            await callback_query.answer("⚠️ Hợp đồng đã nằm trong Blacklist Nợ Xấu từ trước.", show_alert=True)
            return

        contract.status = RentalStatus.BAD_DEBT.value
        if not contract.notes:
            contract.notes = "[BLACKLIST]"
        elif "[BLACKLIST]" not in contract.notes:
            contract.notes = f"{contract.notes}\n[BLACKLIST]"

        db.commit()

        await callback_query.message.edit_text(
            f"⚠️ <b>ĐÃ CHUYỂN HỢP ĐỒNG SANG NỢ XẤU</b>\n\n"
            f"- Mã hợp đồng: <code>{contract_code}</code>\n"
            f"- Tiền thuê còn phải thu: <b>{_rental_fmt_num(contract.rental_debt or 0):,} VNĐ</b>\n"
            f"- Trạng thái mới: <b>Nợ xấu (BLACKLIST)</b>",
            parse_mode=ParseMode.HTML
        )
        LogInfo(f"[RntBad] Contract {contract_code} set to BAD_DEBT by {callback_query.from_user.id}", LogType.SYSTEM_STATUS)
    except Exception as e:
        db.rollback()
        LogError(f"Error in rnt_bad_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi chuyển Nợ Xấu.", show_alert=True)
    finally:
        db.close()


# --- Rental: Payment Confirmed ---
@bot.on_message(filters.command(["rental_payment_confirmed", "rental_xac_nhan_thanh_toan"]) | filters.regex(r"^@\w+\s+/(rental_payment_confirmed|rental_xac_nhan_thanh_toan)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@command_timeout(auto_delete_cmd=True)
async def rental_payment_confirmed_handler(client, message: Message) -> None:
    if not message.reply_to_message:
        await message.reply_text("⚠️ Vui lòng <b>Reply</b> (trả lời) lại tin nhắn THÔNG BÁO ĐÓNG TIỀN THUÊ của Bot để sử dụng lệnh này.", parse_mode=ParseMode.HTML)
        return

    replied_text = message.reply_to_message.text or message.reply_to_message.caption
    if not replied_text or "THÔNG BÁO ĐÓNG TIỀN THUÊ" not in replied_text:
        await message.reply_text("⚠️ Lệnh này chỉ dùng để Reply vào tin nhắn THÔNG BÁO ĐÓNG TIỀN THUÊ của Bot.", parse_mode=ParseMode.HTML)
        return

    import re as _re
    match = _re.search(r"Mã Hợp Đồng:\s*([A-Za-z0-9_\-]+)", replied_text)
    if not match:
        await message.reply_text("❌ Không thể trích xuất Mã Hợp Đồng từ tin nhắn.", parse_mode=ParseMode.HTML)
        return

    contract_code = match.group(1).strip()

    args = message.text.split(maxsplit=1)
    if len(args) < 2:
        await message.reply_text("⚠️ Vui lòng cung cấp số tiền. Lệnh ví dụ: <code>/rental_payment_confirmed 5000000</code>", parse_mode=ParseMode.HTML)
        return

    amount_str = args[1]
    paid_amount = _parse_float_rental(amount_str)

    if paid_amount <= 0:
        await message.reply_text("⚠️ Số tiền thanh toán không hợp lệ.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        from app.models.rental import Rental, RentalCustomer, RentalStatus
        from app.models.telegram import TelegramProjectMember
        import html

        contract = db.query(Rental).filter(Rental.contract_id == contract_code).first()
        if not contract:
            await message.reply_text(f"❌ Không tìm thấy hợp đồng cho thuê <b>{contract_code}</b> trong CSDL.", parse_mode=ParseMode.HTML)
            return

        if contract.status == RentalStatus.CANCELLED.value:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> đã bị hủy, không thể thanh toán.", parse_mode=ParseMode.HTML)
            return

        # Subtract paid amount from rental_debt
        contract.rental_debt = (contract.rental_debt or 0.0) - paid_amount

        # Save payment record for history
        from app.models.rental import RentalPayment
        now = datetime.datetime.now()
        new_payment = RentalPayment(
            contract_id=contract.contract_id,
            payment_date=now.date(),
            payment_time=now,
            payment_amount=paid_amount
        )
        db.add(new_payment)

        db.commit()

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()

        def fmt_num(val):
            if val is None: return 0
            return int(val) if val == int(val) else val

        remaining = fmt_num(contract.rental_debt)
        date_str = datetime.datetime.now().strftime('%d/%m/%Y')
        amount_fmt = fmt_num(paid_amount)

        reply_msg = (
            f"<b>{date_str}</b>\n"
            f"Đã cập nhật thanh toán: <b>{amount_fmt:,} VNĐ</b> vào hợp đồng cho thuê <b>{contract.contract_id}</b>\n"
            f"Khách hàng: <b>{customer.customer_name if customer else 'N/A'}</b>\n"
            # f"Tổng tiền thanh toán HD hiện tại: <b>{remaining:,} VNĐ</b>"
        )

        await message.reply_text(reply_msg, parse_mode=ParseMode.HTML)
        LogInfo(f"[RentalPaymentConfirmed] Contract {contract.contract_id} received {amount_fmt} by user {message.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in rental_payment_confirmed_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình cập nhật thanh toán.")
    finally:
        db.close()


# --- Rental: Update Contract (Interactive) ---
def _build_rental_update_contract_form(contract, customer) -> str:
    """Form cập nhật hợp đồng cho thuê đã điền sẵn dữ liệu hiện tại."""
    def fmt_dt(dt):
        return dt.strftime('%d/%m/%Y') if dt else ""

    return f"""<b>FORM CẬP NHẬT HỢP ĐỒNG CHO THUÊ</b>
Vui lòng sao chép toàn bộ form dưới đây, chỉnh sửa thông tin cần thay đổi và gửi lại:

<pre>/rental_update_contract {contract.contract_id}
Mã Khách Hàng: {customer.customer_id or ""}
Tên Nhóm: {customer.group_name or ""}
Tên Khách Hàng: {customer.customer_name or ""}
Liên Hệ Khách Hàng: {customer.contact_info or ""}
Số Điện Thoại: {customer.number_phone or ""}
Chat ID (Telegram): {customer.chat_id or ""}
Mã Hợp Đồng: {contract.contract_id or ""}
Mã Bất Động Sản: {contract.real_estate_id or ""}
Loại Hợp Đồng: {contract.type_contract or ""}
Ngày Bắt Đầu Thuê (dd/mm/yyyy): {fmt_dt(contract.start_rental)}
Ngày Kết Thúc Thuê (dd/mm/yyyy): {fmt_dt(contract.end_rental)}
Tiền Cọc: {_rental_fmt_num(contract.deposit)}
Tiền Thuê / Tháng: {_rental_fmt_num(contract.monthly_rental)}
Trạng Thái (active/expired/cancelled): {contract.status or 'active'}
</pre>"""


def _ruh_customer_list_keyboard(customers, page):
    """Bàn phím chọn khách hàng để cập nhật hợp đồng: tối đa 10 nút/trang."""
    PAGE_SIZE = 10
    total = len(customers)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_customers = customers[start:min(start + PAGE_SIZE, total)]

    buttons = [
        [InlineKeyboardButton(
            f"{c.customer_id} - {c.customer_name}",
            callback_data=f"ruh_c|{_rsid(c.id)}|0"
        )]
        for c in page_customers
    ]

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"ruh_cp|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="ruh_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"ruh_cp|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="ruh_x")])
    return InlineKeyboardMarkup(buttons)


def _ruh_contract_list_keyboard(contracts, page, customer_hex):
    """Bàn phím chọn hợp đồng cần cập nhật: tối đa 10 nút/trang."""
    PAGE_SIZE = 10
    total = len(contracts)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_contracts = contracts[start:min(start + PAGE_SIZE, total)]

    buttons = [
        [InlineKeyboardButton(
            f"{c.contract_id} ({_rental_status_label(c.status)})",
            callback_data=f"ruh_s|{_rsid(c.id)}"
        )]
        for c in page_contracts
    ]

    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"ruh_c|{customer_hex}|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="ruh_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"ruh_c|{customer_hex}|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="ruh_x")])
    return InlineKeyboardMarkup(buttons)


@bot.on_callback_query(filters.regex(r"^ruh_noop$"))
@command_timeout(auto_delete_cmd=True)
async def ruh_noop_callback(client, callback_query: CallbackQuery):
    await callback_query.answer()


@bot.on_callback_query(filters.regex(r"^ruh_x$"))
@command_timeout(auto_delete_cmd=True)
async def ruh_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


@bot.on_callback_query(filters.regex(r"^ruh_cp\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def ruh_customer_page_callback(client, callback_query: CallbackQuery):
    """Phân trang danh sách khách hàng khi cập nhật hợp đồng."""
    page = int(callback_query.matches[0].group(1))
    db = SessionLocal()
    try:
        customers = _rental_customers_of_chat(db, callback_query.message.chat.id)
        if not customers:
            await _rk_safe_edit(callback_query, "ℹ️ Chưa có khách hàng cho thuê nào trong dự án này.")
            return

        await _rk_safe_edit(
            callback_query,
            "<b>CẬP NHẬT HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng cần cập nhật hợp đồng:",
            _ruh_customer_list_keyboard(customers, page),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in ruh_customer_page_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^ruh_c\|([a-f0-9]{32})\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def ruh_contract_list_callback(client, callback_query: CallbackQuery):
    """Chọn khách hàng -> hiển thị danh sách hợp đồng có thể cập nhật."""
    customer_hex = callback_query.matches[0].group(1)
    page = int(callback_query.matches[0].group(2))
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental, RentalStatus
        customer = db.query(RentalCustomer).filter(RentalCustomer.id == _ruid(customer_hex)).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng.", show_alert=True)
            return

        # Hợp đồng đã hủy không cập nhật được -> không đưa vào danh sách
        contracts = db.query(Rental).filter(
            Rental.customer_id == customer.id,
            Rental.status != RentalStatus.CANCELLED.value
        ).all()
        if not contracts:
            await callback_query.answer("ℹ️ Khách hàng chưa có hợp đồng nào có thể cập nhật.", show_alert=True)
            return

        await _rk_safe_edit(
            callback_query,
            f"<b>CẬP NHẬT HỢP ĐỒNG CHO THUÊ</b>\n"
            f"Khách hàng: <b>{customer.customer_name}</b> ({customer.customer_id or 'N/A'})\n\n"
            f"Chọn hợp đồng cần cập nhật:",
            _ruh_contract_list_keyboard(contracts, page, customer_hex),
        )
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in ruh_contract_list_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^ruh_s\|([a-f0-9]{32})$"))
@command_timeout(timeout_seconds=600, auto_delete_cmd=True)  # Form nhiều trường -> cần thời gian điền
async def ruh_select_contract_callback(client, callback_query: CallbackQuery):
    """Chọn hợp đồng -> hiển thị form cập nhật đã điền sẵn."""
    contract_hex = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental
        contract = db.query(Rental).filter(Rental.id == _ruid(contract_hex)).first()
        if not contract:
            await callback_query.answer("⚠️ Không tìm thấy hợp đồng.", show_alert=True)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        if not customer:
            await callback_query.answer("⚠️ Không tìm thấy khách hàng của hợp đồng này.", show_alert=True)
            return

        form_msg = await callback_query.message.reply_text(
            _build_rental_update_contract_form(contract, customer), parse_mode=ParseMode.HTML
        )
        form_tracker.track(callback_query.message.chat.id, "rental_update_contract", contract.contract_id, form_msg.id)
        await callback_query.message.delete()
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in ruh_select_contract_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


# --- Rental: Update Contract ---
@bot.on_message(filters.command(["rental_update_contract", "rental_cap_nhat_hop_dong"]) | filters.regex(r"^@\w+\s+/(rental_update_contract|rental_cap_nhat_hop_dong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(timeout_seconds=600, auto_delete_cmd=True)  # Form nhiều trường -> cần thời gian điền
async def rental_update_contract_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_update_contract", "rental_cap_nhat_hop_dong"])
    if args is None: return

    if len(args) < 2:
        db = SessionLocal()
        try:
            customers = _rental_customers_of_chat(db, message.chat.id)
            if not customers:
                await message.reply_text("ℹ️ Chưa có khách hàng cho thuê nào trong dự án này.", parse_mode=ParseMode.HTML)
                return

            await message.reply_text(
                "<b>CẬP NHẬT HỢP ĐỒNG CHO THUÊ</b>\n\nChọn khách hàng cần cập nhật hợp đồng:",
                reply_markup=_ruh_customer_list_keyboard(customers, 0),
                parse_mode=ParseMode.HTML
            )
        finally:
            db.close()
        return

    contract_code = args[1]
    chat_id = str(message.chat.id)
    db = SessionLocal()
    try:
        from app.models.telegram import TelegramProjectMember
        from app.models.rental import RentalCustomer, Rental, RentalStatus

        current_group = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()
        project_id = current_group.project_id

        # Get valid contacts for project isolation
        valid_members = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project_id,
            TelegramProjectMember.role == "member"
        ).all()

        valid_groups = []
        for m in valid_members:
            if m.group_name:
                valid_groups.append(m.group_name)

        contract = db.query(Rental).filter(Rental.contract_id == contract_code).first()
        if not contract:
            await message.reply_text(f"⚠️ Không tìm thấy hợp đồng cho thuê nào có mã <b>{contract_code}</b>.", parse_mode=ParseMode.HTML)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        if not customer or not customer.group_name or customer.group_name not in valid_groups:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> không thuộc về nhóm hợp lệ nào trong dự án hiện tại.", parse_mode=ParseMode.HTML)
            return

        if contract.status == RentalStatus.CANCELLED.value:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> đã bị hủy, không thể cập nhật.", parse_mode=ParseMode.HTML)
            return

        # Handle form requests (just /rental_update_contract HD123) vs form submissions (multiple lines)
        lines = message.text.strip().split("\n")
        if len(lines) < 3:
            # Show pre-filled form
            form_msg = await message.reply_text(
                _build_rental_update_contract_form(contract, customer), parse_mode=ParseMode.HTML
            )
            # Track the form message so we can delete it after successful update
            form_tracker.track(message.chat.id, "rental_update_contract", contract.contract_id, form_msg.id)
            return

        # Parse submitted form data
        data = {}
        for line in lines[1:]:
            if ":" in line:
                key, val = line.split(":", 1)
                data[key.strip()] = val.strip()

        parse_float = _parse_float_rental

        def parse_date(date_str: str):
            if not date_str: return None
            try:
                return datetime.datetime.strptime(date_str, "%d/%m/%Y").date()
            except:
                return None

        new_contract_id = data.get("Mã Hợp Đồng", "").strip()
        if not new_contract_id:
            await message.reply_text("⚠️ <b>Mã Hợp Đồng</b> không được để trống.", parse_mode=ParseMode.HTML)
            return

        # Check duplicate contract code
        if new_contract_id != contract.contract_id:
            dup_contract = db.query(Rental).filter(Rental.contract_id == new_contract_id).first()
            if dup_contract:
                await message.reply_text(f"⚠️ Hợp đồng mã <b>{new_contract_id}</b> đã tồn tại trên một hợp đồng khác.", parse_mode=ParseMode.HTML)
                return

        # Update customer info
        group_name = data.get("Tên Nhóm", customer.group_name)
        if not group_name or group_name not in valid_groups:
            await message.reply_text(f"⚠️ Nhóm mới <b>{group_name}</b> không hợp lệ hoặc chưa được đồng bộ.", parse_mode=ParseMode.HTML)
            return
            
        # Chat ID nhóm member: đổi nhóm -> lấy lại theo nhóm mới; nhập tay -> ưu tiên giá trị nhập
        input_chat_id = data.get("Chat ID (Telegram)", "").strip()
        new_chat_id = customer.chat_id

        if group_name != customer.group_name:
            new_chat_id = None
            for m in valid_members:
                if m.group_name == group_name and m.chat_id:
                    new_chat_id = str(m.chat_id)
                    break

        if input_chat_id:
            valid_chat_ids = {str(m.chat_id) for m in valid_members if m.chat_id}
            if valid_chat_ids and input_chat_id not in valid_chat_ids:
                await message.reply_text(f"⚠️ Chat ID <b>{input_chat_id}</b> không thuộc nhóm member nào của dự án. Vui lòng kiểm tra lại hoặc để trống để bot tự lấy theo Tên Nhóm.", parse_mode=ParseMode.HTML)
                return
            new_chat_id = input_chat_id

        customer.group_name = group_name
        customer.customer_name = data.get("Tên Khách Hàng", customer.customer_name)
        customer.contact_info = data.get("Liên Hệ Khách Hàng", customer.contact_info)
        customer.number_phone = data.get("Số Điện Thoại", customer.number_phone)
        customer.chat_id = new_chat_id

        # Update contract fields
        contract.contract_id = new_contract_id
        contract.real_estate_id = data.get("Mã Bất Động Sản", contract.real_estate_id)
        contract.type_contract = data.get("Loại Hợp Đồng", contract.type_contract)
        contract.start_rental = parse_date(data.get("Ngày Bắt Đầu Thuê (dd/mm/yyyy)", "")) or contract.start_rental
        contract.end_rental = parse_date(data.get("Ngày Kết Thúc Thuê (dd/mm/yyyy)", "")) or contract.end_rental
        contract.deposit = parse_float(data.get("Tiền Cọc", "")) or contract.deposit
        contract.monthly_rental = parse_float(data.get("Tiền Thuê / Tháng", "")) or contract.monthly_rental

        debt_str = data.get("Tổng tiền thanh toán HD", "").strip()
        if debt_str:
            contract.rental_debt = parse_float(debt_str)

        # Update status
        status_str = data.get("Trạng Thái (active/expired/cancelled)", "").strip().lower()
        if status_str in [RentalStatus.ACTIVE.value, RentalStatus.EXPIRED.value, RentalStatus.CANCELLED.value]:
            contract.status = status_str

        db.commit()
        await message.reply_text(
            f"✅ Đã cập nhật hợp đồng cho thuê <b>{new_contract_id}</b> thành công!\n"
            f"Chat ID (Telegram): <code>{new_chat_id or 'Chưa có'}</code>",
            parse_mode=ParseMode.HTML
        )
        LogInfo(f"[RentalUpdateContract] Updated rental contract {new_contract_id} (chat_id={new_chat_id}) by {message.from_user.id}", LogType.SYSTEM_STATUS)

        # Delete the form template message after successful update
        form_msg_id = form_tracker.pop(message.chat.id, "rental_update_contract", contract_code)
        if form_msg_id:
            try:
                await client.delete_messages(chat_id=message.chat.id, message_ids=form_msg_id)
            except Exception as del_err:
                LogError(f"Failed to delete rental update form message: {del_err}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in rental_update_contract_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình xử lý.")
    finally:
        db.close()

# --- Revenue Date Selector ---
async def generate_rental_revenue_report(client, message, project_id, start_date, end_date):
    db = SessionLocal()
    try:
        from app.models.telegram import TelegramProjectMember
        from app.models.rental import RentalCustomer, Rental, RentalStatus, RentalPayment
        
        valid_members = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project_id,
            TelegramProjectMember.role == "member"
        ).all()
        
        valid_groups = []
        for m in valid_members:
            if m.group_name:
                valid_groups.append(m.group_name)
            
        customers = db.query(RentalCustomer).filter(RentalCustomer.group_name.in_(valid_groups)).all()
        if not customers:
            await message.reply_text("⚠️ Không có khách hàng nào trong dự án cập nhật thông tin trên hệ thống.", parse_mode=ParseMode.HTML)
            return

        all_rentals = db.query(Rental).filter(Rental.customer_id.in_([c.id for c in customers])).all()
        contract_ids = [c.contract_id for c in all_rentals]
        
        payments = db.query(RentalPayment).filter(
            RentalPayment.contract_id.in_(contract_ids)
        ).all()
        
        import datetime
        valid_payments = []
        for payment in payments:
            record_date = payment.payment_date
            if not record_date:
                continue
            if start_date <= record_date <= end_date:
                valid_payments.append(payment)
                
        from bot.utils.utils import fmt_vn
            
        total_collected = sum([p.payment_amount or 0 for p in valid_payments])
        
        active_rentals = [c for c in all_rentals if c.status != RentalStatus.CANCELLED.value]
        total_remaining = sum([max(0.0, c.rental_debt or 0.0) for c in active_rentals])
        total_receivable = total_collected + total_remaining
        
        rate = (total_collected / total_receivable * 100) if total_receivable > 0 else 0.0
        rate_str = f"{rate:.2f}".replace(".", ",")
        
        today = datetime.date.today()
        paid_contracts = 0
        unpaid_contracts = 0
        overdue_contracts = 0
        
        for c in active_rentals:
            debt = c.rental_debt or 0.0
            if debt <= 0:
                paid_contracts += 1
            else:
                is_overdue = (c.status in [RentalStatus.EXPIRED.value, RentalStatus.BAD_DEBT.value]) or (c.end_rental and c.end_rental < today)
                if is_overdue:
                    overdue_contracts += 1
                else:
                    unpaid_contracts += 1
        
        report_lines = [
            "<b>BÁO CÁO DOANH THU THUÊ NHÀ</b>",
            f"<i>Thời gian: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}</i>",
            f"{'━' * 15}",
            "<b>Doanh thu</b>",
            f"• Tổng tiền thuê phải thu: {fmt_vn(total_receivable)}",
            f"• Tổng tiền thuê đã thu: {fmt_vn(total_collected)}",
            f"• Còn phải thu: {fmt_vn(total_remaining)}",
            f"• Tỷ lệ thu tiền: {rate_str}%",
            "",
            "<b>Thanh toán</b>",
            f"• Đã thanh toán: {paid_contracts} hợp đồng",
            f"• Chưa thanh toán: {unpaid_contracts} hợp đồng",
            f"• Quá hạn: {overdue_contracts} hợp đồng",
        ]
        
        await message.reply_text("\n".join(report_lines), parse_mode=ParseMode.HTML)
        
    except Exception as e:
        LogError(f"Error in generate_rental_revenue_report: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình tính toán báo cáo.")
    finally:
        db.close()


@bot.on_message(filters.command(["rental_revenue", "rental_doanh_thu"]) | filters.regex(r"^@\w+\s+/(rental_revenue|rental_doanh_thu)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(auto_delete_cmd=True)
async def rental_revenue_report_handler(client, message: Message) -> None:
    db = SessionLocal()
    try:
        from app.models.telegram import TelegramProjectMember
        chat_id = str(message.chat.id)
        current_project_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()

        import re
        match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})", message.text)
        if match:
            import datetime
            start_date_str = match.group(1)
            end_date_str = match.group(2)
            try:
                start_date = datetime.datetime.strptime(start_date_str, "%d/%m/%Y").date()
                end_date = datetime.datetime.strptime(end_date_str, "%d/%m/%Y").date()
                await generate_rental_revenue_report(client, message, current_project_member.project_id, start_date, end_date)
                return
            except ValueError:
                await message.reply_text("⚠️ Định dạng ngày không hợp lệ. Vui lòng dùng dd/mm/yyyy - dd/mm/yyyy", parse_mode=ParseMode.HTML)
                return

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("7 ngày qua", callback_data="rental_rev_7d"), InlineKeyboardButton("14 ngày qua", callback_data="rental_rev_14d")],
            [InlineKeyboardButton("21 ngày qua", callback_data="rental_rev_21d"), InlineKeyboardButton("1 tháng qua", callback_data="rental_rev_1m")],
            [InlineKeyboardButton("1 quý qua", callback_data="rental_rev_1q"), InlineKeyboardButton("Năm nay", callback_data="rental_rev_ytd")],
            [InlineKeyboardButton("Năm trước", callback_data="rental_rev_prev")],
            [InlineKeyboardButton("Hủy", callback_data="rental_rev_cancel")]
        ])
        
        await message.reply_text("Vui lòng chọn khoảng thời gian để xem báo cáo doanh thu thuê nhà:", reply_markup=keyboard)
        
    except Exception as e:
        LogError(f"Error in rental_revenue_report_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rental_rev_(.+)$"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_group_role("main")
@command_timeout(auto_delete_cmd=True)
async def rental_revenue_callback_handler(client, callback_query: CallbackQuery):
    db = SessionLocal()
    try:
        from app.models.telegram import TelegramProjectMember
        chat_id = str(callback_query.message.chat.id)
        current_project_member = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()

        data = callback_query.data.split("_", 2)[2]
        
        if data == "cancel":
            await callback_query.message.delete()
            return

        import datetime
        today = datetime.date.today()
        
        start_date = today
        end_date = today
        
        if data == "7d":
            start_date = today - datetime.timedelta(days=7)
        elif data == "14d":
            start_date = today - datetime.timedelta(days=14)
        elif data == "21d":
            start_date = today - datetime.timedelta(days=21)
        elif data == "1m":
            start_date = today - datetime.timedelta(days=30)
        elif data == "1q":
            start_date = today - datetime.timedelta(days=90)
        elif data == "ytd":
            start_date = datetime.date(today.year, 1, 1)
        elif data == "prev":
            start_date = datetime.date(today.year - 1, 1, 1)
            end_date = datetime.date(today.year - 1, 12, 31)
            
        await callback_query.message.delete()
        await generate_rental_revenue_report(client, callback_query.message, current_project_member.project_id, start_date, end_date)
        
    except Exception as e:
        LogError(f"Error in rental_revenue_callback_handler: {e}", LogType.SYSTEM_STATUS)
        await callback_query.message.reply_text("❌ Có lỗi xảy ra.")
    finally:
        db.close()

# --- Rental: Cashflow Report (Paginated) ---
def _rbc_build_report(db, chat_id):
    """Dựng báo cáo dòng tiền: trả về (header_lines, customer_lines, err_text)."""
    from app.models.telegram import TelegramProjectMember
    from app.models.rental import RentalCustomer, Rental, RentalStatus, RentalPayment
    from bot.utils.utils import fmt_vn

    current_project_member = db.query(TelegramProjectMember).filter(
        TelegramProjectMember.chat_id == str(chat_id)
    ).first()
    if not current_project_member:
        return None, None, "⚠️ Nhóm này chưa được đồng bộ vào dự án nào. Vui lòng sử dụng lệnh /syncchat trước."

    valid_members = db.query(TelegramProjectMember).filter(
        TelegramProjectMember.project_id == current_project_member.project_id,
        TelegramProjectMember.role == "member"
    ).all()

    valid_groups = [m.group_name for m in valid_members if m.group_name]
    customers = db.query(RentalCustomer).filter(RentalCustomer.group_name.in_(valid_groups)).all()
    if not customers:
        return None, None, "⚠️ Không có khách hàng nào trong dự án này."

    total_contracts = 0
    total_deposit = 0
    total_debt = 0
    project_monthly_totals = {}
    customer_lines = []

    for customer in customers:
        all_cust_rentals = db.query(Rental).filter(Rental.customer_id == customer.id).all()
        active_rentals = [c for c in all_cust_rentals if c.status in [RentalStatus.ACTIVE.value, RentalStatus.EXPIRED.value]]

        if not active_rentals and not all_cust_rentals:
            continue

        cust_contracts = len(active_rentals)
        cust_deposit = sum([(c.deposit or 0) for c in active_rentals])
        cust_debt = sum([(c.rental_debt or 0) for c in active_rentals])

        total_contracts += cust_contracts
        total_deposit += cust_deposit
        total_debt += cust_debt

        cust_paid_str = ""
        if all_cust_rentals:
            contract_ids = [c.contract_id for c in all_cust_rentals]
            payments = db.query(RentalPayment).filter(RentalPayment.contract_id.in_(contract_ids)).all()
            monthly_totals = {}
            for payment in payments:
                if not payment.payment_date:
                    continue
                ym = payment.payment_date.strftime("%m/%Y")
                monthly_totals[ym] = monthly_totals.get(ym, 0) + (payment.payment_amount or 0)
                project_monthly_totals[ym] = project_monthly_totals.get(ym, 0) + (payment.payment_amount or 0)

            if monthly_totals:
                cust_paid_str = f"<b>{fmt_vn(sum(monthly_totals.values()))}</b>"

        if not cust_paid_str:
            cust_paid_str = "Chưa đóng tiền"

        if cust_contracts == 0 and cust_paid_str == "Chưa đóng tiền":
            continue

        customer_lines.append(
            f"<b>{customer.customer_name}</b> (Mã: {customer.customer_id})\n"
            f"   Hợp đồng: {cust_contracts} | Tiền Cọc: <b>{fmt_vn(cust_deposit)}</b> | Số tiền nợ còn lại: <b>{fmt_vn(cust_debt)}</b>\n"
            f"   Tổng thanh toán: {cust_paid_str}"
        )

    if total_contracts == 0 and not project_monthly_totals:
        return None, None, "ℹ️ Không có dữ liệu hợp đồng/dòng tiền nào trong dự án."

    total_project_paid = sum(project_monthly_totals.values()) if project_monthly_totals else 0
    header_lines = [
        f"<b>BÁO CÁO DÒNG TIỀN DỰ ÁN THUÊ NHÀ</b>",
        f"{'━' * 15}",
        f"<b>Tổng Hợp Đồng Đang Thuê:</b> {total_contracts}",
        f"<b>Tổng Tiền Cọc Đang Giữ:</b> {fmt_vn(total_deposit)}",
        f"<b>Tổng Tiền Thuê Đã Thu:</b> {fmt_vn(total_project_paid)}",
        f"<b>Tổng Số Khách Hàng:</b> {len(customer_lines)}",
        f"{'━' * 15}"
    ]

    return header_lines, customer_lines, None


def _rbc_page_text(header_lines, customer_lines, page):
    """Nội dung 1 trang báo cáo: tối đa 10 khách hàng / trang."""
    PAGE_SIZE = 10
    total = len(customer_lines)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_lines = customer_lines[start:min(start + PAGE_SIZE, total)]

    body = []
    for idx, line in enumerate(page_lines, start + 1):
        body.append(f"\n{idx}. {line}")

    footer = [f"\n{'━' * 15}", f"<i>Trang {page + 1}/{total_pages} — {total} khách hàng</i>"]
    return "\n".join(header_lines + body + footer), total_pages


def _rbc_keyboard(page, total_pages):
    """Nút phân trang báo cáo dòng tiền."""
    buttons = []
    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("<< Trước", callback_data=f"rbc_p|{page - 1}"))
    nav_row.append(InlineKeyboardButton(f"{page + 1}/{total_pages}", callback_data="rbc_noop"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("Sau >>", callback_data=f"rbc_p|{page + 1}"))
    if len(nav_row) > 1:
        buttons.append(nav_row)

    buttons.append([InlineKeyboardButton("Hủy", callback_data="rbc_x")])
    return InlineKeyboardMarkup(buttons)


@bot.on_callback_query(filters.regex(r"^rbc_noop$"))
@command_timeout(auto_delete_cmd=True)
async def rbc_noop_callback(client, callback_query: CallbackQuery):
    await callback_query.answer()


@bot.on_callback_query(filters.regex(r"^rbc_x$"))
@command_timeout(auto_delete_cmd=True)
async def rbc_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


@bot.on_callback_query(filters.regex(r"^rbc_p\|(\d+)$"))
@command_timeout(auto_delete_cmd=True)
async def rbc_page_callback(client, callback_query: CallbackQuery):
    """Chuyển trang báo cáo dòng tiền."""
    page = int(callback_query.matches[0].group(1))
    db = SessionLocal()
    try:
        header_lines, customer_lines, err = _rbc_build_report(db, callback_query.message.chat.id)
        if err:
            await _rk_safe_edit(callback_query, err)
            return

        text, total_pages = _rbc_page_text(header_lines, customer_lines, page)
        await _rk_safe_edit(callback_query, text, _rbc_keyboard(page, total_pages))
        await callback_query.answer()
    except Exception as e:
        LogError(f"Error in rbc_page_callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra.", show_alert=True)
    finally:
        db.close()


# --- Report Cashflow Command ---
@bot.on_message(filters.command(["rental_cashflow_report", "rental_bao_cao_dong_tien"]) | filters.regex(r"^@\w+\s+/(rental_cashflow_report|rental_bao_cao_dong_tien)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(auto_delete_cmd=True)
async def rental_cashflow_report_handler(client, message: Message) -> None:
    db = SessionLocal()
    try:
        header_lines, customer_lines, err = _rbc_build_report(db, message.chat.id)
        if err:
            await message.reply_text(err, parse_mode=ParseMode.HTML)
            return

        text, total_pages = _rbc_page_text(header_lines, customer_lines, 0)
        await message.reply_text(
            text,
            reply_markup=_rbc_keyboard(0, total_pages),
            parse_mode=ParseMode.HTML
        )

    except Exception as e:
        LogError(f"Error in rental_cashflow_report_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình truy xuất báo cáo.")
    finally:
        db.close()


# --- Rental: List Contracts ---
@bot.on_message(filters.command(["rental_list_contract", "rental_danh_sach_hop_dong"]) | filters.regex(r"^@\w+\s+/(rental_list_contract|rental_danh_sach_hop_dong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(auto_delete_cmd=True)
async def rental_list_contract_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_list_contract", "rental_danh_sach_hop_dong"])
    if args is None: return

    db = SessionLocal()
    try:
        from app.models.rental import RentalCustomer, Rental, RentalStatus
        from app.models.telegram import TelegramProjectMember

        chat_id = str(message.chat.id)
        current_group = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()

        if not current_group:
            await message.reply_text("⚠️ Nhóm này chưa được đồng bộ vào dự án nào.", parse_mode=ParseMode.HTML)
            return

        project_id = current_group.project_id

        valid_members = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project_id,
            TelegramProjectMember.role == "member"
        ).all()

        valid_groups = []
        for m in valid_members:
            if m.group_name:
                valid_groups.append(m.group_name)

        customers = db.query(RentalCustomer).filter(RentalCustomer.group_name.in_(valid_groups)).all()
        if not customers:
            await message.reply_text("ℹ️ Không có khách hàng nào trong dự án này.", parse_mode=ParseMode.HTML)
            return

        customer_ids = [c.id for c in customers]
        customer_map = {c.id: c for c in customers}

        all_contracts = db.query(Rental).filter(Rental.customer_id.in_(customer_ids)).all()
        if not all_contracts:
            await message.reply_text("ℹ️ Không có hợp đồng cho thuê nào trong dự án.", parse_mode=ParseMode.HTML)
            return

        STATUS_MAP = {
            RentalStatus.ACTIVE.value: "Đang thuê",
            RentalStatus.EXPIRED.value: "Hết hạn",
            RentalStatus.CANCELLED.value: "Đã hủy",
            RentalStatus.BAD_DEBT.value: "Nợ xấu",
        }

        # Group by status
        grouped = {}
        for c in all_contracts:
            label = STATUS_MAP.get(c.status, "Không rõ")
            grouped.setdefault(label, []).append(c)

        # Sort order
        status_order = ["Đang thuê", "Hết hạn", "Nợ xấu", "Đã hủy", "Không rõ"]

        lines = [
            "DANH SÁCH HỢP ĐỒNG CHO THUÊ",
            f"Tổng: {len(all_contracts)} hợp đồng",
            f"{'━' * 15}",
        ]

        idx = 1
        for status in status_order:
            contracts = grouped.get(status, [])
            if not contracts:
                continue
            lines.append(f"\n{status} ({len(contracts)})")
            for c in contracts:
                cust = customer_map.get(c.customer_id)
                cust_name = cust.customer_name if cust else "N/A"
                lines.append(f"  {idx}. {c.contract_id} - {cust_name}")
                idx += 1

        if len(all_contracts) > 20:
            import io
            txt_content = "\n".join(lines)
            buf = io.BytesIO(txt_content.encode("utf-8"))
            buf.name = f"danh_sach_hop_dong_{datetime.datetime.now().strftime('%d%m%Y')}.txt"
            await message.reply_document(
                document=buf,
                caption=f"<b>DANH SÁCH HỢP ĐỒNG CHO THUÊ</b>\nTổng: <b>{len(all_contracts)}</b> hợp đồng\n\n<i>File đính kèm bên dưới.</i>",
                parse_mode=ParseMode.HTML
            )
        else:
            # Send as inline message with HTML formatting
            html_lines = [
                f"<b>DANH SÁCH HỢP ĐỒNG CHO THUÊ</b>",
                f"Tổng: <b>{len(all_contracts)}</b> hợp đồng",
                f"{'━' * 15}",
            ]
            idx2 = 1
            for status in status_order:
                contracts = grouped.get(status, [])
                if not contracts:
                    continue
                html_lines.append(f"\n<b>{status} ({len(contracts)})</b>")
                for c in contracts:
                    cust = customer_map.get(c.customer_id)
                    cust_name = cust.customer_name if cust else "N/A"
                    html_lines.append(f"  {idx2}. <code>{c.contract_id}</code> - {cust_name}")
                    idx2 += 1

            msg_text = "\n".join(html_lines)
            await message.reply_text(msg_text, parse_mode=ParseMode.HTML)

        LogInfo(f"[RentalListContract] Listed {len(all_contracts)} contracts by user {message.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        LogError(f"Error in rental_list_contract_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra khi truy xuất danh sách hợp đồng.")
    finally:
        db.close()



# --- Rental: Bad Debt / Blacklist ---
@bot.on_message(filters.command(["rental_bad_debt", "rental_xac_nhan_no_xau"]) | filters.regex(r"^@\w+\s+/(rental_bad_debt|rental_xac_nhan_no_xau)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main")
@command_timeout(auto_delete_cmd=True)
async def rental_bad_debt_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["rental_bad_debt", "rental_xac_nhan_no_xau"])
    if args is None: return

    if len(args) < 2:
        await message.reply_text("⚠️ Vui lòng cung cấp mã hợp đồng. Lệnh ví dụ: <code>/rental_bad_debt LMD-HD220101</code>", parse_mode=ParseMode.HTML)
        return

    contract_code = args[1].strip()
    db = SessionLocal()
    try:
        from app.models.rental import Rental, RentalCustomer, RentalStatus
        from app.models.telegram import TelegramProjectMember

        # Verify contract belongs to current project
        chat_id = str(message.chat.id)
        current_group = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.chat_id == chat_id
        ).first()

        if not current_group:
            await message.reply_text("⚠️ Nhóm này chưa được đồng bộ vào dự án nào.", parse_mode=ParseMode.HTML)
            return

        project_id = current_group.project_id

        valid_members = db.query(TelegramProjectMember).filter(
            TelegramProjectMember.project_id == project_id,
            TelegramProjectMember.role == "member"
        ).all()

        valid_groups = []
        for m in valid_members:
            if m.group_name:
                valid_groups.append(m.group_name)

        contract = db.query(Rental).filter(Rental.contract_id == contract_code).first()
        if not contract:
            await message.reply_text(f"❌ Không tìm thấy hợp đồng cho thuê <b>{contract_code}</b> trong CSDL.", parse_mode=ParseMode.HTML)
            return

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        if not customer or not customer.group_name or customer.group_name not in valid_groups:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> không thuộc về nhóm hợp lệ nào trong dự án hiện tại.", parse_mode=ParseMode.HTML)
            return

        if contract.status == RentalStatus.BAD_DEBT.value:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> đã nằm trong Blacklist Nợ Xấu từ trước rồi.", parse_mode=ParseMode.HTML)
            return

        if contract.status == RentalStatus.CANCELLED.value:
            await message.reply_text(f"⚠️ Hợp đồng <b>{contract_code}</b> đã bị hủy, không thể đưa vào Blacklist.", parse_mode=ParseMode.HTML)
            return

        def fmt_num(val):
            if val is None: return 0
            return int(val) if val == int(val) else val

        def fmt_dt(dt):
            if not dt: return "N/A"
            return dt.strftime('%d/%m/%Y')

        if contract.status == RentalStatus.ACTIVE.value:
            status_label = "Đang thuê"
        elif contract.status == RentalStatus.EXPIRED.value:
            status_label = "Hết hạn"
        elif contract.status == RentalStatus.BAD_DEBT.value:
            status_label = "Nợ xấu"
        else:
            status_label = "Không rõ"

        debt = fmt_num(contract.rental_debt)
        monthly = fmt_num(contract.monthly_rental)

        text = (
            f"<b>XÁC NHẬN ĐƯA VÀO BLACKLIST NỢ XẤU</b>\n\n"
            f"<b>Khách hàng:</b> {customer.customer_name}\n"
            f"<b>Liên hệ:</b> {customer.contact_info}\n"
            f"<b>Mã Hợp Đồng:</b> <code>{contract_code}</code>\n"
            f"<b>Trạng thái:</b> {status_label}\n"
            f"<b>Loại HĐ:</b> {contract.type_contract or 'N/A'}\n"
            f"<b>Mã BĐS:</b> {contract.real_estate_id or 'N/A'}\n"
            f"<b>Tiền thuê:</b> {monthly:,} VNĐ/tháng\n"
            # f"<b>Tổng tiền thanh toán HD:</b> {debt:,} VNĐ\n"
            f"<b>Thời gian thuê:</b> {fmt_dt(contract.start_rental)} - {fmt_dt(contract.end_rental)}\n\n"
            f"Bạn có chắc chắn muốn đưa hợp đồng này vào <b>BLACKLIST NỢ XẤU</b> không?"
        )

        buttons = [
            [
                InlineKeyboardButton("Xác nhận nợ xấu", callback_data=f"rbd_confirm_{contract.id}"),
                InlineKeyboardButton("Hủy", callback_data="rbd_cancel")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(buttons)
        await message.reply_text(text, reply_markup=reply_markup, parse_mode=ParseMode.HTML)

    except Exception as e:
        LogError(f"Error in rental_bad_debt_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình xử lý.")
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rbd_confirm_(.*)$"))
@command_timeout(auto_delete_cmd=True)
async def rental_bad_debt_confirm_callback(client, callback_query: CallbackQuery):
    contract_uuid = callback_query.matches[0].group(1)
    db = SessionLocal()
    try:
        from app.models.rental import Rental, RentalCustomer, RentalStatus

        contract = db.query(Rental).filter(Rental.id == contract_uuid).first()
        if not contract:
            await callback_query.message.edit_text("⚠️ Hợp đồng không tồn tại hoặc đã bị xóa.")
            return

        if contract.status == RentalStatus.BAD_DEBT.value:
            await callback_query.message.edit_text(f"⚠️ Hợp đồng <b>{contract.contract_id}</b> đã nằm trong Blacklist Nợ Xấu từ trước.", parse_mode=ParseMode.HTML)
            return

        contract.status = RentalStatus.BAD_DEBT.value
        db.commit()

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()

        def fmt_num(val):
            if val is None: return 0
            return int(val) if val == int(val) else val

        debt = fmt_num(contract.rental_debt)

        await callback_query.message.edit_text(
            f"✅ Đã đưa hợp đồng cho thuê <b>{contract.contract_id}</b> và Khách hàng <b>{customer.customer_name if customer else 'N/A'}</b> vào <b>BLACKLIST NỢ XẤU</b> thành công!\n\n"
            f"<b>Tổng tiền thanh toán HD:</b> {debt:,} VNĐ\n\n"
            f"<i>Thông tin vẫn được lưu giữ nguyên trạng để tiếp tục truy thu hoặc tra soát.</i>",
            parse_mode=ParseMode.HTML
        )
        LogInfo(f"[RentalBadDebt] Contract {contract.contract_id} marked as BAD_DEBT by user {callback_query.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in rbd_confirm callback: {e}", LogType.SYSTEM_STATUS)
        await callback_query.answer("❌ Có lỗi xảy ra khi xác nhận nợ xấu.", show_alert=True)
    finally:
        db.close()


@bot.on_callback_query(filters.regex(r"^rbd_cancel$"))
@command_timeout(auto_delete_cmd=True)
async def rental_bad_debt_cancel_callback(client, callback_query: CallbackQuery):
    await callback_query.message.delete()


# --- Rental: Tạo Phiếu Thu ---
@bot.on_message(filters.command(["rental_tao_phieu_thu"]) | filters.regex(r"^@\w+\s+/rental_tao_phieu_thu\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("Rental")
@require_group_role("main", "member")
@command_timeout(timeout_seconds=600, auto_delete_cmd=True)  # Form nhiều trường -> cần thời gian điền
async def rental_tao_phieu_thu_handler(client, message: Message) -> None:
    lines = message.text.strip().split("\n")

    # If single line -> send blank/prefilled form template
    if len(lines) < 2:
        args = message.text.split(maxsplit=1)
        contract_code_prefill = ""
        if len(args) > 1:
            contract_code_prefill = args[1].strip()

        today_str = datetime.date.today().strftime("%d/%m/%Y")
        
        form_template = f"""<b>FORM TẠO PHIẾU THU HỢP ĐỒNG</b>
Vui lòng sao chép form dưới đây, điền thông tin và gửi lại:

<pre>/rental_tao_phieu_thu
Mã hợp đồng: {contract_code_prefill}
Ngày thanh toán: {today_str}
Số tiền thanh toán: </pre>"""
        await message.reply_text(form_template, parse_mode=ParseMode.HTML)
        return

    # If multi-line -> parse form
    data = {}
    for line in lines[1:]:
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip().lower()] = val.strip()

    contract_id = data.get("mã hợp đồng", "").strip()
    payment_date_str = data.get("ngày thanh toán", "").strip()
    payment_amount_str = data.get("số tiền thanh toán", "").strip()

    if not contract_id or not payment_date_str or not payment_amount_str:
        await message.reply_text(
            "⚠️ <b>Mã hợp đồng</b>, <b>Ngày thanh toán</b> và <b>Số tiền thanh toán</b> là bắt buộc.",
            parse_mode=ParseMode.HTML
        )
        return

    # Parse and validate amount
    try:
        paid_amount = _parse_float_rental(payment_amount_str)
    except Exception:
        paid_amount = 0.0

    if paid_amount <= 0:
        await message.reply_text("⚠️ <b>Số tiền thanh toán</b> không hợp lệ hoặc phải lớn hơn 0.", parse_mode=ParseMode.HTML)
        return

    # Parse and validate date
    payment_date = None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            payment_date = datetime.datetime.strptime(payment_date_str, fmt).date()
            break
        except ValueError:
            continue

    if not payment_date:
        await message.reply_text("⚠️ Định dạng <b>Ngày thanh toán</b> không hợp lệ. Vui lòng dùng <code>dd/mm/yyyy</code>.", parse_mode=ParseMode.HTML)
        return

    db = SessionLocal()
    try:
        from app.models.rental import Rental, RentalPayment, RentalCustomer
        import uuid as _uuid

        # Verify contract exists
        contract = db.query(Rental).filter(Rental.contract_id == contract_id).first()
        if not contract:
            await message.reply_text(f"❌ Không tìm thấy hợp đồng cho thuê nào có mã <b>{contract_id}</b> trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        # Insert new RentalPayment record (DO NOT update rentals table)
        new_payment = RentalPayment(
            id=_uuid.uuid4(),
            contract_id=contract.contract_id,
            payment_date=payment_date,
            payment_time=datetime.datetime.now(),
            payment_amount=paid_amount
        )
        db.add(new_payment)
        db.commit()

        customer = db.query(RentalCustomer).filter(RentalCustomer.id == contract.customer_id).first()
        cust_name = customer.customer_name if customer else "N/A"

        def fmt_money_local(val):
            if val is None: return "0 VNĐ"
            return f"{int(val):,} VNĐ".replace(",", ".")

        success_msg = (
            f"✅ <b>TẠO PHIẾU THU THÀNH CÔNG</b>\n"
            f"{'━' * 15}\n"
            f"<b>Mã hợp đồng:</b> <code>{contract.contract_id}</code>\n"
            f"<b>Khách hàng:</b> <b>{cust_name}</b>\n"
            f"<b>Ngày thanh toán:</b> <code>{payment_date.strftime('%d/%m/%Y')}</code>\n"
            f"<b>Số tiền thanh toán:</b> <code>{fmt_money_local(paid_amount)}</code>\n\n"
            f"<i>Lưu ý: Đã tạo bản ghi thanh toán mới. Không thay đổi dư nợ trong bảng rentals.</i>"
        )
        await message.reply_text(success_msg, parse_mode=ParseMode.HTML)
        LogInfo(f"[RentalTaoPhieuThu] Created payment of {paid_amount} for contract {contract.contract_id} by {message.from_user.id}", LogType.SYSTEM_STATUS)

    except Exception as e:
        db.rollback()
        LogError(f"Error in rental_tao_phieu_thu_handler: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text("❌ Có lỗi xảy ra trong quá trình tạo phiếu thu.")
    finally:
        db.close()


from pyrogram import filters
from bot.utils.enums import CustomTitle
from pyrogram.types import Message
from pyrogram.enums import ParseMode
from bot.utils.bot import bot
from bot.utils.utils import check_command_target, require_user_type, require_project_name, require_custom_title
from bot.utils.enums import UserType
import re

# GGoMoonSin project handlers

# --- Thêm nhân viên ---
@bot.on_message(filters.command(["ggomoonsin_create_employee", "ggomoonsin_tao_nhan_vien"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_create_employee|ggomoonsin_tao_nhan_vien)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_create_employee_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_create_employee", "ggomoonsin_tao_nhan_vien"])
    if args is None: return

    from bot.utils.human_resource import handle_create_employee
    await handle_create_employee(client, message, "/ggomoonsin_create_employee")


# --- Cập nhật nhân viên ---
@bot.on_message(filters.command(["ggomoonsin_update_employee", "ggomoonsin_cap_nhat_nhan_vien"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_update_employee|ggomoonsin_cap_nhat_nhan_vien)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_update_employee_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_update_employee", "ggomoonsin_cap_nhat_nhan_vien"])
    if args is None: return

    from bot.utils.human_resource import handle_update_employee
    await handle_update_employee(client, message, "/ggomoonsin_update_employee")


# --- Xóa nhân viên (soft delete) ---
@bot.on_message(filters.command(["ggomoonsin_delete_employee", "ggomoonsin_xoa_nhan_vien"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_delete_employee|ggomoonsin_xoa_nhan_vien)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_delete_employee_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_delete_employee", "ggomoonsin_xoa_nhan_vien"])
    if args is None: return

    from bot.utils.human_resource import handle_delete_employee
    await handle_delete_employee(client, message, "/ggomoonsin_delete_employee")


# --- Chấm công (Check-in) ---
@bot.on_message(filters.command(["ggomoonsin_check_in", "ggomoonsin_cham_cong"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_check_in|ggomoonsin_cham_cong)\b"))
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.MEMBER_HR)
async def ggomoonsin_check_in_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_check_in", "ggomoonsin_cham_cong"])
    if args is None: return

    from bot.utils.human_resource import handle_check_in
    await handle_check_in(client, message, "/ggomoonsin_check_in")


# --- Tan ca (Check-out) ---
@bot.on_message(filters.command(["ggomoonsin_check_out", "ggomoonsin_tan_ca"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_check_out|ggomoonsin_tan_ca)\b"))
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.MEMBER_HR)
async def ggomoonsin_check_out_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_check_out", "ggomoonsin_tan_ca"])
    if args is None: return

    from bot.utils.human_resource import handle_check_out
    await handle_check_out(client, message, "/ggomoonsin_check_out")


# --- Xin nghỉ phép (Request Leave) ---
@bot.on_message(filters.command(["ggomoonsin_request_leave", "ggomoonsin_xin_nghi_phep"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_request_leave|ggomoonsin_xin_nghi_phep)\b"))
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.MEMBER_HR)
async def ggomoonsin_request_leave_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_request_leave", "ggomoonsin_xin_nghi_phep"])
    if args is None: return

    from bot.utils.human_resource import handle_request_leave
    await handle_request_leave(client, message, "/ggomoonsin_request_leave")


# --- Đăng ký tăng ca (Request Overtime) ---
@bot.on_message(filters.command(["ggomoonsin_request_overtime", "ggomoonsin_dang_ky_tang_ca"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_request_overtime|ggomoonsin_dang_ky_tang_ca)\b"))
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.MEMBER_HR)
async def ggomoonsin_request_overtime_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_request_overtime", "ggomoonsin_dang_ky_tang_ca"])
    if args is None: return

    from bot.utils.human_resource import handle_request_overtime
    await handle_request_overtime(client, message, "/ggomoonsin_request_overtime")


# --- Callback handlers cho HR ---
@bot.on_callback_query(filters.regex(r"^auth_(ci|co|lv|ov)\|(.+)$"))
async def _ggomoonsin_auth_select_callback(client, callback_query):
    from bot.utils.human_resource import handle_authority_callback
    await handle_authority_callback(client, callback_query)

@bot.on_callback_query(filters.regex(r"^lv_req\|(ok|no)\|([-\d]+)\|(\d+)$"))
async def _ggomoonsin_lv_req_callback(client, callback_query):
    from bot.utils.human_resource import handle_leave_request_callback
    await handle_leave_request_callback(client, callback_query)

@bot.on_callback_query(filters.regex(r"^ov_req\|(ok|no)\|([-\d]+)\|(\d+)$"))
async def _ggomoonsin_ov_req_callback(client, callback_query):
    from bot.utils.human_resource import handle_overtime_request_callback
    await handle_overtime_request_callback(client, callback_query)



# --- Xem chấm công (List Check-in) ---
@bot.on_message(filters.command(["ggomoonsin_list_check_in", "ggomoonsin_xem_cham_cong"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_list_check_in|ggomoonsin_xem_cham_cong)\b"))
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.MEMBER_HR)
async def ggomoonsin_list_check_in_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_list_check_in", "ggomoonsin_xem_cham_cong"])
    if args is None: return

    from bot.utils.human_resource import handle_list_check_in
    await handle_list_check_in(client, message, "/ggomoonsin_list_check_in")


# --- Xem danh sách nghỉ phép (List Request Leave) ---
@bot.on_message(filters.command(["ggomoonsin_list_request_leave", "ggomoonsin_xem_nghi_phep"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_list_request_leave|ggomoonsin_xem_nghi_phep)\b"))
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.MEMBER_HR)
async def ggomoonsin_list_request_leave_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_list_request_leave", "ggomoonsin_xem_nghi_phep"])
    if args is None: return

    from bot.utils.human_resource import handle_list_request_leave
    await handle_list_request_leave(client, message, "/ggomoonsin_list_request_leave")


# --- Giao việc (Create Task) ---
@bot.on_message(filters.command(["ggomoonsin_create_task", "ggomoonsin_giao_viec"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_create_task|ggomoonsin_giao_viec)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_create_task_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_create_task", "ggomoonsin_giao_viec"])
    if args is None: return

    from bot.utils.human_resource import handle_create_task
    await handle_create_task(client, message, "/ggomoonsin_create_task")


# --- Xem danh sách task (List Tasks) ---
@bot.on_message(filters.command(["ggomoonsin_list_tasks", "ggomoonsin_xem_cong_viec"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_list_tasks|ggomoonsin_xem_cong_viec)\b"))
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.MEMBER_HR)
async def ggomoonsin_list_tasks_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_list_tasks", "ggomoonsin_xem_cong_viec"])
    if args is None: return

    from bot.utils.human_resource import handle_list_tasks
    await handle_list_tasks(client, message, "/ggomoonsin_list_tasks")


# --- Hủy task (Reply /cancel vào tin nhắn task) ---
@bot.on_message(filters.command(["cancel", "huy_task"]) & filters.reply)
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_cancel_task_handler(client, message: Message) -> None:
    from bot.utils.human_resource import handle_cancel_task_reply
    await handle_cancel_task_reply(client, message)


# --- Xuất bảng lương (Export Payroll) ---
@bot.on_message(filters.command(["export_payroll", "ggomoonsin_xuat_luong", "ggomoonsin_export_payroll"]) | filters.regex(r"^@\w+\s+/(export_payroll|ggomoonsin_xuat_luong|ggomoonsin_export_payroll)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_export_payroll_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["export_payroll", "ggomoonsin_xuat_luong", "ggomoonsin_export_payroll"])
    if args is None: return

    from bot.utils.human_resource import handle_export_payroll
    # Re-construct message.text for the actual handler to parse args
    import re
    cmd = args[0] if args else "export_payroll"
    message.text = f"/{cmd} " + " ".join(args[1:]) if len(args) > 1 else f"/{cmd}"
    await handle_export_payroll(client, message, f"/{cmd}")


# --- Tạo lại bảng chấm công (Recreate Attendance Report) ---
@bot.on_message(filters.command(["ggomoonsin_recreate_attendance_report", "ggomoonsin_tao_lai_bang_cham_cong"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_recreate_attendance_report|ggomoonsin_tao_lai_bang_cham_cong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_recreate_attendance_report_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_recreate_attendance_report", "ggomoonsin_tao_lai_bang_cham_cong"])
    if args is None: return

    from bot.utils.human_resource import handle_recreate_attendance_report
    import re
    cmd = args[0] if args else "ggomoonsin_recreate_attendance_report"
    message.text = f"/{cmd} " + " ".join(args[1:]) if len(args) > 1 else f"/{cmd}"
    await handle_recreate_attendance_report(client, message, f"/{cmd}")


# --- Xem danh sách công việc (Admin - Check Tasks) ---
@bot.on_message(filters.command(["ggomoonsin_check_tasks", "ggomoonsin_danh_sach_cong_viec"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_check_tasks|ggomoonsin_danh_sach_cong_viec)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_check_tasks_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_check_tasks", "ggomoonsin_danh_sach_cong_viec"])
    if args is None: return

    from bot.utils.human_resource import handle_check_tasks
    cmd = args[0] if args else "ggomoonsin_danh_sach_cong_viec"
    clean_cmd = cmd.split('@')[0]
    message.text = cmd + " " + " ".join(args[1:]) if len(args) > 1 else cmd
    await handle_check_tasks(client, message, clean_cmd)


# --- Xuất bảng lương Excel (List Payroll Excel) ---
@bot.on_message(filters.command(["ggomoonsin_list_payroll", "ggomoonsin_xuat_danh_sach_luong"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_list_payroll|ggomoonsin_xuat_danh_sach_luong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_list_payroll_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_list_payroll", "ggomoonsin_xuat_danh_sach_luong"])
    if args is None: return

    from bot.utils.human_resource import handle_list_payroll_excel
    import re
    cmd = args[0] if args else "ggomoonsin_list_payroll"
    message.text = cmd + " " + " ".join(args[1:]) if len(args) > 1 else cmd
    await handle_list_payroll_excel(client, message, cmd)


# --- Xuất danh sách nhân viên Excel (List Employee Excel) ---
@bot.on_message(filters.command(["ggomoonsin_list_employee", "ggomoonsin_danh_sach_nhan_vien"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_list_employee|ggomoonsin_danh_sach_nhan_vien)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_list_employee_handler(client, message: Message) -> None:
    from app.db.session import SessionLocal
    db = SessionLocal()
    try:
        from app.models.employee import Employee
        import tempfile
        import os
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        employees = db.query(Employee).order_by(Employee.last_name, Employee.first_name).all()

        if not employees:
            await message.reply_text("⚠️ Không có nhân viên nào trong hệ thống.", parse_mode=ParseMode.HTML)
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DanhSachNhanVien"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = ["STT", "Họ", "Tên", "Username TG", "SĐT", "Chức Vụ", "Lương CB", "Trạng Thái"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, emp in enumerate(employees, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
            ws.cell(row=row_idx, column=2, value=emp.last_name or "").alignment = left_align
            ws.cell(row=row_idx, column=3, value=emp.first_name or "").alignment = left_align
            ws.cell(row=row_idx, column=4, value=emp.telegram_username or "").alignment = left_align
            ws.cell(row=row_idx, column=5, value=emp.phone or "").alignment = left_align
            ws.cell(row=row_idx, column=6, value=emp.position or "").alignment = left_align
            ws.cell(row=row_idx, column=7, value=emp.base_salary or 0).alignment = center_align
            ws.cell(row=row_idx, column=8, value=emp.status or "").alignment = center_align
            for c in range(1, 9):
                ws.cell(row=row_idx, column=c).border = thin_border

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="ggomoonsin_ds_nv_")
        wb.save(tmp.name)
        tmp.close()

        await message.reply_document(
            document=tmp.name,
            caption=f"📋 <b>DANH SÁCH NHÂN VIÊN GGOMOONSIN</b>\nTổng: {len(employees)} nhân viên",
            parse_mode=ParseMode.HTML
        )
        os.unlink(tmp.name)
    except Exception as e:
        from bot.utils.logger import LogError, LogType
        LogError(f"Error exporting employee list: {e}", LogType.SYSTEM_STATUS)
        await message.reply_text(f"❌ Có lỗi xảy ra: {e}", parse_mode=ParseMode.HTML)
    finally:
        db.close()


# --- Xuất danh sách chấm công (List Attendance Excel) ---
@bot.on_message(filters.command(["ggomoonsin_list_attendance", "ggomoonsin_danh_sach_cham_cong"]) | filters.regex(r"^@\w+\s+/(ggomoonsin_list_attendance|ggomoonsin_danh_sach_cham_cong)\b"))
@require_user_type(UserType.OWNER, UserType.ADMIN)
@require_project_name("GGoMoonSin")
@require_custom_title(CustomTitle.SUPER_MAIN, CustomTitle.MAIN_HR)
async def ggomoonsin_list_attendance_handler(client, message: Message) -> None:
    args = await check_command_target(client, message.text, ["ggomoonsin_list_attendance", "ggomoonsin_danh_sach_cham_cong"])
    if args is None: return

    from bot.utils.human_resource import handle_list_attendance_excel
    import re
    cmd = args[0] if args else "ggomoonsin_list_attendance"
    message.text = cmd + " " + " ".join(args[1:]) if len(args) > 1 else cmd
    await handle_list_attendance_excel(client, message, cmd)
